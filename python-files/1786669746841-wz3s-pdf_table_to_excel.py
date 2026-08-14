#!/usr/bin/env python3
"""
PDF 表格转 Excel 工具
=====================
功能：
  1. 识别 PDF 中的表格（支持多页、多表格）
  2. 按原 PDF 表格形式输出到 Excel（保留合并单元格样式、表头格式）
  3. 对提取的数据进行自动分析（数据类型推断、统计摘要、空值检测）
  4. 将处理好的数据按表头匹配填入模板 Excel

用法：
  # 基础：提取 PDF 中所有表格到 Excel
  python pdf_table_to_excel.py extract input.pdf output.xlsx

  # 提取 + 分析报告
  python pdf_table_to_excel.py analyze input.pdf output.xlsx --report analysis.txt

  # 提取 + 按模板填入（模板第一行为表头）
  python pdf_table_to_excel.py template input.pdf template.xlsx output.xlsx

  # 全流程：提取 + 分析 + 模板填入
  python pdf_table_to_excel.py full input.pdf template.xlsx output.xlsx --report analysis.txt

依赖：pdfplumber, openpyxl, pandas
"""

import argparse
import json
import os
import sys
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


# ──────────────────────────────────────────────
# 1. PDF 表格识别与提取
# ──────────────────────────────────────────────

class PDFTableExtractor:
    """从 PDF 中识别并提取表格"""

    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.tables: List[Dict[str, Any]] = []  # 存储所有提取的表格

    def extract_all_tables(
        self,
        table_settings: Optional[Dict] = None,
        min_rows: int = 2,
        min_cols: int = 2,
    ) -> List[Dict[str, Any]]:
        """
        提取 PDF 中所有页面的所有表格。

        Args:
            table_settings: pdfplumber 的表格检测参数（None 则自动推断）
            min_rows: 最小行数过滤
            min_cols: 最小列数过滤

        Returns:
            表格列表，每个表格包含: page, table_index, data, headers, bbox
        """
        if table_settings is None:
            # 默认使用自动策略 + 横线检测
            table_settings = {"vertical_strategy": "lines", "horizontal_strategy": "lines"}

        extracted = []

        with pdfplumber.open(self.pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                # 先尝试线条检测
                tables = page.extract_tables(table_settings=table_settings)

                # 如果线条检测没有结果，回退到文本策略
                if not tables:
                    text_settings = {
                        "vertical_strategy": "text",
                        "horizontal_strategy": "text",
                        "keep_blank_chars": False,
                    }
                    tables = page.extract_tables(table_settings=text_settings)

                for tbl_idx, table_data in enumerate(tables, 1):
                    # 过滤过小的表格
                    if len(table_data) < min_rows or (
                        table_data and len(table_data[0]) < min_cols
                    ):
                        continue

                    # 清洗数据：去除 None，统一为空字符串
                    cleaned = []
                    for row in table_data:
                        cleaned_row = [
                            str(cell).strip() if cell is not None else "" for cell in row
                        ]
                        cleaned.append(cleaned_row)

                    # 第一行作为表头
                    headers = cleaned[0] if cleaned else []
                    data_rows = cleaned[1:] if len(cleaned) > 1 else []

                    table_info = {
                        "page": page_num,
                        "table_index": tbl_idx,
                        "headers": headers,
                        "data": data_rows,
                        "num_rows": len(data_rows),
                        "num_cols": len(headers),
                        "raw": cleaned,  # 包含表头的完整数据
                    }
                    extracted.append(table_info)

        self.tables = extracted
        return extracted

    def get_table_summary(self) -> str:
        """生成表格提取摘要"""
        if not self.tables:
            return "未检测到表格。"
        lines = [f"共检测到 {len(self.tables)} 个表格：\n"]
        for t in self.tables:
            lines.append(
                f"  - 第{t['page']}页 表{t['table_index']}: "
                f"{t['num_rows']} 行 × {t['num_cols']} 列"
            )
            lines.append(f"    表头: {t['headers']}")
        return "\n".join(lines)


# ──────────────────────────────────────────────
# 2. Excel 输出（保留原 PDF 格式）
# ──────────────────────────────────────────────

class ExcelWriter:
    """将提取的表格写入 Excel，保留原 PDF 格式"""

    # 样式定义
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DATA_FONT = Font(name="微软雅黑", size=10)
    DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    ALT_ROW_FILL = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

    def __init__(self):
        self.workbook = openpyxl.Workbook()
        # 删除默认 Sheet
        self.workbook.remove(self.workbook.active)

    def add_table_sheet(self, table: Dict[str, Any], sheet_name: Optional[str] = None):
        """
        将单个表格写入一个 Sheet。

        Args:
            table: 表格数据字典
            sheet_name: Sheet 名称（None 则自动生成）
        """
        if sheet_name is None:
            sheet_name = f"P{table['page']}_T{table['table_index']}"
        # Sheet 名称最长 31 字符
        sheet_name = sheet_name[:31]
        ws = self.workbook.create_sheet(title=sheet_name)

        headers = table["headers"]
        data = table["data"]

        # 写表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.THIN_BORDER

        # 写数据行
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._parse_cell_value(value))
                cell.font = self.DATA_FONT
                cell.alignment = self.DATA_ALIGN
                cell.border = self.THIN_BORDER
                # 隔行变色
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

        # 自动列宽
        self._auto_column_width(ws, table["raw"])

        # 冻结首行
        ws.freeze_panes = "A2"

    def _parse_cell_value(self, value: str):
        """尝试将字符串转换为数值类型"""
        if not value or value == "":
            return None
        # 尝试转为整数
        try:
            int_val = int(value.replace(",", ""))
            return int_val
        except (ValueError, AttributeError):
            pass
        # 尝试转为浮点数
        try:
            float_val = float(value.replace(",", "").replace("%", ""))
            if "%" in value:
                return float_val / 100  # 百分比转为小数
            return float_val
        except (ValueError, AttributeError):
            pass
        return value

    def _auto_column_width(self, ws, raw_data: List[List[str]]):
        """根据内容自动调整列宽"""
        if not raw_data:
            return
        num_cols = len(raw_data[0])
        for col_idx in range(1, num_cols + 1):
            max_length = 0
            for row in raw_data:
                if col_idx - 1 < len(row):
                    cell_len = len(str(row[col_idx - 1]))
                    # 中文字符按 2 个宽度计算
                    chinese_count = sum(1 for c in str(row[col_idx - 1]) if '\u4e00' <= c <= '\u9fff')
                    cell_len = cell_len + chinese_count
                    max_length = max(max_length, cell_len)
            # 限制列宽范围
            col_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    def save(self, output_path: str):
        """保存 Excel 文件"""
        # 如果没有任何 Sheet，创建一个空 Sheet
        if not self.workbook.sheetnames:
            self.workbook.create_sheet(title="空")
        self.workbook.save(output_path)
        return output_path


# ──────────────────────────────────────────────
# 3. 数据分析
# ──────────────────────────────────────────────

class DataAnalyzer:
    """对提取的表格数据进行分析"""

    @staticmethod
    def analyze_table(table: Dict[str, Any]) -> Dict[str, Any]:
        """
        分析单个表格，返回分析结果。

        包括：数据类型推断、统计摘要、空值检测、唯一值检测
        """
        headers = table["headers"]
        data = table["data"]

        if not data:
            return {"error": "无数据行", "headers": headers}

        # 转为 DataFrame 便于分析
        df = pd.DataFrame(data, columns=headers)

        analysis = {
            "table_location": f"第{table['page']}页 表{table['table_index']}",
            "dimensions": f"{len(df)} 行 × {len(df.columns)} 列",
            "columns": [],
        }

        for col in df.columns:
            col_data = df[col]
            col_info = {"column_name": col}

            # 空值统计
            empty_count = int((col_data == "").sum()) + int(col_data.isna().sum())
            col_info["empty_count"] = empty_count
            col_info["empty_ratio"] = f"{empty_count / len(df) * 100:.1f}%"

            # 尝试数值分析
            numeric_series = pd.to_numeric(
                col_data.str.replace(",", "").str.replace("%", "").str.replace("￥", "").str.strip(),
                errors="coerce"
            )
            non_null_numeric = numeric_series.dropna()

            if len(non_null_numeric) > 0 and len(non_null_numeric) / len(df) > 0.5:
                # 数值型列
                col_info["data_type"] = "数值型"
                col_info["stats"] = {
                    "min": float(non_null_numeric.min()),
                    "max": float(non_null_numeric.max()),
                    "mean": round(float(non_null_numeric.mean()), 2),
                    "median": round(float(non_null_numeric.median()), 2),
                    "sum": round(float(non_null_numeric.sum()), 2),
                    "std": round(float(non_null_numeric.std()), 2) if len(non_null_numeric) > 1 else 0,
                }
            else:
                # 文本型列
                col_info["data_type"] = "文本型"
                unique_vals = int(col_data[col_data != ""].nunique())
                col_info["unique_values"] = unique_vals
                if unique_vals <= 10:
                    dist = col_data[col_data != ""].value_counts()
                    col_info["value_distribution"] = {str(k): int(v) for k, v in dist.items()}

            analysis["columns"].append(col_info)

        # 整体质量评分
        total_cells = len(df) * len(df.columns)
        empty_cells = sum(c["empty_count"] for c in analysis["columns"])
        analysis["data_quality"] = {
            "total_cells": total_cells,
            "empty_cells": empty_cells,
            "completeness": f"{(1 - empty_cells / total_cells) * 100:.1f}%" if total_cells > 0 else "N/A",
        }

        return analysis

    @staticmethod
    def generate_report(analyses: List[Dict[str, Any]]) -> str:
        """生成可读的分析报告"""
        lines = ["=" * 60, "PDF 表格数据分析报告", "=" * 60, ""]

        for a in analyses:
            lines.append(f"【{a.get('table_location', '未知')}】")
            lines.append(f"  维度: {a.get('dimensions', 'N/A')}")

            if "error" in a:
                lines.append(f"  ⚠ {a['error']}")
                lines.append("")
                continue

            lines.append(f"  数据完整度: {a['data_quality']['completeness']} "
                        f"({a['data_quality']['empty_cells']}/{a['data_quality']['total_cells']} 空值)")
            lines.append("")

            for col in a.get("columns", []):
                lines.append(f"  ▸ {col['column_name']} ({col['data_type']})")
                lines.append(f"    空值: {col['empty_count']} ({col['empty_ratio']})")

                if col["data_type"] == "数值型":
                    s = col["stats"]
                    lines.append(f"    最小值: {s['min']}  最大值: {s['max']}")
                    lines.append(f"    均值: {s['mean']}  中位数: {s['median']}")
                    lines.append(f"    总和: {s['sum']}  标准差: {s['std']}")
                else:
                    lines.append(f"    唯一值数: {col.get('unique_values', 'N/A')}")
                    if "value_distribution" in col:
                        lines.append(f"    值分布:")
                        for val, cnt in col["value_distribution"].items():
                            lines.append(f"      {val}: {cnt}次")

            lines.append("")

        lines.append("=" * 60)
        lines.append("报告结束")
        return "\n".join(lines)


# ──────────────────────────────────────────────
# 4. 模板填充
# ──────────────────────────────────────────────

class TemplateFiller:
    """将提取的数据按表头匹配填入模板 Excel"""

    @staticmethod
    def load_template(template_path: str) -> Dict[str, Any]:
        """
        加载模板 Excel，读取所有 Sheet 的表头（第一行）。

        Returns:
            {sheet_name: {"headers": [...], "header_row": 1, "max_col": N}}
        """
        wb = openpyxl.load_workbook(template_path)
        templates = {}

        for ws in wb.worksheets:
            headers = []
            max_col = ws.max_column
            header_row = 1

            # 找到表头行（第一个非空行）
            for row in ws.iter_rows(min_row=1, max_row=5, max_col=max_col):
                row_values = [cell.value for cell in row]
                if any(v is not None and str(v).strip() != "" for v in row_values):
                    header_row = row[0].row
                    headers = [str(v).strip() if v is not None else "" for v in row_values]
                    break

            templates[ws.title] = {
                "headers": headers,
                "header_row": header_row,
                "max_col": max_col,
                "max_row": ws.max_row,
            }

        wb.close()
        return templates

    @staticmethod
    def match_headers(
        pdf_headers: List[str], template_headers: List[str]
    ) -> Dict[int, int]:
        """
        匹配 PDF 表头与模板表头，返回列映射关系。

        Returns:
            {pdf_col_index: template_col_index}
        """
        mapping = {}

        # 精确匹配
        for p_idx, p_header in enumerate(pdf_headers):
            p_clean = p_header.strip().lower()
            for t_idx, t_header in enumerate(template_headers):
                t_clean = t_header.strip().lower()
                if p_clean == t_clean and p_clean != "":
                    mapping[p_idx] = t_idx
                    break

        # 模糊匹配（包含关系）—— 仅对未精确匹配的列
        for p_idx, p_header in enumerate(pdf_headers):
            if p_idx in mapping or not p_header.strip():
                continue
            p_clean = p_header.strip().lower()
            best_score = 0
            best_t_idx = None
            for t_idx, t_header in enumerate(template_headers):
                if t_idx in mapping.values() or not t_header.strip():
                    continue
                t_clean = t_header.strip().lower()
                # 简单的包含关系匹配
                if p_clean in t_clean or t_clean in p_clean:
                    score = min(len(p_clean), len(t_clean)) / max(len(p_clean), len(t_clean))
                    if score > best_score:
                        best_score = score
                        best_t_idx = t_idx
            if best_t_idx is not None and best_score > 0.3:
                mapping[p_idx] = best_t_idx

        return mapping

    @staticmethod
    def fill_template(
        tables: List[Dict[str, Any]],
        template_path: str,
        output_path: str,
    ) -> str:
        """
        将 PDF 表格数据填入模板 Excel。

        策略：
        - 加载模板，读取每个 Sheet 的表头
        - 对每个 PDF 表格，找到表头匹配度最高的模板 Sheet
        - 按列映射关系填入数据
        - 保留模板原有的格式和样式

        Args:
            tables: PDF 提取的表格列表
            template_path: 模板 Excel 路径
            output_path: 输出路径

        Returns:
            输出文件路径
        """
        # 加载模板
        wb = openpyxl.load_workbook(template_path)
        templates = TemplateFiller.load_template(template_path)

        if not templates:
            # 没有模板表头，直接追加数据
            for table in tables:
                sheet_name = f"P{table['page']}_T{table['table_index']}"[:31]
                ws = wb.create_sheet(title=sheet_name)
                # 写入完整数据
                for row_idx, row in enumerate(table["raw"], 1):
                    for col_idx, val in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)
            wb.save(output_path)
            return output_path

        fill_log = []

        for table in tables:
            pdf_headers = table["headers"]
            best_sheet = None
            best_mapping = {}
            best_match_count = 0

            # 找到匹配度最高的模板 Sheet
            for sheet_name, tmpl_info in templates.items():
                mapping = TemplateFiller.match_headers(pdf_headers, tmpl_info["headers"])
                if len(mapping) > best_match_count:
                    best_match_count = len(mapping)
                    best_sheet = sheet_name
                    best_mapping = mapping

            if best_sheet is None or best_match_count == 0:
                # 没有匹配的模板 Sheet，新建一个
                sheet_name = f"P{table['page']}_T{table['table_index']}"[:31]
                ws = wb.create_sheet(title=sheet_name)
                for row_idx, row in enumerate(table["raw"], 1):
                    for col_idx, val in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)
                fill_log.append(f"表(P{table['page']}_T{table['table_index']}): 无匹配模板，已新建 Sheet")
                continue

            # 填入数据到匹配的模板 Sheet
            ws = wb[best_sheet]
            tmpl_info = templates[best_sheet]
            start_row = tmpl_info["header_row"] + 1

            # 如果模板中已有数据行，找到下一个空行
            while ws.cell(row=start_row, column=1).value is not None:
                start_row += 1

            # 写入数据
            for row_data in table["data"]:
                for p_idx, value in enumerate(row_data):
                    if p_idx in best_mapping:
                        t_idx = best_mapping[p_idx]
                        # 转换数值
                        parsed_val = TemplateFiller._parse_value(value)
                        ws.cell(row=start_row, column=t_idx + 1, value=parsed_val)
                start_row += 1

            matched_cols = [pdf_headers[k] for k in best_mapping]
            fill_log.append(
                f"表(P{table['page']}_T{table['table_index']}) → 模板Sheet「{best_sheet}」, "
                f"匹配 {best_match_count}/{len(pdf_headers)} 列: {matched_cols}"
            )

        wb.save(output_path)

        # 打印填充日志
        print("\n模板填充日志：")
        for log in fill_log:
            print(f"  {log}")

        return output_path

    @staticmethod
    def _parse_value(value: str):
        """解析单元格值"""
        if not value or value == "":
            return None
        try:
            return int(value.replace(",", ""))
        except (ValueError, AttributeError):
            pass
        try:
            cleaned = value.replace(",", "").replace("%", "").replace("￥", "").strip()
            float_val = float(cleaned)
            if "%" in value:
                return float_val / 100
            return float_val
        except (ValueError, AttributeError):
            pass
        return value


# ──────────────────────────────────────────────
# 5. 主流程
# ──────────────────────────────────────────────

def cmd_extract(args):
    """提取表格到 Excel"""
    extractor = PDFTableExtractor(args.input)
    tables = extractor.extract_all_tables()

    print(extractor.get_table_summary())

    if not tables:
        print("未检测到表格，退出。")
        return

    writer = ExcelWriter()
    for table in tables:
        writer.add_table_sheet(table)

    writer.save(args.output)
    print(f"\n已保存到: {args.output}")


def cmd_analyze(args):
    """提取表格 + 分析"""
    extractor = PDFTableExtractor(args.input)
    tables = extractor.extract_all_tables()

    print(extractor.get_table_summary())

    if not tables:
        print("未检测到表格，退出。")
        return

    # 分析
    analyzer = DataAnalyzer()
    analyses = [analyzer.analyze_table(t) for t in tables]
    report = analyzer.generate_report(analyses)

    print("\n" + report)

    # 保存分析报告
    if args.report:
        with open(args.report, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"分析报告已保存到: {args.report}")

    # 同时输出 Excel
    writer = ExcelWriter()
    for table in tables:
        writer.add_table_sheet(table)
    writer.save(args.output)
    print(f"Excel 已保存到: {args.output}")


def cmd_template(args):
    """提取表格 + 模板填充"""
    extractor = PDFTableExtractor(args.input)
    tables = extractor.extract_all_tables()

    print(extractor.get_table_summary())

    if not tables:
        print("未检测到表格，退出。")
        return

    TemplateFiller.fill_template(tables, args.template, args.output)
    print(f"\n已保存到: {args.output}")


def cmd_full(args):
    """全流程：提取 + 分析 + 模板填充"""
    extractor = PDFTableExtractor(args.input)
    tables = extractor.extract_all_tables()

    print(extractor.get_table_summary())

    if not tables:
        print("未检测到表格，退出。")
        return

    # 分析
    analyzer = DataAnalyzer()
    analyses = [analyzer.analyze_table(t) for t in tables]
    report = analyzer.generate_report(analyses)

    print("\n" + report)

    if args.report:
        with open(args.report, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"分析报告已保存到: {args.report}")

    # 模板填充
    TemplateFiller.fill_template(tables, args.template, args.output)
    print(f"\n已保存到: {args.output}")

    # 同时输出原始提取结果（用于对比）
    raw_output = args.output.replace(".xlsx", "_raw.xlsx")
    writer = ExcelWriter()
    for table in tables:
        writer.add_table_sheet(table)
    writer.save(raw_output)
    print(f"原始提取结果已保存到: {raw_output}")


def main():
    parser = argparse.ArgumentParser(
        description="PDF 表格转 Excel 工具 — 识别、分析、模板填充",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  # 提取 PDF 表格到 Excel
  python pdf_table_to_excel.py extract report.pdf output.xlsx

  # 提取 + 分析报告
  python pdf_table_to_excel.py analyze report.pdf output.xlsx --report analysis.txt

  # 提取 + 按模板填入
  python pdf_table_to_excel.py template report.pdf template.xlsx output.xlsx

  # 全流程
  python pdf_table_to_excel.py full report.pdf template.xlsx output.xlsx --report analysis.txt
        """,
    )

    subparsers = parser.add_subparsers(dest="command", help="运行模式")

    # extract
    p_extract = subparsers.add_parser("extract", help="提取 PDF 表格到 Excel")
    p_extract.add_argument("input", help="输入 PDF 文件路径")
    p_extract.add_argument("output", help="输出 Excel 文件路径")
    p_extract.set_defaults(func=cmd_extract)

    # analyze
    p_analyze = subparsers.add_parser("analyze", help="提取 + 数据分析")
    p_analyze.add_argument("input", help="输入 PDF 文件路径")
    p_analyze.add_argument("output", help="输出 Excel 文件路径")
    p_analyze.add_argument("--report", help="分析报告输出路径（.txt）")
    p_analyze.set_defaults(func=cmd_analyze)

    # template
    p_template = subparsers.add_parser("template", help="提取 + 模板填充")
    p_template.add_argument("input", help="输入 PDF 文件路径")
    p_template.add_argument("template", help="模板 Excel 文件路径")
    p_template.add_argument("output", help="输出 Excel 文件路径")
    p_template.set_defaults(func=cmd_template)

    # full
    p_full = subparsers.add_parser("full", help="全流程：提取 + 分析 + 模板填充")
    p_full.add_argument("input", help="输入 PDF 文件路径")
    p_full.add_argument("template", help="模板 Excel 文件路径")
    p_full.add_argument("output", help="输出 Excel 文件路径")
    p_full.add_argument("--report", help="分析报告输出路径（.txt）")
    p_full.set_defaults(func=cmd_full)

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    # 检查输入文件
    if not os.path.exists(args.input):
        print(f"错误：输入文件不存在: {args.input}")
        sys.exit(1)

    if hasattr(args, "template") and not os.path.exists(args.template):
        print(f"错误：模板文件不存在: {args.template}")
        sys.exit(1)

    args.func(args)


if __name__ == "__main__":
    main()
