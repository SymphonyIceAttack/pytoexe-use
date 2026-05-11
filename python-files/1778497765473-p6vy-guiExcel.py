import openpyxl
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog

def seleccionar_archivo1():
    file_path1 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry1.delete(0, tk.END)
    entry1.insert(0, file_path1)

def seleccionar_archivo2():
    file_path2 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry2.delete(0, tk.END)
    entry2.insert(0, file_path2)

def copiar_datos():
    file_path1 = entry1.get()
    file_path2 = entry2.get()

    workbook1 = openpyxl.load_workbook(file_path1)
    sheet1 = workbook1.active

    workbook2 = openpyxl.load_workbook(file_path2)
    sheet2 = workbook2.active

    for row1 in range(1, sheet1.max_row + 1):
        cell1A = sheet1.cell(row=row1, column=1)
        for row2 in range(1, sheet2.max_row + 1):
            cell2A = sheet2.cell(row=row2, column=1)
            if cell1A.value is not None and cell2A.value is not None and cell1A.value.lower() == cell2A.value.lower():
                for col in range(1, sheet1.max_column + 1):
                    cell1 = sheet1.cell(row=row1, column=col)
                    cell2 = sheet2.cell(row=row2, column=col)
                    cell1.value = cell2.value
                    
    workbook1.save(file_path1)

root = tk.Tk()
root.title("Copiador de datos de Excel")

frame1 = tk.Frame(root)
frame1.pack(padx=10, pady=5)

entry1 = tk.Entry(frame1, width=50)
entry1.pack(side=tk.LEFT, padx=5)

button1 = tk.Button(frame1, text="Seleccionar Rellenar", command=seleccionar_archivo1)
button1.pack(side=tk.LEFT, padx=5)

frame2 = tk.Frame(root)
frame2.pack(padx=10, pady=5)

entry2 = tk.Entry(frame2, width=50)
entry2.pack(side=tk.LEFT, padx=5)

button2 = tk.Button(frame2, text="Seleccionar Ingredientes", command=seleccionar_archivo2)
button2.pack(side=tk.LEFT, padx=5)

ok_button = tk.Button(root, text="Copiar", command=copiar_datos)
ok_button.pack(padx=10, pady=10)

root.mainloop()
