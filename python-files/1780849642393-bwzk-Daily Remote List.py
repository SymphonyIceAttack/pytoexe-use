import os
import time
import win32com.client as win32
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import shutil

# ----------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------
base_path = r"\\tosan-solutions\shares\APM\Remote List"
output_folder = r"C:\Users\sheini.arman\Documents\Remote\New folder"
file_combine = os.path.join(output_folder, "Combine.xlsx")
file_combine_template = os.path.join(output_folder, "Combine_Template.xlsx")  # Template to preserve format

if not os.path.exists(output_folder):
    os.makedirs(output_folder)

folders = [
    "ProductionTeam", "prd-AAServer", "prd-A-Card", "prd-A-Core", "prd-A-ECom",
    "prd-A-Framework", "prd-A-Interbank", "prd-A-Loan", "prd-A-Modern", "prd-A-Switch",
    "prd-A-Terminal", "prd-ATM", "prd-A-Treasury", "prd-BancoCore", "prd-Bank2b",
    "prd-Bankway", "prd-BoCard", "prd-BPMS-Finance-infrastucture", "prd-BPMSProcess",
    "prd-BusinessAutomation", "prd-BusinessFramework", "prd-CardHost", "prd-CardTest",
    "prd-Chakavak", "prd-ChannelManager", "prd-Chapar", "prd-Cheque", "prd-ChequeTest",
    "prd-Commercial", "prd-CoreExtension", "prd-CoreTest", "prd-CoreUI", "prd-CreditHost",
    "prd-Device", "prd-devops", "prd-EGover", "prd-EGoverTest", "PRD-ExirUI",
    "prd-FinanceTest", "prd-Framework", "prd-FrameworkTest", "prd-Fundamental",
    "prd-InterbankTest", "prd-InternetBank", "prd-LiteCore", "prd-Loan1", "prd-Loan2",
    "prd-LoanTest", "prd-LoanUI", "prd-Mahan-ML", "prd-MobileBank", "prd-MobilFaceet",
    "prd-ModernBo", "prd-ModernTest", "prd-Paya", "prd-Retail", "prd-Satna",
    "prd-Sepanta", "prd-Sipa", "prd-SwiftMXProject", "prd-TelephoneBank",
    "prd-TerminalTest", "prd-TMS", "prd-Trade", "prd-TradeWebProject", "prd-Yek"
]

def find_remote_list_file(folder_name):
    folder_path = os.path.join(base_path, folder_name)
    if not os.path.exists(folder_path): return None
    for file in os.listdir(folder_path):
        if file.lower().endswith((".xlsx", ".xlsm")):
            if os.path.splitext(file)[0].endswith("-Remote-List"):
                return os.path.join(folder_path, file)
    return None

def create_or_load_combine_template():
    """Create a formatted Combine template if it doesn't exist, or load existing one."""
    if not os.path.exists(file_combine_template):
        # Create a new template with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Combine"
        
        # Add headers with formatting
        headers = ["لاین", "تیم", "نام و نام خانوادگی", "علت"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        
        wb.save(file_combine_template)
        wb.close()
        print("✓ Created new Combine template with formatting")
    
    return file_combine_template

# ----------------------------------------------------------------------
# 1. Combine - Preserve format and style
# ----------------------------------------------------------------------
def combine_excel_files():
    """Combine data while preserving the format of the Combine file."""
    try:
        # Ensure template exists
        template_path = create_or_load_combine_template()
        
        # Load the template to preserve formatting
        combined_wb = load_workbook(template_path)
        ws = combined_wb.active
        
        # Clear existing data (keep headers and formatting)
        if ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                for col in range(1, 5):  # Columns A-D
                    ws.cell(row=row, column=col).value = None
        
        # Start from row 2 (after header)
        current_row = 2
        
        for folder_name in folders:
            file_path = find_remote_list_file(folder_name)
            if not file_path:
                continue
            
            try:
                wb = load_workbook(file_path, data_only=True)
                
                # Try to find the "List" sheet, fallback to active sheet
                if "List" in wb.sheetnames:
                    sheet = wb["List"]
                else:
                    sheet = wb.active
                    print(f"⚠ 'List' sheet not found in {folder_name}, using active sheet")
                
                # Get data from each row
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(row):  # If row has any data
                        # Apply the same formatting as header row for data cells
                        for col_idx, value in enumerate(row[:4], 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=str(value) if value else "")
                            cell.font = Font(name='Calibri', size=11)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                        
                        current_row += 1
                
                wb.close()
                print(f"✓ Processed: {folder_name}")
                
            except Exception as e:
                print(f"✗ Error processing {folder_name}: {e}")
                continue
        
        # Save the combined file with preserved formatting
        combined_wb.save(file_combine)
        combined_wb.close()
        print(f"✓ Combined file saved: {file_combine}")
        return True
        
    except Exception as e:
        print(f"✗ Error in combine_excel_files: {e}")
        return False

# ----------------------------------------------------------------------
# 2. Email
# ----------------------------------------------------------------------
def send_outlook_email():
    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = "SAP@tosan-solutions.com"
        # Adding the CC line here
        mail.CC = "APM-Core@tosan-solutions.com; APM-Card@tosan-solutions.com; APM-Finance@tosan-solutions.com; APM-Modern@tosan-solutions.com" 
        mail.Subject = "لیست دورکاری"
        
        # HTML with RTL direction, B Nazanin font, and bold styling
        mail.HTMLBody = """
        <html dir="rtl">
        <head>
        <meta charset="UTF-8">
        <style>
            body {
                font-family: 'B Nazanin', 'B Nazanin Bold', 'Tahoma', sans-serif;
                font-size: 14pt;
                direction: rtl;
            }
            p, div, span, li {
                font-family: 'B Nazanin', 'B Nazanin Bold', 'Tahoma', sans-serif;
                font-weight: normal;
            }
            b, strong, h1, h2, h3, h4, h5, h6 {
                font-family: 'B Nazanin', 'B Nazanin Bold', 'Tahoma', sans-serif;
                font-weight: bold;
            }
        </style>
        </head>
        <body>
        <p>با سلام و درود</p>
        <p>لیست دورکاری امروز تقدیم حضورتان می گردد</p>
        <p>متشکرم.</p>
        </body>
        </html>
        """
        
        mail.Attachments.Add(file_combine)
        mail.Send()
        return True
    except Exception as e:
        print(f"Email Error: {e}")
        return False

# ----------------------------------------------------------------------
# 3. Cleanup - Clear team files and combine file
# ----------------------------------------------------------------------
def clear_team_file_content(file_path):
    """Clears columns C and D in team Excel files."""
    try:
        wb = load_workbook(file_path)
        
        # Find the correct sheet
        if "List" in wb.sheetnames:
            ws = wb["List"]
        else:
            ws = wb.active
        
        # Clear columns C and D (starting from row 2)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=3).value = None  # Clear C
            ws.cell(row=row, column=4).value = None  # Clear D
        
        wb.save(file_path)
        wb.close()
        print(f"   ✓ Cleared columns C-D: {os.path.basename(file_path)}")
    except Exception as e:
        print(f"   ✗ Failed to clear {file_path}: {e}")

def clear_combine_file_content(file_path):
    """Clears columns A, B, C, and D in the Combine file (preserving header)."""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Clear columns A-D (starting from row 2, keep header)
        for row in range(2, ws.max_row + 1):
            for col in range(1, 5):  # Columns A, B, C, D
                ws.cell(row=row, column=col).value = None
        
        wb.save(file_path)
        wb.close()
        print(f"   ✓ Cleared columns A-D in Combine file")
    except Exception as e:
        print(f"   ✗ Failed to clear Combine file: {e}")

# ----------------------------------------------------------------------
# Main Execution
# ----------------------------------------------------------------------
if __name__ == "__main__":
    print("📊 Starting Remote List Processing...")
    print("=" * 50)
    
    if combine_excel_files():
        print("\n📧 Sending email...")
        if send_outlook_email():
            print("✓ Email sent successfully")
            
            print("\n🧹 Starting Cleanup...")
            print("-" * 30)
            
            # Clear team files (columns C and D)
            print("Clearing team files (columns C-D):")
            for folder in folders:
                f = find_remote_list_file(folder)
                if f:
                    clear_team_file_content(f)
            
            # Clear combine file (columns A, B, C, D)
            print("\nClearing Combine file (columns A-D):")
            if os.path.exists(file_combine):
                clear_combine_file_content(file_combine)
            
            print("\n" + "=" * 50)
            print("✨ Process completed successfully!")
        else:
            print("\n❌ Email failed to send.")
    else:
        print("\n❌ Failed to create combined file.")