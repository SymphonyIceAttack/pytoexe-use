import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import sys
from docx import Document
from openpyxl import load_workbook

class PartUpdaterApp:
    def __init__(self, root):
        self.root = root
        root.title("Обновление наименований запчастей из Excel")
        root.geometry("620x520")
        root.resizable(False, False)

        # Переменные для путей
        self.word_path = tk.StringVar()
        self.excel_path = tk.StringVar()

        # Столбцы в Excel (по умолчанию артикул - A, наименование - B)
        self.excel_article_col = tk.StringVar(value="1")
        self.excel_name_col = tk.StringVar(value="2")

        # Столбцы в Word (по умолчанию артикул - 1, наименование - 2)
        self.word_article_col = tk.StringVar(value="1")
        self.word_name_col = tk.StringVar(value="2")

        # Флаг, пропускать ли заголовок в Word (первая строка)
        self.skip_header = tk.BooleanVar(value=True)

        # Создаём виджеты
        self.create_widgets()

    def create_widgets(self):
        # Рамка для выбора файлов
        file_frame = tk.LabelFrame(self.root, text="Файлы", padx=10, pady=10)
        file_frame.pack(fill="x", padx=10, pady=5)

        # Word файл
        tk.Label(file_frame, text="Документ Word:").grid(row=0, column=0, sticky="w")
        tk.Entry(file_frame, textvariable=self.word_path, width=50).grid(row=0, column=1, padx=5)
        tk.Button(file_frame, text="Обзор...", command=self.browse_word).grid(row=0, column=2)

        # Excel файл
        tk.Label(file_frame, text="Файл Excel:").grid(row=1, column=0, sticky="w", pady=(5,0))
        tk.Entry(file_frame, textvariable=self.excel_path, width=50).grid(row=1, column=1, padx=5, pady=(5,0))
        tk.Button(file_frame, text="Обзор...", command=self.browse_excel).grid(row=1, column=2, pady=(5,0))

        # Рамка для настроек столбцов
        col_frame = tk.LabelFrame(self.root, text="Настройка столбцов", padx=10, pady=10)
        col_frame.pack(fill="x", padx=10, pady=5)

        # Excel столбцы
        tk.Label(col_frame, text="Excel: артикул (номер столбца)").grid(row=0, column=0, sticky="w")
        tk.Entry(col_frame, textvariable=self.excel_article_col, width=5).grid(row=0, column=1, padx=5)
        tk.Label(col_frame, text="наименование (номер)").grid(row=0, column=2, padx=(15,0))
        tk.Entry(col_frame, textvariable=self.excel_name_col, width=5).grid(row=0, column=3)

        # Word столбцы
        tk.Label(col_frame, text="Word: артикул (номер столбца)").grid(row=1, column=0, sticky="w", pady=(5,0))
        tk.Entry(col_frame, textvariable=self.word_article_col, width=5).grid(row=1, column=1, padx=5, pady=(5,0))
        tk.Label(col_frame, text="наименование (номер)").grid(row=1, column=2, padx=(15,0), pady=(5,0))
        tk.Entry(col_frame, textvariable=self.word_name_col, width=5).grid(row=1, column=3, pady=(5,0))

        # Чекбокс для пропуска заголовка
        tk.Checkbutton(col_frame, text="Пропустить первую строку в таблицах Word (заголовок)",
                       variable=self.skip_header).grid(row=2, column=0, columnspan=4, sticky="w", pady=(10,0))

        # Кнопка запуска
        self.run_btn = tk.Button(self.root, text="Обновить наименования", command=self.run_update,
                                 bg="#4CAF50", fg="white", font=("Arial", 12), height=1, width=25)
        self.run_btn.pack(pady=15)

        # Текстовое поле для вывода лога
        self.log = scrolledtext.ScrolledText(self.root, height=12, state='normal', wrap=tk.WORD)
        self.log.pack(fill="both", padx=10, pady=5, expand=True)
        self.log.insert(tk.END, "Готов к работе.\n")

    def browse_word(self):
        path = filedialog.askopenfilename(filetypes=[("Word documents", "*.docx")])
        if path:
            self.word_path.set(path)
            self.log_insert(f"Выбран Word: {os.path.basename(path)}\n")

    def browse_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.excel_path.set(path)
            self.log_insert(f"Выбран Excel: {os.path.basename(path)}\n")

    def log_insert(self, text):
        self.log.insert(tk.END, text)
        self.log.see(tk.END)
        self.root.update_idletasks()

    def run_update(self):
        # Проверка наличия файлов
        word_file = self.word_path.get().strip()
        excel_file = self.excel_path.get().strip()
        if not word_file or not excel_file:
            messagebox.showerror("Ошибка", "Выберите оба файла.")
            return
        if not os.path.exists(word_file):
            messagebox.showerror("Ошибка", "Файл Word не найден.")
            return
        if not os.path.exists(excel_file):
            messagebox.showerror("Ошибка", "Файл Excel не найден.")
            return

        # Считываем номера столбцов
        try:
            ex_art = int(self.excel_article_col.get())
            ex_name = int(self.excel_name_col.get())
            w_art = int(self.word_article_col.get())
            w_name = int(self.word_name_col.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Номера столбцов должны быть целыми числами.")
            return

        if ex_art < 1 or ex_name < 1 or w_art < 1 or w_name < 1:
            messagebox.showerror("Ошибка", "Номера столбцов должны быть >= 1.")
            return

        skip_first = self.skip_header.get()

        self.log_insert("\n--- Запуск обновления ---\n")
        self.run_btn.config(state="disabled")
        self.root.update_idletasks()

        try:
            # 1. Чтение Excel
            self.log_insert("Чтение Excel...\n")
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active

            # Строим словарь замен
            mapping = {}
            for row in ws.iter_rows(min_row=1, values_only=True):
                if len(row) < max(ex_art, ex_name):
                    continue
                art = row[ex_art - 1]
                name = row[ex_name - 1]
                if art is not None and name is not None:
                    art_str = str(art).strip()
                    name_str = str(name).strip()
                    if art_str and name_str:
                        mapping[art_str] = name_str

            if not mapping:
                messagebox.showwarning("Предупреждение", "В Excel не найдено пар артикул-наименование.")
                self.log_insert("Словарь пуст. Отмена.\n")
                self.run_btn.config(state="normal")
                return

            self.log_insert(f"Загружено {len(mapping)} записей из Excel.\n")

            # 2. Обработка Word
            self.log_insert("Открытие Word документа...\n")
            doc = Document(word_file)
            replaced_count = 0
            table_count = 0

            for tbl in doc.tables:
                table_count += 1
                # Проверяем, есть ли нужные столбцы
                if len(tbl.columns) < max(w_art, w_name):
                    self.log_insert(f"Таблица {table_count}: пропущена (мало столбцов).\n")
                    continue

                rows = tbl.rows
                start_row = 1 if skip_first else 0  # 0-индексация
                for i in range(start_row, len(rows)):
                    try:
                        cell_art = rows[i].cells[w_art - 1]
                        cell_name = rows[i].cells[w_name - 1]
                    except IndexError:
                        continue

                    art_text = cell_art.text.strip()
                    if art_text in mapping:
                        new_name = mapping[art_text]
                        # Заменяем содержимое ячейки наименования
                        cell_name.text = new_name
                        replaced_count += 1
                        self.log_insert(f"Заменено: {art_text} -> {new_name}\n")

            self.log_insert(f"Обработано таблиц: {table_count}\n")
            self.log_insert(f"Всего замен: {replaced_count}\n")

            # Сохраняем новый документ с суффиксом _updated
            base, ext = os.path.splitext(word_file)
            new_file = base + "_updated" + ext
            doc.save(new_file)
            self.log_insert(f"Документ сохранён как: {new_file}\n")
            messagebox.showinfo("Готово", f"Обновление завершено!\nСохранено: {new_file}\nЗамен: {replaced_count}")

        except Exception as e:
            self.log_insert(f"ОШИБКА: {str(e)}\n")
            messagebox.showerror("Ошибка", f"Произошла ошибка:\n{str(e)}")
        finally:
            self.run_btn.config(state="normal")
            self.log_insert("--- Конец ---\n")

if __name__ == "__main__":
    root = tk.Tk()
    app = PartUpdaterApp(root)
    root.mainloop()