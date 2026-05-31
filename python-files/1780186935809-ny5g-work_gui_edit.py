#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工时编辑工具 - 最终稳定版（K类工单加载不依赖名称）
"""

import os
import sys
import csv
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import subprocess
import threading

# ========== 常量 ==========
EDIT_HEADERS = ['ID.', 'OP代码', '部门代码', '部门名称', '项目代码', '派工单', '作业内容', '工时合计']
DATA_SOURCE_HEADERS = ['进度状态', '派发状态', '结算批次', '部门', '船号', '派工单编码', '派工单名称', '计划工时', '施工班组']

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

def get_tomorrow_str():
    return (datetime.now() + timedelta(days=1)).strftime("%m.%d")

def check_data_source(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        for col_idx, exp in enumerate(DATA_SOURCE_HEADERS, 1):
            val = ws.cell(row=2, column=col_idx).value
            if not val or val.strip() != exp:
                val = ws.cell(row=1, column=col_idx).value
                if not val or val.strip() != exp:
                    return False
        return True
    except:
        return False

def validate_edit_file(filepath):
    tomorrow = get_tomorrow_str()
    if tomorrow not in Path(filepath).stem:
        return False, f"文件名必须包含明天日期 {tomorrow}"
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        for col_idx, exp in enumerate(EDIT_HEADERS, 1):
            val = ws.cell(row=2, column=col_idx).value
            if not val or val.strip() != exp:
                return False, f"第2行第{col_idx}列应为'{exp}'"
        has_person = False
        col = 9
        while col <= ws.max_column:
            if ws.cell(row=1, column=col).value:
                has_person = True
                break
            col += 1
        if not has_person:
            return False, "第1行I列起未找到人员姓名"
        return True, ""
    except Exception as e:
        return False, str(e)

def load_people(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    people = []
    col = 9
    while True:
        val = ws.cell(row=1, column=col).value
        if not val:
            break
        people.append(str(val).strip())
        col += 1
    wb.close()
    return people

def load_data_source_with_progress(filepath, progress_callback):
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    header_row = 2
    if ws.cell(row=2, column=1).value != DATA_SOURCE_HEADERS[0]:
        header_row = 1
    total_rows = sum(1 for _ in ws.iter_rows(min_row=header_row+1))
    if total_rows == 0:
        total_rows = 1
    rows = ws.iter_rows(min_row=header_row+1, values_only=True)
    records = []
    ship_idx = defaultdict(list)
    processed = 0
    for idx, row in enumerate(rows):
        if not row[0]:
            continue
        rec = {
            '进度状态': row[0],
            '派发状态': row[1],
            '结算批次': row[2],
            '部门': row[3],
            '船号': row[4],
            '派工单编码': str(row[5]) if row[5] else '',
            '派工单名称': str(row[6]) if row[6] else '',
            '计划工时': row[7],
            '施工班组': str(row[8]) if row[8] else '',
        }
        records.append(rec)
        ship = rec['船号']
        if ship:
            ship_idx[ship].append(idx)
        processed += 1
        if processed % 1000 == 0:
            progress_callback(processed / total_rows * 100)
    progress_callback(100)
    wb.close()
    return records, ship_idx

def load_teams_from_csv(filepath):
    for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                reader = csv.reader(f)
                teams = []
                for row in reader:
                    if len(row) >= 3:
                        w = row[0].strip()
                        c = row[1].strip()
                        n = row[2].strip()
                        if w and c and n:
                            teams.append((w, c, n))
                if teams:
                    return teams
        except:
            continue
    return []

def load_teams_from_excel(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        teams = []
        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 3:
                w = str(row[0]).strip() if row[0] else ''
                c = str(row[1]).strip() if row[1] else ''
                n = str(row[2]).strip() if row[2] else ''
                if w and c and n:
                    teams.append((w, c, n))
        wb.close()
        return teams
    except:
        return []

def get_team_info(root):
    base = get_base_dir()
    all_teams = []
    for fname in ['info.csv', 'info.xlsx']:
        path = os.path.join(base, fname)
        if os.path.exists(path):
            if fname.endswith('.csv'):
                teams = load_teams_from_csv(path)
            else:
                teams = load_teams_from_excel(path)
            if teams:
                all_teams = teams
                break
    if not all_teams:
        ans = messagebox.askyesno("班组文件", "未自动找到班组配置文件。\n是否手动选择一个包含班组列表的文件？")
        if ans:
            path = filedialog.askopenfilename(title="选择班组文件", filetypes=[("CSV/Excel", "*.csv *.xlsx")])
            if path:
                if path.endswith('.csv'):
                    all_teams = load_teams_from_csv(path)
                else:
                    all_teams = load_teams_from_excel(path)
        if not all_teams:
            team = manual_team_dialog(root)
            if team:
                all_teams = [team]
            else:
                return None, None, None
    code_to_name = {t[1]: t[2] for t in all_teams}
    name_to_code = {t[2]: t[1] for t in all_teams}
    if len(all_teams) == 1:
        selected = all_teams[0]
    else:
        selected = select_team_dialog(root, all_teams)
        if selected is None:
            return None, None, None
    return selected, code_to_name, name_to_code

def select_team_dialog(root, teams):
    team_names = [f"{t[0]} {t[1]} {t[2]}" for t in teams]
    choice = None
    dlg = tk.Toplevel(root)
    dlg.title("选择班组")
    dlg.geometry("400x300")
    dlg.transient(root)
    dlg.grab_set()
    tk.Label(dlg, text="请选择您的班组：").pack(pady=5)
    lb = tk.Listbox(dlg)
    lb.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    for name in team_names:
        lb.insert(tk.END, name)
    def on_ok():
        nonlocal choice
        sel = lb.curselection()
        if sel:
            choice = teams[sel[0]]
        dlg.destroy()
    tk.Button(dlg, text="确定", command=on_ok).pack(pady=5)
    root.wait_window(dlg)
    return choice

def manual_team_dialog(root):
    dlg = tk.Toplevel(root)
    dlg.title("手动输入班组")
    dlg.geometry("400x200")
    dlg.transient(root)
    dlg.grab_set()
    tk.Label(dlg, text="请手动输入您的班组信息：").pack(pady=5)
    tk.Label(dlg, text="车间：").pack()
    e1 = tk.Entry(dlg, width=30)
    e1.pack(pady=2)
    tk.Label(dlg, text="班组代码：").pack()
    e2 = tk.Entry(dlg, width=30)
    e2.pack(pady=2)
    tk.Label(dlg, text="班组名称：").pack()
    e3 = tk.Entry(dlg, width=30)
    e3.pack(pady=2)
    result = [None]
    def on_ok():
        w = e1.get().strip()
        c = e2.get().strip()
        n = e3.get().strip()
        if w and c and n:
            result[0] = (w, c, n)
            dlg.destroy()
        else:
            messagebox.showwarning("警告", "三项都不能为空")
    tk.Button(dlg, text="确定", command=on_ok).pack(pady=10)
    root.wait_window(dlg)
    return result[0]

def match_in_ship_index(ship_idx, records, ship_no, query, match_col):
    if not query:
        return []
    q = query.lower()
    matches = []
    for idx in ship_idx.get(ship_no, []):
        rec = records[idx]
        field = rec['派工单编码'] if match_col == 'F' else rec['派工单名称']
        if q in field.lower():
            display = f"{rec['部门']} {rec['施工班组']} {rec['船号']} {rec['派工单编码']} {rec['派工单名称']}"
            matches.append((idx, display))
    def score(m):
        text = m[1].lower()
        pos = text.find(q)
        return (len(q)/len(text), -pos)
    matches.sort(key=score, reverse=True)
    return matches

def sort_matches_by_workshop(matches, records, workshop_name):
    def key_func(m):
        idx, _ = m
        rec = records[idx]
        if rec.get('部门', '') == workshop_name:
            return 0
        else:
            return 1
    matches.sort(key=key_func)
    return matches

def get_op_code(team_code, rec):
    if rec.get('施工班组', '') == team_code:
        return "A111"
    else:
        return "A112"

class WorkEditApp:
    def __init__(self, root):
        self.root = root
        self.root.title("工时编辑工具")
        self.root.geometry("1400x800")
        self.root.minsize(1000, 600)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.people = []
        self.records = []
        self.ship_idx = {}
        self.team = None
        self.code_to_name = {}
        self.name_to_code = {}
        self.edit_path = ""
        self.tasks = defaultdict(list)
        self.person_idx = 0
        self.task_idx = 0
        self.mode = 'person'
        self._loading = False

        self.create_widgets()
        self.root.after(100, self.init_flow)

    def create_widgets(self):
        main_pane = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        left_frame = ttk.LabelFrame(main_pane, text="人员列表")
        main_pane.add(left_frame, weight=1)
        self.person_lb = tk.Listbox(left_frame, font=("Consolas", 10), selectmode=tk.SINGLE)
        self.person_lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(left_frame, orient=tk.VERTICAL, command=self.person_lb.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.person_lb.config(yscrollcommand=sb.set)

        self.person_lb.bind('<Up>', self.on_up_person)
        self.person_lb.bind('<Down>', self.on_down_person)
        self.person_lb.bind('<ButtonRelease-1>', self.on_click_person)
        self.person_lb.bind('<Return>', lambda e: self.enter_person())

        right_frame = ttk.LabelFrame(main_pane, text="工单列表（当前人员）")
        main_pane.add(right_frame, weight=3)
        columns = ('ship', 'op', 'dept', 'code', 'name', 'hours')
        self.task_tree = ttk.Treeview(right_frame, columns=columns, show='headings', height=20)
        self.task_tree.heading('ship', text='船号')
        self.task_tree.heading('op', text='OP代码')
        self.task_tree.heading('dept', text='部门')
        self.task_tree.heading('code', text='派工单代码')
        self.task_tree.heading('name', text='派工单名称')
        self.task_tree.heading('hours', text='工时')
        self.task_tree.column('ship', width=90, stretch=True)
        self.task_tree.column('op', width=80, stretch=True)
        self.task_tree.column('dept', width=180, stretch=True)
        self.task_tree.column('code', width=110, stretch=True)
        self.task_tree.column('name', width=250, stretch=True)
        self.task_tree.column('hours', width=80, stretch=True)

        vsb = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=self.task_tree.yview)
        hsb = ttk.Scrollbar(right_frame, orient=tk.HORIZONTAL, command=self.task_tree.xview)
        self.task_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.task_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        right_frame.grid_rowconfigure(0, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        self.task_tree.bind('<<TreeviewSelect>>', self.on_task_select)
        self.task_tree.bind('<Up>', lambda e: self.move_task(-1))
        self.task_tree.bind('<Down>', lambda e: self.move_task(1))
        self.task_tree.bind('<Return>', lambda e: self.modify_task())

        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(fill=tk.X, pady=10, padx=5)
        ttk.Button(btn_frame, text="➕ 新增工单", command=self.add_task).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="➖ 删除工单", command=self.del_task).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="✳️ 修改工时", command=self.modify_task).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📊 工时统计", command=self.show_stats).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="💾 保存并导出", command=self.save_and_export).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🔄 重选班组", command=self.restart_for_team).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🚪 放弃退出", command=self.quit_app).pack(side=tk.LEFT, padx=2)

        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, pady=(0,5), padx=5)
        self.status = ttk.Label(status_frame, text="就绪", foreground="blue")
        self.status.pack(side=tk.LEFT)
        self.status.bind('<Double-Button-1>', lambda e: self.show_about())
        self.progress_var = tk.IntVar()
        self.progress_bar = ttk.Progressbar(status_frame, variable=self.progress_var, length=200, mode='determinate')
        self.progress_bar.pack(side=tk.RIGHT, padx=5)

        self.root.bind('<plus>', lambda e: self.add_task())
        self.root.bind('<KP_Add>', lambda e: self.add_task())
        self.root.bind('<minus>', lambda e: self.del_task())
        self.root.bind('<KP_Subtract>', lambda e: self.del_task())
        self.root.bind('<asterisk>', lambda e: self.modify_task())
        self.root.bind('<KP_Multiply>', lambda e: self.modify_task())
        self.root.bind('<Escape>', lambda e: self.escape())

    def show_about(self):
        about = tk.Toplevel(self.root)
        about.title("关于")
        about.geometry("300x200")
        about.transient(self.root)
        about.grab_set()
        about.focus_force()
        about.lift()
        text = "\n\nYSS预录速录工具\nVer. 1.0\nSalute for you from fanbo in silpway workshop, YAMIC"
        tk.Label(about, text=text, justify=tk.LEFT, padx=20, pady=20).pack()
        tk.Button(about, text="关闭", command=about.destroy).pack(pady=10)
        about.bind('<Escape>', lambda e: about.destroy())

    def restart_for_team(self):
        if not messagebox.askyesno("确认", "重新选择班组将放弃当前所有未保存的更改，并重新选择待编辑文件。\n确定继续吗？"):
            return
        self.tasks.clear()
        self.people = []
        self.edit_path = ""
        team_info, code_to_name, name_to_code = get_team_info(self.root)
        if team_info is None:
            messagebox.showerror("错误", "无法获取班组信息，操作取消")
            return
        self.team = team_info
        self.code_to_name = code_to_name
        self.name_to_code = name_to_code
        messagebox.showinfo("班组", f"当前班组：{team_info[0]} {team_info[1]} {team_info[2]}")
        while True:
            path = filedialog.askopenfilename(title="请选择待编辑的xlsx文件", filetypes=[("Excel", "*.xlsx")])
            if not path:
                if messagebox.askyesno("退出", "未选择待编辑文件，是否放弃操作？"):
                    return
                else:
                    continue
            ok, msg = validate_edit_file(path)
            if not ok:
                messagebox.showerror("校验失败", msg)
                continue
            self.edit_path = path
            break
        self.people = load_people(self.edit_path)
        if not self.people:
            messagebox.showerror("错误", "未找到人员姓名列（第1行I列及之后）")
            return
        self.load_tasks_from_file()
        self.refresh_person_display()
        self.refresh_task()
        self.status.config(text="已重新加载文件和班组")

    # ---------- 人员列表手动控制 ----------
    def on_up_person(self, event):
        self.move_person(-1)
        return "break"

    def on_down_person(self, event):
        self.move_person(1)
        return "break"

    def on_click_person(self, event):
        idx = self.person_lb.nearest(event.y)
        if 0 <= idx < len(self.people):
            self.person_idx = idx
            self.mode = 'person'
            self._set_person_selection()
            self.refresh_task()
            self.status.config(text=f"选中：{self.people[self.person_idx]}")

    def _set_person_selection(self):
        self.person_lb.selection_clear(0, tk.END)
        self.person_lb.selection_set(self.person_idx)
        self.person_lb.see(self.person_idx)
        self.person_lb.activate(self.person_idx)
        self.person_lb.focus_set()

    def move_person(self, delta):
        new = self.person_idx + delta
        if 0 <= new < len(self.people):
            self.person_idx = new
            self.mode = 'person'
            self._set_person_selection()
            self.refresh_task()
            self.status.config(text=f"选中：{self.people[self.person_idx]}")

    # ---------- 初始化 ----------
    def init_flow(self):
        while True:
            path = filedialog.askopenfilename(title="选择数据源xlsx", filetypes=[("Excel", "*.xlsx")])
            if not path:
                if messagebox.askyesno("退出", "未选择数据源，是否退出？"):
                    self.root.quit()
                    return
                continue
            if not check_data_source(path):
                messagebox.showerror("错误", "数据源格式不正确")
                continue
            if (datetime.now().date() - datetime.fromtimestamp(os.path.getmtime(path)).date()).days > 7:
                if not messagebox.askyesno("警告", "数据源已超过7天未更新，确定继续吗？"):
                    continue
            self.data_src = path
            break

        self.progress_var.set(0)
        self.progress_bar.pack(side=tk.RIGHT, padx=5)
        self.status.config(text="正在加载数据源...")
        self.root.update_idletasks()

        def load_thread():
            records, ship_idx = load_data_source_with_progress(self.data_src, self.update_progress)
            self.root.after(0, lambda: self.on_data_loaded(records, ship_idx))

        threading.Thread(target=load_thread, daemon=True).start()

    def update_progress(self, percent):
        self.root.after(0, lambda: self.progress_var.set(percent))

    def on_data_loaded(self, records, ship_idx):
        self.records = records
        self.ship_idx = ship_idx
        self.progress_var.set(100)
        self.status.config(text="数据源加载完成")
        self.root.update_idletasks()
        self.select_edit_file()

    def select_edit_file(self):
        while True:
            path = filedialog.askopenfilename(title="选择待编辑xlsx", filetypes=[("Excel", "*.xlsx")])
            if not path:
                if messagebox.askyesno("退出", "未选择待编辑文件，是否退出？"):
                    self.root.quit()
                    return
                continue
            ok, msg = validate_edit_file(path)
            if not ok:
                messagebox.showerror("校验失败", msg)
                continue
            self.edit_path = path
            break

        self.people = load_people(self.edit_path)
        if not self.people:
            messagebox.showerror("错误", "未找到人员姓名（第1行I列起）")
            self.root.quit()
            return

        team_info, code_to_name, name_to_code = get_team_info(self.root)
        if team_info is None:
            messagebox.showerror("错误", "无法获取班组信息，程序退出")
            self.root.quit()
            return
        self.team = team_info
        self.code_to_name = code_to_name
        self.name_to_code = name_to_code
        messagebox.showinfo("班组", f"当前班组：{team_info[0]} {team_info[1]} {team_info[2]}")

        self.load_tasks_from_file()
        self.refresh_person_display()
        self.refresh_task()
        self.progress_bar.pack_forget()
        self.status.config(text="就绪，可使用鼠标或快捷键")

    # ========== 关键修改：加载 K 类工单时不依赖名称，只看 OP 代码 ==========
    def load_tasks_from_file(self):
        wb = load_workbook(self.edit_path)
        ws = wb.active
        person_cols = {}
        col = 9
        for name in self.people:
            person_cols[name] = col
            col += 1
        self.tasks.clear()
        for row in range(3, ws.max_row+1):
            op = ws.cell(row=row, column=2).value
            code = ws.cell(row=row, column=6).value
            name = ws.cell(row=row, column=7).value
            # 如果 OP 是休假代码，强制规范化
            if op == 'K121':
                code = 'K121'
                name = '公休'
            elif op == 'K122':
                code = 'K122'
                name = '事假'
            elif not code:   # 非休假且无派工单代码则跳过
                continue
            dept = ws.cell(row=row, column=4).value or ''
            # 尝试从数据源查找（非休假工单）
            src = None
            if op not in ('K121', 'K122'):
                for rec in self.records:
                    if rec['派工单编码'] == code and rec['派工单名称'] == name:
                        src = rec
                        break
            if src is None:
                src = {'船号': '', '部门': dept, '施工班组': ''}
            # 读取各人员工时
            for pname, pcol in person_cols.items():
                hrs = ws.cell(row=row, column=pcol).value
                if hrs and isinstance(hrs, (int, float)) and hrs != 0:
                    pi = self.people.index(pname)
                    # 查找是否已存在相同工单
                    existing = None
                    for t in self.tasks[pi]:
                        if t['code'] == code and t['name'] == name:
                            existing = t
                            break
                    if existing:
                        existing['hours'] += hrs
                    else:
                        team_name = src.get('施工班组', '')
                        target_code = self.name_to_code.get(team_name, team_name)
                        target_name = team_name
                        self.tasks[pi].append({
                            'ship': src['船号'],
                            'code': code,
                            'name': name,
                            'hours': hrs,
                            'op': op,
                            'dept': src['部门'],
                            'team': src['施工班组'],
                            'target_code': target_code,
                            'target_name': target_name
                        })
        wb.close()

    def save_tasks_to_file(self):
        wb = load_workbook(self.edit_path)
        ws = wb.active
        person_cols = {}
        col = 9
        for name in self.people:
            person_cols[name] = col
            col += 1
        if ws.max_row >= 3:
            ws.delete_rows(3, ws.max_row - 2)
        orders = {}
        for tasks in self.tasks.values():
            for t in tasks:
                if t['op'].startswith('K'):
                    if t['code'] == 'K121':
                        norm_name = '公休'
                    elif t['code'] == 'K122':
                        norm_name = '事假'
                    else:
                        norm_name = t['name']
                    key = ('', norm_name)
                else:
                    key = (t['code'], t['name'])
                if key not in orders:
                    orders[key] = {
                        'code': key[0],
                        'name': key[1],
                        'op': t['op'],
                        'ship': t['ship'],
                        'target_code': t.get('target_code', ''),
                        'target_name': t.get('target_name', ''),
                        'total': 0
                    }
                orders[key]['total'] += t['hours']
        rows = []
        for o in orders.values():
            if o['op'].startswith('K'):
                rows.append({
                    'A': None,
                    'B': o['op'],
                    'C': '',
                    'D': '',
                    'E': '',
                    'F': '',
                    'G': o['name'],
                    'H': o['total'],
                })
            elif o['op'] == 'A111':
                rows.append({
                    'A': None,
                    'B': o['op'],
                    'C': '',
                    'D': '',
                    'E': o['ship'],
                    'F': o['code'],
                    'G': o['name'],
                    'H': o['total'],
                })
            elif o['op'] == 'A112':
                rows.append({
                    'A': None,
                    'B': o['op'],
                    'C': o['target_code'],
                    'D': o['target_name'],
                    'E': o['ship'],
                    'F': o['code'],
                    'G': o['name'],
                    'H': o['total'],
                })
            else:
                rows.append({
                    'A': None,
                    'B': o['op'],
                    'C': self.team[1],
                    'D': self.team[2],
                    'E': o['ship'],
                    'F': o['code'],
                    'G': o['name'],
                    'H': o['total'],
                })
        def order_key(row):
            op = row['B']
            if op == 'A111':
                return 0
            if op == 'A112':
                return 1
            if op == 'K121':
                return 2
            return 3
        rows.sort(key=order_key)
        for idx, row in enumerate(rows, 1):
            row['A'] = idx
            new_row = ws.max_row + 1
            for col_idx, col_letter in enumerate(['A','B','C','D','E','F','G','H'], 1):
                cell = ws.cell(row=new_row, column=col_idx, value=row[col_letter])
                cell.border = THIN_BORDER
        row_map = {}
        for r in range(3, ws.max_row+1):
            f = ws.cell(row=r, column=6).value
            g = ws.cell(row=r, column=7).value
            if f is None:
                f = ''
            if g is None:
                g = ''
            row_map[(f, g)] = r
        for name, c in person_cols.items():
            for r in range(3, ws.max_row+1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.border = THIN_BORDER
        for pi, tasks in self.tasks.items():
            pname = self.people[pi]
            col = person_cols[pname]
            for t in tasks:
                if t['op'].startswith('K'):
                    if t['code'] == 'K121':
                        norm_name = '公休'
                    elif t['code'] == 'K122':
                        norm_name = '事假'
                    else:
                        norm_name = t['name']
                    key = ('', norm_name)
                else:
                    key = (t['code'], t['name'])
                if key in row_map:
                    cell = ws.cell(row=row_map[key], column=col)
                    cell.value = t['hours']
                    cell.border = THIN_BORDER
                else:
                    alt_key = (t['code'], t['name'])
                    if alt_key in row_map:
                        cell = ws.cell(row=row_map[alt_key], column=col)
                        cell.value = t['hours']
                        cell.border = THIN_BORDER
        wb.save(self.edit_path)
        wb.close()

    def can_add_task(self, person_idx, new_task):
        existing = self.tasks[person_idx]
        is_new_holiday = new_task['op'] in ('K121', 'K122')
        has_holiday = any(t['op'] in ('K121', 'K122') for t in existing)
        if is_new_holiday and existing:
            return False, "已有其他工单，不能添加公休/事假"
        if has_holiday and not is_new_holiday:
            return False, "该人员已有公休/事假，不能添加其他工单"
        return True, ""

    def get_work_total(self, person_idx):
        return sum(t['hours'] for t in self.tasks[person_idx] if t['op'] not in ('K121', 'K122'))

    def refresh_person_display(self):
        self._loading = True
        old_idx = self.person_idx
        self.person_lb.delete(0, tk.END)
        for i in range(len(self.people)):
            total = self.get_work_total(i)
            has_holiday = any(t['op'] in ('K121', 'K122') for t in self.tasks[i])
            if has_holiday:
                text = f"{self.people[i]} (工时:{total}) [休]"
            else:
                text = f"{self.people[i]} (工时:{total})"
            self.person_lb.insert(tk.END, text)
            if has_holiday or total == 0:
                self.person_lb.itemconfig(i, fg='green')
            else:
                self.person_lb.itemconfig(i, fg='black')
        if old_idx < len(self.people):
            self.person_idx = old_idx
        else:
            self.person_idx = 0
        self._set_person_selection()
        self._loading = False

    def refresh_task(self):
        for item in self.task_tree.get_children():
            self.task_tree.delete(item)
        if self.person_idx < len(self.people):
            for t in self.tasks[self.person_idx]:
                if t['op'] == 'A111':
                    dept_display = ''
                elif t['op'] == 'A112':
                    code = t.get('target_code', '')
                    name = t.get('target_name', '')
                    dept_display = f"{code}:{name}" if code else name
                else:
                    dept_display = t.get('dept', '')
                item = self.task_tree.insert('', tk.END, values=(
                    t['ship'], t['op'], dept_display, t['code'], t['name'], t['hours']
                ))
                if t['op'] == 'A112':
                    self.task_tree.tag_configure('support', foreground='green')
                    self.task_tree.item(item, tags=('support',))
        if self.mode == 'task' and self.task_idx < len(self.tasks[self.person_idx]):
            children = self.task_tree.get_children()
            if self.task_idx < len(children):
                self.task_tree.selection_set(children[self.task_idx])
                self.task_tree.see(children[self.task_idx])

    def on_task_select(self, e):
        sel = self.task_tree.selection()
        if sel:
            children = self.task_tree.get_children()
            try:
                self.task_idx = children.index(sel[0])
                self.mode = 'task'
                task = self.tasks[self.person_idx][self.task_idx]
                self.status.config(text=f"选中工单：{task['code']}")
            except:
                pass

    def move_task(self, delta):
        if self.mode != 'task' and self.tasks[self.person_idx]:
            self.mode = 'task'
            self.task_idx = 0
        if self.mode == 'task':
            children = self.task_tree.get_children()
            if not children:
                return
            new = self.task_idx + delta
            if 0 <= new < len(children):
                self.task_idx = new
                self.task_tree.selection_set(children[new])
                self.task_tree.see(children[new])
                task = self.tasks[self.person_idx][self.task_idx]
                self.status.config(text=f"选中工单：{task['code']}")
                self.task_tree.focus_set()

    def enter_person(self):
        if self.tasks[self.person_idx]:
            self.mode = 'task'
            self.task_idx = 0
            self.refresh_task()
            self.status.config(text="进入工单列表")
            children = self.task_tree.get_children()
            if children:
                self.task_tree.selection_set(children[0])
                self.task_tree.focus_set()
        else:
            messagebox.showinfo("提示", "暂无工单，请按+新增")

    def escape(self):
        if self.mode == 'task':
            self.mode = 'person'
            self.refresh_task()
            self.status.config(text="返回人员列表")
            self.person_lb.focus_set()

    # ---------- 自定义输入对话框 ----------
    def custom_askstring(self, title, prompt, initialvalue=""):
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.geometry("400x150")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.focus_force()
        dlg.lift()
        ttk.Label(dlg, text=prompt).pack(pady=10)
        entry = ttk.Entry(dlg, width=40)
        entry.pack(pady=5)
        if initialvalue:
            entry.insert(0, initialvalue)
        entry.focus_set()
        entry.select_range(0, tk.END)
        result = [None]
        def on_ok():
            result[0] = entry.get().strip()
            dlg.destroy()
        def on_cancel():
            dlg.destroy()
        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="确定", command=on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.LEFT, padx=5)
        dlg.bind('<Return>', lambda e: on_ok())
        dlg.bind('<Escape>', lambda e: on_cancel())
        self.root.wait_window(dlg)
        return result[0] if result[0] is not None else None

    # ---------- 选择派工单对话框 ----------
    def select_order_dialog(self, title, items, ship_no):
        per_page = 10
        total_pages = (len(items) + per_page - 1) // per_page
        current_page = 0
        selected_item = None
        countdown_active = False
        cancel_timer = None
        countdown_total = 30

        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        max_width = max(len(disp) for _, disp in items) * 8
        width = min(max(500, max_width), 800)
        height = min(400, 50 + per_page * 25)
        dlg.geometry(f"{width}x{height}")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.focus_force()
        dlg.lift()

        ttk.Label(dlg, text=f"船号: {ship_no}", font=("Arial", 10, "bold")).pack(pady=5)

        frame = ttk.Frame(dlg)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        lb = tk.Listbox(frame, font=("Consolas", 10), height=per_page)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=lb.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.config(yscrollcommand=sb.set)

        status_var = tk.StringVar()
        status_label = ttk.Label(dlg, textvariable=status_var, foreground="blue")
        status_label.pack(pady=5)
        timer_var = tk.StringVar()
        timer_label = ttk.Label(dlg, textvariable=timer_var, foreground="red")
        timer_label.pack()

        def update_list():
            lb.delete(0, tk.END)
            start = current_page * per_page
            end = min(start + per_page, len(items))
            for i in range(start, end):
                idx, disp = items[i]
                pos = i - start + 1
                if pos == 10:
                    label = "0"
                else:
                    label = str(pos)
                lb.insert(tk.END, f"{label}. {disp}")
            status_var.set(f"第 {current_page+1}/{total_pages} 页，直接按数字键选择，Q/E翻页，Enter直接确认")
            lb.focus_set()
            dlg.after(10, lb.focus_set)

        def clear_selection():
            nonlocal selected_item, countdown_active, cancel_timer
            if selected_item is not None:
                for i in range(lb.size()):
                    lb.itemconfig(i, fg='black')
                selected_item = None
                countdown_active = False
                if cancel_timer:
                    dlg.after_cancel(cancel_timer)
                timer_var.set("")
                lb.config(state=tk.NORMAL)

        def confirm_selection():
            nonlocal selected_item
            if selected_item is not None:
                dlg.result = selected_item
                dlg.destroy()

        def set_selected_item(item_idx):
            nonlocal selected_item, countdown_active, cancel_timer, countdown_total
            clear_selection()
            selected_item = item_idx
            start = current_page * per_page
            pos = item_idx - start + 1
            if 1 <= pos <= 10:
                list_index = pos - 1
                if list_index < lb.size():
                    lb.itemconfig(list_index, fg='red')
            countdown_total = 30
            countdown_active = True
            lb.config(state=tk.DISABLED)
            def timer_tick():
                nonlocal countdown_total
                if countdown_active and countdown_total > 0:
                    countdown_total -= 1
                    timer_var.set(f"确认中 {countdown_total/10:.1f} 秒，按 Esc 取消，或按 Enter 直接确认")
                    cancel_timer = dlg.after(100, timer_tick)
                elif countdown_active and countdown_total == 0:
                    confirm_selection()
            dlg.after(100, timer_tick)

        def on_key(event):
            nonlocal current_page, selected_item, countdown_active, cancel_timer
            keysym = event.keysym
            if countdown_active:
                if keysym == 'Escape':
                    clear_selection()
                    status_var.set(f"第 {current_page+1}/{total_pages} 页，直接按数字键选择，Q/E翻页")
                    return "break"
                elif keysym == 'Return':
                    confirm_selection()
                    return "break"
                else:
                    return "break"
            if keysym.isdigit():
                num = int(keysym)
                if num == 0:
                    pos = 10
                else:
                    pos = num
                if 1 <= pos <= 10:
                    start = current_page * per_page
                    actual_index = start + pos - 1
                    if actual_index < len(items):
                        set_selected_item(items[actual_index][0])
            elif keysym.lower() == 'q':
                if current_page > 0:
                    current_page -= 1
                    update_list()
                    clear_selection()
            elif keysym.lower() == 'e':
                if current_page < total_pages - 1:
                    current_page += 1
                    update_list()
                    clear_selection()
            elif keysym == 'Escape':
                dlg.result = None
                dlg.destroy()
            return "break"

        dlg.bind('<Key>', on_key)
        update_list()
        dlg.result = None
        self.root.wait_window(dlg)
        return dlg.result

    # ---------- 业务操作 ----------
    def add_task(self):
        if not self.people:
            return
        name = self.people[self.person_idx]
        ship = self.custom_askstring("新增工单", f"为 {name} 添加任务\n请输入船号（单独输入K自动补全）:")
        if not ship:
            return
        ship = ship.strip()
        if ship.upper() == "K":
            weekday = datetime.now().weekday()
            ship = "K122" if weekday < 5 else "K121"
            op = ship
            task = {
                'ship': ship,
                'code': ship,
                'name': '公休' if ship == "K121" else "事假",
                'hours': 8,
                'op': op,
                'dept': '',
                'team': '',
                'target_code': '',
                'target_name': ''
            }
            ok, msg = self.can_add_task(self.person_idx, task)
            if not ok:
                messagebox.showwarning("互斥规则", msg)
                return
            self.add_task_to_person(self.person_idx, task)
            return

        query = self.custom_askstring("新增工单", f"船号：{ship}\n输入派工单代码/名称:")
        if not query:
            return
        query = query.strip()
        match_col = 'F' if re.match(r'^[A-Za-z0-9_\-]+$', query) else 'G'
        matches = match_in_ship_index(self.ship_idx, self.records, ship, query, match_col)
        if not matches:
            messagebox.showwarning("未匹配", "未找到匹配的派工单")
            return
        matches = sort_matches_by_workshop(matches, self.records, self.team[0])

        selected_idx = self.select_order_dialog(f"选择派工单 - 船号:{ship}", matches, ship)
        if selected_idx is None:
            return
        source_rec = self.records[selected_idx]

        hours_str = self.custom_askstring("工时", "请输入工时（整数或几点五）:")
        if not hours_str:
            return
        try:
            if '.' in hours_str:
                hrs = float(hours_str)
                if hrs * 2 != int(hrs * 2):
                    messagebox.showerror("错误", "只接受整数或半整数")
                    return
            else:
                hrs = float(int(hours_str))
        except:
            messagebox.showerror("错误", "无效数值")
            return
        op = get_op_code(self.team[1], source_rec)
        team_name = source_rec['施工班组']
        target_code = self.name_to_code.get(team_name, team_name)
        target_name = team_name
        task = {
            'ship': ship,
            'code': source_rec['派工单编码'],
            'name': source_rec['派工单名称'],
            'hours': hrs,
            'op': op,
            'dept': source_rec['部门'],
            'team': source_rec['施工班组'],
            'target_code': target_code,
            'target_name': target_name
        }
        ok, msg = self.can_add_task(self.person_idx, task)
        if not ok:
            messagebox.showwarning("互斥规则", msg)
            return
        self.add_task_to_person(self.person_idx, task)

    def add_task_to_person(self, pi, task):
        total = self.get_work_total(pi) + (task['hours'] if task['op'] not in ('K121','K122') else 0)
        if total > 24:
            messagebox.showwarning("超限", "该人员实际工作总工时将超过24小时")
            return False
        self.tasks[pi].append(task)
        self.refresh_person_display()
        self.refresh_task()
        self.status.config(text=f"已添加 {task['code']} 工时 {task['hours']}")
        return True

    def del_task(self):
        if self.mode != 'task' or not self.tasks[self.person_idx]:
            messagebox.showinfo("提示", "请先在右侧选择要删除的工单")
            return
        task = self.tasks[self.person_idx][self.task_idx]
        if messagebox.askyesno("确认", f"删除 {task['code']} 吗？"):
            self.tasks[self.person_idx].pop(self.task_idx)
            if self.task_idx >= len(self.tasks[self.person_idx]):
                self.task_idx = max(0, len(self.tasks[self.person_idx])-1)
            if not self.tasks[self.person_idx]:
                self.mode = 'person'
            self.refresh_person_display()
            self.refresh_task()
            self.status.config(text="已删除工单")

    def modify_task(self):
        if self.mode != 'task' or not self.tasks[self.person_idx]:
            messagebox.showinfo("提示", "请先在右侧选择要修改的工单")
            return
        task = self.tasks[self.person_idx][self.task_idx]
        new = self.custom_askstring("修改工时", f"当前工时 {task['hours']}\n输入新工时:", str(task['hours']))
        if not new:
            return
        try:
            if '.' in new:
                hrs = float(new)
                if hrs * 2 != int(hrs * 2):
                    messagebox.showerror("错误", "只接受整数或半整数")
                    return
            else:
                hrs = float(int(new))
        except:
            messagebox.showerror("错误", "无效数值")
            return
        old = task['hours']
        diff = hrs - old
        if task['op'] not in ('K121','K122'):
            total = self.get_work_total(self.person_idx) + diff
            if total > 24:
                messagebox.showwarning("超限", "修改后实际工作总工时将超过24小时")
                return
        task['hours'] = hrs
        self.refresh_person_display()
        self.refresh_task()
        self.status.config(text=f"工时修改为 {hrs}")

    def show_stats(self):
        stats = defaultdict(float)
        for tasks in self.tasks.values():
            for t in tasks:
                key = (t['ship'], t['code'], t['name'])
                stats[key] += t['hours']
        if not stats:
            messagebox.showinfo("统计", "无工单数据")
            return
        lines = []
        for (ship, code, name), total in stats.items():
            lines.append(f"{self.team[0]} {self.team[1]} {ship} {name} {total}")
        text = "\n".join(lines)
        dlg = tk.Toplevel(self.root)
        dlg.title("工时统计")
        dlg.geometry("600x400")
        txt = tk.Text(dlg, wrap=tk.NONE)
        txt.insert(tk.END, text)
        txt.config(state=tk.DISABLED)
        sb_y = ttk.Scrollbar(dlg, orient=tk.VERTICAL, command=txt.yview)
        sb_x = ttk.Scrollbar(dlg, orient=tk.HORIZONTAL, command=txt.xview)
        txt.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        txt.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")
        dlg.grid_rowconfigure(0, weight=1)
        dlg.grid_columnconfigure(0, weight=1)
        ttk.Button(dlg, text="关闭", command=dlg.destroy).grid(row=2, column=0, pady=5)

    def save_and_export(self):
        try:
            self.save_tasks_to_file()
            messagebox.showinfo("成功", "文件已保存")
            if messagebox.askyesno("打开", "是否查看保存的文件？"):
                if sys.platform == 'win32':
                    os.startfile(self.edit_path)
                elif sys.platform == 'darwin':
                    subprocess.run(['open', self.edit_path])
                else:
                    subprocess.run(['xdg-open', self.edit_path])
            if messagebox.askyesno("继续", "是否还有工作需要处理？"):
                self.restart_for_new_file()
            else:
                self.root.quit()
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def restart_for_new_file(self):
        while True:
            path = filedialog.askopenfilename(title="选择新的待编辑文件", filetypes=[("Excel", "*.xlsx")])
            if not path:
                if messagebox.askyesno("退出", "未选择文件，是否退出？"):
                    self.root.quit()
                    return
                continue
            ok, msg = validate_edit_file(path)
            if not ok:
                messagebox.showerror("校验失败", msg)
                continue
            self.edit_path = path
            break
        self.people = load_people(self.edit_path)
        if not self.people:
            messagebox.showerror("错误", "未找到人员姓名")
            self.root.quit()
            return
        self.tasks.clear()
        self.load_tasks_from_file()
        self.person_idx = 0
        self.task_idx = 0
        self.mode = 'person'
        self.refresh_person_display()
        self.refresh_task()
        self.status.config(text="已加载新文件")

    def quit_app(self):
        if messagebox.askyesno("退出", "放弃所有未保存的更改并退出？"):
            self.root.quit()

    def on_closing(self):
        self.quit_app()

if __name__ == "__main__":
    root = tk.Tk()
    app = WorkEditApp(root)
    root.mainloop()