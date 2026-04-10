import tkinter as tk
import openpyxl
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
from datetime import datetime
import webbrowser
import tempfile
from openpyxl import Workbook, load_workbook

DATA_FILE = "stock_data.json"
BACKUP_FOLDER = "backups"
MOT_DE_PASSE = "admin123"  # 🔐 غير هذا الرقم السري كما تريد

class StockApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestion de Stock - Magasin d'Ameublement")
        self.root.geometry("1350x800")
        self.root.configure(bg='#f0f0f0')

        self.load_data_with_migration()
        self.create_notebook()
        self.refresh_all()
        self.update_stats()
        self.check_alerts()

    def load_data_with_migration(self):
        default_article = {
            "nom": "",
            "stock": 0,
            "prix": 0,
            "emplacement": "",
            "seuil": 5,
            "unite": "piece"
        }
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                self.articles = data.get("articles", {})
                self.historique = data.get("historique", [])
                for code, art in self.articles.items():
                    for key, val in default_article.items():
                        if key not in art:
                            art[key] = val
                for mvt in self.historique:
                    if "destinataire" not in mvt:
                        mvt["destinataire"] = ""
                    if "note" not in mvt:
                        mvt["note"] = ""
        else:
            self.articles = {}
            self.historique = []
        self.save_data()

    def save_data(self):
        if not os.path.exists(BACKUP_FOLDER):
            os.makedirs(BACKUP_FOLDER)
        backup_file = os.path.join(BACKUP_FOLDER, f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        if os.path.exists(DATA_FILE):
            import shutil
            shutil.copy(DATA_FILE, backup_file)
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "articles": self.articles,
                "historique": self.historique
            }, f, indent=4, ensure_ascii=False)

    def update_stats(self):
        nb_articles = len(self.articles)
        valeur_totale = sum(art.get("prix", 0) * art["stock"] for art in self.articles.values())
        nb_alertes = sum(1 for art in self.articles.values() if art["stock"] <= art.get("seuil", 0))
        self.stats_label.config(
            text=f"Produits: {nb_articles} | Valeur totale: {valeur_totale:,.0f} DT | Stocks bas: {nb_alertes}"
        )

    def check_alerts(self):
        alertes = []
        for code, art in self.articles.items():
            if art["stock"] <= art.get("seuil", 0):
                alertes.append(f"{code} - {art['nom']} : stock={art['stock']} (seuil={art['seuil']})")
        if alertes:
            messagebox.showwarning("Alerte stock bas", "\n".join(alertes[:10]))

    def create_notebook(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook', background='#f0f0f0')
        style.configure('TNotebook.Tab', font=('Arial', 10, 'bold'), padding=[15, 5])
        style.map('TNotebook.Tab', background=[('selected', '#2E7D32'), ('active', '#81C784')])

        self.stats_label = tk.Label(self.root, text="", font=('Arial', 11, 'bold'), bg='#e0e0e0', relief='sunken', anchor='w', padx=10)
        self.stats_label.pack(fill='x', padx=10, pady=(10,0))

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        self.tab_articles = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_articles, text="Gestion produits")
        self.create_article_tab()

        self.tab_mouvements = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_mouvements, text="Entree / Sortie")
        self.create_mouvement_tab()

        self.tab_stock = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_stock, text="Etat de stock")
        self.create_stock_tab()

        self.tab_hist = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_hist, text="Historique")
        self.create_history_tab()

        self.tab_utils = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_utils, text="Utilitaires")
        self.create_utils_tab()

    def create_article_tab(self):
        frame_form = ttk.LabelFrame(self.tab_articles, text="Ajouter / Modifier un produit", padding=15)
        frame_form.pack(fill='x', padx=15, pady=10)

        fields = [
            ("Code article :", "code"),
            ("Confirmer code :", "code_confirm"),
            ("Nom produit :", "nom"),
            ("Prix unitaire (DT) :", "prix"),
            ("Stock initial :", "stock_init"),
            ("Emplacement (rayon) :", "emplacement"),
            ("Seuil d'alerte (stock mini) :", "seuil"),
            ("Unite :", "unite")
        ]
        self.entries = {}
        row = 0
        for label, key in fields:
            ttk.Label(frame_form, text=label, font=('Arial', 10)).grid(row=row, column=0, padx=10, pady=8, sticky='e')
            self.entries[key] = ttk.Entry(frame_form, width=35, font=('Arial', 10))
            self.entries[key].grid(row=row, column=1, padx=10, pady=8, sticky='w')
            row += 1
        self.entries["stock_init"].insert(0, "0")
        self.entries["seuil"].insert(0, "5")
        self.entries["prix"].insert(0, "0")
        self.entries["unite"].insert(0, "piece")

        btn_frame = tk.Frame(frame_form)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=15)
        tk.Button(btn_frame, text="Ajouter produit", bg='#2E7D32', fg='white', font=('Arial', 10, 'bold'),
                  command=self.ajouter_article, padx=15, pady=5).pack(side='left', padx=10)
        tk.Button(btn_frame, text="Modifier produit", bg='#FF9800', fg='white',
                  command=self.modifier_article, padx=15, pady=5).pack(side='left', padx=10)
        tk.Button(btn_frame, text="Supprimer produit", bg='#f44336', fg='white',
                  command=self.supprimer_article, padx=15, pady=5).pack(side='left', padx=10)

        frame_list = ttk.LabelFrame(self.tab_articles, text="Liste des produits", padding=10)
        frame_list.pack(fill='both', expand=True, padx=15, pady=10)

        self.tree_articles = ttk.Treeview(frame_list, columns=("code", "nom", "stock", "prix", "seuil"), show="headings")
        self.tree_articles.heading("code", text="Code")
        self.tree_articles.heading("nom", text="Nom")
        self.tree_articles.heading("stock", text="Stock")
        self.tree_articles.heading("prix", text="Prix (DT)")
        self.tree_articles.heading("seuil", text="Seuil")
        for col in self.tree_articles['columns']:
            self.tree_articles.column(col, width=140)
        self.tree_articles.pack(fill='both', expand=True)
        self.tree_articles.bind('<<TreeviewSelect>>', self.on_article_select)

    def on_article_select(self, event):
        selected = self.tree_articles.selection()
        if selected:
            code = self.tree_articles.item(selected[0])['values'][0]
            art = self.articles[code]
            self.entries["code"].delete(0, tk.END)
            self.entries["code"].insert(0, code)
            self.entries["code_confirm"].delete(0, tk.END)
            self.entries["code_confirm"].insert(0, code)
            self.entries["nom"].delete(0, tk.END)
            self.entries["nom"].insert(0, art["nom"])
            self.entries["prix"].delete(0, tk.END)
            self.entries["prix"].insert(0, art.get("prix", 0))
            self.entries["stock_init"].delete(0, tk.END)
            self.entries["stock_init"].insert(0, art["stock"])
            self.entries["emplacement"].delete(0, tk.END)
            self.entries["emplacement"].insert(0, art.get("emplacement", ""))
            self.entries["seuil"].delete(0, tk.END)
            self.entries["seuil"].insert(0, art.get("seuil", 5))
            self.entries["unite"].delete(0, tk.END)
            self.entries["unite"].insert(0, art.get("unite", "piece"))

    def ajouter_article(self):
        code = self.entries["code"].get().strip()
        code_confirm = self.entries["code_confirm"].get().strip()
        nom = self.entries["nom"].get().strip()
        if not code or not nom:
            messagebox.showerror("Erreur", "Code et nom requis")
            return
        if code != code_confirm:
            messagebox.showerror("Erreur", "Les codes ne correspondent pas")
            return
        if code in self.articles:
            messagebox.showerror("Erreur", "Ce code existe deja")
            return
        try:
            stock = int(self.entries["stock_init"].get())
            prix = float(self.entries["prix"].get())
            seuil = int(self.entries["seuil"].get())
        except:
            messagebox.showerror("Erreur", "Verifiez les nombres (stock, prix, seuil)")
            return
        self.articles[code] = {
            "nom": nom,
            "stock": stock,
            "prix": prix,
            "emplacement": self.entries["emplacement"].get(),
            "seuil": seuil,
            "unite": self.entries["unite"].get()
        }
        self.save_data()
        self.refresh_all()
        messagebox.showinfo("Succes", f"Produit {code} ajoute")
        self.clear_article_form()

    def modifier_article(self):
        code = self.entries["code"].get().strip()
        if not code or code not in self.articles:
            messagebox.showerror("Erreur", "Selectionnez un produit")
            return
        try:
            stock = int(self.entries["stock_init"].get())
            prix = float(self.entries["prix"].get())
            seuil = int(self.entries["seuil"].get())
        except:
            messagebox.showerror("Erreur", "Verifiez les nombres")
            return
        self.articles[code].update({
            "nom": self.entries["nom"].get(),
            "stock": stock,
            "prix": prix,
            "emplacement": self.entries["emplacement"].get(),
            "seuil": seuil,
            "unite": self.entries["unite"].get()
        })
        self.save_data()
        self.refresh_all()
        messagebox.showinfo("Succes", f"Produit {code} modifie")

    def supprimer_article(self):
        code = self.entries["code"].get().strip()
        if not code or code not in self.articles:
            messagebox.showerror("Erreur", "Selectionnez un produit")
            return
        if messagebox.askyesno("Confirmation", f"Supprimer {code} definitivement ?"):
            del self.articles[code]
            self.save_data()
            self.refresh_all()
            self.clear_article_form()

    def clear_article_form(self):
        for key in ["code", "code_confirm", "nom", "prix", "stock_init", "emplacement", "seuil", "unite"]:
            self.entries[key].delete(0, tk.END)
        self.entries["stock_init"].insert(0, "0")
        self.entries["seuil"].insert(0, "5")
        self.entries["prix"].insert(0, "0")
        self.entries["unite"].insert(0, "piece")

    def create_mouvement_tab(self):
        frame = ttk.LabelFrame(self.tab_mouvements, text="Enregistrer un mouvement", padding=20)
        frame.pack(fill='both', expand=True, padx=20, pady=20)

        ttk.Label(frame, text="Article (code ou nom) :").grid(row=0, column=0, padx=10, pady=10, sticky='e')
        self.combo_article = ttk.Combobox(frame, width=50, font=('Arial', 10))
        self.combo_article.grid(row=0, column=1, padx=10, pady=10, columnspan=2, sticky='w')
        self.combo_article.bind('<KeyRelease>', self.update_article_list)
        self.combo_article.bind('<<ComboboxSelected>>', self.on_article_selected)

        ttk.Label(frame, text="Quantite :").grid(row=1, column=0, padx=10, pady=10, sticky='e')
        self.entry_qty = ttk.Entry(frame, width=15, font=('Arial', 10))
        self.entry_qty.grid(row=1, column=1, padx=10, pady=10, sticky='w')

        ttk.Label(frame, text="Date (AAAA-MM-JJ) :").grid(row=2, column=0, padx=10, pady=10, sticky='e')
        self.entry_date = ttk.Entry(frame, width=15, font=('Arial', 10))
        self.entry_date.insert(0, datetime.today().strftime("%Y-%m-%d"))
        self.entry_date.grid(row=2, column=1, padx=10, pady=10, sticky='w')

        ttk.Label(frame, text="Nom et prenom du destinataire :").grid(row=3, column=0, padx=10, pady=10, sticky='e')
        self.entry_destinataire = ttk.Entry(frame, width=40, font=('Arial', 10))
        self.entry_destinataire.grid(row=3, column=1, padx=10, pady=10, sticky='w')
        ttk.Label(frame, text="(Obligatoire pour une sortie)", font=('Arial', 8, 'italic')).grid(row=4, column=1, padx=10, pady=0, sticky='w')

        ttk.Label(frame, text="Note (optionnel) :").grid(row=5, column=0, padx=10, pady=10, sticky='e')
        self.entry_note = ttk.Entry(frame, width=40, font=('Arial', 10))
        self.entry_note.grid(row=5, column=1, padx=10, pady=10, sticky='w')

        btn_frame = tk.Frame(frame)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=20)
        tk.Button(btn_frame, text="Entree (+)", bg='#2196F3', fg='white', font=('Arial', 10, 'bold'),
                  command=self.entree_stock, padx=20, pady=5).pack(side='left', padx=10)
        tk.Button(btn_frame, text="Sortie (-)", bg='#f44336', fg='white', font=('Arial', 10, 'bold'),
                  command=self.sortie_stock, padx=20, pady=5).pack(side='left', padx=10)

        self.update_article_list()

    def update_article_list(self, event=None):
        texte = self.combo_article.get().lower()
        suggestions = []
        for code, art in self.articles.items():
            display = f"{code} - {art['nom']}"
            if texte in code.lower() or texte in art['nom'].lower():
                suggestions.append(display)
        self.combo_article['values'] = suggestions
        if event and event.keysym not in ('Down', 'Up', 'Return', 'Tab'):
            self.combo_article.set(texte)

    def on_article_selected(self, event=None):
        selected = self.combo_article.get()
        if ' - ' in selected:
            self.selected_code = selected.split(' - ')[0].strip()
        else:
            self.selected_code = None

    def get_selected_code(self):
        if hasattr(self, 'selected_code') and self.selected_code:
            return self.selected_code
        text = self.combo_article.get()
        if ' - ' in text:
            return text.split(' - ')[0].strip()
        for code, art in self.articles.items():
            if text == code or text == art['nom']:
                return code
        return None

    def entree_stock(self):
        self._mouvement("entree", +1, destinataire_requis=False)

    def sortie_stock(self):
        self._mouvement("sortie", -1, destinataire_requis=True)

    def _mouvement(self, type_mvt, signe, destinataire_requis):
        code = self.get_selected_code()
        if not code or code not in self.articles:
            messagebox.showerror("Erreur", "Veuillez selectionner un article valide (code ou nom)")
            return
        try:
            qty = int(self.entry_qty.get())
            if qty <= 0:
                raise ValueError
        except:
            messagebox.showerror("Erreur", "Quantite positive requise")
            return
        date = self.entry_date.get().strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except:
            messagebox.showerror("Erreur", "Date invalide (AAAA-MM-JJ)")
            return
        destinataire = self.entry_destinataire.get().strip()
        if destinataire_requis and not destinataire:
            messagebox.showerror("Erreur", "Le nom et prenom du destinataire sont obligatoires pour une sortie")
            return
        stock_actuel = self.articles[code]["stock"]
        if type_mvt == "sortie" and qty > stock_actuel:
            messagebox.showerror("Erreur", f"Stock insuffisant : {stock_actuel}")
            return
        nouveau = stock_actuel + signe * qty
        self.articles[code]["stock"] = nouveau
        self.historique.append({
            "date": date,
            "code": code,
            "nom": self.articles[code]["nom"],
            "type": type_mvt,
            "quantite": qty,
            "destinataire": destinataire if destinataire else "",
            "note": self.entry_note.get().strip()
        })
        self.save_data()
        self.refresh_all()
        self.entry_qty.delete(0, tk.END)
        self.entry_destinataire.delete(0, tk.END)
        self.entry_note.delete(0, tk.END)
        messagebox.showinfo("Succes", f"{type_mvt.capitalize()} de {qty} {self.articles[code]['unite']}(s) effectuee. Nouveau stock: {nouveau}")

    def create_stock_tab(self):
        frame = ttk.Frame(self.tab_stock)
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        search_frame = tk.Frame(frame)
        search_frame.pack(fill='x', pady=5)
        tk.Label(search_frame, text="Rechercher (code ou nom) :").pack(side='left', padx=5)
        self.search_entry = tk.Entry(search_frame, width=30)
        self.search_entry.pack(side='left', padx=5)
        self.search_entry.bind('<KeyRelease>', self.filter_stock)
        tk.Button(search_frame, text="Effacer", command=self.clear_search).pack(side='left', padx=5)

        self.tree_stock = ttk.Treeview(frame, columns=("code", "nom", "stock", "prix", "valeur", "seuil", "emplacement"), show="headings", height=20)
        self.tree_stock.heading("code", text="Code")
        self.tree_stock.heading("nom", text="Nom")
        self.tree_stock.heading("stock", text="Stock")
        self.tree_stock.heading("prix", text="Prix (DT)")
        self.tree_stock.heading("valeur", text="Valeur (DT)")
        self.tree_stock.heading("seuil", text="Seuil")
        self.tree_stock.heading("emplacement", text="Emplacement")
        for col in self.tree_stock['columns']:
            self.tree_stock.column(col, width=140)
        self.tree_stock.pack(fill='both', expand=True)

    def clear_search(self):
        self.search_entry.delete(0, tk.END)
        self.filter_stock()

    def filter_stock(self, event=None):
        term = self.search_entry.get().lower()
        for row in self.tree_stock.get_children():
            self.tree_stock.delete(row)
        for code, art in self.articles.items():
            if term in code.lower() or term in art['nom'].lower():
                valeur = art.get("prix", 0) * art["stock"]
                self.tree_stock.insert("", "end", values=(code, art["nom"], art["stock"], art.get("prix", 0), valeur, art.get("seuil", 0), art.get("emplacement", "")))
                if art["stock"] <= art.get("seuil", 0):
                    last = self.tree_stock.get_children()[-1]
                    self.tree_stock.tag_configure('alert', background='#FFE0B2')
                    self.tree_stock.item(last, tags=('alert',))

    def refresh_stock_table(self):
        self.filter_stock()

    def create_history_tab(self):
        frame = ttk.Frame(self.tab_hist)
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        btn_frame = tk.Frame(frame)
        btn_frame.pack(fill='x', pady=5)

        tk.Button(btn_frame, text="Supprimer la sélection (avec mot de passe)", bg='#f44336', fg='white',
                  command=self.supprimer_mouvement_avec_securite).pack(side='left', padx=5)

        self.tree_hist = ttk.Treeview(frame, columns=("date", "code", "nom", "type", "quantite", "destinataire", "note"), show="headings", height=18)
        self.tree_hist.heading("date", text="Date")
        self.tree_hist.heading("code", text="Code")
        self.tree_hist.heading("nom", text="Nom")
        self.tree_hist.heading("type", text="Type")
        self.tree_hist.heading("quantite", text="Qte")
        self.tree_hist.heading("destinataire", text="Destinataire")
        self.tree_hist.heading("note", text="Note")
        for col in self.tree_hist['columns']:
            self.tree_hist.column(col, width=140)
        self.tree_hist.pack(fill='both', expand=True)

    def supprimer_mouvement_avec_securite(self):
        selected = self.tree_hist.selection()
        if not selected:
            messagebox.showerror("Erreur", "Veuillez sélectionner une ligne dans l'historique")
            return

        mot_de_passe = simpledialog.askstring("Mot de passe requis", "Entrez le mot de passe pour supprimer cette ligne :", show='*')
        if mot_de_passe != MOT_DE_PASSE:
            messagebox.showerror("Accès refusé", "Mot de passe incorrect. Suppression annulée.")
            return

        selected_values = self.tree_hist.item(selected[0])['values']
        index_a_supprimer = None
        for idx, mvt in enumerate(self.historique):
            if (mvt["date"] == selected_values[0] and
                mvt["code"] == selected_values[1] and
                mvt["nom"] == selected_values[2] and
                mvt["type"] == selected_values[3] and
                mvt["quantite"] == selected_values[4] and
                mvt.get("destinataire", "") == selected_values[5] and
                mvt.get("note", "") == selected_values[6]):
                index_a_supprimer = idx
                break

        if index_a_supprimer is None:
            messagebox.showerror("Erreur", "Impossible de retrouver la ligne dans les données")
            return

        mouvement = self.historique[index_a_supprimer]
        code = mouvement["code"]
        qty = mouvement["quantite"]
        type_mvt = mouvement["type"]

        if code not in self.articles:
            messagebox.showerror("Erreur", f"L'article {code} n'existe plus dans le stock")
            return

        if type_mvt == "entree":
            nouveau_stock = self.articles[code]["stock"] - qty
            if nouveau_stock < 0:
                if not messagebox.askyesno("Stock négatif", f"Après suppression, le stock deviendrait {nouveau_stock}. Continuer quand même ?"):
                    return
        else:
            nouveau_stock = self.articles[code]["stock"] + qty

        self.articles[code]["stock"] = nouveau_stock
        del self.historique[index_a_supprimer]

        self.save_data()
        self.refresh_all()
        messagebox.showinfo("Succès", "La ligne a été supprimée et le stock a été corrigé.")

    def refresh_history_table(self):
        for row in self.tree_hist.get_children():
            self.tree_hist.delete(row)
        for mvt in reversed(self.historique):
            self.tree_hist.insert("", "end", values=(
                mvt["date"], mvt["code"], mvt["nom"], mvt["type"], mvt["quantite"],
                mvt.get("destinataire", ""), mvt.get("note", "")
            ))

    def create_utils_tab(self):
        frame = ttk.LabelFrame(self.tab_utils, text="Outils d'import/export et impression", padding=20)
        frame.pack(fill='both', expand=True, padx=20, pady=20)

        tk.Button(frame, text="Exporter l'etat de stock vers Excel", bg='#4CAF50', fg='white', font=('Arial', 10),
                  command=self.export_stock_excel, padx=10, pady=5, width=40).pack(pady=10)
        tk.Button(frame, text="Exporter l'historique vers Excel", bg='#2196F3', fg='white', font=('Arial', 10),
                  command=self.export_history_excel, padx=10, pady=5, width=40).pack(pady=10)
        tk.Button(frame, text="Importer des articles depuis Excel", bg='#FF9800', fg='white', font=('Arial', 10),
                  command=self.import_articles_from_excel, padx=10, pady=5, width=40).pack(pady=10)
        tk.Button(frame, text="Imprimer l'etat de stock", bg='#9C27B0', fg='white', font=('Arial', 10),
                  command=self.print_stock, padx=10, pady=5, width=40).pack(pady=10)
        tk.Button(frame, text="Rafraichir toutes les donnees", bg='#607D8B', fg='white', font=('Arial', 10),
                  command=self.refresh_all, padx=10, pady=5, width=40).pack(pady=20)

    def refresh_all(self):
        self.refresh_stock_table()
        self.refresh_history_table()
        self.update_article_list()
        self.update_stats()
        self.check_alerts()

    def export_stock_excel(self):
        if not self.articles:
            messagebox.showwarning("Aucune donnee", "Rien a exporter")
            return
        fichier = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not fichier:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock"
        ws.append(["Code", "Nom", "Stock", "Prix unitaire (DT)", "Valeur totale (DT)", "Seuil", "Emplacement", "Unite"])
        for code, art in self.articles.items():
            ws.append([code, art["nom"], art["stock"], art.get("prix",0), art.get("prix",0)*art["stock"], art.get("seuil",0), art.get("emplacement",""), art.get("unite","")])
        wb.save(fichier)
        messagebox.showinfo("Export", f"Stock exporte vers {fichier}")

    def export_history_excel(self):
        if not self.historique:
            messagebox.showwarning("Aucune donnee", "Historique vide")
            return
        fichier = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not fichier:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Historique"
        ws.append(["Date", "Code", "Nom", "Type", "Quantite", "Destinataire", "Note"])
        for mvt in self.historique:
            ws.append([mvt["date"], mvt["code"], mvt["nom"], mvt["type"], mvt["quantite"], mvt.get("destinataire",""), mvt.get("note","")])
        wb.save(fichier)
        messagebox.showinfo("Export", f"Historique exporte vers {fichier}")

    def import_articles_from_excel(self):
        fichier = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not fichier:
            return
        try:
            wb = load_workbook(fichier)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                messagebox.showerror("Erreur", "Fichier vide")
                return
            ajouts = 0
            for i, row in enumerate(rows[1:], start=2):
                code = str(row[0]).strip() if row[0] else ""
                nom = str(row[1]).strip() if row[1] else ""
                if not code or not nom:
                    continue
                if code in self.articles:
                    continue
                try:
                    stock = int(row[2]) if len(row) > 2 and row[2] is not None else 0
                except:
                    stock = 0
                prix = 0
                seuil = 5
                emplacement = ""
                unite = "piece"
                self.articles[code] = {
                    "nom": nom,
                    "stock": stock,
                    "prix": prix,
                    "emplacement": emplacement,
                    "seuil": seuil,
                    "unite": unite
                }
                ajouts += 1
            self.save_data()
            self.refresh_all()
            messagebox.showinfo("Import", f"{ajouts} article(s) importe(s) avec succes.")
        except Exception as e:
            messagebox.showerror("Erreur d'import", str(e))

    def print_stock(self):
        if not self.articles:
            messagebox.showwarning("Aucune donnee", "Rien a imprimer")
            return
        html = """<html><head><meta charset="UTF-8"><title>Etat de stock</title>
        <style>body{font-family:Arial; margin:20px;} table{border-collapse:collapse;width:100%;}
        th,td{border:1px solid #ddd;padding:8px;} th{background:#4CAF50;color:white;}
        .footer{margin-top:30px;font-size:0.8em;color:#777;}</style></head>
        <body><h1>Etat de stock au {date}</h1><table><th>Code</th><th>Nom</th><th>Stock</th><th>Prix (DT)</th><th>Valeur (DT)</th><th>Seuil</th><tr>{rows}
        </table><div class="footer">Genere le {date_heure}</div></body></html>"""
        rows = ""
        for code, art in self.articles.items():
            valeur = art.get("prix",0) * art["stock"]
            rows += f"<tr><td>{code}</td><td>{art['nom']}</td><td>{art['stock']}</td><td>{art.get('prix',0)}</td><td>{valeur}</td><td>{art.get('seuil',0)}</td></tr>"
        now = datetime.now()
        full = html.format(date=now.strftime("%d/%m/%Y"), rows=rows, date_heure=now.strftime("%d/%m/%Y %H:%M:%S"))
        with tempfile.NamedTemporaryFile(mode="w", suffix=".html", delete=False, encoding="utf-8") as f:
            f.write(full)
            path = f.name
        webbrowser.open(path)
        messagebox.showinfo("Impression", "Ouvert dans le navigateur. Ctrl+P pour imprimer.")

if __name__ == "__main__":
    root = tk.Tk()
    app = StockApp(root)
    root.mainloop()
