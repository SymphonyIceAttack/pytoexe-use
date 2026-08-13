
#importaciones
import sys
import os
import json
import shutil
import traceback
from datetime import datetime

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except Exception:
    with open("registro_iva_error.log", "w") as f:
        f.write("No se pudo abrir la parte visual del programa (tkinter).\n")
        f.write("Reinstala Python desde python.org marcando la opcion 'tcl/tk and IDLE'.\n\n")
        f.write(traceback.format_exc())
    sys.exit(1)

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except Exception:
    with open("registro_iva_error.log", "w") as f:
        f.write("Falta instalar una libreria. Abri una terminal y ejecuta: pip install openpyxl\n\n")
        f.write(traceback.format_exc())
    try:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Error al iniciar",
            "Falta instalar 'openpyxl' avisar a tobias"
            "pip install openpyxl\n\nDetalle guardado en registro_iva_error.log"
        )
    except Exception:
        pass
    sys.exit(1)



# GIF animado opcional. Para usarlo: pip install pillow
try:
    from PIL import Image, ImageTk, ImageSequence
except Exception:
    Image = ImageTk = ImageSequence = None


class AnimatedGIF(ttk.Label):
    def __init__(self, parent, path, max_size=(170, 170)):
        super().__init__(parent)
        self.frames = []
        self.index = 0
        self.delay = 80

        if Image is None or ImageTk is None or ImageSequence is None:
            self.configure(text="Instala Pillow para ver el GIF:\npip install pillow", foreground="gray")
            return

        try:
            gif = Image.open(path)
            self.delay = gif.info.get("duration", 80) or 80

            for frame in ImageSequence.Iterator(gif):
                img = frame.copy().convert("RGBA")
                img.thumbnail(max_size, Image.LANCZOS)
                self.frames.append(ImageTk.PhotoImage(img))

            if self.frames:
                self.configure(image=self.frames[0])
                self.after(self.delay, self.animate)
        except Exception:
            self.configure(text="No se pudo cargar el GIF", foreground="gray")

    def animate(self):
        if not self.frames:
            return
        self.index = (self.index + 1) % len(self.frames)
        self.configure(image=self.frames[self.index])
        self.after(self.delay, self.animate)

# ---------------------------------------------------------------------------
# Configuracion / rutas
# ---------------------------------------------------------------------------

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_PATH = os.path.join(BASE_DIR, "registro_iva_config.json")
PROVEEDORES_CACHE_PATH = os.path.join(BASE_DIR, "registro_iva_proveedores_cache.json")
BACKUPS_DIR = os.path.join(BASE_DIR, "backups")
MAX_BACKUPS = 30
DEFAULT_NEW_FILENAME = "Registro_IVA.xlsx"
GIF_FILENAME = "gif_derecha.gif"
GIF_MAX_SIZE = (170, 170)

COLUMNS = {
    "Fecha": 1, "Comprobante": 2, "Proveedor": 3, "Sit iva": 4,
    "Cuit": 5, "Tasa": 6, "Neto": 7, "IVA 21%": 8, "IVA 10,5%": 9,
    "PERC IVA": 10, "REG ESP": 11, "PERC IIBB": 12,
}
NUMERIC_FIELDS = ["Neto", "IVA 21%", "IVA 10,5%", "PERC IVA", "REG ESP", "PERC IIBB", "Tasa"]
TOTAL_COL = 13
FIRST_DATA_ROW = 3
LAST_DATA_ROW = 71
TOTALES_ROW = 72

MESES = [
    ("Enero", 1), ("Febrero", 2), ("Marzo", 3), ("Abril", 4),
    ("Mayo", 5), ("Junio", 6), ("Julio", 7), ("Agosto", 8),
    ("Septiembre", 9), ("Octubre", 10), ("Noviembre", 11), ("Diciembre", 12),
]
MES_A_NUM = {nombre: num for nombre, num in MESES}

SIT_IVA_OPCIONES = ["RI", "Monotributo", "Exento", "Consumidor Final", "Otro (especificar)"]
SIT_IVA_OTRO = "Otro (especificar)"


# ---------------------------------------------------------------------------
# Helpers de datos
# ---------------------------------------------------------------------------

def load_config():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_config(data):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def parse_number(text):
    text = text.strip()
    if not text:
        return None
    return float(text.replace(",", "."))


def find_next_empty_row(ws):
    row = FIRST_DATA_ROW
    while True:
        proveedor = ws.cell(row=row, column=COLUMNS["Proveedor"]).value
        if isinstance(proveedor, str) and proveedor.strip().upper() == "TOTALES":
            raise ValueError("completa todo bobiii")
        fecha = ws.cell(row=row, column=COLUMNS["Fecha"]).value
        if fecha is None and proveedor is None:
            return row
        row += 1


def get_provider_cuit_map(ws):
    """analizar proveedores existentes y guardar en cache"""
    mapping = {}
    row = FIRST_DATA_ROW
    while True:
        proveedor = ws.cell(row=row, column=COLUMNS["Proveedor"]).value
        if isinstance(proveedor, str) and proveedor.strip().upper() == "TOTALES":
            break
        if isinstance(proveedor, str) and proveedor.strip():
            cuit = ws.cell(row=row, column=COLUMNS["Cuit"]).value
            cuit = str(cuit).strip() if cuit else None
            mapping[proveedor.strip()] = cuit
        row += 1
        if row > FIRST_DATA_ROW + 500:  # tope de seguridad
            break
    return mapping


def load_provider_cache():
    if os.path.exists(PROVEEDORES_CACHE_PATH):
        try:
            with open(PROVEEDORES_CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_provider_cache(data):
    try:
        with open(PROVEEDORES_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def scan_for_excel_files():
    """Busca archivos .xlsx en la carpeta del programa (ignora temporales de Excel)."""
    found = []
    try:
        for name in os.listdir(BASE_DIR):
            if name.lower().endswith(".xlsx") and not name.startswith("~$"):
                found.append(os.path.join(BASE_DIR, name))
    except Exception:
        pass
    return sorted(found)


def create_template(path, sheet_name="Hoja1"):
    # genera nuevo excel en caso de no existir ninguno
    headers = ["Fecha", "Comprobante", "Proveedor", "Sit iva", "Cuit", "Tasa",
               "Neto", "IVA 21%", "IVA 10,5%", "PERC IVA", "REG ESP", "PERC IIBB", "Total"]
    widths = [12, 18, 33, 10, 14, 8, 15, 14, 14, 11, 10, 10, 16]
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    currency_fmt = '"$" #,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = w

    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        for col in range(1, 13):
            c = ws.cell(row=row, column=col)
            c.border = border
            if col == 1:
                c.number_format = "dd/mm/yyyy"
            elif col in (7, 8, 9, 10, 11, 12):
                c.number_format = currency_fmt
        total_cell = ws.cell(row=row, column=13, value=f"=G{row}+H{row}+I{row}+J{row}+K{row}+L{row}")
        total_cell.number_format = currency_fmt
        total_cell.border = border

    ws.cell(row=TOTALES_ROW, column=3, value="TOTALES").font = Font(bold=True)
    for col_letter, col in zip("GHIJKLM", range(7, 14)):
        cell = ws.cell(row=TOTALES_ROW, column=col,
                        value=f"=SUM({col_letter}{FIRST_DATA_ROW}:{col_letter}{LAST_DATA_ROW})")
        cell.font = Font(bold=True)
        cell.number_format = currency_fmt
        cell.border = border

    ws.cell(row=74, column=1, value="NOTAS DE CREDITO:").font = Font(bold=True)
    ws.cell(row=75, column=13, value="=SUM(G75:H75)").number_format = currency_fmt
    ws.cell(row=76, column=1, value="EN DOLARES :").font = Font(bold=True)
    ws.cell(row=77, column=13, value="=SUM(G77:L77)").number_format = currency_fmt

    ws.freeze_panes = "A3"
    wb.save(path)


def make_backup(filepath):
    try:
        os.makedirs(BACKUPS_DIR, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = os.path.splitext(os.path.basename(filepath))[0]
        dest = os.path.join(BACKUPS_DIR, f"{name}_{stamp}.xlsx")
        shutil.copy2(filepath, dest)
        backups = sorted(
            [os.path.join(BACKUPS_DIR, f) for f in os.listdir(BACKUPS_DIR) if f.startswith(name + "_")],
            key=os.path.getmtime,
        )
        while len(backups) > MAX_BACKUPS:
            os.remove(backups.pop(0))
    except Exception:
        pass  # un backup fallido no tiene que bloquear el guardado real O HICE ALGO MAL!!


def only_numeric_input(text):
    if text == "":
        return True
    allowed = set("0123456789,.")
    return all(ch in allowed for ch in text)


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("IVANEITOR 9000")
        self.geometry("760x740")
        self.resizable(False, False)

        self.filepath = None
        self.sheet_name = None
        self.entries = {}
        self.session_log = []
        self.provider_cuit_map = {}

        self._build_ui()
        self._try_autoload()

    # ---- Construccion de la gui -------------------------------------

    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        main_frame = ttk.Frame(self)
        main_frame.pack(fill="both", expand=True)

        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side="left", fill="both", expand=True)

        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side="right", fill="y", padx=(0, 10), pady=10)

        gif_path = os.path.join(BASE_DIR, GIF_FILENAME)
        if os.path.exists(gif_path):
            gif_box = ttk.LabelFrame(right_frame, text="")
            gif_box.pack(anchor="n", pady=(35, 0))
            self.side_gif = AnimatedGIF(gif_box, gif_path, max_size=GIF_MAX_SIZE)
            self.side_gif.pack(padx=6, pady=6)
        else:
            self.side_gif = None

        file_frame = ttk.LabelFrame(left_frame, text="Archivo Excel")
        file_frame.pack(fill="x", padx=10, pady=(10, 4))

        self.file_label = ttk.Label(file_frame, text="Buscando archivo...", foreground="gray", wraplength=420)
        self.file_label.pack(side="top", anchor="w", **pad)

        file_btns = ttk.Frame(file_frame)
        file_btns.pack(side="top", anchor="w", **pad)
        ttk.Button(file_btns, text="cambiar archivo...", command=self.select_file).pack(side="left")
        ttk.Button(file_btns, text="Crear excel nuevo...", command=self.create_new_file).pack(side="left", padx=6)

        sheet_row = ttk.Frame(file_frame)
        sheet_row.pack(side="top", anchor="w", fill="x", **pad)
        ttk.Label(sheet_row, text="Hoja:").pack(side="left")
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(sheet_row, textvariable=self.sheet_var, state="readonly", width=30)
        self.sheet_combo.pack(side="left", padx=6)
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_change)

        self.form_frame = ttk.LabelFrame(left_frame, text="Nuevo comprobante")
        self.form_frame.pack(fill="x", padx=10, pady=6)

        vcmd = (self.register(only_numeric_input), "%P")
        row_i = 0

        ttk.Label(self.form_frame, text="Fecha:").grid(row=row_i, column=0, sticky="e", **pad)
        fecha_frame = ttk.Frame(self.form_frame)
        fecha_frame.grid(row=row_i, column=1, sticky="w", **pad)
        today = datetime.now()
        self.dia_var = tk.StringVar(value=str(today.day))
        self.mes_var = tk.StringVar(value=MESES[today.month - 1][0])
        self.anio_var = tk.StringVar(value=str(today.year))
        ttk.Combobox(fecha_frame, textvariable=self.dia_var, values=[str(d) for d in range(1, 32)],
                     width=4, state="readonly").pack(side="left")
        ttk.Combobox(fecha_frame, textvariable=self.mes_var, values=[m[0] for m in MESES],
                     width=11, state="readonly").pack(side="left", padx=4)
        ttk.Combobox(fecha_frame, textvariable=self.anio_var,
                     values=[str(y) for y in range(today.year - 2, today.year + 2)],
                     width=6, state="readonly").pack(side="left")
        row_i += 1

        ttk.Label(self.form_frame, text="Proveedor *:").grid(row=row_i, column=0, sticky="e", **pad)
        self.proveedor_var = tk.StringVar()
        self.proveedor_entry = ttk.Combobox(self.form_frame, textvariable=self.proveedor_var, width=34)
        self.proveedor_entry.grid(row=row_i, column=1, sticky="w", **pad)
        self.proveedor_entry.bind("<KeyRelease>", self._on_proveedor_changed)
        self.proveedor_entry.bind("<<ComboboxSelected>>", self._on_proveedor_changed)
        row_i += 1

        ttk.Label(self.form_frame, text="Comprobante:").grid(row=row_i, column=0, sticky="e", **pad)
        comp_entry = ttk.Entry(self.form_frame, width=34)
        comp_entry.grid(row=row_i, column=1, sticky="w", **pad)
        comp_placeholder = "ej: 0001-00012345"
        comp_entry.insert(0, comp_placeholder)
        comp_entry.bind("<FocusIn>", lambda e, en=comp_entry, ph=comp_placeholder: self._clear_placeholder(en, ph))
        self.entries["Comprobante"] = comp_entry
        row_i += 1

        ttk.Label(self.form_frame, text="Sit iva:").grid(row=row_i, column=0, sticky="e", **pad)
        sitiva_frame = ttk.Frame(self.form_frame)
        sitiva_frame.grid(row=row_i, column=1, sticky="w", **pad)
        self.sitiva_var = tk.StringVar(value="RI")
        self.sitiva_combo = ttk.Combobox(sitiva_frame, textvariable=self.sitiva_var,
                                         values=SIT_IVA_OPCIONES, width=17, state="readonly")
        self.sitiva_combo.pack(side="left")
        self.sitiva_otro_var = tk.StringVar()
        self.sitiva_otro_entry = ttk.Entry(sitiva_frame, textvariable=self.sitiva_otro_var, width=15)
        self.sitiva_combo.bind("<<ComboboxSelected>>", self._on_sitiva_change)
        row_i += 1

        ttk.Label(self.form_frame, text="Cuit:").grid(row=row_i, column=0, sticky="e", **pad)
        cuit_entry = ttk.Entry(self.form_frame, width=34)
        cuit_entry.grid(row=row_i, column=1, sticky="w", **pad)
        cuit_placeholder = "ej: 20-12345678-9"
        cuit_entry.insert(0, cuit_placeholder)
        cuit_entry.bind("<FocusIn>", lambda e, en=cuit_entry, ph=cuit_placeholder: self._clear_placeholder(en, ph))
        self.entries["Cuit"] = cuit_entry
        row_i += 1

        for label in NUMERIC_FIELDS:
            ttk.Label(self.form_frame, text=label + ":").grid(row=row_i, column=0, sticky="e", **pad)
            entry = ttk.Entry(self.form_frame, width=34, validate="key", validatecommand=vcmd)
            entry.grid(row=row_i, column=1, sticky="w", **pad)
            entry.bind("<KeyRelease>", self._check_valid)
            self.entries[label] = entry
            row_i += 1

        ttk.Label(self.form_frame, text="* Proveedor es obligatorio", foreground="gray").grid(
            row=row_i, column=0, columnspan=2, sticky="w", **pad)
        row_i += 1

        self.save_btn = ttk.Button(left_frame, text="GUARDAR FACTURA", command=self.save_entry, state="disabled")
        self.save_btn.pack(fill="x", padx=10, pady=(4, 2), ipady=8)

        self.status_label = ttk.Label(left_frame, text="", foreground="green", font=("", 10, "bold"))
        self.status_label.pack(pady=(2, 6))

        log_frame = ttk.LabelFrame(left_frame, text="Cargados en esta sesion")
        log_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.log_list = tk.Listbox(log_frame, height=8)
        self.log_list.pack(fill="both", expand=True, padx=6, pady=6)

        self._set_form_enabled(False)

    # ---- Comportamiento ----------------------------------------------

    def _on_sitiva_change(self, event=None):
        if self.sitiva_var.get() == SIT_IVA_OTRO:
            self.sitiva_otro_entry.pack(side="left", padx=(4, 0))
            self.sitiva_otro_entry.focus_set()
        else:
            self.sitiva_otro_entry.pack_forget()
            self.sitiva_otro_var.set("")

    def _on_proveedor_changed(self, event=None):
        name = self.proveedor_var.get().strip()
        cuit = self.provider_cuit_map.get(name)
        if not cuit:
            # busqueda sin distinguir mayusculas/minusculas
            for known_name, known_cuit in self.provider_cuit_map.items():
                if known_name.lower() == name.lower() and known_cuit:
                    cuit = known_cuit
                    break
        if cuit:
            self.entries["Cuit"].delete(0, tk.END)
            self.entries["Cuit"].insert(0, cuit)
        self._check_valid()

    def _clear_placeholder(self, entry, placeholder):
        if entry.get() == placeholder:
            entry.delete(0, tk.END)

    def _set_form_enabled(self, enabled):
        state = "normal" if enabled else "disabled"
        combo_state = "readonly" if enabled else "disabled"
        for child in self.form_frame.winfo_children():
            try:
                if child is self.proveedor_entry:
                    child.configure(state=state)
                elif isinstance(child, ttk.Combobox):
                    child.configure(state=combo_state)
                elif isinstance(child, ttk.Frame):
                    for sub in child.winfo_children():
                        if isinstance(sub, ttk.Combobox):
                            sub.configure(state=combo_state)
                        elif isinstance(sub, ttk.Entry):
                            sub.configure(state=state)
                else:
                    child.configure(state=state)
            except tk.TclError:
                pass
        self._check_valid()

    def _check_valid(self, event=None):
        ok = bool(self.filepath) and bool(self.sheet_name) and bool(self.proveedor_var.get().strip())
        self.save_btn.configure(state="normal" if ok else "disabled")

    def _try_autoload(self):
        cfg = load_config()
        filepath = cfg.get("filepath")
        sheet = cfg.get("sheet")
        if filepath and os.path.exists(filepath):
            self._load_file(filepath, preferred_sheet=sheet)
            return

        candidates = scan_for_excel_files()
        if len(candidates) == 1:
            self._load_file(candidates[0])
            messagebox.showinfo(
                "Archivo encontrado",
                f"Se encontro y configuro automaticamente:\n{os.path.basename(candidates[0])}"
            )
        elif len(candidates) > 1:
            self.file_label.config(
                text="Hay varios archivos Excel en esta carpeta. Elegi cual usar.", foreground="orange"
            )
            messagebox.showinfo(
                "Varios archivos encontrados",
                "Hay mas de un archivo Excel en esta carpeta.\n"
                "Elegi cual usar con el boton 'Seleccionar / cambiar archivo...'."
            )
        else:
            new_path = os.path.join(BASE_DIR, DEFAULT_NEW_FILENAME)
            try:
                create_template(new_path)
            except Exception as e:
                self.file_label.config(text="No se pudo crear un Excel nuevo. Avisa a un tecnico.", foreground="red")
                with open(os.path.join(BASE_DIR, "registro_iva_error.log"), "w") as f:
                    f.write(traceback.format_exc())
                messagebox.showerror("Error", f"No se encontro ningun Excel y no se pudo crear uno nuevo:\n{e}")
                return
            self._load_file(new_path)
            messagebox.showinfo(
                "Planilla creada",
                "No se encontro ningun Excel en esta carpeta, asi que se creo uno nuevo:\n"
                f"{DEFAULT_NEW_FILENAME}\n\nYa esta listo para usar."
            )

    def _load_file(self, filepath, preferred_sheet=None):
        try:
            wb = load_workbook(filepath, read_only=True)
        except Exception as e:
            self.file_label.config(text=f"No se pudo abrir {os.path.basename(filepath)}", foreground="red")
            messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")
            return
        self.filepath = filepath
        self.file_label.config(text=filepath, foreground="black")
        self.sheet_combo["values"] = wb.sheetnames
        if preferred_sheet in wb.sheetnames:
            self.sheet_var.set(preferred_sheet)
            self.sheet_name = preferred_sheet
        else:
            self.sheet_combo.current(len(wb.sheetnames) - 1)
            self.sheet_name = self.sheet_var.get()
        wb.close()
        save_config({"filepath": self.filepath, "sheet": self.sheet_name})
        self._refresh_providers()
        self._set_form_enabled(True)

    def select_file(self):
        path = filedialog.askopenfilename(
            title="Seleccionar planilla Excel",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return
        self._load_file(path)

    def create_new_file(self):
        path = filedialog.asksaveasfilename(
            title="Crear planilla nueva",
            defaultextension=".xlsx",
            initialfile=DEFAULT_NEW_FILENAME,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return
        try:
            create_template(path)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear la planilla:\n{e}")
            return
        self._load_file(path)
        messagebox.showinfo("Listo", f"Planilla creada:\n{os.path.basename(path)}")

    def _on_sheet_change(self, event=None):
        self.sheet_name = self.sheet_var.get()
        save_config({"filepath": self.filepath, "sheet": self.sheet_name})
        self._refresh_providers()
        self._check_valid()

    def _refresh_providers(self):
        sheet_map = {}
        try:
            wb = load_workbook(self.filepath, read_only=True)
            ws = wb[self.sheet_name]
            sheet_map = get_provider_cuit_map(ws)
            wb.close()
        except Exception:
            pass

        cache = load_provider_cache()
        for name, cuit in sheet_map.items():
            if cuit:
                cache[name] = cuit
            elif name not in cache:
                cache[name] = cuit
        save_provider_cache(cache)

        self.provider_cuit_map = cache
        self.proveedor_entry["values"] = sorted(cache.keys(), key=str.lower)

    def clear_form(self):
        today = datetime.now()
        self.dia_var.set(str(today.day))
        self.mes_var.set(MESES[today.month - 1][0])
        self.anio_var.set(str(today.year))
        self.proveedor_var.set("")
        self.sitiva_var.set("RI")
        self.sitiva_otro_var.set("")
        self.sitiva_otro_entry.pack_forget()
        placeholder_map = {
            "Comprobante": "ej: 0001-00012345", "Cuit": "ej: 20-12345678-9",
        }
        for label, entry in self.entries.items():
            entry.delete(0, tk.END)
            if label in placeholder_map:
                entry.insert(0, placeholder_map[label])
        self._check_valid()

    def save_entry(self):
        if not self.filepath or not self.sheet_name:
            messagebox.showwarning("Falta configuracion", "Primero seleccioná el archivo y la hoja.")
            return

        proveedor = self.proveedor_var.get().strip()
        if not proveedor:
            messagebox.showwarning("Dato requerido", "El campo Proveedor es obligatorio.")
            return

        try:
            dia = int(self.dia_var.get())
            mes = MES_A_NUM[self.mes_var.get()]
            anio = int(self.anio_var.get())
            fecha = datetime(anio, mes, dia)
        except ValueError:
            messagebox.showerror("Fecha invalida", "Esa combinacion de dia/mes/año no existe (ej: 31 de Abril).")
            return

        values = {}
        placeholder_map = {
            "Comprobante": "ej: 0001-00012345", "Cuit": "ej: 20-12345678-9",
        }
        for label, entry in self.entries.items():
            text = entry.get().strip()
            if text == placeholder_map.get(label):
                text = ""
            values[label] = text

        sitiva_choice = self.sitiva_var.get()
        if sitiva_choice == SIT_IVA_OTRO:
            sit_iva_value = self.sitiva_otro_var.get().strip()
            if not sit_iva_value:
                messagebox.showwarning("Dato requerido", "Elegiste 'Otro' en Sit iva: especificá cuál.")
                return
        else:
            sit_iva_value = sitiva_choice

        try:
            numeric_values = {f: parse_number(values[f]) for f in NUMERIC_FIELDS}
        except ValueError:
            messagebox.showerror("Numero invalido", "Revisá los campos numéricos (Neto, IVA, etc.).")
            return

        if all(numeric_values[f] is None for f in ["Neto", "IVA 21%", "IVA 10,5%"]):
            if not messagebox.askyesno(
                "Comprobante sin montos",
                "No cargaste Neto ni IVA. ¿Sos boludo?"
            ):
                return

        try:
            make_backup(self.filepath)

            wb = load_workbook(self.filepath)
            ws = wb[self.sheet_name]
            row = find_next_empty_row(ws)

            ws.cell(row=row, column=COLUMNS["Fecha"], value=fecha)
            ws.cell(row=row, column=COLUMNS["Comprobante"], value=values["Comprobante"] or None)
            ws.cell(row=row, column=COLUMNS["Proveedor"], value=proveedor)
            ws.cell(row=row, column=COLUMNS["Sit iva"], value=sit_iva_value)
            ws.cell(row=row, column=COLUMNS["Cuit"], value=values["Cuit"] or None)
            for f in NUMERIC_FIELDS:
                ws.cell(row=row, column=COLUMNS[f], value=numeric_values[f])

            total_cell = ws.cell(row=row, column=TOTAL_COL)
            if total_cell.value is None:
                total_cell.value = f"=SUM(G{row}:L{row})"

            wb.save(self.filepath)
        except PermissionError:
            messagebox.showerror(
                "Archivo en uso",
                "No se pudo guardar porque el archivo esta abierto en Excel."
            )
            return
        except ValueError as e:
            messagebox.showerror("No hay lugar", str(e))
            return
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el comprobante:\n{e}\n\nNo se modifico el archivo original.")
            return

        total = sum(v for v in numeric_values.values() if v is not None)
        entry_desc = f"{fecha.strftime('%d/%m/%Y')}  |  {proveedor}  |  ${total:,.2f}  (fila {row})"
        self.session_log.append(entry_desc)
        self.log_list.insert(0, entry_desc)

        self.status_label.config(text=f"✔ Guardado correctamente en la fila {row}")
        self._refresh_providers()
        self.clear_form()
        self.proveedor_entry.focus_set()


if __name__ == "__main__":
    try:
        App().mainloop()
    except Exception:
        with open("registro_iva_error.log", "w") as f:
            f.write(traceback.format_exc())
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "Error inesperado",
                "El programa se cerro por un error.\n\nDetalle guardado en registro_iva_error.log"
            )
        except Exception:
            pass
