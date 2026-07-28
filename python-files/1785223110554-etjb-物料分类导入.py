# -*- coding: utf-8 -*-
"""
物料提报表 - 按采购单分类(P列)将源文件数据导入对应模板
支持命令行拖拽文件或直接指定路径
用法：
  python 物料分类导入.py                            # 交互式输入路径
  python 物料分类导入.py <源文件路径> [模板文件夹路径]  # 直接指定路径

打包为exe后双击运行，也可：
  物料分类导入.exe <源文件路径> [模板文件夹路径]
"""
import os
import sys
import shutil
import traceback
import platform


# ============================================================
# 配置区
# ============================================================
DEFAULT_TEMPLATE_DIR = r"C:\Users\2025057195\Desktop\物料提报表单分类\采购单模板"

OUTPUT_SUBDIR = "分类输出"

CATEGORY_MAP = {
    "标准件": {
        "template": "008模具车间《标准件物料提报表》-模板.xlsm",
        "sheet": "计划",
        "start_row": 2,
    },
    "铸件": {
        "template": "模具车间铸件采购单-模板.xlsm",
        "sheet": "汇总",
        "start_row": 5,
    },
    "精料": {
        "template": "精料采购审批-模板.xlsm",
        "sheet": "明细表",
        "start_row": 6,
    },
}


# ============================================================
# 字段映射
# ============================================================
def fill_standard(wst, row, src_row, serial):
    """标准件提报表"""
    ws = src_row
    wst.cell(row, 1, serial)
    wst.cell(row, 4, f'=IF(B{row}="是","模具标准件",IF(B{row}="否","",""))')
    wst.cell(row, 5, ws.get("B"))
    wst.cell(row, 6, ws.get("I"))
    wst.cell(row, 9, "EA")
    wst.cell(row, 11, ws.get("L"))
    wst.cell(row, 12, ws.get("C"))
    wst.cell(row, 13, "汽车冲压车间")
    wst.cell(row, 14, "飞碟四库")
    wst.cell(row, 15, "2107")
    wst.cell(row, 19, "模具开发")
    wst.cell(row, 23, ws.get("O"))
    wst.cell(row, 24, ws.get("A"))
    wst.cell(row, 25, ws.get("K"))


def adjust_formula(formula, dest_row):
    """将源文件公式中的行号替换为目标行号"""
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula
    import re
    def replace_row(match):
        col = match.group(1)
        return f"{col}{dest_row}"
    return re.sub(r'([A-Z]+)(\d+)', replace_row, formula)


def fill_jingliao(wst, row, src_row, serial):
    """精料采购单"""
    ws = src_row
    wst.cell(row, 1, serial)
    wst.cell(row, 2, ws.get("N"))
    wst.cell(row, 3, ws.get("B"))
    wst.cell(row, 4, ws.get("D"))
    wst.cell(row, 5, ws.get("F"))
    wst.cell(row, 6, ws.get("G"))
    wst.cell(row, 7, ws.get("H"))
    wst.cell(row, 8, ws.get("C"))
    # M列可能是公式（如 =F19*G19*H19*C19*7.85/1000000），需要调整行号
    weight_val = ws.get("M")
    wst.cell(row, 9, adjust_formula(weight_val, row))
    wst.cell(row, 10, ws.get("O"))
    wst.cell(row, 11, ws.get("A"))


def fill_zhujian(wst, row, src_row, serial):
    """铸件采购单"""
    ws = src_row
    wst.cell(row, 1, serial)
    wst.cell(row, 2, ws.get("O"))
    wst.cell(row, 3, ws.get("A"))
    wst.cell(row, 4, ws.get("B"))
    wst.cell(row, 5, ws.get("D"))
    wst.cell(row, 6, ws.get("C"))
    wst.cell(row, 7, ws.get("K"))


FILL_MAP = {
    "标准件": fill_standard,
    "精料": fill_jingliao,
    "铸件": fill_zhujian,
}


# ============================================================
# 主逻辑
# ============================================================
def read_source_data(filepath):
    """读取源文件数据"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    COL_LETTERS = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"]
    categorized = {}
    data_start = 9
    for r in range(data_start, ws.max_row + 1):
        category = ws.cell(r, 16).value
        if not category or str(category).strip() == "":
            continue
        category = str(category).strip()
        if category not in CATEGORY_MAP:
            continue
        row_data = {}
        for idx, letter in enumerate(COL_LETTERS):
            col = idx + 1
            val = ws.cell(r, col).value
            row_data[letter] = val
        categorized.setdefault(category, []).append(row_data)
    wb.close()
    return categorized


def get_input(prompt, default=""):
    """获取用户输入，去除引号"""
    val = input(prompt).strip().strip('"').strip("'")
    if not val and default:
        return default
    return val


def open_folder(path):
    """跨平台打开文件夹（exe在Windows上运行，保留os.startfile兼容）"""
    try:
        if platform.system() == "Windows":
            os.startfile(path)
        elif platform.system() == "Darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}" &')
    except Exception:
        pass


def main():
    print("=" * 60)
    print("  物料提报表 - 分类导入工具 v1.0")
    print("=" * 60)

    # 解析参数
    args = sys.argv[1:]
    source_file = None
    template_dir = None

    if len(args) >= 1:
        source_file = args[0].strip('"').strip("'")
    if len(args) >= 2:
        template_dir = args[1].strip('"').strip("'")

    # 交互式输入
    if not source_file:
        print("\n请拖拽源文件到此窗口，或输入路径：")
        source_file = get_input("源文件路径: ")
    if not source_file or not os.path.isfile(source_file):
        print(f"❌ 文件不存在: {source_file}")
        input("\n按回车键退出...")
        return

    print(f"源文件: {source_file}")

    if not template_dir:
        print(f"\n模板文件夹（直接回车使用默认: {DEFAULT_TEMPLATE_DIR}）")
        template_dir = get_input("模板文件夹: ", DEFAULT_TEMPLATE_DIR)
    if not os.path.isdir(template_dir):
        print(f"❌ 文件夹不存在: {template_dir}")
        input("\n按回车键退出...")
        return

    print(f"模板文件夹: {template_dir}")

    # 验证模板
    for cat, info in CATEGORY_MAP.items():
        tpl_path = os.path.join(template_dir, info["template"])
        if not os.path.exists(tpl_path):
            print(f"❌ 缺少模板: {info['template']}")
            input("\n按回车键退出...")
            return
        print(f"  ✓ {cat}: {info['template']}")

    # 读取源文件
    print("\n正在读取源文件...")
    categorized = read_source_data(source_file)
    total = sum(len(v) for v in categorized.values())
    print(f"共读取 {total} 条:")
    for cat, rows in sorted(categorized.items()):
        print(f"  - {cat}: {len(rows)} 条")

    if total == 0:
        print("⚠ P列未找到标准件/精料/铸件分类数据，退出。")
        input("\n按回车键退出...")
        return

    # 输出目录
    source_dir = os.path.dirname(source_file)
    output_dir = os.path.join(source_dir, OUTPUT_SUBDIR)
    os.makedirs(output_dir, exist_ok=True)

    # 生成文件
    import openpyxl
    from copy import copy as copy_style

    generated_files = []

    for cat, rows in sorted(categorized.items()):
        info = CATEGORY_MAP[cat]
        fill_func = FILL_MAP[cat]
        tpl_path = os.path.join(template_dir, info["template"])
        sheet_name = info["sheet"]
        start_row = info["start_row"]

        base_name = os.path.splitext(os.path.basename(source_file))[0]
        ext = os.path.splitext(info["template"])[1]
        output_name = f"{base_name}_{cat}{ext}"
        output_path = os.path.join(output_dir, output_name)

        print(f"\n生成: {output_name} ...")

        shutil.copy2(tpl_path, output_path)

        wb = openpyxl.load_workbook(output_path, keep_vba=True)
        wst = wb[sheet_name]

        # 采样第一数据行样式
        sample_row = start_row
        ref_cells = {}
        for c in range(1, wst.max_column + 1):
            rc = wst.cell(sample_row, c)
            ref_cells[c] = {
                "font": copy_style(rc.font),
                "fill": copy_style(rc.fill),
                "border": copy_style(rc.border),
                "alignment": copy_style(rc.alignment),
                "number_format": rc.number_format,
            }

        # 计算模板原有数据行数（用于清理）
        original_data_rows = 0
        for r in range(start_row, wst.max_row + 1):
            if wst.cell(r, 1).value is not None:
                original_data_rows += 1
            else:
                break

        # 写入数据
        for idx, src_row in enumerate(rows):
            dest_row = start_row + idx

            # 应用样式
            if idx > 0:
                first_row = start_row
                for c in range(1, wst.max_column + 1):
                    try:
                        target = wst.cell(dest_row, c)
                        ref = wst.cell(first_row, c)
                        target.font = copy_style(ref.font)
                        target.fill = copy_style(ref.fill)
                        target.border = copy_style(ref.border)
                        target.alignment = copy_style(ref.alignment)
                        target.number_format = ref.number_format
                    except:
                        pass

            fill_func(wst, dest_row, src_row)

        # 清理多余旧数据行
        for r in range(start_row + len(rows), start_row + original_data_rows):
            for c in range(1, wst.max_column + 1):
                wst.cell(r, c).value = None

        wb.save(output_path)
        wb.close()
        generated_files.append(output_path)
        print(f"  ✓ {len(rows)} 条 -> {output_path}")

    # 报告
    print("\n" + "=" * 60)
    print("  处理完成！")
    print("=" * 60)
    print(f"\n输出目录: {output_dir}")
    for f in generated_files:
        n = os.path.basename(f)
        print(f"  📄 {n}")
    print(f"\n共处理 {total} 条，生成 {len(generated_files)} 个文件。")

    # 自动打开输出文件夹（跨平台）
    open_folder(output_dir)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        traceback.print_exc()
    try:
        input("\n按回车键退出...")
    except (EOFError, OSError):
        pass
