"""
برنامج جرد المخزون المتكامل 
يعمل بدون إنترنت 
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
import uuid
from datetime import datetime

# ==================== قاعدة البيانات المحلية ====================

DATA_FILE = "inventory_data.json"

DEFAULT_CATEGORIES = ["إلكترونيات", "أدوات مكتبية", "مواد غذائية", "قطع غيار", "مواد تنظيف", "أخرى"]
UNITS = ["قطعة", "كرتون", "كيلو", "لتر", "متر", "علبة", "حزمة"]


def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"products": [], "movements": [], "categories": DEFAULT_CATEGORIES[:]}


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ==================== التطبيق الرئيسي ====================

class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("نظام جرد المخزون")
        self.root.geometry("1100x700")
        self.root.configure(bg="#f0f4f8")

        self.data = load_data()

        # الألوان
        self.colors = {
            "primary": "#1a7a6d",
            "primary_light": "#e0f2f0",
            "bg": "#f0f4f8",
            "card": "#ffffff",
            "text": "#1e293b",
            "text_muted": "#64748b",
            "danger": "#dc2626",
            "danger_light": "#fef2f2",
            "success": "#16a34a",
            "success_light": "#f0fdf4",
            "warning": "#d97706",
            "border": "#e2e8f0",
        }

        self.setup_styles()
        self.build_ui()
        self.refresh_all()

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("Title.TLabel", font=("Arial", 18, "bold"), background=self.colors["card"], foreground=self.colors["text"])
        style.configure("Subtitle.TLabel", font=("Arial", 10), background=self.colors["card"], foreground=self.colors["text_muted"])
        style.configure("Stat.TLabel", font=("Arial", 20, "bold"), background=self.colors["card"], foreground=self.colors["text"])
        style.configure("StatTitle.TLabel", font=("Arial", 9), background=self.colors["card"], foreground=self.colors["text_muted"])

        style.configure("Primary.TButton", font=("Arial", 10, "bold"), background=self.colors["primary"], foreground="white")
        style.map("Primary.TButton", background=[("active", "#15665b")])

        style.configure("Danger.TButton", font=("Arial", 10), background=self.colors["danger"], foreground="white")
        style.map("Danger.TButton", background=[("active", "#b91c1c")])

        style.configure("Treeview", font=("Arial", 10), rowheight=32, background=self.colors["card"], fieldbackground=self.colors["card"])
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"), background=self.colors["bg"], foreground=self.colors["text"])

    def build_ui(self):
        # ===== الشريط العلوي =====
        header = tk.Frame(self.root, bg=self.colors["card"], pady=12, padx=20)
        header.pack(fill="x")

        tk.Label(header, text="📦 نظام جرد المخزون", font=("Arial", 18, "bold"),
                 bg=self.colors["card"], fg=self.colors["primary"]).pack(side="right")
        tk.Label(header, text="إدارة متكاملة - أوفلاين", font=("Arial", 9),
                 bg=self.colors["card"], fg=self.colors["text_muted"]).pack(side="right", padx=(0, 15))

        # أزرار الشريط العلوي
        btn_frame = tk.Frame(header, bg=self.colors["card"])
        btn_frame.pack(side="left")

        buttons = [
            ("➕ إضافة منتج", self.show_add_product, self.colors["primary"]),
            ("🔄 حركة مخزون", self.show_stock_movement, "#4b5563"),
            ("📥 استيراد Excel", self.import_excel, "#0369a1"),
            ("📤 تصدير Excel", self.export_excel, "#0369a1"),
        ]
        for text, cmd, color in buttons:
            btn = tk.Button(btn_frame, text=text, command=cmd, font=("Arial", 9, "bold"),
                           bg=color, fg="white", relief="flat", padx=12, pady=5, cursor="hand2",
                           activebackground=color)
            btn.pack(side="left", padx=3)

        # ===== بطاقات الإحصائيات =====
        self.stats_frame = tk.Frame(self.root, bg=self.colors["bg"], pady=10, padx=20)
        self.stats_frame.pack(fill="x")

        # ===== منطقة البحث والفلترة =====
        filter_frame = tk.Frame(self.root, bg=self.colors["bg"], padx=20, pady=5)
        filter_frame.pack(fill="x")

        # البحث
        tk.Label(filter_frame, text="🔍 بحث:", font=("Arial", 10), bg=self.colors["bg"]).pack(side="right")
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *args: self.refresh_products())
        search_entry = tk.Entry(filter_frame, textvariable=self.search_var, font=("Arial", 11), width=30,
                               relief="solid", bd=1)
        search_entry.pack(side="right", padx=(0, 5))

        # فلتر التصنيف
        tk.Label(filter_frame, text="التصنيف:", font=("Arial", 10), bg=self.colors["bg"]).pack(side="right", padx=(15, 5))
        self.category_var = tk.StringVar(value="الكل")
        cats = ["الكل"] + self.data.get("categories", DEFAULT_CATEGORIES)
        self.category_combo = ttk.Combobox(filter_frame, textvariable=self.category_var, values=cats,
                                           state="readonly", width=15, font=("Arial", 10))
        self.category_combo.pack(side="right")
        self.category_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_products())

        # ===== التبويبات =====
        tab_frame = tk.Frame(self.root, bg=self.colors["bg"], padx=20, pady=5)
        tab_frame.pack(fill="x")

        self.current_tab = tk.StringVar(value="products")
        for text, val in [("📋 المنتجات", "products"), ("📊 سجل الحركات", "movements")]:
            rb = tk.Radiobutton(tab_frame, text=text, variable=self.current_tab, value=val,
                               font=("Arial", 11, "bold"), bg=self.colors["bg"], fg=self.colors["text"],
                               selectcolor=self.colors["primary_light"], indicatoron=0, padx=20, pady=6,
                               relief="flat", cursor="hand2", command=self.switch_tab)
            rb.pack(side="right", padx=2)

        # ===== جدول المنتجات =====
        table_frame = tk.Frame(self.root, bg=self.colors["bg"], padx=20, pady=5)
        table_frame.pack(fill="both", expand=True)

        # جدول المنتجات
        self.products_frame = tk.Frame(table_frame, bg=self.colors["card"])
        self.products_frame.pack(fill="both", expand=True)

        columns = ("name", "sku", "category", "quantity", "min_qty", "price", "unit", "location")
        headers = ("المنتج", "الرمز", "التصنيف", "الكمية", "الحد الأدنى", "السعر", "الوحدة", "الموقع")

        scrollbar = ttk.Scrollbar(self.products_frame)
        scrollbar.pack(side="left", fill="y")

        self.tree = ttk.Treeview(self.products_frame, columns=columns, show="headings",
                                 yscrollcommand=scrollbar.set, height=15)
        scrollbar.config(command=self.tree.yview)

        for col, header in zip(columns, headers):
            self.tree.heading(col, text=header, anchor="e")
            self.tree.column(col, width=120, anchor="e")

        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", lambda e: self.edit_selected_product())

        # أزرار التعديل والحذف
        action_frame = tk.Frame(self.products_frame, bg=self.colors["card"], pady=8)
        action_frame.pack(fill="x")

        tk.Button(action_frame, text="✏️ تعديل المحدد", command=self.edit_selected_product,
                 font=("Arial", 10, "bold"), bg=self.colors["primary"], fg="white",
                 relief="flat", padx=15, pady=4, cursor="hand2").pack(side="right", padx=5)
        tk.Button(action_frame, text="🗑️ حذف المحدد", command=self.delete_selected_product,
                 font=("Arial", 10, "bold"), bg=self.colors["danger"], fg="white",
                 relief="flat", padx=15, pady=4, cursor="hand2").pack(side="right", padx=5)

        # جدول الحركات
        self.movements_frame = tk.Frame(table_frame, bg=self.colors["card"])

        mov_columns = ("product", "type", "quantity", "reason", "date")
        mov_headers = ("المنتج", "النوع", "الكمية", "السبب", "التاريخ")

        mov_scrollbar = ttk.Scrollbar(self.movements_frame)
        mov_scrollbar.pack(side="left", fill="y")

        self.mov_tree = ttk.Treeview(self.movements_frame, columns=mov_columns, show="headings",
                                      yscrollcommand=mov_scrollbar.set, height=18)
        mov_scrollbar.config(command=self.mov_tree.yview)

        for col, header in zip(mov_columns, mov_headers):
            self.mov_tree.heading(col, text=header, anchor="e")
            self.mov_tree.column(col, width=150, anchor="e")

        self.mov_tree.pack(fill="both", expand=True)

    # ==================== الإحصائيات ====================

    def refresh_stats(self):
        for widget in self.stats_frame.winfo_children():
            widget.destroy()

        products = self.data["products"]
        movements = self.data["movements"]

        total = len(products)
        total_value = sum(p.get("price", 0) * p.get("quantity", 0) for p in products)
        low_stock = sum(1 for p in products if p.get("quantity", 0) <= p.get("min_quantity", 5))
        today = datetime.now().strftime("%Y-%m-%d")
        today_moves = sum(1 for m in movements if m.get("date", "").startswith(today))

        stats = [
            ("إجمالي المنتجات", str(total), "📦", self.colors["primary_light"]),
            ("القيمة الإجمالية", f"{total_value:,.0f} ر.س", "💰", self.colors["success_light"]),
            ("مخزون منخفض", str(low_stock), "⚠️", self.colors["danger_light"]),
            ("حركات اليوم", str(today_moves), "📊", "#fef3c7"),
        ]

        for title, value, icon, bg in stats:
            card = tk.Frame(self.stats_frame, bg=bg, padx=20, pady=12, relief="solid", bd=1, highlightbackground=self.colors["border"])
            card.pack(side="right", expand=True, fill="x", padx=5)

            tk.Label(card, text=f"{icon} {title}", font=("Arial", 9), bg=bg, fg=self.colors["text_muted"]).pack(anchor="e")
            tk.Label(card, text=value, font=("Arial", 18, "bold"), bg=bg, fg=self.colors["text"]).pack(anchor="e")

    # ==================== تحديث الجداول ====================

    def refresh_all(self):
        self.refresh_stats()
        self.refresh_products()
        self.refresh_movements()
        self.update_category_filter()

    def update_category_filter(self):
        cats = ["الكل"] + self.data.get("categories", DEFAULT_CATEGORIES)
        self.category_combo["values"] = cats

    def refresh_products(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        search = self.search_var.get().strip().lower()
        cat_filter = self.category_var.get()

        for p in self.data["products"]:
            # فلترة البحث
            if search and search not in p.get("name", "").lower() and search not in p.get("sku", "").lower():
                continue
            # فلترة التصنيف
            if cat_filter != "الكل" and p.get("category") != cat_filter:
                continue

            qty = p.get("quantity", 0)
            min_qty = p.get("min_quantity", 5)

            values = (
                p.get("name", ""),
                p.get("sku", ""),
                p.get("category", ""),
                qty,
                min_qty,
                f"{p.get('price', 0):,.2f}",
                p.get("unit", ""),
                p.get("location", ""),
            )

            tag = "low" if qty <= min_qty else ""
            item_id = self.tree.insert("", "end", values=values, tags=(tag,), iid=p["id"])

        self.tree.tag_configure("low", foreground=self.colors["danger"])

    def refresh_movements(self):
        for item in self.mov_tree.get_children():
            self.mov_tree.delete(item)

        products_map = {p["id"]: p["name"] for p in self.data["products"]}

        for m in self.data["movements"][:50]:
            product_name = products_map.get(m.get("product_id"), "منتج محذوف")
            mov_type = "إدخال ➕" if m.get("type") == "in" else "إخراج ➖"
            date_str = m.get("date", "")[:10]

            self.mov_tree.insert("", "end", values=(
                product_name, mov_type, m.get("quantity", 0), m.get("reason", ""), date_str
            ))

    def switch_tab(self):
        if self.current_tab.get() == "products":
            self.movements_frame.pack_forget()
            self.products_frame.pack(fill="both", expand=True)
        else:
            self.products_frame.pack_forget()
            self.movements_frame.pack(fill="both", expand=True)

    # ==================== إضافة منتج ====================

    def show_add_product(self):
        self._product_dialog(None)

    def edit_selected_product(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("تنبيه", "الرجاء تحديد منتج للتعديل")
            return
        product_id = selected[0]
        product = next((p for p in self.data["products"] if p["id"] == product_id), None)
        if product:
            self._product_dialog(product)

    def _product_dialog(self, product):
        dialog = tk.Toplevel(self.root)
        dialog.title("تعديل المنتج" if product else "إضافة منتج جديد")
        dialog.geometry("450x550")
        dialog.configure(bg=self.colors["card"])
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="تعديل المنتج" if product else "إضافة منتج جديد",
                font=("Arial", 16, "bold"), bg=self.colors["card"], fg=self.colors["primary"]).pack(pady=(15, 10))

        form = tk.Frame(dialog, bg=self.colors["card"], padx=30)
        form.pack(fill="both", expand=True)

        fields = {}

        def add_field(label, key, default="", row=0):
            tk.Label(form, text=label, font=("Arial", 10, "bold"), bg=self.colors["card"],
                    anchor="e").grid(row=row, column=1, sticky="e", pady=5, padx=5)
            var = tk.StringVar(value=str(product.get(key, default)) if product else str(default))
            entry = tk.Entry(form, textvariable=var, font=("Arial", 11), width=25, relief="solid", bd=1)
            entry.grid(row=row, column=0, sticky="w", pady=5)
            fields[key] = var

        add_field("اسم المنتج *", "name", "", 0)
        add_field("رمز المنتج (SKU)", "sku", "", 1)

        # التصنيف
        tk.Label(form, text="التصنيف", font=("Arial", 10, "bold"), bg=self.colors["card"],
                anchor="e").grid(row=2, column=1, sticky="e", pady=5, padx=5)
        cat_var = tk.StringVar(value=product.get("category", DEFAULT_CATEGORIES[0]) if product else DEFAULT_CATEGORIES[0])
        cat_combo = ttk.Combobox(form, textvariable=cat_var, values=self.data.get("categories", DEFAULT_CATEGORIES),
                                  state="readonly", width=23, font=("Arial", 11))
        cat_combo.grid(row=2, column=0, sticky="w", pady=5)
        fields["category"] = cat_var

        add_field("الكمية", "quantity", "0", 3)
        add_field("الحد الأدنى", "min_quantity", "5", 4)
        add_field("السعر", "price", "0", 5)

        # الوحدة
        tk.Label(form, text="الوحدة", font=("Arial", 10, "bold"), bg=self.colors["card"],
                anchor="e").grid(row=6, column=1, sticky="e", pady=5, padx=5)
        unit_var = tk.StringVar(value=product.get("unit", "قطعة") if product else "قطعة")
        unit_combo = ttk.Combobox(form, textvariable=unit_var, values=UNITS, state="readonly", width=23, font=("Arial", 11))
        unit_combo.grid(row=6, column=0, sticky="w", pady=5)
        fields["unit"] = unit_var

        add_field("الموقع", "location", "", 7)
        add_field("ملاحظات", "notes", "", 8)

        def on_save():
            name = fields["name"].get().strip()
            if not name:
                messagebox.showwarning("تنبيه", "اسم المنتج مطلوب!")
                return

            product_data = {
                "name": name,
                "sku": fields["sku"].get().strip(),
                "category": fields["category"].get(),
                "quantity": int(float(fields["quantity"].get() or 0)),
                "min_quantity": int(float(fields["min_quantity"].get() or 5)),
                "price": float(fields["price"].get() or 0),
                "unit": fields["unit"].get(),
                "location": fields["location"].get().strip(),
                "notes": fields["notes"].get().strip(),
            }

            if product:
                # تعديل
                for p in self.data["products"]:
                    if p["id"] == product["id"]:
                        p.update(product_data)
                        p["updated_at"] = datetime.now().isoformat()
                        break
                messagebox.showinfo("نجاح", "تم تعديل المنتج بنجاح ✅")
            else:
                # إضافة
                product_data["id"] = str(uuid.uuid4())
                product_data["created_at"] = datetime.now().isoformat()
                product_data["updated_at"] = datetime.now().isoformat()
                self.data["products"].append(product_data)
                messagebox.showinfo("نجاح", "تمت إضافة المنتج بنجاح ✅")

            save_data(self.data)
            self.refresh_all()
            dialog.destroy()

        btn_frame = tk.Frame(dialog, bg=self.colors["card"], pady=15)
        btn_frame.pack(fill="x", padx=30)

        tk.Button(btn_frame, text="💾 حفظ", command=on_save, font=("Arial", 11, "bold"),
                 bg=self.colors["primary"], fg="white", relief="flat", padx=30, pady=6, cursor="hand2").pack(side="right", padx=5)
        tk.Button(btn_frame, text="إلغاء", command=dialog.destroy, font=("Arial", 11),
                 bg=self.colors["border"], fg=self.colors["text"], relief="flat", padx=20, pady=6, cursor="hand2").pack(side="right", padx=5)

    # ==================== حذف منتج ====================

    def delete_selected_product(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("تنبيه", "الرجاء تحديد منتج للحذف")
            return

        if not messagebox.askyesno("تأكيد الحذف", "هل أنت متأكد من حذف هذا المنتج؟"):
            return

        product_id = selected[0]
        self.data["products"] = [p for p in self.data["products"] if p["id"] != product_id]
        self.data["movements"] = [m for m in self.data["movements"] if m.get("product_id") != product_id]
        save_data(self.data)
        self.refresh_all()
        messagebox.showinfo("تم", "تم حذف المنتج ✅")

    # ==================== حركة مخزون ====================

    def show_stock_movement(self):
        if not self.data["products"]:
            messagebox.showwarning("تنبيه", "لا توجد منتجات. أضف منتجاً أولاً!")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("حركة مخزون")
        dialog.geometry("420x400")
        dialog.configure(bg=self.colors["card"])
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="🔄 حركة مخزون", font=("Arial", 16, "bold"),
                bg=self.colors["card"], fg=self.colors["primary"]).pack(pady=(15, 10))

        form = tk.Frame(dialog, bg=self.colors["card"], padx=30)
        form.pack(fill="both", expand=True)

        # نوع الحركة
        tk.Label(form, text="نوع الحركة", font=("Arial", 10, "bold"), bg=self.colors["card"]).pack(anchor="e", pady=(0, 5))
        type_var = tk.StringVar(value="in")
        type_frame = tk.Frame(form, bg=self.colors["card"])
        type_frame.pack(fill="x", pady=(0, 10))
        tk.Radiobutton(type_frame, text="➕ إدخال", variable=type_var, value="in",
                       font=("Arial", 11), bg=self.colors["card"], selectcolor=self.colors["success_light"]).pack(side="right", padx=10)
        tk.Radiobutton(type_frame, text="➖ إخراج", variable=type_var, value="out",
                       font=("Arial", 11), bg=self.colors["card"], selectcolor=self.colors["danger_light"]).pack(side="right", padx=10)

        # المنتج
        tk.Label(form, text="المنتج *", font=("Arial", 10, "bold"), bg=self.colors["card"]).pack(anchor="e", pady=(0, 5))
        product_names = {p["name"]: p["id"] for p in self.data["products"]}
        product_display = [f"{p['name']} ({p['quantity']} {p['unit']})" for p in self.data["products"]]
        product_var = tk.StringVar()
        product_combo = ttk.Combobox(form, textvariable=product_var, values=product_display,
                                      state="readonly", width=35, font=("Arial", 11))
        product_combo.pack(fill="x", pady=(0, 10))

        # الكمية
        tk.Label(form, text="الكمية *", font=("Arial", 10, "bold"), bg=self.colors["card"]).pack(anchor="e", pady=(0, 5))
        qty_var = tk.StringVar(value="1")
        tk.Entry(form, textvariable=qty_var, font=("Arial", 11), relief="solid", bd=1).pack(fill="x", pady=(0, 10))

        # السبب
        tk.Label(form, text="السبب", font=("Arial", 10, "bold"), bg=self.colors["card"]).pack(anchor="e", pady=(0, 5))
        reason_var = tk.StringVar()
        tk.Entry(form, textvariable=reason_var, font=("Arial", 11), relief="solid", bd=1).pack(fill="x", pady=(0, 10))

        def on_confirm():
            idx = product_combo.current()
            if idx < 0:
                messagebox.showwarning("تنبيه", "الرجاء اختيار منتج")
                return

            qty = int(float(qty_var.get() or 0))
            if qty <= 0:
                messagebox.showwarning("تنبيه", "الكمية يجب أن تكون أكبر من صفر")
                return

            product = self.data["products"][idx]
            mov_type = type_var.get()

            if mov_type == "out" and product["quantity"] < qty:
                messagebox.showerror("خطأ", f"الكمية المتوفرة ({product['quantity']}) غير كافية!")
                return

            # تحديث الكمية
            if mov_type == "in":
                product["quantity"] += qty
            else:
                product["quantity"] -= qty
            product["updated_at"] = datetime.now().isoformat()

            # تسجيل الحركة
            movement = {
                "id": str(uuid.uuid4()),
                "product_id": product["id"],
                "type": mov_type,
                "quantity": qty,
                "reason": reason_var.get().strip(),
                "date": datetime.now().isoformat(),
            }
            self.data["movements"].insert(0, movement)

            save_data(self.data)
            self.refresh_all()
            dialog.destroy()

            action = "إدخال" if mov_type == "in" else "إخراج"
            messagebox.showinfo("تم", f"تم {action} {qty} {product['unit']} بنجاح ✅")

        tk.Button(form, text="✅ تأكيد الحركة", command=on_confirm, font=("Arial", 11, "bold"),
                 bg=self.colors["primary"], fg="white", relief="flat", padx=30, pady=8, cursor="hand2").pack(pady=15)

    # ==================== تصدير Excel ====================

    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messagebox.showerror("خطأ", "الرجاء تثبيت مكتبة openpyxl:\npip install openpyxl")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"جرد_المخزون_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        if not file_path:
            return

        wb = Workbook()

        # ورقة المنتجات
        ws = wb.active
        ws.title = "المنتجات"
        ws.sheet_view.rightToLeft = True

        headers = ["اسم المنتج", "الرمز", "التصنيف", "الكمية", "الحد الأدنى", "السعر", "الوحدة", "الموقع", "ملاحظات"]
        header_fill = PatternFill(start_color="1a7a6d", end_color="1a7a6d", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="right")

        for row, p in enumerate(self.data["products"], 2):
            ws.cell(row=row, column=1, value=p.get("name", ""))
            ws.cell(row=row, column=2, value=p.get("sku", ""))
            ws.cell(row=row, column=3, value=p.get("category", ""))
            ws.cell(row=row, column=4, value=p.get("quantity", 0))
            ws.cell(row=row, column=5, value=p.get("min_quantity", 5))
            ws.cell(row=row, column=6, value=p.get("price", 0))
            ws.cell(row=row, column=7, value=p.get("unit", ""))
            ws.cell(row=row, column=8, value=p.get("location", ""))
            ws.cell(row=row, column=9, value=p.get("notes", ""))

        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 18

        # ورقة الحركات
        ws2 = wb.create_sheet("الحركات")
        ws2.sheet_view.rightToLeft = True
        mov_headers = ["المنتج", "النوع", "الكمية", "السبب", "التاريخ"]
        products_map = {p["id"]: p["name"] for p in self.data["products"]}

        for col, header in enumerate(mov_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="right")

        for row, m in enumerate(self.data["movements"], 2):
            ws2.cell(row=row, column=1, value=products_map.get(m.get("product_id"), "محذوف"))
            ws2.cell(row=row, column=2, value="إدخال" if m.get("type") == "in" else "إخراج")
            ws2.cell(row=row, column=3, value=m.get("quantity", 0))
            ws2.cell(row=row, column=4, value=m.get("reason", ""))
            ws2.cell(row=row, column=5, value=m.get("date", "")[:10])

        for col in range(1, 6):
            ws2.column_dimensions[chr(64 + col)].width = 18

        wb.save(file_path)
        messagebox.showinfo("نجاح", f"تم تصدير البيانات بنجاح ✅\n{file_path}")

    # ==================== استيراد Excel ====================

    def import_excel(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            messagebox.showerror("خطأ", "الرجاء تثبيت مكتبة openpyxl:\npip install openpyxl")
            return

        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return

        if not messagebox.askyesno("تأكيد", "سيتم استبدال جميع البيانات الحالية.\nهل تريد المتابعة؟"):
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            products = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                product = {
                    "id": str(uuid.uuid4()),
                    "name": str(row[0] or ""),
                    "sku": str(row[1] or "") if len(row) > 1 else "",
                    "category": str(row[2] or "أخرى") if len(row) > 2 else "أخرى",
                    "quantity": int(float(row[3] or 0)) if len(row) > 3 else 0,
                    "min_quantity": int(float(row[4] or 5)) if len(row) > 4 else 5,
                    "price": float(row[5] or 0) if len(row) > 5 else 0,
                    "unit": str(row[6] or "قطعة") if len(row) > 6 else "قطعة",
                    "location": str(row[7] or "") if len(row) > 7 else "",
                    "notes": str(row[8] or "") if len(row) > 8 else "",
                    "created_at": datetime.now().isoformat(),
                    "updated_at": datetime.now().isoformat(),
                }
                products.append(product)

            self.data["products"] = products
            self.data["movements"] = []
            save_data(self.data)
            self.refresh_all()

            messagebox.showinfo("نجاح", f"تم استيراد {len(products)} منتج بنجاح ✅")

        except Exception as e:
            messagebox.showerror("خطأ", f"تعذر قراءة الملف:\n{str(e)}")


# ==================== تشغيل البرنامج ====================

if __name__ == "__main__":
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()