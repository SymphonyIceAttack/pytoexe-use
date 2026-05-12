import pandas as pd
import re
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl import load_workbook

def get_program_folder():
    """Возвращает папку, где находится программа"""
    if getattr(sys, 'frozen', False):
        # Запущено как скомпилированный .exe
        return os.path.dirname(sys.executable)
    else:
        # Запущено как скрипт
        return os.path.dirname(os.path.abspath(__file__))

def clean_cell_value(val):
    """Убирает пробелы в начале/конце, заменяет множественные пробелы на один внутри"""
    if pd.isna(val):
        return val
    if isinstance(val, str):
        val = val.strip()
        val = re.sub(r' +', ' ', val)
        return val
    return val

def process_file_preserve_formatting(file_path, report_path):
    """Обрабатывает файл, сохраняя всё форматирование"""
    
    wb = load_workbook(file_path)
    changes_log = []
    total_changes = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        changes_log.append(f'\nЛист: {sheet_name}')
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    old_val = cell.value
                    new_val = clean_cell_value(old_val)
                    
                    if old_val != new_val:
                        cell.value = new_val
                        changes_log.append(
                            f'  Строка {cell.row}, колонка {cell.column_letter}\n'
                            f'    Было: "{old_val}"\n'
                            f'    Стало: "{new_val}"'
                        )
                        total_changes += 1
    
    # Сохраняем исправленный файл
    base, ext = os.path.splitext(file_path)
    new_file_path = f"{base}_исправленный{ext}"
    wb.save(new_file_path)
    
    # Сохраняем отчёт
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(f"Отчёт об исправлениях от {datetime.now()}\n")
        f.write(f"Файл: {os.path.basename(file_path)}\n")
        f.write(f"Всего исправлений: {total_changes}\n")
        if total_changes == 0:
            f.write("\nНет исправлений.\n")
        else:
            f.write("\n".join(changes_log))
    
    return new_file_path, total_changes

def main():
    # Получаем папку с программой
    program_folder = get_program_folder()
    
    # Создаём скрытое главное окно
    root = tk.Tk()
    root.withdraw()
    
    # Делаем окно поверх всех для диалогов
    root.lift()
    root.attributes('-topmost', True)
    
    # Выбираем файл — открываем в папке с программой
    file_path = filedialog.askopenfilename(
        title="Выберите файл Excel для валидации",
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ],
        initialdir=program_folder  # Папка с программой
    )
    
    root.attributes('-topmost', False)
    
    if not file_path:
        messagebox.showinfo("Информация", "Файл не выбран. Выход.")
        root.destroy()
        return
    
    # Отчёт сохраняем автоматически в той же папке, что и исходный файл
    report_path = os.path.join(os.path.dirname(file_path), "отчёт_об_исправлениях.txt")
    
    try:
        # Показываем окно загрузки
        progress_window = tk.Toplevel(root)
        progress_window.title("Обработка")
        progress_window.geometry("300x100")
        progress_window.transient(root)
        progress_window.grab_set()
        
        tk.Label(progress_window, text="Обработка файла...\nПожалуйста, подождите").pack(pady=30)
        root.update()
        
        # Обрабатываем файл
        new_file, changes_count = process_file_preserve_formatting(file_path, report_path)
        
        progress_window.destroy()
        
        # Показываем результат
        result_text = f"✅ Готово!\n\n"
        result_text += f"📊 Исправлений: {changes_count}\n"
        result_text += f"📁 Новый файл: {new_file}\n"
        result_text += f"📄 Отчёт: {report_path}"
        
        messagebox.showinfo("Результат", result_text)
        
    except Exception as e:
        if 'progress_window' in locals():
            progress_window.destroy()
        messagebox.showerror("Ошибка", f"При обработке произошла ошибка:\n{str(e)}")
    
    finally:
        root.destroy()

if __name__ == "__main__":
    main()
