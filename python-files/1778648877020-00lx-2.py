import os
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

# 隐藏主窗口
tk.Tk().withdraw()

def select_file(prompt_msg, window_title, initial_dir):
    """
    打印提示信息，并弹出文件选择对话框
    :param prompt_msg: 终端显示的提示信息
    :param window_title: 文件选择窗口的标题
    :param initial_dir: 初始目录
    """
    print(prompt_msg)
    file_path = filedialog.askopenfilename(
        title=window_title,
        initialdir=initial_dir,
        filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
    )
    if file_path:
        print(f"✅ 选择成功：{os.path.basename(file_path)}\n")
    else:
        print("❌ 未选择文件或已取消\n")
    return file_path


def main():
    # 获取当前脚本所在目录
    current_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    print(f"📁 当前工作目录：{current_dir}\n")
    print("👇 请按提示依次选择所需文件：\n")

    # 1. 选择总表
    master_path = select_file(
        prompt_msg="👉 正在选择：案件研判线索台账（总表）",
        window_title="请选择总表文件（如：案件研判线索台账（共享）.xlsx）",
        initial_dir=current_dir
    )
    if not master_path:
        print("❌ 未选择总表文件，程序退出。")
        return

    # 2. 选择国反文件
    guofan_path = select_file(
        prompt_msg="👉 正在选择：国反.xlsx",
        window_title="请选择 国反.xlsx 文件",
        initial_dir=current_dir
    )
    if not guofan_path:
        print("❌ 未选择国反.xlsx文件，程序退出。")
        return

    # 3. 选择已退回文件
    huitui_path = select_file(
        prompt_msg="👉 正在选择：已退回.xlsx",
        window_title="请选择 已退回.xlsx 文件",
        initial_dir=current_dir
    )
    if not huitui_path:
        print("❌ 未选择已退回.xlsx文件，程序退出。")
        return

    # 4. 选择已下发文件
    xiafa_path = select_file(
        prompt_msg="👉 正在选择：已下发.xlsx",
        window_title="请选择 已下发.xlsx 文件",
        initial_dir=current_dir
    )
    if not xiafa_path:
        print("❌ 未选择已下发.xlsx文件，程序退出。")
        return

    # 检查所有文件是否存在（虽然已选中，但再确认一次）
    files_to_check = [
        (master_path, "总表"),
        (guofan_path, "国反.xlsx"),
        (huitui_path, "已退回.xlsx"),
        (xiafa_path, "已下发.xlsx")
    ]
    for path, name in files_to_check:
        if not os.path.exists(path):
            print(f"❌ 文件不存在：{name} -> {path}")
            return

    # 定义 sheet 映射关系：源文件 -> 总表中的 sheet 名
    file_to_sheet = {
        guofan_path: "案件基本信息（国反录入-每日全量更新）",
        huitui_path: "已退回协同线索",
        xiafa_path: "已下发协同线索"
    }

    # 加载总工作簿（保持样式）
    try:
        book = load_workbook(master_path)
    except Exception as e:
        print(f"❌ 无法打开总表文件：{e}")
        return

    print("\n" + "="*50)
    print("开始处理数据导入...")
    print("="*50)

    for src_file, sheet_name in file_to_sheet.items():
        if sheet_name not in book.sheetnames:
            print(f"❌ 总表中不存在该Sheet：{sheet_name}")
            continue

        print(f"\n📂 处理：{os.path.basename(src_file)} → '{sheet_name}'")

        # 读取源文件，跳过第一行（表头），保留后续所有数据
        try:
            df = pd.read_excel(src_file, skiprows=1, header=None)  # 第二行开始作为数据
        except Exception as e:
            print(f"❌ 读取源文件失败：{src_file}，错误：{e}")
            continue

        if df.empty:
            print(f"⚠️  源文件无数据（跳过）：{src_file}")
            continue

        # 获取目标 sheet
        ws = book[sheet_name]

        # 清空从第2行开始的所有数据（保留第1行表头）
        min_row = 2
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= min_row:
            for row_idx in range(max_row, min_row - 1, -1):
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).value = None
        print(f"✅ 已清空 '{sheet_name}' 的第2行至第{max_row}行数据")

        # 将新数据写入从第2行开始的位置
        start_row = 2
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                ws.cell(row=start_row + i, column=j + 1, value=value)

        print(f"✅ 已从 '{os.path.basename(src_file)}' 导入 {len(df)} 行数据到 '{sheet_name}'")

    # ========== 写入 VLOOKUP 公式 ==========
    print("\n" + "="*50)
    print("开始写入 VLOOKUP 公式...")
    print("="*50)

    target_sheets = [
        "已下发协同线索",
        "已退回协同线索"
    ]
    formula_col = 14  # N列

    for sheet_name in target_sheets:
        if sheet_name not in book.sheetnames:
            print(f"⚠️  跳过公式写入：Sheet 不存在 -> {sheet_name}")
            continue

        ws = book[sheet_name]
        max_row = ws.max_row

        if max_row < 2:
            print(f"⚠️  {sheet_name} 数据少于2行，跳过公式填充")
            continue

        # 从第2行开始写入公式
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=formula_col)
            formula = f'=VLOOKUP(LEFT(L{row_idx},13),案件编号对照表!A:B,2,0)'
            cell.value = formula

        print(f"✅ 已在 '{sheet_name}' 的 N列(第14列) 第2行至第{max_row}行写入公式")

    # 保存总表
    try:
        book.save(master_path)
        print(f"\n🎉 所有操作完成！")
        print(f"💾 文件已保存至：\n{master_path}")
    except Exception as e:
        print(f"❌ 保存文件失败：{e}")

if __name__ == "__main__":
    main()