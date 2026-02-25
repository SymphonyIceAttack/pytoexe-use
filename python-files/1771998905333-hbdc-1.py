import ctypes
import sys
import socket
import os
import datetime
import subprocess
from openpyxl import Workbook

def run_as_admin():
    try:
        if ctypes.windll.shell32.IsUserAnAdmin():
            return
    except:
        pass
    ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 1)
    sys.exit()

def get_hdd_serial():
    try:
        res = subprocess.check_output(
            "wmic diskdrive get serialnumber 2>nul",
            shell=True, text=True
        )
        lines = [l.strip() for l in res.strip().splitlines() if l.strip()]
        if len(lines) >= 2:
            return lines[1]
        return "获取失败"
    except:
        return "获取失败"

def get_local_ip():
    try:
        hostname = socket.gethostname()
        return socket.gethostbyname(hostname)
    except:
        return "获取失败"

def save_to_excel():
    serial = get_hdd_serial()
    ip = get_local_ip()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "硬盘信息"

    ws["A1"] = "获取时间"
    ws["B1"] = "硬盘序列号"
    ws["C1"] = "本机IP"

    ws["A2"] = now
    ws["B2"] = serial
    ws["C2"] = ip

    desktop = os.path.join(os.path.expanduser("Desktop"), "硬盘序列号.xlsx")
    wb.save(desktop)
    return desktop

if __name__ == "__main__":
    run_as_admin()
    print("===== 硬盘序列号采集工具 =====")
    try:
        path = save_to_excel()
        print(f"文件已保存到桌面：硬盘序列号.xlsx")
        print(f"硬盘序列号：{get_hdd_serial()}")
        print(f"本机IP地址：{get_local_ip()}")
    except Exception as e:
        print(f"出错：{e}")
    print("\n按回车键退出...")
    input()
