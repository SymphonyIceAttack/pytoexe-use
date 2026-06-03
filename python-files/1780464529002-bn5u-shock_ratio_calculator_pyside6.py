import csv
import math
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

from matplotlib import rcParams
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from PySide6.QtCore import QAbstractTableModel, QModelIndex, QTimer, Qt
from PySide6.QtGui import QPixmap
from PySide6.QtWidgets import (
    QApplication,
    QAbstractItemView,
    QDoubleSpinBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSlider,
    QSpinBox,
    QSplitter,
    QTableView,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)


APP_TITLE = "后减震器压缩量-后轮跳比计算器"
COMPANY_NAME = "九号科技有限公司"
LOGO_FILE = "企业微信截图_17804561957209.png"
DEFAULT_WINDOW_WIDTH = 1200
DEFAULT_WINDOW_HEIGHT = 960
BG = "#f4f4f2"
FG = "#171717"
GRID = "#d7d5d1"
BLUE = "#111111"
CYAN = "#6b7280"
PINK = "#d1495b"
GREEN = "#2f855a"
AMBER = "#f97316"

rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
rcParams["axes.unicode_minus"] = False


def resource_path(filename: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    return base / filename


def auto_export_dir() -> Path:
    return Path.home() / "Documents" / "ShockRatioCalculator" / "auto_exports"


@dataclass(slots=True)
class Point:
    x: float
    y: float


@dataclass(slots=True)
class GeometryParams:
    w0: Point
    p0: Point
    a: Point
    cmax: float
    steps: int
    theta_search_deg: float
    diagram_c: float


@dataclass(slots=True)
class MechanismState:
    theta_rad: float
    theta_deg: float
    w: Point
    p: Point
    shock_len: float
    shock_comp: float
    wheel_jump: float
    included_angle: float
    shock_angle_deg: float
    shock_angle_change_deg: float


@dataclass(slots=True)
class Row:
    i: int
    shock_comp: float
    theta_deg: float = math.nan
    wheel_jump: float = math.nan
    shock_len: float = math.nan
    px: float = math.nan
    py: float = math.nan
    wx: float = math.nan
    wy: float = math.nan
    avg_wheel_to_shock: float = math.nan
    avg_shock_to_wheel: float = math.nan
    instant_wheel_to_shock: float = math.nan
    instant_shock_to_wheel: float = math.nan
    included_angle: float = math.nan
    shock_angle_deg: float = math.nan
    shock_angle_change_deg: float = math.nan
    lever_ratio_change_rate: float = math.nan
    lever_ratio_change_percent: float = math.nan


@dataclass(slots=True)
class SchemeRecord:
    name: str
    params: GeometryParams
    created_at: str


def is_finite(value: float) -> bool:
    return math.isfinite(value)


def fmt(value: float, digits: int = 3) -> str:
    if not is_finite(value):
        return "-"
    return f"{value:.{digits}f}"


def distance(p1: Point, p2: Point) -> float:
    return math.hypot(p1.x - p2.x, p1.y - p2.y)


def rotate_point(p: Point, theta_rad: float) -> Point:
    c = math.cos(theta_rad)
    s = math.sin(theta_rad)
    return Point(p.x * c - p.y * s, p.x * s + p.y * c)


def dot(a: Point, b: Point) -> float:
    return a.x * b.x + a.y * b.y


def length(v: Point) -> float:
    return math.hypot(v.x, v.y)


def normalize_angle_deg(angle: float) -> float:
    while angle > 180:
        angle -= 360
    while angle < -180:
        angle += 360
    return angle


def shock_abs_angle_deg(p: Point, a: Point) -> float:
    return math.degrees(math.atan2(a.y - p.y, a.x - p.x))


def shock_swingarm_angle_deg(o: Point, p: Point, a: Point) -> float:
    v1 = Point(o.x - p.x, o.y - p.y)
    v2 = Point(a.x - p.x, a.y - p.y)
    l1 = length(v1)
    l2 = length(v2)
    if l1 == 0 or l2 == 0:
        return math.nan
    c = dot(v1, v2) / (l1 * l2)
    c = max(-1.0, min(1.0, c))
    return math.degrees(math.acos(c))


def state_at_theta(
    theta_rad: float,
    w0: Point,
    p0: Point,
    a: Point,
    l0: float,
    shock_angle0_deg: Optional[float] = None,
) -> MechanismState:
    o = Point(0, 0)
    w = rotate_point(w0, theta_rad)
    p = rotate_point(p0, theta_rad)
    shock_len = distance(a, p)
    shock_comp = l0 - shock_len
    wheel_jump = w.y - w0.y
    included_angle = shock_swingarm_angle_deg(o, p, a)
    shock_angle_deg = shock_abs_angle_deg(p, a)
    if shock_angle0_deg is None:
        shock_angle_change_deg = 0.0
    else:
        shock_angle_change_deg = normalize_angle_deg(shock_angle_deg - shock_angle0_deg)

    return MechanismState(
        theta_rad=theta_rad,
        theta_deg=math.degrees(theta_rad),
        w=w,
        p=p,
        shock_len=shock_len,
        shock_comp=shock_comp,
        wheel_jump=wheel_jump,
        included_angle=included_angle,
        shock_angle_deg=shock_angle_deg,
        shock_angle_change_deg=shock_angle_change_deg,
    )


def solve_theta_by_compression(
    target_c: float,
    w0: Point,
    p0: Point,
    a: Point,
    l0: float,
    theta_max_rad: float,
    shock_angle0_deg: Optional[float] = None,
) -> Optional[MechanismState]:
    if abs(target_c) < 1e-9:
        return state_at_theta(0, w0, p0, a, l0, shock_angle0_deg)

    scan_n = 1500
    prev_theta = 0.0
    prev_state = state_at_theta(prev_theta, w0, p0, a, l0, shock_angle0_deg)
    prev_f = prev_state.shock_comp - target_c

    for i in range(1, scan_n + 1):
        theta = theta_max_rad * i / scan_n
        st = state_at_theta(theta, w0, p0, a, l0, shock_angle0_deg)
        f = st.shock_comp - target_c

        if abs(f) < 1e-8:
            return st

        if prev_f * f < 0:
            lo = prev_theta
            hi = theta
            f_lo = prev_f

            for _ in range(80):
                mid = (lo + hi) / 2
                mid_state = state_at_theta(mid, w0, p0, a, l0, shock_angle0_deg)
                f_mid = mid_state.shock_comp - target_c
                if abs(f_mid) < 1e-10:
                    return mid_state
                if f_lo * f_mid <= 0:
                    hi = mid
                else:
                    lo = mid
                    f_lo = f_mid

            return state_at_theta((lo + hi) / 2, w0, p0, a, l0, shock_angle0_deg)

        prev_theta = theta
        prev_f = f

    return None


def calculate_rows(params: GeometryParams) -> Tuple[List[Row], float, float, int]:
    l0 = distance(params.a, params.p0)
    if l0 <= 0:
        raise ValueError("减震器上安装点 A 与下安装点 P 不能重合。")
    if params.steps < 1:
        raise ValueError("计算步数必须大于 0。")

    theta_search_rad = math.radians(params.theta_search_deg)
    shock_angle0_deg = shock_abs_angle_deg(params.p0, params.a)
    rows: List[Row] = []
    failed_count = 0

    for i in range(params.steps + 1):
        target_c = params.cmax * i / params.steps
        state = solve_theta_by_compression(
            target_c,
            params.w0,
            params.p0,
            params.a,
            l0,
            theta_search_rad,
            shock_angle0_deg,
        )

        if state is None:
            failed_count += 1
            rows.append(Row(i=i, shock_comp=target_c))
            continue

        avg_wheel_to_shock = math.nan if target_c == 0 else state.wheel_jump / target_c
        avg_shock_to_wheel = math.nan if state.wheel_jump == 0 else target_c / state.wheel_jump
        rows.append(
            Row(
                i=i,
                shock_comp=target_c,
                theta_deg=state.theta_deg,
                wheel_jump=state.wheel_jump,
                shock_len=state.shock_len,
                px=state.p.x,
                py=state.p.y,
                wx=state.w.x,
                wy=state.w.y,
                avg_wheel_to_shock=avg_wheel_to_shock,
                avg_shock_to_wheel=avg_shock_to_wheel,
                included_angle=state.included_angle,
                shock_angle_deg=state.shock_angle_deg,
                shock_angle_change_deg=state.shock_angle_change_deg,
            )
        )

    for i, row in enumerate(rows):
        if i == 0 and len(rows) > 1:
            left = rows[0]
            right = rows[1]
        elif i == len(rows) - 1 and len(rows) > 1:
            left = rows[i - 1]
            right = rows[i]
        elif 0 < i < len(rows) - 1:
            left = rows[i - 1]
            right = rows[i + 1]
        else:
            continue

        if is_finite(left.wheel_jump) and is_finite(right.wheel_jump):
            dj = right.wheel_jump - left.wheel_jump
            dc = right.shock_comp - left.shock_comp
            row.instant_wheel_to_shock = math.nan if dc == 0 else dj / dc
            row.instant_shock_to_wheel = math.nan if dj == 0 else dc / dj

    for i, row in enumerate(rows):
        if i == 0 and len(rows) > 1:
            left = rows[0]
            right = rows[1]
        elif i == len(rows) - 1 and len(rows) > 1:
            left = rows[i - 1]
            right = rows[i]
        elif 0 < i < len(rows) - 1:
            left = rows[i - 1]
            right = rows[i + 1]
        else:
            continue

        if is_finite(left.instant_wheel_to_shock) and is_finite(right.instant_wheel_to_shock):
            dlr = right.instant_wheel_to_shock - left.instant_wheel_to_shock
            dc = right.shock_comp - left.shock_comp
            row.lever_ratio_change_rate = math.nan if dc == 0 else dlr / dc

    first_lr = next((r.instant_wheel_to_shock for r in rows if is_finite(r.instant_wheel_to_shock)), math.nan)
    for row in rows:
        if is_finite(first_lr) and first_lr != 0 and is_finite(row.instant_wheel_to_shock):
            row.lever_ratio_change_percent = (row.instant_wheel_to_shock - first_lr) / first_lr * 100

    return rows, l0, shock_angle0_deg, failed_count


TABLE_COLUMNS = [
    ("步", lambda r: str(r.i)),
    ("C mm", lambda r: fmt(r.shock_comp, 3)),
    ("θ °", lambda r: fmt(r.theta_deg, 4)),
    ("J mm", lambda r: fmt(r.wheel_jump, 3)),
    ("平均轮跳比", lambda r: fmt(r.avg_wheel_to_shock, 4)),
    ("平均压缩比", lambda r: fmt(r.avg_shock_to_wheel, 4)),
    ("瞬时杠杆比", lambda r: fmt(r.instant_wheel_to_shock, 4)),
    ("瞬时压缩比", lambda r: fmt(r.instant_shock_to_wheel, 4)),
    ("夹角 °", lambda r: fmt(r.included_angle, 3)),
    ("α °", lambda r: fmt(r.shock_angle_deg, 3)),
    ("Δα °", lambda r: fmt(r.shock_angle_change_deg, 3)),
    ("杠杆比变化率", lambda r: fmt(r.lever_ratio_change_rate, 6)),
    ("杠杆比变化 %", lambda r: fmt(r.lever_ratio_change_percent, 3)),
    ("当前长度 mm", lambda r: fmt(r.shock_len, 3)),
    ("下安装点横坐标", lambda r: fmt(r.px, 3)),
    ("下安装点纵坐标", lambda r: fmt(r.py, 3)),
    ("后轮轴横坐标", lambda r: fmt(r.wx, 3)),
    ("后轮轴纵坐标", lambda r: fmt(r.wy, 3)),
]


PARAM_COLUMNS = [
    ("后轮轴横坐标", lambda p: fmt(p.w0.x, 3)),
    ("后轮轴纵坐标", lambda p: fmt(p.w0.y, 3)),
    ("减震器下安装点横坐标", lambda p: fmt(p.p0.x, 3)),
    ("减震器下安装点纵坐标", lambda p: fmt(p.p0.y, 3)),
    ("减震器上安装点横坐标", lambda p: fmt(p.a.x, 3)),
    ("减震器上安装点纵坐标", lambda p: fmt(p.a.y, 3)),
    ("最大压缩量", lambda p: fmt(p.cmax, 3)),
    ("计算步数", lambda p: str(p.steps)),
    ("最大搜索角", lambda p: fmt(p.theta_search_deg, 3)),
    ("示意压缩量", lambda p: fmt(p.diagram_c, 3)),
]


def params_key(params: GeometryParams) -> Tuple[float, ...]:
    return (
        round(params.w0.x, 6),
        round(params.w0.y, 6),
        round(params.p0.x, 6),
        round(params.p0.y, 6),
        round(params.a.x, 6),
        round(params.a.y, 6),
        round(params.cmax, 6),
        float(params.steps),
        round(params.theta_search_deg, 6),
        round(params.diagram_c, 6),
    )


class ResultTableModel(QAbstractTableModel):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.rows: List[Row] = []

    def set_rows(self, rows: List[Row]) -> None:
        self.beginResetModel()
        self.rows = rows
        self.endResetModel()

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return 0 if parent.isValid() else len(self.rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return 0 if parent.isValid() else len(TABLE_COLUMNS)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.ItemDataRole.DisplayRole:
            return TABLE_COLUMNS[index.column()][1](self.rows[index.row()])
        if role == Qt.ItemDataRole.TextAlignmentRole:
            return Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        return None

    def headerData(  # noqa: N802
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.ItemDataRole.DisplayRole,
    ):
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            return TABLE_COLUMNS[section][0]
        return str(section + 1)


class MetricCard(QFrame):
    def __init__(self, title: str, value: str = "-", parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("MetricCard")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 12)
        layout.setSpacing(4)

        self.title_label = QLabel(title)
        self.title_label.setObjectName("MetricTitle")
        self.value_label = QLabel(value)
        self.value_label.setObjectName("MetricValue")
        self.value_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)

        layout.addWidget(self.title_label)
        layout.addWidget(self.value_label)

    def set_value(self, value: str) -> None:
        self.value_label.setText(value)


class MplCanvas(FigureCanvas):
    def __init__(self, min_height: int, parent: Optional[QWidget] = None) -> None:
        self.figure = Figure(facecolor=BG, tight_layout=True)
        self.ax = self.figure.add_subplot(111)
        super().__init__(self.figure)
        self.setParent(parent)
        self.setMinimumHeight(min_height)

    def style_axes(self, xlabel: str = "", ylabel: str = "") -> None:
        self.ax.set_facecolor(BG)
        self.ax.grid(True, color=GRID, alpha=0.55, linewidth=0.8)
        self.ax.tick_params(colors="#56677f")
        self.ax.set_xlabel(xlabel, color=FG)
        self.ax.set_ylabel(ylabel, color=FG)
        for spine in self.ax.spines.values():
            spine.set_color(GRID)

    def show_message(self, text: str) -> None:
        self.ax.clear()
        self.style_axes()
        self.ax.text(0.5, 0.5, text, transform=self.ax.transAxes, ha="center", va="center", color="#7b879f")
        self.ax.set_xticks([])
        self.ax.set_yticks([])
        self.draw_idle()


class MechanismView(MplCanvas):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(420, parent)

    def set_scene(
        self,
        params: Optional[GeometryParams],
        state: Optional[MechanismState],
        l0: float,
        target_c: float,
    ) -> None:
        if params is None or state is None:
            self.show_message("等待计算")
            return

        self.ax.clear()
        self.style_axes("X / mm", "Y / mm")
        o, w0, p0, a, w, p = Point(0, 0), params.w0, params.p0, params.a, state.w, state.p

        self._line(o, w0, "#516070", "--", 1.4, "初始平叉")
        self._line(o, p0, "#516070", "--", 1.4)
        self._line(a, p0, "#516070", "--", 1.4, "初始减震")
        self._line(o, w, BLUE, "-", 3.2, "当前平叉")
        self._line(o, p, CYAN, "-", 2.6)
        self._line(a, p, PINK, "-", 3.2, "当前减震")
        self.ax.annotate("", xy=(w.x, w.y), xytext=(w.x, w0.y), arrowprops={"arrowstyle": "->", "color": GREEN, "lw": 2.2})

        for label, point, color in [("O", o, FG), ("A", a, PINK), ("W0", w0, "#77849a"), ("P0", p0, "#77849a"), ("W", w, BLUE), ("P", p, BLUE)]:
            self.ax.scatter(point.x, point.y, s=52, color=color, edgecolor="#ffffff", linewidth=1.2, zorder=4)
            self.ax.text(point.x, point.y, f" {label}", color=color, fontsize=9, va="bottom")

        c = l0 - state.shock_len
        j = state.wheel_jump
        info = (
            f"C = {fmt(c, 2)} mm\nJ = {fmt(j, 2)} mm\n"
            f"J/C = {fmt(math.nan if c == 0 else j / c, 3)}\n"
            f"θ = {fmt(state.theta_deg, 2)}°   α = {fmt(state.shock_angle_deg, 2)}°"
        )
        self.ax.text(0.02, 0.98, info, transform=self.ax.transAxes, va="top", color=FG, fontsize=9, bbox={"facecolor": "#ffffff", "edgecolor": GRID, "alpha": 0.92})
        self._fit([o, w0, p0, a, w, p])
        self.ax.legend(loc="lower right", frameon=False, labelcolor=FG)
        self.draw_idle()

    def _line(self, a: Point, b: Point, color: str, style: str, width: float, label: Optional[str] = None) -> None:
        self.ax.plot([a.x, b.x], [a.y, b.y], linestyle=style, color=color, linewidth=width, solid_capstyle="round", label=label)

    def _fit(self, pts: List[Point]) -> None:
        xs = [p.x for p in pts]
        ys = [p.y for p in pts]
        dx = max(1.0, max(xs) - min(xs))
        dy = max(1.0, max(ys) - min(ys))
        self.ax.set_xlim(min(xs) - dx * 0.14, max(xs) + dx * 0.14)
        self.ax.set_ylim(min(ys) - dy * 0.18, max(ys) + dy * 0.18)
        self.ax.set_aspect("equal", adjustable="box")


class RatioChart(MplCanvas):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(320, parent)

    def set_rows(self, rows: List[Row]) -> None:
        avg = [r for r in rows if is_finite(r.avg_wheel_to_shock) and r.shock_comp > 0]
        ins = [r for r in rows if is_finite(r.instant_wheel_to_shock)]
        if len(avg) < 2:
            self.show_message("暂无有效曲线数据")
            return

        self.ax.clear()
        self.style_axes("减震器压缩量 C / mm", "轮跳比 / 杠杆比")
        self.ax.plot([r.shock_comp for r in avg], [r.avg_wheel_to_shock for r in avg], color=BLUE, marker="o", linewidth=2.4, markersize=4, label="平均轮跳比")
        self.ax.plot([r.shock_comp for r in ins], [r.instant_wheel_to_shock for r in ins], color=AMBER, marker="s", linewidth=2.0, markersize=3.5, label="瞬时杠杆比")
        self.ax.fill_between([r.shock_comp for r in avg], [r.avg_wheel_to_shock for r in avg], color=BLUE, alpha=0.08)
        self.ax.legend(loc="best", frameon=False, labelcolor=FG)
        self.draw_idle()


class CalculatorWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(DEFAULT_WINDOW_WIDTH, DEFAULT_WINDOW_HEIGHT)

        self.rows: List[Row] = []
        self.params: Optional[GeometryParams] = None
        self.l0 = 0.0
        self.shock_angle0_deg = 0.0
        self.anim_progress = 1.0
        self.anim_started_at = 0.0
        self.updating_slider = False
        self.updating_window_size = False
        self.loading_scheme = False
        self.current_params_saved = True
        self.last_saved_params_key: Optional[Tuple[float, ...]] = None
        self.scheme_records: List[SchemeRecord] = []

        self.timer = QTimer(self)
        self.timer.setInterval(16)
        self.timer.timeout.connect(self._animation_tick)

        self._build_ui()
        self._apply_style()
        self.calculate()
        self.last_saved_params_key = params_key(self.current_params())
        self.current_params_saved = True
        self.update_scheme_status()

    def _build_ui(self) -> None:
        root = QWidget()
        root_layout = QVBoxLayout(root)
        root_layout.setContentsMargins(18, 16, 18, 18)
        root_layout.setSpacing(12)
        self.setCentralWidget(root)

        header = self._build_header()
        root_layout.addWidget(header)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)
        root_layout.addWidget(splitter, 1)

        left = QWidget()
        left.setObjectName("SidePanel")
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(16, 16, 16, 16)
        left_layout.setSpacing(14)
        splitter.addWidget(left)

        scheme_panel = QFrame()
        scheme_panel.setObjectName("SchemePanel")
        scheme_layout = QVBoxLayout(scheme_panel)
        scheme_layout.setContentsMargins(12, 12, 12, 12)
        scheme_layout.setSpacing(8)
        scheme_title = QLabel("方案记录")
        scheme_title.setObjectName("PanelTitle")
        self.scheme_list = QListWidget()
        self.scheme_list.setObjectName("SchemeList")
        self.scheme_list.itemDoubleClicked.connect(self.load_scheme_from_item)
        self.save_scheme_btn = QPushButton("保存当前参数")
        self.save_scheme_btn.clicked.connect(self.save_current_scheme)
        self.scheme_status = QLabel("暂无方案记录。输入参数后点击“保存当前参数”生成记录。")
        self.scheme_status.setObjectName("SchemeStatus")
        self.scheme_status.setWordWrap(True)
        scheme_layout.addWidget(scheme_title)
        scheme_layout.addWidget(self.scheme_list)
        scheme_layout.addWidget(self.save_scheme_btn)
        scheme_layout.addWidget(self.scheme_status)
        left_layout.addWidget(scheme_panel)

        form = QGridLayout()
        form.setHorizontalSpacing(10)
        form.setVerticalSpacing(10)
        left_layout.addLayout(form)

        self.wx = self._double_spin(520)
        self.wy = self._double_spin(0)
        self.px = self._double_spin(260)
        self.py = self._double_spin(20)
        self.ax = self._double_spin(120)
        self.ay = self._double_spin(300)
        self.cmax = self._double_spin(45, maximum=10000)
        self.steps = QSpinBox()
        self.steps.setRange(1, 1000)
        self.steps.setValue(30)
        self.theta_search = self._double_spin(45, minimum=-180, maximum=180)
        self.diagram_c = self._double_spin(45, maximum=10000)

        fields = [
            ("后轮轴横坐标", self.wx),
            ("后轮轴纵坐标", self.wy),
            ("下安装点横坐标", self.px),
            ("下安装点纵坐标", self.py),
            ("上安装点横坐标", self.ax),
            ("上安装点纵坐标", self.ay),
            ("最大压缩量", self.cmax),
            ("计算步数", self.steps),
            ("最大搜索角 θ", self.theta_search),
            ("示意压缩量 C", self.diagram_c),
        ]
        for row, (label_text, widget) in enumerate(fields):
            label = QLabel(label_text)
            label.setObjectName("FormLabel")
            form.addWidget(label, row, 0)
            form.addWidget(widget, row, 1)
        for widget in [self.wx, self.wy, self.px, self.py, self.ax, self.ay, self.cmax, self.steps, self.theta_search, self.diagram_c]:
            widget.valueChanged.connect(self.mark_params_unsaved)

        action_layout = QGridLayout()
        action_layout.setHorizontalSpacing(10)
        left_layout.addLayout(action_layout)

        self.calculate_btn = QPushButton("计算")
        self.calculate_btn.setObjectName("PrimaryButton")
        self.calculate_btn.clicked.connect(self.calculate)
        action_layout.addWidget(self.calculate_btn, 0, 0)

        self.export_btn = QPushButton("导出 CSV")
        self.export_btn.clicked.connect(self.export_csv)
        action_layout.addWidget(self.export_btn, 0, 1)

        metric_grid = QGridLayout()
        metric_grid.setHorizontalSpacing(10)
        metric_grid.setVerticalSpacing(10)
        left_layout.addLayout(metric_grid)

        self.metric_l0 = MetricCard("初始长度 L0")
        self.metric_jump = MetricCard("最大轮跳 J")
        self.metric_avg = MetricCard("平均轮跳比")
        self.metric_lr = MetricCard("终点杠杆比")
        for i, card in enumerate([self.metric_l0, self.metric_jump, self.metric_avg, self.metric_lr]):
            metric_grid.addWidget(card, i // 2, i % 2)

        self.summary = QLabel()
        self.summary.setObjectName("Summary")
        self.summary.setWordWrap(True)
        self.summary.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        left_layout.addWidget(self.summary, 1)

        control = QFrame()
        control.setObjectName("ControlPanel")
        control_layout = QGridLayout(control)
        control_layout.setContentsMargins(12, 12, 12, 12)
        control_layout.setHorizontalSpacing(8)
        control_layout.setVerticalSpacing(8)
        left_layout.addWidget(control)

        self.play_btn = QPushButton("播放")
        self.pause_btn = QPushButton("暂停")
        self.reset_btn = QPushButton("复位")
        self.play_btn.clicked.connect(self.play_animation)
        self.pause_btn.clicked.connect(self.pause_animation)
        self.reset_btn.clicked.connect(self.reset_animation)
        control_layout.addWidget(self.play_btn, 0, 0)
        control_layout.addWidget(self.pause_btn, 0, 1)
        control_layout.addWidget(self.reset_btn, 0, 2)

        self.speed_slider = QSlider(Qt.Orientation.Horizontal)
        self.speed_slider.setRange(30, 300)
        self.speed_slider.setValue(100)
        self.speed_label = QLabel("速度 1.0x")
        self.speed_slider.valueChanged.connect(self._update_speed_label)
        control_layout.addWidget(self.speed_label, 1, 0)
        control_layout.addWidget(self.speed_slider, 1, 1, 1, 2)

        self.progress_slider = QSlider(Qt.Orientation.Horizontal)
        self.progress_slider.setRange(0, 1000)
        self.progress_slider.setValue(1000)
        self.progress_slider.valueChanged.connect(self._manual_progress_changed)
        control_layout.addWidget(QLabel("行程"), 2, 0)
        control_layout.addWidget(self.progress_slider, 2, 1, 1, 2)

        self.window_width = self._int_spin(DEFAULT_WINDOW_WIDTH, 120, 3840, " px")
        self.window_height = self._int_spin(DEFAULT_WINDOW_HEIGHT, 480, 2160, " px")
        self.apply_size_btn = QPushButton("应用尺寸")
        self.apply_size_btn.clicked.connect(self.apply_window_size)
        control_layout.addWidget(QLabel("窗口宽度"), 3, 0)
        control_layout.addWidget(self.window_width, 3, 1, 1, 2)
        control_layout.addWidget(QLabel("窗口高度"), 4, 0)
        control_layout.addWidget(self.window_height, 4, 1, 1, 2)
        control_layout.addWidget(self.apply_size_btn, 5, 0, 1, 3)

        self.anim_info = QLabel("C = -")
        self.anim_info.setObjectName("AnimInfo")
        self.anim_info.setWordWrap(True)
        left_layout.addWidget(self.anim_info)

        tabs = QTabWidget()
        tabs.setObjectName("MainTabs")
        splitter.addWidget(tabs)
        splitter.setSizes([380, 940])

        self.mechanism_view = MechanismView()
        self.chart = RatioChart()
        self.table_model = ResultTableModel(self)
        self.table = QTableView()
        self.table.setModel(self.table_model)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)

        tabs.addTab(self.mechanism_view, "机构示意")
        tabs.addTab(self.chart, "比值曲线")
        tabs.addTab(self.table, "计算表格")

    def _build_header(self) -> QWidget:
        header = QFrame()
        header.setObjectName("Header")
        layout = QHBoxLayout(header)
        layout.setContentsMargins(8, 6, 8, 10)
        layout.setSpacing(10)

        logo_label = QLabel()
        logo_label.setObjectName("CompanyLogo")
        logo_path = resource_path(LOGO_FILE)
        pixmap = QPixmap(str(logo_path))
        if not pixmap.isNull():
            logo_label.setPixmap(
                pixmap.scaled(42, 42, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            )
        else:
            logo_label.setText("9")
            logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_label.setFixedSize(46, 46)

        text_box = QWidget()
        text_layout = QVBoxLayout(text_box)
        text_layout.setContentsMargins(0, 0, 0, 0)
        text_layout.setSpacing(1)
        company = QLabel(COMPANY_NAME)
        company.setObjectName("CompanyName")
        title = QLabel(APP_TITLE)
        title.setObjectName("HeaderTitle")
        text_layout.addWidget(company)
        text_layout.addWidget(title)

        layout.addWidget(logo_label)
        layout.addWidget(text_box, 1)
        return header

    def _double_spin(
        self,
        value: float,
        minimum: float = -100000,
        maximum: float = 100000,
    ) -> QDoubleSpinBox:
        spin = QDoubleSpinBox()
        spin.setRange(minimum, maximum)
        spin.setDecimals(3)
        spin.setSingleStep(1)
        spin.setValue(value)
        spin.setSuffix(" mm")
        return spin

    def _int_spin(self, value: int, minimum: int, maximum: int, suffix: str = "") -> QSpinBox:
        spin = QSpinBox()
        spin.setRange(minimum, maximum)
        spin.setSingleStep(20)
        spin.setValue(value)
        spin.setSuffix(suffix)
        return spin

    def _apply_style(self) -> None:
        self.setStyleSheet(
            """
            QMainWindow {
                background: #f4f4f2;
                color: #171717;
                font-family: "Microsoft YaHei UI", "Segoe UI", sans-serif;
            }
            QLabel { color: #171717; }
            #Header {
                background: #ffffff;
                border: 1px solid #d7d5d1;
                border-radius: 8px;
            }
            #CompanyLogo {
                border: 1px solid #d7d5d1;
                border-radius: 8px;
                background: #ffffff;
            }
            #CompanyName {
                color: #111111;
                font-size: 15px;
                font-weight: 700;
            }
            #HeaderTitle {
                color: #555555;
                font-size: 12px;
            }
            #SidePanel, #ControlPanel, #MetricCard, #SchemePanel {
                background: #ffffff;
                border: 1px solid #d7d5d1;
                border-radius: 8px;
            }
            #PanelTitle {
                color: #111111;
                font-size: 13px;
                font-weight: 700;
            }
            #SchemeList {
                min-height: 120px;
                background: #f7f7f5;
                color: #171717;
                border: 1px solid #dad8d4;
                border-radius: 6px;
                padding: 4px;
            }
            #SchemeList::item {
                padding: 8px;
                border-radius: 4px;
            }
            #SchemeList::item:selected {
                color: #ffffff;
                background: #111111;
            }
            #SchemeStatus {
                color: #777777;
                font-size: 11px;
                line-height: 1.4;
            }
            #FormLabel {
                color: #666666;
                font-size: 12px;
            }
            QDoubleSpinBox, QSpinBox {
                min-height: 30px;
                color: #171717;
                background: #ffffff;
                border: 1px solid #d7d5d1;
                border-radius: 6px;
                padding: 2px 8px;
                selection-background-color: #111111;
            }
            QDoubleSpinBox:focus, QSpinBox:focus {
                border-color: #111111;
            }
            QPushButton {
                min-height: 30px;
                border: 1px solid #d7d5d1;
                border-radius: 6px;
                color: #171717;
                background: #ffffff;
                padding: 4px 12px;
            }
            QPushButton:hover {
                border-color: #111111;
                background: #f2f2f0;
            }
            #PrimaryButton {
                background: #111111;
                border-color: #111111;
                color: white;
                font-weight: 700;
            }
            #MetricTitle {
                color: #666666;
                font-size: 11px;
            }
            #MetricValue {
                color: #111111;
                font-size: 18px;
                font-weight: 700;
            }
            #Summary, #AnimInfo {
                color: #555555;
                line-height: 1.5;
            }
            QSlider::groove:horizontal {
                height: 4px;
                background: #dad8d4;
                border-radius: 2px;
            }
            QSlider::sub-page:horizontal {
                background: #111111;
                border-radius: 2px;
            }
            QSlider::handle:horizontal {
                width: 14px;
                height: 14px;
                margin: -5px 0;
                border-radius: 7px;
                background: #ffffff;
                border: 1px solid #111111;
            }
            QTabWidget::pane {
                border: 1px solid #d7d5d1;
                border-radius: 8px;
                background: #ffffff;
            }
            QTabBar::tab {
                background: #ecebe8;
                color: #666666;
                padding: 9px 18px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                color: #111111;
                background: #ffffff;
                border-bottom: 2px solid #111111;
            }
            QTableView {
                background: #ffffff;
                alternate-background-color: #f7f7f5;
                color: #171717;
                gridline-color: #dad8d4;
                border: none;
                selection-background-color: #eeeeea;
            }
            QHeaderView::section {
                background: #ecebe8;
                color: #171717;
                border: 1px solid #dad8d4;
                padding: 6px;
            }
            """
        )

    def current_params(self) -> GeometryParams:
        return GeometryParams(
            w0=Point(self.wx.value(), self.wy.value()),
            p0=Point(self.px.value(), self.py.value()),
            a=Point(self.ax.value(), self.ay.value()),
            cmax=self.cmax.value(),
            steps=self.steps.value(),
            theta_search_deg=self.theta_search.value(),
            diagram_c=self.diagram_c.value(),
        )

    def set_params(self, params: GeometryParams) -> None:
        self.loading_scheme = True
        self.wx.setValue(params.w0.x)
        self.wy.setValue(params.w0.y)
        self.px.setValue(params.p0.x)
        self.py.setValue(params.p0.y)
        self.ax.setValue(params.a.x)
        self.ay.setValue(params.a.y)
        self.cmax.setValue(params.cmax)
        self.steps.setValue(params.steps)
        self.theta_search.setValue(params.theta_search_deg)
        self.diagram_c.setValue(params.diagram_c)
        self.loading_scheme = False
        self.mark_params_unsaved()

    def mark_params_unsaved(self, *_) -> None:
        if self.loading_scheme:
            return
        current_key = params_key(self.current_params())
        self.current_params_saved = current_key == self.last_saved_params_key
        self.update_scheme_status()

    def save_current_scheme(self) -> None:
        params = self.current_params()
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        name = f"方案 {len(self.scheme_records) + 1}"
        record = SchemeRecord(name=name, params=params, created_at=created_at)
        self.scheme_records.append(record)
        item = QListWidgetItem(f"{name}  {created_at}\n{self.scheme_summary(params)}")
        item.setToolTip(self.scheme_summary(params))
        item.setData(Qt.ItemDataRole.UserRole, len(self.scheme_records) - 1)
        self.scheme_list.addItem(item)
        self.scheme_list.setCurrentItem(item)
        self.last_saved_params_key = params_key(params)
        self.current_params_saved = True
        self.update_scheme_status()

    def load_scheme_from_item(self, item: QListWidgetItem) -> None:
        index = item.data(Qt.ItemDataRole.UserRole)
        if index is None or index < 0 or index >= len(self.scheme_records):
            return
        params = self.scheme_records[index].params
        self.set_params(params)
        self.last_saved_params_key = params_key(params)
        self.current_params_saved = True
        self.update_scheme_status()
        self.calculate()

    def scheme_summary(self, params: GeometryParams) -> str:
        return (
            f"后轮轴=({fmt(params.w0.x, 1)}, {fmt(params.w0.y, 1)})；"
            f"下安装点=({fmt(params.p0.x, 1)}, {fmt(params.p0.y, 1)})；"
            f"上安装点=({fmt(params.a.x, 1)}, {fmt(params.a.y, 1)})；"
            f"最大压缩量={fmt(params.cmax, 1)}；步数={params.steps}"
        )

    def update_scheme_status(self) -> None:
        if self.current_params_saved:
            if self.scheme_records:
                self.scheme_status.setText(f"当前参数已保存，已记录 {len(self.scheme_records)} 组方案参数。")
            else:
                self.scheme_status.setText("暂无方案记录。输入参数后点击“保存当前参数”生成记录。")
        else:
            self.scheme_status.setText("当前参数尚未保存到方案记录，关闭软件时将自动导出 Excel。")

    def calculate(self) -> bool:
        self.pause_animation()
        try:
            self.params = self.current_params()
            self.rows, self.l0, self.shock_angle0_deg, failed_count = calculate_rows(self.params)
        except Exception as exc:
            self.rows = []
            self.table_model.set_rows([])
            QMessageBox.warning(self, "计算失败", str(exc))
            return False

        self._update_metrics(failed_count)
        self.chart.set_rows(self.rows)
        self._fill_table()

        progress = 0.0 if self.params.cmax == 0 else self.params.diagram_c / self.params.cmax
        self.anim_progress = max(0.0, min(1.0, progress))
        self._set_progress_slider(self.anim_progress)
        self._render_progress(self.anim_progress)
        return True

    def _update_metrics(self, failed_count: int) -> None:
        valid_rows = [r for r in self.rows if is_finite(r.theta_deg)]
        if not valid_rows:
            self.metric_l0.set_value(f"{fmt(self.l0, 2)} mm")
            self.metric_jump.set_value("-")
            self.metric_avg.set_value("-")
            self.metric_lr.set_value("-")
            self.summary.setText("没有找到有效解。请检查坐标、最大压缩量或最大搜索角。")
            return

        last = valid_rows[-1]
        first_lr = next((r.instant_wheel_to_shock for r in valid_rows if is_finite(r.instant_wheel_to_shock)), math.nan)
        total_lr_change = math.nan
        if is_finite(first_lr) and first_lr != 0 and is_finite(last.instant_wheel_to_shock):
            total_lr_change = (last.instant_wheel_to_shock - first_lr) / first_lr * 100

        self.metric_l0.set_value(f"{fmt(self.l0, 2)} mm")
        self.metric_jump.set_value(f"{fmt(last.wheel_jump, 2)} mm")
        self.metric_avg.set_value(fmt(last.avg_wheel_to_shock, 3))
        self.metric_lr.set_value(fmt(last.instant_wheel_to_shock, 3))

        lines = [
            f"最大有效压缩量：{fmt(last.shock_comp, 2)} mm",
            f"对应平叉转角 θ：{fmt(last.theta_deg, 3)}°",
            f"平均压缩比：{fmt(last.avg_shock_to_wheel, 4)}",
            f"初始杠杆比：{fmt(first_lr, 3)}    终点杠杆比：{fmt(last.instant_wheel_to_shock, 3)}",
            f"杠杆比总变化：{fmt(total_lr_change, 2)}%",
            f"终点杠杆比变化率：{fmt(last.lever_ratio_change_rate, 5)} 1/mm",
            f"减震器角度变化 Δα：{fmt(last.shock_angle_change_deg, 3)}°",
        ]
        if failed_count:
            lines.append(f"有 {failed_count} 个压缩点未找到解，可能已超过该几何结构可达范围。")
        else:
            lines.append("计算完成。")
        self.summary.setText("\n".join(lines))

    def _fill_table(self) -> None:
        self.table_model.set_rows(self.rows)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(False)

    def export_csv(self) -> None:
        if not self.calculate():
            return
        if not self.rows or self.params is None:
            QMessageBox.warning(self, "导出失败", "当前没有可导出的计算数据。")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出计算数据",
            "shock_ratio_results.csv",
            "CSV 文件 (*.csv)",
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as fp:
                writer = csv.writer(fp)
                writer.writerow(["输入参数", "值"])
                writer.writerows(
                    [
                        ["后轮轴横坐标", fmt(self.params.w0.x, 3)],
                        ["后轮轴纵坐标", fmt(self.params.w0.y, 3)],
                        ["减震器下安装点横坐标", fmt(self.params.p0.x, 3)],
                        ["减震器下安装点纵坐标", fmt(self.params.p0.y, 3)],
                        ["减震器上安装点横坐标", fmt(self.params.a.x, 3)],
                        ["减震器上安装点纵坐标", fmt(self.params.a.y, 3)],
                        ["最大压缩量", fmt(self.params.cmax, 3)],
                        ["计算步数", str(self.params.steps)],
                        ["最大搜索角 θ", fmt(self.params.theta_search_deg, 3)],
                        ["示意压缩量 C", fmt(self.params.diagram_c, 3)],
                        ["初始减震器长度 L0", fmt(self.l0, 3)],
                    ]
                )
                writer.writerow([])
                writer.writerow([title for title, _ in TABLE_COLUMNS])
                for row in self.rows:
                    writer.writerow([getter(row) for _, getter in TABLE_COLUMNS])
        except OSError as exc:
            QMessageBox.warning(self, "导出失败", str(exc))
            return

        QMessageBox.information(self, "导出完成", f"计算数据已导出到：\n{path}")

    def export_excel(self, path: Path, params: GeometryParams, rows: List[Row], l0: float) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        current_sheet = workbook.active
        current_sheet.title = "当前参数"
        current_sheet.append(["参数", "值"])
        for title, getter in PARAM_COLUMNS:
            current_sheet.append([title, getter(params)])
        current_sheet.append(["初始减震器长度", fmt(l0, 3)])

        records_sheet = workbook.create_sheet("方案记录")
        records_sheet.append(["方案名称", "保存时间", *[title for title, _ in PARAM_COLUMNS]])
        for record in self.scheme_records:
            records_sheet.append([record.name, record.created_at, *[getter(record.params) for _, getter in PARAM_COLUMNS]])

        result_sheet = workbook.create_sheet("计算结果")
        result_sheet.append([title for title, _ in TABLE_COLUMNS])
        for row in rows:
            result_sheet.append([getter(row) for _, getter in TABLE_COLUMNS])

        for sheet in workbook.worksheets:
            for column_cells in sheet.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 28)

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)

    def auto_export_unsaved_params(self) -> Optional[Path]:
        params = self.current_params()
        try:
            rows, l0, _, _ = calculate_rows(params)
        except Exception:
            rows, l0 = [], math.nan
        path = auto_export_dir() / f"未保存方案_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.export_excel(path, params, rows, l0)
        self.last_saved_params_key = params_key(params)
        self.current_params_saved = True
        self.update_scheme_status()
        return path

    def _render_progress(self, progress: float) -> None:
        if self.params is None:
            return
        progress = max(0.0, min(1.0, progress))
        c = self.params.cmax * progress
        state = solve_theta_by_compression(
            c,
            self.params.w0,
            self.params.p0,
            self.params.a,
            self.l0,
            math.radians(self.params.theta_search_deg),
            self.shock_angle0_deg,
        )
        self.mechanism_view.set_scene(self.params, state, self.l0, c)
        self._update_anim_info(state, c)

    def _update_anim_info(self, state: Optional[MechanismState], c: float) -> None:
        if state is None:
            self.anim_info.setText(f"C = {fmt(c, 2)} mm，未找到对应机构位置。")
            return
        j = state.wheel_jump
        ratio = math.nan if c == 0 else j / c
        inv_ratio = math.nan if j == 0 else c / j
        self.anim_info.setText(
            f"C = {fmt(c, 2)} mm    J = {fmt(j, 2)} mm\n"
            f"J/C = {fmt(ratio, 3)}    C/J = {fmt(inv_ratio, 4)}\n"
            f"θ = {fmt(state.theta_deg, 2)}°    α = {fmt(state.shock_angle_deg, 2)}°    Δα = {fmt(state.shock_angle_change_deg, 2)}°"
        )

    def _set_progress_slider(self, progress: float) -> None:
        self.updating_slider = True
        self.progress_slider.setValue(int(round(max(0.0, min(1.0, progress)) * 1000)))
        self.updating_slider = False

    def apply_window_size(self) -> None:
        self.resize(self.window_width.value(), self.window_height.value())

    def resizeEvent(self, event) -> None:  # noqa: N802
        super().resizeEvent(event)
        if not hasattr(self, "window_width") or self.updating_window_size:
            return
        self.updating_window_size = True
        self.window_width.setValue(self.width())
        self.window_height.setValue(self.height())
        self.updating_window_size = False

    def closeEvent(self, event) -> None:  # noqa: N802
        if not self.current_params_saved:
            try:
                self.auto_export_unsaved_params()
            except Exception as exc:
                QMessageBox.warning(self, "自动保存失败", f"当前参数未能自动导出为 Excel：\n{exc}")
        super().closeEvent(event)

    def _manual_progress_changed(self, value: int) -> None:
        if self.updating_slider:
            return
        self.pause_animation()
        self.anim_progress = value / 1000
        self._render_progress(self.anim_progress)

    def _update_speed_label(self, value: int) -> None:
        self.speed_label.setText(f"速度 {value / 100:.1f}x")

    def play_animation(self) -> None:
        if self.params is None:
            self.calculate()
        self.anim_progress = 0.0
        self._set_progress_slider(self.anim_progress)
        self.anim_started_at = time.perf_counter()
        self.timer.start()

    def pause_animation(self) -> None:
        if self.timer.isActive():
            self.timer.stop()

    def reset_animation(self) -> None:
        self.pause_animation()
        self.anim_progress = 0.0
        self._set_progress_slider(self.anim_progress)
        self._render_progress(self.anim_progress)

    def _animation_tick(self) -> None:
        speed = max(0.3, self.speed_slider.value() / 100)
        duration = 4.5 / speed
        elapsed = time.perf_counter() - self.anim_started_at
        self.anim_progress = min(1.0, elapsed / duration)
        self._set_progress_slider(self.anim_progress)
        self._render_progress(self.anim_progress)
        if self.anim_progress >= 1.0:
            self.timer.stop()


def main() -> int:
    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    app.setStyle("Fusion")
    window = CalculatorWindow()
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
