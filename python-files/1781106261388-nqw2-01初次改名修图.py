import os
from openpyxl import load_workbook

# 配置区域 文件夹名pic,表名aaa
excel_path = "aaa.xlsx"
file_folder = "pic"


# 读取Excel：B列=原名(无后缀)，A列=新名(无后缀)
name_map = {}
wb = load_workbook(excel_path)
ws = wb.active
# 从第5行开始读取（前面是表头）
for row in ws.iter_rows(min_row=5, values_only=True):
    old_stem = str(row[1]).strip()    # B列：原文件名（不带后缀）
    d_text = str(row[0]).strip()      # A列

    if not old_stem or not d_text:
        continue

    old_stem = str(old_stem).strip()
    new_stem = d_text.strip()
    name_map[old_stem] = new_stem

# 遍历文件夹批量改名
for filename in os.listdir(file_folder):
    old_path = os.path.join(file_folder, filename)
    if os.path.isdir(old_path):
        continue

    # 拆分 文件名 + 后缀
    stem, ext = os.path.splitext(filename)

    # 匹配成功则改名
    if stem in name_map:
        new_filename = name_map[stem] + ext
        new_path = os.path.join(file_folder, new_filename)
        os.rename(old_path, new_path)
        print(f"成功：{filename} → {new_filename}")
    else:
        print(f"无匹配：{filename}")

print("==== 改名完成====")
