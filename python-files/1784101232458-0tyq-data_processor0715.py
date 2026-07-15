# -*- coding: utf-8 -*-
"""
数据处理模块：规则清单匹配核心处理逻辑（pandas 矩阵版）

处理流程：
  步骤1：从【核算规范】读取业务事件 → pandas DataFrame（保留所需列，去除全空行）
  步骤2：清洗附件列，生成"该业务事件下所需附件"（整列操作 + 跳空拼接）
  步骤3：生成三组规则标识及所有中间指标
         报账单组：纯向量运算（.where 条件赋值）
         发票/附件组：整列 zip 并行操作（无 iterrows）
  步骤4：拼接全量标识（三组规则标识整列 @ 拼接）
  步骤5：按"@"拆分全量标识 → 每行一个标识符，输出中间调试表（每次覆盖）
  步骤6：与规则库左连接，按（事件序号+规则编码+规则组编码）去重，无匹配事件完整保留
  输出：标准化审核规则清单 Excel 文件

打包说明（PyInstaller exe 体积优化）：
  打包时可通过 --exclude-module 排除不用的 pandas 可选依赖，例如：
    pyinstaller main.py \\
      --exclude-module sqlalchemy \\
      --exclude-module matplotlib \\
      --exclude-module scipy \\
      --exclude-module PIL \\
      --exclude-module lxml \\
      --onefile
  另可用 UPX（upx.github.io）压缩 exe，一般可减小 40%-60% 体积。
"""

import re
import os
import math
import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 常量定义
# ============================================================

# 核算规范字段别名映射（key=内部字段名, value=Excel 中可能出现的列名列表）
GUIFAN_COL_ALIASES = {
    '适用表单':  ['*适用表单', '适用表单'],
    '场景编码':  ['业务场景编码', '场景编码'],
    '一级流程':  ['*一级流程', '一级流程'],
    '二级流程':  ['*二级流程', '二级流程'],
    '三级流程':  ['*三级流程', '三级流程'],
    '四级流程':  ['*四级流程', '四级流程'],
    '业务大类':  ['*业务大类', '业务大类'],
    '业务小类':  ['*业务小类', '业务小类'],
    '业务场景':  ['*业务场景', '业务场景'],
    '事前申请':  ['事前申请', '申请申请类', '事前申请类'],
    '发票':      ['发票', '发票类'],
    '合同协议':  ['合同/协议', '合同/协议类', '合同协议', '合同协议类'],
    '业务通知':  ['业务通知', '业务通知类'],
    '过程证明':  ['过程证明', '过程证明类'],
    '金额证明':  ['金额证明', '金额证明类', '出入库类'],
    '其他':      ['其他', '其他类'],
}

# 规则库字段别名映射
GUIZEIKU_COL_ALIASES = {
    '规则编码':     ['规则编码'],
    '规则组编码':   ['规则组编码'],
    '规则标识':     ['规则标识'],
    '审核要点':     ['审核要点'],
    '审核对象':     ['审核对象'],
    '附件报账单1':  ['附件/报账单1'],
    '附件报账单2':  ['附件/报账单2'],
    '附件报账字段': ['附件/报账1.字段'],
    '关系符号':     ['关系符号'],
    '附件2字段':    ['附件/报账单2.字段'],
    '预警信息':     ['预警信息'],
    '前置条件':     ['前置条件'],
    '规则间逻辑':   ['规则间逻辑关系'],
    '判断依据':     ['判断依据\n（XX制度、通知、红头文件等）', '判断依据'],
    '控制类型':     ['*控制类型', '控制类型'],
    '规则性质':     ['规则性质'],
    '规则小类':     ['规则小类'],
    '所属业务域':   ['*所属业务域', '所属业务域'],
    '规则适用范围': ['*规则适用范围', '规则适用范围'],
    '适用业务单元': ['适用业务单元（N)', '适用业务单元'],
    '规则适用节点': ['规则适用节点'],
}

# 输出文件前10列固定表头内容（行号 → 列号 → 值）
OUTPUT_HEADER = {
    'row1': {1: '系统信息', 2: '流程及场景', 10: '标准附件'},
    'row2': {
        1: '适用表单', 2: '业务场景编码', 3: '流程',
        7: '场景', 10: '该业务事件下所需附件',
    },
    'row3': {
        3: '*一级流程', 4: '*二级流程', 5: '*三级流程', 6: '*四级流程',
        7: '*业务大类', 8: '*业务小类', 9: '*业务场景',
    },
    'row4': {
        2: '待替换为业务事件编码',
        3: '已有字段', 4: '已有字段', 5: '已有字段',
        6: '已有字段', 7: '已有字段', 8: '已有字段',
        10: '新增字段',
    },
}

OUTPUT_MERGES = ['B1:I1', 'A2:A3', 'B2:B3', 'C2:F2', 'G2:I2', 'J2:J3']
OUTPUT_FIXED_COLS = 10       # 前10列为事件字段，后续列为规则库字段
OUTPUT_DATA_START_ROW = 5    # 数据从第5行开始写入

# 输出前10列对应的 DataFrame 字段名（顺序即列顺序）
OUTPUT_EVENT_FIELDS = [
    '适用表单', '场景编码', '一级流程', '二级流程', '三级流程', '四级流程',
    '业务大类', '业务小类', '业务场景', '该业务事件下所需附件',
]


# ============================================================
# 工具函数
# ============================================================

def _normalize_header(text):
    """
    去除表头文本的前后空格及开头的 * 号，便于跨版本列名匹配。
    例：'*业务大类' → '业务大类'；'  发票  ' → '发票'
    """
    if text is None:
        return ''
    return str(text).strip().lstrip('*').strip()


def _find_col_by_aliases(header_dict, aliases):
    """
    在 {列号: 列名} 字典中，按别名列表查找第一个匹配的列号。
    先精确匹配（去 * 后完全相等），再宽松匹配（别名是列名前缀）。
    例：aliases=['合同/协议', '合同协议'] 在 {5: '合同/协议类'} 中 → 返回 5
    """
    for alias in aliases:
        alias_norm = _normalize_header(alias)
        for col_idx, text in header_dict.items():
            if _normalize_header(text) == alias_norm:
                return col_idx
    for alias in aliases:
        alias_norm = _normalize_header(alias)
        if not alias_norm:
            continue
        for col_idx, text in header_dict.items():
            text_norm = _normalize_header(text)
            if alias_norm in text_norm:
                extra = text_norm[len(alias_norm):]
                if extra and extra not in ['（', '(', '类', '库']:
                    continue
                return col_idx
    return None


def _expand_merged_cells(ws):
    """
    将 openpyxl 工作表的合并单元格区域展开为 {(行,列): 值} 字典。
    合并区域内所有单元格统一使用左上角单元格的值。
    例：A1:C2 合并且值为"发票"，则 (1,1)(1,2)(1,3)(2,1)(2,2)(2,3) 均映射为"发票"。
    """
    merged_map = {}
    for merge_range in ws.merged_cells.ranges:
        top_val = ws.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged_map[(row, col)] = top_val
    return merged_map


def _read_all_values(ws, merged_map):
    """
    读取工作表所有单元格值，合并区域用展开后的值补充。
    返回 {(行,列): 值} 字典，仅包含非空值。
    """
    values = {}
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.value is not None:
                values[(cell.row, cell.column)] = cell.value
    for (r, c), v in merged_map.items():
        if (r, c) not in values and v is not None:
            values[(r, c)] = v
    return values


def _read_values_strip_strike(ws):
    """
    读取工作表所有单元格值，去除删除线格式的字符，并展开合并单元格。
    专用于核算规范文件加载（财务规范中被划删除线的内容视为已废弃，不应参与处理）。

    处理逻辑：
      - 富文本单元格（CellRichText）：仅保留无删除线的 TextBlock 文字段
        例：一格含 '增值税专用发票'（正常）+ '电子普通发票'（删除线）
            → 仅保留 '增值税专用发票'
      - 整格删除线（cell.font.strike=True）：整格视为空，不写入结果
      - 合并单元格：展开，非左上角格继承左上角格处理后的值

    返回: {(行, 列): 值} 字典，仅包含非空值。
    """
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        has_rich_text = True
    except ImportError:
        has_rich_text = False

    def _strip_cell(cell):
        val = cell.value
        if val is None:
            return None
        if has_rich_text and isinstance(val, CellRichText):
            parts = []
            for block in val:
                if isinstance(block, str):
                    parts.append(block)
                elif isinstance(block, TextBlock):
                    if not (block.font and getattr(block.font, 'strike', False)):
                        parts.append(block.text)
            text = ''.join(parts).strip()
            return text if text else None
        if cell.font and getattr(cell.font, 'strike', False):
            return None
        return val

    values = {}
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            v = _strip_cell(cell)
            if v is not None:
                values[(cell.row, cell.column)] = v

    # 展开合并单元格：非左上角格继承处理后的左上角格值
    for merge_range in ws.merged_cells.ranges:
        top_val = _strip_cell(ws.cell(merge_range.min_row, merge_range.min_col))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                if (row, col) not in values and top_val is not None:
                    values[(row, col)] = top_val

    return values


def _split_by_slash(text):
    """
    按顶层斜杠拆分字符串，括号内的斜杠不拆。
    例：'增值税专用发票/电子发票（含增值税/所得税）' → ['增值税专用发票', '电子发票（含增值税/所得税）']
    """
    result, depth, current = [], 0, []
    for ch in text:
        if ch in '（(':
            depth += 1
            current.append(ch)
        elif ch in '）)':
            depth -= 1
            current.append(ch)
        elif ch == '/' and depth == 0:
            part = ''.join(current).strip()
            if part:
                result.append(part)
            current = []
        else:
            current.append(ch)
    part = ''.join(current).strip()
    if part:
        result.append(part)
    return result


def _split_outside_brackets(text, separators):
    """
    按指定分隔符集合拆分字符串，全角/半角括号内的分隔符一律不拆。
    separators 中每个字符均视为独立分隔符。
    例：_split_outside_brackets('a@b（c@d）@e', '@')    → ['a', 'b（c@d）', 'e']
        _split_outside_brackets('甲、乙（丙、丁）', '、') → ['甲', '乙（丙、丁）']
    """
    result, depth, current = [], 0, []
    for ch in text:
        if ch in '（(':
            depth += 1
            current.append(ch)
        elif ch in '）)':
            depth = max(0, depth - 1)
            current.append(ch)
        elif ch in separators and depth == 0:
            part = ''.join(current).strip()
            if part:
                result.append(part)
            current = []
        else:
            current.append(ch)
    part = ''.join(current).strip()
    if part:
        result.append(part)
    return result


def _split_fapiao_names(text):
    """
    将发票字段文本拆分为各独立发票名，分隔符为顿号（、）和斜杠（/），括号内的分隔符不拆。
    例：'增值税专用发票、电子普通发票/增值税普通发票' → ['增值税专用发票', '电子普通发票', '增值税普通发票']
    例：'增值税专用发票（含电子、纸质）、普通发票' → ['增值税专用发票（含电子、纸质）', '普通发票']
    """
    if not text:
        return []
    s = str(text).strip()
    if s in ('', '/'):
        return []
    # 先按顿号拆（括号内不拆），再按斜杠拆（_split_by_slash 已支持括号内不拆）
    parts = _split_outside_brackets(s, '、')
    result = []
    for part in parts:
        part = part.strip()
        if part:
            result.extend(_split_by_slash(part))
    return [p.strip() for p in result if p.strip()]


--- a/_split_attach_items（原实现）
+++ b/_split_attach_items（新实现）
@@
 def _split_attach_items(text):
-    """
-    将附件字段文本按顿号（、）、斜杠（/）、@拆分为独立附件名列表，
-    同时去掉 *[...] 条件说明和空值，括号内的分隔符不拆。
-    """
-    if not text:
-        return []
-    items = []
-    for item in _split_outside_brackets(str(text), '@、/'):
-        item = re.sub(r'\*\[.*?\]', '', item).strip()
-        if item and item != '/':
-            items.append(item)
-    return items
+    """
+    将附件字段文本拆分为独立附件名列表
+    :param text: 原始附件文本
+    :param split_slash: True=按 / 拆分；False=不按 / 拆分（合同/协议等列使用）
+    """
+    if not text:
+        return []
+
+    separators = '@、'
+    if split_slash:
+        separators += '/'
+
+    items = []
+    for item in _split_outside_brackets(str(text), separators):
+        item = re.sub(r'\*\[.*?\]', '', item).strip()
+        if item and item != '/':
+            items.append(item)
+    return items


def _clean_level_val(series):
    """
    清洗业务层级列（业务大类/小类/场景）：fillna + strip，"/" 整体视为空。
    例：Series(['/', '差旅', None]) → Series(['', '差旅', ''])
    后续向量运算可直接用 != '' 判断层级是否有效。
    """
    s = series.fillna('').astype(str).str.strip()
    return s.where(s != '/', '')


# ============================================================
# 数据加载
# ============================================================

def load_规范_df(file_path, progress_cb=None):
    """
    加载财务规范初始化表单 → pandas DataFrame。

    处理逻辑：
      - 合并第2、3行表头（第3行子标题优先），用别名映射定位所需列
      - 跳过"适用表单"列值为表头关键词的行（如 '*适用表单'、'已有字段'）
      - 去掉核心字段（适用表单/业务大类/业务小类/业务场景）全为空或"/"的行

    返回: DataFrame，列名为 GUIFAN_COL_ALIASES 中的内部字段名
    """
    if progress_cb:
        progress_cb(f'正在读取核算规范文件：{os.path.basename(file_path)}')

    wb = openpyxl.load_workbook(file_path, rich_text=True)
    ws = wb.active
    cell_values = _read_values_strip_strike(ws)
    max_row, max_col = ws.max_row, ws.max_column

    # 合并第2、3行表头（第3行子标题优先覆盖第2行）
    h2 = {c: str(v) for c in range(1, max_col + 1) if (v := cell_values.get((2, c)))}
    h3 = {c: str(v) for c in range(1, max_col + 1) if (v := cell_values.get((3, c)))}
    combined = {**h2, **h3}

    col_map = {
        field: idx
        for field, aliases in GUIFAN_COL_ALIASES.items()
        if (idx := _find_col_by_aliases(combined, aliases))
    }

    required = ['适用表单', '一级流程', '二级流程', '业务大类', '业务小类', '业务场景']
    missing = [f for f in required if f not in col_map]
    if missing:
        raise ValueError(f'核算规范缺少必要列：{missing}')

    if progress_cb:
        progress_cb(f'  识别到列：{", ".join(f"{k}(第{v}列)" for k, v in col_map.items())}')

    NEEDED = [
        '适用表单', '场景编码', '一级流程', '二级流程', '三级流程', '四级流程',
        '业务大类', '业务小类', '业务场景',
        '事前申请', '发票', '合同协议', '业务通知', '过程证明', '金额证明', '其他',
    ]
    needed_cols = {f: col_map[f] for f in NEEDED if f in col_map}

    # 过滤表头行：跳过适用表单列为空或为表头关键词的行
    form_col = col_map['适用表单']
    HEADER_LIKE = {'已有字段', '新增字段', '*适用表单', '适用表单', '字段名称', '说明', '备注'}

    rows = []
    for r in range(2, max_row + 1):
        form_val = cell_values.get((r, form_col))
        if not form_val:
            continue
        form_str = str(form_val).strip()
        if form_str.startswith('*') or form_str in HEADER_LIKE:
            continue
        rows.append({field: cell_values.get((r, col)) for field, col in needed_cols.items()})

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # 去掉核心字段全为空/全为"/"的行
    core = [f for f in ['适用表单', '业务大类', '业务小类', '业务场景'] if f in df.columns]
    def _is_blank(v):
        if v is None:
            return True
        return str(v).strip() in ('', '/', 'None', 'nan')

    all_blank = df[core].map(_is_blank).all(axis=1)
    df = df[~all_blank].reset_index(drop=True)

    if progress_cb:
        progress_cb(f'  共读取 {len(df)} 条业务事件数据')

    return df


def load_规则库(file_path, progress_cb=None):
    """
    加载规则库文件 → pandas DataFrame + 输出所需的表头信息。

    处理逻辑：
      - 用第2行表头定位各字段列号
      - 从第一个以"RULE"开头的规则编码行开始读取数据
      - DataFrame 列名为整数列号（1..rule_max_col），另新增 '规则标识'/'规则编码'/'规则组编码' 命名列

    返回: (rules_df, rule_header_data, rule_max_col, rule_header_merges)
      rule_header_data:   {行号: {列号: 值}} — 供输出文件复用表头
      rule_header_merges: [(min_r, min_c, max_r, max_c), ...] — 表头区合并格信息
    """
    if progress_cb:
        progress_cb(f'正在读取规则库文件：{os.path.basename(file_path)}')

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    merged_map = _expand_merged_cells(ws)
    cell_values = _read_all_values(ws, merged_map)
    max_row = ws.max_row
    rule_max_col = ws.max_column

    h2 = {c: str(v) for c in range(1, rule_max_col + 1) if (v := cell_values.get((2, c)))}
    col_map = {
        field: idx
        for field, aliases in GUIZEIKU_COL_ALIASES.items()
        if (idx := _find_col_by_aliases(h2, aliases))
    }

    if progress_cb:
        progress_cb(f'  识别到列：{", ".join(f"{k}(第{v}列)" for k, v in col_map.items())}')

    # 找数据起始行：第一个规则编码以"RULE"开头的行
    data_start = 4
    rule_code_col = col_map.get('规则编码', 1)
    for r in range(3, max_row + 1):
        v = cell_values.get((r, rule_code_col))
        if v and str(v).startswith('RULE'):
            data_start = r
            break

    # 读取表头行数据（供输出文件复用）
    rule_header_data = {
        r: {c: cell_values.get((r, c)) for c in range(1, rule_max_col + 1)}
        for r in range(1, data_start)
    }
    rule_header_merges = [
        (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
        for mr in ws.merged_cells.ranges if mr.max_row < data_start
    ]

    # 读取规则数据行 → DataFrame（列名为整数列号）
    rules_rows = []
    for r in range(data_start, max_row + 1):
        code = cell_values.get((r, rule_code_col))
        if not code:
            continue
        rules_rows.append({c: cell_values.get((r, c)) for c in range(1, rule_max_col + 1)})

    rules_df = pd.DataFrame(rules_rows)

    # 新增命名列，供 merge 时使用
    def _safe_col(col_idx):
        """取整数列号对应的列，不存在时返回空字符串 Series"""
        if col_idx is None or col_idx not in rules_df.columns:
            return pd.Series([''] * len(rules_df), dtype=str)
        return rules_df[col_idx].fillna('').astype(str).str.strip()

    rules_df['规则标识']  = _safe_col(col_map.get('规则标识'))
    rules_df['规则编码']  = _safe_col(col_map.get('规则编码'))
    rules_df['规则组编码'] = _safe_col(col_map.get('规则组编码'))

    if progress_cb:
        progress_cb(f'  共读取 {len(rules_df)} 条规则，规则库共 {rule_max_col} 列')

    return rules_df, rule_header_data, rule_max_col, rule_header_merges


# ============================================================
# 步骤2：清洗附件列 + 生成"该业务事件下所需附件"
# ============================================================

ATTACH_FIELDS = ['事前申请', '发票', '合同协议', '业务通知', '过程证明', '金额证明', '其他']

--- a/step2_清洗与拼接附件（原实现）
+++ b/step2_清洗与拼接附件（新实现）
@@
 ATTACH_FIELDS = ['事前申请', '发票', '合同协议', '业务通知', '过程证明', '金额证明', '其他']
 
+# 以下字段：/ 不再作为分隔符（仅 、@ 拆分）
+_NO_SPLIT_SLASH_FIELDS = {
+    '合同协议', '业务通知', '过程证明', '金额证明'
+}
+
 def step2_清洗与拼接附件(df):
     for field in ATTACH_FIELDS:
         clean_col = field + '_c'
         if field not in df.columns:
             df[clean_col] = ''
             continue
 
         col = df[field].fillna('').astype(str)
         col = col.str.replace(r'\*\[.*?\]', '', regex=True).str.strip()
 
         if field == '事前申请':
             col = col.str.replace(r'[^、/]*（系统表单）[^、/]*', '', regex=True)
             col = col.str.replace(r'^[、/]+|[、/]+$', '', regex=True)
             col = col.str.replace(r'[、/]{2,}', '、', regex=True).str.strip()
 
-        df[clean_col] = col
+        # ✅ 按字段决定是否拆分 /
+        if field in _NO_SPLIT_SLASH_FIELDS:
+            df[clean_col] = col.apply(
+                lambda x: '、'.join(_split_attach_items(x, split_slash=False))
+            )
+        else:
+            df[clean_col] = col.apply(
+                lambda x: '、'.join(_split_attach_items(x, split_slash=True))
+            )
+
+        # 清理首尾及重复分隔符（保险起见）
+        df[clean_col] = (
+            df[clean_col]
+            .str.replace(r'^[、@]+|[、@]+$', '', regex=True)
+            .str.replace(r'[、@]{2,}', '、', regex=True)
+            .str.strip()
+        )
 
     clean_cols = [f + '_c' for f in ATTACH_FIELDS]
     df['该业务事件下所需附件'] = (
         df[clean_cols].apply(lambda r: '+'.join(v for v in r if v), axis=1)
     )
     return df


# ============================================================
# 步骤3：生成三组规则标识
# ============================================================

def step3_报账单标识(df):
    """
    步骤3a：报账单四个层级标识 + 规则标识（纯向量运算，无循环）。

    标识规则：
      - 层级内部用 "-" 拼接，标识之间用 "@" 拼接
      - 业务层级字段值为 "/" 时视为空（_clean_level_val 已处理），不参与拼接
      - 通用标识始终生成；大类/小类/场景标识仅在对应层级不为空时生成

    举例（适用表单="报账单", 业务大类="差旅", 业务小类="出差", 业务场景=""）：
      报账单-通用    = '报账单-通用'
      报账单-业务大类 = '报账单-差旅'
      报账单-业务小类 = '报账单-差旅-出差'
      报账单-业务场景 = ''（场景为空，条件不满足）
      报账单规则标识  = '报账单-通用@报账单-差旅@报账单-差旅-出差'

      【注意：原来的小工具开发步骤中，业务层级字段值为 "/" 时保留了，
      并且大类/小类/场景标识在对应层级为空时也生成，但我认为我的做法更合理。
      后面两个标识使用一样的控制做法】
    """
    表单 = df['适用表单'].fillna('').astype(str).str.strip()
    大类 = _clean_level_val(df['业务大类'])
    小类 = _clean_level_val(df['业务小类'])
    场景 = _clean_level_val(df['业务场景'])

    df['报账单-通用']    = 表单 + '-通用'
    df['报账单-业务大类'] = (表单 + '-' + 大类).where(大类 != '', '')
    df['报账单-业务小类'] = (表单 + '-' + 大类 + '-' + 小类).where(
        (大类 != '') & (小类 != ''), '')
    df['报账单-业务场景'] = (表单 + '-' + 大类 + '-' + 小类 + '-' + 场景).where(
        (大类 != '') & (小类 != '') & (场景 != ''), '')

    # 四列整列 "@" 拼接，再压缩多余 "@" 并去首尾 "@"
    id_cols = ['报账单-通用', '报账单-业务大类', '报账单-业务小类', '报账单-业务场景']
    joined = df[id_cols[0]]
    for c in id_cols[1:]:
        joined = joined + '@' + df[c]
    df['报账单规则标识'] = (
        joined.str.replace(r'@+', '@', regex=True).str.strip('@')
    )
    return df


def step3_发票标识(df):
    """
    步骤3b：发票六个指标（发票附件 + 五级标识 + 规则标识）。

    核心逻辑：一个业务事件的发票字段可能包含多种发票名（用 '、' 或 '/' 分隔），
    每种发票名需单独生成各级标识，同一行的多个结果用 "@" 拼接。
    使用整列 zip 并行操作（zip 遍历各列的 Python list，避免 iterrows 开销）。

    举例（适用表单="报账单", 业务大类="差旅", 小类="出差", 场景="",
          发票="增值税专用发票、电子普通发票"）：
      _split_fapiao_names → ['增值税专用发票', '电子普通发票']

      发票-通用     = '增值税专用发票-通用@电子普通发票-通用'
      发票-报账单   = '增值税专用发票-报账单@电子普通发票-报账单'
      发票-业务大   = '增值税专用发票-差旅@电子普通发票-差旅'
      发票-业务小   = '增值税专用发票-差旅-出差@电子普通发票-差旅-出差'
      发票-业务场景 = ''（场景为空，条件不满足，结果列表为空）
      发票规则标识  = '增值税专用发票-通用@电子普通发票-通用@增值税专用发票-报账单@...'
                      （经 @+ 压缩和首尾 @ 清理）
    """
    表单_s = df['适用表单'].fillna('').astype(str).str.strip()
    大类_s = _clean_level_val(df['业务大类'])
    小类_s = _clean_level_val(df['业务小类'])
    场景_s = _clean_level_val(df['业务场景'])
    fap_s  = df['发票_c'].fillna('') if '发票_c' in df.columns else df.get('发票', pd.Series([''] * len(df))).fillna('')

    # 发票附件：保留原始文本（含 "/" 分隔多种发票）
    df['发票附件'] = fap_s.values

    # 整列并行生成各级标识
    通用_r, 报账单_r, 业务大_r, 业务小_r, 业务场景_r = [], [], [], [], []

    for fap, 表单, 大类, 小类, 场景 in zip(
            fap_s, 表单_s, 大类_s, 小类_s, 场景_s):
        names = _split_fapiao_names(fap)
        if not names:
            通用_r.append('');    报账单_r.append('')
            业务大_r.append('');  业务小_r.append('');  业务场景_r.append('')
            continue
        通用_r.append(   '@'.join(f'{n}-通用'                       for n in names))
        报账单_r.append( '@'.join(f'{n}-{表单}'                      for n in names if 表单 and n != 表单))
        业务大_r.append( '@'.join(f'{n}-{大类}'                      for n in names if 大类))
        业务小_r.append( '@'.join(f'{n}-{大类}-{小类}'               for n in names if 大类 and 小类))
        业务场景_r.append('@'.join(f'{n}-{大类}-{小类}-{场景}'        for n in names if 大类 and 小类 and 场景))

    df['发票-通用']   = pd.array(通用_r,    dtype=object)
    df['发票-报账单'] = pd.array(报账单_r,   dtype=object)
    df['发票-业务大'] = pd.array(业务大_r,   dtype=object)
    df['发票-业务小'] = pd.array(业务小_r,   dtype=object)
    df['发票-业务场景'] = pd.array(业务场景_r, dtype=object)

    id_cols = ['发票-通用', '发票-报账单', '发票-业务大', '发票-业务小', '发票-业务场景']
    joined = df[id_cols[0]]
    for c in id_cols[1:]:
        joined = joined + '@' + df[c]
    df['发票规则标识'] = (
        joined.str.replace(r'@+', '@', regex=True).str.strip('@')
    )
    return df


def step3_附件标识(df):
    """
    步骤3c：附件组（除发票外）六个指标（所需全部附件 + 五级标识 + 规则标识）。

    核心逻辑："所需全部附件（除发票外）"由各清洗列整列 "@" 拼接生成，
    再按 _split_attach_items 拆分为独立附件名列表，逐名生成各级标识，
    同行多个结果用 "@" 拼接。与 step3_发票标识 结构相同，列名和附件来源不同。
    使用整列 zip 并行操作（zip 遍历各列的 Python list，避免 iterrows 开销）。

    举例（适用表单="报账单", 业务大类="差旅", 业务小类="出差", 业务场景="",
          所需全部附件（除发票外）="合同@行程单、住宿凭证"）：
      _split_attach_items → ['合同', '行程单', '住宿凭证']

      附件-通用     = '合同-通用@行程单-通用@住宿凭证-通用'
      附件-报账单   = '合同-报账单@行程单-报账单@住宿凭证-报账单'
      附件-业务大类 = '合同-差旅@行程单-差旅@住宿凭证-差旅'
      附件-业务小类 = '合同-差旅-出差@行程单-差旅-出差@住宿凭证-差旅-出差'
      附件-业务场景 = ''（场景为空，条件不满足，结果列表为空）
      附件规则标识  = '合同-通用@行程单-通用@住宿凭证-通用@合同-报账单@...'
                      （经 @+ 压缩和首尾 @ 清理）
    """
    non_inv = ['事前申请_c', '合同协议_c', '业务通知_c', '过程证明_c', '金额证明_c', '其他_c']
    existing = [c for c in non_inv if c in df.columns]

    # 所需全部附件（除发票外）：过滤空值后用 "@" 拼接
    df['所需全部附件（除发票外）'] = (
        df[existing].apply(lambda r: '@'.join(v for v in r if v), axis=1)
    )

    表单_s = df['适用表单'].fillna('').astype(str).str.strip()
    大类_s = _clean_level_val(df['业务大类'])
    小类_s = _clean_level_val(df['业务小类'])
    场景_s = _clean_level_val(df['业务场景'])
    attach_s = df['所需全部附件（除发票外）']

    通用_r, 报账单_r, 业务大类_r, 业务小类_r, 业务场景_r = [], [], [], [], []

    for attach, 表单, 大类, 小类, 场景 in zip(
            attach_s, 表单_s, 大类_s, 小类_s, 场景_s):
        names = _split_attach_items(attach)
        if not names:
            通用_r.append('');    报账单_r.append('')
            业务大类_r.append(''); 业务小类_r.append(''); 业务场景_r.append('')
            continue
        通用_r.append(   '@'.join(f'{n}-通用'                  for n in names))
        报账单_r.append( '@'.join(f'{n}-{表单}'                 for n in names if 表单))
        业务大类_r.append('@'.join(f'{n}-{大类}'                for n in names if 大类))
        业务小类_r.append('@'.join(f'{n}-{大类}-{小类}'         for n in names if 大类 and 小类))
        业务场景_r.append('@'.join(f'{n}-{大类}-{小类}-{场景}'   for n in names if 大类 and 小类 and 场景))

    df['附件-通用']    = pd.array(通用_r,    dtype=object)
    df['附件-报账单']  = pd.array(报账单_r,   dtype=object)
    df['附件-业务大类'] = pd.array(业务大类_r, dtype=object)
    df['附件-业务小类'] = pd.array(业务小类_r, dtype=object)
    df['附件-业务场景'] = pd.array(业务场景_r, dtype=object)

    id_cols = ['附件-通用', '附件-报账单', '附件-业务大类', '附件-业务小类', '附件-业务场景']
    joined = df[id_cols[0]]
    for c in id_cols[1:]:
        joined = joined + '@' + df[c]
    df['附件规则标识'] = (
        joined.str.replace(r'@+', '@', regex=True).str.strip('@')
    )
    return df


# ============================================================
# 步骤4：全量标识
# ============================================================

def step4_全量标识(df):
    """
    步骤4：三组规则标识整列 "@" 拼接 → 全量标识（去重复 "@"，去首尾 "@"）。

    举例：
      报账单规则标识 = '报账单-通用@报账单-差旅'
      发票规则标识   = '增值税专用发票-通用'
      附件规则标识   = '合同-通用@行程单-通用'
      全量标识       = '报账单-通用@报账单-差旅@增值税专用发票-通用@合同-通用@行程单-通用'

    若某组规则标识为空（如无发票），整列拼接后的多余 "@" 由 str.replace 压缩清理。
    """
    joined = df['报账单规则标识'] + '@' + df['发票规则标识'] + '@' + df['附件规则标识']
    df['全量标识'] = joined.str.replace(r'@+', '@', regex=True).str.strip('@')
    return df


# ============================================================
# 步骤5：拆分全量标识 + 输出中间调试表
# ============================================================

def step5_拆分并输出(df, intermediate_path, progress_cb=None):
    """
    步骤5：按"@"拆分全量标识 → 每行一个标识符（explode），输出调试表。

    举例（某行全量标识 = '报账单-通用@报账单-差旅@合同-通用'）：
      拆分后变为3行，标识符列分别为：
        '报账单-通用' / '报账单-差旅' / '合同-通用'
      每行保留原始业务事件的所有字段（通过 _event_idx 与原行对应）。

    中间调试表每次运行都覆盖输出，用于排查标识符生成是否正确。

    返回: df_split（拆分后 DataFrame，含 _event_idx 和 标识符 列）
    """
    df = df.copy()
    df['_event_idx'] = range(len(df))   # 业务事件行号，用于后续去重和溯源

    df_split = (
        df.assign(标识符=df['全量标识'].str.split('@'))
          .explode('标识符')
    )
    df_split['标识符'] = df_split['标识符'].str.strip()
    df_split = df_split[df_split['标识符'] != ''].reset_index(drop=True)

    try:
        df_split.to_excel(intermediate_path, index=False)
        if progress_cb:
            progress_cb(f'  已输出中间调试表：{os.path.basename(intermediate_path)}（{len(df_split)} 行）')
    except Exception as e:
        if progress_cb:
            progress_cb(f'  警告：中间调试表输出失败 - {e}')

    return df_split


# ============================================================
# 步骤6：与规则库左连接
# ============================================================

def step6_匹配规则(df_split, rules_df, rule_max_col, progress_cb=None):
    """
    步骤6：用标识符在规则库中检索，左连接保留全部业务事件行。

    匹配逻辑：
      - left_on='标识符'，right_on='规则标识'
      - 左连接：无匹配的行保留，规则库字段为 NaN
      - 不做去重，直接将规则库所有字段拼接到结果中
    """
    result = df_split.merge(
        rules_df,
        left_on='标识符',
        right_on='规则标识',
        how='left',
        suffixes=('', '_rule'),
    )

    result = result.sort_values('_event_idx').reset_index(drop=True)

    if '规则编码' in result.columns:
        result['规则编码'] = result['规则编码'].fillna('')
        matched   = (result['规则编码'] != '').sum()
        unmatched = (result['规则编码'] == '').sum()
        if progress_cb:
            progress_cb(f'  匹配完成：成功拼接 {matched} 条，未匹配 {unmatched} 条，共 {len(result)} 行')
    else:
        if progress_cb:
            progress_cb(f'  匹配完成，共 {len(result)} 行')

    if '规则组编码' in result.columns:
        result['规则组编码'] = result['规则组编码'].fillna('')

    # 确保所有规则库整数列都存在（无匹配行该列为 NaN/None）
    for c in range(1, rule_max_col + 1):
        if c not in result.columns:
            result[c] = None

    return result


# ============================================================
# 输出文件生成
# ============================================================

def _get_thin_border():
    """返回灰色细边框对象（复用，不在循环内重复创建）"""
    thin = Side(style='thin', color='BFBFBF')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_header(ws, rule_header_data, rule_max_col, rule_header_merges):
    """
    写入输出文件的4行表头（前10列为固定内容，后续列为规则库表头），
    同时应用合并单元格、填充色、字体、对齐等样式，并设置行高。
    """
    total_cols = OUTPUT_FIXED_COLS + rule_max_col
    thin_border = _get_thin_border()
    WHITE, BLACK = 'FFFFFFFF', 'FF000000'

    for row_idx in range(1, 5):
        row_key = f'row{row_idx}'
        fixed_data = OUTPUT_HEADER.get(row_key, {})
        has_border = (row_idx <= 3)

        for col in range(1, total_cols + 1):
            cell = ws.cell(row_idx, col)
            if col <= OUTPUT_FIXED_COLS:
                cell.value = fixed_data.get(col)
            else:
                cell.value = rule_header_data.get(row_idx, {}).get(col - OUTPUT_FIXED_COLS)

            # 各行各列填充色规则
            if row_idx in (1, 2):
                fgColor = '92D050' if col == 1 else ('70AD47' if col == OUTPUT_FIXED_COLS else '0070C0')
                cell.fill = PatternFill('solid', fgColor=fgColor)
            elif row_idx == 3:
                fgColor = '70AD47' if col == OUTPUT_FIXED_COLS else '0070C0'
                cell.fill = PatternFill('solid', fgColor=fgColor)
            elif row_idx == 4:
                if col == OUTPUT_FIXED_COLS:
                    cell.fill = PatternFill('solid', fgColor='00B050')
                elif col > OUTPUT_FIXED_COLS:
                    cell.fill = PatternFill(fill_type=None)
                else:
                    cell.fill = PatternFill('solid', fgColor='0070C0')

            is_no_fill = (row_idx == 4 and col > OUTPUT_FIXED_COLS)
            font_color = BLACK if is_no_fill else WHITE

            if row_idx == 1:
                cell.font = Font(name='微软雅黑', size=12, bold=True, color=font_color)
            elif row_idx == 2:
                cell.font = Font(name='微软雅黑', size=(10 if col <= 2 else 12), bold=True, color=font_color)
            elif row_idx == 3:
                size = 9 if 3 <= col <= 6 else (10 if 7 <= col <= 9 else 12)
                bold = not (col > OUTPUT_FIXED_COLS)
                cell.font = Font(name='微软雅黑', size=size, bold=bold, color=font_color)
            elif row_idx == 4:
                cell.font = Font(name='微软雅黑', size=10, bold=(col <= OUTPUT_FIXED_COLS), color=font_color)

            cell.border = thin_border if has_border else Border()
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 固定区域合并单元格
    for merge_range in OUTPUT_MERGES:
        ws.merge_cells(merge_range)
    # 规则库表头区合并单元格（列号偏移 OUTPUT_FIXED_COLS）
    for min_r, min_c, max_r, max_c in rule_header_merges:
        out_min = min_c + OUTPUT_FIXED_COLS
        out_max = max_c + OUTPUT_FIXED_COLS
        if min_r != max_r or out_min != out_max:
            ws.merge_cells(start_row=min_r, start_column=out_min,
                           end_row=max_r,   end_column=out_max)

    ws.row_dimensions[1].height = 33.75
    ws.row_dimensions[2].height = 118.15
    ws.row_dimensions[3].height = 178.15
    ws.row_dimensions[4].height = 43.9


def write_output(result_df, output_path, rule_max_col,
                 rule_header_data, rule_header_merges, progress_cb=None):
    """
    将匹配结果 DataFrame 写入输出 Excel 文件。
    前10列：核算规范事件字段（固定顺序 OUTPUT_EVENT_FIELDS）。
    后续列：规则库所有列（按整数列号顺序 1..rule_max_col）。

    性能策略：
      阶段1：预提取各列数据为 Python list（避免 iterrows 逐行 pandas 开销），
             逐行 ws.append 写入值（openpyxl 批量追加，无样式）。
      阶段2：统一批量赋样式（样式对象只创建一次，循环内只做赋值）。
    """
    if progress_cb:
        progress_cb(f'正在生成输出文件：{os.path.basename(output_path)}')

    def _val(v):
        """将单元格值规范化：None/nan/空字符串统一返回 None，其余转 str 并 strip"""
        if v is None:
            return None
        if isinstance(v, float) and math.isnan(v):
            return None
        s = str(v).strip()
        return s if s not in ('', 'None', 'nan') else None

    total_cols = OUTPUT_FIXED_COLS + rule_max_col
    rule_int_cols = list(range(1, rule_max_col + 1))
    n_rows = len(result_df)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    _write_header(ws, rule_header_data, rule_max_col, rule_header_merges)
    ws.freeze_panes = f'A{OUTPUT_DATA_START_ROW}'

    # 阶段1：预提取各列为 Python list，逐行 append（比 iterrows 快，省去 pandas Series 开销）
    event_cols = [
        result_df[f].tolist() if f in result_df.columns else [None] * n_rows
        for f in OUTPUT_EVENT_FIELDS
    ]
    rule_cols = [
        result_df[c].tolist() if c in result_df.columns else [None] * n_rows
        for c in rule_int_cols
    ]
    all_cols = event_cols + rule_cols  # 每个元素是一列的 list

    for i in range(n_rows):
        ws.append([_val(col[i]) for col in all_cols])

    # 阶段2：统一赋样式（样式对象只创建一次）
    last_row = OUTPUT_DATA_START_ROW + n_rows - 1
    for r in range(OUTPUT_DATA_START_ROW, last_row + 1):
        ws.row_dimensions[r].height = 13.85

    for col in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13.0

    wb.save(output_path)
    if progress_cb:
        progress_cb(f'  共写入 {n_rows} 条数据行')


# ============================================================
# 主处理函数
# ============================================================

def process(guifan_path, guizeiku_path, output_path, progress_cb=None):
    """
    主处理函数：完整执行六步骤规则清单匹配流程（pandas 矩阵版）。

    返回: (success: bool, message: str, output_path: str | None)
    """
    def log(msg):
        if progress_cb:
            progress_cb(msg)

    try:
        log('=' * 50)
        log('开始执行规则清单匹配处理（pandas 矩阵版）...')
        log('=' * 50)

        # ---- 步骤1：加载数据 ----
        log('\n【步骤1】加载数据文件')
        df = load_规范_df(guifan_path, progress_cb=log)
        rules_df, rule_header_data, rule_max_col, rule_header_merges = \
            load_规则库(guizeiku_path, progress_cb=log)

        if df.empty:
            return False, '核算规范中没有有效业务事件数据，请检查文件内容。', None
        if rules_df.empty:
            return False, '规则库中没有有效规则数据，请检查文件内容。', None

        # ---- 步骤2：清洗附件列 + 生成所需附件字段 ----
        log('\n【步骤2】清洗附件列，生成"该业务事件下所需附件"')
        df = step2_清洗与拼接附件(df)
        log(f'  完成（{len(df)} 行）')

        # ---- 步骤3：生成三组规则标识 ----
        log('\n【步骤3】生成三组规则标识及中间指标')
        df = step3_报账单标识(df)
        log('  报账单组完成')
        df = step3_发票标识(df)
        log('  发票组完成')
        df = step3_附件标识(df)
        log('  附件组完成')

        # ---- 步骤4：全量标识 ----
        log('\n【步骤4】拼接全量标识')
        df = step4_全量标识(df)
        total_ids = df['全量标识'].str.count('@').sum() + len(df[df['全量标识'] != ''])
        log(f'  全量标识生成完毕，共约 {total_ids} 个标识符')

        # ---- 步骤5：拆分全量标识，输出中间调试表 ----
        log('\n【步骤5】拆分全量标识（输出中间调试表）')
        intermediate_path = output_path.replace('.xlsx', '_中间_全量标识拆分.xlsx')
        df_split = step5_拆分并输出(df, intermediate_path, progress_cb=log)
        log(f'  拆分后共 {len(df_split)} 条（事件×标识符）记录')

        # ---- 步骤6：与规则库匹配拼接 ----
        log('\n【步骤6】与规则库匹配拼接')
        result = step6_匹配规则(df_split, rules_df, rule_max_col, progress_cb=log)

        # ---- 写入输出文件 ----
        log('\n【输出】写入结果文件')
        write_output(result, output_path, rule_max_col,
                     rule_header_data, rule_header_merges, progress_cb=log)

        log('\n' + '=' * 50)
        log(f'处理完成！输出文件：{output_path}')
        log('=' * 50)

        return True, f'处理成功！共生成 {len(result)} 条规则清单记录。', output_path

    except Exception as e:
        error_msg = f'处理出错：{str(e)}\n{traceback.format_exc()}'
        log(f'\n[错误] {error_msg}')
        return False, error_msg, None


# ============================================================
# 测试入口
# ============================================================

if __name__ == '__main__':
    import sys
    sys.stdout.reconfigure(encoding='utf-8')

    base_dir = os.path.dirname(os.path.abspath(__file__))
    guifan   = os.path.join(base_dir, '输入数据', 'XX业务域核算规范.xlsx')
    guizeiku = os.path.join(base_dir, '输入数据', '费报业务域规则库.xlsx')
    output   = os.path.join(base_dir, '输出数据', '测试输出_规则清单.xlsx')

    success, msg, path = process(guifan, guizeiku, output, progress_cb=print)
    print(f'\n结果：{"成功" if success else "失败"} - {msg}')
