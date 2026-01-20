#!/usr/bin/env python3
"""
ema_ectd_backbone_generator.py
Generates and validates eCTD 3.2 and EU M1 3.1 compliant XML backbones.
Version 1.0
"""

import os
import re
import csv
import uuid
import hashlib
from pathlib import Path
from lxml import etree

# =========================================================
# Utility functions
# =========================================================

XLINK_NS_CANONICAL = "http://www.w3.org/1999/xlink"
XLINK_NS_LEGACY = "http://www.w3c.org/1999/xlink"
ICH_NS_DEFAULT = "http://www.ich.org/ectd"
EU_NS_DEFAULT = "http://europa.eu.int"
XML_NS = "http://www.w3.org/XML/1998/namespace"

UTIL_DTD_DIR = Path("util") / "dtd"
UTIL_STYLE_DIR = Path("util") / "style"

ICH_DTD_FILENAME = "ich-ectd-3-2.dtd"
EU_DTD_FILENAME = "eu-regional.dtd"
EU_ENVELOPE_MOD_FILENAME = "eu-envelope.mod"

ICH_XSL_FILENAME = "ectd-2-0.xsl"
EU_XSL_FILENAME = "eu-regional.xsl"

EU_DEFAULTS = {
    "env_country": "at",
    "submission_type": "none",
    "submission_mode": "single",
    "procedure_number": "1",
    "submission_unit_type": "initial",
    "applicant_name": "Unknown applicant",
    "agency_code": "EU-EMA",
    "procedure_type": "national",
    "invented_name": "Unknown product",
    "submission_description": "Not provided",
    "pi_country": "common",
    "pi_lang": "en",
    "pi_type": "other",
}

def md5_checksum(filepath: Path) -> str:
    m = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            m.update(chunk)
    return m.hexdigest()

METADATA_FIELDS = [
    "file_path",
    "title",
    "attributes",
    "operation",
    "modified-leaf",
    "modified-href",
    "ctd_toc",
    "applicant_name",
    "submission_type",
    "sequence_description",
    "eu_country",
    "eu_identifier",
    "eu_submission_type",
    "eu_submission_mode",
    "eu_submission_number",
    "eu_procedure_number",
    "eu_submission_unit_type",
    "eu_agency_code",
    "eu_procedure_type",
    "eu_invented_name",
    "eu_inn",
    "eu_related_sequence",
]

def checksum_for_path(seq_dir: Path, file_path: str) -> tuple[str, bool]:
    """Return checksum for file_path or empty string if missing (with warnings)."""
    file_abs = seq_dir / file_path
    if not file_abs.exists():
        if not file_abs.parent.exists():
            print(f"⚠ Missing directory for leaf: {file_abs.parent}")
        print(f"⚠ Missing file for leaf: {file_path}")
        return "", False
    if file_abs.is_dir():
        print(f"⚠ Skipping checksum for directory path: {file_path}")
        return "", False
    return md5_checksum(file_abs), True

def write_xml_with_doctype_and_xsl(xml_path: Path, root, doctype: str, xsl_href: str) -> None:
    """Write XML with a DOCTYPE and xml-stylesheet PI."""
    etree.indent(root, space="  ")
    body = etree.tostring(root, encoding="utf-8", xml_declaration=False)
    header = b'<?xml version="1.0" encoding="utf-8"?>\n'
    doctype_line = (doctype + "\n").encode("utf-8")
    pi_line = (f'<?xml-stylesheet href="{xsl_href}" type="text/xsl"?>\n').encode("utf-8")
    with open(xml_path, "wb") as f:
        f.write(header)
        f.write(doctype_line)
        f.write(pi_line)
        f.write(body)

def write_index_md5(seq_dir: Path, index_path: Path) -> None:
    """Write MD5 checksum for index.xml to index-md5.txt in sequence root."""
    if not index_path.exists():
        return
    checksum = md5_checksum(index_path)
    rel = index_path.relative_to(seq_dir).as_posix()
    out_path = seq_dir / "index-md5.txt"
    out_path.write_text(f"{checksum}  {rel}\n", encoding="utf-8")

def resolve_previous_leaf_ref(
    base_dir: Path,
    seq_num: int,
    modified_leaf: str | None,
    modified_href: str | None,
    file_path: str,
    fallback_paths: list[str] | None = None,
) -> tuple[int, str, str, str] | None:
    """Resolve prior leaf by explicit ID or by href/path lookup."""
    if modified_leaf:
        return find_previous_leaf_by_id(base_dir, seq_num, modified_leaf)
    if modified_href:
        return find_previous_leaf_id(base_dir, seq_num, [modified_href])
    if fallback_paths:
        return find_previous_leaf_id(base_dir, seq_num, fallback_paths)
    return find_previous_leaf_id(base_dir, seq_num, [file_path])

def load_metadata(xlsx_path: Path, sheet_name: str | None = None) -> list[dict]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read Excel metadata files. "
            "Install it in the current environment."
        ) from exc

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            record[header] = "" if value is None else str(value).strip()
        file_path = (record.get("file_path") or "").strip()
        operation = (record.get("operation") or "").strip().lower()
        modified_href = (record.get("modified-href") or "").strip()
        modified_leaf = (record.get("modified-leaf") or "").strip()
        if not file_path and operation != "delete" and not modified_href and not modified_leaf:
            continue
        out.append(record)
    return out

def ensure_metadata_columns(xlsx_path: Path) -> None:
    """Ensure metadata sheet has all expected columns (including ctd-toc)."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to write Excel metadata files. "
            "Install it in the current environment."
        ) from exc

    if not xlsx_path.exists():
        return
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(h).strip() if h is not None else "" for h in (header_row or [])]
    if not any(headers):
        ws.delete_rows(1, ws.max_row)
        ws.append(METADATA_FIELDS)
        wb.save(xlsx_path)
        return
    updated = False
    for field in METADATA_FIELDS:
        if field == "ctd_toc":
            if "ctd_toc" in headers or "ctd-toc" in headers:
                continue
        if field not in headers:
            headers.append(field)
            ws.cell(row=1, column=len(headers), value=field)
            updated = True
    if updated:
        wb.save(xlsx_path)

def write_metadata_from_scan(seq_dir: Path, xlsx_path: Path) -> int:
    """Scan sequence directory and append missing rows to metadata Excel."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to write Excel metadata files. "
            "Install it in the current environment."
        ) from exc

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in (rows[0] if rows else [])]
        if not headers or "file_path" not in headers:
            ws.delete_rows(1, ws.max_row)
            ws.append(METADATA_FIELDS)
            headers = METADATA_FIELDS[:]
        elif "ctd_toc" not in headers and "ctd-toc" not in headers:
            headers.append("ctd_toc")
            ws.cell(row=1, column=len(headers), value="ctd_toc")
        existing = set()
        for row in rows[1:]:
            if not row:
                continue
            idx = headers.index("file_path")
            value = row[idx] if idx < len(row) else None
            if value:
                existing.add(str(value).strip())
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "metadata"
        ws.append(METADATA_FIELDS)
        headers = METADATA_FIELDS[:]
        existing = set()

    new_rows = 0
    for path in sorted(seq_dir.rglob("*")):
        if not path.is_file():
            continue
        rel_path = path.relative_to(seq_dir).as_posix()
        if "/" not in rel_path:
            continue
        if rel_path == "index.xml" or rel_path == "m1/eu/eu-regional.xml":
            continue
        if rel_path.startswith("util/"):
            continue
        if rel_path in existing:
            continue
        title = path.stem.replace("-", " ").replace("_", " ").strip() or path.name
        row = {
            "file_path": rel_path,
            "title": title,
            "operation": "new",
        }
        ws.append([row.get(field, "") for field in METADATA_FIELDS])
        new_rows += 1

    if new_rows > 0 or not xlsx_path.exists():
        wb.save(xlsx_path)
    return new_rows

def extract_eu_envelope_fields(xml_path: Path) -> dict[str, str]:
    """Extract EU envelope values from eu-regional.xml."""
    if not xml_path.exists():
        return {}
    try:
        root = etree.parse(xml_path).getroot()
    except OSError:
        return {}
    envelope = root.xpath("//*[local-name()='envelope']")
    envelope = envelope[0] if envelope else None
    if envelope is None:
        return {}

    submission = envelope.xpath("./*[local-name()='submission']")
    submission = submission[0] if submission else None
    submission_number = ""
    procedure_number = ""
    submission_type = ""
    submission_mode = ""
    if submission is not None:
        submission_type = submission.get("type") or ""
        submission_mode = submission.get("mode") or ""
        submission_number = submission.xpath("string(./*[local-name()='number'])") or ""
        proc_tracking = submission.xpath("./*[local-name()='procedure-tracking']")
        proc_tracking = proc_tracking[0] if proc_tracking else None
        if proc_tracking is not None:
            procedure_number = proc_tracking.xpath("string(./*[local-name()='number'])") or ""

    submission_unit = envelope.xpath("./*[local-name()='submission-unit']")
    submission_unit = submission_unit[0] if submission_unit else None
    agency = envelope.xpath("./*[local-name()='agency']")
    agency = agency[0] if agency else None
    procedure = envelope.xpath("./*[local-name()='procedure']")
    procedure = procedure[0] if procedure else None

    return {
        "eu_country": envelope.get("country") or "",
        "eu_identifier": envelope.xpath("string(./*[local-name()='identifier'])") or "",
        "eu_submission_type": submission_type,
        "eu_submission_mode": submission_mode,
        "eu_submission_number": submission_number,
        "eu_procedure_number": procedure_number,
        "eu_submission_unit_type": submission_unit.get("type") if submission_unit is not None else "",
        "applicant_name": envelope.xpath("string(./*[local-name()='applicant'])") or "",
        "eu_agency_code": agency.get("code") if agency is not None else "",
        "eu_procedure_type": procedure.get("type") if procedure is not None else "",
        "eu_invented_name": envelope.xpath("string(./*[local-name()='invented-name'])") or "",
        "eu_inn": envelope.xpath("string(./*[local-name()='inn'])") or "",
        "eu_related_sequence": envelope.xpath("string(./*[local-name()='related-sequence'])") or "",
        "sequence_description": envelope.xpath("string(./*[local-name()='submission-description'])") or "",
    }

def _parse_modified_file_ref(modified_file: str) -> tuple[Path | None, str | None]:
    if not modified_file or "#" not in modified_file:
        return None, None
    path_str, leaf_id = modified_file.split("#", 1)
    path_str = path_str.replace("\\", "/").strip()
    if not path_str or not leaf_id:
        return None, None
    return Path(path_str), leaf_id

def _resolve_modified_href(seq_dir: Path, modified_file: str, cache: dict[Path, dict[str, str]]) -> str:
    rel_path, leaf_id = _parse_modified_file_ref(modified_file)
    if rel_path is None or leaf_id is None:
        return ""
    xml_path = (seq_dir / rel_path)
    if not xml_path.exists():
        return ""
    if xml_path not in cache:
        id_map: dict[str, str] = {}
        try:
            xml = etree.parse(xml_path)
        except OSError:
            cache[xml_path] = id_map
            return ""
        leaves = xml.xpath("//*[local-name()='leaf']")
        for leaf in leaves:
            leaf_id_val = leaf.get("ID")
            if not leaf_id_val:
                continue
            href = _leaf_href(leaf)
            if href:
                id_map[leaf_id_val] = href
        cache[xml_path] = id_map
    return cache.get(xml_path, {}).get(leaf_id, "")

def _extract_leaf_rows_from_xml(
    xml_path: Path,
    seq_dir: Path,
    base_prefix: str | None = None,
    include_ctd_toc: bool = True,
    attr_map: dict[str, set[str]] | None = None,
) -> list[dict[str, str]]:
    if not xml_path.exists():
        return []
    rows: list[dict[str, str]] = []
    cache: dict[Path, dict[str, str]] = {}
    xml = etree.parse(xml_path)
    leaves = xml.xpath("//*[local-name()='leaf']")
    for leaf in leaves:
        href = _leaf_href(leaf)
        file_path = href
        if href and base_prefix and not href.startswith(base_prefix + "/"):
            file_path = f"{base_prefix}/{href}"
        title = leaf.xpath("string(./*[local-name()='title'])") or ""
        operation = (leaf.get("operation") or "new").strip().lower()
        modified_file = leaf.get("modified-file") or ""
        _rel_path, modified_leaf = _parse_modified_file_ref(modified_file)
        modified_href = ""
        if modified_file:
            modified_href = _resolve_modified_href(seq_dir, modified_file, cache)
            if modified_href and base_prefix and not modified_href.startswith(base_prefix + "/"):
                modified_href = f"{base_prefix}/{modified_href}"
        if not file_path and modified_href:
            file_path = modified_href
        row = {
            "file_path": file_path,
            "title": title,
            "operation": operation,
            "modified-leaf": modified_leaf or "",
            "modified-href": modified_href,
        }
        if include_ctd_toc:
            parent = leaf.getparent()
            if parent is not None and parent.tag:
                row["ctd_toc"] = parent.tag.split("}", 1)[-1]
        if attr_map is not None:
            parent = leaf.getparent()
            if parent is not None and parent.tag:
                allowed = attr_map.get(parent.tag.split("}", 1)[-1], set())
                attrs: dict[str, str] = {}
                for name in sorted(allowed):
                    if name == "xml:lang":
                        val = parent.get("{%s}lang" % XML_NS)
                    else:
                        val = parent.get(name)
                    if val:
                        attrs[name] = val
                if attrs:
                    row["attributes"] = " ".join(f'{k}="{v}"' for k, v in attrs.items())
        rows.append(row)
    return rows

def extract_metadata_from_xml(seq_dir: Path, xlsx_path: Path) -> tuple[int, int, int]:
    """Update metadata Excel from backbone XML, filling only missing values."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to write Excel metadata files. "
            "Install it in the current environment."
        ) from exc

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "metadata"
        ws.append(METADATA_FIELDS)

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(h).strip() if h is not None else "" for h in (header_row or [])]
    for field in METADATA_FIELDS:
        if field == "ctd_toc":
            if "ctd_toc" in headers or "ctd-toc" in headers:
                continue
        if field not in headers:
            headers.append(field)
            ws.cell(row=1, column=len(headers), value=field)

    col_idx = {name: idx + 1 for idx, name in enumerate(headers) if name}
    if "ctd_toc" not in col_idx and "ctd-toc" in col_idx:
        col_idx["ctd_toc"] = col_idx["ctd-toc"]
    file_col = col_idx.get("file_path")
    existing_by_path: dict[str, int] = {}
    if file_col:
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=file_col).value
            if val is None:
                continue
            key = str(val).strip()
            if key:
                existing_by_path[_norm_href(key)] = row_idx

    updated = 0
    appended = 0
    index_xml = seq_dir / "index.xml"
    eu_xml = seq_dir / "m1" / "eu" / "eu-regional.xml"
    ich_dtd = seq_dir / UTIL_DTD_DIR / ICH_DTD_FILENAME
    eu_dtd = seq_dir / UTIL_DTD_DIR / EU_DTD_FILENAME
    ich_attr_map = _parse_dtd_attlist(ich_dtd) if ich_dtd.exists() else {}
    eu_attr_map = _parse_dtd_attlist(eu_dtd) if eu_dtd.exists() else {}
    rows = _extract_leaf_rows_from_xml(
        index_xml,
        seq_dir,
        include_ctd_toc=True,
        attr_map=ich_attr_map,
    )
    rows += _extract_leaf_rows_from_xml(
        eu_xml,
        seq_dir,
        base_prefix="m1/eu",
        include_ctd_toc=False,
        attr_map=eu_attr_map,
    )

    for row in rows:
        file_path = _norm_href(row.get("file_path") or "")
        if not file_path:
            continue
        if file_path == "index.xml" or file_path == "m1/eu/eu-regional.xml":
            continue
        row_idx = existing_by_path.get(file_path)
        if row_idx is None:
            ws.append([row.get(field, "") for field in METADATA_FIELDS])
            row_idx = ws.max_row
            existing_by_path[file_path] = row_idx
            appended += 1
            continue
        for field, value in row.items():
            if not value:
                continue
            col = col_idx.get(field)
            if not col:
                continue
            cell = ws.cell(row=row_idx, column=col)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = value
                updated += 1

    envelope_fields = extract_eu_envelope_fields(eu_xml)
    envelope_updates = 0
    if envelope_fields:
        target_row = None
        file_col = col_idx.get("file_path")
        if file_col:
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=file_col).value
                if val is None:
                    continue
                key = _norm_href(str(val))
                if key.startswith("m1/eu/"):
                    target_row = row_idx
                    break
        if target_row is None:
            target_row = 2 if ws.max_row >= 2 else None
        if target_row is not None:
            for field, value in envelope_fields.items():
                if not value:
                    continue
                col = col_idx.get(field)
                if not col:
                    continue
                cell = ws.cell(row=target_row, column=col)
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = value
                    envelope_updates += 1

    if updated or appended or envelope_updates:
        wb.save(xlsx_path)
    return updated, appended, envelope_updates

def _norm_href(href: str) -> str:
    """Normalize hrefs for comparison (strip, unify slashes)."""
    href = (href or "").strip()
    return href.replace("\\", "/")

def _leaf_href(leaf) -> str:
    """Extract href from a leaf element (xlink:href or href)."""
    # Prefer explicit xlink href if present
    for k in (
        "{%s}href" % XLINK_NS_CANONICAL,
        "{%s}href" % XLINK_NS_LEGACY,
        "href",
    ):
        if k in leaf.attrib:
            return _norm_href(leaf.attrib.get(k) or "")
    # Fallback: any attribute whose localname is 'href'
    for k, v in leaf.attrib.items():
        if k.endswith("}href") or k == "href":
            return _norm_href(v or "")
    return ""

def _parse_custom_attributes(raw: str) -> dict[str, str]:
    """Parse attributes in the form: key="value" key2='value2'."""
    raw = (raw or "").strip()
    if not raw:
        return {}
    # Normalize smart quotes to ASCII for Excel-friendly input.
    raw = (
        raw.replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2018", "'")
        .replace("\u2019", "'")
    )
    attrs: dict[str, str] = {}
    pattern = r'([^\s=]+)\s*=\s*(?:"([^"]*)"|\'([^\']*)\')'
    for name, v1, v2 in re.findall(pattern, raw):
        value = v1 if v1 is not None else v2
        if name:
            attrs[name] = value
    return attrs

def _tokenize_dtd_attlist(content: str) -> list[str]:
    tokens: list[str] = []
    i = 0
    n = len(content)
    while i < n:
        c = content[i]
        if c.isspace():
            i += 1
            continue
        if c in ("'", '"'):
            quote = c
            i += 1
            start = i
            while i < n and content[i] != quote:
                i += 1
            tokens.append(content[start:i])
            i += 1
            continue
        if c == "(":
            depth = 1
            start = i
            i += 1
            while i < n and depth > 0:
                if content[i] == "(":
                    depth += 1
                elif content[i] == ")":
                    depth -= 1
                i += 1
            tokens.append(content[start:i])
            continue
        start = i
        while i < n and not content[i].isspace():
            i += 1
        tokens.append(content[start:i])
    return tokens

def _parse_dtd_attlist(dtd_path: Path) -> dict[str, set[str]]:
    text = dtd_path.read_text(encoding="utf-8", errors="ignore")
    text = re.sub(r"<!--.*?-->", "", text, flags=re.S)
    entities = _parse_dtd_entities(text)
    text = _expand_entities(text, entities)
    out: dict[str, set[str]] = {}
    for match in re.finditer(r"<!ATTLIST\s+([^\s>]+)\s+([^>]+)>", text, flags=re.S):
        element = match.group(1)
        content = match.group(2)
        tokens = _tokenize_dtd_attlist(content)
        idx = 0
        attrs: set[str] = set()
        while idx < len(tokens):
            name = tokens[idx]
            idx += 1
            if name.startswith("%") and name.endswith(";"):
                continue
            if idx >= len(tokens):
                break
            idx += 1  # attr type
            if idx >= len(tokens):
                break
            default = tokens[idx]
            idx += 1
            if default == "#FIXED" and idx < len(tokens):
                idx += 1
            attrs.add(name)
        if attrs:
            out[element] = attrs
    return out

def _normalize_attr_name(name: str) -> str:
    if name == "xml:lang":
        return "{%s}lang" % XML_NS
    return name

def _is_directory_path(seq_dir: Path, file_path: str) -> bool:
    path_str = (file_path or "").strip()
    if not path_str:
        return False
    if path_str.endswith("/"):
        return True
    candidate = Path(path_str)
    if not candidate.is_absolute():
        candidate = (seq_dir / candidate).resolve()
    return candidate.exists() and candidate.is_dir()

def _allowed_attrs_for_element(attr_map: dict[str, set[str]], element_tag: str) -> set[str]:
    local = element_tag.split("}", 1)[-1]
    return attr_map.get(local, set())

def _filter_allowed_attrs(
    custom_attrs: dict[str, str],
    allowed_attrs: set[str],
    context: str,
    element_tag: str,
) -> dict[str, str]:
    filtered: dict[str, str] = {}
    local = element_tag.split("}", 1)[-1]
    for name, value in custom_attrs.items():
        if name not in allowed_attrs:
            print(f"⚠ Attribute '{name}' not allowed for {local} ({context})")
            continue
        filtered[name] = value
    return filtered


def _referenced_xmls_from_index(prev_dir: Path) -> list[Path]:
    """Return all XML files referenced by index.xml in a sequence.

    We treat a referenced file as an XML candidate if its href ends with '.xml' (case-insensitive)
    and the file exists on disk. Always includes index.xml itself.
    """
    candidates: list[Path] = []

    index_path = prev_dir / "index.xml"
    if index_path.exists():
        candidates.append(index_path)
        try:
            ix = etree.parse(str(index_path))
            leaves = ix.xpath("//*[local-name()='leaf']")
            for leaf in leaves:
                href = _leaf_href(leaf)
                if not href:
                    continue
                if not href.lower().endswith(".xml"):
                    continue
                p = (prev_dir / href).resolve()
                # Only include files that are actually inside the sequence folder
                # (avoid accidental directory traversal)
                try:
                    p.relative_to(prev_dir.resolve())
                except Exception:
                    continue
                if p.exists() and p.is_file():
                    candidates.append(p)
        except Exception:
            # If index.xml is malformed, we at least search it as-is later.
            pass

    # De-duplicate while preserving order
    seen = set()
    out: list[Path] = []
    for p in candidates:
        if p not in seen:
            out.append(p)
            seen.add(p)
    return out


def find_previous_leaf_id(base_dir: Path, current_seq: int, file_paths: list[str]) -> tuple[int, str, str, str] | None:
    """Walk back through previous sequences to locate the most recent matching leaf ID.

    Searches index.xml AND every XML file that is referenced by index.xml (via leaf @xlink:href)
    within each previous sequence.

    Matching strategy (in order):
      1) Exact href match against any of `file_paths`
      2) Basename match (last path component) against any of `file_paths`
         (useful when replace uses a new filename but same document basename).
    """
    wanted = [_norm_href(p) for p in (file_paths or []) if p]
    wanted_basenames = {Path(p).name for p in wanted if p}

    for prev_seq in range(current_seq - 1, -1, -1):
        prev_dir = base_dir / f"{prev_seq:04d}"
        prev_dir_abs = prev_dir.resolve()
        xml_paths = _referenced_xmls_from_index(prev_dir)
        if not xml_paths:
            continue

        for xml_path in xml_paths:
            try:
                xml = etree.parse(str(xml_path))
            except Exception:
                # Skip malformed or non-XML files
                continue
            try:
                rel_xml_path = xml_path.resolve().relative_to(prev_dir_abs).as_posix()
            except Exception:
                rel_xml_path = "index.xml"

            # Use local-name() to be namespace-agnostic
            leaves = xml.xpath("//*[local-name()='leaf']")

            # 1) exact href match
            for leaf in leaves:
                href = _leaf_href(leaf)
                if href and href in wanted:
                    leaf_id = leaf.get("ID")
                    if leaf_id:
                        return prev_seq, rel_xml_path, leaf_id, href

            # 2) basename match (best-effort fallback)
            if wanted_basenames:
                for leaf in leaves:
                    href = _leaf_href(leaf)
                    if href and Path(href).name in wanted_basenames:
                        leaf_id = leaf.get("ID")
                        if leaf_id:
                            return prev_seq, rel_xml_path, leaf_id, href

    return None

def find_previous_leaf_by_id(base_dir: Path, current_seq: int, leaf_id: str) -> tuple[int, str, str, str] | None:
    """Walk back through previous sequences to locate a specific leaf ID."""
    if not leaf_id:
        return None

    for prev_seq in range(current_seq - 1, -1, -1):
        prev_dir = base_dir / f"{prev_seq:04d}"
        prev_dir_abs = prev_dir.resolve()
        xml_paths = _referenced_xmls_from_index(prev_dir)
        if not xml_paths:
            continue

        for xml_path in xml_paths:
            try:
                xml = etree.parse(str(xml_path))
            except Exception:
                continue
            try:
                rel_xml_path = xml_path.resolve().relative_to(prev_dir_abs).as_posix()
            except Exception:
                rel_xml_path = "index.xml"

            leaves = xml.xpath("//*[local-name()='leaf']")
            for leaf in leaves:
                if leaf.get("ID") == leaf_id:
                    href = _leaf_href(leaf)
                    return prev_seq, rel_xml_path, leaf_id, href

    return None
def validate_xml(xml_path: Path, dtd_path: Path):
    """Validate XML against provided DTD, raise if invalid."""
    dtd = etree.DTD(str(dtd_path))
    xml = etree.parse(str(xml_path))
    if not dtd.validate(xml):
        raise ValueError(f"DTD validation failed for {xml_path.name}:\n{dtd.error_log.filter_from_errors()}")
    print(f"✔ Validated {xml_path.name} against {dtd_path.name}")

# =========================================================
# XML generation
# =========================================================

def _normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())

def _parse_dtd_entities(dtd_text: str) -> dict[str, str]:
    entities: dict[str, str] = {}
    for name, v1, v2 in re.findall(
        r'<!ENTITY\s+%\s*([^\s]+)\s+(?:"([^"]*)"|\'([^\']*)\')\s*>',
        dtd_text,
        flags=re.S,
    ):
        entities[name] = v1 or v2 or ""
    return entities

def _expand_entities(text: str, entities: dict[str, str]) -> str:
    for _ in range(5):
        changed = False
        for name, val in entities.items():
            key = f"%{name};"
            if key in text:
                text = text.replace(key, val)
                changed = True
        if not changed:
            break
    return text

def _parse_dtd_structure(dtd_path: Path) -> dict[str, list[str]]:
    text = dtd_path.read_text(encoding="utf-8", errors="ignore")
    text = re.sub(r"<!--.*?-->", "", text, flags=re.S)
    entities = _parse_dtd_entities(text)
    out: dict[str, list[str]] = {}
    for match in re.finditer(r"<!ELEMENT\s+([^\s]+)\s+\(([^>]+)\)>", text, flags=re.S):
        name = match.group(1)
        content = _expand_entities(match.group(2), entities)
        tokens = re.findall(r"[A-Za-z0-9_.:-]+", content)
        children: list[str] = []
        seen = set()
        for tok in tokens:
            if tok == "#PCDATA":
                continue
            if tok not in seen:
                children.append(tok)
                seen.add(tok)
        if children:
            out[name] = children
    return out

def _find_dtd_public_id(text: str) -> str:
    match = re.search(r'PUBLIC\s+"([^"]+)"', text)
    return match.group(1) if match else ""

def _find_dtd_root_element(text: str, preferred: list[str]) -> str:
    for name in preferred:
        if re.search(rf'<!ELEMENT\s+{re.escape(name)}\b', text):
            return name
    match = re.search(r'<!ELEMENT\s+([^\s>]+)\s', text)
    return match.group(1) if match else ""

def _find_dtd_fixed_value(text: str, element: str, attr: str) -> str:
    pattern = rf'<!ATTLIST\s+{re.escape(element)}\s+.*?\b{re.escape(attr)}\b[^#]*#FIXED\s+(?:"([^"]+)"|\'([^\']+)\')'
    match = re.search(pattern, text, flags=re.S)
    if not match:
        return ""
    return match.group(1) or match.group(2) or ""

def _parse_dtd_enum_list(raw: str) -> list[str]:
    cleaned = raw.strip()
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = cleaned[1:-1]
    tokens = []
    for part in cleaned.split("|"):
        tok = part.strip()
        if tok:
            tokens.append(tok)
    return tokens

def _find_dtd_entity_enum(text: str, entity_name: str) -> list[str]:
    pattern = rf'<!ENTITY\s+%\s*{re.escape(entity_name)}\s+(?:"([^"]+)"|\'([^\']+)\')\s*>'
    match = re.search(pattern, text, flags=re.S)
    if not match:
        return []
    raw = match.group(1) or match.group(2) or ""
    return _parse_dtd_enum_list(raw)

def _find_dtd_attlist_enum(text: str, element: str, attr: str) -> list[str]:
    pattern = rf'<!ATTLIST\s+{re.escape(element)}\s+[^>]*?\b{re.escape(attr)}\s+\(([^)]+)\)'
    match = re.search(pattern, text, flags=re.S)
    if not match:
        return []
    return _parse_dtd_enum_list(match.group(1))

def _build_doctype(root_name: str, system_id: str, public_id: str = "") -> str:
    if public_id:
        return f'<!DOCTYPE {root_name} PUBLIC "{public_id}" "{system_id}">'
    return f'<!DOCTYPE {root_name} SYSTEM "{system_id}">'

def _load_ich_dtd_info(seq_dir: Path) -> dict[str, object]:
    info: dict[str, object] = {}
    dtd_path = seq_dir / UTIL_DTD_DIR / ICH_DTD_FILENAME
    if not dtd_path.exists():
        return info
    text = dtd_path.read_text(encoding="utf-8", errors="ignore")
    root_name = _find_dtd_root_element(text, ["ectd:ectd"])
    if root_name:
        info["root_name"] = root_name
        info["ns_ectd"] = _find_dtd_fixed_value(text, root_name, "xmlns:ectd")
        info["ns_xlink"] = _find_dtd_fixed_value(text, root_name, "xmlns:xlink")
        info["dtd_version"] = _find_dtd_fixed_value(text, root_name, "dtd-version")
    info["public_id"] = _find_dtd_public_id(text)
    order_map = _parse_dtd_structure(dtd_path)
    info["order_map"] = order_map
    module_map: dict[str, str] = {}
    module_order: list[str] = []
    if root_name and root_name in order_map:
        for child in order_map[root_name]:
            norm = _normalize_key(child)
            for mk in ("m1", "m2", "m3", "m4", "m5"):
                if norm.startswith(mk):
                    if mk not in module_map:
                        module_map[mk] = child
                        module_order.append(mk)
                    break
    if module_map:
        info["module_map"] = module_map
    if module_order:
        info["module_order"] = module_order
    return info

def _load_eu_dtd_info(seq_dir: Path) -> dict[str, object]:
    info: dict[str, object] = {}
    eu_dtd = seq_dir / UTIL_DTD_DIR / EU_DTD_FILENAME
    env_mod = seq_dir / UTIL_DTD_DIR / EU_ENVELOPE_MOD_FILENAME
    if eu_dtd.exists():
        text = eu_dtd.read_text(encoding="utf-8", errors="ignore")
        root_name = _find_dtd_root_element(text, ["eu:eu-backbone"])
        if root_name:
            info["root_name"] = root_name
            info["ns_eu"] = _find_dtd_fixed_value(text, root_name, "xmlns:eu")
            info["ns_xlink"] = _find_dtd_fixed_value(text, root_name, "xmlns:xlink")
            info["dtd_version"] = _find_dtd_fixed_value(text, root_name, "dtd-version")
        info["public_id"] = _find_dtd_public_id(text)
        info["countries"] = set(_find_dtd_entity_enum(text, "countries"))
        info["languages"] = set(_find_dtd_entity_enum(text, "languages"))
        info["pi_types"] = set(_find_dtd_attlist_enum(text, "pi-doc", "type"))
        order_map = _parse_dtd_structure(eu_dtd)
        info["order_map"] = order_map
        m1_order = order_map.get("m1-eu", [])
        if m1_order:
            info["m1_order"] = m1_order
            child_order_maps: dict[str, dict[str, int]] = {}
            for tag in m1_order:
                children = order_map.get(tag)
                if children:
                    child_order_maps[tag] = {child: idx for idx, child in enumerate(children)}
            info["child_order_maps"] = child_order_maps
    if env_mod.exists():
        text = env_mod.read_text(encoding="utf-8", errors="ignore")
        info["env_countries"] = set(_find_dtd_entity_enum(text, "env-countries"))
        info["submission_types"] = set(_find_dtd_attlist_enum(text, "submission", "type"))
        info["submission_modes"] = set(_find_dtd_attlist_enum(text, "submission", "mode"))
        info["submission_unit_types"] = set(_find_dtd_attlist_enum(text, "submission-unit", "type"))
        info["agency_codes"] = set(_find_dtd_attlist_enum(text, "agency", "code"))
        info["procedure_types"] = set(_find_dtd_attlist_enum(text, "procedure", "type"))
    return info

def _order_for_parent(order_map: dict[str, list[str]], parent_tag: str) -> list[str]:
    tag = parent_tag.split("}", 1)[-1]
    if tag in order_map:
        return order_map[tag]
    norm = _normalize_key(tag)
    for key, children in order_map.items():
        if _normalize_key(key) == norm:
            return children
    return []

def _find_path_to_tag(
    order_map: dict[str, list[str]],
    start_tag: str,
    target_tag: str,
) -> list[str] | None:
    start = start_tag.split("}", 1)[-1]
    target = target_tag.split("}", 1)[-1]
    if start == target:
        return [start]
    visited = set()

    def dfs(cur: str, path: list[str]) -> list[str] | None:
        if cur in visited:
            return None
        visited.add(cur)
        for child in order_map.get(cur, []):
            if child == target:
                return path + [child]
            res = dfs(child, path + [child])
            if res:
                return res
        return None

    return dfs(start, [start])

def _segment_code(segment: str, module_key: str) -> str:
    seg_norm = _normalize_key(segment)
    m = re.match(r"^(\d+)([a-z])?(\d+)?", seg_norm)
    if not m:
        return ""
    digits = m.group(1) or ""
    letter = m.group(2) or ""
    digits2 = m.group(3) or ""
    if module_key and len(module_key) == 2 and module_key[0] == "m" and digits.startswith(module_key[1]) and len(digits) > 1:
        digits = digits[1:]
    return f"{digits}{letter}{digits2}"

def _child_code(child_tag: str) -> str:
    tag = child_tag.split("}", 1)[-1]
    tokens = re.split(r"[^a-z0-9]+", tag.lower())
    tokens = [t for t in tokens if t]
    if tokens and re.fullmatch(r"m[1-5]", tokens[0]):
        tokens = tokens[1:]
    code_tokens: list[str] = []
    for tok in tokens:
        if tok.isdigit():
            code_tokens.append(tok)
            continue
        if len(tok) == 1 and tok.isalpha():
            code_tokens.append(tok)
            continue
        if re.fullmatch(r"\d+[a-z]", tok):
            code_tokens.append(tok)
            continue
        if re.fullmatch(r"[a-z]\d+", tok):
            code_tokens.append(tok)
            continue
        break
    return "".join(code_tokens)

def _match_child(children: list[str], segment: str, module_key: str) -> str | None:
    seg_norm = _normalize_key(segment)
    if not seg_norm:
        return None
    seg_code = _segment_code(segment, module_key)
    best = None
    best_score = -1
    best_len = 10**9
    for child in children:
        c_norm = _normalize_key(child)
        score = 0
        if seg_code:
            c_code = _child_code(child)
            if c_code and (c_code == seg_code or c_code.startswith(seg_code) or seg_code in c_code):
                score = 4
        if score == 0:
            if c_norm == seg_norm:
                score = 3
            elif c_norm.endswith(seg_norm) or seg_norm.endswith(c_norm):
                score = 2
            elif seg_norm in c_norm:
                score = 1
        if score > best_score or (score == best_score and len(c_norm) < best_len):
            best = child
            best_score = score
            best_len = len(c_norm)
    return best if best_score > 0 else None

def _is_dynamic_segment(seg: str) -> bool:
    return bool(re.match(r".*[a-zA-Z].*-\d+$", seg))

def _load_ctd_mapping(mapping_path: Path) -> tuple[dict[int, list[dict]], dict[int, list[dict]]]:
    dir_entries: list[dict] = []
    file_entries: list[dict] = []
    with mapping_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            item_type = (row.get("item_type") or "").strip()
            if item_type not in ("directory", "file"):
                continue
            rel = (row.get("relative_path") or "").strip()
            xml = (row.get("xml_element") or "").strip()
            if not rel or not xml:
                continue
            parts = rel.split("/")
            pattern_parts: list[str] = []
            specificity = 0
            match_len = 0
            for seg in parts:
                if _is_dynamic_segment(seg):
                    pattern_parts.append(r"[^/]+")
                else:
                    pattern_parts.append(re.escape(seg))
                    specificity += 1
            if item_type == "file":
                last = parts[-1]
                base, dot, _ext = last.rpartition(".")
                prefix = base if dot else last
                pattern_parts[-1] = re.escape(prefix) + r"[^/]*"
                match_rel = "/".join(parts[:-1] + [prefix])
                match_len = len(match_rel)
            else:
                match_len = len(rel)
            pattern = "^" + "/".join(pattern_parts) + "$"
            entry = {
                "rel": rel,
                "xml": xml,
                "parts": parts,
                "regex": re.compile(pattern),
                "specificity": specificity,
                "match_len": match_len,
            }
            if item_type == "directory":
                dir_entries.append(entry)
            else:
                file_entries.append(entry)
    dir_by_len: dict[int, list[dict]] = {}
    file_by_len: dict[int, list[dict]] = {}
    for e in dir_entries:
        dir_by_len.setdefault(len(e["parts"]), []).append(e)
    for e in file_entries:
        file_by_len.setdefault(len(e["parts"]), []).append(e)
    return dir_by_len, file_by_len

def _match_mapping_tag(prefix: str, mapping_index: dict[int, list[dict]]) -> str | None:
    if not prefix:
        return None
    parts_len = prefix.count("/") + 1
    candidates = mapping_index.get(parts_len, [])
    best = None
    best_spec = -1
    best_len = -1
    for e in candidates:
        if e["regex"].match(prefix):
            if e["specificity"] > best_spec or (
                e["specificity"] == best_spec and e["match_len"] > best_len
            ):
                best = e
                best_spec = e["specificity"]
                best_len = e["match_len"]
    return best["xml"] if best else None

def create_index_xml(
    seq_dir: Path,
    metadata_rows: list[dict],
    base_dir: Path,
    seq_num: int,
    mapping_path: Path | None = None,
    xsl_path: str | None = None,
):
    """Build index.xml (ICH eCTD 3.2) with support for new/replace/delete/append."""
    dtd_info = _load_ich_dtd_info(seq_dir)
    root_name = dtd_info.get("root_name") or "ectd:ectd"
    NS = dtd_info.get("ns_ectd") or ICH_NS_DEFAULT
    XLINK = dtd_info.get("ns_xlink") or XLINK_NS_CANONICAL
    dtd_version = dtd_info.get("dtd_version") or "3.2"
    root_prefix, root_local = root_name.split(":", 1) if ":" in root_name else ("ectd", root_name)

    root = etree.Element(
        "{%s}%s" % (NS, root_local),
        nsmap={root_prefix: NS, "xlink": XLINK},
        attrib={"dtd-version": dtd_version}
    )

    module_map = dtd_info.get("module_map") or {
        "m1": "m1-administrative-information-and-prescribing-information",
        "m2": "m2-common-technical-document-summaries",
        "m3": "m3-quality",
        "m4": "m4-nonclinical-study-reports",
        "m5": "m5-clinical-study-reports",
    }
    order_map: dict[str, list[str]] = dtd_info.get("order_map") or {}
    module_order = dtd_info.get("module_order") or ["m1", "m2", "m3", "m4", "m5"]
    dtd_path = seq_dir / UTIL_DTD_DIR / ICH_DTD_FILENAME
    attr_map: dict[str, set[str]] = {}
    dir_mapping_index: dict[int, list[dict]] = {}
    file_mapping_index: dict[int, list[dict]] = {}
    if mapping_path is None:
        mapping_path = Path(__file__).with_name("ectd_3_2_2_path_to_xml_element_mapping.csv")
    if mapping_path.exists():
        dir_mapping_index, file_mapping_index = _load_ctd_mapping(mapping_path)
    if dtd_path.exists():
        attr_map = _parse_dtd_attlist(dtd_path)
        if not order_map:
            order_map = _parse_dtd_structure(dtd_path)
        dtd_root_name = None
        for key in order_map.keys():
            if _normalize_key(key) == "ectd":
                dtd_root_name = key
                break
        if dtd_root_name:
            new_order: list[str] = []
            for child in order_map.get(dtd_root_name, []):
                norm = _normalize_key(child)
                for mk in module_map.keys():
                    if norm.startswith(mk):
                        module_map[mk] = child
                        new_order.append(mk)
                        break
            if new_order:
                seen = set(new_order)
                module_order = new_order + [mk for mk in module_order if mk not in seen]

    def _new_leaf_id() -> str:
        return f"N{uuid.uuid4().hex}"

    def _add_leaf(parent, row, file_path: str, href_override: str | None = None):
        operation = (row.get("operation") or "new").strip().lower()
        modified_href = (row.get("modified-href") or "").strip() or None
        modified_leaf = (row.get("modified-leaf") or "").strip() or None
        leaf_id = _new_leaf_id()

        if operation == "delete":
            checksum = ""
        else:
            checksum, _ = checksum_for_path(seq_dir, file_path)

        href_value = href_override or file_path or modified_href or ""

        attrs = {
            "checksum": checksum,
            "checksum-type": "MD5",
            "operation": operation,
            "ID": leaf_id,
        }
        if operation != "delete":
            if not href_value:
                raise ValueError("Non-delete operation requires 'file_path' or 'modified-href'.")
            attrs["{%s}href" % XLINK] = href_value

        if operation in ("replace", "delete", "append"):
            prev_leaf_ref = resolve_previous_leaf_ref(
                base_dir,
                seq_num,
                modified_leaf,
                modified_href,
                file_path,
            )

            if prev_leaf_ref:
                prev_seq, prev_xml_path, prev_id, _prev_href = prev_leaf_ref
                attrs["modified-file"] = f"../{prev_seq:04d}/{prev_xml_path}#{prev_id}"
            else:
                hint = modified_leaf or modified_href or file_path
                print(f"⚠ No previous leaf found for {operation} operation on {hint}")

        order = _order_for_parent(order_map, parent.tag)
        leaf = etree.Element("leaf", attrib=attrs)
        if order and order[0] == "leaf":
            insert_idx = 0
            for ch in parent:
                if ch.tag != "leaf":
                    break
                insert_idx += 1
            parent.insert(insert_idx, leaf)
        else:
            parent.append(leaf)
        title = etree.SubElement(leaf, "title")
        title.text = (row.get("title") or os.path.basename(file_path))

    node_cache: dict[tuple[int, str], etree._Element] = {}

    def _ensure_child(parent, tag_name: str, attrib: dict[str, str] | None = None):
        attrib = attrib or {}
        norm_attrib = {_normalize_attr_name(k): v for k, v in attrib.items()}
        key = (id(parent), tag_name, tuple(sorted(norm_attrib.items())))
        if key in node_cache:
            return node_cache[key]
        for child in parent:
            if child.tag != tag_name:
                continue
            if norm_attrib:
                if all(child.get(k) == v for k, v in norm_attrib.items()):
                    node_cache[key] = child
                    return child
            else:
                if not child.attrib:
                    node_cache[key] = child
                    return child
        order = _order_for_parent(order_map, parent.tag)
        child = etree.Element(tag_name, attrib=norm_attrib)
        if order:
            index = 0
            for name in order:
                if name == tag_name:
                    break
                for ch in parent:
                    if ch.tag == name:
                        index += 1
            parent.insert(index, child)
        else:
            parent.append(child)
        node_cache[key] = child
        return child

    def _ensure_path_by_tags(parent, tags: list[str], last_attrib: dict[str, str] | None = None):
        cur = parent
        for idx, tag in enumerate(tags):
            if last_attrib and idx == len(tags) - 1:
                cur = _ensure_child(cur, tag, attrib=last_attrib)
            else:
                cur = _ensure_child(cur, tag)
        return cur

    def _tags_from_mapping(module_tag: str, file_path: str, is_dir_path: bool) -> list[str]:
        if not dir_mapping_index and not file_mapping_index:
            return []
        tags: list[str] = []
        cur = module_tag.split("}", 1)[-1]
        parts = [p for p in file_path.split("/") if p]
        dir_parts = parts if is_dir_path else parts[:-1]
        if dir_mapping_index:
            for i in range(2, len(dir_parts) + 1):
                prefix = "/".join(dir_parts[:i])
                tag = _match_mapping_tag(prefix, dir_mapping_index)
                if not tag:
                    continue
                if tag == cur:
                    continue
                if tags and tag == tags[-1]:
                    continue
                tags.append(tag)
                cur = tag
        if not is_dir_path and file_mapping_index:
            tag = _match_mapping_tag(file_path, file_mapping_index)
            if tag and tag != cur:
                tags.append(tag)
        return tags

    def _tags_from_order(module_tag: str, file_path: str, is_dir_path: bool, module_key: str) -> list[str]:
        parts = [p for p in file_path.split("/") if p]
        segs = parts[1:] if is_dir_path else parts[1:-1]
        tags: list[str] = []
        cur_tag = module_tag
        for seg in segs:
            order = _order_for_parent(order_map, cur_tag)
            if not order:
                continue
            match = _match_child(order, seg, module_key)
            if not match:
                continue
            tags.append(match)
            cur_tag = match
        return tags

    def _tags_for_row(module_tag: str, file_path: str, ctd_toc: str, is_dir_path: bool, module_key: str) -> list[str]:
        if ctd_toc:
            if order_map:
                path = _find_path_to_tag(order_map, module_tag, ctd_toc)
                if path:
                    return path[1:] if len(path) > 1 else []
            return [ctd_toc] if ctd_toc != module_tag.split("}", 1)[-1] else []
        if dir_mapping_index or file_mapping_index:
            return _tags_from_mapping(module_tag, file_path, is_dir_path)
        return _tags_from_order(module_tag, file_path, is_dir_path, module_key)

    def _ensure_path(parent, path_parts: list[str], module_key: str):
        cur = parent
        for seg in path_parts:
            order = _order_for_parent(order_map, cur.tag)
            if not order:
                continue
            match = _match_child(order, seg, module_key)
            if not match:
                continue
            cur = _ensure_child(cur, match)
        return cur

    def _ensure_path_from_mapping(parent, dir_parts: list[str]):
        cur = parent
        if not dir_mapping_index:
            return cur
        for i in range(2, len(dir_parts) + 1):
            prefix = "/".join(dir_parts[:i])
            tag = _match_mapping_tag(prefix, dir_mapping_index)
            if not tag:
                continue
            if tag == cur.tag.split("}", 1)[-1]:
                continue
            cur = _ensure_child(cur, tag)
        return cur

    eu_rows = []
    rows_by_module: dict[str, list[dict]] = {k: [] for k in module_map}
    for row in metadata_rows:
        operation = (row.get("operation") or "new").strip().lower()
        file_path = (row.get("file_path") or "").strip()
        if not file_path:
            if operation == "delete":
                modified_href = (row.get("modified-href") or "").strip()
                modified_leaf = (row.get("modified-leaf") or "").strip()
                if modified_href:
                    file_path = modified_href
                elif modified_leaf:
                    prev_leaf_ref = find_previous_leaf_by_id(base_dir, seq_num, modified_leaf)
                    if prev_leaf_ref and prev_leaf_ref[3]:
                        file_path = prev_leaf_ref[3]
                if not file_path:
                    raise ValueError("Delete operation requires file_path, modified-href, or modified-leaf.")
            else:
                raise ValueError("CSV row is missing required column 'file_path'")
        row_copy = dict(row)
        row_copy["file_path"] = file_path
        if file_path.startswith("m1/eu/") and file_path != "m1/eu/eu-regional.xml":
            eu_rows.append(row_copy)
            continue
        module_key = file_path.split("/", 1)[0]
        if module_key in rows_by_module:
            rows_by_module[module_key].append(row_copy)

    if eu_rows and not any(
        (row.get("file_path") or "").strip() == "m1/eu/eu-regional.xml"
        for row in rows_by_module["m1"]
    ):
        rows_by_module["m1"].append({
            "file_path": "m1/eu/eu-regional.xml",
            "title": "European Union",
            "operation": "new",
        })

    for module_key in module_order:
        rows = rows_by_module.get(module_key) or []
        if not rows:
            continue
        module_elem = _ensure_child(root, module_map[module_key])
        dir_infos: list[dict[str, object]] = []
        dir_rows = []
        for row in rows:
            file_path = (row.get("file_path") or "").strip()
            if not file_path:
                continue
            is_dir_path = _is_directory_path(seq_dir, file_path)
            if not is_dir_path:
                continue
            dir_rows.append(row)
        if dir_rows:
            dir_rows.sort(
                key=lambda r: len([p for p in (r.get("file_path") or "").strip("/").split("/") if p])
            )
        for row in dir_rows:
            file_path = (row.get("file_path") or "").strip()
            custom_attrs = _parse_custom_attributes(row.get("attributes") or "")
            if not custom_attrs:
                continue
            ctd_toc = (row.get("CTD-TOC") or row.get("ctd_toc") or row.get("ctd-toc") or "").strip()
            tags = _tags_for_row(module_elem.tag, file_path, ctd_toc, True, module_key)
            target_tag = tags[-1] if tags else module_elem.tag
            allowed_attrs = _allowed_attrs_for_element(attr_map, target_tag)
            filtered_attrs = _filter_allowed_attrs(
                custom_attrs,
                allowed_attrs,
                file_path,
                target_tag,
            )
            dir_parts = [p for p in file_path.strip("/").split("/") if p]
            base_elem = module_elem
            remaining_tags = tags[:]
            for info in dir_infos:
                parts = info["parts"]
                if len(parts) <= len(dir_parts) and dir_parts[: len(parts)] == parts:
                    dir_tags = info["tags"]
                    if tags[: len(dir_tags)] == dir_tags:
                        base_elem = info["elem"]
                        remaining_tags = tags[len(dir_tags):]
                        break
            if not remaining_tags:
                for k, v in filtered_attrs.items():
                    base_elem.set(_normalize_attr_name(k), v)
                dir_elem = base_elem
            else:
                last_attrib = {_normalize_attr_name(k): v for k, v in filtered_attrs.items()}
                dir_elem = _ensure_path_by_tags(base_elem, remaining_tags, last_attrib=last_attrib)
            dir_infos.append({"parts": dir_parts, "tags": tags, "elem": dir_elem})
        if dir_infos:
            dir_infos.sort(key=lambda d: len(d["parts"]), reverse=True)
        for row in rows:
            file_path = (row.get("file_path") or "").strip()
            if not file_path:
                continue
            ctd_toc = (row.get("CTD-TOC") or row.get("ctd_toc") or row.get("ctd-toc") or "").strip()
            custom_attrs = _parse_custom_attributes(row.get("attributes") or "")
            is_dir_path = _is_directory_path(seq_dir, file_path)
            file_parts = [p for p in file_path.strip("/").split("/") if p]
            tags = _tags_for_row(module_elem.tag, file_path, ctd_toc, is_dir_path, module_key)
            base_elem = module_elem
            remaining_tags = tags[:]
            for info in dir_infos:
                parts = info["parts"]
                if len(parts) <= len(file_parts) and file_parts[: len(parts)] == parts:
                    dir_tags = info["tags"]
                    if tags[: len(dir_tags)] == dir_tags:
                        base_elem = info["elem"]
                        remaining_tags = tags[len(dir_tags):]
                        break
            toc_parent = _ensure_path_by_tags(base_elem, remaining_tags)
            if custom_attrs:
                allowed_attrs = _allowed_attrs_for_element(attr_map, toc_parent.tag)
                filtered_attrs = _filter_allowed_attrs(
                    custom_attrs,
                    allowed_attrs,
                    file_path,
                    toc_parent.tag,
                )
                if filtered_attrs:
                    merged = dict(toc_parent.attrib)
                    for k, v in filtered_attrs.items():
                        merged[_normalize_attr_name(k)] = v
                    parent = toc_parent.getparent()
                    if parent is None:
                        for k, v in merged.items():
                            toc_parent.set(k, v)
                    else:
                        toc_parent = _ensure_child(parent, toc_parent.tag, attrib=merged)
            if not is_dir_path:
                _add_leaf(toc_parent, row, file_path)

    xml_path = seq_dir / "index.xml"
    if xsl_path is None:
        xsl_path = (UTIL_STYLE_DIR / ICH_XSL_FILENAME).as_posix()
    doctype = _build_doctype(
        root_name,
        (UTIL_DTD_DIR / ICH_DTD_FILENAME).as_posix(),
        dtd_info.get("public_id") or "",
    )
    write_xml_with_doctype_and_xsl(
        xml_path,
        root,
        doctype,
        xsl_path,
    )

    validate_xml(xml_path, seq_dir / UTIL_DTD_DIR / ICH_DTD_FILENAME)

# =========================================================

def create_eu_regional_xml(
    seq_dir: Path,
    metadata_rows: list[dict],
    sequence_num: str,
    eu_mapping_path: Path | None = None,
    xsl_path: str | None = None,
) -> Path | None:
    """Build eu-regional.xml (EU M1 3.1) and map EU files into proper sections."""
    eu_rows: list[dict] = []
    for row in metadata_rows:
        operation = (row.get("operation") or "new").strip().lower()
        file_path = (row.get("file_path") or "").strip()
        if not file_path and operation == "delete":
            modified_href = (row.get("modified-href") or "").strip()
            modified_leaf = (row.get("modified-leaf") or "").strip()
            resolved_href = ""
            if modified_href:
                resolved_href = modified_href
            elif modified_leaf:
                prev_leaf_ref = find_previous_leaf_by_id(seq_dir.parent, int(sequence_num), modified_leaf)
                if prev_leaf_ref and prev_leaf_ref[3]:
                    resolved_href = prev_leaf_ref[3]
            if resolved_href and resolved_href.startswith("m1/eu/"):
                file_path = resolved_href
        if file_path.startswith("m1/eu/") and file_path != "m1/eu/eu-regional.xml":
            row_copy = dict(row)
            row_copy["file_path"] = file_path
            eu_rows.append(row_copy)
    if not eu_rows:
        return None

    dtd_info = _load_eu_dtd_info(seq_dir)
    dtd_path = seq_dir / UTIL_DTD_DIR / EU_DTD_FILENAME
    attr_map = _parse_dtd_attlist(dtd_path) if dtd_path.exists() else {}
    root_name = dtd_info.get("root_name") or "eu:eu-backbone"
    NS_EU = dtd_info.get("ns_eu") or EU_NS_DEFAULT
    XLINK = dtd_info.get("ns_xlink") or XLINK_NS_CANONICAL
    dtd_version = dtd_info.get("dtd_version") or "3.1"
    root_prefix, root_local = root_name.split(":", 1) if ":" in root_name else ("eu", root_name)
    root = etree.Element(
        "{%s}%s" % (NS_EU, root_local),
        nsmap={root_prefix: NS_EU, "xlink": XLINK},
        attrib={"dtd-version": dtd_version}
    )
    dir_mapping_index: dict[int, list[dict]] = {}
    file_mapping_index: dict[int, list[dict]] = {}
    if eu_mapping_path is None:
        eu_mapping_path = Path(__file__).with_name("ectd_eu_m1_path_to_xml_element_mapping.csv")
    if eu_mapping_path.exists():
        dir_mapping_index, file_mapping_index = _load_ctd_mapping(eu_mapping_path)

    def _first_value(keys, default=""):
        for key in keys:
            val = (metadata_rows[0].get(key) or "").strip()
            if val:
                return val
        return default

    def _validated_value(keys, allowed, default):
        value = _first_value(keys, default)
        if allowed and value not in allowed:
            return default
        return value

    def _new_leaf_id() -> str:
        return f"N{uuid.uuid4().hex}"

    def _add_leaf(parent, row, file_path: str, href_override: str):
        operation = (row.get("operation") or "new").strip().lower()
        modified_href = (row.get("modified-href") or "").strip() or None
        modified_leaf = (row.get("modified-leaf") or "").strip() or None
        leaf_id = _new_leaf_id()

        if operation == "delete":
            checksum = ""
        else:
            checksum, _ = checksum_for_path(seq_dir, file_path)

        href_value = href_override or file_path or modified_href or ""

        attrs = {
            "checksum": checksum,
            "checksum-type": "MD5",
            "operation": operation,
            "ID": leaf_id,
        }
        if operation != "delete":
            if not href_value:
                raise ValueError("Non-delete operation requires 'file_path' or 'modified-href'.")
            attrs["{%s}href" % XLINK] = href_value

        if operation in ("replace", "delete", "append"):
            prev_leaf_ref = resolve_previous_leaf_ref(
                seq_dir.parent,
                int(sequence_num),
                modified_leaf,
                modified_href,
                file_path,
                fallback_paths=[file_path, href_override],
            )

            if prev_leaf_ref:
                prev_seq, prev_xml_path, prev_id, _prev_href = prev_leaf_ref
                attrs["modified-file"] = f"../{prev_seq:04d}/{prev_xml_path}#{prev_id}"
            else:
                hint = modified_leaf or modified_href or file_path
                print(f"⚠ No previous leaf found for {operation} operation on {hint}")

        leaf = etree.SubElement(parent, "leaf", attrib=attrs)
        title = etree.SubElement(leaf, "title")
        title.text = (row.get("title") or os.path.basename(file_path))

    def _ensure_child(parent, tag, attrib=None, order_map=None):
        attrib = attrib or {}
        for child in parent:
            if child.tag == tag and all(child.get(k) == v for k, v in attrib.items()):
                return child
        new_elem = etree.Element(tag, attrib=attrib)
        if order_map and tag in order_map:
            insert_at = len(parent)
            new_order = order_map[tag]
            for idx, child in enumerate(list(parent)):
                child_order = order_map.get(child.tag)
                if child_order is not None and child_order > new_order:
                    insert_at = idx
                    break
            parent.insert(insert_at, new_elem)
        else:
            parent.append(new_elem)
        return new_elem

    countries = dtd_info.get("countries") or {
        "at", "be", "bg", "common", "cy", "cz", "de", "dk", "edqm", "ee", "el", "es", "ema",
        "fi", "fr", "hr", "hu", "ie", "is", "it", "li", "lt", "lu", "lv", "mt", "nl", "no",
        "pl", "pt", "ro", "se", "si", "sk", "uk", "xi",
    }
    languages = dtd_info.get("languages") or {
        "bg", "cs", "da", "de", "el", "en", "es", "et", "fi", "fr", "ga", "hr", "hu", "is",
        "it", "lt", "lv", "mt", "nl", "no", "pl", "pt", "ro", "sk", "sl", "sv",
    }
    pi_types = dtd_info.get("pi_types") or {"spc", "annex2", "outer", "interpack", "impack", "other", "pl", "combined"}
    env_countries = dtd_info.get("env_countries") or countries
    submission_types = dtd_info.get("submission_types") or set()
    submission_modes = dtd_info.get("submission_modes") or set()
    submission_unit_types = dtd_info.get("submission_unit_types") or set()
    agency_codes = dtd_info.get("agency_codes") or set()
    procedure_types = dtd_info.get("procedure_types") or set()

    eu_envelope = etree.SubElement(root, "eu-envelope")
    env_country = _validated_value(["eu_country"], env_countries, EU_DEFAULTS["env_country"])
    if env_country not in env_countries:
        env_country = EU_DEFAULTS["env_country"]
    envelope = etree.SubElement(eu_envelope, "envelope", attrib={"country": env_country})
    etree.SubElement(envelope, "identifier").text = _first_value(["eu_identifier"], str(uuid.uuid4()))

    submission = etree.SubElement(
        envelope,
        "submission",
        attrib={
            "type": _validated_value(["eu_submission_type"], submission_types, EU_DEFAULTS["submission_type"]),
            "mode": _validated_value(["eu_submission_mode"], submission_modes, EU_DEFAULTS["submission_mode"]),
        },
    )
    submission_number = _first_value(["eu_submission_number"], "")
    if submission_number:
        etree.SubElement(submission, "number").text = submission_number
    proc_tracking = etree.SubElement(submission, "procedure-tracking")
    etree.SubElement(proc_tracking, "number").text = _first_value(
        ["eu_procedure_number"],
        EU_DEFAULTS["procedure_number"],
    )

    etree.SubElement(
        envelope,
        "submission-unit",
        attrib={"type": _validated_value(
            ["eu_submission_unit_type"],
            submission_unit_types,
            EU_DEFAULTS["submission_unit_type"],
        )},
    )
    etree.SubElement(envelope, "applicant").text = _first_value(
        ["applicant_name"],
        EU_DEFAULTS["applicant_name"],
    )
    etree.SubElement(
        envelope,
        "agency",
        attrib={"code": _validated_value(
            ["eu_agency_code"],
            agency_codes,
            EU_DEFAULTS["agency_code"],
        )},
    )
    etree.SubElement(
        envelope,
        "procedure",
        attrib={"type": _validated_value(
            ["eu_procedure_type"],
            procedure_types,
            EU_DEFAULTS["procedure_type"],
        )},
    )
    etree.SubElement(envelope, "invented-name").text = _first_value(
        ["eu_invented_name"],
        EU_DEFAULTS["invented_name"],
    )
    inn_value = _first_value(["eu_inn"], "")
    if inn_value:
        etree.SubElement(envelope, "inn").text = inn_value
    etree.SubElement(envelope, "sequence").text = sequence_num
    etree.SubElement(envelope, "related-sequence").text = _first_value(
        ["eu_related_sequence"], sequence_num
    )
    etree.SubElement(envelope, "submission-description").text = _first_value(
        ["sequence_description"],
        EU_DEFAULTS["submission_description"],
    )

    m1_eu = etree.SubElement(root, "m1-eu")
    m1_order = dtd_info.get("m1_order") or [
        "m1-0-cover",
        "m1-2-form",
        "m1-3-pi",
        "m1-4-expert",
        "m1-5-specific",
        "m1-6-environrisk",
        "m1-7-orphan",
        "m1-8-pharmacovigilance",
        "m1-9-clinical-trials",
        "m1-10-paediatrics",
        "m1-responses",
        "m1-additional-data",
    ]
    m1_order_map = {tag: idx for idx, tag in enumerate(m1_order)}
    child_order_maps = dtd_info.get("child_order_maps") or {
        "m1-3-pi": {
            "m1-3-1-spc-label-pl": 0,
            "m1-3-2-mockup": 1,
            "m1-3-3-specimen": 2,
            "m1-3-4-consultation": 3,
            "m1-3-5-approved": 4,
            "m1-3-6-braille": 5,
        },
        "m1-4-expert": {
            "m1-4-1-quality": 0,
            "m1-4-2-non-clinical": 1,
            "m1-4-3-clinical": 2,
        },
        "m1-5-specific": {
            "m1-5-1-bibliographic": 0,
            "m1-5-2-generic-hybrid-bio-similar": 1,
            "m1-5-3-data-market-exclusivity": 2,
            "m1-5-4-exceptional-circumstances": 3,
            "m1-5-5-conditional-ma": 4,
        },
        "m1-6-environrisk": {
            "m1-6-1-non-gmo": 0,
            "m1-6-2-gmo": 1,
        },
        "m1-7-orphan": {
            "m1-7-1-similarity": 0,
            "m1-7-2-market-exclusivity": 1,
        },
        "m1-8-pharmacovigilance": {
            "m1-8-1-pharmacovigilance-system": 0,
            "m1-8-2-risk-management-system": 1,
        },
    }

    def _find_token(parts, allowed, default):
        for part in parts:
            if part in allowed:
                return part
        return default

    def _map_m1_section(section: str, parts: list[str]):
        if section == "additional-data":
            return "m1-additional-data", None, "specific"
        if section == "responses":
            return "m1-responses", None, "specific"
        if section.startswith("110"):
            return "m1-10-paediatrics", None, "leaf"
        if section.startswith("10"):
            return "m1-0-cover", None, "specific"
        if section.startswith("12"):
            return "m1-2-form", None, "specific"
        if section.startswith("13"):
            pi_child_map = {
                "31": "m1-3-1-spc-label-pl",
                "32": "m1-3-2-mockup",
                "33": "m1-3-3-specimen",
                "34": "m1-3-4-consultation",
                "35": "m1-3-5-approved",
                "36": "m1-3-6-braille",
            }
            child_key = parts[1][:2] if len(parts) > 1 else ""
            return "m1-3-pi", pi_child_map.get(child_key, "m1-3-1-spc-label-pl"), "pi-doc"
        if section.startswith("14"):
            expert_map = {
                "41": "m1-4-1-quality",
                "42": "m1-4-2-non-clinical",
                "43": "m1-4-3-clinical",
            }
            child_key = parts[1][:2] if len(parts) > 1 else ""
            return "m1-4-expert", expert_map.get(child_key, "m1-4-1-quality"), "leaf"
        if section.startswith("15"):
            specific_map = {
                "51": "m1-5-1-bibliographic",
                "52": "m1-5-2-generic-hybrid-bio-similar",
                "53": "m1-5-3-data-market-exclusivity",
                "54": "m1-5-4-exceptional-circumstances",
                "55": "m1-5-5-conditional-ma",
            }
            child_key = parts[1][:2] if len(parts) > 1 else ""
            return "m1-5-specific", specific_map.get(child_key, "m1-5-1-bibliographic"), "leaf"
        if section.startswith("16"):
            env_map = {
                "61": "m1-6-1-non-gmo",
                "62": "m1-6-2-gmo",
            }
            child_key = parts[1][:2] if len(parts) > 1 else ""
            return "m1-6-environrisk", env_map.get(child_key, "m1-6-1-non-gmo"), "leaf"
        if section.startswith("17"):
            orphan_map = {
                "71": "m1-7-1-similarity",
                "72": "m1-7-2-market-exclusivity",
            }
            child_key = parts[1][:2] if len(parts) > 1 else ""
            return "m1-7-orphan", orphan_map.get(child_key, "m1-7-1-similarity"), "leaf"
        if section.startswith("18"):
            pv_map = {
                "81": "m1-8-1-pharmacovigilance-system",
                "82": "m1-8-2-risk-management-system",
            }
            child_key = parts[1][:2] if len(parts) > 1 else ""
            return "m1-8-pharmacovigilance", pv_map.get(child_key, "m1-8-1-pharmacovigilance-system"), "leaf"
        if section.startswith("19"):
            return "m1-9-clinical-trials", None, "leaf"
        return "m1-additional-data", None, "specific"

    def _container_type_for_section(section_tag: str) -> str:
        order_map = dtd_info.get("order_map")
        if order_map and section_tag in order_map:
            children = order_map[section_tag]
            if "specific" in children:
                return "specific"
            if "pi-doc" in children:
                return "pi-doc"
            return "leaf"
        if section_tag in {
            "m1-0-cover",
            "m1-2-form",
            "m1-5-specific",
            "m1-responses",
            "m1-additional-data",
        }:
            return "specific"
        if section_tag == "m1-3-pi":
            return "pi-doc"
        return "leaf"

    def _mapped_section_tags(file_path: str) -> tuple[str | None, str | None]:
        if not dir_mapping_index and not file_mapping_index:
            return None, None
        parts = [p for p in file_path.split("/") if p]
        tags: list[str] = []
        if dir_mapping_index:
            for i in range(2, len(parts)):
                prefix = "/".join(parts[:i])
                tag = _match_mapping_tag(prefix, dir_mapping_index)
                if not tag:
                    continue
                if tags and tag == tags[-1]:
                    continue
                tags.append(tag)
        if file_mapping_index:
            file_tag = _match_mapping_tag(file_path, file_mapping_index)
            if file_tag and (not tags or file_tag != tags[-1]):
                tags.append(file_tag)
        tags = [t for t in tags if t != "m1-eu"]
        section_tag = tags[0] if tags else None
        child_tag = tags[1] if len(tags) > 1 else None
        return section_tag, child_tag

    dir_infos: list[dict[str, object]] = []
    for row in eu_rows:
        file_path = (row.get("file_path") or "").strip()
        if not file_path:
            continue
        if not _is_directory_path(seq_dir, file_path):
            continue
        custom_attrs = _parse_custom_attributes(row.get("attributes") or "")
        if not custom_attrs:
            continue
        rel_path = file_path[len("m1/eu/"):]
        parts = [p for p in rel_path.split("/") if p]
        if not parts:
            continue

        section_tag, child_tag = _mapped_section_tags(file_path)
        if not section_tag:
            section = parts[0]
            section_tag, child_tag, container_type = _map_m1_section(section, parts)
        else:
            container_type = _container_type_for_section(section_tag)
        section_elem = _ensure_child(m1_eu, section_tag, order_map=m1_order_map)
        if child_tag:
            section_elem = _ensure_child(
                section_elem,
                child_tag,
                order_map=child_order_maps.get(section_tag),
            )

        if container_type == "specific":
            country = parts[1] if len(parts) > 1 else "common"
            if country not in countries:
                country = "common"
            container = _ensure_child(section_elem, "specific", {"country": country})
        elif container_type == "pi-doc":
            pi_country = _find_token(parts, countries, EU_DEFAULTS["pi_country"])
            pi_lang = _find_token(parts, languages, EU_DEFAULTS["pi_lang"])
            pi_type = _find_token(parts, pi_types, EU_DEFAULTS["pi_type"])
            container = _ensure_child(
                section_elem,
                "pi-doc",
                {"country": pi_country, "type": pi_type, "{http://www.w3.org/XML/1998/namespace}lang": pi_lang},
            )
        else:
            container = section_elem

        allowed_attrs = _allowed_attrs_for_element(attr_map, container.tag)
        filtered_attrs = _filter_allowed_attrs(
            custom_attrs,
            allowed_attrs,
            file_path,
            container.tag,
        )
        if filtered_attrs:
            merged = dict(container.attrib)
            for k, v in filtered_attrs.items():
                merged[_normalize_attr_name(k)] = v
            parent = container.getparent()
            if parent is not None:
                container = _ensure_child(parent, container.tag, attrib=merged)
            else:
                for k, v in merged.items():
                    container.set(k, v)

        dir_parts = [p for p in file_path.strip("/").split("/") if p]
        dir_infos.append({"parts": dir_parts, "elem": container})

    if dir_infos:
        dir_infos.sort(key=lambda d: len(d["parts"]), reverse=True)

    for row in eu_rows:
        file_path = (row.get("file_path") or "").strip()
        rel_path = file_path[len("m1/eu/"):]
        parts = [p for p in rel_path.split("/") if p]
        if not parts:
            continue
        custom_attrs = _parse_custom_attributes(row.get("attributes") or "")
        is_dir_path = _is_directory_path(seq_dir, file_path)

        file_parts = [p for p in file_path.strip("/").split("/") if p]
        container = None
        for info in dir_infos:
            parts_prefix = info["parts"]
            if len(parts_prefix) <= len(file_parts) and file_parts[: len(parts_prefix)] == parts_prefix:
                container = info["elem"]
                break
        if container is None:
            section_tag, child_tag = _mapped_section_tags(file_path)
            if not section_tag:
                section = parts[0]
                section_tag, child_tag, container_type = _map_m1_section(section, parts)
            else:
                container_type = _container_type_for_section(section_tag)
            section_elem = _ensure_child(m1_eu, section_tag, order_map=m1_order_map)
            if child_tag:
                section_elem = _ensure_child(
                    section_elem,
                    child_tag,
                    order_map=child_order_maps.get(section_tag),
                )

            if container_type == "specific":
                country = parts[1] if len(parts) > 1 else "common"
                if country not in countries:
                    country = "common"
                container = _ensure_child(section_elem, "specific", {"country": country})
            elif container_type == "pi-doc":
                pi_country = _find_token(parts, countries, EU_DEFAULTS["pi_country"])
                pi_lang = _find_token(parts, languages, EU_DEFAULTS["pi_lang"])
                pi_type = _find_token(parts, pi_types, EU_DEFAULTS["pi_type"])
                container = _ensure_child(
                    section_elem,
                    "pi-doc",
                    {"country": pi_country, "type": pi_type, "{http://www.w3.org/XML/1998/namespace}lang": pi_lang},
                )
            else:
                container = section_elem
        if custom_attrs:
            allowed_attrs = _allowed_attrs_for_element(attr_map, container.tag)
            filtered_attrs = _filter_allowed_attrs(
                custom_attrs,
                allowed_attrs,
                file_path,
                container.tag,
            )
            if filtered_attrs:
                merged = dict(container.attrib)
                for k, v in filtered_attrs.items():
                    merged[_normalize_attr_name(k)] = v
                parent = container.getparent()
                if parent is not None:
                    container = _ensure_child(parent, container.tag, attrib=merged)
                else:
                    for k, v in merged.items():
                        container.set(k, v)

        if not is_dir_path:
            _add_leaf(container, row, file_path, rel_path)

    m1_eu_dir = seq_dir / "m1" / "eu"
    m1_eu_dir.mkdir(parents=True, exist_ok=True)
    xml_path = m1_eu_dir / "eu-regional.xml"

    if xsl_path is None:
        xsl_path = (UTIL_STYLE_DIR / EU_XSL_FILENAME).as_posix()
    doctype = _build_doctype(
        root_name,
        (Path("..") / ".." / UTIL_DTD_DIR / EU_DTD_FILENAME).as_posix(),
        dtd_info.get("public_id") or "",
    )
    write_xml_with_doctype_and_xsl(
        xml_path,
        root,
        doctype,
        xsl_path,
    )

    if dtd_path.exists():
        validate_xml(xml_path, dtd_path)
    else:
        print(f"⚠️  EU regional DTD not found at {dtd_path}; skipped validation for eu-regional.xml")
    return xml_path

# =========================================================
# Main entry
# =========================================================

def main(
    base_dir: str,
    sequence_num: str,
    scan: bool = False,
    mapfile: str | None = None,
    extract_xml: bool = False,
    eu_mapfile: str | None = None,
    xsl_path: str | None = None,
    eu_xsl_path: str | None = None,
):
    base = Path(base_dir)
    seq_dir = base / sequence_num
    xlsx_file = base / f"metadata-{sequence_num}.xlsx"
    mapping_path = Path(mapfile) if mapfile else None
    eu_mapping_path = Path(eu_mapfile) if eu_mapfile else None

    if scan:
        count = write_metadata_from_scan(seq_dir, xlsx_file)
        print(f"✔ Wrote {count} rows to {xlsx_file}")

    if xlsx_file.exists():
        ensure_metadata_columns(xlsx_file)

    if extract_xml:
        updated, appended, envelope_updates = extract_metadata_from_xml(seq_dir, xlsx_file)
        print(
            f"✔ Extracted XML into {xlsx_file}: "
            f"{appended} new rows, {updated} filled cells, {envelope_updates} envelope cells"
        )

    if scan or extract_xml:
        print("ℹ Skipped backbone generation due to -scan/-extractXML.")
        return

    if not xlsx_file.exists():
        raise FileNotFoundError(f"Metadata Excel missing: {xlsx_file}")

    metadata = load_metadata(xlsx_file)
    seq_int = int(sequence_num)

    print(f"Starting backbone generation for sequence {sequence_num} …")
    create_eu_regional_xml(
        seq_dir,
        metadata,
        sequence_num,
        eu_mapping_path=eu_mapping_path,
        xsl_path=eu_xsl_path,
    )
    create_index_xml(
        seq_dir,
        metadata,
        base,
        seq_int,
        mapping_path=mapping_path,
        xsl_path=xsl_path,
    )
    write_index_md5(seq_dir, seq_dir / "index.xml")
    print("✅ All XML backbones created and validated!")

if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(
        description=(
            "Generate eCTD XML backbones from Excel metadata (metadata-<seq>.xlsx). "
            "Excel columns must include file_path, title, attributes (backbone element attrs), "
            "operation, modified-leaf, modified-href and optional ctd-toc (override target XML element tag), "
            "plus EU envelope fields (eu_country, eu_identifier, eu_submission_type, eu_submission_mode, "
            "eu_submission_number, eu_procedure_number, eu_submission_unit_type, eu_agency_code, "
            "eu_procedure_type, eu_invented_name, eu_inn, eu_related_sequence)."
        )
    )
    ap.add_argument("base_directory")
    ap.add_argument("sequence_number")
    ap.add_argument(
        "-scan",
        action="store_true",
        help="Scan sequence folder and append missing file_path rows to metadata-<seq>.xlsx.",
    )
    ap.add_argument(
        "-extractXML",
        action="store_true",
        help="Fill missing metadata fields from backbone XML files (index.xml, eu-regional.xml).",
    )
    ap.add_argument(
        "-eu_mapfile",
        default=None,
        help=(
            "Path to EU M1 mapping CSV for eu-regional.xml "
            "(default: ectd_eu_m1_path_to_xml_element_mapping.csv). "
            "Required columns: item_type (directory|file), relative_path, xml_element."
        ),
    )
    ap.add_argument(
        "-mapfile",
        default=None,
        help=(
            "Path to mapping CSV for path->XML element tags "
            "(default: ectd_3_2_2_path_to_xml_element_mapping.csv). "
            "Required columns: item_type (directory|file), relative_path, xml_element."
        ),
    )
    ap.add_argument(
        "-xsl",
        default=None,
        help="Path to XSL for index.xml (default: util/style/ectd-2-0.xsl).",
    )
    ap.add_argument(
        "-eu_xsl",
        default=None,
        help="Path to XSL for eu-regional.xml (default: util/style/eu-regional.xsl).",
    )
    args = ap.parse_args()
    main(
        args.base_directory,
        args.sequence_number,
        scan=args.scan,
        mapfile=args.mapfile,
        extract_xml=args.extractXML,
        eu_mapfile=args.eu_mapfile,
        xsl_path=args.xsl,
        eu_xsl_path=args.eu_xsl,
    )
