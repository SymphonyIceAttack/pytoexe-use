# -*- coding: utf-8 -*-
"""
Генератор Excel-файла с системой скоринга дистрибьюторов
Создаёт трёхлистовую структуру:
- Лист1: исходные данные (36 столбцов)
- Скоринг: автоматический расчёт баллов по 6 блокам
- Методика: описание критериев оценки

Запуск: python generate_scoring_system.py
Зависимости: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ──────────────────────────────────────────────
# 0. СТИЛИ ОФОРМЛЕНИЯ
# ──────────────────────────────────────────────
BLUE_DARK = PatternFill("solid", fgColor="1F4E79")
BLUE_MED = PatternFill("solid", fgColor="2E75B6")
BLUE_LIGHT = PatternFill("solid", fgColor="D6E4F0")
GREEN_LIGHT = PatternFill("solid", fgColor="E2EFDA")
ORANGE_LIGHT = PatternFill("solid", fgColor="FCE4D6")
YELLOW_LIGHT = PatternFill("solid", fgColor="FFF2CC")
RED_LIGHT = PatternFill("solid", fgColor="F8D7DA")
GRAY_LIGHT = PatternFill("solid", fgColor="F2F2F2")

WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
WHITE_BOLD_12 = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
DARK_BOLD = Font(name="Calibri", bold=True, color="1F4E79", size=10)
DARK_NORMAL = Font(name="Calibri", color="333333", size=10)
SCORE_FONT = Font(name="Calibri", bold=True, color="1F4E79", size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_cell(cell, fill=None, font=None, alignment=None, border=THIN_BORDER):
    if fill: cell.fill = fill
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if border: cell.border = border


# ──────────────────────────────────────────────
# 1. СТРУКТУРА СТОЛБЦОВ ЛИСТА "ЛИСТ1" (36 столбцов)
# ──────────────────────────────────────────────
COLUMNS_MAIN = [
    ("№", 5, "num"),
    ("Основная компания", 30, "text"),
    ("Аффилированные компании", 25, "text"),
    ("ИНН", 14, "text"),
    ("Руководство (ФИО, должность, контактные данные)", 35, "text"),
    ("Менеджеры (ФИО, должность, контактные данные)", 35, "text"),
    ("Выделенный менеджер под нашу компанию", 20, "list_yesno"),
    ("Главный офис (город)", 18, "text"),
    ("Федеральный округ", 18, "text"),
    ("Основной регион", 20, "text"),
    ("Дополнительный регион", 20, "text"),
    ("Перспективный регион (экспансия)", 22, "text"),
    ("Год регистрации", 10, "num"),
    ("Стаж работы с компанией", 12, "num"),
    ("Отсутствие в реестре недобросовестных поставщиков", 18, "list_rnp"),
    ("Оборот за 2025 г., млн. руб.", 16, "num"),
    ("Динамика оборота (рост/падение к пред. году), %", 18, "num"),
    ("Дебиторская задолженность (млн руб.)", 16, "num"),
    ("Общее количество обслуживаемых ЛПУ", 12, "num"),
    ("Название ЛПУ с которыми работают (сокращенное)", 30, "text"),
    ("Количество целевых ЛПУ", 12, "num"),
    ("Название целевых ЛПУ (сокращенное)", 30, "text"),
    ("Охват целевых ЛПУ, %", 12, "num"),
    ("Самостоятельная подготовка аукционов", 18, "list_auction"),
    ("Направления работы", 25, "num"),
    ("Складская логистика", 18, "list_warehouse"),
    ("Уровень сервиса (логистика, сроки доставки, наличие консультирующих специалистов)", 18, "num"),
    ("Маркетинговая активность (готовность дистрибьютора проводить маркетинговые мероприятия для ЛПУ, участие в выставках/конференциях)", 18, "list_marketing"),
    ("Продажи общей линейки", 16, "num"),
    ("Продажи линейки АК", 16, "num"),
    ("Продажи линейки TS", 16, "num"),
    ("Основные конкуренты в портфеле (перечень)", 30, "text"),
    ("Конфликт интересов", 12, "num"),
    ("Коммуникация", 16, "list_communication"),
    ("Планирование работы, проведенная работа, следующий шаг (звонок, отправить письмо, встреча иди др.) с указанием сроков, отчет о выполнении шага)", 35, "text"),
    ("Комментарии менеджера", 35, "text"),
]

TOTAL_COLS_MAIN = len(COLUMNS_MAIN)
DATA_START_ROW = 3
DATA_END_ROW = 52


# ──────────────────────────────────────────────
# 2. СОЗДАНИЕ КНИГИ И ЛИСТОВ
# ──────────────────────────────────────────────
wb = openpyxl.Workbook()

# Лист 1: Основные данные
ws1 = wb.active
ws1.title = "Лист1"
ws1.sheet_properties.tabColor = "1F4E79"

# Лист 2: Скоринг
ws2 = wb.create_sheet("Скоринг")
ws2.sheet_properties.tabColor = "E2EFDA"

# Лист 3: Методика
ws3 = wb.create_sheet("Методика")
ws3.sheet_properties.tabColor = "FCE4D6"


# ──────────────────────────────────────────────
# 3. ЛИСТ "ЛИСТ1" - ШАПКА И НАСТРОЙКИ
# ──────────────────────────────────────────────
# Заголовки столбцов
for i, col_def in enumerate(COLUMNS_MAIN, start=1):
    cell = ws1.cell(row=1, column=i, value=col_def[0])
    style_cell(cell, fill=BLUE_DARK, font=WHITE_BOLD_12, alignment=CENTER)
    ws1.column_dimensions[get_column_letter(i)].width = col_def[1]

# Высота шапки и закрепление
ws1.row_dimensions[1].height = 40
ws1.freeze_panes = "A3"


# ──────────────────────────────────────────────
# 4. ВЫПАДАЮЩИЕ СПИСКИ ДЛЯ ЛИСТА "ЛИСТ1"
# ──────────────────────────────────────────────
def add_dv(ws, col_letter, formula, start_row, end_row):
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Ошибка",
        error="Выберите значение из списка"
    )
    dv.prompt = "Выберите из списка"
    dv.promptTitle = "Ввод данных"
    dv.showInputMessage = True
    dv.showDropDown = False
    ws.add_data_validation(dv)
    for r in range(start_row, end_row + 1):
        dv.add(ws[f"{col_letter}{r}"])

# Применение выпадающих списков
for i, col_def in enumerate(COLUMNS_MAIN, start=1):
    letter = get_column_letter(i)
    dtype = col_def[2]
    if dtype == "list_yesno":
        add_dv(ws1, letter, '"Да,Нет,Частично"', DATA_START_ROW, DATA_END_ROW)
    elif dtype == "list_rnp":
        add_dv(ws1, letter, '"Отсутствует,Уточняется,В реестре"', DATA_START_ROW, DATA_END_ROW)
    elif dtype == "list_auction":
        add_dv(ws1, letter, '"Полностью,Частично,Не участвует"', DATA_START_ROW, DATA_END_ROW)
    elif dtype == "list_warehouse":
        add_dv(ws1, letter, '"А,В,С,Нет"', DATA_START_ROW, DATA_END_ROW)
    elif dtype == "list_marketing":
        add_dv(ws1, letter, '"Активно,Точечно,Не проводит"', DATA_START_ROW, DATA_END_ROW)
    elif dtype == "list_communication":
        add_dv(ws1, letter, '"Проактивен,В течение дня,Игнорирует"', DATA_START_ROW, DATA_END_ROW)


# ──────────────────────────────────────────────
# 5. ЛИСТ "ЛИСТ1" - НУМЕРАЦИЯ И ФОРМАТИРОВАНИЕ
# ──────────────────────────────────────────────
for r in range(DATA_START_ROW, DATA_END_ROW + 1):
    # Номер строки
    ws1.cell(row=r, column=1, value=r - DATA_START_ROW + 1)
    style_cell(ws1.cell(row=r, column=1), fill=GRAY_LIGHT, font=DARK_NORMAL, alignment=CENTER)
    
    # Остальные ячейки
    for c in range(2, TOTAL_COLS_MAIN + 1):
        cell = ws1.cell(row=r, column=c)
        cell.border = THIN_BORDER
        cell.font = DARK_NORMAL
        cell.alignment = LEFT if COLUMNS_MAIN[c-1][2] == "text" else CENTER
        # Чередование цвета строк
        if (r - DATA_START_ROW) % 2 == 1:
            cell.fill = PatternFill("solid", fgColor="F8F9FA")


# ──────────────────────────────────────────────
# 6. ЛИСТ "СКОРИНГ" - СТРУКТУРА СТОЛБЦОВ
# ──────────────────────────────────────────────
SCORING_COLUMNS = [
    ("№", 5),
    ("Основная компания", 30),
    ("ИНН", 14),
    ("Б1: Идентификация (0-15)", 12),
    ("Б2: Юр. безопасность (0-25)", 12),
    ("Б3: Клиентская база (0-25)", 12),
    ("Б4: Ассортимент (0-15)", 12),
    ("Б5: Конкуренты (0-10)", 12),
    ("Б6: Коммуникация (0-10)", 12),
    ("ИТОГО (0-100)", 12),
    ("Категория", 20),
]

# Заголовки листа "Скоринг"
for i, (name, width) in enumerate(SCORING_COLUMNS, start=1):
    cell = ws2.cell(row=1, column=i, value=name)
    style_cell(cell, fill=BLUE_DARK, font=WHITE_BOLD_12, alignment=CENTER)
    ws2.column_dimensions[get_column_letter(i)].width = width

ws2.row_dimensions[1].height = 40
ws2.freeze_panes = "A2"


# ──────────────────────────────────────────────
# 7. ЛИСТ "СКОРИНГ" - ФОРМУЛЫ РАСЧЁТА БАЛЛОВ
# ──────────────────────────────────────────────
def col_l(n):
    return get_column_letter(n)

for r in range(DATA_START_ROW, DATA_END_ROW + 1):
    # № п/п
    ws2.cell(row=r, column=1, value=r - DATA_START_ROW + 1)
    style_cell(ws2.cell(row=r, column=1), fill=GRAY_LIGHT, font=DARK_NORMAL, alignment=CENTER)
    
    # Название компании (ссылка на Лист1)
    f_name = f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";ИНДЕКС('Лист1'!B:B;ПОИСКПОЗ(A{r};'Лист1'!A:A;0)))"
    ws2.cell(row=r, column=2, value=f_name)
    
    # ИНН (ссылка на Лист1)
    f_inn = f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";ИНДЕКС('Лист1'!D:D;ПОИСКПОЗ(A{r};'Лист1'!A:A;0)))"
    ws2.cell(row=r, column=3, value=f_inn)
    
    # ── БЛОК 1: Идентификация и ресурсы (0-15) ──
    # Столбцы Лист1: G=Выделенный менеджер, E=Руководство, F=Менеджеры, D=ИНН, J-L=Регионы
    f_b1 = (
        f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";"
        f"("
        # ИНН в базе (столбец D)
        f"ЕСЛИ('Лист1'!D{r}<>\"\";5;0)+"
        # Руководство (столбец E) - наличие данных
        f"ЕСЛИ('Лист1'!E{r}<>\"\";5;0)+"
        # Менеджеры (столбец F) - кол-во человек
        f"ЕСЛИ('Лист1'!F{r}>=5;5;ЕСЛИ('Лист1'!F{r}>=2;3;0))+"
        # Выделенный менеджер (столбец G)
        f"ЕСЛИ('Лист1'!G{r}=\"Да\";5;ЕСЛИ('Лист1'!G{r}=\"Частично\";3;0))+"
        # География (столбцы J, K, L)
        f"ЕСЛИ(СЧЁТЕСЛИ('Лист1'!J{r}:L{r};\"<>\")>=3;5;ЕСЛИ(СЧЁТЕСЛИ('Лист1'!J{r}:L{r};\"<>\")>=2;3;0))"
        f"))"
    )
    ws2.cell(row=r, column=4, value=f_b1)
    
    # ── БЛОК 2: Юр. и фин. безопасность (0-25) ──
    # Столбцы Лист1: M=Год регистрации, N=Стаж, O=РНП, P=Оборот, Q=Динамика, R=Дебиторка
    f_b2 = (
        f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";"
        f"("
        # Год регистрации (возраст > 7 лет)
        f"ЕСЛИ(2026-'Лист1'!M{r}>7;5;ЕСЛИ(2026-'Лист1'!M{r}>=3;3;0))+"
        # Стаж работы с компанией
        f"ЕСЛИ('Лист1'!N{r}>=5;5;ЕСЛИ('Лист1'!N{r}>=2;3;0))+"
        # Отсутствие в РНП
        f"ЕСЛИ('Лист1'!O{r}=\"Отсутствует\";5;ЕСЛИ('Лист1'!O{r}=\"Уточняется\";3;0))+"
        # Оборот
        f"ЕСЛИ('Лист1'!P{r}>500;5;ЕСЛИ('Лист1'!P{r}>=100;3;0))+"
        # Динамика оборота
        f"ЕСЛИ('Лист1'!Q{r}>=10;5;ЕСЛИ('Лист1'!Q{r}>=-10;3;0))+"
        # Дебиторская задолженность (< 5% от оборота)
        f"ЕСЛИ(И('Лист1'!P{r}>0;'Лист1'!R{r}/'Лист1'!P{r}<0.05);5;"
        f"ЕСЛИ(И('Лист1'!P{r}>0;'Лист1'!R{r}/'Лист1'!P{r}<0.15);3;0))"
        f"))"
    )
    ws2.cell(row=r, column=5, value=f_b2)
    
    # ── БЛОК 3: Клиентская база (0-25) ──
    # Столбцы Лист1: S=Общее кол-во ЛПУ, U=Кол-во целевых ЛПУ, W=Охват %, X=Аукционы, Y=Направления
    f_b3 = (
        f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";"
        f"("
        # Общее кол-во ЛПУ
        f"ЕСЛИ('Лист1'!S{r}>100;5;ЕСЛИ('Лист1'!S{r}>=30;3;0))+"
        # Кол-во целевых ЛПУ
        f"ЕСЛИ('Лист1'!U{r}>50;5;ЕСЛИ('Лист1'!U{r}>=20;3;0))+"
        # Охват целевых ЛПУ
        f"ЕСЛИ('Лист1'!W{r}>=80;5;ЕСЛИ('Лист1'!W{r}>=40;3;0))+"
        # Самостоятельная подготовка аукционов
        f"ЕСЛИ('Лист1'!X{r}=\"Полностью\";5;ЕСЛИ('Лист1'!X{r}=\"Частично\";3;0))+"
        # Направления работы
        f"ЕСЛИ('Лист1'!Y{r}>=5;5;ЕСЛИ('Лист1'!Y{r}>=2;3;0))"
        f"))"
    )
    ws2.cell(row=r, column=6, value=f_b3)
    
    # ── БЛОК 4: Работа с ассортиментом (0-15) ──
    # Столбцы Лист1: Z=Склад, AA=Сервис, AB=Маркетинг, AC-AE=Продажи линеек
    f_b4 = (
        f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";"
        f"("
        # Складская логистика
        f"ЕСЛИ('Лист1'!Z{r}=\"А\";5;ЕСЛИ('Лист1'!Z{r}=\"В\";3;0))+"
        # Уровень сервиса
        f"ЕСЛИ('Лист1'!AA{r}>=4;5;ЕСЛИ('Лист1'!AA{r}>=3;3;0))+"
        # Маркетинговая активность
        f"ЕСЛИ('Лист1'!AB{r}=\"Активно\";5;ЕСЛИ('Лист1'!AB{r}=\"Точечно\";3;0))+"
        # Продажи по линейкам (сумма AC + AD + AE)
        f"ЕСЛИ(('Лист1'!AC{r}+'Лист1'!AD{r}+'Лист1'!AE{r})>2000;5;"
        f"ЕСЛИ(('Лист1'!AC{r}+'Лист1'!AD{r}+'Лист1'!AE{r})>=500;3;0))"
        f"))"
    )
    ws2.cell(row=r, column=7, value=f_b4)
    
    # ── БЛОК 5: Конкурентная среда (0-10) ──
    # Столбцы Лист1: AG=Конфликт интересов (1-5), AF=Основные конкуренты
    f_b5 = (
        f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";"
        f"ЕСЛИ('Лист1'!AG{r}=1;10;ЕСЛИ('Лист1'!AG{r}<=3;5;0)))"
    )
    ws2.cell(row=r, column=8, value=f_b5)
    
    # ── БЛОК 6: Статус и коммуникация (0-10) ──
    # Столбцы Лист1: AH=Коммуникация, AI=Планирование, AJ=Комментарии
    f_b6 = (
        f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";"
        f"("
        # Коммуникация
        f"ЕСЛИ('Лист1'!AH{r}=\"Проактивен\";5;ЕСЛИ('Лист1'!AH{r}=\"В течение дня\";3;0))+"
        # Планирование
        f"ЕСЛИ(И(ДЛНСТР('Лист1'!AI{r})>20;ПОИСК(\"срок\";'Лист1'!AI{r})>0);5;"
        f"ЕСЛИ(ДЛНСТР('Лист1'!AI{r})>10;3;0))"
        f"))"
    )
    ws2.cell(row=r, column=9, value=f_b6)
    
    # ── ИТОГО ──
    f_total = f"=ЕСЛИ('Лист1'!A{r}=\"\";\"\";СУММ(D{r}:I{r}))"
    ws2.cell(row=r, column=10, value=f_total)
    
    # ── Категория ──
    f_cat = (
        f"=ЕСЛИ(J{r}=\"\";\"\";"
        f"ЕСЛИ(J{r}>=80;\"★ Приоритетный\";"
        f"ЕСЛИ(J{r}>=50;\"● Перспективный\";"
        f"ЕСЛИ(J{r}>=25;\"▲ Рабочий\";\"✖ Кандидат на замену\"))))"
    )
    ws2.cell(row=r, column=11, value=f_cat)
    
    # Стилизация ячеек скоринга
    for sc in range(4, 12):
        cell = ws2.cell(row=r, column=sc)
        cell.font = SCORE_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if sc == 10:
            cell.fill = YELLOW_LIGHT
        elif sc == 11:
            cell.fill = GREEN_LIGHT
    
    # Стилизация для названия и ИНН
    for sc in [2, 3]:
        cell = ws2.cell(row=r, column=sc)
        cell.border = THIN_BORDER
        cell.font = DARK_NORMAL
        cell.alignment = LEFT if sc == 2 else CENTER


# ──────────────────────────────────────────────
# 8. УСЛОВНОЕ ФОРМАТИРОВАНИЕ ЛИСТА "СКОРИНГ"
# ──────────────────────────────────────────────
total_range = f"J{DATA_START_ROW}:J{DATA_END_ROW}"

# Зеленый >= 80
ws2.conditional_formatting.add(total_range, CellIsRule(
    operator="greaterThanOrEqual", formula=["80"],
    fill=PatternFill("solid", fgColor="C6EFCE"),
    font=Font(bold=True, color="006100")
))
# Желтый 50-79
ws2.conditional_formatting.add(total_range, CellIsRule(
    operator="between", formula=["50", "79"],
    fill=PatternFill("solid", fgColor="FFEB9C"),
    font=Font(bold=True, color="9C6500")
))
# Оранжевый 25-49
ws2.conditional_formatting.add(total_range, CellIsRule(
    operator="between", formula=["25", "49"],
    fill=PatternFill("solid", fgColor="FCE4D6"),
    font=Font(bold=True, color="974706")
))
# Красный < 25
ws2.conditional_formatting.add(total_range, CellIsRule(
    operator="lessThan", formula=["25"],
    fill=PatternFill("solid", fgColor="FFC7CE"),
    font=Font(bold=True, color="9C0006")
))

# Условное форматирование категории
cat_range = f"K{DATA_START_ROW}:K{DATA_END_ROW}"
ws2.conditional_formatting.add(cat_range, CellIsRule(
    operator="containsText", formula=['"Приоритетный"'],
    fill=PatternFill("solid", fgColor="C6EFCE")
))
ws2.conditional_formatting.add(cat_range, CellIsRule(
    operator="containsText", formula=['"Перспективный"'],
    fill=PatternFill("solid", fgColor="FFEB9C")
))
ws2.conditional_formatting.add(cat_range, CellIsRule(
    operator="containsText", formula=['"Рабочий"'],
    fill=PatternFill("solid", fgColor="FCE4D6")
))
ws2.conditional_formatting.add(cat_range, CellIsRule(
    operator="containsText", formula=['"замену"'],
    fill=PatternFill("solid", fgColor="FFC7CE")
))


# ──────────────────────────────────────────────
# 9. ЛИСТ "МЕТОДИКА" - ОПИСАНИЕ КРИТЕРИЕВ
# ──────────────────────────────────────────────
methodology = [
    ["МЕТОДИКА БАЛЛЬНОЙ ОЦЕНКИ ДИСТРИБЬЮТОРОВ", "", ""],
    ["", "", ""],
    ["Блок", "Критерий", "Баллы"],
    ["", "", ""],
    ["Б1: Идентификация и ресурсы (макс. 15)", "", ""],
    ["", "ИНН есть в базе 1С: Есть = 5, Нет = 0", ""],
    ["", "Руководство (опыт в отрасли): Есть данные = 5, Нет = 0", ""],
    ["", "Менеджеры (кол-во): ≥5 = 5, 2-4 = 3, <2 = 0", ""],
    ["", "Выделенный менеджер: Да = 5, Частично = 3, Нет = 0", ""],
    ["", "География присутствия: 3+ региона = 5, 2 региона = 3, 1 регион = 0", ""],
    ["", "", ""],
    ["Б2: Юр. и фин. безопасность (макс. 25)", "", ""],
    ["", "Возраст компании: >7 лет = 5, 3-7 лет = 3, <3 лет = 0", ""],
    ["", "Стаж работы с нами: ≥5 лет = 5, 2-4 года = 3, <2 лет = 0", ""],
    ["", "Отсутствие в РНП: Отсутствует = 5, Уточняется = 3, В реестре = 0", ""],
    ["", "Оборот за 2025: >500 млн = 5, 100-500 млн = 3, <100 млн = 0", ""],
    ["", "Динамика оборота: ≥10% = 5, -10%...+10% = 3, <-10% = 0", ""],
    ["", "Дебиторка: <5% оборота = 5, 5-15% = 3, >15% = 0", ""],
    ["", "", ""],
    ["Б3: Клиентская база (макс. 25)", "", ""],
    ["", "Общее кол-во ЛПУ: >100 = 5, 30-100 = 3, <30 = 0", ""],
    ["", "Кол-во целевых ЛПУ: >50 = 5, 20-50 = 3, <20 = 0", ""],
    ["", "Охват целевых ЛПУ: ≥80% = 5, 40-79% = 3, <40% = 0", ""],
    ["", "Аукционы: Полностью = 5, Частично = 3, Не участвует = 0", ""],
    ["", "Направления работы: ≥5 = 5, 2-4 = 3, 1 = 0", ""],
    ["", "", ""],
    ["Б4: Работа с ассортиментом (макс. 15)", "", ""],
    ["", "Складская логистика: Класс А = 5, Класс В = 3, С/Нет = 0", ""],
    ["", "Уровень сервиса: ≥4 = 5, 3 = 3, <3 = 0", ""],
    ["", "Маркетинговая активность: Активно = 5, Точечно = 3, Не проводит = 0", ""],
    ["", "Продажи по линейкам: >2000 тыс = 5, 500-2000 тыс = 3, <500 тыс = 0", ""],
    ["", "", ""],
    ["Б5: Конкурентная среда (макс. 10)", "", ""],
    ["", "Конфликт интересов (1-5): 1 = 10, 2-3 = 5, 4-5 = 0", ""],
    ["", "(1 = нет конфликта, 5 = прямой конкурент в приоритете)", ""],
    ["", "", ""],
    ["Б6: Статус и коммуникация (макс. 10)", "", ""],
    ["", "Коммуникация: Проактивен = 5, В течение дня = 3, Игнорирует = 0", ""],
    ["", "Планирование работы: Есть сроки = 5, Есть план = 3, Нет = 0", ""],
    ["", "", ""],
    ["КАТЕГОРИИ ПАРТНЁРОВ", "", ""],
    ["", "80-100 баллов → ★ Приоритетный (стратегический)", ""],
    ["", "50-79 баллов  → ● Перспективный (развивать)", ""],
    ["", "25-49 баллов  → ▲ Рабочий (контролировать)", ""],
    ["", "0-24 баллов   → ✖ Кандидат на замену", ""],
]

for row_idx, row_data in enumerate(methodology, start=1):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font = DARK_NORMAL
        cell.alignment = LEFT
        if row_idx == 1:
            cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        elif row_idx == 3:
            cell.font = WHITE_BOLD
            cell.fill = BLUE_MED
        elif val and val.startswith("Б") and ":" in val:
            cell.font = Font(name="Calibri", bold=True, size=11, color="2E75B6")
            cell.fill = BLUE_LIGHT
        elif val and val.startswith("КАТЕГОРИИ"):
            cell.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")

ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 65
ws3.column_dimensions["C"].width = 15


# ──────────────────────────────────────────────
# 10. СОХРАНЕНИЕ ФАЙЛА
# ──────────────────────────────────────────────
filename = "Критерии_оценки_дистрибьюторов_со_скорингом.xlsx"
wb.save(filename)

print("=" * 70)
print(f"✅ Файл успешно создан: {filename}")
print("=" * 70)
print(f"\n📋 Структура файла:")
print(f"   • Лист 'Лист1': {TOTAL_COLS_MAIN} столбцов, {DATA_END_ROW - DATA_START_ROW + 1} строк для данных")
print(f"   • Лист 'Скоринг': автоматический расчет баллов по 6 блокам")
print(f"   • Лист 'Методика': полное описание критериев оценки")
print(f"\n📊 Система оценки:")
print(f"   • Максимальный балл: 100")
print(f"   • Блоков оценки: 6")
print(f"   • Категорий партнёров: 4")
print(f"\n🔧 Инструкция:")
print(f"   1. Откройте файл в Excel")
print(f"   2. Заполните данные на листе 'Лист1'")
print(f"   3. Баллы автоматически рассчитаются на листе 'Скоринг'")
print(f"   4. Категория партнёра определяется по итоговому баллу")
print(f"\n💡 Примечание:")
print(f"   • Исходная таблица 'Лист1' не изменяется")
print(f"   • Все расчёты происходят на отдельном листе 'Скоринг'")
print(f"   • Формулы используют ссылки ИНДЕКС/ПОИСКПОЗ")
print("=" * 70)