#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Генератор XML для Аршин и Росаккредитации
Полная версия на Python, воспроизводящая функционал оригинального HTML/JS приложения.

Функции:
- Генерация XML для ФГИС «Аршин» (поверка средств измерений)
- Генерация XML для Росаккредитации (сведения о результатах поверки)
- Чтение данных из Excel (XLSX, XLS)
- Создание шаблонов Excel для обоих типов
- Автоматическое форматирование дат
- Корректная обработка пригодных/непригодных СИ
- Экранирование спецсимволов для XML
"""

import re
import os
import sys
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple, Union

# Попытка импорта openpyxl, если не установлена — выводим инструкцию
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Ошибка: библиотека 'openpyxl' не найдена. Установите её командой: pip install openpyxl")
    sys.exit(1)

# ----------------------------------------------------------------------
# Вспомогательные функции
# ----------------------------------------------------------------------

def safe_xml(text: str) -> str:
    """Экранирование спецсимволов для XML."""
    if not isinstance(text, str):
        text = str(text)
    replacements = {
        '&': '&amp;',
        '<': '&lt;',
        '>': '&gt;',
        '"': '&quot;',
        "'": '&apos;'
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text

def format_date(date_value: Any) -> str:
    """
    Преобразование даты из различных форматов в YYYY-MM-DD.
    Поддерживает: строки (dd.mm.yyyy, dd-mm-yyyy, yyyy-mm-dd, yyyy/mm/dd),
    числа (Excel serial), объекты datetime.
    """
    if not date_value:
        return ""

    # Если уже строка
    if isinstance(date_value, str):
        date_str = date_value.strip()
        if not date_str:
            return ""
        # Убираем лишние пробелы, заменяем разделители
        cleaned = re.sub(r'[./]', '-', date_str)
        # Пробуем разные форматы
        patterns = [
            (r'^(\d{4})-(\d{1,2})-(\d{1,2})$', False),  # YYYY-MM-DD
            (r'^(\d{1,2})-(\d{1,2})-(\d{4})$', True),   # DD-MM-YYYY
            (r'^(\d{4})-(\d{1,2})-(\d{1,2})$', False),  # повтор
        ]
        for pattern, is_dmy in patterns:
            match = re.match(pattern, cleaned)
            if match:
                if is_dmy:
                    day, month, year = match.group(1), match.group(2), match.group(3)
                else:
                    year, month, day = match.group(1), match.group(2), match.group(3)
                try:
                    dt = datetime(int(year), int(month), int(day))
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue
        # Попробуем через datetime общий парсинг
        for fmt in ("%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y.%m.%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
        # Если ничего не подошло, возвращаем исходную строку (может быть ошибка)
        return date_str

    # Если число (Excel serial)
    if isinstance(date_value, (int, float)):
        # Excel serial (1900 system) -> datetime
        try:
            # Дата в Excel: 1 января 1900 = 1
            # datetime.fromordinal работает с 1 годом, поэтому используем timedelta
            from datetime import timedelta
            excel_base = datetime(1899, 12, 30)
            delta = timedelta(days=date_value)
            dt = excel_base + delta
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return str(date_value)

    # Если уже datetime
    if isinstance(date_value, datetime):
        return date_value.strftime("%Y-%m-%d")

    return str(date_value)

def is_unfit(value: Any) -> bool:
    """Определяет, является ли значение признаком 'непригодно'."""
    if not value:
        return False
    val = str(value).strip().lower()
    return val in ('непригодно', 'непригоден', 'брак', 'false', 'нет', '0')

def normalize_verification_result(value: Any) -> str:
    """Нормализует результат поверки в 'пригодно' или 'непригодно'."""
    if not value:
        return 'пригодно'
    val = str(value).strip().lower()
    if val in ('не пригоден', 'непригоден', 'непригодно', 'брак', 'false', 'нет', '0'):
        return 'непригодно'
    if val in ('пригодно', 'годен', 'true', 'да', '1', '+'):
        return 'пригодно'
    return val  # fallback

def split_full_name(full_name: str) -> Dict[str, str]:
    """Разделяет ФИО на фамилию и имя (первое слово и второе)."""
    if not full_name:
        return {'lastName': 'Неизвестно', 'firstName': 'Неизвестно'}
    parts = str(full_name).strip().split()
    if len(parts) >= 2:
        return {'lastName': parts[0], 'firstName': parts[1]}
    return {'lastName': 'Неизвестно', 'firstName': 'Неизвестно'}

# ----------------------------------------------------------------------
# Генерация XML для Аршин
# ----------------------------------------------------------------------

def generate_arshin_xml(data_rows: List[List[str]]) -> str:
    """
    Генерирует XML для ФГИС «Аршин» на основе данных строк.
    Каждая строка — список из 22 элементов (согласно таблице).
    """
    xml_lines = ['<?xml version="1.0" encoding="utf-8"?>']
    xml_lines.append('<gost:application xmlns:gost="urn://fgis-arshin.gost.ru/module-verifications/import/2020-06-19">')

    for row in data_rows:
        # Пропускаем пустые строки без номера ГРСИ
        if not row or not row[0]:
            continue

        # Извлекаем значения (до 22)
        values = row + [''] * (22 - len(row))
        num_grsi = values[0]
        serial_number = values[1]
        year = values[2]
        modification = values[3]
        sign_cipher = values[4] or ''
        owner = values[5] or ''
        vrf_date = format_date(values[6])
        valid_date = format_date(values[7]) if not is_unfit(values[8]) and values[7] else ''
        applicability = values[8] or ''
        reason_unfit = values[9] or ''
        doc_title = values[10] or ''
        metrologist = values[11] or ''
        etalon_number = values[12] or ''
        # Дополнительные СИ (типы и номера)
        mi_types = [values[13] or '', values[14] or '', values[15] or '']
        mi_serials = [values[16] or '', values[17] or '', values[18] or '']
        temperature = values[19] or ''
        pressure = values[20] or ''
        humidity = values[21] or ''

        # Начало блока result
        xml_lines.append('    <gost:result>')
        xml_lines.append('        <gost:miInfo>')
        xml_lines.append('            <gost:singleMI>')
        xml_lines.append(f'                <gost:mitypeNumber>{safe_xml(num_grsi)}</gost:mitypeNumber>')
        xml_lines.append(f'                <gost:manufactureNum>{safe_xml(serial_number)}</gost:manufactureNum>')
        if year:
            xml_lines.append(f'                <gost:manufactureYear>{safe_xml(year)}</gost:manufactureYear>')
        if modification:
            xml_lines.append(f'                <gost:modification>{safe_xml(modification)}</gost:modification>')
        xml_lines.append('            </gost:singleMI>')
        xml_lines.append('        </gost:miInfo>')
        if sign_cipher:
            xml_lines.append(f'        <gost:signCipher>{safe_xml(sign_cipher)}</gost:signCipher>')
        if owner:
            xml_lines.append(f'        <gost:miOwner>{safe_xml(owner)}</gost:miOwner>')
        xml_lines.append(f'        <gost:vrfDate>{vrf_date}</gost:vrfDate>')
        if valid_date:
            xml_lines.append(f'        <gost:validDate>{valid_date}</gost:validDate>')
        xml_lines.append('        <gost:type>2</gost:type>')
        xml_lines.append('        <gost:calibration>false</gost:calibration>')

        if is_unfit(applicability):
            xml_lines.append('        <gost:inapplicable>')
            reason = reason_unfit if reason_unfit else 'Причина не указана'
            xml_lines.append(f'            <gost:reasons>{safe_xml(reason)}</gost:reasons>')
            xml_lines.append('        </gost:inapplicable>')
        else:
            xml_lines.append('        <gost:applicable>')
            xml_lines.append('            <gost:signPass>true</gost:signPass>')
            xml_lines.append('            <gost:signMi>true</gost:signMi>')
            xml_lines.append('        </gost:applicable>')

        if doc_title:
            xml_lines.append(f'        <gost:docTitle>{safe_xml(doc_title)}</gost:docTitle>')
        if metrologist:
            xml_lines.append(f'        <gost:metrologist>{safe_xml(metrologist)}</gost:metrologist>')

        # Средства измерений (эталоны и дополнительные)
        xml_lines.append('        <gost:means>')
        if etalon_number:
            xml_lines.append('            <gost:mieta>')
            xml_lines.append(f'                <gost:number>{safe_xml(etalon_number)}</gost:number>')
            xml_lines.append('            </gost:mieta>')

        # Дополнительные СИ
        has_extra = any(t and s for t, s in zip(mi_types, mi_serials))
        if has_extra:
            xml_lines.append('            <gost:mis>')
            for mi_type, mi_serial in zip(mi_types, mi_serials):
                if mi_type and mi_serial:
                    xml_lines.append('                <gost:mi>')
                    xml_lines.append(f'                    <gost:typeNum>{safe_xml(mi_type)}</gost:typeNum>')
                    xml_lines.append(f'                    <gost:manufactureNum>{safe_xml(mi_serial)}</gost:manufactureNum>')
                    xml_lines.append('                </gost:mi>')
            xml_lines.append('            </gost:mis>')
        xml_lines.append('        </gost:means>')

        # Условия поверки (температура, давление, влажность)
        if temperature or pressure or humidity:
            xml_lines.append('        <gost:conditions>')
            if temperature:
                # замена точки на запятую
                temp_val = temperature.replace('.', ',')
                xml_lines.append(f'            <gost:temperature>{safe_xml(temp_val)}</gost:temperature>')
            if pressure:
                press_val = pressure.replace('.', ',')
                xml_lines.append(f'            <gost:pressure>{safe_xml(press_val)}</gost:pressure>')
            if humidity:
                hum_val = humidity.replace('.', ',')
                xml_lines.append(f'            <gost:hymidity>{safe_xml(hum_val)}</gost:hymidity>')
            xml_lines.append('        </gost:conditions>')

        xml_lines.append('    </gost:result>')

    xml_lines.append('</gost:application>')
    return '\n'.join(xml_lines)

# ----------------------------------------------------------------------
# Генерация XML для Росаккредитации
# ----------------------------------------------------------------------

def generate_accreditation_xml(data_rows: List[List[str]]) -> str:
    """
    Генерирует XML для Росаккредитации (структура VerificationMeasuringInstrumentData).
    Каждая строка: номер, тип СИ, дата поверки, дата действия, результат, СНИЛС, ФИО.
    """
    xml_lines = ["<?xml version='1.0' encoding='UTF-8'?>"]
    xml_lines.append('<Message xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="schema.xsd">')
    xml_lines.append('  <VerificationMeasuringInstrumentData>')

    for row in data_rows:
        if not row or not row[0]:
            continue
        # Дополняем до 7 элементов
        values = row + [''] * (7 - len(row))
        number_verification = values[0]
        type_mi = values[1]
        date_verification = format_date(values[2])
        date_end = format_date(values[3])
        result_raw = values[4] or ''
        snils = values[5] or ''
        full_name = values[6] or ''

        result_code = '1' if normalize_verification_result(result_raw) == 'пригодно' else '2'
        name_parts = split_full_name(full_name)

        xml_lines.append('    <VerificationMeasuringInstrument>')
        xml_lines.append(f'      <NumberVerification>{safe_xml(number_verification)}</NumberVerification>')
        xml_lines.append(f'      <DateVerification>{date_verification}</DateVerification>')
        xml_lines.append(f'      <DateEndVerification>{date_end}</DateEndVerification>')
        xml_lines.append(f'      <TypeMeasuringInstrument>{safe_xml(type_mi)}</TypeMeasuringInstrument>')
        xml_lines.append('      <ApprovedEmployees>')
        xml_lines.append('        <Name>')
        xml_lines.append(f'          <Last>{safe_xml(name_parts["lastName"])}</Last>')
        xml_lines.append(f'          <First>{safe_xml(name_parts["firstName"])}</First>')
        xml_lines.append('        </Name>')
        xml_lines.append(f'        <SNILS>{safe_xml(snils)}</SNILS>')
        xml_lines.append('      </ApprovedEmployees>')
        xml_lines.append(f'      <ResultVerification>{result_code}</ResultVerification>')
        xml_lines.append('    </VerificationMeasuringInstrument>')

    xml_lines.append('  </VerificationMeasuringInstrumentData>')
    xml_lines.append('  <SaveMethod>1</SaveMethod>')
    xml_lines.append('</Message>')
    return '\n'.join(xml_lines)

# ----------------------------------------------------------------------
# Работа с Excel
# ----------------------------------------------------------------------

def create_arshin_template(filepath: str) -> None:
    """Создаёт шаблон Excel для Аршин с заголовками и примерами данных."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные для Аршин"

    headers = [
        "Номер в ГРСИ", "Заводской номер", "Год выпуска", "Модификация",
        "Шифр знака", "Владелец СИ", "Дата поверки", "Дата действия",
        "Пригодность", "Причина непригодности", "Наименование документа", "ФИО метролога",
        "Номер СИ/эталона", "Тип СИ 1", "Тип СИ 2", "Тип СИ 3",
        "Заводской номер СИ 1", "Заводской номер СИ 2", "Заводской номер СИ 3",
        "Температура", "Давление", "Влажность"
    ]

    # Пример данных (пригодный прибор)
    example_data = [
        "47957-11", "3026171", "2013", "ТОП-0,66 У3", "ГЗО",
        "Восточное производственное отделение", "18.07.2025", "17.07.2030",
        "пригодно", "", "ГОСТ 8.217-2024", "Наседкин Дмитрий Александрович",
        "27007.04.2Р.00564940", "14883-95", "46434-11", "59851-15",
        "99936", "ОВ-51", "52719", "21,4", "101,8", "30,7"
    ]

    # Пример непригодного прибора
    unfit_example = [
        "2746-71", "327567", "1975", "М416", "ГЗС",
        'филиал ПАО "Россети Волга" - "Самарские РС", Самарское ПО, Красноярский РЭС',
        "17.11.2025", "", "непригодно", "Превышение допустимой погрешности", "ГОСТ 8.409-81",
        "Кондрашин Дмитрий Витальевич", "6332.77.4Р.00304693", "46434-11", "", "",
        "ОВ-51", "", "", "22,9", "756,5", "32,8"
    ]

    # Запись заголовков
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Запись примеров
    for row_idx, row_data in enumerate([example_data, unfit_example], start=2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Автоширина колонок
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    wb.save(filepath)

def create_accreditation_template(filepath: str) -> None:
    """Создаёт шаблон Excel для аккредитации."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные для аккредитации"

    headers = [
        "Номер Аршин", "Тип поверяемого СИ", "Дата результатов поверки",
        "Дата действия результатов поверки", "Результат поверки", "СНИЛС", "Фамилия и Имя"
    ]

    example = [
        "С-ГЗО/18-07-2025/462546997", "ЗНОМ-35-65 У1", "18.07.2025", "17.07.2030",
        "пригодно", "10786907691", "Наседкин Дмитрий"
    ]
    unfit_example = [
        "С-ГЗО/18-07-2025/462546998", "ЗНОМ-35-65 У1", "18.07.2025", "", "непригодно",
        "10786907692", "Иванов Петр"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate([example, unfit_example], start=2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(filepath)

def read_excel_data(filepath: str, sheet_name: Optional[str] = None, skip_header: bool = True) -> List[List[str]]:
    """
    Читает данные из Excel и возвращает список строк (каждая строка — список значений).
    По умолчанию пропускает первую строку (заголовки).
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Файл {filepath} не найден.")

    wb = load_workbook(filepath, data_only=True)
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    rows_data = []
    for row in ws.iter_rows(values_only=True):
        # Преобразуем None в пустую строку, остальное в строку
        str_row = [str(cell) if cell is not None else '' for cell in row]
        rows_data.append(str_row)

    if skip_header and len(rows_data) > 1:
        rows_data = rows_data[1:]  # убираем заголовки

    # Удаляем полностью пустые строки
    rows_data = [row for row in rows_data if any(cell.strip() for cell in row)]

    return rows_data

# ----------------------------------------------------------------------
# Интерфейс командной строки
# ----------------------------------------------------------------------

def print_menu():
    print("\n" + "=" * 60)
    print("ГЕНЕРАТОР XML ДЛЯ АРШИН И РОСАККРЕДИТАЦИИ")
    print("=" * 60)
    print("1. Сгенерировать XML для Аршин из Excel файла")
    print("2. Сгенерировать XML для Росаккредитации из Excel файла")
    print("3. Скачать шаблон Excel для Аршин")
    print("4. Скачать шаблон Excel для Росаккредитации")
    print("5. Выход")
    print("=" * 60)

def save_xml_file(content: str, default_name: str) -> None:
    """Сохраняет XML содержимое в файл с запросом имени."""
    print("\nВведите имя файла для сохранения (без расширения) или оставьте пустым для имени по умолчанию:")
    user_name = input("Имя файла: ").strip()
    if not user_name:
        file_name = default_name
    else:
        file_name = user_name
    if not file_name.endswith('.xml'):
        file_name += '.xml'

    with open(file_name, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"Файл успешно сохранён как {file_name}")

def main():
    if not OPENPYXL_AVAILABLE:
        print("Невозможно продолжить: требуется библиотека openpyxl.")
        sys.exit(1)

    while True:
        print_menu()
        choice = input("Выберите действие (1-5): ").strip()

        if choice == '1':
            file_path = input("Введите путь к Excel файлу для Аршин: ").strip()
            if not file_path:
                print("Путь не указан.")
                continue
            try:
                rows = read_excel_data(file_path)
                if not rows:
                    print("В файле нет данных (после удаления заголовков).")
                    continue
                # Обработка дат и нормализация пригодности
                processed_rows = []
                for row in rows:
                    # Дополняем до 22 колонок
                    processed = list(row)
                    # Если есть колонка с датой поверки (7я по счёту, индекс 6)
                    if len(processed) > 6:
                        processed[6] = format_date(processed[6])
                    if len(processed) > 7:
                        processed[7] = format_date(processed[7])
                    if len(processed) > 8:
                        if is_unfit(processed[8]):
                            processed[8] = 'непригодно'
                        else:
                            processed[8] = 'пригодно'
                    processed_rows.append(processed)
                xml_content = generate_arshin_xml(processed_rows)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                save_xml_file(xml_content, f"arshin_export_{timestamp}.xml")
                print("Генерация XML для Аршин завершена.")
            except Exception as e:
                print(f"Ошибка: {e}")

        elif choice == '2':
            file_path = input("Введите путь к Excel файлу для Росаккредитации: ").strip()
            if not file_path:
                print("Путь не указан.")
                continue
            try:
                rows = read_excel_data(file_path)
                if not rows:
                    print("В файле нет данных.")
                    continue
                processed_rows = []
                for row in rows:
                    processed = list(row)
                    # Даты: индексы 2 и 3 (третья и четвёртая колонки)
                    if len(processed) > 2:
                        processed[2] = format_date(processed[2])
                    if len(processed) > 3:
                        processed[3] = format_date(processed[3])
                    if len(processed) > 4:
                        processed[4] = normalize_verification_result(processed[4])
                    processed_rows.append(processed)
                xml_content = generate_accreditation_xml(processed_rows)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                save_xml_file(xml_content, f"accreditation_export_{timestamp}.xml")
                print("Генерация XML для Росаккредитации завершена.")
            except Exception as e:
                print(f"Ошибка: {e}")

        elif choice == '3':
            file_path = input("Введите имя файла для сохранения шаблона Аршин (по умолчанию: шаблон_Аршин.xlsx): ").strip()
            if not file_path:
                file_path = "шаблон_Аршин.xlsx"
            elif not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            try:
                create_arshin_template(file_path)
                print(f"Шаблон для Аршин успешно создан: {file_path}")
            except Exception as e:
                print(f"Ошибка при создании шаблона: {e}")

        elif choice == '4':
            file_path = input("Введите имя файла для сохранения шаблона аккредитации (по умолчанию: шаблон_аккредитация.xlsx): ").strip()
            if not file_path:
                file_path = "шаблон_аккредитация.xlsx"
            elif not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            try:
                create_accreditation_template(file_path)
                print(f"Шаблон для аккредитации успешно создан: {file_path}")
            except Exception as e:
                print(f"Ошибка при создании шаблона: {e}")

        elif choice == '5':
            print("До свидания!")
            break
        else:
            print("Неверный выбор. Пожалуйста, введите число от 1 до 5.")

if __name__ == "__main__":
    main()