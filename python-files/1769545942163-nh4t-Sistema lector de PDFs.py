import os
import re
import pdfplumber
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ruta
ruta_base = os.path.dirname(os.path.abspath(__file__))
pdf_folder = os.path.join(ruta_base, "PDFs_a_Procesar")
output_excel = os.path.join(ruta_base, "Sistema_lector_de_PDFs.xlsx")

# Entidades
def obtener_entidad(nombre):
    nombre = nombre.upper()
    if "CONTROLADORES" in nombre:
        return "CON"
    if "STAFE" in nombre: 
        return "STAFE"
    if "TEC" in nombre:
        return "TEC"
    return "OTROS"

# AGUINALDO 
def obtener_aguinaldo(nombre):
    nombre = nombre.upper()
    if "AGUINALDO" in nombre:
        return "AGUINALDO"  

#DETECTAR AGUINALDO
def detectar_tipo(tabla):
    texto = " ".join(
        " ".join(str(c) for c in fila if c)
        for fila in tabla
    ).upper()

    if "AGUINALD" in texto:
        return "AGUINALDO"
    return None

# Crear carpeta si no existe
if not os.path.exists(pdf_folder):
    os.makedirs(pdf_folder)
    print(f"CARPETA CREADA: {pdf_folder}")
    exit()

def extraer_fecha_del_pdf(page):
    try:
        texto_completo = page.extract_text()
        if texto_completo:
            patron_fecha_rango = r'(\d{2}-\w{3}-\d{4})\s*-\s*(\d{2}-\w{3}-\d{4})'
            match = re.search(patron_fecha_rango, texto_completo)
            if match:
                return f"{match.group(1)}_{match.group(2)}"
            
            patron_fecha = r'\d{2}-\w{3}-\d{4}'
            match = re.search(patron_fecha, texto_completo)
            if match:
                return match.group(0)
    except:
        pass
    return None

def extraer_numero_nomina(filename):
    match = re.search(r'NOMINA[_\s]*(\d+)', filename, re.IGNORECASE)
    if match:
        return match.group(1).zfill(2)
    match = re.search(r'^\d{7}', filename)
    return match.group(0) if match else "00"

def es_fila_encabezado(fila):
    if not fila:
        return False
    texto_fila = ' '.join(str(celda) for celda in fila if celda).upper()
    palabras_encabezado = [
        'CONCEPTOS', 'PERCEPCIONES', 'DEDUCCIONES', 'EXENTOS',
        'SEMANA', 'BASE GRAVABLE', 'IMPUESTO', 'NO. NOMINA'#, 'AGUINALDO'
    ]
    return any(palabra in texto_fila for palabra in palabras_encabezado)

def limpiar_encabezado(celda):
    if not celda or pd.isna(celda):
        return ""
    celda_str = str(celda).strip()
    match_semana = re.search(r'(Semana\s*\d+)', celda_str, re.IGNORECASE)
    if match_semana:
        return match_semana.group(1)
    
    patrones_fecha = [
        r'\d{2} -[A-Z]{3} -\d{4}\s* -\s*\d{2} -[A-Z]{3} -\d{4}',
        r'\d{1} -[A-Z]{3} -\d{4}\s* -\s*\d{2} -[A-Z]{3} -\d{4}',
        r'\d{2}-[A-Z]{3}-\d{4}\s*-\s*\d{2}-[A-Z]{3}-\d{4}',
        r'\d{1}-[A-Z]{3}-\d{4}\s*-\s*\d{2}-[A-Z]{3}-\d{4}',
        r'\d{2}/\d{2}/\d{4}\s*-\s*\d{2}/\d{2}/\d{4}',
        r'\d{2}-[A-Z]{3}-\d{4}',
        r'\d{1}-[A-Z]{3}-\d{4}',
        r'\d{1} -[A-Z]{3} -\d{4}',
        r'\d{2}/\d{2}/\d{4}',
    ]
    texto_limpio = celda_str
    for patron in patrones_fecha:
        texto_limpio = re.sub(patron, '', texto_limpio, flags=re.IGNORECASE)
    
    return texto_limpio.strip()

def limpiar_celda_datos(celda):
    if not celda or pd.isna(celda):
        return ""
    
    celda_str = str(celda).strip()
    
    # Borra si es un solo número
    if re.match(r'^\d+[.,]?\d*$', celda_str):
        if len(celda_str) == 1:
            return ""
        return celda_str
    
    if not re.search(r'\d{2}[-/]\w{3}[-/]\d{4}|\d{2}[-/]\d{2}[-/]\d{4}', celda_str):
        return celda_str
    # VARIANTE DE FECHAS
    patrones_fecha = [
        r'\d{2} -[A-Z]{3} -\d{4}\s* -\s*\d{2} -[A-Z]{3} -\d{4}',
        r'\d{2}-[A-Z]{3}-\d{4}\s*-\s*\d{2}-[A-Z]{3}-\d{4}',
        r'\d{1}-[A-Z]{3}-\d{4}\s*-\s*\d{2}-[A-Z]{3}-\d{4}',
        r'\d{2}/\d{2}/\d{4}\s*-\s*\d{2}/\d{2}/\d{4}',
        r'\d{2}-[A-Z]{3}-\d{4}',
        r'\d{1}-[A-Z]{3}-\d{4}',
        r'\d{2}/\d{2}/\d{4}',
        r'\d{1}/\d{2}/\d{4}',
    ]
    
    texto_limpio = celda_str
    for patron in patrones_fecha:
        texto_limpio = re.sub(patron, '', texto_limpio, flags=re.IGNORECASE)
    
    texto_limpio = re.sub(r'\s+', ' ', texto_limpio).strip()
    texto_limpio = re.sub(r'^\s*-\s*|\s*-\s*$', '', texto_limpio).strip()
    
    return texto_limpio

def normalizar_tabla(tabla):
    if not tabla or len(tabla) == 0:
        return []
    
    tabla_normalizada = []
    for fila in tabla:
        if not fila or all(not celda for celda in fila):
            continue
        
        es_encabezado = es_fila_encabezado(fila)
        fila_procesada = []
        for celda in fila:
            if not celda:
                fila_procesada.append('')
                continue
            
            celda_str = str(celda).strip()
            celda_limpia = limpiar_encabezado(celda_str) if es_encabezado else limpiar_celda_datos(celda_str)
            celda_limpia = celda_limpia.replace('\n', ' ')
            celda_limpia = re.sub(r'\s+', ' ', celda_limpia).strip()
            fila_procesada.append(celda_limpia)
        
        tabla_normalizada.append(fila_procesada)
    return tabla_normalizada

def aplicar_formato_automatico(worksheet, num_columnas):
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    filas_encabezado = []
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=50), 1):
        texto_fila = ' '.join(str(cell.value or '') for cell in row).upper()
        if any(palabra in texto_fila for palabra in ['CONCEPTOS', 'PERCEPCIONES', 'DEDUCCIONES', 'SEMANA']):#, 'AGUINALDO']):
            filas_encabezado.append(row_idx)
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1), 1):
        if row_idx in filas_encabezado:
            fill_color = "366092"
        else:
            fill_color = "E7E6E6" if row_idx % 2 == 0 else "FFFFFF"
        
        for cell in row:
            if row_idx in filas_encabezado:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            cell.border = border
            if cell.value and isinstance(cell.value, str):
                valor_limpio = cell.value.replace(',', '').replace('$', '').strip()
                if re.match(r'^-?\d+\.?\d*$', valor_limpio):
                    try:
                        if '.' in valor_limpio:
                            cell.value = float(valor_limpio)
                            cell.number_format = '#,##0.00'
                        else:
                            cell.value = int(valor_limpio)
                            cell.number_format = '#,##0'
                    except: pass

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 3, 40)

#NOMBRE DE HOJA
def generar_nombre_hoja_unico(filename, page_num, fecha, nombres_existentes):
    for char in "[]:*?/\\":
        filename = filename.replace(char, "")
    num_nomina = extraer_numero_nomina(filename)
    if fecha:
        fecha_corta = re.sub(r'-(\w{3})-\d{4}', r'\1', fecha).replace('_', '-')
        nombre_base = f"{num_nomina}_{fecha_corta}_{obtener_entidad(filename)}"
    else:
        nombre_base = num_nomina
    
    nombre_propuesto = nombre_base[:31]
    if nombre_propuesto not in nombres_existentes:
        return nombre_propuesto
    
    suffix = f"_P{page_num}"
    nombre_final = nombre_base[:31-len(suffix)] + suffix
    return nombre_final[:31]

table_settings = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines", 
    "snap_tolerance": 1,
    "join_tolerance": 3,
    "edge_min_length": 10,
    "intersection_tolerance": 3,
}

archivos_pdf = [f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")]
if not archivos_pdf:
    print("No se encontraron archivos PDF.")
    exit()

with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    contador_hojas = 0
    nombres_hojas_usados = set()

    for num_archivo, file in enumerate(archivos_pdf, 1):
        pdf_path = os.path.join(pdf_folder, file)
        print(f"[{num_archivo}/{len(archivos_pdf)}] Procesando: {file}")
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    fecha = extraer_fecha_del_pdf(page)
                    tables = page.extract_tables(table_settings=table_settings)
                    
                    if not tables: continue
                    
                    tablas_finales = []
                    for idx, tabla in enumerate(tables, 1):
                        tabla_norm = normalizar_tabla(tabla)
                        if not tabla_norm: continue
                        
                        df = pd.DataFrame(tabla_norm)
                        df = df.dropna(how='all').dropna(axis=1, how='all')
                        if df.empty: continue
                        
                        df = df.map(lambda x: "".join(ch for ch in str(x) if ch.isprintable()) if x else "")
                        df.columns = [f'Col{i+1}' for i in range(len(df.columns))]
                        
                        if len(tables) > 1:
                            sep = pd.DataFrame([[f'=== TABLA {idx} ==='] + ['']*(len(df.columns)-1)], columns=df.columns)
                            tablas_finales.extend([sep, df, pd.DataFrame([['']*len(df.columns)], columns=df.columns)])
                        else:
                            tablas_finales.append(df)
                    
                    if tablas_finales:
                        df_final = pd.concat(tablas_finales, ignore_index=True).fillna('')
                        nombre_hoja = generar_nombre_hoja_unico(file, page_num, fecha, nombres_hojas_usados)
                        nombres_hojas_usados.add(nombre_hoja)
                        df_final.to_excel(writer, sheet_name=nombre_hoja, index=False)
                        contador_hojas += 1
        except Exception as e:
            print(f"  ERROR en {file}: {e}")

    if contador_hojas > 0:
        wb = writer.book
        for sheet_name in wb.sheetnames:
            aplicar_formato_automatico(wb[sheet_name], wb[sheet_name].max_column)

print(f"\nPROCESO COMPLETADO. Total hojas: {contador_hojas}")