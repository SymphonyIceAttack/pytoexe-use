import argparse
import base64
import csv
import getpass
import imaplib
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from email import policy
from email.parser import BytesParser
from email.utils import parsedate_to_datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple
from urllib.parse import parse_qs, unquote, urlencode, urljoin, urlparse

import requests


IMAP_HOST = "imap.exmail.qq.com"
IMAP_PORT = 993
MAILBOX_DEFAULT = "INBOX"
MAILBOX_RE_WANGBEI = "&UXZO1mWHTvZZOQ-/RE-&c4uNHQ-"

LINK_PREFIX = "https://edms.gulf.co.th/dms/login.asp"
SINCE_DEFAULT = ""

STATE_FILE_DEFAULT = "pipeline_state.json"
OUTPUT_DIR_DEFAULT = "pipeline_output"
CREDENTIALS_FILE = "credentials.json"

# Hardcoded credentials as requested.
IMAP_USER_FIXED = "y1.zhang@sepec.com.cn"
IMAP_PASSWORD_FIXED = "Zy09190011#"
EDMS_USER_FIXED = "BEI.W"
EDMS_PASSWORD_FIXED = "Wang-8232405"


def log(msg: str) -> None:
    print(msg, flush=True)


@dataclass
class MessageItem:
    mailbox: str
    uid: str
    subject: str
    sender: str
    date: str
    html: str


class LinkAndTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: List[str] = []
        self.tables: List[List[List[str]]] = []
        self._in_table = 0
        self._in_row = False
        self._in_cell = False
        self._cell_buf: List[str] = []
        self._row: List[str] = []
        self._table: List[List[str]] = []

    def handle_starttag(self, tag: str, attrs: Sequence[Tuple[str, str]]) -> None:
        t = tag.lower()
        attrs_dict = dict(attrs)
        if t == "a":
            href = (attrs_dict.get("href") or "").strip()
            if href:
                self.links.append(href)
        elif t == "table":
            self._in_table += 1
            if self._in_table == 1:
                self._table = []
        elif self._in_table and t == "tr":
            self._in_row = True
            self._row = []
        elif self._in_table and t in ("td", "th"):
            self._in_cell = True
            self._cell_buf = []
        elif self._in_table and self._in_cell and t == "br":
            self._cell_buf.append("\n")

    def handle_endtag(self, tag: str) -> None:
        t = tag.lower()
        if self._in_table and self._in_cell and t in ("td", "th"):
            text = "".join(self._cell_buf)
            text = re.sub(r"[\t\r\n]+", " ", text)
            text = re.sub(r"\s{2,}", " ", text).strip()
            self._row.append(text)
            self._in_cell = False
            self._cell_buf = []
        elif self._in_table and self._in_row and t == "tr":
            if any(cell for cell in self._row):
                self._table.append(self._row)
            self._in_row = False
            self._row = []
        elif t == "table" and self._in_table:
            self._in_table -= 1
            if self._in_table == 0 and self._table:
                self.tables.append(self._table)
                self._table = []

    def handle_data(self, data: str) -> None:
        if self._in_table and self._in_cell:
            self._cell_buf.append(data)


def get_html_body(msg) -> str:
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                return part.get_content()
    elif msg.get_content_type() == "text/html":
        return msg.get_content()
    return ""


def decode_modutf7(s: str) -> str:
    out: List[str] = []
    i = 0
    while i < len(s):
        if s[i] == "&":
            j = s.find("-", i)
            if j == -1:
                out.append("&")
                i += 1
                continue
            tok = s[i + 1 : j]
            if tok == "":
                out.append("&")
            else:
                b64 = tok.replace(",", "/")
                pad = "=" * ((4 - len(b64) % 4) % 4)
                try:
                    out.append(base64.b64decode(b64 + pad).decode("utf-16-be", errors="replace"))
                except Exception:
                    out.append("&" + tok + "-")
            i = j + 1
        else:
            out.append(s[i])
            i += 1
    return "".join(out)


def parse_since_time(since_text: str) -> datetime:
    return datetime.strptime(since_text, "%Y-%m-%d %H:%M:%S")


def message_after_since(msg_date: str, since_dt: datetime) -> bool:
    if not msg_date:
        return False
    try:
        dt = parsedate_to_datetime(msg_date)
        # Convert to local naive for comparison with user input naive timestamp.
        if dt.tzinfo is not None:
            dt = dt.astimezone().replace(tzinfo=None)
        return dt >= since_dt
    except Exception:
        return False


def message_matches(msg, since_dt: Optional[datetime]) -> bool:
    if since_dt is not None:
        if not message_after_since(msg.get("Date", ""), since_dt):
            return False
    return True


def load_state(path: Path) -> Dict[str, List[str]]:
    if not path.exists():
        return {"processed_keys": []}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {"processed_keys": []}
        keys = data.get("processed_keys")
        if not isinstance(keys, list):
            # backward compatible
            keys = data.get("processed_uids", [])
        if not isinstance(keys, list):
            keys = []
        return {"processed_keys": [str(x) for x in keys]}
    except Exception:
        return {"processed_keys": []}


def save_state(path: Path, processed_keys: Set[str]) -> None:
    payload = {"processed_keys": sorted(processed_keys)}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def list_mailboxes(imap: imaplib.IMAP4_SSL) -> List[str]:
    typ, boxes = imap.list()
    if typ != "OK" or not boxes:
        return ["INBOX"]
    names: List[str] = []
    for b in boxes:
        line = b.decode("utf-8", errors="ignore")
        m = re.match(r'.* "([^"]+)"$', line)
        if not m:
            continue
        name = m.group(1)
        # Skip non-selectable parent.
        if "\\NoSelect" in line:
            continue
        names.append(name)
    return names or ["INBOX"]


def fetch_candidates(
    imap_user: str,
    imap_password: str,
    mailboxes: List[str],
    processed_keys: Set[str],
    full_scan: bool,
    since_dt: Optional[datetime],
    max_emails: int,
) -> List[MessageItem]:
    items: List[MessageItem] = []
    with imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT) as imap:
        log("[INFO] 正在登录 IMAP ...")
        imap.login(imap_user, imap_password)
        for mailbox in mailboxes:
            log(f"[INFO] 扫描文件夹: {mailbox}")
            if imap.select(f"\"{mailbox}\"")[0] != "OK":
                log(f"[WARN] 无法选择文件夹: {mailbox}")
                continue
            if since_dt is not None:
                since_str = since_dt.strftime("%d-%b-%Y")
                typ, data = imap.uid("SEARCH", None, "SINCE", since_str)
            else:
                typ, data = imap.uid("SEARCH", None, "ALL")
            if typ != "OK" or not data or not data[0]:
                log(f"[INFO] 文件夹无邮件: {mailbox}")
                continue

            all_uids = [x.decode("utf-8", errors="ignore") for x in data[0].split()]
            log(f"[INFO] 文件夹候选邮件数: {len(all_uids)}")
            checked = 0
            for uid in all_uids:
                if max_emails > 0 and len(items) >= max_emails:
                    log(f"[INFO] 达到 max-emails={max_emails}，停止扫描")
                    return items
                key = f"{mailbox}|{uid}"
                if not full_scan and key in processed_keys:
                    continue
                typ, msg_data = imap.uid("FETCH", uid, "(RFC822)")
                if typ != "OK" or not msg_data or not msg_data[0]:
                    continue
                checked += 1
                if checked % 20 == 0:
                    log(f"[INFO] 已检查 {checked}/{len(all_uids)}，当前命中 {len(items)}")
                raw = msg_data[0][1]
                msg = BytesParser(policy=policy.default).parsebytes(raw)
                html = get_html_body(msg)
                if not html:
                    continue
                if not message_matches(msg, since_dt=since_dt):
                    continue
                items.append(
                    MessageItem(
                        mailbox=mailbox,
                        uid=uid,
                        subject=msg.get("Subject", ""),
                        sender=msg.get("From", ""),
                        date=msg.get("Date", ""),
                        html=html,
                    )
                )
    return items


def find_document_table(tables: List[List[List[str]]]) -> List[List[str]]:
    required = ["document", "rev", "title", "date"]
    for table in tables:
        for row in table:
            joined = " ".join(cell.lower() for cell in row if cell)
            if all(k in joined for k in required) and "rev" in joined:
                return table
    return []


def find_header_index(table: List[List[str]]) -> int:
    for i, row in enumerate(table):
        joined = " ".join(c.lower() for c in row if c)
        if "document" in joined and "rev" in joined and "date" in joined:
            return i
    return -1


def build_column_map(header_row: List[str]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for i, col in enumerate(header_row):
        key = col.strip().lower()
        if "document" in key:
            m["document"] = i
        elif key == "rev" or "rev" in key:
            m["rev"] = i
        elif "title/description" in key or key == "title" or "title" in key:
            m["title"] = i
        elif key == "code" or key == "return" or "code" in key:
            m["code"] = i
        elif "date/ref" in key or key == "date" or "date" in key:
            m["date"] = i
    return m


def extract_doc_rows(table: List[List[str]]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    hidx = find_header_index(table)
    if hidx < 0:
        return rows
    cmap = build_column_map(table[hidx])
    required = ("document", "rev", "title", "code", "date")
    if not all(k in cmap for k in required):
        return rows

    for row in table[hidx + 1 :]:
        first = row[0].strip() if row else ""
        if not first.isdigit():
            continue
        rows.append(
            {
                "Document": row[cmap["document"]].strip() if cmap["document"] < len(row) else "",
                "Rev": row[cmap["rev"]].strip() if cmap["rev"] < len(row) else "",
                "Title/Description": row[cmap["title"]].strip() if cmap["title"] < len(row) else "",
                "Code": row[cmap["code"]].strip() if cmap["code"] < len(row) else "",
                "Date/Ref": row[cmap["date"]].strip() if cmap["date"] < len(row) else "",
            }
        )
    return rows


def extract_login_links(html: str, anchor_links: List[str]) -> List[str]:
    links = set()
    for link in anchor_links:
        if link.startswith(LINK_PREFIX):
            links.add(link)
    pattern = re.compile(r"https://edms\.gulf\.co\.th/dms/login\.asp[^\s\"'<>]*", re.I)
    for m in pattern.findall(html):
        if m.startswith(LINK_PREFIX):
            links.add(m)
    return sorted(links)


def category_from_document(doc_code: str) -> str:
    parts = doc_code.split("-")
    if len(parts) >= 3:
        return parts[2].strip()
    return ""


def extract_title_from_subject(subject: str) -> str:
    text = (subject or "").strip()
    if not text:
        return ""

    # 1) Prefer multiline tail (many subjects put the real title on the last line).
    lines = [ln.strip() for ln in re.split(r"[\r\n]+", text) if ln and ln.strip()]
    for ln in reversed(lines):
        if re.search(r"\b(return|re|fw|fwd)\s*:", ln, re.I):
            continue
        if "[id:" in ln.lower():
            continue
        return re.sub(r"\s{2,}", " ", ln).strip()

    # 2) Folded/one-line subject: remove noisy prefixes and keep text after last doc code.
    folded = re.sub(r"\s+", " ", text).strip()
    folded = re.sub(r"^\s*(re|fw|fwd|return)\s*:\s*", "", folded, flags=re.I).strip()
    folded = re.sub(r"\[ID:\s*\d+\]", "", folded, flags=re.I).strip()
    folded = re.sub(r"\b[A-Z]:[A-Z ]+\b", "", folded, flags=re.I).strip()

    doc_code_re = re.compile(
        r"\b[A-Z]{2,5}-\d{3}-[A-Z](?:-[A-Z0-9]{2,6}){2,7}-\d+\b",
        re.I,
    )
    matches = list(doc_code_re.finditer(folded))
    if matches:
        tail = folded[matches[-1].end() :].strip(" -_:;")
        if tail:
            return re.sub(r"\s{2,}", " ", tail).strip()

    return re.sub(r"\s{2,}", " ", folded).strip()


def resolve_title_description(table_title: str, email_subject: str) -> str:
    subject_title = extract_title_from_subject(email_subject)
    if subject_title:
        return subject_title
    return (table_title or "").strip()


def extract_rows_from_message(item: MessageItem) -> Tuple[List[Dict[str, str]], List[str]]:
    parser = LinkAndTableParser()
    parser.feed(item.html)
    table = find_document_table(parser.tables)
    doc_rows = extract_doc_rows(table)
    links = extract_login_links(item.html, parser.links)

    out_rows: List[Dict[str, str]] = []
    link_joined = ";".join(links)
    for doc in doc_rows:
        code = doc["Document"]
        resolved_title = resolve_title_description(
            table_title=doc["Title/Description"],
            email_subject=item.subject,
        )
        out_rows.append(
            {
                "类别": category_from_document(code),
                "编码": code,
                "版本": doc["Rev"],
                "文件名": resolved_title,
                "状态": doc["Code"],
                "链接": link_joined,
                "邮件主题": item.subject,
                "邮件发件人": item.sender,
                "邮件日期": item.date,
                "邮件UID": item.uid,
                "邮箱文件夹": item.mailbox,
            }
        )
    return out_rows, links


def write_excel_or_csv(rows: List[Dict[str, str]], out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    headers = ["类别", "编码", "版本", "文件名", "状态", "链接", "邮件主题", "邮件发件人", "邮件日期", "邮件UID", "邮箱文件夹"]
    day = datetime.now().strftime("%Y%m%d")
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore

        xlsx_path = out_dir / f"OE意见清单_{day}.xlsx"
        if xlsx_path.exists():
            wb = load_workbook(xlsx_path)
            ws = wb.active
            if ws.max_row < 1:
                ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "data"
            ws.append(headers)

        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        wb.save(xlsx_path)
        return xlsx_path
    except Exception:
        csv_path = out_dir / f"OE意见清单_{day}.csv"
        mode = "a" if csv_path.exists() else "w"
        with csv_path.open(mode, newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            if mode == "w":
                writer.writeheader()
            writer.writerows(rows)
        return csv_path


def extract_docid(login_url: str) -> str:
    q = parse_qs(urlparse(login_url).query)
    return (q.get("docid") or [""])[0]


def encode_password(key_digits: str, password: str) -> str:
    localkey = key_digits * 32
    parts = []
    for i, ch in enumerate(password):
        kk = localkey[i * 2 : i * 2 + 2]
        if len(kk) < 2:
            raise ValueError("login key length invalid")
        parts.append(str(int(kk) + ord(ch)))
    return ":" + ":".join(parts)


def get_login_key(session: requests.Session, base_url: str, username: str) -> str:
    params = {"gkey": "1", "login": username, "myrnd": "0.123456"}
    resp = session.get(urljoin(base_url, "login.asp"), params=params, timeout=30)
    resp.raise_for_status()
    digits = re.sub(r"\D", "", resp.text)
    if not digits:
        raise RuntimeError("failed to fetch login key")
    return digits


def login_and_open_viewdoc(
    session: requests.Session,
    login_url: str,
    username: str,
    password: str,
) -> str:
    parsed = urlparse(login_url)
    base_url = f"{parsed.scheme}://{parsed.netloc}/dms/"
    docid = extract_docid(login_url)

    init = session.get(login_url, timeout=30)
    init.raise_for_status()

    key_digits = get_login_key(session, base_url, username)
    pwden = encode_password(key_digits, password)

    payload = {
        "login": username,
        "password": "",
        "allowapi": "Y",
        "loginmode": "local",
        "postback": "1",
        "pwden": pwden,
    }
    post_url = urljoin(base_url, "login.asp")
    if docid:
        post_url = f"{post_url}?{urlencode({'docid': docid})}"
    resp = session.post(post_url, data=payload, timeout=30, allow_redirects=True)
    resp.raise_for_status()

    if "logout" not in resp.text.lower() and "viewdoc.asp" not in resp.url.lower():
        raise RuntimeError("EDMS login failed")

    if docid:
        return urljoin(base_url, f"viewdoc.asp?docid={docid}")
    return urljoin(base_url, "viewdoc.asp")


def is_login_page_response(resp: requests.Response) -> bool:
    url_l = (resp.url or "").lower()
    text_l = (resp.text or "").lower()
    if "login.asp" in url_l and "viewdoc.asp" not in url_l:
        return True
    return ("name=\"login\"" in text_l) or ("loginmode" in text_l and "pwden" in text_l)


def get_viewdoc_url_from_login_link(login_url: str) -> str:
    parsed = urlparse(login_url)
    base_url = f"{parsed.scheme}://{parsed.netloc}/dms/"
    docid = extract_docid(login_url)
    if docid:
        return urljoin(base_url, f"viewdoc.asp?docid={docid}")
    return urljoin(base_url, "viewdoc.asp")


def open_viewdoc_with_relogin(
    session: requests.Session,
    login_link: str,
    edms_user: str,
    edms_password: str,
) -> Tuple[str, requests.Response]:
    # Reuse existing session first; relogin only when session is invalid.
    viewdoc_url = get_viewdoc_url_from_login_link(login_link)
    page = session.get(viewdoc_url, timeout=30, allow_redirects=True)
    if page.status_code >= 400 or is_login_page_response(page):
        log(f"[INFO] 会话未登录或已失效，重新登录: {login_link}")
        viewdoc_url = login_and_open_viewdoc(session, login_link, edms_user, edms_password)
        page = session.get(viewdoc_url, timeout=30, allow_redirects=True)
        page.raise_for_status()
        if is_login_page_response(page):
            raise RuntimeError("EDMS relogin succeeded but still not authorized")
    else:
        page.raise_for_status()
    return viewdoc_url, page


def parse_getfile_links(html: str, page_url: str) -> List[Tuple[str, str]]:
    links: List[Tuple[str, str]] = []
    seen: Set[str] = set()
    pattern = re.compile(
        r"<a[^>]*href\s*=\s*['\"]?([^'\" >]+)[^>]*>(.*?)</a>",
        flags=re.IGNORECASE | re.DOTALL,
    )
    for m in pattern.finditer(html):
        href = m.group(1).strip()
        txt = re.sub(r"<[^>]+>", "", m.group(2))
        txt = re.sub(r"\s+", " ", txt).strip()
        abs_url = urljoin(page_url, href)
        if urlparse(abs_url).path.lower().endswith("/getfile.asp"):
            if abs_url not in seen:
                seen.add(abs_url)
                links.append((abs_url, txt))
    return links


def filename_from_content_disposition(header_value: str) -> str:
    if not header_value:
        return ""
    m = re.search(r"filename\*?=(?:UTF-8''|\"?)([^\";]+)", header_value, re.I)
    if not m:
        return ""
    return unquote(m.group(1)).strip()


def guess_name_from_url(url: str) -> str:
    q = parse_qs(urlparse(url).query)
    name = (q.get("filename") or [""])[0]
    return unquote(name).strip() if name else ""


def guess_name_from_text(text: str) -> str:
    if not text:
        return ""
    m = re.search(r"[A-Za-z0-9._-]+\.[A-Za-z0-9]{1,8}", text)
    return m.group(0) if m else ""


def sanitize_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]+', "_", name).strip()


def sanitize_dirname(name: str) -> str:
    s = re.sub(r'[<>:"/\\|?*]+', "_", name).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:120] if len(s) > 120 else s


def dedupe_path(target_dir: Path, name: str) -> Path:
    p = target_dir / name
    if not p.exists():
        return p
    stem, suffix = p.stem, p.suffix
    i = 2
    while True:
        p2 = target_dir / f"{stem}_{i}{suffix}"
        if not p2.exists():
            return p2
        i += 1


def should_keep_attachment(name: str) -> bool:
    return "CONSOL" in (name or "").upper()


def stream_download_with_resume(
    session: requests.Session,
    file_url: str,
    final_path: Path,
    max_retries: int = 3,
) -> Tuple[bool, int, str]:
    part_path = final_path.with_suffix(final_path.suffix + ".part")
    if final_path.exists() and final_path.stat().st_size > 0:
        return True, 0, "exists"

    last_error = ""
    for attempt in range(1, max_retries + 1):
        try:
            resume_from = part_path.stat().st_size if part_path.exists() else 0
            headers = {}
            if resume_from > 0:
                headers["Range"] = f"bytes={resume_from}-"
            resp = session.get(file_url, timeout=120, stream=True, headers=headers)
            if resp.status_code not in (200, 206):
                last_error = f"HTTP {resp.status_code}"
                continue

            if resp.status_code == 200 and resume_from > 0:
                # Server ignored range; restart download.
                part_path.unlink(missing_ok=True)
                resume_from = 0

            mode = "ab" if (resp.status_code == 206 and resume_from > 0) else "wb"
            with part_path.open(mode) as f:
                for chunk in resp.iter_content(chunk_size=1024 * 64):
                    if chunk:
                        f.write(chunk)

            # Basic size sanity check when content-length exists.
            content_len = resp.headers.get("Content-Length")
            if content_len and content_len.isdigit():
                expected = resume_from + int(content_len) if resp.status_code == 206 else int(content_len)
                actual = part_path.stat().st_size
                if actual < expected:
                    last_error = f"incomplete {actual}/{expected}"
                    continue

            part_path.replace(final_path)
            return True, attempt, ""
        except Exception as ex:
            last_error = str(ex)
    return False, max_retries, last_error


def download_attachments_for_login_link(
    session: requests.Session,
    login_link: str,
    edms_user: str,
    edms_password: str,
    run_download_dir: Path,
    email_uid: str,
    email_subject: str,
    email_mailbox: str,
) -> List[Dict[str, str]]:
    log(f"[INFO] 打开文档: {login_link}")
    viewdoc_url, page = open_viewdoc_with_relogin(
        session=session,
        login_link=login_link,
        edms_user=edms_user,
        edms_password=edms_password,
    )
    file_links = parse_getfile_links(page.text, viewdoc_url)
    log(f"[INFO] docid={extract_docid(viewdoc_url)} 附件链接数: {len(file_links)}")

    docid = extract_docid(viewdoc_url) or "unknown_docid"
    target_dir = run_download_dir

    saved_items: List[Dict[str, str]] = []
    for file_url, text_hint in file_links:
        # First pass to resolve real filename without downloading body.
        head_resp = session.get(file_url, timeout=60, stream=True)
        if head_resp.status_code not in (200, 206):
            saved_items.append(
                {
                    "邮件UID": email_uid,
                    "邮箱文件夹": email_mailbox,
                    "邮件主题": email_subject,
                    "docid": docid,
                    "登录链接": login_link,
                    "查看链接": viewdoc_url,
                    "下载链接": file_url,
                    "附件名": "",
                    "保存路径": "",
                    "下载状态": "failed",
                    "重试次数": 0,
                    "错误": f"HTTP {head_resp.status_code}",
                }
            )
            continue
        real_name = (
            filename_from_content_disposition(head_resp.headers.get("Content-Disposition", ""))
            or guess_name_from_url(file_url)
            or guess_name_from_text(text_hint)
        )
        head_resp.close()
        if not real_name:
            saved_items.append(
                {
                    "邮件UID": email_uid,
                    "邮箱文件夹": email_mailbox,
                    "邮件主题": email_subject,
                    "docid": docid,
                    "登录链接": login_link,
                    "查看链接": viewdoc_url,
                    "下载链接": file_url,
                    "附件名": "",
                    "保存路径": "",
                    "下载状态": "failed",
                    "重试次数": 0,
                    "错误": "missing filename",
                }
            )
            continue
        safe_name = sanitize_filename(real_name)
        if (not safe_name) or (not should_keep_attachment(safe_name)):
            continue
        save_path = dedupe_path(target_dir, safe_name)
        log(f"[INFO] 下载: {safe_name}")
        ok, attempts, err = stream_download_with_resume(session, file_url, save_path, max_retries=3)
        saved_items.append(
            {
                "邮件UID": email_uid,
                "邮箱文件夹": email_mailbox,
                "邮件主题": email_subject,
                "docid": docid,
                "登录链接": login_link,
                "查看链接": viewdoc_url,
                "下载链接": file_url,
                "附件名": safe_name,
                "保存路径": str(save_path.resolve()) if ok else "",
                "下载状态": "success" if ok else "failed",
                "重试次数": attempts,
                "错误": err,
            }
        )
    return saved_items


def write_attachment_mapping(items: List[Dict[str, str]], out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "邮件附件映射.csv"
    headers = [
        "邮件UID",
        "邮箱文件夹",
        "邮件主题",
        "docid",
        "登录链接",
        "查看链接",
        "下载链接",
        "附件名",
        "保存路径",
        "下载状态",
        "重试次数",
        "错误",
    ]
    mode = "a" if path.exists() else "w"
    with path.open(mode, newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        if mode == "w":
            writer.writeheader()
        writer.writerows(items)
    return path


def create_run_dir(base_dir: Path) -> Path:
    base_dir.mkdir(parents=True, exist_ok=True)
    day_name = f"OE意见_{datetime.now().strftime('%Y%m%d')}"
    run_dir = base_dir / day_name
    if not run_dir.exists():
        run_dir.mkdir(parents=True, exist_ok=True)
        return run_dir
    i = 2
    while True:
        p = base_dir / f"{day_name}_{i}"
        if not p.exists():
            p.mkdir(parents=True, exist_ok=True)
            return p
        i += 1


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def load_credentials(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {}
        return {
            "imap_user": str(data.get("imap_user", "")),
            "imap_password": str(data.get("imap_password", "")),
            "edms_user": str(data.get("edms_user", "")),
            "edms_password": str(data.get("edms_password", "")),
        }
    except Exception:
        return {}


def save_credentials(path: Path, values: Dict[str, str]) -> None:
    payload = {
        "imap_user": values.get("imap_user", ""),
        "imap_password": values.get("imap_password", ""),
        "edms_user": values.get("edms_user", ""),
        "edms_password": values.get("edms_password", ""),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="IMAP筛选邮件 + 提取表格到Excel + 自动登录EDMS下载附件；首跑全量，后续增量。"
    )
    parser.add_argument("--imap-user", default=IMAP_USER_FIXED)
    parser.add_argument("--imap-password", default=IMAP_PASSWORD_FIXED)
    parser.add_argument("--edms-user", default=EDMS_USER_FIXED)
    parser.add_argument("--edms-password", default=EDMS_PASSWORD_FIXED)
    parser.add_argument("--mailbox", default=MAILBOX_RE_WANGBEI, help="默认 RE-王贝 文件夹")
    parser.add_argument("--all-mailboxes", action="store_true", help="扫描所有可选文件夹（推荐）")
    parser.add_argument("--single-mailbox", dest="all_mailboxes", action="store_false", help="只扫描 --mailbox 指定文件夹")
    parser.set_defaults(all_mailboxes=False)
    parser.add_argument("--since", default=SINCE_DEFAULT, help="处理开始时间，例如 2026-02-16 00:00:00；留空表示不限时间")
    parser.add_argument("--no-download", action="store_true", help="只提取邮件到结果文件，不下载附件")
    parser.add_argument("--output-dir", default=OUTPUT_DIR_DEFAULT)
    parser.add_argument("--state-file", default=STATE_FILE_DEFAULT)
    parser.add_argument("--max-emails", type=int, default=0, help="最多处理多少封邮件，0 表示不限制")
    parser.add_argument("--interactive", action="store_true", help="交互输入参数模式")
    parser.add_argument("--remember-creds", action="store_true", help="记住账号密码到项目目录 credentials.json")
    parser.add_argument("--forget-creds", action="store_true", help="清除已保存账号密码")
    parser.add_argument(
        "--mode",
        choices=["auto", "full", "incremental"],
        default="auto",
        help="auto: 首次无状态文件全量，后续增量",
    )
    args = parser.parse_args()

    cred_path = app_dir() / CREDENTIALS_FILE
    if args.forget_creds and cred_path.exists():
        cred_path.unlink(missing_ok=True)

    auto_interactive = len(sys.argv) == 1
    if args.interactive or auto_interactive:
        log("时间参数格式：YYYY-MM-DD HH:MM:SS（例如 2026-02-16 00:00:00）")
        since_in = input("请输入开始时间（留空表示不限时间）: ").strip()
        if since_in:
            args.since = since_in

    missing = []
    if not args.imap_user:
        missing.append("--imap-user")
    if not args.imap_password:
        missing.append("--imap-password")
    if not args.edms_user:
        missing.append("--edms-user")
    if not args.edms_password:
        missing.append("--edms-password")
    if missing:
        raise ValueError(
            "缺少必填参数: "
            + ", ".join(missing)
            + "。请先在项目目录准备 credentials.json，或命令行传参并加 --remember-creds 保存。"
        )

    if args.remember_creds:
        save_credentials(
            cred_path,
            {
                "imap_user": args.imap_user,
                "imap_password": args.imap_password,
                "edms_user": args.edms_user,
                "edms_password": args.edms_password,
            },
        )

    base_output_dir = Path(args.output_dir)
    output_dir = create_run_dir(base_output_dir)
    run_download_dir = output_dir / f"OE意见_{datetime.now().strftime('%Y%m%d')}"
    run_download_dir.mkdir(parents=True, exist_ok=True)
    state_path = Path(args.state_file)
    state = load_state(state_path)
    processed_keys = set(state.get("processed_keys", []))
    since_dt = parse_since_time(args.since) if args.since.strip() else None
    log(f"[INFO] 时间参数 since: {args.since or '不限'}")

    if args.mode == "full":
        full_scan = True
    elif args.mode == "incremental":
        full_scan = False
    else:
        full_scan = not state_path.exists()

    with imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT) as imap_probe:
        log("[INFO] 正在读取文件夹列表 ...")
        imap_probe.login(args.imap_user, args.imap_password)
        if args.all_mailboxes:
            mailboxes = list_mailboxes(imap_probe)
        else:
            target = args.mailbox
            if target == "RE-王贝":
                target = MAILBOX_RE_WANGBEI
            # Best effort: resolve by decoded display name.
            all_names = list_mailboxes(imap_probe)
            resolved = target
            for n in all_names:
                if n == target:
                    resolved = n
                    break
                decoded = decode_modutf7(n)
                if decoded == target or decoded.endswith("/" + target):
                    resolved = n
                    break
            mailboxes = [resolved]
    log(f"[INFO] 本次扫描文件夹数: {len(mailboxes)}")

    candidates = fetch_candidates(
        imap_user=args.imap_user,
        imap_password=args.imap_password,
        mailboxes=mailboxes,
        processed_keys=processed_keys,
        full_scan=full_scan,
        since_dt=since_dt,
        max_emails=args.max_emails,
    )

    all_rows: List[Dict[str, str]] = []
    success_keys: Set[str] = set()
    total_downloaded = 0
    attachment_items: List[Dict[str, str]] = []
    with requests.Session() as edms_session:
        for item in candidates:
            try:
                log(f"[INFO] 处理邮件 {item.mailbox}|{item.uid}")
                rows, links = extract_rows_from_message(item)
                all_rows.extend(rows)
                if not args.no_download:
                    for link in links:
                        saved = download_attachments_for_login_link(
                            session=edms_session,
                            login_link=link,
                            edms_user=args.edms_user,
                            edms_password=args.edms_password,
                            run_download_dir=run_download_dir,
                            email_uid=item.uid,
                            email_subject=item.subject,
                            email_mailbox=item.mailbox,
                        )
                        attachment_items.extend(saved)
                        total_downloaded += len([x for x in saved if x.get("下载状态") == "success"])
                success_keys.add(f"{item.mailbox}|{item.uid}")
            except Exception as ex:
                log(f"[WARN] {item.mailbox}|{item.uid} failed: {ex}")

    result_file = write_excel_or_csv(all_rows, output_dir)
    mapping_file = write_attachment_mapping(attachment_items, output_dir)
    processed_keys.update(success_keys)
    save_state(state_path, processed_keys)

    log(f"mode: {'full' if full_scan else 'incremental'}")
    log(f"mailboxes_scanned: {len(mailboxes)}")
    log(f"matched_emails: {len(candidates)}")
    log(f"extracted_rows: {len(all_rows)}")
    log(f"downloaded_files: {total_downloaded}")
    log(f"run_dir: {output_dir.resolve()}")
    log(f"download_dir: {run_download_dir.resolve()}")
    log(f"result_file: {result_file.resolve()}")
    log(f"attachment_mapping: {mapping_file.resolve()}")
    log(f"state_file: {state_path.resolve()}")


if __name__ == "__main__":
    try:
        main()
    except Exception as ex:
        log(f"[ERROR] {ex}")
        if len(sys.argv) == 1:
            input("按回车退出...")
        else:
            raise
