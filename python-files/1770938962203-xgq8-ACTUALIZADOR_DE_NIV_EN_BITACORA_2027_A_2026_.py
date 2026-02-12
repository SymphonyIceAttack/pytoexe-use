import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import os
import time
import random
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import logging

# Configuración de logs
OUTPUT_DIR = r"C:\DESARROLLO\ACTUALIZAR_NO_INSCRITOS_EN_BITACORA_OK\NO_INSCRITOS_LOG"
os.makedirs(OUTPUT_DIR, exist_ok=True)
LOG_FILE = os.path.join(OUTPUT_DIR, "actualizacion.log")

logging.basicConfig(
    filename='debug_gui.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def log_error(vin, msg):
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"{vin} => {msg}\n")


# =====================================================
# LOGIN (TU LÓGICA ORIGINAL)
# =====================================================
def login(page, url, user, pwd):
    page.goto(url)
    page.wait_for_selector("#email", timeout=10000)
    page.fill("#email", user)
    page.fill("#password", pwd)
    time.sleep(1)
    page.press("#password", "Enter")
    page.wait_for_load_state("domcontentloaded")

    page.wait_for_selector("a[href='#/mostrarNoInscritos']", timeout=15000)
    page.reload()
    page.wait_for_timeout(2000)

    page.click("a[href='#/mostrarNoInscritos']")
    page.wait_for_url("**/mostrarNoInscritos", timeout=15000)

    page.wait_for_selector(
        "text=Procesando información de vehiculos no inscritos",
        state="detached",
        timeout=30000
    )

    page.wait_for_selector("button:has(svg.MuiSvgIcon-root)", timeout=15000)
    page.click("button:has(svg.MuiSvgIcon-root)")

    page.wait_for_selector("input[aria-label='Search']", timeout=10000)


# =====================================================
# ACTUALIZAR VIN (con diálogo simplificado)
# =====================================================
def actualizar(page, vin):
    try:
        page.fill("input[aria-label='Search']", vin)
        page.keyboard.press("Enter")
        page.wait_for_timeout(3000)

        coincidencias = page.locator("button:has-text('Editar')")
        cantidad = coincidencias.count()

        if cantidad == 0:
            log_error(vin, "Sin coincidencias")
            return "SIN COINCIDENCIA"

        for idx in range(cantidad):
            try:
                coincidencias.nth(idx).click()
                page.wait_for_selector("button:has-text('Guardar')", timeout=10000)

                # ===== CAMBIAR AÑO =====
                campo_anio = page.locator("input[name='anio_model']")
                if campo_anio.count() > 0:
                    campo_anio = campo_anio.first
                    campo_anio.scroll_into_view_if_needed()
                    page.wait_for_timeout(500)

                    campo_anio.click()
                    campo_anio.press("Control+A")
                    campo_anio.press("Backspace")
                    campo_anio.type("2026", delay=100)
                    campo_anio.press("Tab")

                    page.wait_for_timeout(800)

                # ==============================
                # GUARDAR (VERSIÓN SIMPLIFICADA)
                # ==============================

                # Manejo simplificado de diálogo
                def aceptar_dialogo(dialog):
                    print(f"📢 Alerta recibida: {dialog.message}")
                    log_error(vin, f"Dialogo: {dialog.message}")
                    time.sleep(1)  # 🔥 Solo 1 segundo visible
                    dialog.accept()

                page.once("dialog", aceptar_dialogo)

                guardar_btn = page.locator("button:has-text('Guardar')")
                guardar_btn.scroll_into_view_if_needed()
                guardar_btn.click(force=True)

                # Solo pequeña pausa para estabilidad
                page.wait_for_timeout(800)

            except Exception as ed:
                log_error(vin, f"Error al editar coincidencia {idx + 1}: {str(ed)[:100]}")

        return f"ACTUALIZADO ({cantidad})"

    except Exception as e:
        log_error(vin, str(e)[:100])
        captura = os.path.join(OUTPUT_DIR, f"ERROR_{vin}.png")
        try:
            page.screenshot(path=captura)
        except:
            pass
        return "ERROR"


# =====================================================
# PROCESAR EXCEL
# =====================================================
def procesar(excel_path, user, pwd, output):
    wb = load_workbook(excel_path)
    ws = wb.active
    resultados = {"ACTUALIZADO": 0, "SIN COINCIDENCIA": 0, "ERROR": 0}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.set_default_timeout(20000)

        login(page, "http://201.116.142.155/inscritos/#/", user, pwd)

        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=2), start=2):
            vin = str(row[0].value).strip() if row[0].value else ""

            if len(vin) != 17:
                row[1].value = "VIN INVÁLIDO"
                continue

            resultado = actualizar(page, vin)
            row[1].value = resultado

            clave = resultado.split()[0]
            if clave in resultados:
                resultados[clave] += 1

            if i % 2 == 0:
                wb.save(excel_path)

            time.sleep(random.uniform(1.5, 3.0))

        wb.save(excel_path)
        browser.close()

    resumen = (
        f"Actualizados: {resultados['ACTUALIZADO']}\n"
        f"Con error: {resultados['ERROR']}\n"
        f"Sin coincidencia: {resultados['SIN COINCIDENCIA']}"
    )

    output.set(resumen)


# =====================================================
# GUI
# =====================================================
def gui():
    def iniciar():
        ruta = entrada_excel.get()
        usuario = entrada_usuario.get()
        clave = entrada_clave.get()

        if not ruta or not usuario or not clave:
            messagebox.showwarning("Faltan campos", "Completa todos los campos.")
            return

        salida.set("Procesando...")
        threading.Thread(
            target=lambda: procesar(ruta, usuario, clave, salida)
        ).start()

    ventana = tk.Tk()
    ventana.title("Actualizador VIN GUI")
    ventana.geometry("500x400")

    def mostrar_bienvenida():
        mensaje = """ACTUALIZADOR DE VIN EN BITÁCORA
Versión GUI 1.0

Actualiza el campo año_model a 2026 antes de guardar.
Los errores se registran en log y se generan capturas.

FIRMA:
Desarrollado por el área de sistemas — Julio 2025"""
        messagebox.showinfo("Bienvenida", mensaje)

    ventana.after(500, mostrar_bienvenida)

    salida = tk.StringVar()

    tk.Label(ventana, text="Usuario:").pack()
    entrada_usuario = tk.Entry(ventana, width=40)
    entrada_usuario.pack()

    tk.Label(ventana, text="Contraseña:").pack()
    entrada_clave = tk.Entry(ventana, width=40, show="*")
    entrada_clave.pack()

    tk.Label(ventana, text="Archivo Excel con VINs:").pack()
    entrada_excel = tk.Entry(ventana, width=40)
    entrada_excel.pack()

    def seleccionar_archivo():
        archivo = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        entrada_excel.delete(0, tk.END)
        entrada_excel.insert(0, archivo)

    tk.Button(ventana, text="Seleccionar Excel", command=seleccionar_archivo).pack(pady=5)
    tk.Button(ventana, text="Actualizar VINs", bg="green", fg="white", command=iniciar).pack(pady=10)
    tk.Label(ventana, textvariable=salida, fg="blue").pack(pady=10)

    ventana.mainloop()


if __name__ == "__main__":
    gui()
