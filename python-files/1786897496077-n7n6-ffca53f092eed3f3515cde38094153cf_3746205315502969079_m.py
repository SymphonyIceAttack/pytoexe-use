#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BigSeller 拣货单(Pick List) PDF 批量转 Excel 工具  v2

用法:
    pick_list_pdf_to_excel.exe <PDF文件夹或单个PDF路径> [输出Excel路径] [--no-images]

v2 更新:
    - 自动从PDF正文识别语言（简体中文 / 英文），表头列名跟随PDF语言，
      不再强制中英双语混排
    - 提取每行商品图片（Item列缩略图），直接嵌入Excel"明细"sheet对应行
    - 支持 --no-images 关闭图片嵌入（处理量很大时可加速、减小文件体积）

依赖: pdfplumber, openpyxl, Pillow, pypdfium2 (pdfplumber渲染用)
打包: pyinstaller --onefile --collect-all pdfplumber --collect-all pypdfium2 pick_list_pdf_to_excel.py
"""

import sys
import re
import glob
import os
import tempfile
import shutil
import traceback

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ---------------------------------------------------------------------------
# 语言检测 & 双语文案表
# ---------------------------------------------------------------------------

CJK_RE = re.compile(r"[\u4e00-\u9fff]")


def detect_language(text):
    return "zh" if CJK_RE.search(text or "") else "en"


LABELS = {
    "en": {
        "source_file": "Source File",
        "picklist_no": "Pick List No",
        "warehouse": "Warehouse",
        "wave_type": "Wave Type",
        "print_time": "Print Time",
        "package_qty": "Package Qty",
        "sku_qty": "SKU Qty",
        "product_qty": "Product Quantity",
        "no": "#",
        "item_image": "Item Image",
        "shelf": "Shelf",
        "sku": "SKU",
        "qty": "Qty",
        "sheet_detail": "Detail",
        "sheet_summary": "Pick List Summary",
        "sheet_errors": "Errors",
        "col_errors": "Error Message",
    },
    "zh": {
        "source_file": "源文件",
        "picklist_no": "拣货单号",
        "warehouse": "仓库",
        "wave_type": "波次类型",
        "print_time": "打印时间",
        "package_qty": "包裹数量",
        "sku_qty": "SKU数量",
        "product_qty": "商品数量",
        "no": "序号",
        "item_image": "商品图片",
        "shelf": "货架",
        "sku": "SKU",
        "qty": "数量",
        "sheet_detail": "明细",
        "sheet_summary": "拣货单汇总",
        "sheet_errors": "解析异常",
        "col_errors": "异常信息",
    },
}

HEADER_PATTERNS = {
    "warehouse": re.compile(r"(?:Warehouse|仓库)[:：]\s*([^\n]+)"),
    "wave_type": re.compile(r"(?:Wave Type|波次类型)[:：]\s*([^\n]+)"),
    "print_time": re.compile(r"(?:Print Time|打印时间)[:：]\s*([^\n]+)"),
    "package_qty": re.compile(r"(?:Package Qty|包裹数量)[:：]\s*(\d+)"),
    "sku_qty": re.compile(r"(?:SKU Qty|SKU数量)[:：]\s*(\d+)"),
    "product_qty": re.compile(r"(?:Product Quantity|商品数量)[:：]\s*(\d+)"),
}
PICKLIST_NO_PATTERN = re.compile(r"\n(\S+)\s+(?:Wave Type|波次类型)")


def parse_page_header(text, fallback):
    header = dict(fallback) if fallback else {}
    if "Pick List" in text or "拣货单" in text:
        m = PICKLIST_NO_PATTERN.search(text)
        if m:
            header["picklist_no"] = m.group(1).strip()
        for key, pattern in HEADER_PATTERNS.items():
            m = pattern.search(text)
            if m:
                header[key] = m.group(1).strip()
    return header


# ---------------------------------------------------------------------------
# 图片裁剪：把每一行 Item 缩略图 渲染成独立PNG，按纵向位置对应到每行
# ---------------------------------------------------------------------------

def crop_row_images(page, header_top, expected_rows, tmp_dir, page_tag):
    """
    header_top: 表头文字("SKU"/"Item")所在的 top 坐标，只取比它靠下的图片
    expected_rows: 这一页应该有几行数据（用于数量校验）
    返回: 图片文件路径列表，按从上到下顺序；若数量对不上则返回 []（不嵌入，但不报错阻断）
    """
    imgs = [im for im in page.images if im["top"] > header_top + 2]
    imgs.sort(key=lambda im: im["top"])
    if not imgs or len(imgs) != expected_rows:
        return []

    paths = []
    for i, im in enumerate(imgs):
        bbox = (im["x0"], im["top"], im["x1"], im["bottom"])
        try:
            cropped = page.crop(bbox)
            pil_img = cropped.to_image(resolution=150)
            out_path = os.path.join(tmp_dir, f"{page_tag}_row{i}.png")
            pil_img.save(out_path)
            paths.append(out_path)
        except Exception:
            return []  # 任意一张裁剪失败就整批放弃，保证数据不出错位
    return paths


# ---------------------------------------------------------------------------
# 解析单个PDF
# ---------------------------------------------------------------------------

def parse_pdf(pdf_path, tmp_dir, want_images=True):
    detail_rows = []
    summary_rows = []
    errors = []
    seen_picklist_nos = set()
    lang_votes = []

    fname = os.path.basename(pdf_path)
    header = {}

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_idx, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                lang_votes.append(detect_language(text))
                header = parse_page_header(text, header)

                if header.get("picklist_no") and header["picklist_no"] not in seen_picklist_nos:
                    seen_picklist_nos.add(header["picklist_no"])
                    summary_rows.append(dict(
                        source_file=fname,
                        picklist_no=header.get("picklist_no", ""),
                        warehouse=header.get("warehouse", ""),
                        wave_type=header.get("wave_type", ""),
                        print_time=header.get("print_time", ""),
                        package_qty=header.get("package_qty", ""),
                        sku_qty=header.get("sku_qty", ""),
                        product_qty=header.get("product_qty", ""),
                    ))

                header_top = None
                for w in page.extract_words():
                    if w["text"] in ("SKU", "Item", "货架", "序号"):
                        header_top = w["top"] if header_top is None else min(header_top, w["top"])

                page_rows = []
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        first_cell = (row[0] or "").strip()
                        if not first_cell.isdigit():
                            continue
                        cells = [(c or "").replace("\n", " ").strip() for c in row]
                        non_empty = [c for c in cells[1:] if c]
                        if len(non_empty) < 2:
                            errors.append(f"{fname} p{page_idx}: 行数据列数不足 -> {row}")
                            continue
                        qty = non_empty[-1]
                        sku = non_empty[-2]
                        shelf = " ".join(non_empty[:-2]) if len(non_empty) > 2 else ""
                        page_rows.append(dict(
                            source_file=fname,
                            picklist_no=header.get("picklist_no", ""),
                            warehouse=header.get("warehouse", ""),
                            no=first_cell,
                            shelf=shelf,
                            sku=sku,
                            qty=qty,
                            _image_path=None,
                        ))

                if want_images and page_rows and header_top is not None:
                    img_paths = crop_row_images(
                        page, header_top, len(page_rows), tmp_dir,
                        page_tag=f"{fname}_p{page_idx}".replace(" ", "_"),
                    )
                    if img_paths:
                        for r, p in zip(page_rows, img_paths):
                            r["_image_path"] = p
                    elif page_rows:
                        errors.append(
                            f"{fname} p{page_idx}: 图片数量与行数不匹配，该页未嵌入图片"
                        )

                detail_rows.extend(page_rows)
    except Exception as e:
        errors.append(f"{fname}: 解析失败 - {e}")

    if not detail_rows and not errors:
        errors.append(f"{fname}: 未识别到任何拣货单明细，请检查PDF格式是否与模板一致")

    lang = max(set(lang_votes), key=lang_votes.count) if lang_votes else "en"
    return detail_rows, summary_rows, errors, lang


# ---------------------------------------------------------------------------
# Excel 输出
# ---------------------------------------------------------------------------

def style_header_row(ws, row_idx=1):
    fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws, skip_cols=()):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        if col_letter in skip_cols:
            continue
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)


def write_detail_sheet(wb, title, rows, L, with_images):
    ws = wb.create_sheet(title)
    cols = [L["source_file"], L["picklist_no"], L["warehouse"], L["no"]]
    if with_images:
        cols.append(L["item_image"])
    cols += [L["shelf"], L["sku"], L["qty"]]
    ws.append(cols)

    img_col_letter = get_column_letter(cols.index(L["item_image"]) + 1) if with_images else None

    for i, r in enumerate(rows, start=2):
        row_vals = [r["source_file"], r["picklist_no"], r["warehouse"], r["no"]]
        if with_images:
            row_vals.append("")
        row_vals += [r["shelf"], r["sku"], r["qty"]]
        ws.append(row_vals)

        if with_images and r.get("_image_path"):
            try:
                xl_img = XLImage(r["_image_path"])
                xl_img.width = 45
                xl_img.height = 45
                ws.add_image(xl_img, f"{img_col_letter}{i}")
                ws.row_dimensions[i].height = 34
            except Exception:
                pass

    if rows:
        style_header_row(ws)
        ws.freeze_panes = "A2"
        skip = {img_col_letter} if img_col_letter else set()
        autosize_columns(ws, skip_cols=skip)
        if img_col_letter:
            ws.column_dimensions[img_col_letter].width = 8
    return ws


def write_summary_sheet(wb, title, rows, L):
    ws = wb.create_sheet(title)
    cols = [L["source_file"], L["picklist_no"], L["warehouse"], L["wave_type"],
            L["print_time"], L["package_qty"], L["sku_qty"], L["product_qty"]]
    ws.append(cols)
    for r in rows:
        ws.append([r["source_file"], r["picklist_no"], r["warehouse"], r["wave_type"],
                   r["print_time"], r["package_qty"], r["sku_qty"], r["product_qty"]])
    if rows:
        style_header_row(ws)
        ws.freeze_panes = "A2"
        autosize_columns(ws)
    return ws


def write_error_sheet(wb, title, errors, L):
    ws = wb.create_sheet(title)
    ws.append([L["col_errors"]])
    for e in errors:
        ws.append([e])
    style_header_row(ws)
    autosize_columns(ws)
    return ws


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    want_images = "--no-images" not in sys.argv[1:]

    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.getcwd()

    default_input = base_dir
    default_output = os.path.join(base_dir, "拣货单汇总.xlsx")

    input_path = args[0] if len(args) > 0 else default_input
    output_path = args[1] if len(args) > 1 else default_output

    if os.path.isdir(input_path):
        pdf_files = sorted(glob.glob(os.path.join(input_path, "*.pdf")))
    else:
        pdf_files = [input_path]

    if not pdf_files:
        print(f"未在 {input_path} 找到任何PDF文件")
        sys.exit(1)

    tmp_dir = tempfile.mkdtemp(prefix="picklist_imgs_")
    all_details, all_summaries, all_errors, langs = [], [], [], []

    try:
        for pdf_path in pdf_files:
            details, summaries, errors, lang = parse_pdf(pdf_path, tmp_dir, want_images)
            all_details.extend(details)
            all_summaries.extend(summaries)
            all_errors.extend(errors)
            langs.append(lang)
            print(f"已处理: {os.path.basename(pdf_path)} -> {len(details)} 条明细, "
                  f"{len(errors)} 条异常, 语言={lang}")

        overall_lang = max(set(langs), key=langs.count) if langs else "en"
        L = LABELS[overall_lang]

        wb = Workbook()
        wb.remove(wb.active)

        write_detail_sheet(wb, L["sheet_detail"], all_details, L, want_images)
        write_summary_sheet(wb, L["sheet_summary"], all_summaries, L)
        if all_errors:
            write_error_sheet(wb, L["sheet_errors"], all_errors, L)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)

        print(f"\n完成！共处理 {len(pdf_files)} 个PDF文件，"
              f"{len(all_summaries)} 张拣货单，{len(all_details)} 条明细。"
              f"（识别语言: {overall_lang}）")
        print(f"输出文件: {output_path}")
        if all_errors:
            print(f"注意: 有 {len(all_errors)} 条解析异常，详见Excel中的 '{L['sheet_errors']}' sheet")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _pause_before_exit():
    """打包成exe双击运行时，窗口默认处理完就自动关闭，来不及看结果/报错。
    这里统一暂停等待按键，方便双击用户看清输出。命令行运行时按一下回车即可。"""
    try:
        input("\n按回车键关闭窗口...")
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        _pause_before_exit()
        raise
    except Exception:
        print("\n运行出错，详细信息如下：\n")
        traceback.print_exc()
        _pause_before_exit()
        sys.exit(1)
    else:
        _pause_before_exit()
