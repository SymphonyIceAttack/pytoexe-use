from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, Reference
from datetime import datetime

# =========================
# 🎨 стили
# =========================
header_fill = PatternFill(start_color="2F2F2F", end_color="2F2F2F", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
center = Alignment(horizontal="center")

# =========================
# 📊 1. ОБЩИЙ ФАЙЛ
# =========================
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Операции"

headers = ["Дата", "Тип", "Категория", "Подкатегория", "Сумма (₽)", "Комментарий"]
ws1.append(headers)

for col in range(1, len(headers)+1):
    cell = ws1.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

# пример строки
ws1.append(["2026-04-15", "Расход", "Еда", "Кафе", 1200, "Обед"])

# сводка
ws1_s = wb1.create_sheet("Сводка")

ws1_s["A1"] = "Доход"
ws1_s["A2"] = "Расход"
ws1_s["A3"] = "Баланс"

ws1_s["B1"] = '=SUMIF(Операции!B:B,"Доход",Операции!E:E)'
ws1_s["B2"] = '=SUMIF(Операции!B:B,"Расход",Операции!E:E)'
ws1_s["B3"] = "=B1-B2"

# график расходов
pie = PieChart()
labels = Reference(ws1, min_col=3, min_row=2, max_row=10)
data = Reference(ws1, min_col=5, min_row=1, max_row=10)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Расходы по категориям"

ws1_s.add_chart(pie, "D2")

wb1.save("Финансы_общий.xlsx")