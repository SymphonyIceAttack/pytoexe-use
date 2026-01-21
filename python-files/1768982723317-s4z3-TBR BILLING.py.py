
import tkinter as tk
from tkinter import messagebox
import openpyxl
import os
from datetime import datetime

# ------------------ APP SETUP ------------------
root = tk.Tk()
root.title("Billing Software")
root.geometry("900x650")

items = []
invoice_no = 1
FILE = "Billing_Data.xlsx"

# ------------------ EXCEL SETUP ------------------
def init_excel():
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.append(["Invoice", "Date", "Customer", "Mobile",
               "Item", "Qty", "Price", "Amount", "GST", "Total"])
    wb.save(FILE)

if not os.path.exists(FILE):
    init_excel()

# ------------------ FUNCTIONS ------------------
def add_item():
    name = entry_item.get()
    qty = entry_qty.get()
    price = entry_price.get()

    if not name or not qty or not price:
        messagebox.showerror("Error", "Enter all item details")
        return

    amount = int(qty) * float(price)
    items.append((name, int(qty), float(price), amount))

    listbox.insert(tk.END, f"{name:<20}{qty:<10}{price:<10}{amount:<10}")

    entry_item.delete(0, tk.END)
    entry_qty.delete(0, tk.END)
    entry_price.delete(0, tk.END)
    entry_item.focus()

def generate_bill():
    global invoice_no
    customer = entry_customer.get()
    mobile = entry_mobile.get()

    if not customer or not mobile or not items:
        messagebox.showerror("Error", "Missing customer or items")
        return

    total = sum(i[3] for i in items)
    gst_rate = float(gst_var.get())
    gst_amt = total * gst_rate / 100
    grand = total + gst_amt

    bill_text.delete("1.0", tk.END)
    bill_text.insert(tk.END, "🧾 BILL RECEIPT\n")
    bill_text.insert(tk.END, f"Invoice : {invoice_no}\n")
    bill_text.insert(tk.END, f"Date    : {datetime.now().strftime('%d-%m-%Y')}\n")
    bill_text.insert(tk.END, f"Customer: {customer}\n")
    bill_text.insert(tk.END, f"Mobile  : {mobile}\n")
    bill_text.insert(tk.END, "-"*50 + "\n")

    for it in items:
        bill_text.insert(tk.END, f"{it[0]} x{it[1]} = ₹{it[3]}\n")

    bill_text.insert(tk.END, "-"*50 + "\n")
    bill_text.insert(tk.END, f"Subtotal : ₹{total}\n")
    bill_text.insert(tk.END, f"GST {gst_rate}% : ₹{gst_amt:.2f}\n")
    bill_text.insert(tk.END, f"TOTAL    : ₹{grand:.2f}\n")

    # Save to Excel
    wb = openpyxl.load_workbook(FILE)
    sh = wb.active
    for it in items:
        sh.append([
            invoice_no,
            datetime.now().strftime('%d-%m-%Y'),
            customer, mobile,
            it[0], it[1], it[2], it[3],
            gst_rate, grand
        ])
    wb.save(FILE)

    invoice_no += 1

def clear_all():
    items.clear()
    listbox.delete(0, tk.END)
    bill_text.delete("1.0", tk.END)
    entry_customer.delete(0, tk.END)
    entry_mobile.delete(0, tk.END)
    entry_item.delete(0, tk.END)
    entry_qty.delete(0, tk.END)
    entry_price.delete(0, tk.END)
    entry_customer.focus()

def show_history():
    hist = tk.Toplevel(root)
    hist.title("Bill History")
    hist.geometry("600x400")

    frame = tk.Frame(hist)
    frame.pack(fill=tk.BOTH, expand=True)

    scroll = tk.Scrollbar(frame)
    scroll.pack(side=tk.RIGHT, fill=tk.Y)

    lb = tk.Listbox(frame, font=("Courier New", 10), yscrollcommand=scroll.set)
    lb.pack(fill=tk.BOTH, expand=True)
    scroll.config(command=lb.yview)

    wb = openpyxl.load_workbook(FILE)
    sh = wb.active

    seen = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        inv = row[0]
        seen.setdefault(inv, row)

    for k, v in seen.items():
        lb.insert(tk.END, f"Invoice {k} | {v[1]} | {v[2]} | ₹{v[9]}")

def delete_all_history():
    confirm = messagebox.askyesno(
        "Confirm Delete",
        "⚠️ This will DELETE ALL billing history.\nThis action cannot be undone.\n\nContinue?"
    )
    if not confirm:
        return

    if os.path.exists(FILE):
        os.remove(FILE)
        init_excel()

    messagebox.showinfo("Done", "All billing history deleted successfully.")

# ------------------ UI ------------------
tk.Label(root, text="Customer Name").place(x=20, y=20)
entry_customer = tk.Entry(root)
entry_customer.place(x=150, y=20)

tk.Label(root, text="Mobile").place(x=20, y=50)
entry_mobile = tk.Entry(root)
entry_mobile.place(x=150, y=50)

tk.Label(root, text="Item").place(x=20, y=90)
entry_item = tk.Entry(root)
entry_item.place(x=150, y=90)

tk.Label(root, text="Qty").place(x=20, y=120)
entry_qty = tk.Entry(root)
entry_qty.place(x=150, y=120)

tk.Label(root, text="Price").place(x=20, y=150)
entry_price = tk.Entry(root)
entry_price.place(x=150, y=150)

tk.Label(root, text="GST %").place(x=20, y=180)
gst_var = tk.StringVar(value="5")
tk.OptionMenu(root, gst_var, "0", "5", "12", "18", "28").place(x=150, y=175)

tk.Button(root, text="Add Item", command=add_item).place(x=20, y=220)
tk.Button(root, text="Generate Bill", command=generate_bill).place(x=120, y=220)
tk.Button(root, text="Clear / New Bill", command=clear_all).place(x=260, y=220)
tk.Button(root, text="History", command=show_history).place(x=420, y=220)
tk.Button(root, text="Delete All History", fg="red",
          command=delete_all_history).place(x=520, y=220)

# ------------------ LISTBOX (SCROLL) ------------------
frame1 = tk.Frame(root)
frame1.place(x=20, y=260)

sb1 = tk.Scrollbar(frame1)
sb1.pack(side=tk.RIGHT, fill=tk.Y)

listbox = tk.Listbox(frame1, width=60, height=8,
                     font=("Courier New", 10),
                     yscrollcommand=sb1.set)
listbox.pack(side=tk.LEFT)
sb1.config(command=listbox.yview)

# ------------------ BILL TEXT (SCROLL) ------------------
frame2 = tk.Frame(root)
frame2.place(x=20, y=420)

sb2 = tk.Scrollbar(frame2)
sb2.pack(side=tk.RIGHT, fill=tk.Y)

bill_text = tk.Text(frame2, width=75, height=10,
                    yscrollcommand=sb2.set)
bill_text.pack(side=tk.LEFT)
sb2.config(command=bill_text.yview)

root.mainloop()
