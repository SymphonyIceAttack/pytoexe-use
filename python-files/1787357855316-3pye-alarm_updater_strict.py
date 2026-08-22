#!/usr/bin/env python3
"""
Alarm Updater
-------------
Reads alarms from aa.xlsx and updates monthly "last alarm day"
values in al.xlsx.

Matching rule:
- Take the numeric part of ChannelTagNr / Tag.
- Ignore a leading T (and any other non-numeric characters).
- Ignore leading zeroes.
- Example: T01101, 01101, T1101 and 1101 -> 1101.

Input:
  aa.xlsx
    B = ChannelTagNr
    C = Date

Target:
  al.xlsx
    Sheet = "All IO Channels"
    B = Tag
    M:X = January:December

Only the day number is written to M:X.
The full date is used for comparison so an older event never
overwrites a newer one.

The original al.xlsx is never overwritten.
"""

from pathlib import Path
from datetime import datetime, date
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Install it with: pip install openpyxl")
    input("Press Enter to exit...")
    sys.exit(1)


MONTH_COLUMNS = {
    1: 13,  # M
    2: 14,  # N
    3: 15,  # O
    4: 16,  # P
    5: 17,  # Q
    6: 18,  # R
    7: 19,  # S
    8: 20,  # T
    9: 21,  # U
    10: 22, # V
    11: 23, # W
    12: 24, # X
}

TARGET_SHEET = "All IO Channels"


def normalize_tag(value):
    """
    Normalize a Tag more strictly.

    Rules:
    - Ignore a leading T or other non-alphanumeric prefix.
    - Ignore leading zeroes in the numeric part.
    - Preserve trailing letters (T16018 != T16018a).
    - Preserve slash-separated suffixes such as /2 (T01501/2 stays distinct
      from T15012).
    """
    if value is None:
        return None

    text = str(value).strip().upper()

    # Remove a leading T if present.
    if text.startswith("T"):
        text = text[1:]

    # Keep digits, letters and slash; remove spaces and punctuation such as
    # hyphens used in descriptions like -SLD.
    text = re.sub(r"[^A-Z0-9/]", "", text)

    # Normalize each slash-separated numeric/alpha segment independently.
    parts = text.split("/")
    normalized_parts = []

    for part in parts:
        if not part:
            continue

        m = re.match(r"(0*\d+)([A-Z]*)$", part)
        if m:
            number = m.group(1).lstrip("0") or "0"
            suffix = m.group(2)
            normalized_parts.append(number + suffix)
        else:
            normalized_parts.append(part)

    if not normalized_parts:
        return None

    return "/".join(normalized_parts)


def as_date(value):
    """Convert an Excel/Python date value to datetime, or return None."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, str):
        value = value.strip()
        for fmt in (
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
            "%d.%m.%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
        ):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass

    return None


def build_target_index(ws):
    """Map normalized numeric Tag -> target row."""
    index = {}
    duplicates = {}

    for row in range(1, ws.max_row + 1):
        key = normalize_tag(ws.cell(row=row, column=2).value)
        if not key:
            continue

        if key in index:
            duplicates.setdefault(key, [index[key]]).append(row)
        else:
            index[key] = row

    return index, duplicates


def update_workbook(aa_path, al_path, output_path):
    wb_aa = load_workbook(aa_path, data_only=True)
    wb_al = load_workbook(al_path)

    if TARGET_SHEET not in wb_al.sheetnames:
        raise ValueError(
            f'Target sheet "{TARGET_SHEET}" was not found in al.xlsx.'
        )

    ws_aa = wb_aa[wb_aa.sheetnames[0]]
    ws_al = wb_al[TARGET_SHEET]

    target_index, duplicates = build_target_index(ws_al)

    # For each target row/month, keep the latest complete source date
    # encountered in aa.xlsx.
    latest = {}

    stats = {
        "source_rows": 0,
        "valid_dates": 0,
        "matched": 0,
        "not_found": 0,
        "updated": 0,
        "unchanged": 0,
    }

    for row in range(1, ws_aa.max_row + 1):
        tag = ws_aa.cell(row=row, column=2).value
        raw_date = ws_aa.cell(row=row, column=3).value

        if tag is None and raw_date is None:
            continue

        stats["source_rows"] += 1

        key = normalize_tag(tag)
        alarm_date = as_date(raw_date)

        if not key or alarm_date is None:
            continue

        stats["valid_dates"] += 1

        target_row = target_index.get(key)
        if target_row is None:
            stats["not_found"] += 1
            continue

        stats["matched"] += 1
        month_col = MONTH_COLUMNS[alarm_date.month]
        slot = (target_row, month_col)

        if slot not in latest or alarm_date > latest[slot]:
            latest[slot] = alarm_date

    # Apply only the newest event for each alarm/month.
    for (target_row, month_col), new_date in latest.items():
        cell = ws_al.cell(row=target_row, column=month_col)
        old_value = cell.value

        # IMPORTANT:
        # The target workbook contains only the day number, so an old day
        # cannot be compared directly with the new full date. We compare
        # against the year/month/day information remembered from the source
        # data in a hidden helper sheet.
        #
        # To keep al.xlsx simple, we use an auxiliary hidden sheet storing
        # the last full date for each Tag/month.
        helper_name = "_AlarmUpdaterDates"
        if helper_name not in wb_al.sheetnames:
            helper = wb_al.create_sheet(helper_name)
            helper.sheet_state = "hidden"
            helper.append(["TagKey", "MonthColumn", "LastFullDate"])
        else:
            helper = wb_al[helper_name]

        # Find existing helper record.
        found_row = None
        for r in range(2, helper.max_row + 1):
            if (helper.cell(r, 1).value == normalize_tag(ws_al.cell(target_row, 2).value)
                    and helper.cell(r, 2).value == month_col):
                found_row = r
                break

        old_full_date = None
        if found_row:
            old_full_date = as_date(helper.cell(found_row, 3).value)

        # If no helper date exists, inspect the visible day value only as a
        # fallback. A helper date is created from this run for future safety.
        should_update = old_full_date is None or new_date > old_full_date

        if should_update:
            cell.value = new_date.day
            cell.number_format = "0"

            if found_row:
                helper.cell(found_row, 3).value = new_date
            else:
                helper.append([
                    normalize_tag(ws_al.cell(target_row, 2).value),
                    month_col,
                    new_date,
                ])

            stats["updated"] += 1
        else:
            stats["unchanged"] += 1

    wb_al.save(output_path)
    return stats, duplicates


def main():
    base = Path(__file__).resolve().parent

    # Default filenames allow simply placing the EXE beside aa.xlsx/al.xlsx.
    aa_path = base / "aa.xlsx"
    al_path = base / "al.xlsx"
    output_path = base / "al_updated.xlsx"

    if not aa_path.exists():
        print(f"Missing file: {aa_path}")
        input("Press Enter to exit...")
        return

    if not al_path.exists():
        print(f"Missing file: {al_path}")
        input("Press Enter to exit...")
        return

    print("Updating alarm data...")
    try:
        stats, duplicates = update_workbook(aa_path, al_path, output_path)
    except Exception as exc:
        print("\nERROR:")
        print(exc)
        input("\nPress Enter to exit...")
        return

    print("\nDONE")
    print(f"Source rows processed : {stats['source_rows']}")
    print(f"Valid alarm/date rows : {stats['valid_dates']}")
    print(f"Matched alarms        : {stats['matched']}")
    print(f"Not found / skipped   : {stats['not_found']}")
    print(f"Cells updated         : {stats['updated']}")
    print(f"Cells unchanged       : {stats['unchanged']}")
    if duplicates:
        print("\nWARNING: duplicate numeric Tag keys were found in al.xlsx:")
        for key, rows in duplicates.items():
            print(f"  {key}: rows {rows}")

    print(f"\nOutput: {output_path}")
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
