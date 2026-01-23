import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

def generar_excel(balance_file):
    wb_orig = load_workbook(balance_file, data_only=True)
    df_balance = pd.read_excel(balance_file, sheet_name="Balance")

    # Nuevo libro de Excel
    wb = Workbook()
    wb.remove(wb.active)

    teams = ["Alegría City","Robloxia","Serafín","Banana","Inverciudad","Perrito"]

    # Copiar hoja Balance
    ws_balance = wb.create_sheet("Balance")
    for row in wb_orig["Balance"].iter_rows(values_only=True):
        ws_balance.append(row)

    # Crear hojas por equipo
    for team in teams:
        ws = wb.create_sheet(team)
        ws.append(["Empresa","Monto Invertido","Año Compra",
                   "Valor Inicial","Valor Actual","Rentabilidad %","Ganancia"])
        for r in range(2, 12):
            ws[f"D{r}"] = f'=SI.ERROR(BUSCARX(A{r}&C{r},Balance!$A:$A&Balance!$B:$B,Balance!$C:$C),0)'
            ws[f"E{r}"] = f'=SI.ERROR(BUSCARX(A{r},Balance!$A:$A,Balance!$D:$D),0)'
            ws[f"F{r}"] = f'=SI(D{r}=0,0,(E{r}-D{r})/D{r}*100)'
            ws[f"G{r}"] = f'=B{r}*(F{r}/100)'

        # Resumen
        ws["J2"]="Total Invertido"; ws["K2"]="=SUMA(B2:B11)"
        ws["J3"]="Valor Actual Total"; ws["K3"]="=SUMA(E2:E11)"
        ws["J4"]="Ganancia Total"; ws["K4"]="=SUMA(G2:G11)"
        ws["J5"]="Rent Promedio %"; ws["K5"]="=SI(K2=0,0,K4/K2*100)"

        # Estado
        ws["J7"]="Estado:"; ws["K7"]='=SI(K4>0,"😊 Vas Ganando","😕 Estás en Pérdida")'
        ws["J7"].font = Font(bold=True)

        # Gráfico
        chart = BarChart(); chart.title=f"Rentabilidad % - {team}"
        data = Reference(ws, min_col=6, min_row=1, max_row=11)
        cats = Reference(ws, min_col=1, min_row=2, max_row=11)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J10")

    # Dashboard global
    ws_g = wb.create_sheet("Dashboard Global")
    ws_g.append(["Equipo","Total Invertido","Valor Actual",
                 "Ganancia Total","Rent Promedio %"])
    row = 2
    for team in teams:
        ws_g.append([team,
                     f"='{team}'!K2",
                     f"='{team}'!K3",
                     f"='{team}'!K4",
                     f"='{team}'!K5"])
        row += 1

    chart2 = BarChart(); chart2.title="Valor Actual por Equipo"
    data2 = Reference(ws_g, min_col=3, min_row=1, max_row=row)
    cats2 = Reference(ws_g, min_col=1, min_row=2, max_row=row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws_g.add_chart(chart2, "F2")

    output = "HDD_Version_2_Dashboard.xlsx"
    wb.save(output)
    return output

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: generador.exe [archivo_balance.xlsx]")
    else:
        balance_file = sys.argv[1]
        result = generar_excel(balance_file)
        print("Archivo creado:", result)
