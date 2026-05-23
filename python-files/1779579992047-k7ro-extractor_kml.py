Python 3.14.0 (tags/v3.14.0:ebf955d, Oct  7 2025, 10:15:03) [MSC v.1944 64 bit (AMD64)] on win32
Type "help", "copyright", "credits" or "license" for more information.
>>> import os
import zipfile
import re
from xml.etree import ElementTree
import openpyxl
from openpyxl.styles import Font, PatternFill

def parse_kml_coordinates(coord_text):
    """Limpia y extrae las coordenadas de texto del KML."""
    coords = []
    # Eliminar espacios en blanco innecesarios y separar por puntos de coordenadas
    parts = coord_text.strip().split()
    for part in parts:
        if ',' in part:
            sub_parts = part.split(',')
            if len(sub_parts) >= 2:
                try:
                    lon = float(sub_parts[0])
                    lat = float(sub_parts[1])
                    coords.append((lon, lat))
                except ValueError:
                    continue
    return coords

def extract_metadata(xml_content):
    """Extrae los elementos Placemark con su respectivo nombre y coordenadas."""
    # Remover namespaces del XML para facilitar la búsqueda con ElementTree
    xml_content = re.sub(r'xmlns="[^"]+"', '', xml_content, count=1)
    root = ElementTree.fromstring(xml_content)
    
    data = []
    # Buscar todos los Placemarks en el archivo
    for placemark in root.findall('.//Placemark'):
        name_node = placemark.find('name')
        name = name_node.text.strip() if name_node is not None and name_node.text else "Sin_Nombre"
        
        # Intentar extraer geometría de punto
        coord_node = placemark.find('.//Point/coordinates')
        if coord_node is not None and coord_node.text:
            coords = parse_kml_coordinates(coord_node.text)
            if coords:
                data.append({
                    'nombre': name,
                    'tipo': 'Punto / Poste / Nodo',
                    'lon': coords[0][0],
                    'lat': coords[0][1]
                })
                continue
                
        # Intentar extraer geometría de línea (tramos de fibra)
        line_node = placemark.find('.//LineString/coordinates')
        if line_node is not None and line_node.text:
            coords = parse_kml_coordinates(line_node.text)
            if coords:
                # Tomamos el punto inicial para la metadata básica en Excel
                data.append({
                    'nombre': name,
                    'tipo': 'Línea / Tramo Fibra',
                    'lon': coords[0][0],
                    'lat': coords[0][1]
                })
    return data

def process_file(file_path):
    """Detecta el tipo de archivo (KML/KMZ) y extrae su contenido XML."""
    if file_path.lower().endswith('.kmz'):
        with zipfile.ZipFile(file_path, 'r') as z:
            kml_files = [f for f in z.namelist() if f.lower().endswith('.kml')]
            if not kml_files:
                return None
            return z.read(kml_files[0]).decode('utf-8', errors='ignore')
    else:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

def create_excel(data, output_xlsx):
    """Genera un archivo Excel ordenado y formateado con la metadata."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metadata Fibra"
    
    # Encabezados
    headers = ["Elemento", "Tipo", "Longitud (X)", "Latitud (Y)"]
    ws.append(headers)
    
    # Estilo básico para encabezados
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
    
    # Insertar datos ordenados alfabéticamente por nombre del elemento
    sorted_data = sorted(data, key=lambda x: x['nombre'])
    for item in sorted_data:
        ws.append([item['nombre'], item['tipo'], item['lon'], item['lat']])
        
    # Autoajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output_xlsx)

def create_autocad_script(data, output_scr):
    """Genera el archivo .scr para AutoCAD utilizando coordenadas WGS84."""
    with open(output_scr, 'w', encoding='utf-8') as f:
        f.write("; Script generado automáticamente para diseño de Fibra Óptica\n")
        f.write("PDMODE 34\n")  # Define el estilo del punto (cruz con círculo)
        f.write("PDSIZE 0.5\n")  # Tamaño del punto en pantalla
        
        for item in data:
            x = item['lon']
            y = item['lat']
            nombre = item['nombre'].replace(" ", "_") # Evita problemas de espacios en comandos de AutoCAD
            
            # Dibujar el Punto
            f.write(f"_POINT {x},{y}\n")
            
            # Insertar el Texto identificador justo al lado del punto
            # Formato: _TEXT [Punto inserción] [Altura] [Rotación] [Texto]
            f.write(f"_TEXT {x + 0.00005},{y + 0.00005} 0.0001 0 {nombre}\n")
            
        f.write("_ZOOM _E\n")  # Hace un Zoom Extents al finalizar la carga

def main():
    print("=" * 60)
    print(" EXTRACTOR DE METADATA KML/KMZ PARA PLANOS DE FIBRA ÓPTICA")
    print("=" * 60)
    
    file_path = input("\nArrastra o introduce la ruta del archivo (.kml o .kmz): ").strip('" ')
    
    if not os.path.exists(file_path):
        print("[Error] El archivo especificado no existe.")
        input("\nPresiona Enter para salir...")
        return
        
    print("\n[1/3] Procesando archivo y extrayendo datos...")
    xml_content = process_file(file_path)
    
    if not xml_content:
        print("[Error] No se pudo leer el contenido del KML.")
        input("\nPresiona Enter para salir...")
        return
        
    metadata = extract_metadata(xml_content)
    
    if not metadata:
        print("[Advertencia] No se encontraron elementos con coordenadas válidas.")
        input("\nPresiona Enter para salir...")
        return
        
    base_path = os.path.splitext(file_path)[0]
    output_xlsx = base_path + "_Metadata.xlsx"
    output_scr = base_path + "_AutoCAD.scr"
    
    print("[2/3] Generando reporte estructurado en Excel...")
    create_excel(metadata, output_xlsx)
    
    print("[3/3] Creando script de inserción rápida (.scr) para AutoCAD...")
    create_autocad_script(metadata, output_scr)
    
    print("\n" + "=" * 60)
    print(" ¡PROCESO COMPLETADO CON ÉXITO!")
    print("=" * 60)
    print(f" Excel generado: {os.path.basename(output_xlsx)}")
    print(f" Script AutoCAD: {os.path.basename(output_scr)}")
    print("=" * 60)
    
    # Nota técnica sobre coordenadas
    print("\n> NOTA PARA AUTOCAD:")
    print("  Este script exporta las coordenadas nativas en formato geográfico (Longitud/Latitud).")
    print("  Para que coincida con tus planos UTM de topografía o catastro, asegúrate de")
    print("  configurar la geoubicación en AutoCAD (comando GEOGRAPHICLOCATION) en el huso")
    print("  correspondiente antes de correr el archivo .scr.")
    
    input("\nPresiona Enter para cerrar el programa...")

if __name__ == "__main__":
    main()
