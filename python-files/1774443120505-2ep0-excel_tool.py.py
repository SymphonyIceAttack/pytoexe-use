import openpyxl
import sys
import os
from openpyxl.styles import Alignment, Font

def num2chinese_upper_wanyuan(amount):
    if amount is None or str(amount).strip() == "":
        return "零元整"
    try:
        amount = round(float(amount) * 10000, 2)
    except:
        return "无效金额"
    if amount == 0:
        return "零元整"

    digits = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]
    units = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿"]
    yuan = int(amount)
    jiao = int((amount - yuan) * 10)
    fen = int(round((amount - yuan) * 100 - jiao * 10, 0))

    def convert_section(n):
        section = ""
        has_zero = False
        for i in range(4):
            d = n // (10 ** (3 - i)) % 10
            if d != 0:
                if has_zero:
                    section += "零"
                section += digits[d] + units[3 - i]
                has_zero = False
            else:
                has_zero = True if i < 3 else has_zero
        return section

    yuan_str = ""
    if yuan >= 100000000:
        yuan_str += convert_section(yuan // 100000000) + "亿"
        yuan = yuan % 100000000
    if yuan >= 10000:
        yuan_str += convert_section(yuan // 10000) + "万"
        yuan = yuan % 10000
    if yuan > 0:
        yuan_str += convert_section(yuan)
    yuan_str = yuan_str if yuan_str else "零"
    yuan_str += "元"

    if jiao == 0 and fen == 0:
        yuan_str += "整"
    else:
        if jiao > 0:
            yuan_str += digits[jiao] + "角"
        if fen > 0:
            yuan_str += digits[fen] + "分"
    return yuan_str

def process_file(input_file):
    folder = os.path.dirname(input_file)
    filename = os.path.basename(input_file)
    name, ext = os.path.splitext(filename)
    output_file = os.path.join(folder, f"{name}_大写{ext}")

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 取消2、3行合并
    merged = list(ws.merged_cells.ranges)
    for m in merged:
        if 2 <= m.max_row and 2 >= m.min_row or 3 <= m.max_row and 3 >= m.min_row:
            ws.unmerge_cells(str(m))

    # 取消M列合并
    merged = list(ws.merged_cells.ranges)
    for m in merged:
        if m.min_col <= 13 <= m.max_col:
            ws.unmerge_cells(str(m))

    # G列左侧插入大写金额列
    insert_col = 7
    ws.insert_cols(insert_col)
    ws.cell(row=5, column=insert_col, value="大写金额")
    ws.cell(row=5, column=insert_col).font = Font(bold=True)
    ws.cell(row=5, column=insert_col).alignment = Alignment(horizontal="center", vertical="center")

    for r in range(6, ws.max_row + 1):
        val = ws.cell(row=r, column=insert_col + 1).value
        upper_val = num2chinese_upper_wanyuan(val)
        ws.cell(row=r, column=insert_col, value=upper_val)
        ws.cell(row=r, column=insert_col).alignment = Alignment(horizontal="left", vertical="center")

    # 合并
    ws.merge_cells('A1:N1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C4:N4')
    ws['C4'].alignment = Alignment(horizontal='left', vertical='center')

    merge_list_left = [
        "A2:B2", "C2:I2", "K2:L2", "M2:N2",
        "A3:B3", "C3:I3", "K3:L3", "M3:N3"
    ]
    for m_range in merge_list_left:
        try:
            ws.merge_cells(m_range)
            start = ws[m_range.split(":")[0]]
            start.alignment = Alignment(horizontal="left", vertical="center")
        except:
            pass

    # 列宽
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["L"].width = 8
    ws.column_dimensions["N"].width = 14

    wb.save(output_file)
    wb.close()
    return output_file

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("请将Excel文件拖拽到本程序上运行！")
        os.system("pause")
        sys.exit()

    input_path = sys.argv[1]
    try:
        out = process_file(input_path)
        print(f"处理完成！\n输出文件：{out}")
    except Exception as e:
        print(f"出错：{e}")
    os.system("pause")