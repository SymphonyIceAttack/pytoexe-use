import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import os

# ===================== 配置参数 =====================
METRIC_SHEET_INDEX = 1
COL_NAME = "员工姓名"
COL_DAILY_WAGE = "本日工资总额"

BASE_WAGE = {
    "男_试用": 2200,
    "男_转正": 2400,
    "女_试用": 2000,
    "女_转正": 2200
}
NIGHT_SUBSIDY = 5
FULL_ATTEND_DAY = 15
SENIORITY_MONTH = 20

ATTEND_HEADERS = ["姓名", "性别", "试用状态", "当月出勤天数", "夜班天数", "工龄满一年"]
# ====================================================

def auto_fit_column(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value if cell.value is not None else "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max_len + 4

def read_metric_single(file_path):
    res = defaultdict(float)
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        if len(wb.worksheets) < METRIC_SHEET_INDEX + 1:
            messagebox.showwarning("警告", f"{os.path.basename(file_path)} 工作表不足2个，跳过")
            wb.close()
            return res
        ws = wb.worksheets[METRIC_SHEET_INDEX]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if COL_NAME not in headers or COL_DAILY_WAGE not in headers:
            messagebox.showwarning("警告", f"{os.path.basename(file_path)} 缺少姓名/本日工资总额列，跳过")
            wb.close()
            return res
        idx_name = headers.index(COL_NAME)
        idx_wage = headers.index(COL_DAILY_WAGE)
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[idx_name]).strip() if row[idx_name] else ""
            wage = float(row[idx_wage]) if row[idx_wage] is not None else 0.0
            if name:
                res[name] += wage
        wb.close()
    except Exception as e:
        messagebox.showerror("读取失败", f"{os.path.basename(file_path)}：{str(e)}")
    return res

def batch_metric(file_list, progress_var):
    total_data = defaultdict(float)
    total = len(file_list)
    for i, f in enumerate(file_list):
        progress_var.set(f"计量文件处理：{i+1}/{total}")
        day_data = read_metric_single(f)
        for name, val in day_data.items():
            total_data[name] += val
    return dict(total_data)

def read_attend_single(file_path):
    res = defaultdict(float)
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for h in ATTEND_HEADERS:
            if h not in headers:
                messagebox.showerror("考勤格式错误", f"{os.path.basename(file_path)} 缺失列：{h}")
                wb.close()
                return res
        idx_name = headers.index("姓名")
        idx_gender = headers.index("性别")
        idx_trial = headers.index("试用状态")
        idx_workday = headers.index("当月出勤天数")
        idx_night = headers.index("夜班天数")
        idx_senior = headers.index("工龄满一年")
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[idx_name]).strip()
            gender = str(row[idx_gender]).strip()
            trial = str(row[idx_trial]).strip()
            work_days = int(row[idx_workday]) if row[idx_workday] else 0
            night_days = int(row[idx_night]) if row[idx_night] else 0
            senior_flag = str(row[idx_senior]).strip()
            base = BASE_WAGE.get(f"{gender}_{trial}", 0)
            night_money = night_days * NIGHT_SUBSIDY
            full_money = FULL_ATTEND_DAY * work_days if work_days >= 26 else 0
            senior_money = SENIORITY_MONTH if senior_flag == "是" else 0
            total_attend = base + night_money + full_money + senior_money
            res[name] = total_attend
        wb.close()
    except Exception as e:
        messagebox.showerror("考勤读取失败", f"{os.path.basename(file_path)}：{str(e)}")
    return res

def batch_attend(file_list, progress_var):
    total_data = defaultdict(float)
    total = len(file_list)
    for i, f in enumerate(file_list):
        progress_var.set(f"考勤文件处理：{i+1}/{total}")
        data = read_attend_single(f)
        for name, val in data.items():
            total_data[name] += val
    return dict(total_data)

def export_summary(metric_dict, attend_dict, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "月度工资汇总"
    font_header = Font(bold=True, color="FFFFFFFF")
    font_normal = Font()
    fill_header = PatternFill("solid", fgColor="FF366092")
    fill_total = PatternFill("solid", fgColor="FFD0D7E3")
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    titles = ["序号", "员工姓名", "月度计量计件工资(元)", "月度考勤计时工资(元)", "当月总工资(元)"]
    for col, title in enumerate(titles, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border

    all_names = sorted(set(metric_dict.keys()) | set(attend_dict.keys()))
    row = 2
    sum_metric = 0.0
    sum_attend = 0.0
    sum_all = 0.0

    for seq, name in enumerate(all_names):
        m_wage = round(metric_dict.get(name, 0.0), 2)
        a_wage = round(attend_dict.get(name, 0.0), 2)
        total_wage = m_wage + a_wage
        sum_metric += m_wage
        sum_attend += a_wage
        sum_all += total_wage

        ws.cell(row=row, column=1, value=seq+1).alignment = align_center
        ws.cell(row=row, column=2, value=name).font = font_normal
        c3 = ws.cell(row=row, column=3, value=m_wage)
        c3.number_format = "#,##0.00"
        c3.font = font_normal
        c4 = ws.cell(row=row, column=4, value=a_wage)
        c4.number_format = "#,##0.00"
        c4.font = font_normal
        c5 = ws.cell(row=row, column=5, value=total_wage)
        c5.number_format = "#,##0.00"
        c5.font = font_normal
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    total_row = row
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=round(sum_metric,2)).number_format = "#,##0.00"
    ws.cell(row=total_row, column=4, value=round(sum_attend,2)).number_format = "#,##0.00"
    ws.cell(row=total_row, column=5, value=round(sum_all,2)).number_format = "#,##0.00"
    for c in range(1,6):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = fill_total
        cell.alignment = align_center
        cell.border = border

    auto_fit_column(ws)
    ws.freeze_panes = "A2"
    wb.save(save_path)
    return True

# ===================== 主界面（使用 Text 组件） =====================
if __name__ == "__main__":
    os.environ["TK_SILENCE_DEPRECATION"] = "1"
    metric_files = []
    attend_files = []

    root = tk.Tk()
    root.title("车间工资核算工具 (Mac稳定版)")
    root.geometry("720x750")
    root.lift()
    root.focus_force()

    progress_text = tk.StringVar(value="等待操作")

    tk.Label(root, text="月度工资自动核算", font=("",14,"bold")).pack(pady=10)

    # ---------- 计量文件显示（Text） ----------
    tk.Label(root, text="【计量工资文件列表】", font=("",11,"bold")).pack()
    frame_metric = tk.Frame(root)
    frame_metric.pack(fill=tk.BOTH, expand=True, pady=2, padx=10)
    scroll_metric = tk.Scrollbar(frame_metric)
    scroll_metric.pack(side=tk.RIGHT, fill=tk.Y)
    text_metric = tk.Text(frame_metric, height=6, yscrollcommand=scroll_metric.set, wrap=tk.NONE)
    text_metric.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scroll_metric.config(command=text_metric.yview)
    text_metric.insert(tk.END, "暂无计量文件\n")

    # ---------- 考勤文件显示（Text） ----------
    tk.Label(root, text="【考勤数据文件列表】", font=("",11,"bold")).pack()
    frame_attend = tk.Frame(root)
    frame_attend.pack(fill=tk.BOTH, expand=True, pady=2, padx=10)
    scroll_attend = tk.Scrollbar(frame_attend)
    scroll_attend.pack(side=tk.RIGHT, fill=tk.Y)
    text_attend = tk.Text(frame_attend, height=6, yscrollcommand=scroll_attend.set, wrap=tk.NONE)
    text_attend.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scroll_attend.config(command=text_attend.yview)
    text_attend.insert(tk.END, "暂无考勤文件\n")

    # 更新函数
    def update_metric_display():
        text_metric.delete(1.0, tk.END)
        if not metric_files:
            text_metric.insert(tk.END, "暂无计量文件\n")
        else:
            for p in metric_files:
                name = os.path.basename(p) or p
                text_metric.insert(tk.END, name + "\n")
        # 终端输出调试
        print("计量文件列表：", [os.path.basename(p) for p in metric_files])

    def update_attend_display():
        text_attend.delete(1.0, tk.END)
        if not attend_files:
            text_attend.insert(tk.END, "暂无考勤文件\n")
        else:
            for p in attend_files:
                name = os.path.basename(p) or p
                text_attend.insert(tk.END, name + "\n")
        print("考勤文件列表：", [os.path.basename(p) for p in attend_files])

    # 选择文件
    def select_metric():
        global metric_files
        root.lift()
        root.focus_force()
        paths = filedialog.askopenfilenames(
            parent=root,
            title="批量选择计量Excel",
            filetypes=[("Excel表格", "*.xlsx")]
        )
        if not paths:
            return
        metric_files = list(paths)
        update_metric_display()
        label_metric_count.config(text=f"计量文件总数：{len(metric_files)}")

    def select_attend():
        global attend_files
        root.lift()
        root.focus_force()
        paths = filedialog.askopenfilenames(
            parent=root,
            title="批量选择考勤Excel",
            filetypes=[("Excel表格", "*.xlsx")]
        )
        if not paths:
            return
        attend_files = list(paths)
        update_attend_display()
        label_attend_count.config(text=f"考勤文件总数：{len(attend_files)}")

    # 清空
    def clear_metric():
        global metric_files
        metric_files = []
        update_metric_display()
        label_metric_count.config(text="计量文件总数：0")

    def clear_attend():
        global attend_files
        attend_files = []
        update_attend_display()
        label_attend_count.config(text="考勤文件总数：0")

    # ---------- 计算函数（保持原有逻辑） ----------
    def calc_only_metric():
        global metric_files
        if not metric_files:
            messagebox.showinfo("提示", "请先上传计量工资Excel文件！")
            return
        progress_text.set("正在单独计算计件工资...")
        metric_data = batch_metric(metric_files, progress_text)
        empty_attend = {}
        save_path = filedialog.asksaveasfilename(
            parent=root,
            defaultextension=".xlsx",
            initialfile="仅计件工资汇总.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            progress_text.set("操作取消")
            return
        try:
            export_summary(metric_data, empty_attend, save_path)
            messagebox.showinfo("完成", "仅计件工资表格导出成功！")
            progress_text.set("单独计件计算完成")
        except Exception as e:
            messagebox.showerror("错误", str(e))
            progress_text.set("导出失败")

    def calc_only_attend():
        global attend_files
        if not attend_files:
            messagebox.showinfo("提示", "请先上传考勤Excel文件！")
            return
        progress_text.set("正在单独计算考勤工资...")
        attend_data = batch_attend(attend_files, progress_text)
        empty_metric = {}
        save_path = filedialog.asksaveasfilename(
            parent=root,
            defaultextension=".xlsx",
            initialfile="仅考勤工资汇总.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            progress_text.set("操作取消")
            return
        try:
            export_summary(empty_metric, attend_data, save_path)
            messagebox.showinfo("完成", "仅考勤工资表格导出成功！")
            progress_text.set("导出完成")
        except Exception as e:
            messagebox.showerror("错误", str(e))
            progress_text.set("导出失败")

    def calc_all():
        global metric_files, attend_files
        if not metric_files and not attend_files:
            messagebox.showinfo("提示", "请至少上传一类Excel文件！")
            return
        progress_text.set("正在合并计算全部工资...")
        metric_data = batch_metric(metric_files, progress_text) if metric_files else {}
        attend_data = batch_attend(attend_files, progress_text) if attend_files else {}
        save_path = filedialog.asksaveasfilename(
            parent=root,
            defaultextension=".xlsx",
            initialfile="计件+考勤总工资汇总.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            progress_text.set("操作取消")
            return
        try:
            export_summary(metric_data, attend_data, save_path)
            messagebox.showinfo("完成", "总工资表导出成功！")
            progress_text.set("合并计算完成")
        except Exception as e:
            messagebox.showerror("错误", str(e))
            progress_text.set("导出失败")

    # ---------- 按钮布局 ----------
    frame1 = tk.Frame(root)
    frame1.pack(pady=5)
    tk.Button(frame1, text="选择计量文件", command=select_metric).pack(side=tk.LEFT, padx=3)
    tk.Button(frame1, text="清空计量", command=clear_metric).pack(side=tk.LEFT, padx=3)
    tk.Button(frame1, text="仅算计件", command=calc_only_metric).pack(side=tk.LEFT, padx=3)
    label_metric_count = tk.Label(frame1, text="计量文件总数：0", padx=8)
    label_metric_count.pack(side=tk.LEFT)

    frame2 = tk.Frame(root)
    frame2.pack(pady=5)
    tk.Button(frame2, text="选择考勤文件", command=select_attend).pack(side=tk.LEFT, padx=3)
    tk.Button(frame2, text="清空考勤", command=clear_attend).pack(side=tk.LEFT, padx=3)
    tk.Button(frame2, text="仅算考勤", command=calc_only_attend).pack(side=tk.LEFT, padx=3)
    label_attend_count = tk.Label(frame2, text="考勤文件总数：0", padx=8)
    label_attend_count.pack(side=tk.LEFT)

    tk.Label(root, textvariable=progress_text, fg="#0055cc").pack(pady=10)

    tk.Button(root, text="合并计算总工资", width=32, height=2, command=calc_all).pack(pady=6)
    tk.Button(root, text="退出程序", width=32, command=root.quit).pack()

    root.mainloop()