import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import openpyxl.utils

import os
import shutil
import tkinter as tk
from tkinter import ttk
import win32com.client
import ctypes
import time

# ---------- Enable High-DPI Awareness (Windows only) ----------
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

# ---------- Force-close Excel ----------
def close_excel_file(file_path):
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        for wb in excel.Workbooks:
            if os.path.abspath(wb.FullName) == file_path:
                wb.Close(SaveChanges=True)
                break
        excel.Quit()
    except Exception as e:
        print(f"Could not close Excel automatically: {e}")

# ---------- Create and Open Excel ----------
def create_excel(tables_data, gare_x="X", poste_y="Y"):
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    import os
    import shutil
    import time
    import win32com.client
    import ctypes
    from openpyxl.cell.cell import MergedCell

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AvancementEssais"

    start_row = 5
    start_col = 3
    gap_between_tables = 2
    current_col = start_col
    max_height = 0

    # Fills
    header_fill = PatternFill(start_color="DBDBDB", end_color="DBDBDB", fill_type="solid")
    top_rows_fill = PatternFill(start_color="bec1c6", end_color="bec1c6", fill_type="solid")
    full_green_fill = PatternFill(start_color="00c853", end_color="00c853", fill_type="solid")
    full_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    full_yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Border
    thin_black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    table_averages = []

    # -------------------- Create tables horizontally --------------------
    for t_index, (data_rows, cols, first_cell, values) in enumerate(tables_data):
        total_rows = data_rows + 2
        max_height = max(max_height, total_rows)

        # --- Calculate average of second column ---
        avg_percentage = ""
        avg_float = None
        if cols >= 2:
            perc_values = []
            for r in range(data_rows):
                try:
                    val = values[r][1]
                    if val != "":
                        val_clean = str(val).strip().replace("%", "")
                        perc_values.append(float(val_clean))
                except:
                    pass
            if perc_values:
                avg_value = sum(perc_values) / len(perc_values)
                avg_percentage = f"{round(avg_value, 2)}%"
                avg_float = avg_value
        table_averages.append((first_cell, avg_float))

        # --- Fill table cells ---
        for r in range(total_rows):
            for c in range(cols):
                cell_value = ""
                if r == 0:  # Header row
                    if c == 0:
                        cell_value = first_cell
                    elif c == 1:
                        cell_value = "Avancement & Réalisation"
                    elif c == 2:
                        cell_value = "Remarques"
                elif r == total_rows - 1:  # Total row
                    if c == 0:
                        cell_value = "Total"
                    elif c == 1 and avg_percentage:
                        cell_value = avg_percentage
                else:  # Data rows
                    try:
                        cell_value = values[r - 1][c]
                    except:
                        cell_value = ""

                cell = ws.cell(row=start_row + r, column=current_col + c, value=cell_value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_black_border

                # Coloring
                if r == 0:
                    cell.fill = header_fill
                elif r != total_rows - 1 and c == 1:
                    try:
                        val_float_cell = float(str(cell_value).strip().replace("%", ""))
                        if val_float_cell == 100:
                            cell.fill = full_green_fill
                        elif val_float_cell == 0:
                            cell.fill = full_red_fill
                        else:
                            cell.fill = full_yellow_fill
                    except:
                        cell.fill = full_yellow_fill
                elif r == total_rows - 1 and not (cols >= 3 and c == 2):
                    cell.fill = header_fill

        current_col += cols + gap_between_tables

    # -------------------- Top title rows --------------------
    merged_columns = 80  # Fixed number of columns to merge for the top title

    for row in [1, 2]:
        for col in range(1, merged_columns + 1):
            ws.cell(row=row, column=col).fill = top_rows_fill

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=merged_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Avancement Essais GARE {gare_x} POSTE {poste_y}"  # still variable
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True, size=14)

    # -------------------- Auto-adjust column widths --------------------
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column_cells:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                break
        if column_letter is None:
            continue
        for cell in column_cells:
            if not isinstance(cell, MergedCell) and cell.row > 2 and cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 3

    from openpyxl.chart import BarChart3D, PieChart3D, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.utils import get_column_letter

    # -------------------- Bar Chart Code --------------------
    if table_averages:

        # -------------------- Create / reuse hidden sheet --------------------
        if "ChartData" in wb.sheetnames:
            chart_ws = wb["ChartData"]
        else:
            chart_ws = wb.create_sheet("ChartData")
            chart_ws.sheet_state = 'hidden'

        # Write chart data as percentages (0–100)
        for idx, (name, avg_val) in enumerate(table_averages, start=1):
            chart_ws.cell(row=idx, column=1, value=name)
            safe_val = 0 if avg_val is None else round(avg_val, 2)  # <-- keep 0–100
            chart_ws.cell(row=idx, column=2, value=safe_val)

        data = Reference(chart_ws, min_col=2, min_row=1, max_row=len(table_averages))
        categories = Reference(chart_ws, min_col=1, min_row=1, max_row=len(table_averages))

        # -------------------- 3D Bar Chart --------------------
        bar_chart = BarChart3D()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "AVANCEMENT ESSAIS PAR TYPE D'INSTALATION DE SÉCURITÉ"
        bar_chart.add_data(data, titles_from_data=False)
        bar_chart.set_categories(categories)
        bar_chart.gapWidth = 0

        # Remove axis titles and grid lines
        bar_chart.y_axis.title = None
        bar_chart.x_axis.title = None
        bar_chart.y_axis.majorGridlines = None

        # Dark grey background
        dark_color = "595959"
        bar_chart.graphical_properties = GraphicalProperties(solidFill=dark_color)
        bar_chart.plot_area.graphical_properties = GraphicalProperties(solidFill=dark_color)

        # -------------------- Data Labels --------------------
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties

        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        bar_chart.dataLabels.number_format = '0%'
        bar_chart.dataLabels.showSerName = False
        bar_chart.dataLabels.showCatName = False


        # Increase font size and bold
        bar_chart.dataLabels.txPr = RichText(
            p=[
                Paragraph(
                    pPr=ParagraphProperties(
                        defRPr=CharacterProperties(
                            sz=1800,  # 18 pt, adjust bigger/smaller if needed
                            b=True
                        )
                    )
                )
            ]
        )

        # -------------------- Determine bar chart position --------------------
        if len(tables_data) >= 3:
            table_index = 1
        else:
            table_index = len(tables_data) - 1

        table_start_col = start_col + sum([t[1] + gap_between_tables for t in tables_data[:table_index]])
        table_cols = tables_data[table_index][1]

        chart_start_col = table_start_col + table_cols // 2
        chart_col_letter = get_column_letter(chart_start_col)

        max_rows = max([t[0] for t in tables_data])
        chart_start_row = start_row + max_rows + 3

        # -------------------- Add Bar Chart to worksheet --------------------
        ws.add_chart(
            bar_chart,
            f"{get_column_letter(chart_start_col)}{chart_start_row}"
        )

        # -------------------- 3D Pie Chart (Global Average) --------------------

        from openpyxl.chart import PieChart3D, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.layout import Layout, ManualLayout
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.utils import get_column_letter

        # ---------- Create / reuse hidden sheet ----------
        bar_chart = BarChart3D()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "AVANCEMENT ESSAIS PAR TYPE D'INSTALATION DE SÉCURITÉ"
        bar_chart.add_data(data, titles_from_data=False)
        bar_chart.set_categories(categories)
        bar_chart.gapWidth = 0

        # Remove axis titles and grid lines
        bar_chart.y_axis.title = None
        bar_chart.x_axis.title = None
        bar_chart.y_axis.majorGridlines = None

        # Dark grey background
        dark_color = "595959"
        bar_chart.graphical_properties = GraphicalProperties(solidFill=dark_color)
        bar_chart.plot_area.graphical_properties = GraphicalProperties(solidFill=dark_color)

        # Data Labels
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
        # -------------------- Data Labels --------------------
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        bar_chart.dataLabels.showPercent = True
        bar_chart.dataLabels.number_format = '0\\%'
        bar_chart.dataLabels.showSerName = False
        bar_chart.dataLabels.showCatName = False
        bar_chart.dataLabels.showLeaderLines = False
        bar_chart.dataLabels.showLegendKey = False
        bar_chart.dataLabels.txPr = RichText(
            p=[
                Paragraph(
                    pPr=ParagraphProperties(
                        defRPr=CharacterProperties(
                            sz=1200,  # 13 pt
                            b=True  # bold
                        )
                    )
                )
            ]
        )


        # Determine bar chart position
        if len(tables_data) >= 3:
            table_index = 1
        else:
            table_index = len(tables_data) - 1

        table_start_col = start_col + sum(
            [t[1] + gap_between_tables for t in tables_data[:table_index]]
        )
        table_cols = tables_data[table_index][1]

        chart_start_col = table_start_col + table_cols // 2

        max_rows = max([t[0] for t in tables_data])
        chart_start_row = start_row + max_rows + 3

        # Add Bar Chart to worksheet
        ws.add_chart(
            bar_chart,
            f"{get_column_letter(chart_start_col)}{chart_start_row}"
        )

        # -------------------- 3D Pie Chart (Global Average) --------------------

        # -------------------- Create / reuse hidden sheet --------------------
        if "PieChartData" in wb.sheetnames:
            pie_ws = wb["PieChartData"]
        else:
            pie_ws = wb.create_sheet("PieChartData")
            pie_ws.sheet_state = 'hidden'

        # -------------------- Calculate global average --------------------
        all_vals = [0 if avg is None else avg for _, avg in table_averages]
        global_avg = round(sum(all_vals) / len(all_vals), 2) if all_vals else 0
        remaining = 100 - global_avg

        # -------------------- Write chart data --------------------
        # Row 1 = empty header → prevents Excel from creating "Series 1"
        pie_ws.cell(row=1, column=2, value="")

        # Only the first slice will have a label
        category_name = f"Avancement Essais {gare_x} Poste {poste_y}"
        pie_ws.cell(row=2, column=1, value=category_name)
        pie_ws.cell(row=2, column=2, value=global_avg / 100)

        # Remaining slice has no category (label hidden)
        pie_ws.cell(row=3, column=1, value=None)
        pie_ws.cell(row=3, column=2, value=remaining / 100)

        pie_data = Reference(pie_ws, min_col=2, min_row=1, max_row=3)
        pie_categories = Reference(pie_ws, min_col=1, min_row=2, max_row=3)

        # -------------------- Create PieChart3D --------------------
        pie_chart = PieChart3D()
        pie_chart.title = "AVANCEMENT GLOBAL ESSAIS"

        # Use header row (empty) so Excel does NOT generate "Series 1"
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_categories)

        pie_chart.graphical_properties = GraphicalProperties(solidFill=dark_color)
        pie_chart.plot_area.graphical_properties = GraphicalProperties(solidFill=dark_color)

        # -------------------- Data Labels --------------------
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties

        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showVal = False
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showCatName = True
        pie_chart.dataLabels.showLeaderLines = False
        pie_chart.dataLabels.number_format = '0%'
        pie_chart.dataLabels.separator = " "  # removes semicolon
        pie_chart.dataLabels.showLegendKey = False  # hides the little square

        # Only show label for first slice and format it
        pie_chart.dataLabels.txPr = RichText(
            p=[
                Paragraph(
                    pPr=ParagraphProperties(
                        defRPr=CharacterProperties(
                            sz=1200,  # 13 pt
                            b=True  # bold
                        )
                    )
                )
            ]
        )

        # -------------------- Remove legend --------------------
        pie_chart.legend = None

        # -------------------- SPACING BETWEEN CHARTS --------------------
        spacing_columns = 6
        pie_start_col = chart_start_col + 15 + spacing_columns
        pie_start_row = chart_start_row

        # -------------------- Add chart to worksheet --------------------
        ws.add_chart(
            pie_chart,
            f"{get_column_letter(pie_start_col)}{pie_start_row}"
        )

        # Optional: if you want exact EMU-based positioning (advanced), you can convert offsets here

    # -------------------- Save and open Excel --------------------
    file_name = "AvancementEssais.xlsx"
    script_folder = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_folder, file_name)
    temp_path = os.path.join(script_folder, "temp.xlsx")

    try:
        # ---- Save to temp file first ----
        wb.save(temp_path)

        # ---- Close workbook automatically if already open ----
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            for wb_open in excel.Workbooks:
                if wb_open.FullName.lower() == file_path.lower():
                    wb_open.Close(SaveChanges=False)
                    break
        except:
            pass  # Excel not running

        # ---- Replace existing file safely ----
        if os.path.exists(file_path):
            os.remove(file_path)

        shutil.move(temp_path, file_path)

        # ---- Open Excel ----
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True

        # Open workbook in editable mode (not ReadOnly)
        wb_opened = excel.Workbooks.Open(
            file_path,
            ReadOnly=False,  # Ensure it's editable
            Editable=True,  # Explicitly set editable
            IgnoreReadOnlyRecommended=True  # Ignore any "Open as Read-Only Recommended"
        )

        # Optional: unprotect workbook (if protected)
        try:
            wb_opened.Unprotect()
        except:
            pass

        # ---- Bring to foreground ----
        try:
            hwnd = ctypes.windll.user32.GetForegroundWindow()
            ctypes.windll.user32.SetForegroundWindow(hwnd)
        except:
            pass

    except Exception as e:
        print(f"Error saving Excel file: {e}")


# ---------- Bring Tkinter window to front ----------
def bring_window_to_front(root):
    try:
        hwnd = root.winfo_id()
        ctypes.windll.user32.SetForegroundWindow(hwnd)
        root.lift()
    except Exception:
        root.lift()

# ---------- GUI ----------
class ModernExcelGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.overrideredirect(True)
        self.geometry("1000x700")
        self.configure(bg="#161c29")
        self.resizable(False, False)

        self.table_frames = []
        self._offsetx = 0
        self._offsety = 0
        self.is_maximized = False
        self._previous_geometry = self.geometry()

        self.create_header()
        self.create_controls()
        self.create_buttons()
        self.create_scrollable_area()

    def _on_mousewheel(self, event):
        # Shift + wheel → horizontal scroll
        if event.state & 0x1:
            self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
        else:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ---------- Generate boxes ----------
    def generate_table_boxes(self, frame):
        for widget in frame.box_container.winfo_children():
            widget.destroy()

        rows = frame.rows_var.get()
        cols = frame.cols_var.get()

        if rows < 1 or cols < 1:
            return

        frame.box_entries = []
        for r in range(rows):
            row_entries = []
            for c in range(cols):
                entry = ttk.Entry(frame.box_container, width=10)
                entry.grid(row=r, column=c, padx=2, pady=2)
                row_entries.append(entry)
            frame.box_entries.append(row_entries)

        for c in range(cols):
            frame.box_container.grid_columnconfigure(c, weight=1)

    # ---------- Header ----------
    def create_header(self):
        header = tk.Frame(self, bg="#ffc9a5", height=80)
        header.pack(fill="x", side="top")
        header.bind("<Button-1>", self.start_move)
        header.bind("<B1-Motion>", self.do_move)

        tk.Label(header, text="Avancement Essais", bg="#ffc9a5",
                 fg="#161c29", font=("Helvetica", 20, "bold")).pack(side="left", padx=25, pady=15)

        btn_frame = tk.Frame(header, bg="#ffc9a5")
        btn_frame.pack(side="right", padx=15, pady=15)

        def create_circle(color, command):
            canvas = tk.Canvas(btn_frame, width=25, height=25,
                               bg="#ffc9a5", highlightthickness=0)
            canvas.create_oval(2, 2, 23, 23,
                               fill=color, outline=color, width=1)
            canvas.pack(side="right", padx=6)
            canvas.bind("<Button-1>", lambda e: command())
            def on_enter(e): canvas.itemconfig(1, outline="#353333")
            def on_leave(e): canvas.itemconfig(1, outline=color)
            canvas.bind("<Enter>", on_enter)
            canvas.bind("<Leave>", on_leave)
            return canvas

        create_circle("#f86402", self.close_window)
        create_circle("#00c853", self.toggle_maximize)
        create_circle("#ffc900", self.minimize_window)

    def start_move(self, event):
        self._offsetx = event.x
        self._offsety = event.y

    def do_move(self, event):
        if not self.is_maximized:
            x = self.winfo_pointerx() - self._offsetx
            y = self.winfo_pointery() - self._offsety
            self.geometry(f"+{x}+{y}")

    def close_window(self):
        self.destroy()

    def minimize_window(self):
        self.overrideredirect(False)
        self.iconify()
        self.after(100, self._check_restore)

    def _check_restore(self):
        if self.state() == "normal":
            self.overrideredirect(True)
            self.lift()
        else:
            self.after(100, self._check_restore)

    def toggle_maximize(self):
        if not self.is_maximized:
            self._previous_geometry = self.geometry()
            self.geometry(f"{self.winfo_screenwidth()}x{self.winfo_screenheight()}+0+0")
            self.is_maximized = True
        else:
            self.geometry(self._previous_geometry)
            self.is_maximized = False

    # ---------- Controls ----------
    def create_controls(self):
        self.controls = tk.Frame(self, bg="#161c29", pady=12)
        self.controls.pack(fill="x", padx=35, pady=12)

        tk.Label(self.controls, text="Nombre de tables:", bg="#161c29",
                 fg="#ffc9a5", font=("Helvetica", 16)).grid(row=0, column=0, padx=8, pady=6, sticky="w")
        self.num_tables_var = tk.IntVar(value=1)
        ttk.Entry(self.controls, textvariable=self.num_tables_var,
                  width=6, font=("Helvetica", 14)).grid(row=0, column=1, padx=8, pady=6, sticky="w")

        tk.Label(self.controls, text="GARE X:", bg="#161c29",
                 fg="#ffc9a5", font=("Helvetica", 16)).grid(row=0, column=2, padx=8, pady=6, sticky="w")
        self.gare_x_var = tk.StringVar(value="X")
        ttk.Entry(self.controls, textvariable=self.gare_x_var,
                  width=6, font=("Helvetica", 14)).grid(row=0, column=3, padx=8, pady=6)

        tk.Label(self.controls, text="POSTE Y:", bg="#161c29",
                 fg="#ffc9a5", font=("Helvetica", 16)).grid(row=0, column=4, padx=8, pady=6, sticky="w")
        self.poste_y_var = tk.StringVar(value="Y")
        ttk.Entry(self.controls, textvariable=self.poste_y_var,
                  width=6, font=("Helvetica", 14)).grid(row=0, column=5, padx=8, pady=6)

    # ---------- Buttons ----------
    def create_buttons(self):
        self.btn_frame = tk.Frame(self, bg="#161c29")
        self.btn_frame.pack(fill="x", padx=35, pady=10)

        style_btn = {
            "font": ("Helvetica", 14, "bold"),
            "bg": "#ffc9a5",
            "fg": "#161c29",
            "activebackground": "#f86402",
            "activeforeground": "#161c29",
            "bd": 0,
            "padx": 18,
            "pady": 10
        }

        tk.Button(self.btn_frame,
                  text="Créer les entrées de tables",
                  command=self.create_table_inputs,
                  **style_btn).pack(side="left", padx=12)

        tk.Button(self.btn_frame,
                  text="Générer l'excel",
                  command=self.regenerate_excel,
                  **style_btn).pack(side="left", padx=12)

    # ---------- Scrollable area ----------
    def create_scrollable_area(self):
        self.container = tk.Frame(self, bg="#161c29")
        self.container.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.canvas = tk.Canvas(self.container, bg="#161c29", highlightthickness=0)
        self.scrollable_frame = tk.Frame(self.canvas, bg="#161c29")

        self.v_scroll = tk.Scrollbar(self.container, orient="vertical",
                                     command=self.canvas.yview)
        self.h_scroll = tk.Scrollbar(self.container, orient="horizontal",
                                     command=self.canvas.xview)

        self.canvas.configure(yscrollcommand=self.v_scroll.set,
                              xscrollcommand=self.h_scroll.set)

        self.v_scroll.pack(side="right", fill="y")
        self.h_scroll.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.canvas.create_window((0, 0),
                                  window=self.scrollable_frame,
                                  anchor="nw")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        # 🔥 Enable mouse wheel scrolling (only when cursor is inside canvas)
        self.canvas.bind("<Enter>", lambda e: self.canvas.bind_all("<MouseWheel>", self._on_mousewheel))
        self.canvas.bind("<Leave>", lambda e: self.canvas.unbind_all("<MouseWheel>"))

    # ---------- Table Inputs ----------
    def create_table_inputs(self):
        for frame in self.table_frames:
            frame.destroy()
        self.table_frames.clear()

        try:
            num = self.num_tables_var.get()
            if num < 1 or num > 50:
                return
        except:
            return

        for t in range(num):
            frame = tk.Frame(self.scrollable_frame,
                             bg="#353333", bd=0, relief="flat")
            frame.pack(fill="x", padx=10, pady=8, ipady=12)

            tk.Label(frame, text=f"Table {t + 1}", bg="#353333", fg="#ffc9a5",
                     font=("Helvetica", 14, "bold")).grid(row=0, column=0, padx=6, pady=6, sticky="w")
            tk.Label(frame, text="Lignes:", bg="#353333", fg="#ffc9a5",
                     font=("Helvetica", 12)).grid(row=1, column=0, padx=6, pady=6)
            tk.Label(frame, text="Colonnes:", bg="#353333", fg="#ffc9a5",
                     font=("Helvetica", 12)).grid(row=1, column=2, padx=6, pady=6)
            tk.Label(frame, text="Première cellule", bg="#353333", fg="#ffc9a5",
                     font=("Helvetica", 12)).grid(row=1, column=4, padx=6, pady=6)

            rows_var = tk.IntVar(value=10)
            cols_var = tk.IntVar(value=2)
            first_cell_var = tk.StringVar(value=f"Table{t + 1}_R1_C1")

            rows_entry = ttk.Entry(frame, textvariable=rows_var, width=6, font=("Helvetica", 12))
            rows_entry.grid(row=1, column=1, padx=6, pady=6)
            rows_entry.bind("<Return>", lambda e, f=frame: self.generate_table_boxes(f))

            cols_entry = ttk.Entry(frame, textvariable=cols_var, width=6, font=("Helvetica", 12))
            cols_entry.grid(row=1, column=3, padx=6, pady=6)
            cols_entry.bind("<Return>", lambda e, f=frame: self.generate_table_boxes(f))

            ttk.Entry(frame, textvariable=first_cell_var, width=18, font=("Helvetica", 12)).grid(row=1, column=5, padx=6, pady=6)

            frame.rows_var = rows_var
            frame.cols_var = cols_var
            frame.first_cell_var = first_cell_var

            frame.box_container = tk.Frame(frame, bg="#353333")
            frame.box_container.grid(row=2, column=0, columnspan=6, padx=6, pady=(10, 5), sticky="w")

            self.table_frames.append(frame)

    def regenerate_excel(self):
        tables_data = []

        for frame in self.table_frames:
            rows = frame.rows_var.get()
            cols = frame.cols_var.get()
            first_cell = frame.first_cell_var.get()
            if rows < 1 or cols < 1:
                continue

            values = []
            for row_entries in getattr(frame, "box_entries", []):
                values.append([e.get() for e in row_entries])

            tables_data.append((rows, cols, first_cell, values))

        create_excel(tables_data, self.gare_x_var.get(), self.poste_y_var.get())
        bring_window_to_front(self)

if __name__ == "__main__":
    app = ModernExcelGUI()
    app.mainloop()
