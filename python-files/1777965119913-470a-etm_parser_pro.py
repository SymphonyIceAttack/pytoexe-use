import time
import random
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

FILE_NAME = "материалы на Зенит.xlsx"
OUTPUT_FILE = "result.xlsx"
LOG_FILE = "log.txt"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
}

def log(text):
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(text + "\n")

def get_usd_rub():
    try:
        r = requests.get("https://www.cbr-xml-daily.ru/daily_json.js")
        return r.json()["Valute"]["USD"]["Value"]
    except:
        return 90  # fallback

USD_RUB = get_usd_rub()

def normalize_price(price_text):
    price_text = price_text.replace(" ", "").replace(",", ".")
    
    if "USD" in price_text:
        value = float(price_text.replace("USD", ""))
        return round(value * USD_RUB, 2)
    
    if "€" in price_text:
        value = float(price_text.replace("€", ""))
        return round(value * (USD_RUB * 1.1), 2)

    return float(''.join(filter(lambda x: x.isdigit() or x == '.', price_text)))

def find_product_link(article):
    url = f"https://www.etm.ru/search/?search={article}"
    
    r = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(r.text, "html.parser")

    items = soup.select(".e-search-result__item")

    for item in items:
        code = item.select_one(".e-product-card__code")
        link = item.select_one("a")

        if code and code.text.strip() == article:
            return "https://www.etm.ru" + link["href"]

    return None

def get_price_from_card(url):
    r = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(r.text, "html.parser")

    price = soup.select_one(".e-product-price__value")

    if price:
        return normalize_price(price.text.strip())

    return None

wb = load_workbook(FILE_NAME)
ws = wb.active

for i in range(1, 36):
    article = ws[f"B{i}"].value

    if not article:
        continue

    article = str(article).strip()
    print(f"[{i}] Поиск: {article}")
    log(f"Поиск: {article}")

    try:
        link = find_product_link(article)

        if not link:
            ws[f"F{i}"] = "нет на сайте"
            log(" -> не найден")
            continue

        price = get_price_from_card(link)

        if price:
            ws[f"D{i}"] = price
            log(f" -> цена: {price}")
        else:
            ws[f"F{i}"] = "нет цены"
            log(" -> нет цены")

    except Exception as e:
        ws[f"F{i}"] = "ошибка"
        log(f" -> ошибка: {e}")

    time.sleep(random.uniform(1.5, 3.5))

wb.save(OUTPUT_FILE)

print("Готово!")