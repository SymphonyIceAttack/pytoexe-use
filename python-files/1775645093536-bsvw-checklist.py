import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
import json
import os

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ChecklistApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Лист самопроверки конструктора")
        self.root.geometry("1000x700")
        self.root.minsize(800, 600)

        # Цветовая схема
        self.colors = {
            "primary": "#2c3e50",
            "secondary": "#e74c3c",
            "success": "#27ae60",
            "danger": "#e67e22",
            "bg": "#ecf0f1",
            "card": "#ffffff",
            "text": "#2c3e50",
            "text_light": "#7f8c8d",
            "header": "#2980b9"
        }
        self.root.configure(bg=self.colors["bg"])

        # Данные
        self.items = []               # список кортежей (тип, текст)
        self.excel_row_map = {}       # индекс пункта -> номер строки в Excel
        self.current = 0
        self.item_status = {}         # индекс -> отметка ("✓" или "Нет")
        self.state_file = "checklist_progress.json"
        self.excel_file_path = None
        self.mark_column = 'C'        # столбец для отметок (по умолчанию C)

        # UI элементы
        self.setup_ui()

        if EXCEL_AVAILABLE:
            self.ask_for_excel()
        else:
            messagebox.showwarning(
                "Библиотека не установлена",
                "openpyxl не найден. Установите: pip install openpyxl\nПрограмма будет работать без сохранения в Excel."
            )
            self.use_default_items()

        self.update_display()

    def setup_ui(self):
        # Основной контейнер
        self.main_frame = tk.Frame(self.root, bg=self.colors["bg"])
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Верхняя карточка
        self.top_card = tk.Frame(self.main_frame, bg=self.colors["card"], relief="flat")
        self.top_card.pack(fill=tk.X, pady=(0, 20))
        top_inner = tk.Frame(self.top_card, bg=self.colors["card"])
        top_inner.pack(fill=tk.X, padx=20, pady=15)

        self.section_var = tk.StringVar()
        self.section_label = tk.Label(
            top_inner, textvariable=self.section_var,
            font=("Segoe UI", 18, "bold"), fg=self.colors["primary"], bg=self.colors["card"]
        )
        self.section_label.pack(anchor="w")

        self.progress_var = tk.StringVar()
        self.progress_label = tk.Label(
            top_inner, textvariable=self.progress_var,
            font=("Segoe UI", 11), fg=self.colors["text_light"], bg=self.colors["card"]
        )
        self.progress_label.pack(anchor="w", pady=(5, 0))

        # Кнопки справа
        btn_frame = tk.Frame(top_inner, bg=self.colors["card"])
        btn_frame.pack(side=tk.RIGHT, anchor="n")

        self.save_btn = tk.Button(
            btn_frame, text="💾 Сохранить", font=("Segoe UI", 10),
            bg=self.colors["primary"], fg="white", relief="flat",
            command=self.save_state, padx=15, cursor="hand2"
        )
        self.save_btn.pack(side=tk.LEFT, padx=5)

        if EXCEL_AVAILABLE:
            self.excel_btn = tk.Button(
                btn_frame, text="📂 Выбрать Excel", font=("Segoe UI", 10),
                bg=self.colors["header"], fg="white", relief="flat",
                command=self.select_excel_file, padx=15, cursor="hand2"
            )
            self.excel_btn.pack(side=tk.LEFT, padx=5)

            self.excel_status = tk.Label(
                btn_frame, text="", font=("Segoe UI", 9),
                bg=self.colors["card"], fg=self.colors["success"]
            )
            self.excel_status.pack(side=tk.LEFT, padx=10)

        # Карточка с содержанием
        self.content_card = tk.Frame(self.main_frame, bg=self.colors["card"], relief="flat")
        self.content_card.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        content_inner = tk.Frame(self.content_card, bg=self.colors["card"])
        content_inner.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        self.content_var = tk.StringVar()
        self.content_label = tk.Label(
            content_inner, textvariable=self.content_var,
            font=("Segoe UI", 22), wraplength=800, justify="center",
            bg=self.colors["card"], fg=self.colors["text"]
        )
        self.content_label.pack(fill=tk.BOTH, expand=True)

        # Панель отметок
        marks_card = tk.Frame(self.main_frame, bg=self.colors["card"], relief="flat")
        marks_card.pack(fill=tk.X, pady=(0, 15))
        marks_inner = tk.Frame(marks_card, bg=self.colors["card"])
        marks_inner.pack(fill=tk.X, padx=20, pady=15)

        self.btn_na = tk.Button(
            marks_inner, text="✗ Нет / Недоступно", font=("Segoe UI", 13, "bold"),
            bg=self.colors["danger"], fg="white", relief="flat",
            command=lambda: self.mark_item("Нет"), cursor="hand2"
        )
        self.btn_na.pack(side=tk.LEFT, padx=(0, 10), expand=True, fill=tk.X, ipady=8)

        self.btn_done = tk.Button(
            marks_inner, text="✓ Выполнено", font=("Segoe UI", 13, "bold"),
            bg=self.colors["success"], fg="white", relief="flat",
            command=lambda: self.mark_item("✓"), cursor="hand2"
        )
        self.btn_done.pack(side=tk.RIGHT, expand=True, fill=tk.X, ipady=8)

        # Панель навигации
        nav_card = tk.Frame(self.main_frame, bg=self.colors["card"], relief="flat")
        nav_card.pack(fill=tk.X)
        nav_inner = tk.Frame(nav_card, bg=self.colors["card"])
        nav_inner.pack(fill=tk.X, padx=20, pady=15)

        self.prev_btn = tk.Button(
            nav_inner, text="←  Назад", font=("Segoe UI", 12),
            bg=self.colors["primary"], fg="white", relief="flat",
            command=self.prev, width=12, cursor="hand2"
        )
        self.prev_btn.pack(side=tk.LEFT)

        self.next_btn = tk.Button(
            nav_inner, text="Далее  →", font=("Segoe UI", 12),
            bg=self.colors["primary"], fg="white", relief="flat",
            command=self.next, width=12, cursor="hand2"
        )
        self.next_btn.pack(side=tk.RIGHT)

        # Горячие клавиши
        self.root.bind("<Left>", lambda e: self.prev())
        self.root.bind("<Right>", lambda e: self.next())
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def ask_for_excel(self):
        """При запуске предложить выбрать Excel-файл и столбец для отметок"""
        answer = messagebox.askyesno(
            "Выбор файла",
            "Хотите выбрать Excel-файл с чек-листом?\n(Да — выберите файл, Нет — будет использован встроенный список)"
        )
        if answer:
            self.select_excel_file()
        else:
            self.use_default_items()

    def use_default_items(self):
        """Резервный встроенный список (можно расширить при необходимости)"""
        # Здесь можно оставить полный список, но для краткости – базовый пример
        self.items = [
            ("ЗАГОЛОВОК", "1. Определение уровня \"чистого\" пола"),
            ("ПУНКТ", "1.1 От метровой отметки"),
            ("ПУНКТ", "1.2 По конвектору с уточнением, выставлен ли он в горизонт"),
            ("ЗАГОЛОВОК", "2. Определение уровня \"чистого\" потолка"),
            ("ПУНКТ", "2.1 От метровой отметки"),
            ("ПУНКТ", "2.2 Наличие светильников"),
        ]
        self.excel_row_map = {}
        self.current = 0
        self.item_status = {}
        self.load_state()

    def load_items_from_excel(self, filepath, mark_column):
        """
        Загружает пункты и заголовки из Excel-файла.
        Возвращает (items, row_map), где:
          items: список (тип, текст)
          row_map: dict {индекс_пункта: номер_строки}
        """
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        items = []
        row_map = {}
        index = 0  # общий индекс для всех элементов (и заголовков, и пунктов)

        for row in ws.iter_rows(min_row=1, max_col=2):
            cell_a = row[0]  # столбец A (номер)
            cell_b = row[1]  # столбец B (текст)
            val_a = cell_a.value
            val_b = cell_b.value

            if val_b is None or str(val_b).strip() == "":
                continue

            text_b = str(val_b).strip()
            item_type = None

            # Определяем тип строки
            if val_a is not None:
                a_str = str(val_a).strip()
                # Пытаемся преобразовать в число
                try:
                    num = float(a_str)
                    if '.' in a_str:
                        # Число с точкой -> пункт
                        item_type = "ПУНКТ"
                    else:
                        # Целое число -> заголовок раздела
                        item_type = "ЗАГОЛОВОК"
                except ValueError:
                    # Не число, но может быть номером вида "1.1" (если сохранено как строка)
                    if '.' in a_str and a_str.replace('.', '').isdigit():
                        item_type = "ПУНКТ"
                    else:
                        # Непонятное значение — пропускаем
                        continue
            else:
                # Если первый столбец пуст, но второй не пуст — считаем пунктом
                item_type = "ПУНКТ"

            if item_type is None:
                continue

            items.append((item_type, text_b))
            if item_type == "ПУНКТ":
                # Запоминаем строку, куда потом писать отметку
                row_map[len(items)-1] = cell_b.row
            index += 1

        wb.close()
        return items, row_map

    def select_excel_file(self):
        if not EXCEL_AVAILABLE:
            messagebox.showwarning("Внимание", "Установите библиотеку: pip install openpyxl")
            return

        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return

        # Запрашиваем столбец для отметок
        column = simpledialog.askstring(
            "Столбец для отметок",
            "Введите букву столбца (например, C, D, E):",
            initialvalue=self.mark_column
        )
        if not column:
            column = self.mark_column
        else:
            column = column.strip().upper()

        try:
            items, row_map = self.load_items_from_excel(path, column)
            if not items:
                messagebox.showerror(
                    "Ошибка",
                    "Не удалось найти пункты самопроверки в файле.\n"
                    "Убедитесь, что в столбце A указаны номера (1, 1.1, 2 и т.д.),\n"
                    "а в столбце B — описание."
                )
                return

            self.items = items
            self.excel_row_map = row_map
            self.excel_file_path = path
            self.mark_column = column
            self.current = 0
            self.item_status = {}
            self.save_state()

            self.excel_status.config(text=f"📎 {os.path.basename(path)} (столбец {column})", fg=self.colors["success"])
            пунктов = sum(1 for t, _ in items if t == "ПУНКТ")
            messagebox.showinfo("Файл загружен", f"Загружено разделов: {len(items)-пунктов}\nПунктов для проверки: {пунктов}")

            self.update_display()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл:\n{e}")

    def get_current_section(self):
        for i in range(self.current, -1, -1):
            if self.items[i][0] == "ЗАГОЛОВОК":
                return self.items[i][1]
        return "Раздел не указан"

    def update_display(self):
        if not self.items:
            return

        type_, text = self.items[self.current]
        self.section_var.set(f"📌 {self.get_current_section()}")
        self.content_var.set(text)

        if type_ == "ЗАГОЛОВОК":
            self.content_label.config(font=("Segoe UI", 26, "bold"), fg=self.colors["header"])
            self.btn_done.config(state=tk.DISABLED, bg="#bdc3c7", cursor="arrow")
            self.btn_na.config(state=tk.DISABLED, bg="#bdc3c7", cursor="arrow")
        else:
            self.content_label.config(font=("Segoe UI", 20), fg=self.colors["text"])
            self.btn_done.config(state=tk.NORMAL, cursor="hand2")
            self.btn_na.config(state=tk.NORMAL, cursor="hand2")

            status = self.item_status.get(self.current)
            if status == "✓":
                self.btn_done.config(bg=self.colors["success"])
                self.btn_na.config(bg="#bdc3c7", fg="white")
            elif status == "Нет":
                self.btn_na.config(bg=self.colors["danger"])
                self.btn_done.config(bg="#bdc3c7", fg="white")
            else:
                self.btn_done.config(bg=self.colors["success"])
                self.btn_na.config(bg=self.colors["danger"])

        total_items = sum(1 for t, _ in self.items if t == "ПУНКТ")
        done_items = sum(1 for v in self.item_status.values() if v == "✓")
        self.progress_var.set(f"Пункт {self.current + 1} из {len(self.items)}  |  ✅ Выполнено: {done_items} / {total_items}")
        self.root.title(f"Лист самопроверки — {self.current + 1}/{len(self.items)}")
        self.prev_btn.config(state=tk.NORMAL if self.current > 0 else tk.DISABLED)
        self.next_btn.config(state=tk.NORMAL if self.current < len(self.items) - 1 else tk.DISABLED)

    def mark_item(self, mark):
        if self.items[self.current][0] != "ПУНКТ":
            return
        self.item_status[self.current] = mark
        self.save_state()
        if self.excel_file_path and EXCEL_AVAILABLE:
            self.update_excel()
        self.update_display()
        self.root.after(300, self.next)

    def update_excel(self):
        """Записывает отметки в выбранный столбец Excel"""
        if not self.excel_file_path or not EXCEL_AVAILABLE:
            return
        try:
            wb = openpyxl.load_workbook(self.excel_file_path)
            ws = wb.active

            # Преобразуем букву столбца в номер (A->1, B->2, ...)
            col_idx = ord(self.mark_column.upper()) - ord('A') + 1

            for idx, mark in self.item_status.items():
                if idx in self.excel_row_map:
                    row_num = self.excel_row_map[idx]
                    ws.cell(row=row_num, column=col_idx, value=mark)

            wb.save(self.excel_file_path)
            self.excel_status.config(text=f"✅ {os.path.basename(self.excel_file_path)} (обновлено)")
            self.root.after(2000, lambda: self.excel_status.config(
                text=f"📎 {os.path.basename(self.excel_file_path)} (столбец {self.mark_column})"))
        except PermissionError:
            messagebox.showwarning("Файл заблокирован", "Закройте Excel-файл и попробуйте снова.")
        except Exception as e:
            messagebox.showerror("Ошибка Excel", f"Не удалось обновить файл:\n{e}")

    def load_state(self):
        if os.path.exists(self.state_file):
            try:
                with open(self.state_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if data.get("items_count") == len(self.items) and data.get("excel_file") == self.excel_file_path:
                    self.current = data.get("current", 0)
                    self.item_status = {int(k): v for k, v in data.get("status", {}).items()}
                else:
                    self.current = 0
                    self.item_status = {}
            except Exception:
                pass

    def save_state(self):
        try:
            data = {
                "current": self.current,
                "status": self.item_status,
                "items_count": len(self.items),
                "excel_file": self.excel_file_path,
                "mark_column": self.mark_column
            }
            with open(self.state_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def next(self):
        if self.current < len(self.items) - 1:
            self.current += 1
            self.update_display()

    def prev(self):
        if self.current > 0:
            self.current -= 1
            self.update_display()

    def on_close(self):
        self.save_state()
        self.root.destroy()


if __name__ == "__main__":
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    root = tk.Tk()
    app = ChecklistApp(root)
    root.mainloop()