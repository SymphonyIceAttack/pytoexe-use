"""
CBC SCHOOL MANAGEMENT SYSTEM - FULL CBC (PRIMARY & JSS) - FINAL
===============================================================
- Grades 1-6 Primary, 7-9 JSS
- Primary descriptive reports (no ranking)
- JSS reports with ranking and total points
- School logo on dashboard and PDFs
- Level selector filters all operations
- Add Student grade dropdown adapts to selected level
- All original buttons present and working
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import hashlib
import csv
import os
import shutil
from datetime import datetime
from fpdf import FPDF
import json
import secrets

# Optional Excel export
try:
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional PIL for images
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# ---------- CONFIGURATION ----------
CONFIG_FILE = 'school_config.json'

def load_config():
    default_config = {
        'school_name': 'My School',
        'logo_path': ''
    }
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except:
            return default_config
    return default_config

def save_config(config):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=4)

config = load_config()

# Add default colors to config if they don't exist (for backward compatibility)
if 'colors' not in config:
    config['colors'] = {
        'header_bg': 'lightgray',
        'header_fg': 'darkblue',
        'button_bg': 'green',
        'button_fg': 'white'
    }

# ---------- TOOLTIP CLASS ----------
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind('<Enter>', self.show_tip)
        widget.bind('<Leave>', self.hide_tip)

    def show_tip(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                         background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                         font=("Arial", "9", "normal"))
        label.pack()

    def hide_tip(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

# ---------- TREEVIEW SORTING ----------
def treeview_sort_column(tree, col, reverse):
    data = [(tree.set(child, col), child) for child in tree.get_children('')]
    data.sort(reverse=reverse)
    for index, (val, child) in enumerate(data):
        tree.move(child, '', index)
    tree.heading(col, command=lambda: treeview_sort_column(tree, col, not reverse))

# ---------- PASSWORD HELPER ----------
def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(8)
    hashed = hashlib.sha256((password + salt).encode()).hexdigest()
    return f"{salt}${hashed}"

def verify_password(stored, provided):
    salt, hashed = stored.split('$')
    return hashed == hashlib.sha256((provided + salt).encode()).hexdigest()

# ---------- DATABASE SETUP ----------
conn = sqlite3.connect('cbc_school.db')
c = conn.cursor()
c.execute("PRAGMA foreign_keys = ON")

c.execute("""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE,
    password TEXT,
    role TEXT,
    fullname TEXT
)""")

# Students table with school_level
c.execute("""
CREATE TABLE IF NOT EXISTS students (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    adm_no TEXT UNIQUE,
    name TEXT,
    grade INTEGER,
    stream TEXT,
    school_level TEXT DEFAULT 'jss'
)""")

c.execute("""
CREATE TABLE IF NOT EXISTS subjects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    grade INTEGER,
    UNIQUE(name, grade)
)""")

c.execute("""
CREATE TABLE IF NOT EXISTS terms (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    year INTEGER,
    active INTEGER DEFAULT 0
)""")

# Marks table with teacher_comments
c.execute("""
CREATE TABLE IF NOT EXISTS marks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    student_id INTEGER,
    subject_id INTEGER,
    term_id INTEGER,
    score REAL,
    indicator TEXT,
    points INTEGER,
    teacher_comments TEXT,
    FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE,
    FOREIGN KEY(subject_id) REFERENCES subjects(id) ON DELETE CASCADE,
    FOREIGN KEY(term_id) REFERENCES terms(id) ON DELETE CASCADE,
    UNIQUE(student_id, subject_id, term_id)
)""")

c.execute("""
CREATE TABLE IF NOT EXISTS archived_marks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    student_id INTEGER,
    subject_id INTEGER,
    term_id INTEGER,
    score REAL,
    indicator TEXT,
    points INTEGER,
    teacher_comments TEXT,
    archive_date TEXT
)""")

# Default admin
c.execute("SELECT * FROM users WHERE username='admin'")
if not c.fetchone():
    pwd = hash_password('admin123')
    c.execute("INSERT INTO users (username, password, role, fullname) VALUES (?,?,?,?)",
              ('admin', pwd, 'admin', 'Administrator'))

# Default subjects for primary (Grades 1-6) and JSS (Grades 7-9)
primary_subjects = [
    'English', 'Kiswahili', 'Mathematics', 
    'Integrated Science', 'Agriculture', 'Home Science',
    'Creative Arts', 'Physical Education', 'Religious Education',
    'Social Studies', 'Indigenous Language'
]
jss_subjects = [
    'English','Kiswahili','Mathematics','Integrated Science',
    'Pretechnical Studies','Agriculture & Nutrition','CRE',
    'Social Studies','Creative Arts & Sports'
]

for grade in range(1, 7):
    for s in primary_subjects:
        try:
            c.execute("INSERT INTO subjects (name, grade) VALUES (?,?)", (s, grade))
        except sqlite3.IntegrityError:
            pass

for grade in range(7, 10):
    for s in jss_subjects:
        try:
            c.execute("INSERT INTO subjects (name, grade) VALUES (?,?)", (s, grade))
        except sqlite3.IntegrityError:
            pass

# Default term
year = datetime.now().year
c.execute("SELECT * FROM terms WHERE name='Term 1' AND year=?", (year,))
if not c.fetchone():
    c.execute("INSERT INTO terms (name, year, active) VALUES (?,?,?)", ('Term 1', year, 1))

conn.commit()

# ---------- CBC GRADING ----------
def get_indicator(score):
    try:
        s = float(score)
        if s >= 90: return ('EE1',8)
        if s >= 75: return ('EE2',7)
        if s >= 58: return ('ME1',6)
        if s >= 41: return ('ME2',5)
        if s >= 31: return ('AE1',4)
        if s >= 21: return ('AE2',3)
        if s >= 11: return ('BE1',2)
        return ('BE2',1)
    except:
        return ('?',0)

def get_student_rank(student_id, term_id):
    """For JSS only; returns rank and total students in grade."""
    c.execute("""
        SELECT s.id, AVG(m.score) as avg_score
        FROM students s
        LEFT JOIN marks m ON s.id = m.student_id AND m.term_id = ?
        WHERE s.grade = (SELECT grade FROM students WHERE id = ?)
        GROUP BY s.id
        HAVING avg_score IS NOT NULL
        ORDER BY avg_score DESC
    """, (term_id, student_id))
    rankings = c.fetchall()
    for rank, (sid, _) in enumerate(rankings, start=1):
        if sid == student_id:
            return rank, len(rankings)
    return "N/A", 0

def center(win, w, h):
    win.update_idletasks()
    ws = win.winfo_screenwidth()
    hs = win.winfo_screenheight()
    x = (ws//2) - (w//2)
    y = (hs//2) - (h//2)
    win.geometry(f'{w}x{h}+{x}+{y}')

# ---------- VALIDATION HELPERS ----------
def validate_grade(grade_str):
    try:
        g = int(grade_str)
        if 1 <= g <= 9:
            return (True, g)
        else:
            return (False, "Grade must be between 1 and 9.")
    except ValueError:
        return (False, "Grade must be a number.")

def validate_score(score_str):
    try:
        s = float(score_str)
        if 0 <= s <= 100:
            return (True, s)
        else:
            return (False, "Score must be between 0 and 100.")
    except ValueError:
        return (False, "Score must be a number.")

def validate_year(year_str):
    try:
        y = int(year_str)
        if 2000 <= y <= 2100:
            return (True, y)
        else:
            return (False, "Year must be between 2000 and 2100.")
    except ValueError:
        return (False, "Year must be a number.")

# ---------- LOGIN ----------
class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("CBC School System")
        center(self.root, 400,300)
        self.root.resizable(False, False)
        self.show_login()
        self.root.mainloop()

    def show_login(self):
        tk.Label(self.root, text="CBC SCHOOL", font=('Arial',18, 'bold'), fg='darkblue').pack(pady=20)
        tk.Label(self.root, text="Username").pack()
        self.user = tk.Entry(self.root, font=('Arial',12))
        self.user.pack(pady=5)
        tk.Label(self.root, text="Password").pack()
        self.pwd = tk.Entry(self.root, font=('Arial',12), show="*")
        self.pwd.pack(pady=5)
        login_btn = tk.Button(self.root, text="Login", command=self.do_login,
                              bg=config['colors']['button_bg'], fg=config['colors']['button_fg'],
                              font=('Arial',12, 'bold'), width=15)
        login_btn.pack(pady=20)
        ToolTip(login_btn, "Click to login (Ctrl+Enter)")
        self.root.bind('<Return>', lambda e: self.do_login())
        self.user.focus_set()

    def do_login(self):
        u = self.user.get()
        p = self.pwd.get()
        c.execute("SELECT * FROM users WHERE username=?", (u,))
        user = c.fetchone()
        if user and verify_password(user[2], p):
            self.root.destroy()
            Dashboard(user)
        else:
            messagebox.showerror("Error", "Invalid username or password.")

# ---------- SCHOOL SETTINGS WINDOW (with logo) ----------
class SchoolSettingsWindow:
    def __init__(self, dashboard):
        self.dashboard = dashboard
        self.win = tk.Toplevel(dashboard.root)
        self.win.title("School Settings")
        self.win.geometry("500x400")
        center(self.win, 500, 400)
        self.win.resizable(False, False)

        tk.Label(self.win, text="EDIT SCHOOL SETTINGS", font=('Arial',16,'bold'), fg='darkblue').pack(pady=10)

        name_frame = tk.Frame(self.win)
        name_frame.pack(pady=5)
        tk.Label(name_frame, text="School Name:", font=('Arial',12)).pack(side='left', padx=10)
        self.school_name = tk.Entry(name_frame, font=('Arial',12), width=30)
        self.school_name.insert(0, config['school_name'])
        self.school_name.pack(side='left')

        logo_frame = tk.Frame(self.win)
        logo_frame.pack(pady=10)
        tk.Label(logo_frame, text="Logo (GIF/PNG/JPG):", font=('Arial',12)).grid(row=0, column=0, padx=5)
        self.logo_path = tk.Entry(logo_frame, font=('Arial',12), width=30)
        self.logo_path.insert(0, config.get('logo_path', ''))
        self.logo_path.grid(row=0, column=1, padx=5)
        btn_browse = tk.Button(logo_frame, text="Browse", command=self.browse_logo)
        btn_browse.grid(row=0, column=2, padx=5)

        self.preview_label = tk.Label(self.win, text="No preview", bg='gray', width=30, height=6)
        self.preview_label.pack(pady=5)
        self.update_preview()

        btn_save = tk.Button(self.win, text="SAVE", command=self.save_settings,
                             bg=config['colors']['button_bg'], fg=config['colors']['button_fg'],
                             font=('Arial',12,'bold'), width=10)
        btn_save.pack(pady=10)
        ToolTip(btn_save, "Save settings (Ctrl+Enter)")

        self.win.bind('<Return>', lambda e: self.save_settings())
        self.win.bind('<Escape>', lambda e: self.win.destroy())

    def browse_logo(self):
        filename = filedialog.askopenfilename(
            title="Select Logo",
            filetypes=[("Image files", "*.gif *.png *.jpg *.jpeg *.bmp *.ppm")]
        )
        if filename:
            self.logo_path.delete(0, tk.END)
            self.logo_path.insert(0, filename)
            self.update_preview()

    def update_preview(self):
        path = self.logo_path.get()
        if path and os.path.exists(path):
            try:
                if PIL_AVAILABLE:
                    pil_img = Image.open(path)
                    pil_img.thumbnail((100, 100))
                    self.tk_img = ImageTk.PhotoImage(pil_img)
                    self.preview_label.config(image=self.tk_img, text='', bg='white')
                else:
                    if path.lower().endswith('.gif'):
                        self.tk_img = tk.PhotoImage(file=path)
                        self.preview_label.config(image=self.tk_img, text='', bg='white')
                    else:
                        self.preview_label.config(image='', text="Preview only for GIF\n(install PIL for more formats)", bg='lightgray')
            except Exception as e:
                self.preview_label.config(image='', text=f"Error: {str(e)[:30]}", bg='gray')
        else:
            self.preview_label.config(image='', text="No preview", bg='gray')

    def save_settings(self):
        global config
        config['school_name'] = self.school_name.get()
        config['logo_path'] = self.logo_path.get()
        # Preserve colors if they exist
        if 'colors' not in config:
            config['colors'] = {
                'header_bg': 'lightgray',
                'header_fg': 'darkblue',
                'button_bg': 'green',
                'button_fg': 'white'
            }
        save_config(config)
        messagebox.showinfo("Success", "Settings updated!")
        self.win.destroy()
        self.dashboard.refresh_branding()

# ---------- DASHBOARD ----------
class Dashboard:
    def __init__(self, user):
        self.user = user
        self.root = tk.Tk()
        self.root.title(f"CBC School System - {user[4]}")
        self.root.geometry('1000x650')
        center(self.root, 1000, 650)

        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # File Menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Change Password", command=self.change_password, accelerator="Ctrl+P")
        file_menu.add_command(label="Logout", command=self.logout, accelerator="Ctrl+L")
        file_menu.add_command(label="Backup", command=self.backup)
        file_menu.add_command(label="Restore", command=self.restore)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit, accelerator="Ctrl+Q")

        # Students Menu
        students_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Students", menu=students_menu)
        students_menu.add_command(label="Add Student", command=self.add_student, accelerator="Ctrl+N")
        students_menu.add_command(label="Manage Students", command=self.manage_students)
        students_menu.add_command(label="Promote Students", command=self.promote_students)
        students_menu.add_command(label="Import CSV", command=self.import_students)
        students_menu.add_command(label="Export CSV", command=self.export_students)

        # Marks Menu
        marks_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Marks", menu=marks_menu)
        marks_menu.add_command(label="Enter Marks", command=self.enter_marks)
        marks_menu.add_command(label="View Marks", command=self.view_marks)
        marks_menu.add_command(label="Import Marks (CSV)", command=self.import_marks)
        marks_menu.add_command(label="Export Marks (CSV)", command=self.export_marks)
        marks_menu.add_command(label="Export Marks (Excel)", command=self.export_marks_excel)

        # Reports Menu
        reports_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Reports", menu=reports_menu)
        reports_menu.add_command(label="Report Card (PDF)", command=self.report_card, accelerator="Ctrl+R")
        reports_menu.add_command(label="Batch Print Report Cards", command=self.batch_report_cards)
        reports_menu.add_command(label="Class Performance", command=self.class_performance)
        reports_menu.add_command(label="Previous Results", command=self.previous_results)
        reports_menu.add_command(label="Grading Guide", command=self.grading_guide)

        # Admin Menu (admin only)
        if user[3] == 'admin':
            admin_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Admin", menu=admin_menu)
            admin_menu.add_command(label="Manage Subjects", command=self.manage_subjects)
            admin_menu.add_command(label="Manage Terms", command=self.manage_terms)
            admin_menu.add_command(label="Archive Term", command=self.archive_term)
            admin_menu.add_command(label="Manage Users", command=self.manage_users)
            admin_menu.add_separator()
            admin_menu.add_command(label="School Settings", command=self.school_settings)
            admin_menu.add_command(label="Appearance Settings", command=self.appearance_settings)

        # Header with logo and level selector
        self.header_frame = tk.Frame(self.root, height=80)
        self.header_frame.pack(fill='x', padx=10, pady=5)
        self.header_frame.pack_propagate(False)
        self.refresh_branding()

        # Main content frame
        self.main_content = tk.Frame(self.root)
        self.main_content.pack(fill='both', expand=True, padx=10, pady=5)

        # Welcome and user info
        self.welcome_frame = tk.Frame(self.main_content)
        self.welcome_frame.pack(pady=10)
        tk.Label(self.welcome_frame, text=f"Welcome {user[4]}", font=('Arial',20, 'bold'), fg='darkblue').pack()
        tk.Label(self.welcome_frame, text=f"Role: {user[3].title()}", font=('Arial',12)).pack()

        # Current term info
        term = get_active_term()
        if term:
            term_label = tk.Label(self.main_content, text=f"Current Term: {term[1]} {term[2]}", font=('Arial', 12, 'italic'))
            term_label.pack(pady=5)

        # Stats area
        self.stats_frame = tk.Frame(self.main_content)
        self.stats_frame.pack(pady=10, fill='x')
        self.refresh_stats()

        # Quick actions frame
        actions_frame = tk.LabelFrame(self.main_content, text="Quick Actions", font=('Arial', 12, 'bold'))
        actions_frame.pack(pady=10, fill='x')

        btn_add_student = tk.Button(actions_frame, text="Add Student", command=self.add_student,
                                    bg=config['colors']['button_bg'], fg=config['colors']['button_fg'], width=15)
        btn_add_student.pack(side='left', padx=10, pady=5)
        btn_enter_marks = tk.Button(actions_frame, text="Enter Marks", command=self.enter_marks,
                                    bg=config['colors']['button_bg'], fg=config['colors']['button_fg'], width=15)
        btn_enter_marks.pack(side='left', padx=10, pady=5)
        btn_report = tk.Button(actions_frame, text="Report Card", command=self.report_card,
                               bg=config['colors']['button_bg'], fg=config['colors']['button_fg'], width=15)
        btn_report.pack(side='left', padx=10, pady=5)
        btn_batch = tk.Button(actions_frame, text="Batch Print", command=self.batch_report_cards,
                              bg=config['colors']['button_bg'], fg=config['colors']['button_fg'], width=15)
        btn_batch.pack(side='left', padx=10, pady=5)

        # Keyboard shortcuts
        self.root.bind('<Control-p>', lambda e: self.change_password())
        self.root.bind('<Control-l>', lambda e: self.logout())
        self.root.bind('<Control-q>', lambda e: self.root.quit())
        self.root.bind('<Control-n>', lambda e: self.add_student())
        self.root.bind('<Control-r>', lambda e: self.report_card())

        self.root.mainloop()

    def load_logo_image(self, max_size=(80,80)):
        path = config.get('logo_path', '')
        if not path or not os.path.exists(path):
            return None
        try:
            if PIL_AVAILABLE:
                pil_img = Image.open(path)
                pil_img.thumbnail(max_size)
                return ImageTk.PhotoImage(pil_img)
            else:
                if path.lower().endswith('.gif'):
                    return tk.PhotoImage(file=path)
                else:
                    return None
        except Exception:
            return None

    def refresh_branding(self):
        for widget in self.header_frame.winfo_children():
            widget.destroy()

        colors = config.get('colors', {})
        header_bg = colors.get('header_bg', 'lightgray')
        header_fg = colors.get('header_fg', 'darkblue')
        
        self.header_frame.config(bg=header_bg)

        logo_img = self.load_logo_image()
        if logo_img:
            self.header_logo = logo_img
            lbl_logo = tk.Label(self.header_frame, image=logo_img, bg=header_bg)
            lbl_logo.pack(side='left', padx=10)

        lbl_name = tk.Label(self.header_frame, text=config['school_name'],
                            font=('Arial', 24, 'bold'),
                            bg=header_bg,
                            fg=header_fg)
        lbl_name.pack(side='left', expand=True)

        self.level_var = tk.StringVar(value='All')
        self.level_combo = ttk.Combobox(self.header_frame, textvariable=self.level_var,
                                        values=['All', 'Primary', 'JSS'],
                                        state='readonly', width=10)
        self.level_combo.pack(side='right', padx=10)
        self.level_combo.bind('<<ComboboxSelected>>', self.on_level_change)
        ToolTip(self.level_combo, "Filter students by level (affects all operations)")

    def on_level_change(self, event=None):
        self.refresh_stats()

    def refresh_stats(self):
        for widget in self.stats_frame.winfo_children():
            widget.destroy()

        c.execute("""
            SELECT grade, school_level, COUNT(*) 
            FROM students 
            GROUP BY grade, school_level
            ORDER BY grade
        """)
        data = c.fetchall()

        primary_counts = {}
        jss_counts = {}
        for grade, level, count in data:
            if level == 'primary':
                primary_counts[grade] = count
            else:
                jss_counts[grade] = count

        primary_frame = tk.LabelFrame(self.stats_frame, text="Primary (Grades 1-6)", font=('Arial', 12, 'bold'))
        primary_frame.pack(side='left', padx=10, fill='both', expand=True)

        jss_frame = tk.LabelFrame(self.stats_frame, text="Junior Secondary (Grades 7-9)", font=('Arial', 12, 'bold'))
        jss_frame.pack(side='left', padx=10, fill='both', expand=True)

        row = 0
        for g in range(1, 7):
            count = primary_counts.get(g, 0)
            tk.Label(primary_frame, text=f"Grade {g}: {count} students", font=('Arial', 11)).grid(row=row, column=0, sticky='w', padx=5, pady=2)
            row += 1
        total_primary = sum(primary_counts.values())
        tk.Label(primary_frame, text=f"Total Primary: {total_primary}", font=('Arial', 11, 'bold')).grid(row=row, column=0, sticky='w', padx=5, pady=5)

        row = 0
        for g in range(7, 10):
            count = jss_counts.get(g, 0)
            tk.Label(jss_frame, text=f"Grade {g}: {count} students", font=('Arial', 11)).grid(row=row, column=0, sticky='w', padx=5, pady=2)
            row += 1
        total_jss = sum(jss_counts.values())
        tk.Label(jss_frame, text=f"Total JSS: {total_jss}", font=('Arial', 11, 'bold')).grid(row=row, column=0, sticky='w', padx=5, pady=5)

    # ---------- USER ACCOUNT ----------
    def change_password(self):
        win = tk.Toplevel(self.root)
        win.title("Change Password")
        win.geometry("300x200")
        center(win, 300, 200)
        win.resizable(False, False)

        tk.Label(win, text="Current Password:").pack(pady=5)
        e_old = tk.Entry(win, show="*")
        e_old.pack(pady=5)

        tk.Label(win, text="New Password:").pack(pady=5)
        e_new = tk.Entry(win, show="*")
        e_new.pack(pady=5)

        tk.Label(win, text="Confirm New:").pack(pady=5)
        e_confirm = tk.Entry(win, show="*")
        e_confirm.pack(pady=5)

        def save():
            old = e_old.get()
            new = e_new.get()
            confirm = e_confirm.get()
            if not verify_password(self.user[2], old):
                messagebox.showerror("Error", "Current password incorrect.")
                return
            if new != confirm:
                messagebox.showerror("Error", "New passwords do not match.")
                return
            hashed = hash_password(new)
            c.execute("UPDATE users SET password=? WHERE id=?", (hashed, self.user[0]))
            conn.commit()
            messagebox.showinfo("Success", "Password changed.")
            win.destroy()

        btn_save = tk.Button(win, text="Change Password", command=save,
                             bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
        btn_save.pack(pady=10)
        ToolTip(btn_save, "Change your password")
        win.bind('<Return>', lambda e: save())
        win.bind('<Escape>', lambda e: win.destroy())

    # ---------- APPEARANCE SETTINGS ----------
    def appearance_settings(self):
        if self.user[3] != 'admin':
            messagebox.showerror("Error", "Admin access required.")
            return

        win = tk.Toplevel(self.root)
        win.title("Appearance Settings")
        win.geometry("400x300")
        center(win, 400, 300)
        win.resizable(False, False)

        colors = config.get('colors', {})

        tk.Label(win, text="Header Background Color:").pack(pady=5)
        e_hdr_bg = tk.Entry(win, width=30)
        e_hdr_bg.insert(0, colors.get('header_bg', 'lightgray'))
        e_hdr_bg.pack(pady=5)

        tk.Label(win, text="Header Text Color:").pack(pady=5)
        e_hdr_fg = tk.Entry(win, width=30)
        e_hdr_fg.insert(0, colors.get('header_fg', 'darkblue'))
        e_hdr_fg.pack(pady=5)

        tk.Label(win, text="Button Background Color:").pack(pady=5)
        e_btn_bg = tk.Entry(win, width=30)
        e_btn_bg.insert(0, colors.get('button_bg', 'green'))
        e_btn_bg.pack(pady=5)

        tk.Label(win, text="Button Text Color:").pack(pady=5)
        e_btn_fg = tk.Entry(win, width=30)
        e_btn_fg.insert(0, colors.get('button_fg', 'white'))
        e_btn_fg.pack(pady=5)

        def save():
            config['colors'] = {
                'header_bg': e_hdr_bg.get(),
                'header_fg': e_hdr_fg.get(),
                'button_bg': e_btn_bg.get(),
                'button_fg': e_btn_fg.get()
            }
            save_config(config)
            messagebox.showinfo("Success", "Appearance saved. Restart to see full effect.")
            self.refresh_branding()
            win.destroy()

        btn_save = tk.Button(win, text="Save Settings", command=save,
                             bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
        btn_save.pack(pady=20)
        ToolTip(btn_save, "Save appearance settings")
        win.bind('<Return>', lambda e: save())
        win.bind('<Escape>', lambda e: win.destroy())

    # ---------- OTHER METHODS ----------
    def school_settings(self):
        SchoolSettingsWindow(self)

    def logout(self):
        if messagebox.askyesno("Logout", "Are you sure?"):
            self.root.destroy()
            App()

    def backup(self):
        try:
            shutil.copyfile('cbc_school.db', f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db')
            messagebox.showinfo("Success", "Backup created.")
        except Exception as e:
            messagebox.showerror("Error", f"Backup failed: {e}")

    def restore(self):
        f = filedialog.askopenfilename(filetypes=[("DB files", "*.db")])
        if f:
            if messagebox.askyesno("Confirm", "This will overwrite current database. Continue?"):
                try:
                    shutil.copyfile(f, 'cbc_school.db')
                    messagebox.showinfo("Success", "Database restored. Please restart.")
                    self.root.quit()
                except Exception as e:
                    messagebox.showerror("Error", f"Restore failed: {e}")

    # ---------- BULK OPERATIONS ----------
    def promote_students(self):
        win = tk.Toplevel(self.root)
        win.title("Promote Students")
        win.geometry("400x200")
        center(win, 400, 200)
        win.resizable(False, False)

        tk.Label(win, text="PROMOTE STUDENTS", font=('Arial',16,'bold'), fg='darkblue').pack(pady=20)

        frame = tk.Frame(win)
        frame.pack(pady=10)
        tk.Label(frame, text="From Grade:").grid(row=0, column=0, padx=5, pady=5)
        from_grade = ttk.Combobox(frame, values=list(range(1,10)), state='readonly', width=5)
        from_grade.grid(row=0, column=1, padx=5)
        from_grade.set(7)

        tk.Label(frame, text="To Grade:").grid(row=0, column=2, padx=5, pady=5)
        to_grade = ttk.Combobox(frame, values=list(range(2,10)), state='readonly', width=5)
        to_grade.grid(row=0, column=3, padx=5)
        to_grade.set(8)

        def do_promote():
            fg = int(from_grade.get())
            tg = int(to_grade.get())
            if fg >= tg:
                messagebox.showerror("Error", "To grade must be higher than from grade.")
                return
            level_filter = self.level_var.get()
            if level_filter != 'All':
                c.execute("UPDATE students SET grade = ? WHERE grade = ? AND school_level = ?",
                          (tg, fg, level_filter.lower()))
            else:
                c.execute("UPDATE students SET grade = ? WHERE grade = ?", (tg, fg))
            conn.commit()
            messagebox.showinfo("Success", f"Students promoted to Grade {tg}.")
            self.refresh_stats()
            win.destroy()

        btn_promote = tk.Button(win, text="PROMOTE", command=do_promote,
                                bg=config['colors']['button_bg'], fg=config['colors']['button_fg'],
                                font=('Arial',12,'bold'), width=15)
        btn_promote.pack(pady=20)
        ToolTip(btn_promote, "Promote all students in selected grade (respects level filter)")

        win.bind('<Return>', lambda e: do_promote())
        win.bind('<Escape>', lambda e: win.destroy())

    # ---------- STUDENT MANAGEMENT ----------
    def add_student(self):
        win = tk.Toplevel(self.root)
        win.title("Add Student")
        win.geometry("400x400")
        center(win, 400, 400)
        win.resizable(False, False)

        tk.Label(win, text="ADD STUDENT", font=('Arial',16,'bold')).pack(pady=10)

        frame = tk.Frame(win)
        frame.pack(pady=10)

        tk.Label(frame, text="Admission No:").grid(row=0, column=0, pady=5, sticky='w')
        e_adm = tk.Entry(frame, width=25)
        e_adm.grid(row=0, column=1, pady=5)

        tk.Label(frame, text="Name:").grid(row=1, column=0, pady=5, sticky='w')
        e_name = tk.Entry(frame, width=25)
        e_name.grid(row=1, column=1, pady=5)

        tk.Label(frame, text="Grade:").grid(row=2, column=0, pady=5, sticky='w')
        current_level = self.level_var.get()
        if current_level == 'Primary':
            grade_values = list(range(1,7))
        elif current_level == 'JSS':
            grade_values = list(range(7,10))
        else:
            grade_values = list(range(1,10))
        grade_var = tk.StringVar()
        grade_cb = ttk.Combobox(frame, textvariable=grade_var, values=grade_values, state='readonly', width=23)
        grade_cb.grid(row=2, column=1, pady=5)
        if grade_values:
            grade_cb.set(grade_values[0])

        tk.Label(frame, text="Stream:").grid(row=3, column=0, pady=5, sticky='w')
        e_stream = tk.Entry(frame, width=25)
        e_stream.grid(row=3, column=1, pady=5)

        tk.Label(frame, text="School Level:").grid(row=4, column=0, pady=5, sticky='w')
        level_var = tk.StringVar()
        level_cb = ttk.Combobox(frame, textvariable=level_var, values=['primary', 'jss'], state='readonly')
        level_cb.grid(row=4, column=1, pady=5)
        if current_level != 'All':
            level_cb.set(current_level.lower())
        else:
            level_cb.set('primary')

        # Button frame
        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)
        save_btn = tk.Button(btn_frame, text="Save", command=lambda: save(),
                             bg=config['colors']['button_bg'], fg=config['colors']['button_fg'], width=10)
        save_btn.pack(side='left', padx=5)
        ToolTip(save_btn, "Save student (Ctrl+Enter)")
        cancel_btn = tk.Button(btn_frame, text="Cancel", command=win.destroy, bg='red', fg='white', width=10)
        cancel_btn.pack(side='left', padx=5)
        ToolTip(cancel_btn, "Close window (Esc)")

        def save():
            adm = e_adm.get().strip()
            name = e_name.get().strip()
            grade_str = grade_var.get()
            stream = e_stream.get().strip()
            level = level_var.get()

            if not adm:
                messagebox.showerror("Error", "Admission number is required.")
                return
            if not name:
                messagebox.showerror("Error", "Student name is required.")
                return
            if not grade_str:
                messagebox.showerror("Error", "Please select a grade.")
                return
            grade_val = int(grade_str)
            if grade_val <= 6 and level != 'primary':
                if not messagebox.askyesno("Confirm", f"Grade {grade_val} is primary. Set level to primary?"):
                    return
                level = 'primary'
            elif grade_val >= 7 and level != 'jss':
                if not messagebox.askyesno("Confirm", f"Grade {grade_val} is JSS. Set level to jss?"):
                    return
                level = 'jss'

            try:
                c.execute("INSERT INTO students (adm_no, name, grade, stream, school_level) VALUES (?,?,?,?,?)",
                          (adm, name, grade_val, stream, level))
                conn.commit()
                messagebox.showinfo("Success", "Student added.")
                self.refresh_stats()
                win.destroy()
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "Admission number already exists.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

        win.bind('<Return>', lambda e: save())
        win.bind('<Escape>', lambda e: win.destroy())

    def manage_students(self):
        win = tk.Toplevel(self.root)
        win.title("Manage Students")
        win.geometry("1000x550")
        center(win, 1000, 550)
        win.resizable(True, True)

        tree_frame = tk.Frame(win)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

        scrollbar = tk.Scrollbar(tree_frame)
        scrollbar.pack(side='right', fill='y')

        tree = ttk.Treeview(tree_frame, columns=('ID','Adm','Name','Grade','Stream','Level'),
                            show='headings', yscrollcommand=scrollbar.set)
        tree.heading('ID', text='ID', command=lambda: treeview_sort_column(tree, 'ID', False))
        tree.heading('Adm', text='Adm No', command=lambda: treeview_sort_column(tree, 'Adm', False))
        tree.heading('Name', text='Name', command=lambda: treeview_sort_column(tree, 'Name', False))
        tree.heading('Grade', text='Grade', command=lambda: treeview_sort_column(tree, 'Grade', False))
        tree.heading('Stream', text='Stream', command=lambda: treeview_sort_column(tree, 'Stream', False))
        tree.heading('Level', text='Level', command=lambda: treeview_sort_column(tree, 'Level', False))

        tree.column('ID', width=50)
        tree.column('Adm', width=100)
        tree.column('Name', width=250)
        tree.column('Grade', width=60)
        tree.column('Stream', width=100)
        tree.column('Level', width=80)
        tree.pack(fill='both', expand=True)

        scrollbar.config(command=tree.yview)

        tree.tag_configure('odd', background='#f0f0f0')
        tree.tag_configure('even', background='#ffffff')

        def load_data():
            for row in tree.get_children():
                tree.delete(row)
            level_filter = self.level_var.get()
            query = "SELECT id, adm_no, name, grade, stream, school_level FROM students"
            params = []
            if level_filter != 'All':
                query += " WHERE school_level=?"
                params.append(level_filter.lower())
            query += " ORDER BY grade, name"
            c.execute(query, params)
            for i, row in enumerate(c.fetchall()):
                tag = 'even' if i % 2 == 0 else 'odd'
                tree.insert('', 'end', values=row, tags=(tag,))

        load_data()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=5)

        def edit_student():
            selected = tree.selection()
            if not selected:
                messagebox.showerror("Error", "Select a student first.")
                return

            item = tree.item(selected[0])
            sid, adm, name, grade, stream, level = item['values']

            edit_win = tk.Toplevel(win)
            edit_win.title("Edit Student")
            edit_win.geometry("400x400")
            center(edit_win, 400, 400)

            tk.Label(edit_win, text="EDIT STUDENT", font=('Arial',14,'bold')).pack(pady=10)

            frame = tk.Frame(edit_win)
            frame.pack(pady=10)

            tk.Label(frame, text="Admission No:").grid(row=0, column=0, pady=5, sticky='w')
            e_adm = tk.Entry(frame, width=25)
            e_adm.insert(0, adm)
            e_adm.config(state='readonly')
            e_adm.grid(row=0, column=1, pady=5)

            tk.Label(frame, text="Name:").grid(row=1, column=0, pady=5, sticky='w')
            e_name = tk.Entry(frame, width=25)
            e_name.insert(0, name)
            e_name.grid(row=1, column=1, pady=5)

            tk.Label(frame, text="Grade:").grid(row=2, column=0, pady=5, sticky='w')
            e_grade = tk.Entry(frame, width=25)
            e_grade.insert(0, grade)
            e_grade.grid(row=2, column=1, pady=5)

            tk.Label(frame, text="Stream:").grid(row=3, column=0, pady=5, sticky='w')
            e_stream = tk.Entry(frame, width=25)
            e_stream.insert(0, stream)
            e_stream.grid(row=3, column=1, pady=5)

            tk.Label(frame, text="School Level:").grid(row=4, column=0, pady=5, sticky='w')
            level_var = tk.StringVar()
            level_cb = ttk.Combobox(frame, textvariable=level_var, values=['primary', 'jss'], state='readonly')
            level_cb.grid(row=4, column=1, pady=5)
            level_cb.set(level)

            def save_edit():
                new_name = e_name.get().strip()
                new_grade_str = e_grade.get().strip()
                new_stream = e_stream.get().strip()
                new_level = level_var.get()

                if not new_name:
                    messagebox.showerror("Error", "Name cannot be empty.")
                    return
                try:
                    new_grade = int(new_grade_str)
                    if new_grade < 1 or new_grade > 9:
                        messagebox.showerror("Error", "Grade must be between 1 and 9.")
                        return
                except ValueError:
                    messagebox.showerror("Error", "Grade must be a number.")
                    return

                if new_grade <= 6 and new_level != 'primary':
                    if not messagebox.askyesno("Confirm", f"Grade {new_grade} is primary. Set level to primary?"):
                        return
                    new_level = 'primary'
                elif new_grade >= 7 and new_level != 'jss':
                    if not messagebox.askyesno("Confirm", f"Grade {new_grade} is JSS. Set level to jss?"):
                        return
                    new_level = 'jss'

                try:
                    c.execute("""
                        UPDATE students 
                        SET name=?, grade=?, stream=?, school_level=?
                        WHERE id=?
                    """, (new_name, new_grade, new_stream, new_level, sid))
                    conn.commit()
                    messagebox.showinfo("Success", "Student updated.")
                    edit_win.destroy()
                    load_data()
                    self.refresh_stats()
                except Exception as e:
                    messagebox.showerror("Error", str(e))

            btn_frame = tk.Frame(edit_win)
            btn_frame.pack(pady=10)
            save_btn = tk.Button(btn_frame, text="Save", command=save_edit,
                                 bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
            save_btn.pack(side='left', padx=5)
            ToolTip(save_btn, "Save changes (Ctrl+Enter)")
            cancel_btn = tk.Button(btn_frame, text="Cancel", command=edit_win.destroy, bg='red', fg='white')
            cancel_btn.pack(side='left', padx=5)
            ToolTip(cancel_btn, "Close window (Esc)")
            edit_win.bind('<Return>', lambda e: save_edit())
            edit_win.bind('<Escape>', lambda e: edit_win.destroy())

        def delete_student():
            selected = tree.selection()
            if not selected:
                messagebox.showerror("Error", "Select a student first.")
                return

            if messagebox.askyesno("Confirm", "Delete this student? All marks will also be deleted."):
                item = tree.item(selected[0])
                sid = item['values'][0]
                c.execute("DELETE FROM students WHERE id=?", (sid,))
                conn.commit()
                load_data()
                self.refresh_stats()

        def delete_all_marks():
            selected = tree.selection()
            if not selected:
                messagebox.showerror("Error", "Select a student first.")
                return
            if messagebox.askyesno("Confirm", "Delete ALL marks for this student? This cannot be undone."):
                item = tree.item(selected[0])
                sid = item['values'][0]
                c.execute("DELETE FROM marks WHERE student_id=?", (sid,))
                conn.commit()
                messagebox.showinfo("Success", "All marks deleted for this student.")

        edit_btn = tk.Button(btn_frame, text="Edit Selected", command=edit_student,
                             bg='orange', fg='white', width=15)
        edit_btn.pack(side='left', padx=5)
        ToolTip(edit_btn, "Edit selected student")
        del_btn = tk.Button(btn_frame, text="Delete Selected", command=delete_student,
                            bg='red', fg='white', width=15)
        del_btn.pack(side='left', padx=5)
        ToolTip(del_btn, "Delete selected student")
        delm_btn = tk.Button(btn_frame, text="Delete All Marks", command=delete_all_marks,
                              bg='purple', fg='white', width=15)
        delm_btn.pack(side='left', padx=5)
        ToolTip(delm_btn, "Delete all marks for selected student")
        ref_btn = tk.Button(btn_frame, text="Refresh", command=load_data,
                            bg='blue', fg='white', width=15)
        ref_btn.pack(side='left', padx=5)
        ToolTip(ref_btn, "Refresh list")
        close_btn = tk.Button(btn_frame, text="Close", command=win.destroy,
                              bg='gray', fg='white', width=15)
        close_btn.pack(side='left', padx=5)
        ToolTip(close_btn, "Close window (Esc)")
        win.bind('<Escape>', lambda e: win.destroy())

    def import_students(self):
        f = filedialog.askopenfilename(filetypes=[("CSV", "*.csv")])
        if not f: return
        success = 0
        errors = []
        with open(f) as file:
            reader = csv.reader(file)
            header = next(reader, None)
            for row_num, row in enumerate(reader, start=2):
                if len(row) >= 4:
                    adm, name, grade_str, level = row[0], row[1], row[2], row[3] if len(row)>3 else 'primary'
                    try:
                        grade = int(grade_str)
                        if grade < 1 or grade > 9:
                            errors.append(f"Row {row_num}: Grade must be 1-9")
                            continue
                    except ValueError:
                        errors.append(f"Row {row_num}: Invalid grade")
                        continue
                    if level not in ['primary','jss']:
                        level = 'primary' if grade <=6 else 'jss'
                    try:
                        c.execute("INSERT INTO students (adm_no, name, grade, school_level) VALUES (?,?,?,?)",
                                  (adm, name, grade, level))
                        success += 1
                    except sqlite3.IntegrityError:
                        errors.append(f"Row {row_num}: Duplicate admission number {adm}")
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                else:
                    errors.append(f"Row {row_num}: Not enough columns")
            conn.commit()
        self.refresh_stats()
        if errors:
            msg = f"Imported {success} students, {len(errors)} errors.\n\nFirst 10 errors:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                msg += f"\n... and {len(errors)-10} more"
            messagebox.showwarning("Import", msg)
        else:
            messagebox.showinfo("Import", f"Imported {success} students successfully.")

    def export_students(self):
        f = filedialog.asksaveasfilename(defaultextension=".csv")
        if not f: return
        c.execute("SELECT adm_no, name, grade, school_level FROM students")
        rows = c.fetchall()
        with open(f, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['admission_no', 'name', 'grade', 'school_level'])
            writer.writerows(rows)
        messagebox.showinfo("Done", f"Exported {len(rows)} students.")

    # ---------- MARKS ENTRY ----------
    def enter_marks(self):
        win = tk.Toplevel(self.root)
        win.title("Enter Marks")
        win.geometry("850x650")
        center(win, 850, 650)
        win.resizable(True, True)

        tk.Label(win, text="Grade:").pack()
        grade_var = tk.StringVar()
        grade_cb = ttk.Combobox(win, textvariable=grade_var, values=list(range(1,10)), state='readonly')
        grade_cb.pack()
        grade_cb.set(7)
        ToolTip(grade_cb, "Select grade to filter students")

        tk.Label(win, text="Student:").pack()
        student_var = tk.StringVar()
        student_cb = ttk.Combobox(win, textvariable=student_var, state='readonly')
        student_cb.pack()
        ToolTip(student_cb, "Select student")

        tk.Label(win, text="Term:").pack()
        term_var = tk.StringVar()
        term_cb = ttk.Combobox(win, textvariable=term_var, state='readonly')
        c.execute("SELECT name, year FROM terms")
        terms = [f"{t[0]} {t[1]}" for t in c.fetchall()]
        term_cb['values'] = terms
        if terms:
            term_cb.set(terms[0])
        term_cb.pack()
        ToolTip(term_cb, "Select term")

        tree_frame = tk.Frame(win)
        tree_frame.pack(fill='both', expand=True, pady=10, padx=10)

        scrollbar = tk.Scrollbar(tree_frame)
        scrollbar.pack(side='right', fill='y')

        tree = ttk.Treeview(tree_frame, columns=('Subject', 'Score', 'Indicator', 'Points', 'Comments'),
                            show='headings', yscrollcommand=scrollbar.set)
        tree.heading('Subject', text='Subject')
        tree.heading('Score', text='Score')
        tree.heading('Indicator', text='Indicator')
        tree.heading('Points', text='Points')
        tree.heading('Comments', text='Teacher Comments')
        tree.column('Subject', width=200)
        tree.column('Score', width=60, anchor='center')
        tree.column('Indicator', width=80, anchor='center')
        tree.column('Points', width=60, anchor='center')
        tree.column('Comments', width=200)
        tree.pack(fill='both', expand=True)

        scrollbar.config(command=tree.yview)

        tree.tag_configure('odd', background='#f0f0f0')
        tree.tag_configure('even', background='#ffffff')

        def load_students():
            grade = grade_var.get()
            if grade:
                level_filter = self.level_var.get()
                query = "SELECT id, name FROM students WHERE grade=?"
                params = [int(grade)]
                if level_filter != 'All':
                    query += " AND school_level=?"
                    params.append(level_filter.lower())
                c.execute(query, params)
                students = [f"{s[1]} (ID:{s[0]})" for s in c.fetchall()]
                student_cb['values'] = students
                if students:
                    student_cb.set(students[0])
                    load_subjects()

        def load_subjects():
            for row in tree.get_children():
                tree.delete(row)
            if not student_var.get() or not term_var.get():
                return
            stu_id = int(student_var.get().split('ID:')[1][:-1])
            c.execute("SELECT grade FROM students WHERE id=?", (stu_id,))
            grade = c.fetchone()[0]
            c.execute("SELECT id, name FROM subjects WHERE grade=? ORDER BY name", (grade,))
            subjects = c.fetchall()
            tparts = term_var.get().split()
            tname = ' '.join(tparts[:-1])
            tyear = int(tparts[-1])
            c.execute("SELECT id FROM terms WHERE name=? AND year=?", (tname, tyear))
            term_id = c.fetchone()[0]
            c.execute("SELECT subject_id, score, indicator, points, teacher_comments FROM marks WHERE student_id=? AND term_id=?", (stu_id, term_id))
            marks = {m[0]: (m[1], m[2], m[3], m[4]) for m in c.fetchall()}
            for i, (sid, sname) in enumerate(subjects):
                tag = 'even' if i % 2 == 0 else 'odd'
                if sid in marks:
                    sc, ind, pts, comm = marks[sid]
                    tree.insert('', 'end', iid=sid, values=(sname, sc, ind, pts, comm or ''), tags=(tag,))
                else:
                    tree.insert('', 'end', iid=sid, values=(sname, '', '', '', ''), tags=(tag,))

        grade_cb.bind('<<ComboboxSelected>>', lambda e: load_students())
        student_cb.bind('<<ComboboxSelected>>', lambda e: load_subjects())
        term_cb.bind('<<ComboboxSelected>>', lambda e: load_subjects())

        def edit_score(event):
            item = tree.selection()[0]
            subj = tree.item(item, 'values')[0]
            current_score = tree.item(item, 'values')[1]
            current_comments = tree.item(item, 'values')[4]
            d = tk.Toplevel(win)
            d.title(f"Score for {subj}")
            d.geometry("300x200")
            center(d, 300, 200)

            tk.Label(d, text="Score (0-100):").pack(pady=5)
            e_score = tk.Entry(d)
            e_score.insert(0, current_score)
            e_score.pack(pady=5)

            tk.Label(d, text="Teacher Comments:").pack(pady=5)
            e_comments = tk.Entry(d, width=40)
            e_comments.insert(0, current_comments if current_comments else '')
            e_comments.pack(pady=5)

            def save():
                score_val = e_score.get().strip()
                comments_val = e_comments.get().strip()
                if score_val:
                    ok, score = validate_score(score_val)
                    if not ok:
                        messagebox.showerror("Error", score)
                        return
                    ind, pts = get_indicator(score)
                else:
                    score_val = None
                    ind, pts = '', ''
                tree.item(item, values=(subj, score_val, ind, pts, comments_val))
                d.destroy()

            save_btn = tk.Button(d, text="Save", command=save,
                                 bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
            save_btn.pack(pady=5)
            ToolTip(save_btn, "Save (Enter)")
            d.bind('<Return>', lambda e: save())
            d.bind('<Escape>', lambda e: d.destroy())

        tree.bind('<Double-1>', edit_score)

        def save_marks():
            if not student_var.get() or not term_var.get():
                messagebox.showerror("Error", "Please select student and term.")
                return
            stu_id = int(student_var.get().split('ID:')[1][:-1])
            tparts = term_var.get().split()
            tname = ' '.join(tparts[:-1])
            tyear = int(tparts[-1])
            c.execute("SELECT id FROM terms WHERE name=? AND year=?", (tname, tyear))
            term_id = c.fetchone()[0]
            saved = 0
            errors = 0
            for item in tree.get_children():
                vals = tree.item(item, 'values')
                sid = int(item)
                score_str = vals[1]
                comments = vals[4] if len(vals) > 4 else ''
                if score_str:
                    ok, score = validate_score(score_str)
                    if not ok:
                        messagebox.showerror("Error", f"Invalid score for {vals[0]}: {score}")
                        return
                    ind, pts = get_indicator(score)
                else:
                    score = None
                    ind, pts = '', 0
                try:
                    if score is not None:
                        c.execute("""
                            INSERT OR REPLACE INTO marks (student_id, subject_id, term_id, score, indicator, points, teacher_comments)
                            VALUES (?,?,?,?,?,?,?)
                        """, (stu_id, sid, term_id, score, ind, pts, comments))
                    else:
                        c.execute("DELETE FROM marks WHERE student_id=? AND subject_id=? AND term_id=?", (stu_id, sid, term_id))
                    saved += 1
                except Exception as e:
                    errors += 1
                    print(e)
            conn.commit()
            if messagebox.askyesno("Success", f"{saved} marks saved, {errors} errors.\nDo you want to close?"):
                win.destroy()
            else:
                load_subjects()

        save_btn = tk.Button(win, text="Save Marks", command=save_marks,
                             bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
        save_btn.pack(pady=10)
        ToolTip(save_btn, "Save all marks (Ctrl+Enter)")
        win.bind('<Control-Return>', lambda e: save_marks())
        win.bind('<Escape>', lambda e: win.destroy())
        load_students()

    def view_marks(self):
        win = tk.Toplevel(self.root)
        win.title("View Marks")
        win.geometry("900x550")
        center(win, 900, 550)
        win.resizable(True, True)

        sel_frame = tk.Frame(win)
        sel_frame.pack(pady=10)

        tk.Label(sel_frame, text="Student:").grid(row=0, column=0, padx=5)
        student_var = tk.StringVar()
        student_cb = ttk.Combobox(sel_frame, textvariable=student_var, state='readonly', width=30)
        student_cb.grid(row=0, column=1, padx=5)
        ToolTip(student_cb, "Select student")

        tk.Label(sel_frame, text="Term:").grid(row=0, column=2, padx=5)
        term_var = tk.StringVar()
        term_cb = ttk.Combobox(sel_frame, textvariable=term_var, state='readonly', width=15)
        term_cb.grid(row=0, column=3, padx=5)
        ToolTip(term_cb, "Select term")

        view_btn = tk.Button(sel_frame, text="View Marks", command=lambda: load_marks(),
                             bg='blue', fg='white', font=('Arial', 10, 'bold'))
        view_btn.grid(row=0, column=4, padx=10)
        ToolTip(view_btn, "Load marks (Ctrl+V)")

        level_filter = self.level_var.get()
        if level_filter != 'All':
            c.execute("SELECT id, name, adm_no FROM students WHERE school_level=?", (level_filter.lower(),))
        else:
            c.execute("SELECT id, name, adm_no FROM students")
        students = c.fetchall()
        student_dict = {f"{s[1]} ({s[2]})": s[0] for s in students}
        student_cb['values'] = list(student_dict.keys())
        if student_dict:
            student_cb.set(list(student_dict.keys())[0])

        c.execute("SELECT id, name, year FROM terms")
        terms = c.fetchall()
        term_dict = {f"{t[1]} {t[2]}": t[0] for t in terms}
        term_cb['values'] = list(term_dict.keys())
        if term_dict:
            term_cb.set(list(term_dict.keys())[0])

        tree_frame = tk.Frame(win)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

        scrollbar = tk.Scrollbar(tree_frame)
        scrollbar.pack(side='right', fill='y')

        tree = ttk.Treeview(tree_frame, columns=('Subject', 'Score', 'Indicator', 'Points', 'Comments'),
                            show='headings', yscrollcommand=scrollbar.set)
        tree.heading('Subject', text='Subject', command=lambda: treeview_sort_column(tree, 'Subject', False))
        tree.heading('Score', text='Score', command=lambda: treeview_sort_column(tree, 'Score', False))
        tree.heading('Indicator', text='Indicator', command=lambda: treeview_sort_column(tree, 'Indicator', False))
        tree.heading('Points', text='Points', command=lambda: treeview_sort_column(tree, 'Points', False))
        tree.heading('Comments', text='Comments', command=lambda: treeview_sort_column(tree, 'Comments', False))
        tree.column('Subject', width=200)
        tree.column('Score', width=60, anchor='center')
        tree.column('Indicator', width=80, anchor='center')
        tree.column('Points', width=60, anchor='center')
        tree.column('Comments', width=200)
        tree.pack(fill='both', expand=True)

        scrollbar.config(command=tree.yview)

        tree.tag_configure('odd', background='#f0f0f0')
        tree.tag_configure('even', background='#ffffff')

        total_label = tk.Label(win, text="Total Points: 0", font=('Arial',12,'bold'), fg='darkblue')
        total_label.pack(pady=5)

        def load_marks():
            for row in tree.get_children():
                tree.delete(row)

            student_key = student_var.get()
            term_key = term_var.get()
            if not student_key or not term_key:
                messagebox.showerror("Error", "Please select both student and term.")
                return

            sid = student_dict[student_key]
            term_id = term_dict[term_key]

            c.execute("""
                SELECT sub.name, m.score, m.indicator, m.points, m.teacher_comments
                FROM marks m
                JOIN subjects sub ON m.subject_id = sub.id
                WHERE m.student_id = ? AND m.term_id = ?
                ORDER BY sub.name
            """, (sid, term_id))

            marks = c.fetchall()
            total_points = 0

            if not marks:
                messagebox.showinfo("No Data", "No marks found for this student in the selected term.")
                total_label.config(text="Total Points: 0")
                return

            for i, (sub, sc, ind, pts, comm) in enumerate(marks):
                tag = 'even' if i % 2 == 0 else 'odd'
                tree.insert('', 'end', values=(sub, sc, ind, pts, comm or ''), tags=(tag,))
                total_points += pts if pts else 0

            total_label.config(text=f"Total Points: {total_points}")

        def export_csv():
            student_key = student_var.get()
            term_key = term_var.get()
            if not student_key or not term_key:
                messagebox.showerror("Error", "Please select student and term first.")
                return
            sid = student_dict[student_key]
            term_id = term_dict[term_key]
            c.execute("""
                SELECT sub.name, m.score, m.indicator, m.points, m.teacher_comments
                FROM marks m
                JOIN subjects sub ON m.subject_id = sub.id
                WHERE m.student_id = ? AND m.term_id = ?
                ORDER BY sub.name
            """, (sid, term_id))
            marks = c.fetchall()
            if not marks:
                messagebox.showinfo("No Data", "No marks to export.")
                return
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                initialfile=f"marks_{student_key.split('(')[0].strip()}_{term_key}.csv"
            )
            if filename:
                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Subject', 'Score', 'Indicator', 'Points', 'Comments'])
                    writer.writerows(marks)
                messagebox.showinfo("Success", f"Exported to {filename}.")

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=5)

        export_btn = tk.Button(btn_frame, text="Export to CSV", command=export_csv,
                               bg='green', fg='white', font=('Arial', 10, 'bold'), width=15)
        export_btn.pack(side='left', padx=5)
        ToolTip(export_btn, "Export displayed marks to CSV")
        close_btn = tk.Button(btn_frame, text="Close", command=win.destroy,
                              bg='red', fg='white', font=('Arial', 10, 'bold'), width=15)
        close_btn.pack(side='left', padx=5)
        ToolTip(close_btn, "Close window (Esc)")
        win.bind('<Escape>', lambda e: win.destroy())
        win.bind('<Control-v>', lambda e: load_marks())

    # ---------- IMPORT/EXPORT MARKS ----------
    def import_marks(self):
        f = filedialog.askopenfilename(filetypes=[("CSV", "*.csv")])
        if not f:
            return
        success = 0
        errors = []
        try:
            with open(f) as file:
                reader = csv.reader(file)
                header = next(reader)
                expected = ['admission_no', 'subject', 'score', 'term', 'year', 'comments']
                if header != expected:
                    messagebox.showerror("Error", f"CSV must have columns: {', '.join(expected)}")
                    return
                for row_num, row in enumerate(reader, start=2):
                    if len(row) < 5:
                        errors.append(f"Row {row_num}: Wrong column count")
                        continue
                    adm, subject, score_str, term, year_str = row[:5]
                    comments = row[5] if len(row) > 5 else ''
                    ok_score, score = validate_score(score_str)
                    if not ok_score:
                        errors.append(f"Row {row_num}: Invalid score '{score_str}'")
                        continue
                    ok_year, year = validate_year(year_str)
                    if not ok_year:
                        errors.append(f"Row {row_num}: Invalid year '{year_str}'")
                        continue
                    c.execute("SELECT id, grade FROM students WHERE adm_no=?", (adm,))
                    stu = c.fetchone()
                    if not stu:
                        errors.append(f"Row {row_num}: Student admission {adm} not found")
                        continue
                    stu_id, grade = stu
                    c.execute("SELECT id FROM subjects WHERE name=? AND grade=?", (subject, grade))
                    subj = c.fetchone()
                    if not subj:
                        errors.append(f"Row {row_num}: Subject '{subject}' not found for grade {grade}")
                        continue
                    subj_id = subj[0]
                    c.execute("SELECT id FROM terms WHERE name=? AND year=?", (term, year))
                    term_rec = c.fetchone()
                    if not term_rec:
                        errors.append(f"Row {row_num}: Term '{term} {year}' not found")
                        continue
                    term_id = term_rec[0]
                    ind, pts = get_indicator(score)
                    try:
                        c.execute("""
                            INSERT OR REPLACE INTO marks (student_id, subject_id, term_id, score, indicator, points, teacher_comments)
                            VALUES (?,?,?,?,?,?,?)
                        """, (stu_id, subj_id, term_id, score, ind, pts, comments))
                        success += 1
                    except Exception as e:
                        errors.append(f"Row {row_num}: DB error - {str(e)}")
                conn.commit()
        except Exception as e:
            messagebox.showerror("Error", f"Import failed: {e}")
            return
        if errors:
            msg = f"Imported {success} marks, {len(errors)} errors.\n\nFirst 10 errors:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                msg += f"\n... and {len(errors)-10} more"
            messagebox.showwarning("Import", msg)
        else:
            messagebox.showinfo("Import", f"Imported {success} marks successfully.")

    def export_marks(self):
        f = filedialog.asksaveasfilename(defaultextension=".csv")
        if not f:
            return
        try:
            c.execute("""
                SELECT s.adm_no, sub.name, m.score, t.name, t.year, m.teacher_comments
                FROM marks m
                JOIN students s ON m.student_id = s.id
                JOIN subjects sub ON m.subject_id = sub.id
                JOIN terms t ON m.term_id = t.id
                ORDER BY t.year, t.name, s.name
            """)
            rows = c.fetchall()
            with open(f, 'w', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(['admission_no', 'subject', 'score', 'term', 'year', 'comments'])
                writer.writerows(rows)
            messagebox.showinfo("Export", f"Exported {len(rows)} marks.")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")

    def export_marks_excel(self):
        if not EXCEL_AVAILABLE:
            messagebox.showinfo("Excel Export", "openpyxl not installed. Please install it or use CSV export.")
            return
        f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not f:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Marks"
            ws.append(['Admission No', 'Student Name', 'Grade', 'Subject', 'Score', 'Indicator', 'Points', 'Comments', 'Term', 'Year'])
            c.execute("""
                SELECT s.adm_no, s.name, s.grade, sub.name, m.score, m.indicator, m.points, m.teacher_comments, t.name, t.year
                FROM marks m
                JOIN students s ON m.student_id = s.id
                JOIN subjects sub ON m.subject_id = sub.id
                JOIN terms t ON m.term_id = t.id
                ORDER BY t.year, t.name, s.grade, s.name, sub.name
            """)
            for row in c.fetchall():
                ws.append(row)
            wb.save(f)
            messagebox.showinfo("Export", f"Exported to {f}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")

    def archive_term(self):
        if messagebox.askyesno("Archive Term", "This will move all marks for the active term to archive and clear them. Continue?"):
            active = get_active_term()
            if not active:
                messagebox.showerror("Error", "No active term set.")
                return
            term_id, term_name, year = active
            c.execute("""
                INSERT INTO archived_marks (student_id, subject_id, term_id, score, indicator, points, teacher_comments, archive_date)
                SELECT student_id, subject_id, term_id, score, indicator, points, teacher_comments, date('now')
                FROM marks WHERE term_id = ?
            """, (term_id,))
            c.execute("DELETE FROM marks WHERE term_id = ?", (term_id,))
            conn.commit()
            messagebox.showinfo("Success", f"Term {term_name} {year} archived.")

    # ---------- REPORTS ----------
    def add_logo_to_pdf(self, pdf):
        path = config.get('logo_path', '')
        if path and os.path.exists(path):
            try:
                pdf.image(path, x=10, y=8, w=30)
                pdf.set_y(25)
            except Exception as e:
                print("Logo error:", e)
                pdf.set_y(10)
        else:
            pdf.set_y(10)

    def generate_primary_report(self, student_id, term_id, student_name, adm, grade, term_name, term_year):
        c.execute("""
            SELECT sub.name, m.score, m.indicator, m.points, m.teacher_comments
            FROM marks m
            JOIN subjects sub ON m.subject_id = sub.id
            WHERE m.student_id = ? AND m.term_id = ?
            ORDER BY sub.name
        """, (student_id, term_id))
        marks = c.fetchall()
        if not marks:
            return None

        pdf = FPDF()
        pdf.add_page()
        self.add_logo_to_pdf(pdf)
        
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, f"{config['school_name']}", 0, 1, 'C')
        pdf.ln(10)
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(0, 10, "PRIMARY SCHOOL PROGRESS REPORT", 0, 1, 'C')
        pdf.ln(5)
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 8, f"Name: {student_name}", 0, 1)
        pdf.cell(0, 8, f"Admission No: {adm}", 0, 1)
        pdf.cell(0, 8, f"Grade: {grade}", 0, 1)
        pdf.cell(0, 8, f"Term: {term_name} {term_year}", 0, 1)
        pdf.ln(5)
        pdf.set_font("Arial", 'B', 11)
        pdf.cell(0, 10, "LEARNING AREA PERFORMANCE", 0, 1, 'L')
        pdf.ln(2)
        for sub, score, indicator, points, comments in marks:
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(0, 8, sub, 0, 1)
            pdf.set_font("Arial", '', 10)
            pdf.cell(0, 6, f"Achievement Level: {indicator} ({points} points)  Score: {int(score)}%", 0, 1)
            pdf.cell(0, 6, f"Teacher's Comments: {comments if comments else 'Continuing to make progress'}", 0, 1)
            pdf.ln(3)
        pdf.ln(5)
        pdf.set_font("Arial", 'B', 11)
        pdf.cell(0, 8, "STRENGTHS AND AREAS FOR GROWTH", 0, 1)
        pdf.set_font("Arial", '', 10)
        pdf.cell(0, 6, "Strengths: ________________________________________________", 0, 1)
        pdf.cell(0, 6, "Areas for improvement: ____________________________________", 0, 1)
        pdf.ln(10)
        pdf.cell(90, 8, "Class Teacher: ____________________", 0, 0)
        pdf.cell(90, 8, "Parent/Guardian: ____________________", 0, 1)
        pdf.cell(90, 8, "Date: ____________________", 0, 0)
        return pdf

    def generate_jss_report(self, student_id, term_id, student_name, adm, grade, term_name, term_year):
        c.execute("""
            SELECT sub.name, m.score, m.indicator, m.points
            FROM marks m
            JOIN subjects sub ON m.subject_id = sub.id
            WHERE m.student_id = ? AND m.term_id = ?
            ORDER BY sub.name
        """, (student_id, term_id))
        marks = c.fetchall()
        if not marks:
            return None

        rank, total_students = get_student_rank(student_id, term_id)

        pdf = FPDF()
        pdf.add_page()
        self.add_logo_to_pdf(pdf)
        
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, f"{config['school_name']}", 0, 1, 'C')
        pdf.ln(10)
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(0, 10, "JSS REPORT CARD", 0, 1, 'C')
        pdf.ln(5)
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 8, f"Name: {student_name}", 0, 1)
        pdf.cell(0, 8, f"Admission No: {adm}", 0, 1)
        pdf.cell(0, 8, f"Grade: {grade}", 0, 1)
        pdf.cell(0, 8, f"Term: {term_name} {term_year}", 0, 1)
        pdf.ln(5)

        pdf.set_font("Arial", 'B', 10)
        pdf.cell(80, 10, "Subject", 1, 0, 'C')
        pdf.cell(30, 10, "Score", 1, 0, 'C')
        pdf.cell(40, 10, "Indicator", 1, 0, 'C')
        pdf.cell(30, 10, "Points", 1, 1, 'C')

        pdf.set_font("Arial", '', 10)
        total_points = 0
        for sub, sc, ind, pts in marks:
            pdf.cell(80, 8, sub, 1, 0)
            pdf.cell(30, 8, str(int(sc)), 1, 0, 'C')
            pdf.cell(40, 8, ind, 1, 0, 'C')
            pdf.cell(30, 8, str(pts), 1, 1, 'C')
            total_points += pts

        pdf.ln(5)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, f"Total Points: {total_points}", 0, 1)
        pdf.cell(0, 8, f"Class Position: {rank} out of {total_students}", 0, 1)

        pdf.ln(10)
        pdf.set_font("Arial", '', 10)
        pdf.cell(90, 8, "Class Teacher: ____________________", 0, 0)
        pdf.cell(90, 8, "Headteacher: ____________________", 0, 1)
        pdf.cell(90, 8, "Date: ____________________", 0, 0)
        return pdf

    def report_card(self):
        win = tk.Toplevel(self.root)
        win.title("Report Card")
        win.geometry("400x350")
        center(win, 400, 350)

        tk.Label(win, text="Select Student", font=('Arial', 11)).pack(pady=5)
        level_filter = self.level_var.get()
        if level_filter != 'All':
            c.execute("SELECT id, name, adm_no FROM students WHERE school_level=?", (level_filter.lower(),))
        else:
            c.execute("SELECT id, name, adm_no FROM students")
        students = c.fetchall()
        student_dict = {f"{s[1]} ({s[2]})": s[0] for s in students}
        student_var = tk.StringVar()
        student_cb = ttk.Combobox(win, textvariable=student_var, values=list(student_dict.keys()), state='readonly')
        student_cb.pack(pady=5)
        ToolTip(student_cb, "Select student")

        tk.Label(win, text="Select Term", font=('Arial', 11)).pack(pady=5)
        c.execute("SELECT id, name, year FROM terms ORDER BY year DESC, name")
        terms = c.fetchall()
        term_dict = {f"{t[1]} {t[2]}": t[0] for t in terms}
        term_var = tk.StringVar()
        term_cb = ttk.Combobox(win, textvariable=term_var, values=list(term_dict.keys()), state='readonly')
        term_cb.pack(pady=5)
        ToolTip(term_cb, "Select term")

        def generate():
            student_key = student_var.get()
            term_key = term_var.get()
            if not student_key or not term_key:
                messagebox.showerror("Error", "Select student and term.")
                return
            sid = student_dict[student_key]
            term_id = term_dict[term_key]
            term_parts = term_key.split()
            tname = ' '.join(term_parts[:-1])
            tyear = term_parts[-1]

            c.execute("SELECT name, adm_no, grade, school_level FROM students WHERE id=?", (sid,))
            name, adm, grade, level = c.fetchone()

            if level == 'primary':
                pdf = self.generate_primary_report(sid, term_id, name, adm, grade, tname, tyear)
                if not pdf:
                    messagebox.showinfo("No marks", "No marks found for this student in the selected term.")
                    return
                default_filename = f"{adm}_{tname}_{tyear}_primary.pdf"
            else:
                pdf = self.generate_jss_report(sid, term_id, name, adm, grade, tname, tyear)
                if not pdf:
                    messagebox.showinfo("No marks", "No marks found for this student in the selected term.")
                    return
                default_filename = f"{adm}_{tname}_{tyear}_jss.pdf"

            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                initialfile=default_filename,
                filetypes=[("PDF files", "*.pdf")]
            )
            if filename:
                pdf.output(filename)
                messagebox.showinfo("Done", f"PDF saved to {filename}.")
            win.destroy()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)
        gen_btn = tk.Button(btn_frame, text="Generate PDF", command=generate,
                            bg=config['colors']['button_bg'], fg=config['colors']['button_fg'],
                            font=('Arial', 10, 'bold'), width=15)
        gen_btn.pack(side='left', padx=5)
        ToolTip(gen_btn, "Generate PDF report card (Ctrl+Enter)")
        close_btn = tk.Button(btn_frame, text="Close", command=win.destroy,
                              bg='red', fg='white', font=('Arial', 10, 'bold'), width=15)
        close_btn.pack(side='left', padx=5)
        ToolTip(close_btn, "Close window (Esc)")
        win.bind('<Return>', lambda e: generate())
        win.bind('<Escape>', lambda e: win.destroy())

    def batch_report_cards(self):
        win = tk.Toplevel(self.root)
        win.title("Batch Print Report Cards")
        win.geometry("400x250")
        center(win, 400, 250)

        tk.Label(win, text="BATCH PRINT", font=('Arial',16,'bold'), fg='darkblue').pack(pady=20)

        frame = tk.Frame(win)
        frame.pack(pady=10)

        tk.Label(frame, text="Grade:").grid(row=0, column=0, padx=5)
        grade_var = tk.StringVar()
        grade_cb = ttk.Combobox(frame, textvariable=grade_var, values=list(range(1,10)), state='readonly', width=5)
        grade_cb.grid(row=0, column=1, padx=5)
        grade_cb.set(7)

        tk.Label(frame, text="Term:").grid(row=1, column=0, padx=5)
        term_var = tk.StringVar()
        term_cb = ttk.Combobox(frame, textvariable=term_var, state='readonly', width=15)
        c.execute("SELECT name, year FROM terms")
        terms = [f"{t[0]} {t[1]}" for t in c.fetchall()]
        term_cb['values'] = terms
        if terms:
            term_cb.set(terms[0])
        term_cb.grid(row=1, column=1, padx=5)

        def do_batch():
            grade = grade_var.get()
            term = term_var.get()
            if not grade or not term:
                messagebox.showerror("Error", "Select grade and term.")
                return
            tparts = term.split()
            tname = ' '.join(tparts[:-1])
            tyear = int(tparts[-1])
            c.execute("SELECT id FROM terms WHERE name=? AND year=?", (tname, tyear))
            term_id = c.fetchone()
            if not term_id:
                messagebox.showerror("Error", "Term not found.")
                return
            term_id = term_id[0]
            level_filter = self.level_var.get()
            query = "SELECT id, name, adm_no, school_level FROM students WHERE grade=?"
            params = [int(grade)]
            if level_filter != 'All':
                query += " AND school_level=?"
                params.append(level_filter.lower())
            c.execute(query, params)
            students = c.fetchall()
            if not students:
                messagebox.showinfo("No students", "No students in this grade matching the selected level.")
                return
            outdir = filedialog.askdirectory(title="Select folder to save PDFs")
            if not outdir:
                return
            count = 0
            errors = 0
            for sid, sname, adm, level in students:
                try:
                    if level == 'primary':
                        pdf = self.generate_primary_report(sid, term_id, sname, adm, grade, tname, tyear)
                        if not pdf:
                            errors += 1
                            continue
                        filename = os.path.join(outdir, f"{adm}_{tname}_{tyear}_primary.pdf")
                    else:
                        pdf = self.generate_jss_report(sid, term_id, sname, adm, grade, tname, tyear)
                        if not pdf:
                            errors += 1
                            continue
                        filename = os.path.join(outdir, f"{adm}_{tname}_{tyear}_jss.pdf")
                    pdf.output(filename)
                    count += 1
                except Exception as e:
                    errors += 1
                    print(f"Error printing {adm}: {e}")
            messagebox.showinfo("Batch Print", f"Generated {count} PDFs, {errors} errors.\nSaved to: {outdir}")
            win.destroy()

        btn_print = tk.Button(win, text="GENERATE PDFs", command=do_batch,
                              bg=config['colors']['button_bg'], fg=config['colors']['button_fg'],
                              font=('Arial',12,'bold'), width=15)
        btn_print.pack(pady=20)
        ToolTip(btn_print, "Generate PDF report cards for all students in grade (respects level filter)")

        win.bind('<Return>', lambda e: do_batch())
        win.bind('<Escape>', lambda e: win.destroy())

    def class_performance(self):
        win = tk.Toplevel(self.root)
        win.title("Class Performance")
        win.geometry("700x550")
        center(win, 700, 550)
        win.resizable(True, True)

        tk.Label(win, text="Select Grade:").pack(pady=5)
        grade_var = tk.StringVar()
        grade_cb = ttk.Combobox(win, textvariable=grade_var, values=list(range(1,10)), state='readonly')
        grade_cb.pack()
        grade_cb.set(7)
        ToolTip(grade_cb, "Select grade")

        tk.Label(win, text="Select Term:").pack(pady=5)
        term_var = tk.StringVar()
        term_cb = ttk.Combobox(win, textvariable=term_var, state='readonly')
        c.execute("SELECT name, year FROM terms ORDER BY year DESC")
        terms = [f"{t[0]} {t[1]}" for t in c.fetchall()]
        term_cb['values'] = terms
        if terms:
            term_cb.set(terms[0])
        term_cb.pack()
        ToolTip(term_cb, "Select term")

        show_rank_var = tk.IntVar(value=1)
        chk_rank = tk.Checkbutton(win, text="Show student ranking (JSS only)", variable=show_rank_var)
        chk_rank.pack(pady=5)

        notebook = ttk.Notebook(win)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)

        frame1 = ttk.Frame(notebook)
        notebook.add(frame1, text="Subject Averages")
        tree1 = ttk.Treeview(frame1, columns=('Subject','Average'), show='headings')
        tree1.heading('Subject', text='Subject', command=lambda: treeview_sort_column(tree1, 'Subject', False))
        tree1.heading('Average', text='Average Score', command=lambda: treeview_sort_column(tree1, 'Average', False))
        tree1.column('Subject', width=200)
        tree1.column('Average', width=100)
        tree1.pack(fill='both', expand=True)

        frame2 = ttk.Frame(notebook)
        notebook.add(frame2, text="Student Performance")
        tree2 = ttk.Treeview(frame2, columns=('Name','Adm','Average'), show='headings')
        tree2.heading('Name', text='Name', command=lambda: treeview_sort_column(tree2, 'Name', False))
        tree2.heading('Adm', text='Adm No', command=lambda: treeview_sort_column(tree2, 'Adm', False))
        tree2.heading('Average', text='Average', command=lambda: treeview_sort_column(tree2, 'Average', False))
        tree2.column('Name', width=200)
        tree2.column('Adm', width=100)
        tree2.column('Average', width=100)
        tree2.pack(fill='both', expand=True)

        tree1.tag_configure('odd', background='#f0f0f0')
        tree1.tag_configure('even', background='#ffffff')
        tree2.tag_configure('odd', background='#f0f0f0')
        tree2.tag_configure('even', background='#ffffff')

        def load_data():
            grade = grade_var.get()
            term = term_var.get()
            if not grade or not term:
                messagebox.showerror("Error", "Please select grade and term.")
                return
            tparts = term.split()
            tname = ' '.join(tparts[:-1])
            tyear = int(tparts[-1])
            c.execute("SELECT id FROM terms WHERE name=? AND year=?", (tname, tyear))
            term_id = c.fetchone()[0]

            for row in tree1.get_children():
                tree1.delete(row)
            for row in tree2.get_children():
                tree2.delete(row)

            level_filter = self.level_var.get()
            level_condition = ""
            level_params = []
            if level_filter != 'All':
                level_condition = " AND s.school_level = ?"
                level_params.append(level_filter.lower())

            query = """
                SELECT sub.name, AVG(m.score) as avg_score
                FROM subjects sub
                LEFT JOIN marks m ON sub.id = m.subject_id AND m.term_id=?
                LEFT JOIN students s ON m.student_id = s.id AND s.grade=?
                WHERE sub.grade=?
            """ + level_condition + """
                GROUP BY sub.name
                ORDER BY sub.name
            """
            params = [term_id, int(grade), int(grade)] + level_params
            c.execute(query, params)
            for i, (sub, avg) in enumerate(c.fetchall()):
                tag = 'even' if i % 2 == 0 else 'odd'
                tree1.insert('', 'end', values=(sub, f"{avg:.2f}" if avg else "-"), tags=(tag,))

            query = """
                SELECT s.name, s.adm_no, AVG(m.score) as avg_score
                FROM students s
                LEFT JOIN marks m ON s.id = m.student_id AND m.term_id=?
                WHERE s.grade=?
            """ + level_condition + """
                GROUP BY s.id
                HAVING avg_score IS NOT NULL
                ORDER BY avg_score DESC
            """
            params = [term_id, int(grade)] + level_params
            c.execute(query, params)
            if show_rank_var.get() == 1:
                tree2['columns'] = ('Rank','Name','Adm','Average')
                tree2.heading('Rank', text='Rank')
                tree2.column('Rank', width=50)
                for i, (name, adm, avg) in enumerate(c.fetchall(), 1):
                    tag = 'even' if i % 2 == 0 else 'odd'
                    tree2.insert('', 'end', values=(i, name, adm, f"{avg:.2f}"), tags=(tag,))
            else:
                tree2['columns'] = ('Name','Adm','Average')
                for i, (name, adm, avg) in enumerate(c.fetchall()):
                    tag = 'even' if i % 2 == 0 else 'odd'
                    tree2.insert('', 'end', values=(name, adm, f"{avg:.2f}"), tags=(tag,))

        load_btn = tk.Button(win, text="Load Performance", command=load_data,
                             bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
        load_btn.pack(pady=5)
        ToolTip(load_btn, "Load performance data (Ctrl+L)")
        win.bind('<Return>', lambda e: load_data())
        win.bind('<Escape>', lambda e: win.destroy())
        win.bind('<Control-l>', lambda e: load_data())

    def previous_results(self):
        win = tk.Toplevel(self.root)
        win.title("Previous Results")
        win.geometry("900x550")
        center(win, 900, 550)
        win.resizable(True, True)

        sel_frame = tk.Frame(win)
        sel_frame.pack(pady=10)

        tk.Label(sel_frame, text="Student:").pack(side='left', padx=5)
        student_var = tk.StringVar()
        student_cb = ttk.Combobox(sel_frame, textvariable=student_var, state='readonly', width=30)
        student_cb.pack(side='left', padx=5)

        load_btn = tk.Button(sel_frame, text="Load Results", command=lambda: load_results(),
                             bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
        load_btn.pack(side='left', padx=10)
        ToolTip(load_btn, "Load previous results for student")

        level_filter = self.level_var.get()
        if level_filter != 'All':
            c.execute("SELECT id, name, adm_no FROM students WHERE school_level=?", (level_filter.lower(),))
        else:
            c.execute("SELECT id, name, adm_no FROM students")
        students = c.fetchall()
        student_dict = {f"{s[1]} ({s[2]})": s[0] for s in students}
        student_cb['values'] = list(student_dict.keys())
        if student_dict:
            student_cb.set(list(student_dict.keys())[0])

        tree_frame = tk.Frame(win)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

        scrollbar = tk.Scrollbar(tree_frame)
        scrollbar.pack(side='right', fill='y')

        tree = ttk.Treeview(tree_frame, columns=('Term', 'Subject', 'Score', 'Indicator', 'Points', 'Comments'),
                            show='headings', yscrollcommand=scrollbar.set)
        tree.heading('Term', text='Term')
        tree.heading('Subject', text='Subject')
        tree.heading('Score', text='Score')
        tree.heading('Indicator', text='Indicator')
        tree.heading('Points', text='Points')
        tree.heading('Comments', text='Comments')
        tree.column('Term', width=100)
        tree.column('Subject', width=200)
        tree.column('Score', width=60, anchor='center')
        tree.column('Indicator', width=80, anchor='center')
        tree.column('Points', width=60, anchor='center')
        tree.column('Comments', width=200)
        tree.pack(fill='both', expand=True)

        scrollbar.config(command=tree.yview)

        tree.tag_configure('odd', background='#f0f0f0')
        tree.tag_configure('even', background='#ffffff')

        def load_results():
            for row in tree.get_children():
                tree.delete(row)
            student_key = student_var.get()
            if not student_key:
                return
            sid = student_dict[student_key]
            c.execute("""
                SELECT t.name || ' ' || t.year, sub.name, m.score, m.indicator, m.points, m.teacher_comments
                FROM marks m
                JOIN terms t ON m.term_id = t.id
                JOIN subjects sub ON m.subject_id = sub.id
                WHERE m.student_id = ?
                ORDER BY t.year, t.name, sub.name
            """, (sid,))
            rows = c.fetchall()
            for i, row in enumerate(rows):
                tag = 'even' if i % 2 == 0 else 'odd'
                tree.insert('', 'end', values=row, tags=(tag,))

        close_btn = tk.Button(win, text="Close", command=win.destroy, bg='red', fg='white')
        close_btn.pack(pady=5)
        ToolTip(close_btn, "Close window (Esc)")
        win.bind('<Escape>', lambda e: win.destroy())
        win.bind('<Return>', lambda e: load_results())

    def grading_guide(self):
        win = tk.Toplevel(self.root)
        win.title("CBC Grading Guide")
        win.geometry("500x350")
        center(win, 500, 350)
        win.resizable(False, False)

        text = """
        CBC 8-LEVEL GRADING SYSTEM (Primary & JSS)
        ===========================================
        
        Score Range    Indicator    Points    Description
        --------------------------------------------------
        90 - 100       EE1          8         Exceeding expectations
        75 - 89        EE2          7         Exceeding expectations
        58 - 74        ME1          6         Meeting expectations
        41 - 57        ME2          5         Meeting expectations
        31 - 40        AE1          4         Approaching expectations
        21 - 30        AE2          3         Approaching expectations
        11 - 20        BE1          2         Below Expectations
        1 - 10         BE2          1         Below expectations

        Note: Primary reports are descriptive (no ranking).
        JSS reports include class position and total points.
        """

        tk.Label(win, text=text, justify='left', font=('Courier', 10)).pack(pady=20)
        close_btn = tk.Button(win, text="Close", command=win.destroy, bg='blue', fg='white')
        close_btn.pack(pady=10)
        ToolTip(close_btn, "Close window (Esc)")
        win.bind('<Escape>', lambda e: win.destroy())

    # ---------- ADMIN ----------
    def manage_subjects(self):
        if self.user[3] != 'admin':
            messagebox.showerror("Error", "Admin access required.")
            return

        win = tk.Toplevel(self.root)
        win.title("Manage Subjects")
        win.geometry("600x400")
        center(win, 600, 400)
        win.resizable(True, True)

        tk.Label(win, text="Select Grade:").pack(pady=5)
        grade_var = tk.StringVar()
        grade_cb = ttk.Combobox(win, textvariable=grade_var, values=list(range(1,10)), state='readonly')
        grade_cb.pack()
        grade_cb.set(7)
        ToolTip(grade_cb, "Select grade to manage subjects")

        tree = ttk.Treeview(win, columns=('ID','Subject'), show='headings')
        tree.heading('ID', text='ID', command=lambda: treeview_sort_column(tree, 'ID', False))
        tree.heading('Subject', text='Subject', command=lambda: treeview_sort_column(tree, 'Subject', False))
        tree.column('ID', width=50)
        tree.column('Subject', width=300)
        tree.pack(fill='both', expand=True, padx=10, pady=10)

        tree.tag_configure('odd', background='#f0f0f0')
        tree.tag_configure('even', background='#ffffff')

        def load_subjects():
            for row in tree.get_children():
                tree.delete(row)
            grade = grade_var.get()
            c.execute("SELECT id, name FROM subjects WHERE grade=? ORDER BY name", (int(grade),))
            for i, row in enumerate(c.fetchall()):
                tag = 'even' if i % 2 == 0 else 'odd'
                tree.insert('', 'end', values=row, tags=(tag,))

        grade_cb.bind('<<ComboboxSelected>>', lambda e: load_subjects())
        load_subjects()

        add_frame = tk.Frame(win)
        add_frame.pack(pady=5)

        tk.Label(add_frame, text="New Subject:").pack(side='left')
        new_subj = tk.Entry(add_frame, width=25)
        new_subj.pack(side='left', padx=5)

        def add_subject():
            grade = grade_var.get()
            name = new_subj.get().strip()
            if not name:
                messagebox.showerror("Error", "Subject name cannot be empty.")
                return
            try:
                g = int(grade)
                if g < 1 or g > 9:
                    messagebox.showerror("Error", "Invalid grade.")
                    return
            except ValueError:
                messagebox.showerror("Error", "Invalid grade.")
                return
            try:
                c.execute("INSERT INTO subjects (name, grade) VALUES (?,?)", (name, g))
                conn.commit()
                new_subj.delete(0, tk.END)
                load_subjects()
                messagebox.showinfo("Success", "Subject added.")
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "Subject already exists for this grade.")

        add_btn = tk.Button(add_frame, text="Add", command=add_subject,
                            bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
        add_btn.pack(side='left')
        ToolTip(add_btn, "Add new subject (Enter)")
        add_frame.bind('<Return>', lambda e: add_subject())

        def delete_subject():
            selected = tree.selection()
            if not selected:
                messagebox.showerror("Error", "Select a subject to delete.")
                return
            item = tree.item(selected[0])
            sid = item['values'][0]
            if messagebox.askyesno("Confirm", "Delete this subject? Marks will also be deleted."):
                c.execute("DELETE FROM subjects WHERE id=?", (sid,))
                conn.commit()
                load_subjects()

        del_btn = tk.Button(win, text="Delete Selected", command=delete_subject, bg='red', fg='white')
        del_btn.pack(pady=5)
        ToolTip(del_btn, "Delete selected subject")
        close_btn = tk.Button(win, text="Close", command=win.destroy, bg='gray', fg='white')
        close_btn.pack(pady=5)
        ToolTip(close_btn, "Close window (Esc)")
        win.bind('<Escape>', lambda e: win.destroy())

    def manage_terms(self):
        if self.user[3] != 'admin':
            messagebox.showerror("Error", "Admin access required.")
            return

        win = tk.Toplevel(self.root)
        win.title("Manage Terms")
        win.geometry("500x350")
        center(win, 500, 350)
        win.resizable(True, True)

        tree = ttk.Treeview(win, columns=('ID','Term','Year','Active'), show='headings')
        tree.heading('ID', text='ID', command=lambda: treeview_sort_column(tree, 'ID', False))
        tree.heading('Term', text='Term', command=lambda: treeview_sort_column(tree, 'Term', False))
        tree.heading('Year', text='Year', command=lambda: treeview_sort_column(tree, 'Year', False))
        tree.heading('Active', text='Active', command=lambda: treeview_sort_column(tree, 'Active', False))
        tree.column('ID', width=50)
        tree.column('Term', width=100)
        tree.column('Year', width=80)
        tree.column('Active', width=60)
        tree.pack(fill='both', expand=True, padx=10, pady=10)

        tree.tag_configure('odd', background='#f0f0f0')
        tree.tag_configure('even', background='#ffffff')

        def load_terms():
            for row in tree.get_children():
                tree.delete(row)
            c.execute("SELECT id, name, year, active FROM terms ORDER BY year DESC, name")
            for i, row in enumerate(c.fetchall()):
                active = "Yes" if row[3] else "No"
                tag = 'even' if i % 2 == 0 else 'odd'
                tree.insert('', 'end', values=(row[0], row[1], row[2], active), tags=(tag,))

        load_terms()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=5)

        def add_term():
            add_win = tk.Toplevel(win)
            add_win.title("Add Term")
            add_win.geometry("300x200")
            center(add_win, 300, 200)

            tk.Label(add_win, text="Term Name:").pack(pady=5)
            e_name = tk.Entry(add_win)
            e_name.pack(pady=5)

            tk.Label(add_win, text="Year:").pack(pady=5)
            e_year = tk.Entry(add_win)
            e_year.insert(0, str(datetime.now().year))
            e_year.pack(pady=5)

            def save():
                name = e_name.get().strip()
                year_str = e_year.get().strip()
                if not name:
                    messagebox.showerror("Error", "Term name cannot be empty.")
                    return
                ok, y = validate_year(year_str)
                if not ok:
                    messagebox.showerror("Error", y)
                    return
                try:
                    c.execute("INSERT INTO terms (name, year) VALUES (?,?)", (name, y))
                    conn.commit()
                    add_win.destroy()
                    load_terms()
                except sqlite3.IntegrityError:
                    messagebox.showerror("Error", "Term already exists.")

            save_btn = tk.Button(add_win, text="Save", command=save,
                                 bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
            save_btn.pack(pady=10)
            ToolTip(save_btn, "Save term (Enter)")
            add_win.bind('<Return>', lambda e: save())
            add_win.bind('<Escape>', lambda e: add_win.destroy())

        def set_active():
            selected = tree.selection()
            if not selected:
                messagebox.showerror("Error", "Select a term.")
                return
            item = tree.item(selected[0])
            tid = item['values'][0]
            c.execute("UPDATE terms SET active=0")
            c.execute("UPDATE terms SET active=1 WHERE id=?", (tid,))
            conn.commit()
            load_terms()
            messagebox.showinfo("Success", "Active term updated.")

        add_btn = tk.Button(btn_frame, text="Add Term", command=add_term,
                            bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
        add_btn.pack(side='left', padx=5)
        ToolTip(add_btn, "Add a new term")
        active_btn = tk.Button(btn_frame, text="Set Active", command=set_active, bg='orange', fg='white')
        active_btn.pack(side='left', padx=5)
        ToolTip(active_btn, "Set selected term as active")
        close_btn = tk.Button(btn_frame, text="Close", command=win.destroy, bg='gray', fg='white')
        close_btn.pack(side='left', padx=5)
        ToolTip(close_btn, "Close window (Esc)")
        win.bind('<Escape>', lambda e: win.destroy())

    def manage_users(self):
        if self.user[3] != 'admin':
            messagebox.showerror("Error", "Admin access required.")
            return

        win = tk.Toplevel(self.root)
        win.title("Manage Users")
        win.geometry("600x400")
        center(win, 600, 400)
        win.resizable(True, True)

        tree = ttk.Treeview(win, columns=('ID','Username','Full Name','Role'), show='headings')
        tree.heading('ID', text='ID', command=lambda: treeview_sort_column(tree, 'ID', False))
        tree.heading('Username', text='Username', command=lambda: treeview_sort_column(tree, 'Username', False))
        tree.heading('Full Name', text='Full Name', command=lambda: treeview_sort_column(tree, 'Full Name', False))
        tree.heading('Role', text='Role', command=lambda: treeview_sort_column(tree, 'Role', False))
        tree.column('ID', width=50)
        tree.column('Username', width=120)
        tree.column('Full Name', width=200)
        tree.column('Role', width=100)
        tree.pack(fill='both', expand=True, padx=10, pady=10)

        tree.tag_configure('odd', background='#f0f0f0')
        tree.tag_configure('even', background='#ffffff')

        def load_users():
            for row in tree.get_children():
                tree.delete(row)
            c.execute("SELECT id, username, fullname, role FROM users")
            for i, row in enumerate(c.fetchall()):
                tag = 'even' if i % 2 == 0 else 'odd'
                tree.insert('', 'end', values=row, tags=(tag,))

        load_users()

        frame = tk.Frame(win)
        frame.pack(pady=5)

        tk.Label(frame, text="Username:").grid(row=0, column=0, padx=5)
        e_user = tk.Entry(frame)
        e_user.grid(row=0, column=1, padx=5)

        tk.Label(frame, text="Password:").grid(row=1, column=0, padx=5)
        e_pass = tk.Entry(frame, show="*")
        e_pass.grid(row=1, column=1, padx=5)

        tk.Label(frame, text="Full Name:").grid(row=2, column=0, padx=5)
        e_full = tk.Entry(frame)
        e_full.grid(row=2, column=1, padx=5)

        tk.Label(frame, text="Role:").grid(row=3, column=0, padx=5)
        role_var = tk.StringVar()
        role_cb = ttk.Combobox(frame, textvariable=role_var, values=['admin','teacher'], state='readonly')
        role_cb.grid(row=3, column=1, padx=5)
        role_cb.set('teacher')

        def add_user():
            user = e_user.get().strip()
            pwd = e_pass.get()
            full = e_full.get().strip()
            role = role_var.get()
            if not user or not pwd:
                messagebox.showerror("Error", "Username and password are required.")
                return
            hashed = hash_password(pwd)
            try:
                c.execute("INSERT INTO users (username, password, role, fullname) VALUES (?,?,?,?)",
                          (user, hashed, role, full))
                conn.commit()
                messagebox.showinfo("Success", "User added.")
                e_user.delete(0, tk.END)
                e_pass.delete(0, tk.END)
                e_full.delete(0, tk.END)
                load_users()
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "Username already exists.")

        add_btn = tk.Button(win, text="Add User", command=add_user,
                            bg=config['colors']['button_bg'], fg=config['colors']['button_fg'])
        add_btn.pack(pady=5)
        ToolTip(add_btn, "Add new user")

        def delete_user():
            selected = tree.selection()
            if not selected:
                messagebox.showerror("Error", "Select a user.")
                return
            item = tree.item(selected[0])
            uid = item['values'][0]
            if uid == 1:
                messagebox.showerror("Error", "Cannot delete default admin.")
                return
            if messagebox.askyesno("Confirm", "Delete this user?"):
                c.execute("DELETE FROM users WHERE id=?", (uid,))
                conn.commit()
                load_users()

        del_btn = tk.Button(win, text="Delete Selected", command=delete_user, bg='red', fg='white')
        del_btn.pack(pady=5)
        ToolTip(del_btn, "Delete selected user")
        close_btn = tk.Button(win, text="Close", command=win.destroy, bg='gray', fg='white')
        close_btn.pack(pady=5)
        ToolTip(close_btn, "Close window (Esc)")
        win.bind('<Escape>', lambda e: win.destroy())

# ---------- HELPER ----------
def get_active_term():
    c.execute("SELECT id, name, year FROM terms WHERE active=1")
    return c.fetchone()

# ---------- MAIN ----------
if __name__ == "__main__":
    App()