"""
Excel 多条件对比工具（增强版）
匹配键：货品名称 + 客户 + 批号 + 入库类型（4项精确匹配）
模糊比对：规格（相似度 >= 阈值视为一致）
数值比对：验收数量
输出：带颜色高亮的 xlsx + 汇总统计
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
import os

# ============================================================
# ① 用户配置区 —— 改成你自己的文件路径和列名
# ============================================================
# 桌面路径（固定为 C:\Users\Administrator\Desktop）
desktop = r'C:\Users\Administrator\Desktop'

file_a = os.path.join(desktop, '验收入库单数据.xlsx')
file_b = os.path.join(desktop, '马上放心数据.xlsx')

# 沙盒/调试回退：若桌面文件不存在，尝试当前目录
import os as _os
if not _os.path.isfile(file_a):
    _candidate = '验收入库单数据.xlsx'
    if _os.path.isfile(_candidate):
        file_a = _candidate
if not _os.path.isfile(file_b):
    _candidate = '马上放心数据.xlsx'
    if _os.path.isfile(_candidate):
        file_b = _candidate

col_product  = '货品名称'
col_customer = '客户'
col_batch    = '批号'
col_in_type  = '入库类型'      # ← 新增：第4个精确匹配条件
col_spec     = '规格'
col_qty      = '验收数量'

SPEC_SIMILARITY_THRESHOLD = 0.85   # 规格相似度阈值（0~1），可调低至0.7

output_path = r'C:\Users\Administrator\Desktop\对比结果_增强版.xlsx'

# 沙盒/调试回退：当前系统不是 Windows 时，落到当前目录
import sys as _sys
if _sys.platform != 'win32':
    output_path = _os.path.join(_os.getcwd(), '对比结果_增强版.xlsx')

# ============================================================
# ② 颜色 / 样式定义
# ============================================================
FILL_GREEN  = PatternFill('solid', start_color='C6EFCE')  # 完全一致
FONT_GREEN  = Font(color='006100')
FILL_RED    = PatternFill('solid', start_color='FFC7CE')  # 有差异
FONT_RED    = Font(color='9C0006')
FILL_YELLOW = PatternFill('solid', start_color='FFEB9C')  # 仅A有
FONT_YELLOW = Font(color='9C6500')
FILL_BLUE   = PatternFill('solid', start_color='BDD7EE')  # 仅B有
FONT_BLUE   = Font(color='1F4E79')

thin_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)

HEADER_FILL = PatternFill('solid', start_color='305496')
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
TITLE_FONT  = Font(name='Microsoft YaHei', bold=True, size=14, color='305496')

# ============================================================
# ③ 读取数据
# ============================================================
def load(path):
    df = pd.read_excel(path, dtype=str)
    # 去空格 + 空字符串转 NaN
    df = df.apply(lambda col: col.str.strip() if col.dtype == 'object' else col)
    df.replace(r'^\s*$', pd.NA, regex=True, inplace=True)
    # 数值列转 float（方便比较）
    for c in [col_qty]:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    return df

df_a = load(file_a)
df_b = load(file_b)

# ============================================================
# ④ 构建复合键（4项精确匹配）
# ============================================================
key_cols = [col_product, col_customer, col_batch, col_in_type]

def make_key(row):
    parts = []
    for c in key_cols:
        v = row.get(c)
        parts.append(str(v).strip().lower() if pd.notna(v) else '')
    return '||'.join(parts)

df_a['__key'] = df_a.apply(make_key, axis=1)
df_b['__key'] = df_b.apply(make_key, axis=1)

keys_a = set(df_a['__key'].dropna())
keys_b = set(df_b['__key'].dropna())
common_keys = keys_a & keys_b
only_a_keys = keys_a - keys_b
only_b_keys = keys_b - keys_a

# ============================================================
# ⑤ 逐条比对
# ============================================================
def spec_similarity(sa, sb):
    if pd.isna(sa) or pd.isna(sb):
        return 0.0
    return SequenceMatcher(None, str(sa), str(sb)).ratio()

def classify(row_a, row_b):
    """返回 (状态标签, 差异说明列表)"""
    diffs = []
    # 规格模糊比对
    sim = spec_similarity(row_a.get(col_spec), row_b.get(col_spec))
    spec_ok = sim >= SPEC_SIMILARITY_THRESHOLD
    if not spec_ok:
        diffs.append(f'规格不符(相似度{sim*100:.1f}%)')
    # 数量比对
    qa, qb = row_a.get(col_qty), row_b.get(col_qty)
    qty_ok = False
    if pd.notna(qa) and pd.notna(qb):
        qty_ok = (float(qa) == float(qb))
        if not qty_ok:
            diffs.append(f'数量不符(验收入库单={qa}, 马上放心={qb})')
    else:
        diffs.append('数量缺失')
    if not diffs:
        return '完全一致', []
    return '存在差异', diffs

rows_out = []

# 共有键
for k in sorted(common_keys):
    ra = df_a[df_a['__key'] == k].iloc[0]
    rb = df_b[df_b['__key'] == k].iloc[0]
    status, diffs = classify(ra, rb)
    rows_out.append({
        '比对结果': status,
        '差异说明': '; '.join(diffs),
        '货品名称': ra.get(col_product),
        '客户': ra.get(col_customer),
        '批号': ra.get(col_batch),
        '入库类型': ra.get(col_in_type),
        '验收入库单_规格': ra.get(col_spec),
        '马上放心_规格': rb.get(col_spec),
        '规格相似度': '',
        '验收入库单_验收数量': ra.get(col_qty) if pd.notna(ra.get(col_qty)) else '',
        '马上放心_验收数量': rb.get(col_qty) if pd.notna(rb.get(col_qty)) else '',
        '数量差异': '',
    })
    # 填相似度 / 数量差异
    sim = spec_similarity(ra.get(col_spec), rb.get(col_spec))
    rows_out[-1]['规格相似度'] = f'{sim*100:.1f}%'
    qa, qb = ra.get(col_qty), rb.get(col_qty)
    if pd.notna(qa) and pd.notna(qb):
        rows_out[-1]['数量差异'] = float(qa) - float(qb)

# 仅A有
for k in sorted(only_a_keys):
    ra = df_a[df_a['__key'] == k].iloc[0]
    rows_out.append({
        '比对结果': '仅验收入库单存在',
        '差异说明': '马上放心数据中未找到（按名称+客户+批号+入库类型匹配）',
        '货品名称': ra.get(col_product),
        '客户': ra.get(col_customer),
        '批号': ra.get(col_batch),
        '入库类型': ra.get(col_in_type),
        '验收入库单_规格': ra.get(col_spec), '马上放心_规格': '', '规格相似度': '',
        '验收入库单_验收数量': ra.get(col_qty) if pd.notna(ra.get(col_qty)) else '',
        '马上放心_验收数量': '', '数量差异': '',
    })

# 仅B有
for k in sorted(only_b_keys):
    rb = df_b[df_b['__key'] == k].iloc[0]
    rows_out.append({
        '比对结果': '仅马上放心存在',
        '差异说明': '验收入库单中未找到（按名称+客户+批号+入库类型匹配）',
        '货品名称': rb.get(col_product),
        '客户': rb.get(col_customer),
        '批号': rb.get(col_batch),
        '入库类型': rb.get(col_in_type),
        '验收入库单_规格': '', '马上放心_规格': rb.get(col_spec), '规格相似度': '',
        '验收入库单_验收数量': '',
        '马上放心_验收数量': rb.get(col_qty) if pd.notna(rb.get(col_qty)) else '',
        '数量差异': '',
    })

result_df = pd.DataFrame(rows_out)

# ============================================================
# ⑥ 写入 xlsx（带颜色 + 格式）
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = '对比结果'

headers = list(result_df.columns)
ws.append(headers)

# 表头样式
for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border

# 数据行
for r_idx, row in result_df.iterrows():
    ws.append([row[h] for h in headers])
    excel_row = r_idx + 2
    status = row['比对结果']

    if status == '完全一致':
        fill, font = FILL_GREEN, FONT_GREEN
    elif status == '存在差异':
        fill, font = FILL_RED, FONT_RED
    elif status == '仅验收入库单存在':
        fill, font = FILL_YELLOW, FONT_YELLOW
    elif status == '仅马上放心存在':
        fill, font = FILL_BLUE, FONT_BLUE
    else:
        fill, font = None, None

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=excel_row, column=col_idx)
        c.border = thin_border
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if fill and font:
            c.fill = fill
            c.font = font

    # 数量差异列特别标红加粗
    diff_val = row.get('数量差异', '')
    if diff_val not in ('', 0, 0.0) and pd.notna(diff_val):
        c = ws.cell(row=excel_row, column=headers.index('数量差异') + 1)
        c.font = Font(bold=True, color='C00000')

# 列宽自适应（中文按2倍宽度估算）
for col_idx, h in enumerate(headers, 1):
    max_len = max(
        [len(str(h))] +
        [len(str(v)) for v in result_df[h].tolist()[:200]]
    )
    # 粗略估算：中文/全角按2
    def vis_len(s):
        s = str(s)
        return sum(2 if ord(c) > 127 else 1 for c in s)
    max_len = max(vis_len(h), max((vis_len(v) for v in result_df[h].tolist()[:200]), default=0))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

# 冻结首行 + 自动筛选
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# ---------- Sheet2 汇总统计 ----------
ws2 = wb.create_sheet('汇总统计')
summary_headers = ['统计项', '数量']
counts = {
    '完全一致': (result_df['比对结果'] == '完全一致').sum(),
    '存在差异': (result_df['比对结果'] == '存在差异').sum(),
    '仅验收入库单存在': (result_df['比对结果'] == '仅验收入库单存在').sum(),
    '仅马上放心存在': (result_df['比对结果'] == '仅马上放心存在').sum(),
    '合计比对条目': len(result_df),
}
ws2.append(summary_headers)
for col_idx in range(1, 3):
    c = ws2.cell(row=1, column=col_idx)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

for label, val in counts.items():
    ws2.append([label, int(val)])

# 给差异行加红色粗体
for r in range(2, ws2.max_row + 1):
    label = ws2.cell(row=r, column=1).value
    if label == '存在差异':
        for col_idx in range(1, 3):
            c = ws2.cell(row=r, column=col_idx)
            c.font = Font(bold=True, color='C00000')
            c.fill = FILL_RED

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 12

wb.save(output_path)
print(f'✅ 对比完成 → {output_path}')
print(f'   完全一致: {counts["完全一致"]}  存在差异: {counts["存在差异"]}  '
      f'仅A: {counts["仅验收入库单存在"]}  仅B: {counts["仅马上放心存在"]}')
