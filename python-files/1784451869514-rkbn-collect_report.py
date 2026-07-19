# === ГЛОБАЛЬНЫЙ ПЕРЕХВАТ ОШИБОК ===
import sys
import traceback
import os

def log_and_pause():
    """Выводит ошибку на экран, записывает в лог и ждёт нажатия Enter."""
    print("\n" + "=" * 60)
    print("КРИТИЧЕСКАЯ ОШИБКА! Окно не закроется, пока не нажмёте Enter.")
    traceback.print_exc()
    try:
        os.makedirs(r"C:\Temp", exist_ok=True)
        with open(r"C:\Temp\error.log", "w", encoding="utf-8") as f:
            traceback.print_exc(file=f)
        print(f"\nПодробности также записаны в C:\\Temp\\error.log")
    except:
        pass
    input("\nНажмите Enter для выхода...")
    sys.exit(1)

# Ловим ВСЕ ошибки, включая ошибки импорта модулей
try:
    import platform
    import socket
    import getpass
    import subprocess
    from datetime import datetime, timedelta
    from io import BytesIO
    from pathlib import Path

    import psutil
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from Crypto.Cipher import AES
    from Crypto.Protocol.KDF import PBKDF2
    from Crypto.Random import get_random_bytes

except Exception:
    log_and_pause()

# === ПАРАМЕТРЫ ШИФРОВАНИЯ ===
PASSWORD = "OtP1nd2j3hL9dxeREJOlQkUvMRPdICH0m2K7BQh8sVuY8XRT3ugUpPHdw1xnLEvh"
SALT = b"MyS@lt!2024_"
ITERATIONS = 10_000
KEY_SIZE = 32
OUTPUT_DIR = r"C:\Temp"

def get_hostname():
    return platform.node()

def get_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

def get_user():
    return getpass.getuser()

# ==================== ФУНКЦИИ СБОРА ====================
def collect_overview():
    data = {
        "Report Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Hostname": get_hostname(),
        "IP Address": get_ip(),
        "User": get_user(),
        "OS": platform.platform(),
        "Architecture": platform.architecture()[0],
        "Processor": platform.processor(),
    }
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                            r"SOFTWARE\Microsoft\Windows NT\CurrentVersion") as key:
            data["OS Name"] = winreg.QueryValueEx(key, "ProductName")[0]
            try:
                data["ReleaseId"] = winreg.QueryValueEx(key, "DisplayVersion")[0]
            except:
                data["ReleaseId"] = ""
            data["Build"] = winreg.QueryValueEx(key, "CurrentBuild")[0]
    except:
        data["OS Name"] = platform.platform()
    try:
        out = subprocess.check_output("systeminfo", shell=True, text=True)
        for line in out.splitlines():
            if "Domain:" in line:
                data["Domain"] = line.split(":")[1].strip()
                break
    except:
        data["Domain"] = "Unknown"
    return data

def collect_system_info():
    info = {}
    try:
        cs = psutil.win_sysinfo()
        info["Manufacturer"] = cs.manufacturer
        info["Model"] = cs.model
        info["Total RAM (GB)"] = round(psutil.virtual_memory().total / (1024**3), 1)
    except:
        pass
    try:
        out = subprocess.check_output("wmic bios get serialnumber", shell=True, text=True)
        lines = out.strip().splitlines()
        if len(lines) > 1 and lines[1].strip():
            info["Serial Number"] = lines[1].strip()
    except:
        pass
    try:
        out = subprocess.check_output("wmic os get installdate", shell=True, text=True)
        lines = out.strip().splitlines()
        if len(lines) > 1 and lines[1].strip():
            dt = datetime.strptime(lines[1].strip()[:14], "%Y%m%d%H%M%S")
            info["Install Date"] = dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        pass
    return info

def collect_users_and_groups():
    users = []
    admins = []
    try:
        out = subprocess.check_output(
            'wmic useraccount where "localaccount=true" get name',
            shell=True, text=True
        )
        for line in out.splitlines():
            name = line.strip()
            if name and name != "Name":
                users.append(name)
    except:
        pass
    try:
        out = subprocess.check_output("net localgroup administrators", shell=True, text=True)
        capture = False
        for line in out.splitlines():
            if "---" in line:
                capture = True
                continue
            if capture and line.strip() and not line.startswith("Команда выполнена"):
                admins.append(line.strip())
    except:
        pass
    groups = []
    try:
        out = subprocess.check_output("net localgroup", shell=True, text=True)
        for line in out.splitlines():
            if line.startswith("*"):
                group_name = line[1:].strip()
                members = []
                try:
                    members_out = subprocess.check_output(
                        f'net localgroup "{group_name}"', shell=True, text=True
                    )
                    capture = False
                    for mline in members_out.splitlines():
                        if "---" in mline:
                            capture = True
                            continue
                        if capture and mline.strip() and not mline.startswith("Команда выполнена"):
                            members.append(mline.strip())
                    groups.append({"Group": group_name, "Members": ", ".join(members)})
                except:
                    groups.append({"Group": group_name, "Members": "Error retrieving"})
    except:
        pass
    user_list = [{"Username": u, "IsAdmin": u in admins} for u in users]
    return user_list, groups

def collect_network():
    rdp = {}
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                            r"SYSTEM\CurrentControlSet\Control\Terminal Server") as key:
            fDeny = winreg.QueryValueEx(key, "fDenyTSConnections")[0]
        rdp["RDP Enabled"] = fDeny == 0
    except:
        rdp["RDP Enabled"] = "Unknown"
    netstat = []
    try:
        out = subprocess.check_output("netstat -ano", shell=True, text=True)
        netstat = [{"Line": line.strip()} for line in out.splitlines() if line.strip()]
    except:
        pass
    hosts = []
    hosts_path = r"C:\Windows\System32\drivers\etc\hosts"
    if os.path.exists(hosts_path):
        with open(hosts_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    hosts.append({"Entry": line})
    dns_cache = []
    try:
        out = subprocess.check_output("ipconfig /displaydns", shell=True, text=True)
        record = {}
        for line in out.splitlines():
            if not line.strip():
                if record:
                    dns_cache.append(record)
                    record = {}
            elif ":" in line:
                k, v = line.split(":", 1)
                record[k.strip()] = v.strip()
        if record:
            dns_cache.append(record)
    except:
        pass
    return rdp, netstat, hosts, dns_cache

def collect_installed_software():
    import winreg
    software = []
    paths = [
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall",
        r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"
    ]
    for uninstall_path in paths:
        try:
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, uninstall_path, 0, winreg.KEY_READ)
            i = 0
            while True:
                try:
                    subkey_name = winreg.EnumKey(key, i)
                    subkey = winreg.OpenKey(key, subkey_name)
                    try:
                        name = winreg.QueryValueEx(subkey, "DisplayName")[0]
                    except:
                        name = None
                    if name:
                        version = publisher = install_date = ""
                        try: version = winreg.QueryValueEx(subkey, "DisplayVersion")[0]
                        except: pass
                        try: publisher = winreg.QueryValueEx(subkey, "Publisher")[0]
                        except: pass
                        try: install_date = winreg.QueryValueEx(subkey, "InstallDate")[0]
                        except: pass
                        software.append({
                            "Name": name,
                            "Version": version,
                            "Publisher": publisher,
                            "InstallDate": install_date,
                        })
                    winreg.CloseKey(subkey)
                    i += 1
                except OSError:
                    break
            winreg.CloseKey(key)
        except:
            pass
    return software

def collect_portable_apps():
    known_names = set()
    installed = collect_installed_software()
    for sw in installed:
        known_names.add(sw["Name"].lower().replace(" ", ""))
    portable_paths = [
        os.environ.get("ProgramFiles", "C:\\Program Files"),
        os.environ.get("ProgramFiles(x86)", "C:\\Program Files (x86)"),
        os.path.join(os.environ.get("LOCALAPPDATA", ""), "Programs"),
    ]
    portable = []
    for base in portable_paths:
        if not base or not os.path.exists(base):
            continue
        try:
            for entry in os.scandir(base):
                if entry.is_dir():
                    exe_file = None
                    for f in os.scandir(entry.path):
                        if f.is_file() and f.name.lower().endswith(".exe"):
                            exe_file = f.path
                            break
                    if exe_file:
                        name = os.path.splitext(os.path.basename(exe_file))[0]
                        clean = name.lower().replace(" ", "")
                        if clean in known_names:
                            continue
                        version = manufacturer = ""
                        try:
                            escaped = exe_file.replace("\\", "\\\\")
                            info = subprocess.check_output(
                                f'wmic datafile where name="{escaped}" get version, manufacturer',
                                shell=True, text=True
                            )
                            lines = info.strip().splitlines()
                            if len(lines) > 1:
                                parts = lines[1].split(None, 1)
                                version = parts[0] if parts else ""
                                manufacturer = parts[1] if len(parts) > 1 else ""
                        except:
                            pass
                        portable.append({
                            "Name": name,
                            "Version": version,
                            "Publisher": manufacturer,
                            "Path": entry.path,
                        })
        except:
            pass
    return portable

def collect_autoruns():
    import winreg
    entries = []
    hives = {
        "HKCU": winreg.HKEY_CURRENT_USER,
        "HKLM": winreg.HKEY_LOCAL_MACHINE,
    }
    subpaths = [
        r"Software\Microsoft\Windows\CurrentVersion\Run",
        r"Software\Microsoft\Windows\CurrentVersion\RunOnce",
        r"Software\Microsoft\Windows\CurrentVersion\RunServices",
        r"Software\Microsoft\Windows\CurrentVersion\RunServicesOnce",
    ]
    for hive_name, hive in hives.items():
        for subpath in subpaths:
            try:
                key = winreg.OpenKey(hive, subpath, 0, winreg.KEY_READ)
                i = 0
                while True:
                    try:
                        name, value, _ = winreg.EnumValue(key, i)
                        entries.append({
                            "Hive": hive_name,
                            "Path": subpath,
                            "Name": name,
                            "Value": str(value)
                        })
                        i += 1
                    except OSError:
                        break
                winreg.CloseKey(key)
            except:
                pass
    special = [
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Winlogon", "Shell"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Winlogon", "Userinit"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Windows", "AppInit_DLLs"),
    ]
    for hive, path, value_name in special:
        try:
            key = winreg.OpenKey(hive, path, 0, winreg.KEY_READ)
            value, _ = winreg.QueryValueEx(key, value_name)
            entries.append({
                "Hive": "HKLM",
                "Path": path,
                "Name": value_name,
                "Value": str(value)
            })
            winreg.CloseKey(key)
        except:
            pass
    return entries

def collect_processes():
    proc_list = []
    for proc in psutil.process_iter(['pid', 'name', 'exe', 'username']):
        try:
            proc_list.append({
                "PID": proc.info['pid'],
                "Name": proc.info['name'],
                "Path": proc.info['exe'] or "",
                "User": proc.info['username'] or "",
            })
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass
    return proc_list

def collect_services():
    services = []
    for svc in psutil.win_service_iter():
        try:
            services.append({
                "Name": svc.name(),
                "DisplayName": svc.display_name(),
                "Status": svc.status(),
                "StartType": svc.start_type(),
                "BinaryPath": svc.binpath() or "",
            })
        except:
            pass
    return services

def collect_drivers():
    drivers = []
    try:
        out = subprocess.check_output("driverquery /v /fo csv", shell=True, text=True)
        lines = out.splitlines()
        if lines:
            headers = lines[0].replace('"', '').split(',')
            for line in lines[1:]:
                if line.strip():
                    vals = [v.strip('" ') for v in line.split(',')]
                    if len(vals) >= len(headers):
                        drivers.append(dict(zip(headers, vals)))
    except:
        pass
    return drivers

def collect_scheduled_tasks():
    tasks = []
    try:
        out = subprocess.check_output('schtasks /query /fo CSV /v', shell=True, text=True)
        lines = out.splitlines()
        if not lines:
            return []
        headers = lines[0].replace('"', '').split(',')
        for line in lines[1:]:
            if line.strip():
                vals = [v.strip('" ') for v in line.split(',')]
                if len(vals) >= len(headers):
                    tasks.append(dict(zip(headers, vals)))
    except:
        pass
    return tasks

def collect_security_events():
    events = []
    try:
        query = "*[System[(EventID=4624 or EventID=4625 or EventID=4672 or EventID=1102 or EventID=7045 or EventID=7036 or EventID=1001 or EventID=41) and TimeCreated[timediff(@SystemTime) <= 86400000]]]"
        cmd = f'wevtutil qe Security /q:"{query}" /c:100 /rd:true /f:text'
        out = subprocess.check_output(cmd, shell=True, text=True)
        record = {}
        for line in out.splitlines():
            if not line.strip():
                if record:
                    events.append(record)
                    record = {}
            elif ":" in line:
                k, v = line.split(":", 1)
                record[k.strip()] = v.strip()
        if record:
            events.append(record)
    except:
        pass
    return events

def collect_disk_encryption():
    disks = []
    for part in psutil.disk_partitions():
        try:
            usage = psutil.disk_usage(part.mountpoint)
            disks.append({
                "Device": part.device,
                "Mount": part.mountpoint,
                "FS": part.fstype,
                "TotalGB": round(usage.total / (1024**3), 2),
                "FreeGB": round(usage.free / (1024**3), 2),
            })
        except:
            pass
    bitlocker = []
    try:
        out = subprocess.check_output("manage-bde -status", shell=True, text=True)
        vol = {}
        for line in out.splitlines():
            if "Volume:" in line:
                if vol:
                    bitlocker.append(vol)
                vol = {"Volume": line.split(":", 1)[1].strip()}
            elif "Conversion Status:" in line:
                vol["ConversionStatus"] = line.split(":", 1)[1].strip()
            elif "Protection Method:" in line:
                vol["ProtectionMethod"] = line.split(":", 1)[1].strip()
        if vol:
            bitlocker.append(vol)
    except:
        pass
    return disks, bitlocker

def collect_environment():
    return [{"Variable": k, "Value": v} for k, v in os.environ.items()]

def collect_usb_history():
    import winreg
    devices = []
    path = r"SYSTEM\CurrentControlSet\Enum\USBSTOR"
    try:
        key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, path, 0, winreg.KEY_READ)
        i = 0
        while True:
            try:
                subkey_name = winreg.EnumKey(key, i)
                devices.append({"Device": subkey_name})
                i += 1
            except OSError:
                break
        winreg.CloseKey(key)
    except:
        pass
    return devices

def collect_unsigned_drivers():
    r"""Поиск неподписанных .sys файлов в System32\drivers (PowerShell)."""
    unsigned = []
    try:
        ps_cmd = r'Get-ChildItem -Path C:\Windows\System32\drivers -Filter *.sys -ErrorAction SilentlyContinue | Get-AuthenticodeSignature | Where-Object { $_.Status -ne "Valid" } | Select-Object -ExpandProperty Path'
        out = subprocess.check_output(["powershell", "-Command", ps_cmd], text=True)
        for line in out.splitlines():
            if line.strip():
                unsigned.append({"Path": line.strip()})
    except:
        pass
    return unsigned

def collect_legacy_tasks():
    r"""Старые задачи в C:\Windows\Tasks и System32\Tasks."""
    files = []
    for folder in [r"C:\Windows\Tasks", r"C:\Windows\System32\Tasks"]:
        if os.path.exists(folder):
            for root, dirs, filenames in os.walk(folder):
                for f in filenames:
                    files.append({"Path": os.path.join(root, f)})
    return files

def collect_applocker():
    rules = []
    try:
        out = subprocess.check_output(
            'powershell -Command "Get-AppLockerPolicy -Effective -ErrorAction SilentlyContinue | Select-Object -ExpandProperty RuleCollections | Out-String"',
            shell=True, text=True
        )
        if out.strip():
            rules = [{"RawPolicy": out.strip()}]
    except:
        pass
    return rules

# ==================== ФОРМИРОВАНИЕ EXCEL ====================
def style_sheet(ws, headers=None):
    if headers:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    ws.freeze_panes = 'A2'

def create_workbook(overview_data, blocks):
    """Создаёт Excel с титульным листом Overview и остальными блоками."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Титульный лист Overview
    ws_overview = wb.create_sheet("Overview", 0)
    ws_overview.merge_cells('A1:B1')
    title_cell = ws_overview['A1']
    title_cell.value = "Host Audit Report"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_overview.row_dimensions[1].height = 30

    ws_overview.merge_cells('A2:B2')
    ws_overview['A2'].value = f"Generated: {overview_data.get('Report Generated', '')}"
    ws_overview['A2'].font = Font(italic=True, size=10)

    row = 4
    for key, value in overview_data.items():
        if key == "Report Generated":
            continue
        ws_overview.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws_overview.cell(row=row, column=2, value=str(value))
        row += 1

    ws_overview.column_dimensions['A'].width = 25
    ws_overview.column_dimensions['B'].width = 50
    ws_overview.freeze_panes = 'A4'

    # Остальные листы
    for block_name, (headers, rows) in blocks.items():
        ws = wb.create_sheet(block_name)
        if rows:
            ws.append(headers)
            for row_dict in rows:
                ws.append([row_dict.get(h, "") for h in headers])
            style_sheet(ws, headers)
        else:
            ws.append(["No data"])
            ws.merge_cells('A1:Z1')
            ws['A1'].font = Font(italic=True, color="808080")
    return wb

def encrypt_data(data):
    key = PBKDF2(PASSWORD, SALT, dkLen=KEY_SIZE, count=ITERATIONS)
    iv = get_random_bytes(16)
    cipher = AES.new(key, AES.MODE_CBC, iv)
    pad_len = 16 - len(data) % 16
    if pad_len == 0:
        pad_len = 16
    padded = data + bytes([pad_len] * pad_len)
    encrypted = cipher.encrypt(padded)
    return iv + encrypted

def main():
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        hostname = get_hostname()
        enc_path = os.path.join(OUTPUT_DIR, f"{hostname}-report.enc")

        print("Сбор информации о системе...")
        # Собираем overview отдельно
        overview = collect_overview()
        # Остальные данные
        system_info = collect_system_info()
        users, groups = collect_users_and_groups()
        rdp, netstat, hosts, dns_cache = collect_network()
        installed = collect_installed_software()
        portable = collect_portable_apps()
        autoruns = collect_autoruns()
        processes = collect_processes()
        services = collect_services()
        drivers = collect_drivers()
        tasks = collect_scheduled_tasks()
        events = collect_security_events()
        disks, bitlocker = collect_disk_encryption()
        env = collect_environment()
        usb = collect_usb_history()
        unsigned = collect_unsigned_drivers()
        legacy = collect_legacy_tasks()
        applocker = collect_applocker()

        # Блоки данных для листов (каждый элемент – кортеж (заголовки, строки))
        data_blocks = {
            "System Info": (["Key", "Value"], [{"Key": k, "Value": v} for k, v in system_info.items()]),
            "Users & Groups": (["Group", "Members"], groups),
            "Local Users": (["Username", "IsAdmin"], users),
            "RDP": (["Property", "Value"], [{"Property": "RDP Enabled", "Value": str(rdp.get("RDP Enabled", ""))}]),
            "Netstat": (["Line"], netstat),
            "Hosts": (["Entry"], hosts),
            "DNS Cache": (["Record"], [{"Record": str(r)} for r in dns_cache]),
            "Installed Software": (["Name", "Version", "Publisher", "InstallDate"], installed),
            "Portable Apps": (["Name", "Version", "Publisher", "Path"], portable),
            "Autoruns": (["Hive", "Path", "Name", "Value"], autoruns),
            "Processes": (["PID", "Name", "Path", "User"], processes),
            "Services": (["Name", "DisplayName", "Status", "StartType", "BinaryPath"], services),
            "Drivers": (["Module Name", "Display Name", "Driver Type", "Link Date"],
                         [{"Module Name": d.get("Module Name",""), "Display Name": d.get("Display Name",""),
                           "Driver Type": d.get("Driver Type",""), "Link Date": d.get("Link Date","")} for d in drivers]),
            "Scheduled Tasks": (list(tasks[0].keys()) if tasks else [], tasks),
            "Security Events": (list(events[0].keys()) if events else [], events),
            "Disks": (list(disks[0].keys()) if disks else [], disks),
            "BitLocker": (["Volume", "ConversionStatus", "ProtectionMethod"], bitlocker),
            "Environment": (["Variable", "Value"], env),
            "USB History": (["Device"], usb),
            "Unsigned Drivers": (["Path"], unsigned),
            "Legacy Tasks": (["Path"], legacy),
            "AppLocker": (["RawPolicy"], applocker),
        }

        print("Создание Excel...")
        wb = create_workbook(overview, data_blocks)
        xlsx_buffer = BytesIO()
        wb.save(xlsx_buffer)
        xlsx_data = xlsx_buffer.getvalue()

        print("Шифрование...")
        encrypted = encrypt_data(xlsx_data)
        with open(enc_path, "wb") as f:
            f.write(encrypted)
        print(f"Зашифрованный отчёт сохранён: {enc_path}")
    except Exception:
        log_and_pause()
    finally:
        input("Нажмите Enter для выхода...")

if __name__ == "__main__":
    main()