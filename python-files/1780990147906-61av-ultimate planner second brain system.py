import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# Style constants (Muted, desaturated slate/sage palette easy on the eyes)
font_family = "Segoe UI"
header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid") # Slate grey
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
title_font = Font(name=font_family, size=16, bold=True, color="2D3748")
section_font = Font(name=font_family, size=12, bold=True, color="4A5568")
body_font = Font(name=font_family, size=10, color="2D3748")
hint_font = Font(name=font_family, size=9, italic=True, color="718096")

accent_fill_1 = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid") # Light slate grey
accent_fill_2 = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid") # Light mint/teal
accent_fill_3 = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid") # Soft yellow alert

thin_border_side = Side(style='thin', color='CBD5E0')
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

# -------------------------------------------------------------
# TAB 1: MASTER REPOSITORY & INTERACTIVE APP DECK
# -------------------------------------------------------------
ws1 = wb.active
ws1.title = "App Architecture & 20 Templates"
ws1.views.sheetView[0].showGridLines = True

# Title block
ws1['A1'] = "Ultimate Planner & Second Brain System - Application Architecture"
ws1['A1'].font = title_font
ws1['A2'] = "A collection of 20 modular dashboard templates cross-optimized for Tablet UI and multi-language deployment."
ws1['A2'].font = hint_font

# Translation Selector UI simulation
ws1['A4'] = "Selected System Language:"
ws1['A4'].font = Font(name=font_family, size=10, bold=True)
ws1['B4'] = "English (Auto-Translate Enabled: EL, SV, FI, NL, DA, RU...)"
ws1['B4'].font = Font(name=font_family, size=10, color="1A202C")
ws1['B4'].fill = accent_fill_2

headers_1 = ["Category", "Template ID", "Template Name", "Tablet UI Optimization Layout", "Key Data Properties / Columns", "Core Functional Widgets Included"]
for col_idx, text in enumerate(headers_1, start=1):
    cell = ws1.cell(row=6, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

templates_data = [
    ("1. Agendas & Calendars", "T01", "Daily Alignment Matrix", "Split-pane (Left: Timeline, Right: Focus Elements)", "Time Block, Task Linked, Energy Level, Notes", "Hourly blocks, Reflection Prompt, Quick Capture"),
    ("1. Agendas & Calendars", "T02", "Weekly Sprint Planner", "3-4 Grid block view optimized for landscape orientation", "Days of Week, Weekly Objectives, Win Metrics, Blockers", "Weekly Rollup, Habit Streaks, Task Drag-Drop target"),
    ("1. Agendas & Calendars", "T03", "Monthly Horizon Overview", "Full-width calendar matrix with expandable sidebar", "Date, Project Deadline, Milestone Marker, Budget Check", "Monthly goal progression bar, high-priority flagger"),
    ("1. Agendas & Calendars", "T04", "Quarterly Strategy Map", "Stacked horizontal lanes for multi-month alignment", "Quarterly OKRs, Key Results, Resource allocation, Status", "Progress bars, OKR linking widget, review schedule"),
    
    ("2. Digital Notebooks", "T05", "Quick Notes Inbox (GTD)", "Single column continuous feed with swipe actions", "Timestamp, Capture Text, Tag, Status (Processed/Raw)", "Voice-to-text input area, fast tagger buttons"),
    ("2. Digital Notebooks", "T06", "Structured Knowledge Vault", "Two-column master-detail layout (Left: Navigation tree, Right: Content)", "Topic, Category, Resources, Core Synthesis, Backlinks", "Markdown block editor, automated backlinks module"),
    ("2. Digital Notebooks", "T07", "Project Research Canvas", "Card-based grid view with cover previews", "Research Project, Hypotheses, Evidence, Source Links, Status", "Web-clipper parser viewport, file attachment bin"),
    ("2. Digital Notebooks", "T08", "Meeting Minutes Hub", "Linear template rows with quick action checkboxes", "Date, Attendees, Action Items, Key Decisions, Next Steps", "Calendar integration link, automated notification dispatcher"),
    
    ("3. Advanced Task Engines", "T09", "ADHD-Friendly Matrix (Eisenhower)", "2x2 Quad layout optimized for touch interaction", "Urgency, Importance, Dopamine Reward Level, Micro-Step", "Focus Timer toggle, micro-step breaker, energy-matching filter"),
    ("3. Advanced Task Engines", "T10", "Kanban Sprint Board", "Horizontal scrollable columns with touch cards", "To-Do, In Progress, Review, Blocked, Done, Effort Points", "Assignee avatars, drag triggers, subtask inline expansion"),
    ("3. Advanced Task Engines", "T11", "Someday/Maybe Bucket", "Accordion lists categorized by life domain", "Idea Name, Domain, Motivation Level, Review Date, Status", "Activation button (moves directly to active Sprint board)"),
    ("3. Advanced Task Engines", "T12", "Recurring Routines Tracker", "Compact inline grid list with status bubble matrices", "Routine Name, Frequency, Last Executed, Target Time, Streak", "Reset trigger automation, push notification config"),
    
    ("4. Life Planners & Hubs", "T13", "Financial Ledger & Budget", "Split view: top summary metrics cards, bottom transaction ledger", "Category, Allocation, Spent, Remaining, Forecast, Alert Level", "Formulaic overspend visual indicator, recurring bill alarms"),
    ("4. Life Planners & Hubs", "T14", "Habit Loop Architect", "Circular visual trackers or horizontal monthly matrices", "Habit Name, Cue, Reward, Monthly Progress Vector, Multi-Day Grid", "Streak tracker formula, accountability text dispatcher"),
    ("4. Life Planners & Hubs", "T15", "Meal & Nutrition Blueprint", "Weekly 7-day grid with macro-nutrient sidebar breakdown", "Day, Meal Type, Recipe Link, Ingredients, Caloric Value, Prep Status", "Auto-generated shopping list constructor module"),
    ("4. Life Planners & Hubs", "T16", "Fitness & Recovery Log", "Line chart layout space alongside precise input fields", "Workout Type, Volume/Intensity, Sleep Quality, HRV, Muscle soreness", "Progressive overload calculator, recovery recommendation indicator"),
    
    ("5. ADHD Second Brain Modules", "T17", "Brain Dump Processing Center", "Large freeform canvas or long text parsing column", "Raw Text, Extracted Tasks, Extracted Insights, Processing Status", "One-click parser converting raw sentences into structured database items"),
    ("5. ADHD Second Brain Modules", "T18", "Hyperfocus Project Sandbox", "Full screen focused UI dashboard hiding non-essential panels", "Active Obsession, Resource Pool, Milestones, Active Sprint, Session Count", "Dopamine milestone reward popups, visual progress meters"),
    ("5. ADHD Second Brain Modules", "T19", "Emotional & Energy Regulator", "Color-coded micro-buttons designed for fast touch interaction", "Current Mood, Energy Level (1-5), Environmental Triggers, Mitigation Action", "Adaptive dashboard filter adjusting task suggestions according to energy"),
    ("5. ADHD Second Brain Modules", "T20", "Digital Decutter & Vault Cleaner", "Step-by-step wizard interface layout", "Area, Last Review Date, Archive Status, Importance Index, Action Plan", "Automated archival engine suggestion engine")
]

for row_idx, data in enumerate(templates_data, start=7):
    for col_idx, val in enumerate(data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = accent_fill_1
        if col_idx in [1, 2]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

# -------------------------------------------------------------
# TAB 2: EXECUTABLE PRODUCT 1 - ADHD SECOND BRAIN CORE
# -------------------------------------------------------------
ws2 = wb.create_sheet(title="P1-ADHD Second Brain")
ws2.views.sheetView[0].showGridLines = True

ws2['A1'] = "Product 1: ADHD Advanced Second Brain Core Hub"
ws2['A1'].font = title_font
ws2['A2'] = "Designed for dopamine-driven capture, hyperfocus grouping, and instant structural filtering."
ws2['A2'].font = hint_font

# Quick Capture Interface
ws2['A4'] = "[ QUICK CAPTURE INBOX ]"
ws2['A4'].font = section_font
headers_p1 = ["Capture ID", "Idea/Task Description", "Initial Processing Category", "Dopamine/Reward Tier", "Action Link"]
for col_idx, text in enumerate(headers_p1, start=1):
    cell = ws2.cell(row=5, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

p1_data = [
    ("QC-001", "Research framework for the app architecture modules", "Projects", "High", "LINK-T07"),
    ("QC-002", "Buy meal prep containers for next week's clean eating plan", "Health & Nutrition", "Low", "LINK-T15"),
    ("QC-003", "Fix bug in multi-language translation localization module", "Active Sprints", "High", "LINK-T10"),
    ("QC-004", "Review continuous subscription overhead and recurring ledger", "Finance", "Medium", "LINK-T13")
]
for row_idx, data in enumerate(p1_data, start=6):
    for col_idx, val in enumerate(data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        if col_idx == 4 and val == "High":
            cell.fill = accent_fill_3

# Active Hyperfocus Sandbox section
ws2['A12'] = "[ ACTIVE HYPERFOCUS SANDBOX ]"
ws2['A12'].font = section_font
ws2['A13'] = "Project Focus:"
ws2['B13'] = "Mobile Application Development MVP"
ws2['B13'].font = Font(name=font_family, size=11, bold=True)
ws2['A14'] = "Current Energy Match Filter:"
ws2['B14'] = "High Energy - Creative Mode"
ws2['B14'].fill = accent_fill_2

# -------------------------------------------------------------
# TAB 3: EXECUTABLE PRODUCT 2 - TIME BLOCKING MATRIX
# -------------------------------------------------------------
ws3 = wb.create_sheet(title="P2-Time Block Matrix")
ws3.views.sheetView[0].showGridLines = True

ws3['A1'] = "Product 2: Tablet-Optimized Daily Time Blocking Matrix"
ws3['A1'].font = title_font

headers_p2 = ["Time Slot", "Scheduled Activity / Focus Block", "Linked Context Block", "Energy Required", "Done"]
for col_idx, text in enumerate(headers_p2, start=1):
    cell = ws3.cell(row=4, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

p2_data = [
    ("07:00 - 08:00", "Morning Routine & Hydration Track", "Life Routine", "Low", "Y"),
    ("08:00 - 10:00", "Deep Work: Architecture Engine & Templates", "Hyperfocus Sandbox", "High", "Y"),
    ("10:00 - 10:30", "Inbox Sweep & Processing Dump", "Quick Notes Inbox", "Medium", "Y"),
    ("10:30 - 12:30", "UI/UX Layout Adaptation for Tablet Views", "Projects", "High", ""),
    ("12:30 - 13:30", "Nutritional Re-fuel & Walk", "Health & Nutrition", "Low", ""),
    ("13:30 - 15:30", "Localization Verification & Translation Matrix", "Active Sprints", "Medium", "")
]
for row_idx, data in enumerate(p2_data, start=5):
    for col_idx, val in enumerate(data, start=1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        if col_idx == 5:
            cell.alignment = Alignment(horizontal="center")

# -------------------------------------------------------------
# TAB 4: EXECUTABLE PRODUCT 3 - RECURRING ROUTINES ARCHITECT
# -------------------------------------------------------------
ws4 = wb.create_sheet(title="P3-Routine Architect")
ws4.views.sheetView[0].showGridLines = True

ws4['A1'] = "Product 3: Recurring Routines & Habit Loops"
ws4['A1'].font = title_font

headers_p3 = ["Habit Loop / Routine", "Target Cadence", "Cue Trigger", "Reward System Set", "Current Streak (Days)", "Completion Index"]
for col_idx, text in enumerate(headers_p3, start=1):
    cell = ws4.cell(row=4, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill

p3_data = [
    ("Morning Mind Dump", "Daily - 07:15", "Sitting down with morning tea", "15 mins guilt-free web browsing", 14, "93%"),
    ("Macro Meal Prep", "Weekly - Sun", "Calendar notification alarm", "Pre-funded cheat meal budget", 4, "100%"),
    ("Financial Ledger Review", "Bi-Weekly", "Paycheck clear notification", "Visualization of investment growth", 6, "85%"),
    ("Digital Decutter Sweep", "Monthly - Last Fri", "Desktop file stack exceeds 15", "Clean, anxiety-free desktop layout", 2, "66%")
]
for row_idx, data in enumerate(p3_data, start=5):
    for col_idx, val in enumerate(data, start=1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border

# -------------------------------------------------------------
# TAB 5: EXECUTABLE PRODUCT 4 - HOLISTIC LIFE RADAR
# -------------------------------------------------------------
ws5 = wb.create_sheet(title="P4-Holistic Life Radar")
ws5.views.sheetView[0].showGridLines = True

ws5['A1'] = "Product 4: Holistic Life Radar & Domain Hub"
ws5['A1'].font = title_font

headers_p4 = ["Life Category Domain", "Active Core Target / OKR", "Allocated Tracker Template", "Current Status", "Progress Meter"]
for col_idx, text in enumerate(headers_p4, start=1):
    cell = ws5.cell(row=4, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill

p4_data = [
    ("Career & Dev", "Launch standalone productivity application", "T10 - Kanban Sprint Board", "Active", "65%"),
    ("Personal Finance", "Maintain savings rate above 35%", "T13 - Financial Ledger & Budget", "On Track", "90%"),
    ("Health & Fitness", "Consistent 4-day workout distribution split", "T16 - Fitness & Recovery Log", "Needs Attention", "45%"),
    ("Knowledge / Intellect", "Synthesize 12 core books into knowledge base", "T06 - Structured Knowledge Vault", "Active", "50%")
]
for row_idx, data in enumerate(p4_data, start=5):
    for col_idx, val in enumerate(data, start=1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        if col_idx == 4:
            if val == "On Track": cell.fill = accent_fill_2
            if val == "Needs Attention": cell.fill = accent_fill_3

# -------------------------------------------------------------
# TAB 6: EXECUTABLE PRODUCT 5 & GLOBAL TRANSLATION LOCALIZATION MATRIX
# -------------------------------------------------------------
ws6 = wb.create_sheet(title="P5-Translation Engine")
ws6.views.sheetView[0].showGridLines = True

ws6['A1'] = "Product 5 & Global Multi-Language Translation Localization Engine"
ws6['A1'].font = title_font
ws6['A2'] = "Provides absolute semantic keys mapped directly to specialized target languages for prompt dynamic switching."
ws6['A2'].font = hint_font

headers_p6 = ["System Key ID", "English (EN)", "Greek (EL)", "Swedish (SV)", "Finnish (FI)", "Dutch (NL)", "Danish (DA)", "Russian (RU)"]
for col_idx, text in enumerate(headers_p6, start=1):
    cell = ws6.cell(row=4, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

translation_data = [
    ("TXT_INBOX", "Quick Inbox", "Γρήγορα Εισερχόμενα", "Snabbinkorg", "Pikainboksi", "Snel Postvak In", "Hurtig Indbakke", "Быстрое входящее"),
    ("TXT_SECBRAIN", "Second Brain", "Δεύτερος Εγκέφαλος", "Andra Hjärna", "Toiset Aivot", "Tweede Brein", "Anden Hjerne", "Второй мозг"),
    ("TXT_HYPERFOCUS", "Hyperfocus Zone", "Ζώνη Υπερεστίασης", "Hyperfokuszon", "Hyperfokusointialue", "Hyperfocus Zone", "Hyperfokus Zone", "Зона гиперфокуса"),
    ("TXT_DAILY_MAT", "Daily Matrix", "Ημερήσιος Πίνακας", "Daglig Matris", "Päivittäinen Matriisi", "Dagelijkse Matrix", "Daglig Matrix", "Ежедневная матрица"),
    ("TXT_TASKS", "Task Management", "Διαχείριση Εργασιών", "Uppgiftshantering", "Tehtävienhallinta", "Taakbeheer", "Opgavestyring", "Управление задачами"),
    ("TXT_ROUTINES", "Habit Loops", "Βρόχοι Συνηθειών", "Vanemönster", "Tottumussilmukat", "Gewoonte Loops", "Vaneloops", "Циклы привычек"),
    ("TXT_FINANCE", "Financial Ledger", "Οικονομικό Καθολικό", "Finansiell Ledge", "Talouskirjanpito", "Financieel Grootboek", "Finansiel Ledger", "Финансовая книга"),
    ("TXT_ENERGY", "Energy Filter", "Φίλτρο Ενέργειας", "Energifilter", "Energiasuodatin", "Energie Filter", "Energifilter", "Фильтр энергии"),
    ("TXT_STATUS", "Status Tracker", "Παρακολούθηση Κατάστασης", "Statusmätare", "Tilaseuranta", "Status Tracker", "Status Tracker", "Отслеживание статуса"),
    ("TXT_DUMP", "Brain Dump", "Εκκένωση Μυαλού", "Hjärndump", "Aivoriihi / Tyhjennys", "Brain Dump", "Brain Dump", "Разгрузка мозга")
]

for row_idx, data in enumerate(translation_data, start=5):
    for col_idx, val in enumerate(data, start=1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = thin_border
        if col_idx == 1:
            cell.font = Font(name=font_family, size=10, bold=True, color="4A5568")
            cell.fill = accent_fill_1

# Auto-adjust column widths uniformly across all sheets
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Basic length handling with safety limit
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

# Save the configured structural engine workbook
wb.save("application_architecture_and_templates.xlsx")