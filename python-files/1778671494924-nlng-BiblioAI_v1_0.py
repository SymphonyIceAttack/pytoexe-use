"""
BiblioAI v3.5 — PySide6 Desktop App
+ Keywords extraction from search APIs
+ Screening tab uses Title, Keywords, Abstract (instead of Authors)
"""

import sys
import os
import re
import glob
import time
import json
import hashlib
import pickle
import multiprocessing
import platform
import tempfile
import xml.etree.ElementTree as ET
from io import BytesIO, StringIO
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import pandas as pd
import feedparser
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QTextEdit, QComboBox, QCheckBox,
    QSpinBox, QSlider, QProgressBar, QTabWidget, QTableWidget,
    QTableWidgetItem, QFileDialog, QGroupBox, QScrollArea, QSplitter,
    QFrame, QMessageBox, QSizePolicy, QStatusBar, QAbstractItemView,
    QHeaderView, QPlainTextEdit, QRadioButton, QButtonGroup, QDoubleSpinBox
)
from PySide6.QtCore import (
    Qt, QThread, Signal, QObject, QTimer, QSettings
)
from PySide6.QtGui import QFont, QColor, QPalette, QBrush, QIcon, QPalette

# ─────────────────────────────────────────────
#  dependencies 
# ─────────────────────────────────────────────
try:
    import cloudscraper
    CLOUDSCRAPER_AVAILABLE = True
except ImportError:
    CLOUDSCRAPER_AVAILABLE = False

try:
    import ollama as _ollama_lib
    OLLAMA_AVAILABLE = True
except ImportError:
    OLLAMA_AVAILABLE = False

try:
    from fulltext_article_downloader import download_article
    FULLTEXT_AVAILABLE = True
except ImportError:
    FULLTEXT_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import litellm
    from litellm import completion
    LITELLM_AVAILABLE = True
except ImportError:
    LITELLM_AVAILABLE = False

# macOS fork safety
if __name__ == '__main__':
    try:
        if multiprocessing.get_start_method(allow_none=True) != 'spawn':
            multiprocessing.set_start_method('spawn', force=True)
    except RuntimeError:
        pass
os.environ["OBJC_DISABLE_INITIALIZE_FORK_SAFETY"] = "YES"

# =============================================
# search, download, AI, export
# =============================================

def safe_str(value, default="N/A"):
    if value is None or value is Ellipsis:
        return default
    return str(value)

def clean_abstract(text):
    if not text:
        return "No abstract available."
    clean = re.sub(r'<[^>]+>', '', str(text))
    return " ".join(clean.split())

def extract_keywords_openalex(item):
    """Extract concept names from OpenAlex 'concepts' list"""
    concepts = item.get('concepts', [])
    keywords = [c.get('display_name', '') for c in concepts if c.get('display_name')]
    return ", ".join(keywords) if keywords else ""

def extract_keywords_scopus(entry):
    """Scopus: try authkeywords or index terms"""
    keywords = []
    # check for authkeywords
    if 'authkeywords' in entry:
        kw_list = entry['authkeywords'].split('|')
        keywords.extend([k.strip() for k in kw_list if k.strip()])
    # also check for idxterms
    if 'idxterms' in entry:
        for term in entry['idxterms'].get('term', []):
            if isinstance(term, str):
                keywords.append(term)
    return ", ".join(keywords) if keywords else ""

def extract_keywords_crossref(item):
    """Crossref often lacks keywords; try 'subject' or 'keyword' array"""
    keywords = []
    if 'subject' in item and isinstance(item['subject'], list):
        keywords.extend(item['subject'])
    if 'keyword' in item and isinstance(item['keyword'], list):
        keywords.extend([kw.get('value', '') for kw in item['keyword'] if kw.get('value')])
    return ", ".join(keywords) if keywords else ""

def extract_keywords_semantic(paper):
    """Semantic Scholar: fieldsOfStudy"""
    fields = paper.get('fieldsOfStudy', [])
    return ", ".join(fields) if fields else ""

def extract_keywords_arxiv(entry):
    """arXiv: categories"""
    if hasattr(entry, 'tags'):
        cats = [tag['term'] for tag in entry.tags if 'term' in tag]
        return ", ".join(cats)
    return ""

def extract_keywords_pubmed(article):
    """PubMed: MeSH headings"""
    keywords = []
    mesh_headings = article.findall(".//MeshHeading/DescriptorName")
    for mh in mesh_headings:
        if mh.text:
            keywords.append(mh.text)
    return ", ".join(keywords) if keywords else ""

def search_openalex(query, limit=500):
    all_data = []
    per_page = min(200, limit)
    pages = (limit + per_page - 1) // per_page
    for page in range(1, pages + 1):
        url = "https://api.openalex.org/works"
        params = {"search": query, "per-page": per_page, "page": page}
        try:
            r = requests.get(url, params=params, timeout=10)
            if r.status_code != 200:
                continue
            results = r.json().get('results', [])
            if not results:
                break
            for item in results:
                title = item.get('display_name') or item.get('title') or "No title"
                authors_list = []
                for auth in item.get('authorships', []):
                    author = auth.get('author', {})
                    name = author.get('display_name') or author.get('name')
                    if name:
                        authors_list.append(name)
                authors = ", ".join(authors_list) if authors_list else "N/A"
                journal = item.get('host_venue', {}).get('display_name') or "N/A"
                year = str(item.get('publication_year', '')) if item.get('publication_year') else ""
                doi_raw = item.get('doi', '')
                doi = doi_raw.replace('https://doi.org/', '') if doi_raw else ''
                abstract = "No abstract available."
                inv_idx = item.get('abstract_inverted_index')
                if inv_idx:
                    words = {}
                    for word, pos_list in inv_idx.items():
                        for pos in pos_list:
                            words[pos] = word
                    abstract = " ".join(words[i] for i in sorted(words.keys()))
                # Keywords extraction
                keywords = extract_keywords_openalex(item)
                all_data.append({
                    'Source': 'OpenAlex',
                    'Title': safe_str(title, "No title"),
                    'Authors': safe_str(authors, "N/A"),
                    'Journal': safe_str(journal, "N/A"),
                    'Year': safe_str(year, ""),
                    'DOI': safe_str(doi, ""),
                    'Abstract': safe_str(clean_abstract(abstract), "No abstract available."),
                    'Keywords': keywords
                })
            time.sleep(0.5)
            if len(all_data) >= limit:
                all_data = all_data[:limit]
                break
        except Exception:
            continue
    return all_data

def search_scopus(query, api_key, limit=500):
    if not api_key:
        return []
    all_data = []
    per_page = min(200, limit)
    start = 0
    url = "https://api.elsevier.com/content/search/scopus"
    headers = {"X-ELS-APIKey": api_key, "Accept": "application/json"}
    while start < limit:
        params = {"query": query, "count": per_page, "start": start}
        try:
            r = requests.get(url, headers=headers, params=params, timeout=10)
            if r.status_code in (401, 403):
                break
            if r.status_code == 429:
                break
            if r.status_code != 200:
                break
            data = r.json().get('search-results', {})
            entries = data.get('entry', [])
            if not entries:
                break
            for item in entries:
                keywords = extract_keywords_scopus(item)
                all_data.append({
                    'Source': 'Scopus',
                    'Title': safe_str(item.get('dc:title'), "No title"),
                    'Authors': safe_str(item.get('dc:creator'), "N/A"),
                    'Journal': safe_str(item.get('prism:publicationName'), "N/A"),
                    'Year': safe_str(item.get('prism:coverDate', '')[:4], ""),
                    'DOI': safe_str(item.get('prism:doi'), ""),
                    'Abstract': "Restricted by Scopus (API does not return abstract)",
                    'Keywords': keywords
                })
            start += per_page
            time.sleep(0.5)
            if len(all_data) >= limit:
                all_data = all_data[:limit]
                break
        except Exception:
            break
    return all_data

def search_crossref(query, limit=500):
    all_data = []
    per_page = min(100, limit)
    offset = 0
    url = "https://api.crossref.org/works"
    while offset < limit:
        params = {"query": query, "rows": per_page, "offset": offset}
        try:
            r = requests.get(url, params=params, timeout=15)
            if r.status_code != 200:
                break
            items = r.json().get("message", {}).get("items", [])
            if not items:
                break
            for item in items:
                title = item.get("title", [""])[0] if item.get("title") else "No title"
                authors = []
                for a in item.get("author", []):
                    given = a.get("given", "")
                    family = a.get("family", "")
                    authors.append(f"{given} {family}".strip())
                abstract = item.get("abstract", "No abstract available.")
                if abstract:
                    abstract = clean_abstract(abstract)
                keywords = extract_keywords_crossref(item)
                all_data.append({
                    'Source': 'Crossref',
                    'Title': safe_str(title, "No title"),
                    'Authors': safe_str(", ".join(authors) if authors else "N/A"),
                    'Journal': safe_str(item.get("container-title", [""])[0] if item.get("container-title") else "N/A"),
                    'Year': safe_str(str(item.get("published-print", {}).get("date-parts", [[0]])[0][0]) if item.get("published-print") else "", ""),
                    'DOI': safe_str(item.get("DOI", ""), ""),
                    'Abstract': safe_str(abstract, "No abstract available."),
                    'Keywords': keywords
                })
            offset += per_page
            time.sleep(0.5)
            if len(all_data) >= limit:
                all_data = all_data[:limit]
                break
        except Exception:
            break
    return all_data

def search_semantic_scholar(query, limit=500):
    all_data = []
    per_page = min(100, limit)
    offset = 0
    url = "https://api.semanticscholar.org/graph/v1/paper/search"
    session = requests.Session()
    session.headers.update({"User-Agent": "BiblioAI/3.5"})
    while offset < limit:
        params = {
            "query": query, "limit": per_page, "offset": offset,
            "fields": "title,abstract,authors,venue,year,doi,fieldsOfStudy"
        }
        success = False
        for attempt in range(3):
            try:
                r = session.get(url, params=params, timeout=15)
                if r.status_code == 429:
                    wait = int(r.headers.get("Retry-After", 2 ** attempt))
                    time.sleep(wait)
                    continue
                if r.status_code != 200:
                    break
                data = r.json()
                papers = data.get("data", [])
                if not papers:
                    offset = limit
                    success = True
                    break
                for paper in papers:
                    authors = [a.get("name", "") for a in paper.get("authors", [])]
                    keywords = extract_keywords_semantic(paper)
                    all_data.append({
                        'Source': 'Semantic Scholar',
                        'Title': safe_str(paper.get("title", "No title")),
                        'Authors': safe_str(", ".join(authors) if authors else "N/A"),
                        'Journal': safe_str(paper.get("venue", "N/A")),
                        'Year': safe_str(str(paper.get("year", "")), ""),
                        'DOI': safe_str(paper.get("doi", ""), ""),
                        'Abstract': safe_str(clean_abstract(paper.get("abstract", "No abstract available."))),
                        'Keywords': keywords
                    })
                offset += per_page
                success = True
                break
            except Exception:
                time.sleep(1)
                continue
        if not success:
            offset += per_page
        time.sleep(0.5)
        if len(all_data) >= limit:
            all_data = all_data[:limit]
            break
    return all_data

def search_arxiv(query, limit=500):
    all_data = []
    per_page = min(100, limit)
    start = 0
    base_url = "http://export.arxiv.org/api/query"
    while start < limit:
        params = {
            "search_query": f"all:{query}",
            "start": start, "max_results": per_page, "sortBy": "relevance"
        }
        try:
            r = requests.get(base_url, params=params, timeout=15)
            if r.status_code != 200:
                break
            feed = feedparser.parse(r.text)
            entries = feed.entries
            if not entries:
                break
            for entry in entries:
                arxiv_id = entry.id.split('/abs/')[-1] if '/abs/' in entry.id else entry.id
                doi = f"10.48550/arXiv.{arxiv_id}"
                title = entry.title
                authors = ", ".join([author.name for author in entry.authors]) if hasattr(entry, 'authors') else "N/A"
                year = str(entry.published_parsed.tm_year) if hasattr(entry, 'published_parsed') else ""
                abstract = clean_abstract(entry.summary)
                keywords = extract_keywords_arxiv(entry)
                all_data.append({
                    'Source': 'arXiv',
                    'Title': safe_str(title, "No title"),
                    'Authors': safe_str(authors, "N/A"),
                    'Journal': "arXiv preprint",
                    'Year': safe_str(year, ""),
                    'DOI': safe_str(doi, ""),
                    'Abstract': safe_str(abstract, "No abstract available."),
                    'Keywords': keywords
                })
            start += per_page
            time.sleep(0.5)
            if len(all_data) >= limit:
                all_data = all_data[:limit]
                break
        except Exception:
            break
    return all_data

def search_pubmed(query, limit=500):
    all_data = []
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    fetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    search_params = {"db": "pubmed", "term": query, "retmax": min(limit, 10000), "retmode": "json"}
    try:
        r = requests.get(base_url, params=search_params, timeout=15)
        if r.status_code != 200:
            return []
        data = r.json()
        ids = data.get("esearchresult", {}).get("idlist", [])
        if not ids:
            return []
        batch_size = 200
        for i in range(0, min(len(ids), limit), batch_size):
            batch_ids = ids[i:i+batch_size]
            fetch_params = {"db": "pubmed", "id": ",".join(batch_ids), "retmode": "xml"}
            resp = requests.get(fetch_url, params=fetch_params, timeout=30)
            if resp.status_code != 200:
                continue
            root = ET.fromstring(resp.content)
            for article in root.findall(".//PubmedArticle"):
                title_elem = article.find(".//ArticleTitle")
                title = title_elem.text if title_elem is not None else "No title"
                abstract_elem = article.find(".//Abstract/AbstractText")
                abstract = abstract_elem.text if abstract_elem is not None else "No abstract available."
                journal_elem = article.find(".//Journal/Title")
                journal = journal_elem.text if journal_elem is not None else "N/A"
                year_elem = article.find(".//PubDate/Year")
                year = year_elem.text if year_elem is not None else "N/A"
                doi_elem = article.find(".//ArticleId[@IdType='doi']")
                doi = doi_elem.text if doi_elem is not None else ""
                authors = []
                for author in article.findall(".//Author"):
                    last = author.find("LastName")
                    fore = author.find("ForeName")
                    if last is not None:
                        name = last.text
                        if fore is not None:
                            name = f"{fore.text} {name}"
                        authors.append(name)
                keywords = extract_keywords_pubmed(article)
                all_data.append({
                    'Source': 'PubMed',
                    'Title': safe_str(title, "No title"),
                    'Authors': safe_str(", ".join(authors) if authors else "N/A"),
                    'Journal': safe_str(journal, "N/A"),
                    'Year': safe_str(year, ""),
                    'DOI': safe_str(doi, ""),
                    'Abstract': safe_str(clean_abstract(abstract), "No abstract available."),
                    'Keywords': keywords
                })
            time.sleep(0.5)
            if len(all_data) >= limit:
                all_data = all_data[:limit]
                break
    except Exception:
        pass
    return all_data

# ── Download function ───────────────────────

def download_one_pdf(doi, scopus_key, wiley_token, download_dir):
    safe_name = doi.replace('/', '_').replace(':', '_') + '.pdf'
    filepath = os.path.join(download_dir, safe_name)
    if os.path.exists(filepath):
        return doi, True, "Already downloaded"
    if not FULLTEXT_AVAILABLE:
        return doi, False, "fulltext-article-downloader not installed"
    try:
        output_path = download_article(doi, output_dir=download_dir, output_filename=safe_name)
        if output_path and os.path.exists(output_path):
            return doi, True, "Success"
        else:
            return doi, False, "Library could not retrieve the article"
    except Exception as e:
        return doi, False, f"Download error: {str(e)}"

# ── PDF text extraction & caching ───────────

def extract_pdf_text_fast(pdf_path, max_pages=None, max_chars=20000):
    try:
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        pages_to_read = doc if max_pages is None or max_pages == 0 else doc[:max_pages]
        text_parts = [page.get_text() for page in pages_to_read if page.get_text()]
        doc.close()
        text = "\n".join(text_parts)
        if max_chars and len(text) > max_chars:
            text = text[:max_chars] + "\n...[truncated]..."
        return text, total_pages
    except:
        return None, 0

def get_cached_text(pdf_path, cache_dir, max_pages, max_chars):
    mtime = os.path.getmtime(pdf_path)
    cache_key = hashlib.md5(f"{pdf_path}_{mtime}_{max_pages}_{max_chars}".encode()).hexdigest()
    cache_file = os.path.join(cache_dir, cache_key)
    if os.path.exists(cache_file):
        try:
            with open(cache_file, "rb") as f:
                return pickle.load(f)
        except:
            pass
    text, pages = extract_pdf_text_fast(pdf_path, max_pages, max_chars)
    if text is not None:
        with open(cache_file, "wb") as f:
            pickle.dump((text, pages), f)
    return text, pages

# ── Prompt builder ───────────────────────────

def get_prompt(prompt_mode, sum_length="standard", template="exec", custom_query=""):
    if prompt_mode == "Quick Summary":
        return f"Summarize in {sum_length} sentences."
    elif prompt_mode == "Template":
        templates = {
            "Executive Summary": "Provide an executive summary (Purpose, Findings, Implications).",
            "Key Findings": "List the key findings with supporting evidence.",
            "Methods": "Describe the methodology and data collection techniques.",
            "Research Questions": "What research questions does this address?",
            "Limitations": "What limitations and future work are mentioned?"
        }
        return templates.get(template, "Summarize.")
    return custom_query if custom_query.strip() else "Summarize."

# ── LLM call ─────────────────────────────────

def call_llm(prompt, backend, model_name=None, ollama_model=None,
             litellm_model=None, api_key=None, provider=None):
    if backend == "ollama":
        if not OLLAMA_AVAILABLE:
            return ("Ollama Python library not installed. "
                    "Run: pip install ollama  — then restart the app.")
        try:
            response = _ollama_lib.chat(
                model=ollama_model,
                messages=[{"role": "user", "content": prompt}])
            return response["message"]["content"]
        except Exception as e:
            return f"Ollama error: {str(e)}"
    elif backend == "litellm":
        if not LITELLM_AVAILABLE:
            return "LiteLLM not installed. Please run: pip install litellm"
        if not litellm_model or not api_key:
            return "Missing LiteLLM model name or API key."
        if provider == "openai":
            os.environ["OPENAI_API_KEY"] = api_key
        elif provider == "anthropic":
            os.environ["ANTHROPIC_API_KEY"] = api_key
        elif provider == "groq":
            os.environ["GROQ_API_KEY"] = api_key
        elif provider == "google":
            os.environ["GOOGLE_API_KEY"] = api_key
        else:
            os.environ["OPENAI_API_KEY"] = api_key
        try:
            response = completion(model=litellm_model,
                                  messages=[{"role": "user", "content": prompt}])
            return response["choices"][0]["message"]["content"]
        except Exception as e:
            return f"LiteLLM error: {str(e)}"
    return "Unknown backend."

# ── AI processing functions ──────────────────

def process_one_pdf_structured(pdf_path, base_prompt, output_spec,
                                backend_config, max_pages, max_chars, cache_dir):
    content, num_pages = get_cached_text(pdf_path, cache_dir, max_pages, max_chars)
    if not content:
        return {"file": os.path.basename(pdf_path), "pages": num_pages,
                "status": "ERROR", "data": None}
    full_prompt = f"""{base_prompt}

OUTPUT FORMAT INSTRUCTION:
{output_spec}

Based on the above description, generate a JSON object for each paper with the exact column names as keys.
Return a SINGLE JSON object with keys being column names and values being the extracted information.
Example for a specification "Columns: Title, Finding, YES/NO question": 
{{"Title": "...", "Finding": "...", "YES/NO question": "YES"}}
Only return valid JSON. Do not add extra text.

CONTENT:
{content}"""
    raw = call_llm(full_prompt, **backend_config)
    cleaned = re.sub(r'```json\s*|\s*```', '', raw)
    start = cleaned.find('{')
    end = cleaned.rfind('}')
    if start != -1 and end != -1:
        json_str = cleaned[start:end+1]
        try:
            data = json.loads(json_str)
            return {"file": os.path.basename(pdf_path), "pages": num_pages,
                    "status": "OK", "data": data}
        except:
            json_str = re.sub(r',\s*}', '}', json_str)
            json_str = re.sub(r',\s*]', ']', json_str)
            try:
                data = json.loads(json_str)
                return {"file": os.path.basename(pdf_path), "pages": num_pages,
                        "status": "OK", "data": data}
            except:
                pass
    return {"file": os.path.basename(pdf_path), "pages": num_pages,
            "status": "PARSE_ERROR", "raw": raw, "data": None}

def process_one_pdf_unstructured(pdf_path, base_prompt, backend_config,
                                  max_pages, max_chars, cache_dir):
    content, num_pages = get_cached_text(pdf_path, cache_dir, max_pages, max_chars)
    if not content:
        return {"file": os.path.basename(pdf_path), "pages": num_pages,
                "response": "Extraction failed", "status": "ERROR"}
    full_prompt = f"{base_prompt}\n\nCONTENT:\n{content}"
    resp = call_llm(full_prompt, **backend_config)
    return {"file": os.path.basename(pdf_path), "pages": num_pages,
            "response": resp, "status": "OK"}

# ── Export functions ─────────────────────────

def create_excel_with_formatting(results, output_spec):
    all_keys = set()
    valid_results = [r for r in results if r.get("data")]
    for r in valid_results:
        all_keys.update(r["data"].keys())
    if not all_keys:
        return None
    rows = []
    for r in valid_results:
        row = {"File": r["file"], "Pages": r["pages"]}
        for k in all_keys:
            val = r["data"].get(k, "")
            if isinstance(val, str):
                val_upper = val.strip().upper()
                if val_upper in ["YES", "Y", "TRUE"]:
                    val = "YES"
                elif val_upper in ["NO", "N", "FALSE"]:
                    val = "NO"
            row[k] = val
        rows.append(row)
    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Analysis', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Analysis']
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A'
            unique_vals = df[col_name].dropna().astype(str).str.upper().unique()
            if set(unique_vals).issubset({"YES", "NO", ""}):
                start_row = 2
                end_row = start_row + len(df) - 1
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                worksheet.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    FormulaRule(formula=[f'{col_letter}{start_row}="YES"'], fill=green_fill))
                worksheet.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    FormulaRule(formula=[f'{col_letter}{start_row}="NO"'], fill=red_fill))
    output.seek(0)
    return output

def create_csv(results, output_spec):
    all_keys = set()
    valid_results = [r for r in results if r.get("data")]
    for r in valid_results:
        all_keys.update(r["data"].keys())
    if not all_keys:
        return None
    rows = []
    for r in valid_results:
        row = {"File": r["file"], "Pages": r["pages"]}
        for k in all_keys:
            row[k] = r["data"].get(k, "")
        rows.append(row)
    df = pd.DataFrame(rows)
    output = StringIO()
    df.to_csv(output, index=False)
    return output.getvalue()

def create_pdf_report(results, output_spec):
    if not REPORTLAB_AVAILABLE:
        return None, "ReportLab not installed."
    valid_results = [r for r in results if r.get("data")]
    if not valid_results:
        return None, "No valid data"
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    try:
        if platform.system() == "Windows":
            font_path = "C:/Windows/Fonts/arial.ttf"
        elif platform.system() == "Darwin":
            font_path = "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"
        else:
            font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        pdfmetrics.registerFont(TTFont('UnicodeFont', font_path))
        style_normal = ParagraphStyle('CustomNormal', parent=styles['Normal'],
                                       fontName='UnicodeFont', fontSize=10)
        style_heading = ParagraphStyle('CustomHeading', parent=styles['Heading1'],
                                        fontName='UnicodeFont', fontSize=14)
    except:
        style_normal = styles['Normal']
        style_heading = styles['Heading1']
    story = []
    story.append(Paragraph("BiblioAI Analysis Report", style_heading))
    story.append(Spacer(1, 0.2 * inch))
    for r in valid_results:
        story.append(Paragraph(f"<b>File:</b> {r['file']} (Pages: {r['pages']})", style_normal))
        story.append(Spacer(1, 0.1 * inch))
        for k, v in r['data'].items():
            safe_k = str(k).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            safe_v = str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            story.append(Paragraph(f"<b>{safe_k}:</b> {safe_v}", style_normal))
        story.append(Spacer(1, 0.2 * inch))
    doc.build(story)
    pdf_data = buffer.getvalue()
    buffer.close()
    return pdf_data, None

# ─────────────────────────────────────────────
# Worker threads (background tasks)
# ─────────────────────────────────────────────

class SearchWorker(QThread):
    progress = Signal(str)
    finished = Signal(list)

    def __init__(self, query, limit, db_flags, scopus_key):
        super().__init__()
        self.query = query
        self.limit = limit
        self.db_flags = db_flags
        self.scopus_key = scopus_key

    def run(self):
        all_results = []
        flags = self.db_flags
        if flags.get("openalex"):
            self.progress.emit("Searching OpenAlex…")
            all_results.extend(search_openalex(self.query, self.limit))
        if flags.get("scopus") and self.scopus_key:
            self.progress.emit("Searching Scopus…")
            all_results.extend(search_scopus(self.query, self.scopus_key, self.limit))
        if flags.get("crossref"):
            self.progress.emit("Searching Crossref…")
            all_results.extend(search_crossref(self.query, self.limit))
        if flags.get("semantic"):
            self.progress.emit("Searching Semantic Scholar…")
            all_results.extend(search_semantic_scholar(self.query, self.limit))
        if flags.get("arxiv"):
            self.progress.emit("Searching arXiv…")
            all_results.extend(search_arxiv(self.query, self.limit))
        if flags.get("pubmed"):
            self.progress.emit("Searching PubMed…")
            all_results.extend(search_pubmed(self.query, self.limit))
        self.finished.emit(all_results)

class DownloadWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(list)

    def __init__(self, dois, scopus_key, wiley_token, download_dir, max_workers):
        super().__init__()
        self.dois = dois
        self.scopus_key = scopus_key
        self.wiley_token = wiley_token
        self.download_dir = download_dir
        self.max_workers = max_workers

    def run(self):
        results = []
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {
                executor.submit(download_one_pdf, doi, self.scopus_key,
                                self.wiley_token, self.download_dir): doi
                for doi in self.dois
            }
            for i, future in enumerate(as_completed(futures)):
                doi, success, msg = future.result()
                results.append(f"{doi}: {msg}")
                self.progress.emit(i + 1, len(self.dois), msg)
        self.finished.emit(results)

class AIWorker(QThread):
    progress = Signal(int, int)
    finished = Signal(list)

    def __init__(self, pdf_files, base_prompt, output_spec, backend_config,
                 max_pages, max_chars, max_workers, cache_dir, structured):
        super().__init__()
        self.pdf_files = pdf_files
        self.base_prompt = base_prompt
        self.output_spec = output_spec
        self.backend_config = backend_config
        self.max_pages = max_pages
        self.max_chars = max_chars
        self.max_workers = max_workers
        self.cache_dir = cache_dir
        self.structured = structured

    def run(self):
        results = []
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            if self.structured:
                futures = [
                    executor.submit(process_one_pdf_structured, f, self.base_prompt,
                                    self.output_spec, self.backend_config,
                                    self.max_pages, self.max_chars, self.cache_dir)
                    for f in self.pdf_files
                ]
            else:
                futures = [
                    executor.submit(process_one_pdf_unstructured, f, self.base_prompt,
                                    self.backend_config, self.max_pages,
                                    self.max_chars, self.cache_dir)
                    for f in self.pdf_files
                ]
            for i, fut in enumerate(as_completed(futures)):
                results.append(fut.result())
                self.progress.emit(i + 1, len(futures))
        self.finished.emit(results)

# ── NEW: ScreeningWorker for YES/NO classification using Keywords ─────────────────
class ScreeningWorker(QThread):
    progress = Signal(int, int, str)   # current, total, status_message
    finished = Signal(list)            # list of dicts with Screening_Result

    def __init__(self, papers, user_prompt, backend_config, max_workers):
        super().__init__()
        self.papers = papers
        self.user_prompt = user_prompt
        self.backend_config = backend_config
        self.max_workers = max_workers

    def run(self):
        results = []
        total = len(self.papers)

        def screen_one(paper):
            title = paper.get("Title", "")
            keywords = paper.get("Keywords", "")
            abstract = paper.get("Abstract", "")
            prompt = f"""You are a strict assistant that answers only YES or NO.
Based on the following paper information, answer the user's question with exactly one word: YES or NO.
Do not add any explanation or extra text.

User question: {self.user_prompt}

Paper title: {title}
Paper keywords: {keywords}
Paper abstract: {abstract}

Answer (YES/NO):"""

            raw = call_llm(prompt, **self.backend_config)
            raw_clean = raw.strip().upper()
            if "YES" in raw_clean and "NO" not in raw_clean[:4]:
                result = "YES"
            elif "NO" in raw_clean:
                result = "NO"
            else:
                result = "ERROR"
            return {
                "DOI": paper.get("DOI", ""),
                "Title": title,
                "Keywords": keywords,
                "Abstract": abstract,
                "Screening_Result": result
            }

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(screen_one, p): idx for idx, p in enumerate(self.papers)}
            for i, future in enumerate(as_completed(futures)):
                res = future.result()
                results.append(res)
                self.progress.emit(i + 1, total, f"Screened {i+1}/{total} – {res['Screening_Result']}")
        self.finished.emit(results)

# =============================================
# THEMES
# =============================================

DARK_STYLE = """
QMainWindow, QWidget {
    background-color: #1e1e2e;
    color: #cdd6f4;
}
QTabWidget::pane {
    border: 1px solid #313244;
    background-color: #181825;
    border-radius: 8px;
}
QTabBar::tab {
    background-color: #2a2a3a;
    color: #a6adc8;
    padding: 8px 18px;
    border-radius: 6px;
    margin: 2px;
}
QTabBar::tab:selected {
    background-color: #3b3b5c;
    color: #ffffff;
}
QPushButton {
    background-color: #313244;
    color: #cdd6f4;
    border: 1px solid #45475a;
    border-radius: 8px;
    padding: 6px 14px;
    font-weight: normal;
}
QPushButton:hover {
    background-color: #3b3b5c;
    border-color: #585b70;
}
QPushButton#primary {
    background-color: #6c5ce7;
    color: white;
    border: none;
}
QPushButton#primary:hover {
    background-color: #5a4bcf;
}
QPushButton#danger {
    background-color: #d14b4b;
    color: white;
    border: none;
}
QPushButton#danger:hover {
    background-color: #b33e3e;
}
QLineEdit, QTextEdit, QPlainTextEdit, QComboBox {
    background-color: #2a2a3a;
    color: #cdd6f4;
    border: 1px solid #3a3a4a;
    border-radius: 6px;
    padding: 5px 8px;
}
QComboBox:hover {
    border-color: #6c5ce7;
}
QComboBox::drop-down {
    border: none;
    width: 24px;
}
QComboBox::down-arrow {
    image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid #a6adc8;
    margin-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #2a2a3a;
    color: #cdd6f4;
    selection-background-color: #3b3b5c;
    selection-color: #ffffff;
    border: 1px solid #3a3a4a;
    border-radius: 6px;
}
QCheckBox, QRadioButton {
    color: #cdd6f4;
    spacing: 6px;
}
QGroupBox {
    border: 1px solid #313244;
    border-radius: 8px;
    margin-top: 8px;
    padding-top: 8px;
    color: #b4befe;
    font-weight: normal;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 6px;
}
QTableWidget {
    background-color: #181825;
    color: #cdd6f4;
    gridline-color: #2a2a3a;
    border: 1px solid #313244;
    border-radius: 6px;
}
QTableWidget::item:selected {
    background-color: #3b3b5c;
    color: white;
}
QHeaderView::section {
    background-color: #2a2a3a;
    color: #a6adc8;
    border: 1px solid #313244;
    padding: 5px;
}
QProgressBar {
    background-color: #2a2a3a;
    border: 1px solid #3a3a4a;
    border-radius: 6px;
    text-align: center;
    color: #cdd6f4;
}
QProgressBar::chunk {
    background-color: #6c5ce7;
    border-radius: 5px;
}
QScrollBar:vertical {
    background-color: #1e1e2e;
    width: 10px;
}
QScrollBar::handle:vertical {
    background-color: #3a3a4a;
    border-radius: 5px;
    min-height: 20px;
}
QLabel#metric {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
        stop:0 #5a4bcf, stop:1 #6c5ce7);
    border-radius: 14px;
    padding: 12px;
    color: white;
    font-size: 14px;
    font-weight: 500;
}
QStatusBar {
    background-color: #181825;
    color: #a6adc8;
    border-top: 1px solid #313244;
}
"""

LIGHT_STYLE = """
QMainWindow, QWidget {
    background-color: #f4f6f9;
    color: #1e293b;
}
QTabWidget::pane {
    border: 1px solid #d1d5db;
    background-color: #ffffff;
    border-radius: 8px;
}
QTabBar::tab {
    background-color: #e2e8f0;
    color: #475569;
    padding: 8px 18px;
    border-radius: 6px;
    margin: 2px;
}
QTabBar::tab:selected {
    background-color: #cbd5e1;
    color: #0f172a;
}
QPushButton {
    background-color: #ffffff;
    color: #1e293b;
    border: 1px solid #cbd5e1;
    border-radius: 8px;
    padding: 6px 14px;
    font-weight: normal;
}
QPushButton:hover {
    background-color: #f1f5f9;
    border-color: #94a3b8;
}
QPushButton#primary {
    background-color: #3b82f6;
    color: white;
    border: none;
}
QPushButton#primary:hover {
    background-color: #2563eb;
}
QPushButton#danger {
    background-color: #ef4444;
    color: white;
    border: none;
}
QPushButton#danger:hover {
    background-color: #dc2626;
}
QLineEdit, QTextEdit, QPlainTextEdit, QComboBox {
    background-color: #ffffff;
    color: #1e293b;
    border: 1px solid #cbd5e1;
    border-radius: 6px;
    padding: 5px 8px;
}
QComboBox:hover {
    border-color: #3b82f6;
}
QComboBox::drop-down {
    border: none;
    width: 24px;
}
QComboBox::down-arrow {
    image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid #64748b;
    margin-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #ffffff;
    color: #1e293b;
    selection-background-color: #e2e8f0;
    selection-color: #0f172a;
    border: 1px solid #cbd5e1;
    border-radius: 6px;
}
QCheckBox, QRadioButton {
    color: #1e293b;
    spacing: 6px;
}
QGroupBox {
    border: 1px solid #d1d5db;
    border-radius: 8px;
    margin-top: 8px;
    padding-top: 8px;
    color: #475569;
    font-weight: normal;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 6px;
}
QTableWidget {
    background-color: #ffffff;
    color: #1e293b;
    gridline-color: #e2e8f0;
    border: 1px solid #d1d5db;
    border-radius: 6px;
}
QTableWidget::item:selected {
    background-color: #dbeafe;
    color: #1e40af;
}
QHeaderView::section {
    background-color: #f8fafc;
    color: #475569;
    border: 1px solid #e2e8f0;
    padding: 5px;
}
QProgressBar {
    background-color: #e2e8f0;
    border: 1px solid #cbd5e1;
    border-radius: 6px;
    text-align: center;
    color: #1e293b;
}
QProgressBar::chunk {
    background-color: #3b82f6;
    border-radius: 5px;
}
QScrollBar:vertical {
    background-color: #f1f5f9;
    width: 10px;
}
QScrollBar::handle:vertical {
    background-color: #cbd5e1;
    border-radius: 5px;
    min-height: 20px;
}
QLabel#metric {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
        stop:0 #3b82f6, stop:1 #06b6d4);
    border-radius: 14px;
    padding: 12px;
    color: white;
    font-size: 14px;
    font-weight: 500;
}
QStatusBar {
    background-color: #ffffff;
    color: #64748b;
    border-top: 1px solid #e2e8f0;
}
"""

# ─────────────────────────────────────────────
# Main Window
# ─────────────────────────────────────────────

class BiblioAIWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📚 BiblioAI v3.5")
        self.setMinimumSize(1200, 800)

        # ── App state ──
        self.settings = QSettings("BiblioAI", "v3.5")
        self.base_work_dir = self.settings.value(
            "base_work_dir", os.path.expanduser("~/BiblioMineR_data"))
        self._setup_dirs()

        self.results_df = None
        self.pdf_files = []
        self.ai_results = []
        self.dl_queue_dois = []
        self.screening_results = []   # store latest screening output

        # default dark
        self.current_theme = self.settings.value("theme", "dark")
        self.apply_theme(self.current_theme)  

        self._build_ui()

    # ── Directory management ────────────────

    def _setup_dirs(self):
        self.download_dir = os.path.join(self.base_work_dir, "downloads")
        self.cache_dir    = os.path.join(self.base_work_dir, "pdf_cache")
        self.saved_dir    = os.path.join(self.base_work_dir, "saved_searches")
        for d in [self.download_dir, self.cache_dir, self.saved_dir]:
            os.makedirs(d, exist_ok=True)

    # ── themes ─────────────────────────

    def apply_theme(self, theme):
        theme = theme.lower()
        if theme == "dark":
            stylesheet = DARK_STYLE
        elif theme == "light":
            stylesheet = LIGHT_STYLE
        else:
            if hasattr(QApplication.instance().styleHints(), 'colorScheme'):
                color_scheme = QApplication.instance().styleHints().colorScheme()
                if color_scheme == Qt.ColorScheme.Dark:
                    stylesheet = DARK_STYLE
                else:
                    stylesheet = LIGHT_STYLE
            else:
                palette = QApplication.palette()
                bg = palette.color(QPalette.Window)
                if bg.lightness() < 128:
                    stylesheet = DARK_STYLE
                else:
                    stylesheet = LIGHT_STYLE
        self.setStyleSheet(stylesheet)
        self.current_theme = theme
        self.settings.setValue("theme", theme)

    def _on_theme_changed(self, theme_text):
        if theme_text == "Dark":
            self.apply_theme("dark")
        elif theme_text == "Light":
            self.apply_theme("light")
        else:  # System
            self.apply_theme("system")

    # ── UI construction ─────────────────────

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(16, 12, 16, 8)
        root.setSpacing(8)

        # Header
        header = QLabel("📚 BiblioAI")
        header.setStyleSheet("font-size:26px; font-weight:700; color:#cdd6f4;")
        sub = QLabel("All-in-one solution for integrating AI into bibliographic research and screening")
        sub.setStyleSheet("font-size:13px; color:#a6adc8;")
        root.addWidget(header)
        root.addWidget(sub)

        # Metric bar
        self.lbl_pdfs     = QLabel("0\n📄 PDFs Ready")
        self.lbl_analyzed = QLabel("0\n✅ Analyzed")
        self.lbl_queue    = QLabel("0\n⚡ Queue")
        metric_row = QHBoxLayout()
        for lbl in (self.lbl_pdfs, self.lbl_analyzed, self.lbl_queue):
            lbl.setObjectName("metric")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setMinimumHeight(60)
            metric_row.addWidget(lbl)
        root.addLayout(metric_row)

        # Main splitter: sidebar | tabs
        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(2)
        splitter.addWidget(self._build_sidebar())
        splitter.addWidget(self._build_tabs())
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([280, 900])
        root.addWidget(splitter, 1)

        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage(
            "BiblioAI v1.0")

    # ── Sidebar (con selettore tema) ─────────────────────────────

    def _build_sidebar(self):
        sidebar = QScrollArea()
        sidebar.setWidgetResizable(True)
        sidebar.setMaximumWidth(300)
        sidebar.setMinimumWidth(260)

        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)

        # Tema
        grp_theme = QGroupBox("🎨 Theme interface")
        theme_layout = QVBoxLayout(grp_theme)
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Dark", "Light", "System"])
        if self.current_theme == "dark":
            self.theme_combo.setCurrentText("Dark")
        elif self.current_theme == "light":
            self.theme_combo.setCurrentText("Light")
        else:
            self.theme_combo.setCurrentText("System")
        self.theme_combo.currentTextChanged.connect(self._on_theme_changed)
        theme_layout.addWidget(self.theme_combo)
        theme_layout.addWidget(QLabel("Change app theme"))
        lay.addWidget(grp_theme)

        # Working directory
        grp_dir = QGroupBox("📁 Working Directory")
        g = QVBoxLayout(grp_dir)
        self.txt_workdir = QLineEdit(self.base_work_dir)
        btn_workdir = QPushButton("Apply")
        btn_workdir.clicked.connect(self._apply_workdir)
        g.addWidget(self.txt_workdir)
        g.addWidget(btn_workdir)
        lay.addWidget(grp_dir)

        # API Keys
        grp_keys = QGroupBox("🔑 API Keys (optional)")
        gk = QVBoxLayout(grp_keys)
        self.txt_scopus_key  = QLineEdit(); self.txt_scopus_key.setPlaceholderText("Scopus API Key")
        self.txt_scopus_key.setEchoMode(QLineEdit.Password)
        self.txt_wiley_token = QLineEdit(); self.txt_wiley_token.setPlaceholderText("Wiley TDM Token")
        self.txt_wiley_token.setEchoMode(QLineEdit.Password)
        self.txt_unpaywall   = QLineEdit(); self.txt_unpaywall.setPlaceholderText("Unpaywall Email")
        gk.addWidget(self.txt_scopus_key)
        gk.addWidget(self.txt_wiley_token)
        gk.addWidget(self.txt_unpaywall)
        lay.addWidget(grp_keys)

        # Database selection
        grp_db = QGroupBox("🔍 Databases")
        gd = QVBoxLayout(grp_db)
        self.chk_openalex = QCheckBox("OpenAlex");         self.chk_openalex.setChecked(True)
        self.chk_scopus   = QCheckBox("Scopus (API key)"); self.chk_scopus.setChecked(False)
        self.chk_crossref = QCheckBox("Crossref");         self.chk_crossref.setChecked(True)
        self.chk_semantic = QCheckBox("Semantic Scholar"); self.chk_semantic.setChecked(True)
        self.chk_arxiv    = QCheckBox("arXiv (preprints)");self.chk_arxiv.setChecked(True)
        self.chk_pubmed   = QCheckBox("PubMed (biomedical)");self.chk_pubmed.setChecked(True)
        for c in (self.chk_openalex, self.chk_scopus, self.chk_crossref,
                  self.chk_semantic, self.chk_arxiv, self.chk_pubmed):
            gd.addWidget(c)
        lay.addWidget(grp_db)

        # LLM Backend
        grp_llm = QGroupBox("🤖 LLM Backend")
        gl = QVBoxLayout(grp_llm)
        self.radio_ollama  = QRadioButton("Local Ollama")
        self.radio_litellm = QRadioButton("Cloud (LiteLLM)")
        self.radio_ollama.setChecked(True)
        self.radio_ollama.toggled.connect(self._toggle_llm_ui)
        gl.addWidget(self.radio_ollama)
        gl.addWidget(self.radio_litellm)

        # Ollama widgets
        self.ollama_widget = QWidget()
        ow = QVBoxLayout(self.ollama_widget)
        ow.setContentsMargins(0, 0, 0, 0)
        if not OLLAMA_AVAILABLE:
            warn = QLabel("⚠️ ollama not installed.\npip install ollama")
            warn.setStyleSheet("color:#f38ba8; font-size:11px;")
            warn.setWordWrap(True)
            ow.addWidget(warn)
        self.txt_ollama_url = QLineEdit("http://localhost:11434")
        self.cmb_ollama_model = QComboBox()
        self.cmb_ollama_model.addItems(["llama3.2", "gemma:2b", "phi4", "mistral", "llama3.2:3b"])
        ow.addWidget(QLabel("Ollama URL:"))
        ow.addWidget(self.txt_ollama_url)
        ow.addWidget(QLabel("Model:"))
        ow.addWidget(self.cmb_ollama_model)
        gl.addWidget(self.ollama_widget)

        # LiteLLM widgets
        self.litellm_widget = QWidget()
        lw = QVBoxLayout(self.litellm_widget)
        lw.setContentsMargins(0,0,0,0)
        self.cmb_provider = QComboBox()
        self.cmb_provider.addItems(["openai","anthropic","groq","google"])
        self.txt_llm_api_key  = QLineEdit(); self.txt_llm_api_key.setPlaceholderText("API Key")
        self.txt_llm_api_key.setEchoMode(QLineEdit.Password)
        self.txt_llm_model    = QLineEdit(); self.txt_llm_model.setPlaceholderText("e.g., gpt-3.5-turbo")
        lw.addWidget(QLabel("Provider:")); lw.addWidget(self.cmb_provider)
        lw.addWidget(QLabel("API Key:"));  lw.addWidget(self.txt_llm_api_key)
        lw.addWidget(QLabel("Model:"));    lw.addWidget(self.txt_llm_model)
        gl.addWidget(self.litellm_widget)
        self.litellm_widget.setVisible(False)
        lay.addWidget(grp_llm)

        # Performance
        grp_perf = QGroupBox("⚡ Performance")
        gp = QVBoxLayout(grp_perf)
        gp.addWidget(QLabel("Parallel Threads (AI):"))
        self.spin_threads = QSpinBox(); self.spin_threads.setRange(1,8); self.spin_threads.setValue(4)
        gp.addWidget(self.spin_threads)
        gp.addWidget(QLabel("Max Pages (0=all):"))
        self.spin_maxpages = QSpinBox(); self.spin_maxpages.setRange(0,500); self.spin_maxpages.setValue(5)
        gp.addWidget(self.spin_maxpages)
        gp.addWidget(QLabel("Max Characters:"))
        self.spin_maxchars = QSpinBox(); self.spin_maxchars.setRange(1000,100000); self.spin_maxchars.setValue(20000)
        self.spin_maxchars.setSingleStep(1000)
        gp.addWidget(self.spin_maxchars)
        lay.addWidget(grp_perf)

        lay.addStretch()
        sidebar.setWidget(w)
        return sidebar

    def _toggle_llm_ui(self, ollama_selected):
        self.ollama_widget.setVisible(ollama_selected)
        self.litellm_widget.setVisible(not ollama_selected)

    def _apply_workdir(self):
        new = self.txt_workdir.text().strip()
        if not new:
            return
        self.base_work_dir = new
        self.settings.setValue("base_work_dir", new)
        self._setup_dirs()
        self.status_bar.showMessage(f"Working directory set to: {new}")

    # ── Tabs ─────────────────────────────────

    def _build_tabs(self):
        tabs = QTabWidget()
        tabs.addTab(self._build_discovery_tab(),  "🔍 Discovery")
        tabs.addTab(self._build_screening_tab(),  "🔬 LLM Screening")   # NEW
        tabs.addTab(self._build_download_tab(),   "📥 Download Queue")
        tabs.addTab(self._build_ai_tab(),         "🧠 AI Analysis")
        return tabs

    # ── Discovery tab (with Keywords column) ──────────

    def _build_discovery_tab(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)

        row = QHBoxLayout()
        self.txt_query = QLineEdit()
        self.txt_query.setPlaceholderText("e.g., 'deep learning' AND 'healthcare'")
        self.txt_query.returnPressed.connect(self._start_search)
        self.spin_limit = QSpinBox()
        self.spin_limit.setRange(5, 2000); self.spin_limit.setValue(500)
        self.spin_limit.setFixedWidth(90)
        btn_search = QPushButton("🚀 Search All")
        btn_search.setObjectName("primary")
        btn_search.clicked.connect(self._start_search)
        row.addWidget(QLabel("Query:")); row.addWidget(self.txt_query, 1)
        row.addWidget(QLabel("Max/DB:")); row.addWidget(self.spin_limit)
        row.addWidget(btn_search)
        lay.addLayout(row)

        self.search_progress = QProgressBar()
        self.search_progress.setRange(0, 0)
        self.search_progress.setVisible(False)
        self.search_status = QLabel("")
        lay.addWidget(self.search_progress)
        lay.addWidget(self.search_status)

        self.results_table = QTableWidget()
        self.results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.setSortingEnabled(True)
        lay.addWidget(self.results_table, 1)

        bottom = QHBoxLayout()
        self.txt_save_name = QLineEdit()
        self.txt_save_name.setPlaceholderText("Save name (without .csv)")
        btn_save = QPushButton("💾 Save")
        btn_save.clicked.connect(self._save_search)
        self.cmb_load = QComboBox()
        self._refresh_saved_list()
        btn_load = QPushButton("📂 Load")
        btn_load.clicked.connect(self._load_search)
        btn_queue = QPushButton("📌 Add all to Download Queue")
        btn_queue.setObjectName("primary")
        btn_queue.clicked.connect(self._add_all_to_queue)
        bottom.addWidget(self.txt_save_name)
        bottom.addWidget(btn_save)
        bottom.addWidget(self.cmb_load)
        bottom.addWidget(btn_load)
        bottom.addStretch()
        bottom.addWidget(btn_queue)
        lay.addLayout(bottom)

        return w

    def _start_search(self):
        query = self.txt_query.text().strip()
        if not query:
            self.status_bar.showMessage("Enter a search query first.")
            return
        scopus_key = self.txt_scopus_key.text().strip()
        if scopus_key:
            os.environ["ELSEVIER_API_KEY"] = scopus_key
        wiley = self.txt_wiley_token.text().strip()
        if wiley:
            os.environ["WILEY_API_KEY"] = wiley
        email = self.txt_unpaywall.text().strip()
        if email:
            os.environ["UNPAYWALL_EMAIL"] = email

        db_flags = {
            "openalex":  self.chk_openalex.isChecked(),
            "scopus":    self.chk_scopus.isChecked(),
            "crossref":  self.chk_crossref.isChecked(),
            "semantic":  self.chk_semantic.isChecked(),
            "arxiv":     self.chk_arxiv.isChecked(),
            "pubmed":    self.chk_pubmed.isChecked(),
        }
        self.search_progress.setVisible(True)
        self.search_status.setText("Starting search…")
        self.results_table.setRowCount(0)

        self._search_worker = SearchWorker(
            query, self.spin_limit.value(), db_flags, scopus_key)
        self._search_worker.progress.connect(lambda msg: self.search_status.setText(msg))
        self._search_worker.finished.connect(self._on_search_finished)
        self._search_worker.start()

    def _on_search_finished(self, all_results):
        self.search_progress.setVisible(False)
        if not all_results:
            self.search_status.setText("No results found.")
            return
        self.results_df = pd.DataFrame(all_results)
        self._populate_results_table(self.results_df)
        self.search_status.setText(f"✅ Found {len(all_results)} papers.")
        self._update_metrics()
        self._enable_screen_button()   # enable screening tab button

    def _populate_results_table(self, df):
        self.results_table.setRowCount(0)
        cols = list(df.columns)
        self.results_table.setColumnCount(len(cols))
        self.results_table.setHorizontalHeaderLabels(cols)
        for row_idx, row in df.iterrows():
            r = self.results_table.rowCount()
            self.results_table.insertRow(r)
            for col_idx, col in enumerate(cols):
                item = QTableWidgetItem(str(row[col]))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.results_table.setItem(r, col_idx, item)
        self.results_table.resizeColumnsToContents()

    def _save_search(self):
        name = self.txt_save_name.text().strip()
        if not name or self.results_df is None:
            self.status_bar.showMessage("Enter a name and run a search first.")
            return
        path = os.path.join(self.saved_dir, f"{name}.csv")
        self.results_df.to_csv(path, index=False)
        self.status_bar.showMessage(f"Saved to {path}")
        self._refresh_saved_list()

    def _load_search(self):
        selected = self.cmb_load.currentText()
        if not selected or selected == "(none)":
            return
        path = os.path.join(self.saved_dir, f"{selected}.csv")
        if not os.path.exists(path):
            return
        self.results_df = pd.read_csv(path)
        self._populate_results_table(self.results_df)
        self.status_bar.showMessage(f"Loaded: {selected}")
        self._update_metrics()
        self._enable_screen_button()   # enable screening tab button

    def _refresh_saved_list(self):
        self.cmb_load.clear()
        files = glob.glob(os.path.join(self.saved_dir, "*.csv"))
        names = [os.path.basename(f).replace('.csv', '') for f in files]
        self.cmb_load.addItem("(none)")
        self.cmb_load.addItems(names)

    def _add_all_to_queue(self):
        if self.results_df is None:
            return
        dois = self.results_df['DOI'].dropna().unique().tolist()
        dois = [d for d in dois if d]
        self.dl_queue_dois = list(set(self.dl_queue_dois + dois))
        self.txt_dl_queue.setPlainText("\n".join(self.dl_queue_dois))
        self.status_bar.showMessage(f"Added {len(dois)} DOIs to download queue.")
        self._update_metrics()

    # ── New Screening Tab (using Keywords instead of Authors) ──────────

    def _build_screening_tab(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)

        info = QLabel(
            "Run a custom LLM screening on the current search results.\n"
            "The model will answer YES or NO based on your question and each paper's title, keywords, and abstract."
        )
        info.setWordWrap(True)
        lay.addWidget(info)

        grp_prompt = QGroupBox("✍️ Screening Question (YES/NO)")
        gl = QVBoxLayout(grp_prompt)
        self.screening_prompt = QTextEdit()
        self.screening_prompt.setPlaceholderText(
            "Example: Does this paper propose a novel deep learning architecture for medical imaging?"
        )
        self.screening_prompt.setMaximumHeight(100)
        gl.addWidget(self.screening_prompt)
        lay.addWidget(grp_prompt)

        btn_row = QHBoxLayout()
        self.btn_screen = QPushButton("▶️ Run LLM Screening")
        self.btn_screen.setObjectName("primary")
        self.btn_screen.clicked.connect(self._run_screening)
        self.btn_screen.setEnabled(False)
        self.screen_progress = QProgressBar()
        self.screen_progress.setVisible(False)
        self.screen_status = QLabel("")
        btn_row.addWidget(self.btn_screen)
        btn_row.addWidget(self.screen_progress)
        btn_row.addStretch()
        lay.addLayout(btn_row)
        lay.addWidget(self.screen_status)

        self.screen_table = QTableWidget()
        self.screen_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.screen_table.setSortingEnabled(True)
        lay.addWidget(self.screen_table, 1)

        bottom = QHBoxLayout()
        self.btn_export_screen = QPushButton("📥 Export Screening Results (CSV/Excel)")
        self.btn_export_screen.clicked.connect(self._export_screening)
        self.btn_export_screen.setEnabled(False)
        self.btn_transfer_yes = QPushButton("✅ Add YES DOIs to Download Queue")
        self.btn_transfer_yes.setObjectName("primary")
        self.btn_transfer_yes.clicked.connect(self._transfer_yes_to_queue)
        self.btn_transfer_yes.setEnabled(False)
        bottom.addWidget(self.btn_export_screen)
        bottom.addWidget(self.btn_transfer_yes)
        lay.addLayout(bottom)

        return w

    def _enable_screen_button(self):
        has_results = self.results_df is not None and not self.results_df.empty
        self.btn_screen.setEnabled(has_results)

    def _run_screening(self):
        if self.results_df is None or self.results_df.empty:
            QMessageBox.warning(self, "No data", "Please run a search first.")
            return

        prompt = self.screening_prompt.toPlainText().strip()
        if not prompt:
            QMessageBox.warning(self, "Missing prompt", "Please enter a screening question.")
            return

        # Build papers list using Title, Keywords, Abstract
        papers = []
        for _, row in self.results_df.iterrows():
            papers.append({
                "DOI": row.get("DOI", ""),
                "Title": row.get("Title", ""),
                "Keywords": row.get("Keywords", ""),
                "Abstract": row.get("Abstract", "")
            })

        backend_config = self._get_backend_config()
        if backend_config["backend"] == "litellm" and (not backend_config.get("litellm_model") or not backend_config.get("api_key")):
            QMessageBox.critical(self, "Config Error", "Please provide LiteLLM model name and API key.")
            return

        self.screen_progress.setMaximum(len(papers))
        self.screen_progress.setValue(0)
        self.screen_progress.setVisible(True)
        self.screen_status.setText("Running screening...")
        self.btn_screen.setEnabled(False)
        self.btn_export_screen.setEnabled(False)
        self.btn_transfer_yes.setEnabled(False)

        self._screen_worker = ScreeningWorker(
            papers, prompt, backend_config, self.spin_threads.value()
        )
        self._screen_worker.progress.connect(self._on_screen_progress)
        self._screen_worker.finished.connect(self._on_screen_finished)
        self._screen_worker.start()

    def _on_screen_progress(self, current, total, msg):
        self.screen_progress.setValue(current)
        self.screen_status.setText(msg)

    def _on_screen_finished(self, results):
        self.screening_results = results
        self.screen_progress.setVisible(False)
        self.btn_screen.setEnabled(True)
        self.btn_export_screen.setEnabled(True)
        self.btn_transfer_yes.setEnabled(True)

        # Populate table
        self.screen_table.setRowCount(0)
        cols = ["DOI", "Title", "Keywords", "Abstract", "Screening_Result"]
        self.screen_table.setColumnCount(len(cols))
        self.screen_table.setHorizontalHeaderLabels(cols)
        for r in results:
            row = self.screen_table.rowCount()
            self.screen_table.insertRow(row)
            self.screen_table.setItem(row, 0, QTableWidgetItem(r["DOI"]))
            self.screen_table.setItem(row, 1, QTableWidgetItem(r["Title"]))
            self.screen_table.setItem(row, 2, QTableWidgetItem(r["Keywords"]))
            self.screen_table.setItem(row, 3, QTableWidgetItem(r["Abstract"]))
            res_item = QTableWidgetItem(r["Screening_Result"])
            if r["Screening_Result"] == "YES":
                res_item.setBackground(QColor("#C6EFCE"))
                res_item.setForeground(QColor("#276221"))
            elif r["Screening_Result"] == "NO":
                res_item.setBackground(QColor("#FFC7CE"))
                res_item.setForeground(QColor("#9C0006"))
            else:
                res_item.setBackground(QColor("#FFEB9C"))
            self.screen_table.setItem(row, 4, res_item)
        self.screen_table.resizeColumnsToContents()
        yes_count = sum(1 for r in results if r["Screening_Result"] == "YES")
        self.screen_status.setText(f"Screening completed. {yes_count} papers marked YES.")

    def _export_screening(self):
        if not self.screening_results:
            QMessageBox.warning(self, "No data", "No screening results to export.")
            return
        df = pd.DataFrame(self.screening_results)
        fmt, _ = QFileDialog.getSaveFileName(self, "Save Screening Results", "screening_results.csv", "CSV (*.csv);;Excel (*.xlsx)")
        if not fmt:
            return
        if fmt.endswith(".csv"):
            df.to_csv(fmt, index=False)
        elif fmt.endswith(".xlsx"):
            with pd.ExcelWriter(fmt, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="Screening", index=False)
                worksheet = writer.sheets["Screening"]
                for row in range(2, len(df)+2):
                    cell = worksheet.cell(row, 5)  # col E = Screening_Result
                    if cell.value == "YES":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif cell.value == "NO":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.status_bar.showMessage(f"Exported to {fmt}")

    def _transfer_yes_to_queue(self):
        if not self.screening_results:
            QMessageBox.warning(self, "No data", "Run screening first.")
            return
        yes_dois = [r["DOI"] for r in self.screening_results if r["Screening_Result"] == "YES" and r["DOI"]]
        if not yes_dois:
            QMessageBox.information(self, "No YES papers", "No DOIs with YES result were found.")
            return
        # Merge with existing queue (avoid duplicates)
        existing = set(self.dl_queue_dois)
        new_dois = [d for d in yes_dois if d not in existing]
        self.dl_queue_dois.extend(new_dois)
        # Update the download queue text area (if already built)
        if hasattr(self, 'txt_dl_queue'):
            self.txt_dl_queue.setPlainText("\n".join(self.dl_queue_dois))
        self.status_bar.showMessage(f"Added {len(new_dois)} YES DOIs to download queue.")
        self._update_metrics()

    # ── Download Tab ─────────────────────────────

    def _build_download_tab(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)

        lay.addWidget(QLabel("Queue (one DOI per line):"))
        self.txt_dl_queue = QPlainTextEdit()
        self.txt_dl_queue.setPlaceholderText("10.1000/xyz123\n10.1016/j.example...")
        lay.addWidget(self.txt_dl_queue, 1)

        row = QHBoxLayout()
        self.spin_dl_threads = QSpinBox(); self.spin_dl_threads.setRange(1,10); self.spin_dl_threads.setValue(5)
        row.addWidget(QLabel("Parallel downloads:")); row.addWidget(self.spin_dl_threads); row.addStretch()
        lay.addLayout(row)

        btn_row = QHBoxLayout()
        btn_dl  = QPushButton("▶️ Start Bulk Download")
        btn_dl.setObjectName("primary")
        btn_dl.clicked.connect(self._start_download)
        btn_clr = QPushButton("🗑️ Clear Queue")
        btn_clr.setObjectName("danger")
        btn_clr.clicked.connect(self._clear_queue)
        btn_row.addWidget(btn_dl); btn_row.addWidget(btn_clr)
        lay.addLayout(btn_row)

        self.dl_progress = QProgressBar(); self.dl_progress.setVisible(False)
        self.dl_status   = QLabel("")
        lay.addWidget(self.dl_progress)
        lay.addWidget(self.dl_status)

        lay.addWidget(QLabel("Download log:"))
        self.dl_log = QPlainTextEdit(); self.dl_log.setReadOnly(True); self.dl_log.setMaximumHeight(160)
        lay.addWidget(self.dl_log)

        return w

    def _start_download(self):
        if not FULLTEXT_AVAILABLE:
            QMessageBox.critical(self, "Missing library",
                "fulltext-article-downloader is not installed.\n"
                "Run: pip install git+https://github.com/computron/fulltext-article-downloader.git")
            return
        raw = self.txt_dl_queue.toPlainText()
        dois = list(set([d.strip() for d in raw.splitlines() if d.strip()]))
        existing = {f.replace('_', '/').replace('.pdf', '')
                    for f in os.listdir(self.download_dir) if f.endswith('.pdf')}
        dois = [d for d in dois if d not in existing]
        if not dois:
            self.dl_status.setText("All DOIs already downloaded.")
            return
        self.dl_progress.setMaximum(len(dois))
        self.dl_progress.setValue(0)
        self.dl_progress.setVisible(True)
        self.dl_log.clear()

        scopus_key  = self.txt_scopus_key.text().strip()
        wiley_token = self.txt_wiley_token.text().strip()

        self._dl_worker = DownloadWorker(
            dois, scopus_key, wiley_token,
            self.download_dir, self.spin_dl_threads.value())
        self._dl_worker.progress.connect(self._on_dl_progress)
        self._dl_worker.finished.connect(self._on_dl_finished)
        self._dl_worker.start()

    def _on_dl_progress(self, done, total, msg):
        self.dl_progress.setValue(done)
        self.dl_status.setText(f"Progress: {done}/{total} — {msg}")
        self.dl_log.appendPlainText(msg)

    def _on_dl_finished(self, results):
        self.dl_progress.setVisible(False)
        ok = sum(1 for r in results if "Success" in r)
        self.dl_status.setText(f"Done. {ok}/{len(results)} succeeded.")

    def _clear_queue(self):
        self.txt_dl_queue.clear()
        self.dl_queue_dois = []
        self._update_metrics()

    # ── AI Analysis Tab (unchanged) ─────────────────────

    def _build_ai_tab(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)

        row = QHBoxLayout()
        self.txt_ai_folder = QLineEdit(self.download_dir)
        btn_browse = QPushButton("Browse…")
        btn_browse.clicked.connect(self._browse_ai_folder)
        btn_scan   = QPushButton("🔍 Scan for PDFs")
        btn_scan.clicked.connect(self._scan_pdfs)
        row.addWidget(QLabel("📁 Folder:"))
        row.addWidget(self.txt_ai_folder, 1)
        row.addWidget(btn_browse)
        row.addWidget(btn_scan)
        lay.addLayout(row)

        self.lbl_pdf_count = QLabel("PDFs found: 0")
        lay.addWidget(self.lbl_pdf_count)

        grp_prompt = QGroupBox("✍️ 1. Content Prompt")
        gp = QVBoxLayout(grp_prompt)

        mode_row = QHBoxLayout()
        self.radio_quick    = QRadioButton("Quick Summary")
        self.radio_template = QRadioButton("Template")
        self.radio_custom   = QRadioButton("Custom Query")
        self.radio_quick.setChecked(True)
        bg = QButtonGroup(self); bg.addButton(self.radio_quick)
        bg.addButton(self.radio_template); bg.addButton(self.radio_custom)
        for r in (self.radio_quick, self.radio_template, self.radio_custom):
            r.toggled.connect(self._update_prompt_ui)
            mode_row.addWidget(r)
        gp.addLayout(mode_row)

        sub_row = QHBoxLayout()
        self.cmb_sum_len = QComboBox(); self.cmb_sum_len.addItems(["3","5","10"])
        self.cmb_sum_len.setCurrentIndex(1)
        self.cmb_template = QComboBox()
        self.cmb_template.addItems(["Executive Summary","Key Findings","Methods",
                                     "Research Questions","Limitations"])
        self.cmb_template.setVisible(False)
        self.txt_custom_query = QTextEdit(); self.txt_custom_query.setMaximumHeight(60)
        self.txt_custom_query.setPlaceholderText("Free-text query…")
        self.txt_custom_query.setVisible(False)
        sub_row.addWidget(QLabel("Summary length:")); sub_row.addWidget(self.cmb_sum_len)
        sub_row.addWidget(self.cmb_template)
        gp.addLayout(sub_row)
        gp.addWidget(self.txt_custom_query)
        lay.addWidget(grp_prompt)

        grp_out = QGroupBox("📊 2. Output Formatting")
        go = QVBoxLayout(grp_out)
        self.chk_formatting = QCheckBox("Enable custom output format (table / Excel / PDF)")
        self.chk_formatting.toggled.connect(self._toggle_format_ui)
        go.addWidget(self.chk_formatting)

        self.format_widget = QWidget()
        fw = QVBoxLayout(self.format_widget)
        fw.setContentsMargins(0,0,0,0)
        fw.addWidget(QLabel("Output table specification:"))
        self.txt_output_spec = QTextEdit()
        self.txt_output_spec.setPlaceholderText(
            "Example: Columns: Paper Title, Main Finding, Supports Hypothesis (YES/NO). "
            "Color YES cells green, NO cells red.")
        self.txt_output_spec.setMaximumHeight(80)
        fw.addWidget(self.txt_output_spec)
        fw.addWidget(QLabel("Export format:"))
        self.cmb_file_type = QComboBox()
        self.cmb_file_type.addItems(["Excel (with colors)", "CSV"])
        if REPORTLAB_AVAILABLE:
            self.cmb_file_type.addItem("PDF report")
        fw.addWidget(self.cmb_file_type)
        self.format_widget.setVisible(False)
        go.addWidget(self.format_widget)
        lay.addWidget(grp_out)

        btn_row = QHBoxLayout()
        self.btn_run_ai = QPushButton("🚀 Start AI Batch Analysis")
        self.btn_run_ai.setObjectName("primary")
        self.btn_run_ai.clicked.connect(self._start_ai)
        btn_row.addWidget(self.btn_run_ai)
        lay.addLayout(btn_row)

        self.ai_progress = QProgressBar(); self.ai_progress.setVisible(False)
        self.ai_status   = QLabel("")
        lay.addWidget(self.ai_progress)
        lay.addWidget(self.ai_status)

        self.ai_results_table = QTableWidget()
        self.ai_results_table.horizontalHeader().setStretchLastSection(True)
        self.ai_results_table.setSortingEnabled(True)
        lay.addWidget(self.ai_results_table, 1)

        self.btn_export = QPushButton("📥 Export Results")
        self.btn_export.setObjectName("primary")
        self.btn_export.setVisible(False)
        self.btn_export.clicked.connect(self._export_ai_results)
        lay.addWidget(self.btn_export)

        return w

    def _browse_ai_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder", self.download_dir)
        if folder:
            self.txt_ai_folder.setText(folder)

    def _scan_pdfs(self):
        folder = self.txt_ai_folder.text().strip()
        self.pdf_files = glob.glob(os.path.join(folder, "*.pdf"))
        self.lbl_pdf_count.setText(f"PDFs found: {len(self.pdf_files)}")
        self.ai_results = []
        self._update_metrics()

    def _update_prompt_ui(self):
        quick = self.radio_quick.isChecked()
        tmpl  = self.radio_template.isChecked()
        cust  = self.radio_custom.isChecked()
        self.cmb_sum_len.setVisible(quick)
        self.cmb_template.setVisible(tmpl)
        self.txt_custom_query.setVisible(cust)

    def _toggle_format_ui(self, enabled):
        self.format_widget.setVisible(enabled)

    def _get_backend_config(self):
        if self.radio_ollama.isChecked():
            os.environ["OLLAMA_HOST"] = self.txt_ollama_url.text().strip()
            return {
                "backend": "ollama",
                "ollama_model": self.cmb_ollama_model.currentText()
            }
        else:
            return {
                "backend": "litellm",
                "litellm_model": self.txt_llm_model.text().strip(),
                "api_key": self.txt_llm_api_key.text().strip(),
                "provider": self.cmb_provider.currentText()
            }

    def _start_ai(self):
        if not self.pdf_files:
            QMessageBox.warning(self, "No PDFs", "Click 'Scan for PDFs' first.")
            return
        backend_config = self._get_backend_config()
        if (backend_config["backend"] == "litellm" and
                (not backend_config.get("litellm_model") or not backend_config.get("api_key"))):
            QMessageBox.critical(self, "Config Error",
                                 "Please provide LiteLLM model name and API key.")
            return

        if self.radio_quick.isChecked():
            prompt_mode = "Quick Summary"
        elif self.radio_template.isChecked():
            prompt_mode = "Template"
        else:
            prompt_mode = "Custom Query"
        base_prompt = get_prompt(
            prompt_mode,
            self.cmb_sum_len.currentText(),
            self.cmb_template.currentText(),
            self.txt_custom_query.toPlainText()
        )

        structured  = self.chk_formatting.isChecked()
        output_spec = self.txt_output_spec.toPlainText() if structured else ""
        if structured and not output_spec.strip():
            QMessageBox.warning(self, "Missing spec",
                                "Please provide an output specification.")
            return

        self.ai_progress.setMaximum(len(self.pdf_files))
        self.ai_progress.setValue(0)
        self.ai_progress.setVisible(True)
        self.ai_status.setText("Running AI analysis…")
        self.btn_run_ai.setEnabled(False)
        self.btn_export.setVisible(False)

        self._ai_worker = AIWorker(
            self.pdf_files, base_prompt, output_spec, backend_config,
            self.spin_maxpages.value(), self.spin_maxchars.value(),
            self.spin_threads.value(), self.cache_dir, structured)
        self._ai_worker.progress.connect(
            lambda d, t: (self.ai_progress.setValue(d),
                          self.ai_status.setText(f"Analyzed {d}/{t}…")))
        self._ai_worker.finished.connect(self._on_ai_finished)
        self._ai_worker.start()

    def _on_ai_finished(self, results):
        self.ai_results = results
        self.ai_progress.setVisible(False)
        self.btn_run_ai.setEnabled(True)
        self._update_metrics()

        structured = self.chk_formatting.isChecked()
        if structured:
            valid = [r for r in results if r.get("data")]
            parse_errors = [r for r in results if r.get("status") == "PARSE_ERROR"]
            if parse_errors:
                self.ai_status.setText(
                    f"Done. {len(valid)} OK, {len(parse_errors)} parse errors.")
            else:
                self.ai_status.setText(f"Done. {len(valid)} papers analyzed.")
            if valid:
                all_keys = list({k for r in valid for k in r["data"].keys()})
                cols = ["File", "Pages"] + all_keys
                self.ai_results_table.setColumnCount(len(cols))
                self.ai_results_table.setHorizontalHeaderLabels(cols)
                self.ai_results_table.setRowCount(0)
                for r in valid:
                    row_idx = self.ai_results_table.rowCount()
                    self.ai_results_table.insertRow(row_idx)
                    self.ai_results_table.setItem(row_idx, 0, QTableWidgetItem(r["file"]))
                    self.ai_results_table.setItem(row_idx, 1, QTableWidgetItem(str(r["pages"])))
                    for ci, k in enumerate(all_keys, start=2):
                        val = str(r["data"].get(k, ""))
                        item = QTableWidgetItem(val)
                        if val.upper() == "YES":
                            item.setBackground(QColor("#C6EFCE"))
                            item.setForeground(QColor("#276221"))
                        elif val.upper() == "NO":
                            item.setBackground(QColor("#FFC7CE"))
                            item.setForeground(QColor("#9C0006"))
                        self.ai_results_table.setItem(row_idx, ci, item)
                self.ai_results_table.resizeColumnsToContents()
                self.btn_export.setVisible(True)
        else:
            self.ai_results_table.setColumnCount(3)
            self.ai_results_table.setHorizontalHeaderLabels(["File","Pages","Response"])
            self.ai_results_table.setRowCount(0)
            for r in results:
                ri = self.ai_results_table.rowCount()
                self.ai_results_table.insertRow(ri)
                self.ai_results_table.setItem(ri, 0, QTableWidgetItem(r["file"]))
                self.ai_results_table.setItem(ri, 1, QTableWidgetItem(str(r["pages"])))
                self.ai_results_table.setItem(ri, 2, QTableWidgetItem(r.get("response","")))
            self.ai_results_table.resizeColumnsToContents()
            self.ai_status.setText(f"Done. {len(results)} papers processed.")
            self.btn_export.setVisible(True)

    def _export_ai_results(self):
        fmt = self.cmb_file_type.currentText() if self.chk_formatting.isChecked() else "CSV"
        output_spec = self.txt_output_spec.toPlainText()

        if fmt == "Excel (with colors)":
            data = create_excel_with_formatting(self.ai_results, output_spec)
            if not data:
                QMessageBox.warning(self, "Export", "No valid structured data.")
                return
            path, _ = QFileDialog.getSaveFileName(
                self, "Save Excel", "analysis_report.xlsx",
                "Excel Files (*.xlsx)")
            if path:
                with open(path, "wb") as f:
                    f.write(data.read())
                self.status_bar.showMessage(f"Excel saved: {path}")

        elif fmt == "CSV":
            csv_data = create_csv(self.ai_results, output_spec)
            if not csv_data:
                df = pd.DataFrame(self.ai_results)
                csv_data = df.to_csv(index=False)
            path, _ = QFileDialog.getSaveFileName(
                self, "Save CSV", "analysis_report.csv",
                "CSV Files (*.csv)")
            if path:
                with open(path, "w", encoding="utf-8") as f:
                    f.write(csv_data)
                self.status_bar.showMessage(f"CSV saved: {path}")

        elif fmt == "PDF report":
            pdf_data, err = create_pdf_report(self.ai_results, output_spec)
            if not pdf_data:
                QMessageBox.warning(self, "Export", f"No PDF generated. {err or ''}")
                return
            path, _ = QFileDialog.getSaveFileName(
                self, "Save PDF", "analysis_report.pdf",
                "PDF Files (*.pdf)")
            if path:
                with open(path, "wb") as f:
                    f.write(pdf_data)
                self.status_bar.showMessage(f"PDF saved: {path}")

    def _update_metrics(self):
        self.lbl_pdfs.setText(f"{len(self.pdf_files)}\n📄 PDFs Ready")
        self.lbl_analyzed.setText(f"{len(self.ai_results)}\n✅ Analyzed")
        q = self.txt_dl_queue.toPlainText() if hasattr(self, 'txt_dl_queue') else ""
        n = len([l for l in q.splitlines() if l.strip()])
        self.lbl_queue.setText(f"{n}\n⚡ Queue")


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("BiblioAI")
    app.setOrganizationName("BiblioAI")
    win = BiblioAIWindow()
    win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()