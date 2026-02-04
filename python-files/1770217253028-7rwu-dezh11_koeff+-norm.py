import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import random
import math

def create_duty_schedule():
    # Загрузка данных из Excel файла
    try:
        # Читаем данные с листа 'Даты_дежурств'
        book = load_workbook('Дежурства.xlsx')
        ws_dates = book['Даты_дежурств']
        
        # Получаем праздничные дни из ячейки F11 (предполагается одностолбчатая таблица)
        holiday_dates = []
        holiday_cell = ws_dates['F12']
        
        if holiday_cell and holiday_cell.value:
            # Если в F11 есть значение, ищем таблицу праздничных дней
            current_row = 12
            while ws_dates[f'F{current_row}'].value is not None:
                cell_value = ws_dates[f'F{current_row}'].value
                try:
                    # Пробуем преобразовать в дату
                    if isinstance(cell_value, datetime):
                        holiday_date = cell_value
                    else:
                        # Пробуем разные форматы дат
                        try:
                            holiday_date = datetime.strptime(str(cell_value), '%d.%m.%Y')
                        except:
                            holiday_date = datetime.strptime(str(cell_value), '%Y-%m-%d')
                    
                    holiday_dates.append(holiday_date.date())
                    
                except Exception as e:
                    print(f"⚠ Предупреждение: не удалось распознать дату в ячейке F{current_row}: {cell_value}")
                
                current_row += 1
        
        # Читаем общее количество рабочих и выходных дней для распределения в строках "пспрс"
        try:
            total_workdays_to_distribute = int(ws_dates['P14'].value) if ws_dates['P14'].value is not None else 0
            total_weekends_to_distribute = int(ws_dates['Q14'].value) if ws_dates['Q14'].value is not None else 0
        except:
            print("⚠ Не удалось прочитать количество дней для распределения из ячеек P14, Q14")
            total_workdays_to_distribute = total_weekends_to_distribute = 0
        
        # Читаем весовые коэффициенты для ПС и ПРС в строках "пспрс"
        try:
            ps_weight = float(ws_dates['N8'].value) if ws_dates['N8'].value is not None else 0.5
            prs_weight = float(ws_dates['N9'].value) if ws_dates['N9'].value is not None else 0.5
        except:
            print("⚠ Не удалось прочитать весовые коэффициенты из ячеек N8, N9")
            ps_weight = prs_weight = 0.5
        
        # Нормализуем веса, чтобы сумма была 1
        total_weight = ps_weight + prs_weight
        if total_weight > 0:
            ps_weight_normalized = ps_weight / total_weight
            prs_weight_normalized = prs_weight / total_weight
        else:
            ps_weight_normalized = prs_weight_normalized = 0.5
        
        # Рассчитываем количество дежурств ПС и ПРС на основе весов (только для строк "пспрс")
        ps_workdays = int(round(total_workdays_to_distribute * ps_weight_normalized))
        prs_workdays = total_workdays_to_distribute - ps_workdays
        
        ps_weekends = int(round(total_weekends_to_distribute * ps_weight_normalized))
        prs_weekends = total_weekends_to_distribute - ps_weekends
        
        # Читаем количество дежурств ПС для строк с типом "пс" (старая логика)
        try:
            ps_for_ps_type_workdays = int(ws_dates['P16'].value) if ws_dates['P16'].value is not None else 0
            ps_for_ps_type_weekends = int(ws_dates['Q16'].value) if ws_dates['Q16'].value is not None else 0
        except:
            print("⚠ Не удалось прочитать количество дежурств ПС для строк 'пс' из ячеек P16, Q16")
            ps_for_ps_type_workdays = ps_for_ps_type_weekends = 0
        
        # Читаем кафедры и их весовые коэффициенты из столбцов M и N
        cathedras = []
        weights = []
        row = 21
        
        print(f"📊 Загружаю кафедры и весовые коэффициенты...")
        
        while True:
            cathedra_cell = ws_dates[f'M{row}']
            weight_cell = ws_dates[f'N{row}']
            
            # Если обе ячейки пустые, прекращаем чтение
            if (cathedra_cell.value is None or cathedra_cell.value == '') and (weight_cell.value is None or weight_cell.value == ''):
                break
            
            # Проверяем, есть ли название кафедры
            if cathedra_cell.value is not None and cathedra_cell.value != '':
                cathedra_name = str(cathedra_cell.value).strip()
                
                # Получаем весовой коэффициент
                if weight_cell.value is not None and weight_cell.value != '':
                    try:
                        weight = float(weight_cell.value)
                    except:
                        print(f"⚠ Ошибка преобразования веса для кафедры '{cathedra_name}', установлен вес 1.0")
                        weight = 1.0
                else:
                    print(f"⚠ Для кафедры '{cathedra_name}' не указан вес, установлен вес 1.0")
                    weight = 1.0
                
                cathedras.append(cathedra_name)
                weights.append(weight)
                print(f"   Кафедра: '{cathedra_name}', вес: {weight}")
            
            row += 1
        
        if not cathedras:
            print("⚠ Не найдено кафедр для распределения")
            cathedras = ['каф. 28', 'каф. 41', 'каф. 42', 'каф. 43']
            weights = [0.25, 0.35, 0.20, 0.20]  # Пример нормализованных весов
            print(f"   Использую кафедры по умолчанию: {cathedras}")
        
        # Нормализуем веса, чтобы сумма была 1
        total_weight_cathedras = sum(weights)
        if total_weight_cathedras > 0:
            normalized_weights = [w / total_weight_cathedras for w in weights]
        else:
            normalized_weights = [1.0 / len(cathedras) for _ in cathedras]
        
        print(f"\n📊 Нормализованные веса кафедр (сумма = 1):")
        for cathedra, norm_weight in zip(cathedras, normalized_weights):
            print(f"   {cathedra}: {norm_weight:.3f}")
        
        print(f"\n📊 Параметры распределения для строк 'пспрс':")
        print(f"   Всего дней для распределения:")
        print(f"     Рабочие дни: {total_workdays_to_distribute}")
        print(f"     Выходные дни: {total_weekends_to_distribute}")
        print(f"   Весовые коэффициенты:")
        print(f"     ПС (N8): {ps_weight} -> нормализованный: {ps_weight_normalized:.3f}")
        print(f"     ПРС (N9): {prs_weight} -> нормализованный: {prs_weight_normalized:.3f}")
        print(f"   Распределение дней между ПС и ПРС:")
        print(f"     ПС: {ps_workdays} рабочих, {ps_weekends} выходных")
        print(f"     ПРС: {prs_workdays} рабочих, {prs_weekends} выходных")
        
        print(f"\n📊 Параметры для строк 'пс':")
        print(f"   ПС: {ps_for_ps_type_workdays} рабочих, {ps_for_ps_type_weekends} выходных")
        print(f"   Всего кафедр: {len(cathedras)}")
        
        # Закрываем книгу для чтения через pandas
        book.close()
        
        # Читаем данные через pandas
        df = pd.read_excel('Дежурства.xlsx', sheet_name='Даты_дежурств')
        
        # Проверяем наличие необходимых столбцов
        required_columns = ['Тип назначаемого л/с', 'Вид', 'Дата начала', 'Дата окончания']
        for col in required_columns:
            if col not in df.columns:
                print(f"Ошибка: отсутствует столбец '{col}'")
                return
        
        # Преобразуем даты в формат datetime, если они еще не в нем
        df['Дата начала'] = pd.to_datetime(df['Дата начала'], dayfirst=True)
        df['Дата окончания'] = pd.to_datetime(df['Дата окончания'], dayfirst=True)
        
        # Находим минимальную и максимальную дату для создания полного диапазона
        min_date = df['Дата начала'].min()
        max_date = df['Дата окончания'].max()
        
        # Создаем список всех дат в диапазоне
        date_range = pd.date_range(start=min_date, end=max_date, freq='D')
        
        # ФИЛЬТРУЕМ ТОЛЬКО ТЕ СТРОКИ, КОТОРЫЕ НУЖНО ОТОБРАЖАТЬ
        # Убираем строки с пустым "Вид" или с типом, который не нужно отображать
        display_df = df.copy()
        
        # Убираем строки, где "Вид" пустой или NaN
        display_df = display_df[display_df['Вид'].notna()]
        display_df = display_df[display_df['Вид'].astype(str).str.strip() != '']
        
        # Также убираем строки, где тип дежурства пустой
        display_df = display_df[display_df['Тип назначаемого л/с'].notna()]
        display_df = display_df[display_df['Тип назначаемого л/с'].astype(str).str.strip() != '']
        
        
        if len(display_df) == 0:
            print("⚠ Ошибка: Нет строк для отображения в расписании")
            return
        
        # Создаем пустой DataFrame для расписания
        schedule_data = []
        
        # Для каждой строки в ОТФИЛЬТРОВАННЫХ данных
        for idx, row in display_df.iterrows():
            duty_type = row['Тип назначаемого л/с']
            duty_kind = row['Вид']
            start_date = row['Дата начала']
            end_date = row['Дата окончания']
            
            # Создаем словарь для текущей строки
            row_data = {
                'Тип назначаемого л/с': duty_type,
                'Вид': duty_kind
            }
            
            # Для каждой даты в общем диапазоне
            for date in date_range:
                date_str = date.strftime('%d.%m.%Y')
                
                # Проверяем, входит ли дата в диапазон дежурства
                if start_date <= date <= end_date:
                    # Проверяем, является ли день выходным (пятница, суббота, воскресенье)
                    if date.weekday() >= 4:  # Выходной (4=пт, 5=сб, 6=вс)
                        row_data[date_str] = 'В'
                    elif date.date() in holiday_dates:  # Проверяем, является ли день праздничным
                        row_data[date_str] = 'В'
                    else:  # Рабочий день
                        row_data[date_str] = 'Р'
                else:
                    row_data[date_str] = ''
            
            schedule_data.append(row_data)
        
        # Создаем DataFrame для расписания
        schedule_df = pd.DataFrame(schedule_data)
        
        print(f"\n📊 Создано расписание с {len(schedule_df)} строками")
        
        # Находим строки с типом "пспрс"
        pspc_rows = schedule_df[schedule_df['Тип назначаемого л/с'].str.lower() == 'пспрс']
        
        # Находим строки с типом "пс"
        ps_rows = schedule_df[schedule_df['Тип назначаемого л/с'].str.lower() == 'пс']
        
        # ВАЖНОЕ ИЗМЕНЕНИЕ: ВМЕСТО РАЗДЕЛЬНОГО РАСПРЕДЕЛЕНИЯ СОБИРАЕМ ВСЕ ДЕЖУРСТВА В ОБЩИЙ ПУЛ
        
        # Создаем копию DataFrame для распределения
        schedule_with_distribution = schedule_df.copy()
        
        # Собираем ВСЕ дежурства, которые нужно распределить по кафедрам
        all_cathedra_duties_to_distribute = []  # Будет хранить (индекс_строки, дата, тип_дня)
        
        print(f"\n🔍 Собираю все дежурства для распределения по кафедрам...")
        
        # 1. Дежурства ПС из строк "пспрс" (рабочие и выходные)
        if not pspc_rows.empty:
            print(f"   Для строк 'пспрс': найдено {len(pspc_rows)} строк")
            
            for idx in pspc_rows.index:
                # Получаем доступные даты для дежурства в этой строки
                available_workdays = []
                available_weekends = []
                
                # Собираем рабочие и выходные дни для этой строки
                for date in date_range:
                    date_str = date.strftime('%d.%m.%Y')
                    if schedule_df.at[idx, date_str] == 'Р':
                        available_workdays.append(date_str)
                    elif schedule_df.at[idx, date_str] == 'В':
                        available_weekends.append(date_str)
                
                print(f"   Строка {idx+1}: {len(available_workdays)} рабочих, {len(available_weekends)} выходных дней")
                
                # Распределяем ПС в рабочие дни (из них потом будут кафедры)
                if available_workdays:
                    ps_workday_dates = random.sample(available_workdays, min(ps_workdays, len(available_workdays)))
                    for date_str in ps_workday_dates:
                        all_cathedra_duties_to_distribute.append((idx, date_str, 'workday'))
                    available_workdays = [d for d in available_workdays if d not in ps_workday_dates]
                
                # Распределяем ПС в выходные дни (из них потом будут кафедры)
                if available_weekends:
                    ps_weekend_dates = random.sample(available_weekends, min(ps_weekends, len(available_weekends)))
                    for date_str in ps_weekend_dates:
                        all_cathedra_duties_to_distribute.append((idx, date_str, 'weekend'))
                    available_weekends = [d for d in available_weekends if d not in ps_weekend_dates]
                
                # Оставшиеся дни распределяем как ПРС или оставляем пустыми
                # ПРС в рабочие дни
                if available_workdays:
                    prs_workday_dates = random.sample(available_workdays, min(prs_workdays, len(available_workdays)))
                    for date_str in prs_workday_dates:
                        schedule_with_distribution.at[idx, date_str] = 'прс'
                    available_workdays = [d for d in available_workdays if d not in prs_workday_dates]
                
                # ПРС в выходные дни
                if available_weekends:
                    prs_weekend_dates = random.sample(available_weekends, min(prs_weekends, len(available_weekends)))
                    for date_str in prs_weekend_dates:
                        schedule_with_distribution.at[idx, date_str] = 'прс'
                    available_weekends = [d for d in available_weekends if d not in prs_weekend_dates]
        
        # 2. Дежурства из строк "пс" (рабочие и выходные) - собираем ВСЕ доступные дни
        if not ps_rows.empty:
            print(f"   Для строк 'пс': найдено {len(ps_rows)} строк")
            
            for idx in ps_rows.index:
                # Получаем доступные даты для дежурства в этой строки
                available_workdays = []
                available_weekends = []
                
                # Собираем все рабочие и выходные дни для этой строки
                for date in date_range:
                    date_str = date.strftime('%d.%m.%Y')
                    if schedule_df.at[idx, date_str] == 'Р':
                        available_workdays.append(date_str)
                    elif schedule_df.at[idx, date_str] == 'В':
                        available_weekends.append(date_str)
                
                print(f"   Строка {idx+1}: {len(available_workdays)} рабочих, {len(available_weekends)} выходных дней")
                
                # Собираем ВСЕ доступные дни для распределения по кафедрам
                for date_str in available_workdays:
                    all_cathedra_duties_to_distribute.append((idx, date_str, 'workday'))
                
                for date_str in available_weekends:
                    all_cathedra_duties_to_distribute.append((idx, date_str, 'weekend'))
        
        print(f"\n📊 Всего собрано дежурств для распределения по кафедрам: {len(all_cathedra_duties_to_distribute)}")
        
        # РАСПРЕДЕЛЯЕМ ВСЕ СОБРАННЫЕ ДЕЖУРСТВА ПРОПОРЦИОНАЛЬНО ВЕСАМ КАФЕДР
        
        if all_cathedra_duties_to_distribute and cathedras:
            print(f"\n🔍 Пропорциональное распределение всех дежурств по кафедрам...")
            print(f"   Всего дежурств для распределения: {len(all_cathedra_duties_to_distribute)}")
            print(f"   Количество кафедр: {len(cathedras)}")
            
            # Перемешиваем даты для случайного распределения
            random.shuffle(all_cathedra_duties_to_distribute)
            
            # Рассчитываем, сколько дежурств должна получить каждая кафедра пропорционально весу
            cathedra_target_counts = []
            
            for i, (cathedra, norm_weight) in enumerate(zip(cathedras, normalized_weights)):
                # Рассчитываем целевое количество дежурств пропорционально нормализованному весу
                target = norm_weight * len(all_cathedra_duties_to_distribute)
                cathedra_target_counts.append({
                    'cathedra': cathedra,
                    'norm_weight': norm_weight,
                    'target': target,
                    'assigned': 0,
                    'min_assigned': math.floor(target),  # Минимальное количество для назначения
                    'remainder': target - math.floor(target)  # Дробная часть для учета
                })
            
            # Выводим информацию о целевых значениях
            print(f"\n   Целевое распределение для кафедр:")
            for target_info in cathedra_target_counts:
                print(f"     {target_info['cathedra']}: вес {target_info['norm_weight']:.3f}, "
                      f"целевое {target_info['target']:.1f} (мин: {target_info['min_assigned']}, "
                      f"остаток: {target_info['remainder']:.2f})")
            
            # Создаем словарь для отслеживания фактического количества дежурств по кафедрам
            cathedra_actual_counts = {cathedra: 0 for cathedra in cathedras}
            
            # Шаг 1: Назначаем минимальное количество каждой кафедре
            duty_index = 0
            print(f"\n   Шаг 1: Назначаем минимальное количество дежурств каждой кафедре...")
            
            for target_info in cathedra_target_counts:
                for _ in range(target_info['min_assigned']):
                    if duty_index < len(all_cathedra_duties_to_distribute):
                        idx, date_str, day_type = all_cathedra_duties_to_distribute[duty_index]
                        schedule_with_distribution.at[idx, date_str] = target_info['cathedra']
                        cathedra_actual_counts[target_info['cathedra']] += 1
                        target_info['assigned'] += 1
                        duty_index += 1
            
            print(f"   После шага 1 распределено: {duty_index} дежурств")
            
            # Шаг 2: Распределяем оставшиеся дежурства по кафедрам с наибольшей дробной частью
            remaining_duties = len(all_cathedra_duties_to_distribute) - duty_index
            
            if remaining_duties > 0:
                print(f"\n   Шаг 2: Распределяем оставшиеся {remaining_duties} дежурств...")
                
                # Сортируем кафедры по дробной части (в порядке убывания)
                cathedra_target_counts.sort(key=lambda x: x['remainder'], reverse=True)
                
                for i in range(remaining_duties):
                    if i < len(cathedra_target_counts):
                        target_info = cathedra_target_counts[i]
                        idx, date_str, day_type = all_cathedra_duties_to_distribute[duty_index]
                        schedule_with_distribution.at[idx, date_str] = target_info['cathedra']
                        cathedra_actual_counts[target_info['cathedra']] += 1
                        target_info['assigned'] += 1
                        duty_index += 1
                    else:
                        # Если кафедр меньше, чем оставшихся дежурств, распределяем циклически
                        target_info = cathedra_target_counts[i % len(cathedra_target_counts)]
                        idx, date_str, day_type = all_cathedra_duties_to_distribute[duty_index]
                        schedule_with_distribution.at[idx, date_str] = target_info['cathedra']
                        cathedra_actual_counts[target_info['cathedra']] += 1
                        target_info['assigned'] += 1
                        duty_index += 1
            
            print(f"\n   ✅ Всего распределено: {duty_index} дежурств")
            
            # Обновляем основной DataFrame
            schedule_df = schedule_with_distribution
            
            # ВЫВОДИМ ИТОГОВУЮ СТАТИСТИКУ
            print(f"\n📊 ИТОГОВОЕ РАСПРЕДЕЛЕНИЕ ДЕЖУРСТВ ПО КАФЕДРАМ:")
            print("=" * 60)
            
            # Сортируем кафедры по количеству дежурств
            sorted_cathedras = sorted(cathedra_actual_counts.items(), key=lambda x: x[1], reverse=True)
            
            for cathedra, count in sorted_cathedras:
                # Находим целевую информацию для этой кафедры
                target_info = next((t for t in cathedra_target_counts if t['cathedra'] == cathedra), None)
                if target_info:
                    target_val = target_info['target']
                    deviation = count - target_val
                    deviation_percent = (deviation / target_val * 100) if target_val > 0 else 0
                    deviation_str = f"{deviation:+.1f} ({deviation_percent:+.1f}%)"
                    
                    # Определяем статус в зависимости от отклонения
                    if abs(deviation_percent) <= 5:
                        status = "✅ Отлично"
                    elif abs(deviation_percent) <= 10:
                        status = "⚠ Приемлемо"
                    else:
                        status = "❌ Большое отклонение"
                    
                    print(f"   {cathedra}:")
                    print(f"     Факт: {count} дежурств")
                    print(f"     Цель: {target_val:.1f} дежурств")
                    print(f"     Отклонение: {deviation_str}")
                    print(f"     Статус: {status}")
                    print()
            
            total_actual = sum(cathedra_actual_counts.values())
            print(f"\n   Всего распределено: {total_actual} дежурств")
            print(f"   Общее доступное: {len(all_cathedra_duties_to_distribute)} дежурств")
            
            # Проверяем балансировку
            print(f"\n📈 Сбалансированность распределения:")
            for target_info in cathedra_target_counts:
                cathedra = target_info['cathedra']
                actual = cathedra_actual_counts[cathedra]
                target = target_info['target']
                if target > 0:
                    percentage = (actual / target) * 100
                    print(f"   {cathedra}: {actual}/{target:.1f} = {percentage:.1f}% от цели")
            
            print("=" * 60)
        else:
            print("⚠ Нет дежурств для распределения по кафедрам")
            schedule_df = schedule_with_distribution
        
        # Загружаем книгу для сохранения на отдельный лист
        book = load_workbook('Дежурства.xlsx')
        
        # Удаляем лист 'Распределение', если он существует
        if 'Распределение' in book.sheetnames:
            std = book['Распределение']
            book.remove(std)
        
        # Создаем новый лист
        schedule_sheet = book.create_sheet('Распределение')
        
        # Определяем стили заливки
        workday_fill = PatternFill(
            start_color='C6EFCE',  # Светло-зеленый
            end_color='C6EFCE',
            fill_type='solid'
        )
        
        weekend_fill = PatternFill(
            start_color='FFC7CE',  # Светло-красный
            end_color='FFC7CE',
            fill_type='solid'
        )
        
        # Шрифты для разных типов дежурств
        ps_font = Font(name='Arial', size=10, bold=True, color='000000')
        prs_font = Font(name='Arial', size=10, bold=True, color='000000')
        
        # Цвета шрифтов для кафедр
        cathedra_colors = [
            '0000FF',  # Синий
            '008000',  # Зеленый
            '800080',  # Фиолетовый
            'FF6600',  # Оранжевый
            'FF0000',  # Красный
            '00CED1',  # Бирюзовый
            'FF1493',  # Розовый
            '4B0082',  # Индиго
        ]
        
        # Создаем словарь шрифтов для кафедр
        cathedra_fonts = {}
        for i, cathedra in enumerate(cathedras):
            color = cathedra_colors[i % len(cathedra_colors)]
            cathedra_fonts[cathedra] = Font(name='Arial', size=10, bold=True, color=color)
        
        # Определяем границы для ячеек
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Центрируем текст
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Записываем DataFrame на лист
        for r_idx, row in enumerate(dataframe_to_rows(schedule_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = schedule_sheet.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                
                # Для заголовков (первая строка) применяем жирный шрифт и центрирование
                if r_idx == 1:
                    cell.font = cell.font.copy(bold=True)
                    cell.alignment = center_alignment
                    # Для дат в заголовке добавляем перенос текста
                    if c_idx > 2:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Для данных (кроме первых двух столбцов)
                    if c_idx > 2 and r_idx > 1:
                        cell.alignment = center_alignment
                        
                        # Получаем тип строки
                        row_type = schedule_df.at[r_idx-2, 'Тип назначаемого л/с']
                        if isinstance(row_type, str):
                            row_type_lower = row_type.lower()
                            
                            # Для строк "пспрс"
                            if row_type_lower == 'пспрс':
                                if value in cathedras:
                                    # Определяем цвет заливки по типу дня
                                    date_str = schedule_df.columns[c_idx-1]
                                    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                                    if date_obj.weekday() >= 4 or date_obj.date() in holiday_dates:
                                        cell.fill = weekend_fill  # Кафедры в выходных днях
                                    else:
                                        cell.fill = workday_fill  # Кафедры в рабочих днях
                                    # Применяем соответствующий шрифт
                                    if value in cathedra_fonts:
                                        cell.font = cathedra_fonts[value]
                                elif value == 'пс':
                                    # Определяем цвет заливки по типу дня
                                    date_str = schedule_df.columns[c_idx-1]
                                    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                                    if date_obj.weekday() >= 4 or date_obj.date() in holiday_dates:
                                        cell.fill = weekend_fill
                                        cell.font = ps_font
                                    else:
                                        cell.fill = workday_fill
                                        cell.font = ps_font
                                elif value == 'прс':
                                    # Определяем цвет заливки по типу дня
                                    date_str = schedule_df.columns[c_idx-1]
                                    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                                    if date_obj.weekday() >= 4 or date_obj.date() in holiday_dates:
                                        cell.fill = weekend_fill
                                    else:
                                        cell.fill = workday_fill
                                    cell.font = prs_font
                                elif value == 'Р':
                                    cell.fill = workday_fill
                                    cell.value = ''
                                elif value == 'В':
                                    cell.fill = weekend_fill
                                    cell.value = ''
                            
                            # Для строк "пс"
                            elif row_type_lower == 'пс':
                                if value in cathedras:
                                    # Определяем цвет заливки по типу дня
                                    date_str = schedule_df.columns[c_idx-1]
                                    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                                    if date_obj.weekday() >= 4 or date_obj.date() in holiday_dates:
                                        cell.fill = weekend_fill  # Кафедры в выходных днях
                                    else:
                                        cell.fill = workday_fill  # Кафедры в рабочих днях
                                    # Применяем соответствующий шрифт
                                    if value in cathedra_fonts:
                                        cell.font = cathedra_fonts[value]
                                elif value == 'пс':
                                    # Определяем цвет заливки по типу дня
                                    date_str = schedule_df.columns[c_idx-1]
                                    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                                    if date_obj.weekday() >= 4 or date_obj.date() in holiday_dates:
                                        cell.fill = weekend_fill
                                        cell.font = ps_font
                                    else:
                                        cell.fill = workday_fill
                                        cell.font = ps_font
                                elif value == 'Р':
                                    cell.fill = workday_fill
                                    cell.value = ''
                                elif value == 'В':
                                    cell.fill = weekend_fill
                                    cell.value = ''
                            
                            # Для обычных строк (не пспрс и не пс)
                            else:
                                if value == 'Р':
                                    cell.fill = workday_fill
                                    cell.value = ''
                                elif value == 'В':
                                    cell.fill = weekend_fill
                                    cell.value = ''
                    elif c_idx <= 2 and r_idx > 1:
                        # Для первых двух столбцов центрируем текст
                        cell.alignment = Alignment(vertical='center')
        
        # Настройка ширины столбцов
        column_widths = {}
        
        # Для первых двух столбцов
        column_widths['A'] = max(schedule_df['Тип назначаемого л/с'].astype(str).apply(len).max(), 20)
        column_widths['B'] = max(schedule_df['Вид'].astype(str).apply(len).max(), 15)
        
        # Для столбцов с датами
        for i, date in enumerate(date_range, 3):
            col_letter = get_column_letter(i)
            # Ширина для дат - немного больше для отображения текста кафедр
            column_widths[col_letter] = 12
        
        # Применяем настройки ширины
        for col_letter, width in column_widths.items():
            schedule_sheet.column_dimensions[col_letter].width = width
        
        # Настройка высоты строк
        schedule_sheet.row_dimensions[1].height = 40  # Заголовки повыше
        for row in range(2, len(schedule_df) + 2):
            schedule_sheet.row_dimensions[row].height = 25
        
        # Добавляем легенду
        legend_row = len(schedule_df) + 3
        
        # Заголовок легенды
        schedule_sheet.cell(row=legend_row, column=1, value="Легенда:").font = Font(bold=True)
        
        # Рабочий день
        schedule_sheet.cell(row=legend_row, column=2, value="Рабочий день")
        schedule_sheet.cell(row=legend_row, column=3).fill = workday_fill
        schedule_sheet.cell(row=legend_row, column=3).border = thin_border
        
        # Выходной день
        schedule_sheet.cell(row=legend_row, column=4, value="Выходной день (пт, сб, вс)")
        schedule_sheet.cell(row=legend_row, column=5).fill = weekend_fill
        schedule_sheet.cell(row=legend_row, column=5).border = thin_border
        
        # Праздничный день
        schedule_sheet.cell(row=legend_row+1, column=4, value="Праздничный день")
        schedule_sheet.cell(row=legend_row+1, column=5).fill = weekend_fill
        schedule_sheet.cell(row=legend_row+1, column=5).border = thin_border
        
        # ПС в рабочем дне
        schedule_sheet.cell(row=legend_row, column=6, value="ПС")
        schedule_sheet.cell(row=legend_row, column=7).fill = workday_fill
        schedule_sheet.cell(row=legend_row, column=7).border = thin_border
        schedule_sheet.cell(row=legend_row, column=7).font = ps_font
        schedule_sheet.cell(row=legend_row, column=7).value = "ПС"
        
        # ПС в выходном дне
        schedule_sheet.cell(row=legend_row+1, column=6, value="ПС в выходной")
        schedule_sheet.cell(row=legend_row+1, column=7).fill = weekend_fill
        schedule_sheet.cell(row=legend_row+1, column=7).border = thin_border
        schedule_sheet.cell(row=legend_row+1, column=7).font = ps_font
        schedule_sheet.cell(row=legend_row+1, column=7).value = "ПС"
        
        # ПРС в рабочем дне
        schedule_sheet.cell(row=legend_row, column=8, value="ПРС")
        schedule_sheet.cell(row=legend_row, column=9).fill = workday_fill
        schedule_sheet.cell(row=legend_row, column=9).border = thin_border
        schedule_sheet.cell(row=legend_row, column=9).font = prs_font
        schedule_sheet.cell(row=legend_row, column=9).value = "ПРС"
        
        # ПРС в выходном дне
        schedule_sheet.cell(row=legend_row+1, column=8, value="ПРС в выходной")
        schedule_sheet.cell(row=legend_row+1, column=9).fill = weekend_fill
        schedule_sheet.cell(row=legend_row+1, column=9).border = thin_border
        schedule_sheet.cell(row=legend_row+1, column=9).font = prs_font
        schedule_sheet.cell(row=legend_row+1, column=9).value = "ПРС"
        
        # Добавляем кафедры в легенду
        cathedra_start_col = 10
        for i, cathedra in enumerate(cathedras):
            col_offset = (i // 2) * 2  # 2 кафедры в строке
            row_offset = i % 2  # 0 для первой строки, 1 для второй
            
            # Название кафедры
            schedule_sheet.cell(row=legend_row + row_offset, column=cathedra_start_col + col_offset, value=cathedra)
            
            # Пример ячейки с цветом
            cell = schedule_sheet.cell(row=legend_row + row_offset, column=cathedra_start_col + col_offset + 1)
            cell.fill = workday_fill
            cell.border = thin_border
            cell.value = cathedra
            if cathedra in cathedra_fonts:
                cell.font = cathedra_fonts[cathedra]
        
        # Пустая ячейка
        empty_cell_col = cathedra_start_col + ((len(cathedras) + 1) // 2) * 2
        schedule_sheet.cell(row=legend_row, column=empty_cell_col, value="Нет дежурства")
        schedule_sheet.cell(row=legend_row, column=empty_cell_col + 1).border = thin_border
        
        # Добавляем информацию о распределении
        info_row = legend_row + 3
        schedule_sheet.cell(row=info_row, column=1, value="Параметры распределения:").font = Font(bold=True)
        
        # Информация для строк "пспрс" (НОВАЯ логика)
        schedule_sheet.cell(row=info_row+1, column=1, value="Для строк 'пспрс':").font = Font(bold=True)
        schedule_sheet.cell(row=info_row+2, column=2, value="Всего рабочих дней для распределения (P14):")
        schedule_sheet.cell(row=info_row+2, column=3, value=total_workdays_to_distribute)
        schedule_sheet.cell(row=info_row+3, column=2, value="Всего выходных дней для распределения (Q14):")
        schedule_sheet.cell(row=info_row+3, column=3, value=total_weekends_to_distribute)
        schedule_sheet.cell(row=info_row+4, column=2, value="Вес ПС (N8):")
        schedule_sheet.cell(row=info_row+4, column=3, value=ps_weight)
        schedule_sheet.cell(row=info_row+5, column=2, value="Вес ПРС (N9):")
        schedule_sheet.cell(row=info_row+5, column=3, value=prs_weight)
        schedule_sheet.cell(row=info_row+6, column=2, value="Нормализованный вес ПС:")
        schedule_sheet.cell(row=info_row+6, column=3, value=f"{ps_weight_normalized:.3f}")
        schedule_sheet.cell(row=info_row+7, column=2, value="Нормализованный вес ПРС:")
        schedule_sheet.cell(row=info_row+7, column=3, value=f"{prs_weight_normalized:.3f}")
        schedule_sheet.cell(row=info_row+8, column=2, value="ПС в рабочие дни:")
        schedule_sheet.cell(row=info_row+8, column=3, value=ps_workdays)
        schedule_sheet.cell(row=info_row+9, column=2, value="ПС в выходные дни:")
        schedule_sheet.cell(row=info_row+9, column=3, value=ps_weekends)
        schedule_sheet.cell(row=info_row+10, column=2, value="ПРС в рабочие дни:")
        schedule_sheet.cell(row=info_row+10, column=3, value=prs_workdays)
        schedule_sheet.cell(row=info_row+11, column=2, value="ПРС в выходные дни:")
        schedule_sheet.cell(row=info_row+11, column=3, value=prs_weekends)
        
        # Информация для строк "пс" (СТАРАЯ логика)
        schedule_sheet.cell(row=info_row+13, column=1, value="Для строк 'пс' (старая логика):").font = Font(bold=True)
        schedule_sheet.cell(row=info_row+14, column=2, value="ПС в рабочие дни (P16):")
        schedule_sheet.cell(row=info_row+14, column=3, value=ps_for_ps_type_workdays)
        schedule_sheet.cell(row=info_row+15, column=2, value="ПС в выходные дни (Q16):")
        schedule_sheet.cell(row=info_row+15, column=3, value=ps_for_ps_type_weekends)
        
        # Добавляем информацию о весовых коэффициентах кафедр
        schedule_sheet.cell(row=info_row+1, column=5, value="Весовые коэффициенты кафедр:").font = Font(bold=True)
        
        for i, (cathedra, weight) in enumerate(zip(cathedras, normalized_weights)):
            schedule_sheet.cell(row=info_row+2+i, column=5, value=f"{cathedra}:")
            schedule_sheet.cell(row=info_row+2+i, column=6, value=f"{weight:.3f}")
        
        # Добавляем информацию об объединенном распределении
        schedule_sheet.cell(row=info_row+len(cathedras)+4, column=1, value="ОБЪЕДИНЕННОЕ РАСПРЕДЕЛЕНИЕ:").font = Font(bold=True, color='FF0000')
        schedule_sheet.cell(row=info_row+len(cathedras)+5, column=2, value="Все дежурства (из строк 'пспрс' и 'пс')")
        schedule_sheet.cell(row=info_row+len(cathedras)+6, column=2, value="собраны в общий пул и распределены")
        schedule_sheet.cell(row=info_row+len(cathedras)+7, column=2, value="пропорционально весам кафедр")
        
        # Добавляем итоговую статистику распределения
        stats_row = info_row + len(cathedras) + 9
        schedule_sheet.cell(row=stats_row, column=1, value="Итоговое распределение дежурств по кафедрам:").font = Font(bold=True)
        
        if 'cathedra_actual_counts' in locals():
            for i, (cathedra, count) in enumerate(sorted_cathedras):
                schedule_sheet.cell(row=stats_row+1+i, column=1, value=cathedra)
                schedule_sheet.cell(row=stats_row+1+i, column=2, value=count)
                
                # Находим целевую информацию для этой кафедры
                target_info = next((t for t in cathedra_target_counts if t['cathedra'] == cathedra), None)
                if target_info:
                    schedule_sheet.cell(row=stats_row+1+i, column=3, value=f"цель: {target_info['target']:.1f}")
                    deviation = count - target_info['target']
                    deviation_percent = (deviation / target_info['target'] * 100) if target_info['target'] > 0 else 0
                    schedule_sheet.cell(row=stats_row+1+i, column=4, value=f"отклонение: {deviation:+.1f} ({deviation_percent:+.1f}%)")
        
        # Сохраняем изменения
        book.save('Дежурства.xlsx')
        
        print("\n" + "=" * 70)
        print("РАСПИСАНИЕ ДЕЖУРСТВ УСПЕШНО СОЗДАНО")
        print("=" * 70)
        print(f"📅 Период: {min_date.strftime('%d.%m.%Y')} - {max_date.strftime('%d.%m.%Y')}")
        print(f"📊 Всего записей в исходных данных: {len(df)}")
        print(f"📋 Записей отображено в расписании: {len(schedule_df)}")
        print(f"📆 Количество дней в расписании: {len(date_range)}")
        
        if holiday_dates:
            print(f"🎉 Количество праздничных дней: {len(holiday_dates)}")
        
        if not pspc_rows.empty:
            print(f"\n🎯 Распределено для 'пспрс' (НОВАЯ логика): {len(pspc_rows)} строк")
            print(f"📈 Всего дней для распределения: {total_workdays_to_distribute} рабочих, {total_weekends_to_distribute} выходных")
            print(f"📈 Веса: ПС={ps_weight}, ПРС={prs_weight}")
            print(f"📈 Итог: ПС={ps_workdays} рабочих, {ps_weekends} выходных")
            print(f"📈 Итог: ПРС={prs_workdays} рабочих, {prs_weekends} выходных")
        
        if not ps_rows.empty:
            print(f"\n🎯 Собрано для строк 'пс': {len(ps_rows)} строк")
            print(f"📈 Все доступные дни собраны в общий пул")
        
        print(f"\n📊 ВСЕ дежурства (из строк 'пспрс' и 'пс') собраны в общий пул")
        print(f"📊 И распределены пропорционально весам кафедр")
        print(f"💾 Файл сохранен: Дежурства.xlsx (лист 'Распределение')")
        
    except FileNotFoundError:
        print("❌ Ошибка: Файл 'Дежурства.xlsx' не найден.")
    except Exception as e:
        print(f"❌ Произошла ошибка: {str(e)}")
        import traceback
        traceback.print_exc()

# Остальные функции остаются без изменений
def create_sample_file_with_holidays_and_psprc():
    """Функция для создания тестового файла с праздничными днями и параметрами для пспрс"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    print("📝 Создаю тестовый файл 'Дежурства.xlsx' с праздничными днями и параметрами для пспрс...")
    
    # Создаем тестовые данные
    sample_data = {
        'Тип назначаемого л/с': [
            'пспрс', 
            'Офицер-оператор', 
            'Сержант-наблюдатель', 
            'пспрс',
            'пс',
            'пс',
            'Офицер связи'
        ],
        'Вид': [
            'Смешанный', 
            'Оперативный', 
            'Наблюдательный', 
            'Смешанный',
            'Кафедральный',
            'Кафедральный',
            'Оперативный'
        ],
        'Дата начала': [
            '01.01.2024', 
            '03.01.2024', 
            '05.01.2024', 
            '07.01.2024',
            '01.01.2024',
            '01.01.2024',
            '10.01.2024'
        ],
        'Дата окончания': [
            '31.01.2024', 
            '12.01.2024', 
            '15.01.2024', 
            '20.01.2024',
            '31.01.2024',
            '31.01.2024',
            '18.01.2024'
        ]
    }
    
    # Праздничные дни (пример)
    holidays = [
        '01.01.2024',  # Новый год
        '02.01.2024',  # Новогодние каникулы
        '03.01.2024',  # Новогодние каникулы
        '07.01.2024',  # Рождество
        '08.01.2024',  # Новогодние каникулы
    ]
    
    # Данные о кафедрах и их весах (пример - веса в сумме дают 1)
    cathedras_data = [
        ('Кафедра 28', 0.30),  # 30% дежурств
        ('Кафедра 41', 0.35),  # 35% дежурств
        ('Кафедра 42', 0.20),  # 20% дежурств
        ('Кафедра 43', 0.15),  # 15% дежурств
    ]
    
    # Создаем Excel файл с тестовыми данными
    wb = Workbook()
    ws = wb.active
    ws.title = "Даты_дежурств"
    
    # Записываем заголовки
    headers = list(sample_data.keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)
    
    # Записываем данные
    for row_idx in range(len(sample_data['Тип назначаемого л/с'])):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx+2, column=col_idx, value=sample_data[key][row_idx])
    
    # Добавляем заголовок для праздничных дней
    ws['F10'] = "Праздничные дни"
    ws['F10'].font = ws['F10'].font.copy(bold=True)
    
    # Добавляем праздничные дни
    for i, holiday in enumerate(holidays, 11):  # Начинаем с 11 строки
        ws[f'F{i}'] = holiday
    
    # Добавляем параметры распределения для пспрс (НОВАЯ логика)
    ws['N7'] = "Распределение дней в строках 'пспрс'"
    ws['N7'].font = ws['N7'].font.copy(bold=True)
    
    ws['N8'] = "Вес для ПС"
    ws['N9'] = "Вес для ПРС"
    
    ws['O8'] = 0.6  # Вес для ПС (60%)
    ws['O9'] = 0.4  # Вес для ПРС (40%)
    
    ws['P13'] = "Общее количество дней для распределения"
    ws['P13'].font = ws['P13'].font.copy(bold=True)
    
    ws['P14'] = "Рабочие дни"
    ws['Q14'] = "Выходные дни"
    ws['P14'].font = ws['Q14'].font = ws['P14'].font.copy(bold=True)
    
    ws['P15'] = 20  # Всего рабочих дней для распределения
    ws['Q15'] = 8   # Всего выходных дней для распределения
    
    # Добавляем параметры распределения для строк "пс" (СТАРАЯ логика)
    ws['O16'] = "ПС (для строк 'пс') - старая логика"
    ws['P16'] = 8   # ПС в рабочие дни для строк "пс"
    ws['Q16'] = 2   # ПС в выходные дни для строк "пс"
    
    # Добавляем кафедры и их весовые коэффициенты
    ws['M20'] = "Кафедры и весовые коэффициенты"
    ws['M20'].font = ws['M20'].font.copy(bold=True)
    
    ws['M21'] = "Кафедра"
    ws['N21'] = "Вес (сумма = 1)"
    ws['M21'].font = ws['N21'].font = ws['M21'].font.copy(bold=True)
    
    for i, (cathedra, weight) in enumerate(cathedras_data, 22):
        ws[f'M{i}'] = cathedra
        ws[f'N{i}'] = weight
    
    # Настраиваем ширину столбцов
    column_widths = [25, 20, 15, 15, 10, 15, 5, 5, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Сохраняем файл
    wb.save('Дежурства.xlsx')
    print("✅ Тестовый файл успешно создан!")
    print("📌 Праздничные дни добавлены в столбец F, начиная с ячейки F11")
    print("📊 Веса для распределения ПС/ПРС (новая логика для 'пспрс') добавлены в ячейки N8, N9, O8, O9")
    print("📊 Общее количество дней для распределения (P14, Q14) для строк 'пспрс'")
    print("📊 Параметры распределения для строк 'пс' (старая логика) добавлены в ячейки P16, Q16")
    print("📊 Кафедры и их веса добавлены в столбцы M и N, начиная с строки 21")
    print("   Пример весов, сумма которых = 1:")
    for cathedra, weight in cathedras_data:
        print(f"   {cathedra}: {weight} ({weight*100:.0f}%)")

def print_sample_instructions():
    """Выводит инструкцию по использованию"""
    print("\n" + "=" * 70)
    print("ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ")
    print("=" * 70)
    print("1. Подготовьте файл 'Дежурства.xlsx' с листом 'Даты_дежурств'")
    print("2. Структура таблицы должна содержать столбцы:")
    print("   - 'Тип назначаемого л/с' (для пспрс укажите 'пспрс', для кафедр укажите 'пс')")
    print("   - 'Вид'")
    print("   - 'Дата начала' (в формате DD.MM.YYYY)")
    print("   - 'Дата окончания' (в формате DD.MM.YYYY)")
    print("3. Ниже ячейки F11 разместите праздничные дни")
    print("\n4. Логика для строк 'пспрс':")
    print("   - P14: общее количество рабочих дней для распределения")
    print("   - Q14: общее количество выходных/праздничных дней для распределения")
    print("   - N8: вес для ПС (например, 0.6 для 60%)")
    print("   - N9: вес для ПРС (например, 0.4 для 40%)")
    print("   - Сумма весов может быть любой, программа их нормализует")
    print("\n5. Логика для строк 'пс':")
    print("   - Все доступные дни в строках 'пс' будут распределены по кафедрам")
    print("\n6. В столбцах M и N, начиная со строки 21, укажите кафедры и их весовые коэффициенты:")
    print("   - Столбец M: названия кафедр (например, 'каф. 28', 'каф. 41')")
    print("   - Столбец N: весовые коэффициенты (рекомендуется, чтобы сумма была = 1)")

    print("   Все дежурства (из строк 'пспрс' и 'пс') собираются в общий пул")
    print("   и распределяются пропорционально весам кафедр")
    print("\n7. Запустите программу")
    print("8. Результат будет на листе 'Распределение'")

if __name__ == "__main__":
    print_sample_instructions()
    
    # Проверяем существование файла
    try:
        create_duty_schedule()
    except FileNotFoundError:
        print("\n" + "=" * 70)
        print("⚠  Файл 'Дежурства.xlsx' не найден!")
        print("=" * 70)
        create_sample = input("Создать тестовый файл с параметрами для пспрс и кафедр? (да/нет): ").lower()
        if create_sample in ['да', 'yes', 'y', 'д']:
            create_sample_file_with_holidays_and_psprc()
            print("\n" + "=" * 70)
            print("🚀 Запускаю создание расписания...")
            print("=" * 70)
            create_duty_schedule()
        else:
            print("❌ Программа завершена. Создайте файл 'Дежурства.xlsx' вручную.")