import math
import datetime
import os

# Константы
R = 8314.462618  # Универсальная газовая постоянная, Дж/(кмоль·К)
P_std = 0.101325  # Стандартное давление, МПа
T_std = 293.15    # Стандартная температура, K (20°C)

# База данных компонентов с полными свойствами
components_db = [
    {"name": "Метан", "eng": "Methane", "formula": "CH4", 
     "M": 16.043, "Tc": 190.56, "Pc": 4.599, "w": 0.011, 
     "H0": 50.01, "H0_vol": 35.88, "rho_std": 0.716},
    
    {"name": "Азот", "eng": "Nitrogen", "formula": "N2", 
     "M": 28.013, "Tc": 126.19, "Pc": 3.396, "w": 0.037, 
     "H0": 0.0, "H0_vol": 0.0, "rho_std": 1.250},
    
    {"name": "Диоксид углерода", "eng": "CO2", "formula": "CO2", 
     "M": 44.010, "Tc": 304.13, "Pc": 7.377, "w": 0.224, 
     "H0": 0.0, "H0_vol": 0.0, "rho_std": 1.977},
    
    {"name": "Этан", "eng": "Ethane", "formula": "C2H6", 
     "M": 30.070, "Tc": 305.32, "Pc": 4.872, "w": 0.099, 
     "H0": 47.52, "H0_vol": 63.78, "rho_std": 1.356},
    
    {"name": "Пропан", "eng": "Propane", "formula": "C3H8", 
     "M": 44.097, "Tc": 369.83, "Pc": 4.248, "w": 0.152, 
     "H0": 46.35, "H0_vol": 91.25, "rho_std": 2.010},
    
    {"name": "и-Бутан", "eng": "iC4", "formula": "C4H10", 
     "M": 58.123, "Tc": 408.14, "Pc": 3.648, "w": 0.176, 
     "H0": 45.59, "H0_vol": 118.6, "rho_std": 2.703},
    
    {"name": "н-Бутан", "eng": "nC4", "formula": "C4H10", 
     "M": 58.123, "Tc": 425.12, "Pc": 3.796, "w": 0.193, 
     "H0": 45.75, "H0_vol": 118.6, "rho_std": 2.703},
    
    {"name": "и-Пентан", "eng": "iC5", "formula": "C5H12", 
     "M": 72.150, "Tc": 460.43, "Pc": 3.381, "w": 0.227, 
     "H0": 45.11, "H0_vol": 145.5, "rho_std": 3.325},
    
    {"name": "н-Пентан", "eng": "nC5", "formula": "C5H12", 
     "M": 72.150, "Tc": 469.70, "Pc": 3.370, "w": 0.251, 
     "H0": 45.11, "H0_vol": 145.5, "rho_std": 3.325}
]

comp_dict = {comp["eng"]: comp for comp in components_db}

def calculate_properties_simple(composition, T_K, P_MPa):
    """
    Упрощенный расчет свойств газа (запасной вариант)
    """
    M_mix = 0
    Tc_mix = 0
    Pc_mix = 0
    Hi_mix = 0
    Hs_mix = 0
    
    for eng, yi in composition.items():
        if eng in comp_dict:
            comp = comp_dict[eng]
            yi_mol = yi / 100
            M_mix += yi_mol * comp["M"]
            Tc_mix += yi_mol * comp["Tc"]
            Pc_mix += yi_mol * comp["Pc"]
            Hi_mix += yi_mol * comp["H0_vol"]
            
            if comp["eng"] == "Methane":
                Hs_mix += yi_mol * comp["H0_vol"] * 1.11
            elif comp["eng"] in ["Ethane", "Propane", "iC4", "nC4", "iC5", "nC5"]:
                Hs_mix += yi_mol * comp["H0_vol"] * 1.09
            else:
                Hs_mix += yi_mol * comp["H0_vol"]
    
    Tpr = T_K / Tc_mix
    Ppr = P_MPa / Pc_mix
    
    Z_work = 1 - 0.185 * Ppr / Tpr + 0.012 * (Ppr / Tpr)**2
    Z_work = max(0.85, min(1.0, Z_work))
    Z_std = 0.998
    
    rho_work = (P_MPa * 1e6 * M_mix) / (Z_work * R * T_K)
    rho_std = (P_std * 1e6 * M_mix) / (Z_std * R * T_std)
    
    d = rho_std / 1.293
    Wobbe = Hi_mix / math.sqrt(d)
    K_vol = (P_std / P_MPa) * (T_K / T_std) * (Z_work / Z_std)
    kappa = 1.275
    R_specific = 8314 / M_mix
    c_sound = math.sqrt(kappa * Z_work * R_specific * T_K)
    mu = 1.12e-5 * (T_K / 293.15) ** 0.67
    
    return {
        "M_mix": M_mix,
        "Z_std": Z_std,
        "Z_work": Z_work,
        "rho_std": rho_std,
        "rho_work": rho_work,
        "d": d,
        "Hi": Hi_mix,
        "Hs": Hs_mix,
        "Wobbe": Wobbe,
        "kappa": kappa,
        "c_sound": c_sound,
        "mu": mu,
        "K_vol": K_vol,
        "source": "simple"
    }

def calculate_properties_aga8(composition, T_K, P_MPa):
    """
    Расчет свойств газа с использованием библиотеки gascompressibility
    """
    try:
        import gascompressibility as gc
        
        total = 0
        for eng, val in composition.items():
            total += val / 100
        
        M_mix = 0
        Tc_mix = 0
        Pc_mix = 0
        Hi_mix = 0
        Hs_mix = 0
        
        for eng, yi in composition.items():
            if eng in comp_dict:
                comp = comp_dict[eng]
                yi_mol = yi / 100 / total
                M_mix += yi_mol * comp["M"]
                Tc_mix += yi_mol * comp["Tc"]
                Pc_mix += yi_mol * comp["Pc"]
                Hi_mix += yi_mol * comp["H0_vol"]
                
                if comp["eng"] == "Methane":
                    Hs_mix += yi_mol * comp["H0_vol"] * 1.11
                elif comp["eng"] in ["Ethane", "Propane", "iC4", "nC4", "iC5", "nC5"]:
                    Hs_mix += yi_mol * comp["H0_vol"] * 1.09
                else:
                    Hs_mix += yi_mol * comp["H0_vol"]
        
        rho_gas_std = (P_std * 1e6 * M_mix) / (0.998 * R * T_std)
        sg = rho_gas_std / 1.293
        
        T_F = (T_K - 273.15) * 9/5 + 32
        P_psia = P_MPa * 145.038
        T_std_F = 68.0
        P_std_psia = P_std * 145.038
        
        Z_work = gc.calc_z(P=P_psia, T=T_F, sg=sg)
        Z_std = gc.calc_z(P=P_std_psia, T=T_std_F, sg=sg)
        
        rho_work = (P_MPa * 1e6 * M_mix) / (Z_work * R * T_K)
        rho_std = (P_std * 1e6 * M_mix) / (Z_std * R * T_std)
        
        d = rho_std / 1.293
        Wobbe = Hi_mix / math.sqrt(d)
        K_vol = (P_std / P_MPa) * (T_K / T_std) * (Z_work / Z_std)
        kappa = 1.275
        R_specific = 8314 / M_mix
        c_sound = math.sqrt(kappa * Z_work * R_specific * T_K)
        mu = 1.12e-5 * (T_K / 293.15) ** 0.67
        
        return {
            "M_mix": M_mix,
            "Z_std": Z_std,
            "Z_work": Z_work,
            "rho_std": rho_std,
            "rho_work": rho_work,
            "d": d,
            "Hi": Hi_mix,
            "Hs": Hs_mix,
            "Wobbe": Wobbe,
            "kappa": kappa,
            "c_sound": c_sound,
            "mu": mu,
            "K_vol": K_vol,
            "source": "gascompressibility"
        }
        
    except Exception as e:
        print(f"\n⚠️ Ошибка в gascompressibility: {e}")
        print("⚠️ Переключаюсь на упрощенные формулы")
        return calculate_properties_simple(composition, T_K, P_MPa)

def pipeline_calculation(Q_std, D_out_mm, wall_thickness_mm, props):
    """
    Расчет параметров трубопровода
    """
    Q_work = Q_std * props["K_vol"]
    G = Q_std * props["rho_std"]
    
    D_in_m = (D_out_mm - 2 * wall_thickness_mm) / 1000
    A = math.pi * (D_in_m ** 2) / 4
    
    v = (Q_work / 3600) / A
    Re = (props["rho_work"] * v * D_in_m) / props["mu"]
    
    if Re < 2300:
        regime = "ламинарный"
    elif Re < 4000:
        regime = "переходный"
    else:
        regime = "турбулентный"
    
    return {
        "Q_std": Q_std,
        "Q_work": Q_work,
        "G": G,
        "D_in_mm": D_in_m * 1000,
        "A": A,
        "v": v,
        "Re": Re,
        "regime": regime
    }

def distribute_remaining(composition, remaining):
    """
    Равномерно распределяет оставшиеся проценты
    """
    if len(composition) == 0 or remaining <= 0:
        return composition
    
    print(f"\n⚠️ Не хватает {remaining:.2f}%")
    choice = input("Распределить равномерно? (д/н): ")
    
    if choice.lower() == 'д':
        add_each = remaining / len(composition)
        for comp in composition:
            composition[comp] += add_each
        print(f"✅ Добавлено по {add_each:.3f}%")
    
    return composition

def export_to_excel(props, pipe, composition, T_c, P_MPa, Q_std, D_out, wall):
    """
    Экспорт результатов в Excel с красивым форматированием
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("\n⚠️ pandas не установлен. Для экспорта выполните: pip3 install pandas openpyxl")
        return
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"gas_calc_results_{timestamp}.xlsx"
    
    # Создаем книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Расчет газа"
    
    # Настройки стилей
    title_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # ===== ЗАГОЛОВОК =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1)
    cell.value = "КАЛЬКУЛЯТОР СВОЙСТВ ПРИРОДНОГО ГАЗА"
    cell.font = title_font
    cell.alignment = center_align
    cell.fill = blue_fill
    row += 2
    
    # ===== ИСХОДНЫЕ ДАННЫЕ =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1)
    cell.value = "ИСХОДНЫЕ ДАННЫЕ"
    cell.font = header_font
    cell.alignment = center_align
    cell.fill = light_blue_fill
    cell.border = thin_border
    row += 1
    
    # Состав газа
    ws.cell(row=row, column=1, value="Состав газа").font = bold_font
    ws.cell(row=row, column=1).fill = light_yellow_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="Мольная доля, %").font = bold_font
    ws.cell(row=row, column=2).fill = light_yellow_fill
    ws.cell(row=row, column=2).border = thin_border
    row += 1
    
    for comp, val in sorted(composition.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=comp).font = normal_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=round(val, 3)).font = normal_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = right_align
        row += 1
    
    ws.cell(row=row, column=1, value="СУММА").font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=round(sum(composition.values()), 3)).font = bold_font
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = right_align
    row += 2
    
    # Параметры процесса
    process_data = [
        ("Температура, °C", T_c),
        ("Давление, МПа (абс.)", P_MPa),
        ("Расход при ст.усл., м³/ч", Q_std),
        ("Наружный диаметр трубы, мм", D_out),
        ("Толщина стенки, мм", wall)
    ]
    
    for param, value in process_data:
        ws.cell(row=row, column=1, value=param).font = normal_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=value).font = normal_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = right_align
        row += 1
    row += 1
    
    # ===== РЕЗУЛЬТАТЫ РАСЧЕТА =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1)
    cell.value = "РЕЗУЛЬТАТЫ РАСЧЕТА"
    cell.font = header_font
    cell.alignment = center_align
    cell.fill = light_green_fill
    cell.border = thin_border
    row += 1
    
    # Свойства газа
    gas_props = [
        ("Метод расчета", props['source']),
        ("Молярная масса, кг/кмоль", props['M_mix']),
        ("Z при ст. условиях", props['Z_std']),
        ("Z при раб. условиях", props['Z_work']),
        ("Z_раб/Z_ст", props['Z_work']/props['Z_std']),
        ("Плотность при ст.усл., кг/м³", props['rho_std']),
        ("Плотность при раб.усл., кг/м³", props['rho_work']),
        ("Относительная плотность", props['d']),
        ("Низшая теплота сгорания, МДж/м³", props['Hi']),
        ("Высшая теплота сгорания, МДж/м³", props['Hs']),
        ("Число Воббе, МДж/м³", props['Wobbe']),
        ("Скорость звука, м/с", props['c_sound']),
        ("Вязкость, Па·с", props['mu']),
        ("Коэффициент пересчета K_vol", props['K_vol'])
    ]
    
    for param, value in gas_props:
        ws.cell(row=row, column=1, value=param).font = normal_font
        ws.cell(row=row, column=1).border = thin_border
        if isinstance(value, float):
            if "МДж" in param or "кг/м³" in param or "м/с" in param:
                ws.cell(row=row, column=2, value=round(value, 2)).font = normal_font
            elif "Z" in param:
                ws.cell(row=row, column=2, value=round(value, 5)).font = normal_font
            elif "вязкость" in param.lower():
                ws.cell(row=row, column=2, value=f"{value:.3e}").font = normal_font
            else:
                ws.cell(row=row, column=2, value=round(value, 4)).font = normal_font
        else:
            ws.cell(row=row, column=2, value=value).font = normal_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = right_align
        row += 1
    row += 1
    
    # Трубопроводный расчет
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1)
    cell.value = "ТРУБОПРОВОДНЫЙ РАСЧЕТ"
    cell.font = header_font
    cell.alignment = center_align
    cell.fill = light_yellow_fill
    cell.border = thin_border
    row += 1
    
    pipe_props = [
        ("Внутренний диаметр, мм", pipe['D_in_mm']),
        ("Площадь сечения, м²", pipe['A']),
        ("Расход при раб.усл., м³/ч", pipe['Q_work']),
        ("Массовый расход, кг/ч", pipe['G']),
        ("Скорость газа, м/с", pipe['v']),
        ("Число Рейнольдса", f"{pipe['Re']:.0f}"),
        ("Режим течения", pipe['regime'])
    ]
    
    for param, value in pipe_props:
        ws.cell(row=row, column=1, value=param).font = normal_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=value).font = normal_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = right_align
        row += 1
    
    # Автоподбор ширины колонок
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 25
    
    # Сохраняем файл
    wb.save(filename)
    
    print(f"\n✅ Результаты сохранены: {filename}")
    print(f"   Полный путь: {os.path.abspath(filename)}")
    print(f"   Файл содержит все исходные данные и результаты на одном листе")

# ==================== ОСНОВНАЯ ПРОГРАММА ====================

print("="*80)
print("     КАЛЬКУЛЯТОР СВОЙСТВ ПРИРОДНОГО ГАЗА И ТРУБОПРОВОДА")
print("="*80)
print("\nВвод состава газа (мольные %, сумма 100%)")
print("-"*80)

composition = {}
total = 0

for i, comp in enumerate(components_db):
    if abs(total - 100.0) < 0.0001:
        print(f"\n✅ Сумма 100%. Остальные пропущены.")
        break
    
    print(f"\n[{i+1}/{len(components_db)}] Текущая сумма: {total:.3f}%")
    value_str = input(f"{comp['name']} ({comp['formula']}) [%]: ")
    
    if value_str.lower() == 'стоп':
        break
    
    if value_str.strip() == "" or value_str.strip() == "0":
        continue
    
    try:
        value = float(value_str)
        if value < 0:
            print("   ⚠️ Отрицательное значение")
            continue
        
        if total + value > 100.0 + 0.01:
            print(f"   ⚠️ Максимум: {100-total:.3f}%")
            continue
        
        composition[comp['eng']] = value
        total += value
        print(f"   ✅ Добавлено: {value}% (сумма: {total:.3f}%)")
        
    except ValueError:
        print("   ⚠️ Некорректное число")

if total == 0:
    print("\n❌ Состав не введен. Использую 100% метан.")
    composition = {'Methane': 100.0}
    total = 100.0
elif total < 99.99:
    remaining = 100 - total
    composition = distribute_remaining(composition, remaining)

print("\n" + "="*80)
print("     ИТОГОВЫЙ СОСТАВ")
print("="*80)
for comp, val in sorted(composition.items(), key=lambda x: x[1], reverse=True):
    print(f"   {comp:15} : {val:6.3f}%")
print(f"\n   {'СУММА':15} : {sum(composition.values()):6.3f}%")

print("\n" + "-"*80)
print("ВВОД ПАРАМЕТРОВ")
print("-"*80)
print("\n⚠️ Давление вводится АБСОЛЮТНОЕ (МПа)")
print("   Если показания манометра, добавьте 0.101325 МПа")

try:
    T_c = float(input("\nТемпература (°C): "))
    P_MPa = float(input("Давление (МПа): "))
except:
    print("❌ Ошибка ввода. Использую T=20°C, P=0.1 МПа")
    T_c = 20
    P_MPa = 0.1

T_K = T_c + 273.15

# Расчет свойств газа
print("\n" + "="*80)
print("     РАСЧЕТ СВОЙСТВ ГАЗА")
print("="*80)

props = calculate_properties_aga8(composition, T_K, P_MPa)

print("\n" + "="*80)
print("     СВОЙСТВА ГАЗА")
print("="*80)

Z_ratio = props['Z_work'] / props['Z_std']

print(f"\n📊 СОСТАВ:")
print(f"   Молярная масса: {props['M_mix']:.3f} кг/кмоль")
print(f"\n📊 СЖИМАЕМОСТЬ:")
print(f"   Z ст.усл.: {props['Z_std']:.5f}")
print(f"   Z раб.усл.: {props['Z_work']:.5f}")
print(f"   Z_раб/Z_ст = {Z_ratio:.5f}")
print(f"\n📊 ПЛОТНОСТЬ:")
print(f"   ст.усл.: {props['rho_std']:.3f} кг/м³")
print(f"   раб.усл.: {props['rho_work']:.3f} кг/м³")
print(f"   отн.: {props['d']:.4f}")
print(f"\n🔥 ТЕПЛОТА СГОРАНИЯ:")
print(f"   Низшая: {props['Hi']:.2f} МДж/м³")
print(f"   Высшая: {props['Hs']:.2f} МДж/м³")
print(f"   Число Воббе: {props['Wobbe']:.2f} МДж/м³")
print(f"\n🎵 ТЕРМОДИНАМИКА:")
print(f"   Скорость звука: {props['c_sound']:.2f} м/с")
print(f"   Вязкость: {props['mu']:.3e} Па·с")
print(f"\n🔄 ПЕРЕСЧЕТ ОБЪЕМА:")
print(f"   K_vol = {props['K_vol']:.5f}")

print("\n" + "="*80)
print("     ТРУБОПРОВОДНЫЙ РАСЧЕТ")
print("="*80)

try:
    Q_std = float(input("\nРасход при ст.усл. (м³/ч): ") or "1000")
    D_out = float(input("Наружный диаметр (мм): ") or "720")
    wall = float(input("Толщина стенки (мм): ") or "12")
except:
    Q_std, D_out, wall = 1000, 720, 12

pipe = pipeline_calculation(Q_std, D_out, wall, props)

print(f"\n📊 РЕЗУЛЬТАТЫ:")
print(f"   Внутренний диаметр: {pipe['D_in_mm']:.1f} мм")
print(f"   Площадь сечения: {pipe['A']:.4f} м²")
print(f"   Расход раб.усл.: {pipe['Q_work']:.3f} м³/ч")
print(f"   Массовый расход: {pipe['G']:.1f} кг/ч")
print(f"   Скорость газа: {pipe['v']:.2f} м/с")
print(f"   Число Рейнольдса: {pipe['Re']:.0f}")
print(f"   Режим течения: {pipe['regime']}")

print("\n" + "="*80)
print("     ЭКСПОРТ РЕЗУЛЬТАТОВ")
print("="*80)

if input("\nСохранить в Excel? (д/н): ").lower() == 'д':
    export_to_excel(props, pipe, composition, T_c, P_MPa, Q_std, D_out, wall)

print("\n" + "="*80)
print("✅ РАСЧЕТ ЗАВЕРШЕН")
print("="*80)
