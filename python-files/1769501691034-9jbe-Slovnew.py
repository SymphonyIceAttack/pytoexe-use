from openpyxl import load_workbook
from tkinter import Tk, filedialog
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import os
import re

# --- ШАГ 0: выбрать файл ---
Tk().withdraw()
file_path = filedialog.askopenfilename(
    title="Выберите Excel файл",
    filetypes=[("Excel files", "*.xlsx")]
)

if not file_path:
    print("Файл не выбран")
    exit()

wb = load_workbook(file_path)

# --- Параметры ---
split_words = ["SGL", "DBL", "TPL", "QDP"]  # слова-разделители
pattern = r"\b(" + "|".join(split_words) + r")\b"

replacements = {
    "+  EXB +  EXB +  EXB +  EXB": "+ 4EXB",
    "+  EXB +  EXB +  EXB": "+ 3EXB",
    " (EUR)": " ",
    "+  EXB +  EXB": "+ 2EXB",
    "SHB": "HBT Small",
    "LHB": "HBT Large",
    "SFB": "FBT Small",
    "LFB": "FBT Large",
    "VF1": "FBT1 Vis Vita",
    "VF2": "FBT2 Vis Vita",
    "HB1": "HBT1 Vis Vita",
    "HB2": "HBT2 Vis Vita",
    "-17)": "-17,99)",
    "-12)": "-12,99)",
    "-11)": "-11,99)",
    "-9)": "-9,99)",
    "-8)": "-8,99)",
    "-7)": "-7,99)",
    "-6)": "-6,99)",
    "-5)": "-5,99)",
    "-4)": "-4,99)",
    "-3)": "-2,99)",
    "-2)": "-2,99)",
    "-1)": "-1,99)",
    "wob": "sharing",
    # можно добавлять новые
}
# сортируем по длине ключа, чтобы избежать конфликтов
replacements = dict(sorted(replacements.items(), key=lambda x: len(x[0]), reverse=True))

highlight_words = [
    "BB", "HB", "FB", "AI", "RMB", "ROM",
    "HBT Small", "HBT Large",
    "FBT Small", "FBT Large",
    "FBT1 Vis Vita", "FBT2 Vis Vita",
    "HBT1 Vis Vita", "HBT2 Vis Vita"
]

yellow_fill = PatternFill(
    start_color="FFFF99",
    end_color="FFFF99",
    fill_type="solid"
)

# --- ШАГ 1: удалить строки без слова-разделителя и с пустым столбцом B ---
for sheet in wb.worksheets:
    for row in range(sheet.max_row, 0, -1):  # идём с конца
        cell_a = sheet[f"A{row}"]
        cell_b = sheet[f"B{row}"]
        
        # проверка слова-разделителя в столбце A
        if isinstance(cell_a.value, str):
            a_has_word = bool(re.search(pattern, cell_a.value))
        else:
            a_has_word = False
        
        # проверка пустого столбца B
        b_empty = (cell_b.value is None or str(cell_b.value).strip() == "")
        
        # удаляем только если A не содержит слово и B пустой
        if not a_has_word and b_empty:
            sheet.delete_rows(row)

# --- ШАГ 2: сделать замены ---
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for old, new in replacements.items():
                    if old in cell.value:
                        cell.value = cell.value.replace(old, new)

# --- ШАГ 3: вставка нового столбца B и копирование A → B ---
for sheet in wb.worksheets:
    sheet.insert_cols(2)  # новый B, старый сдвигается
    max_row = sheet.max_row
    for row in range(1, max_row + 1):
        sheet[f'B{row}'] = sheet[f'A{row}'].value

# --- ШАГ 4: разделение строк по ключевым словам (только столбец A) ---
for sheet in wb.worksheets:
    max_row = sheet.max_row
    for row in range(1, max_row + 1):
        cell_a = sheet[f"A{row}"]
        if not isinstance(cell_a.value, str):
            continue
        match = re.search(pattern, cell_a.value)
        if not match:
            continue
        before = cell_a.value[:match.start()].strip()  # всё до слова
        after = cell_a.value[match.start():].strip()   # слово-разделитель + остаток
        sheet[f"A{row}"] = before
        sheet[f"B{row}"] = after

# --- ШАГ 5: безопасное чередование в новый лист ---
for sheet in wb.worksheets:
    if sheet.title.endswith("_mixed"):
        continue

    new_sheet = wb.create_sheet(title=f"{sheet.title}_mixed")
    new_row = 1
    max_col = sheet.max_column

    for row in range(1, sheet.max_row + 1):
        # A → отдельная строка
        val_a = sheet[f"A{row}"].value
        if val_a not in (None, ""):
            new_sheet[f"A{new_row}"] = val_a
            new_row += 1

        # B, C, D, E... → одна строка
        write_col = 1
        has_data = False

        for col in range(2, max_col + 1):
            val = sheet.cell(row=row, column=col).value
            if val not in (None, ""):
                new_sheet.cell(row=new_row, column=write_col, value=val)
                write_col += 1
                has_data = True

        if has_data:
            new_row += 1

# --- ШАГ 6: подсветка столбца A на _mixed листах + подсветка столбца A на _mixed листах ---
for sheet in wb.worksheets:
    if not sheet.title.endswith("_mixed"):
        continue

    # идём снизу вверх, чтобы вставка строк не сбивала индексы
    for row in range(sheet.max_row, 0, -1):
        cell = sheet[f"A{row}"]

        if not isinstance(cell.value, str):
            continue

        text = cell.value.lower()

        for word in highlight_words:
            # \b — граница слова
            # re.escape чтобы экранировать спецсимволы в слов
            if re.search(r'\b' + re.escape(word) + r'\b', text, flags=re.IGNORECASE):
                # вставка 2 пустых строк перед строкой
                sheet.insert_rows(row, amount=2)
                # подсветка смещенной строки
                sheet[f"A{row + 2}"].fill = yellow_fill
                break  # если слово найдено, больше искать не нужно


# --- ШАГ 7: сохранить новый файл рядом с оригиналом ---
dir_name = os.path.dirname(file_path)
base_name = os.path.splitext(os.path.basename(file_path))[0]
new_file = os.path.join(dir_name, f"{base_name}_new.xlsx")

wb.save(new_file)
print(f"Готово! Файл сохранён как:\n{new_file}")
