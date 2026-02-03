import fitz
import os
import re
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from copy import copy

# =============================
# CONFIGURACIÓN
# =============================
CARPETA_PDFS = r"C:\Users\caherna5\OneDrive - Ryder\Desktop\GUIAS"
ARCHIVO_EXCEL = r"C:\Users\caherna5\Ryder\OPERACIÓN CARRIER - General\BITACORA FEDEX.xlsx"
HOJA = "VALIDACION 2026"
TABLA = "Validacion"


# =============================
# LECTURA DE PDF
# =============================
def extraer_texto_pdf(ruta):
    paginas = []
    with fitz.open(ruta) as doc:
        for page in doc:
            paginas.append(page.get_text())
    return paginas


# =============================
# PROCESO PRINCIPAL
# =============================
def ejecutar_proceso():
    try:
        ordenes = {}
        guias = {}

        archivos = [f for f in os.listdir(CARPETA_PDFS) if f.lower().endswith(".pdf")]

        if not archivos:
            messagebox.showwarning("Aviso", "No hay PDFs en la carpeta.")
            return

        # -----------------------------
        # PROCESAR PDFs
        # -----------------------------
        for archivo in archivos:
            ruta = os.path.join(CARPETA_PDFS, archivo)

            try:
                paginas = extraer_texto_pdf(ruta)
            except:
                continue

            for texto in paginas:
                texto_limpio = re.sub(r'(?<=\d)\s+(?=\d)', '', texto)

                # REFERENCIA (9 dígitos)
                ordenes_encontradas = re.findall(r'\b53\d{7}\b', texto_limpio)

                # PO (11 dígitos)
                pos = re.findall(r'\b53\d{9}\b', texto_limpio)
                po_detectado = pos[0] if pos else ""

                # NOMBRE
                nombre = ""
                nombre_match = re.search(
                    r'\d{4}/\d{2}/\d{2}\s+(?:Sr\.|Sra\.|SR\.|SRA\.)?\s*([A-Za-zÁÉÍÓÚÑáéíóúñ\s]{8,80})',
                    texto
                )
                if nombre_match:
                    nombre = re.sub(r'\s{2,}', ' ', nombre_match.group(1).strip())

                for orden in ordenes_encontradas:
                    if orden not in ordenes:
                        ordenes[orden] = {"po": po_detectado, "nombre": nombre}

                # REF en PDF de guías
                ref_match = re.search(r'REF[:\s]*?(54\d{7})', texto_limpio)
                if not ref_match:
                    continue

                ref = ref_match.group(1)

                # GUIAS (88 + 12 dígitos aunque estén separadas)
                for linea in texto.splitlines():
                    if sum(c.isdigit() for c in linea) >= 10:
                        limpio = re.sub(r'\D', '', linea)
                        if limpio.startswith("88") and len(limpio) == 12:
                            guias.setdefault(ref, set()).add(limpio)

        # -----------------------------
        # ARMAR FILAS FINALES
        # -----------------------------
        filas = []

        for orden, datos in ordenes.items():
            po = datos["po"]
            nombre = datos["nombre"]
            lista_guias = sorted(guias.get(orden, set()))

            if lista_guias:
                for guia in lista_guias:
                    filas.append([orden, po, nombre, guia])
            else:
                filas.append([orden, po, nombre, ""])

        if not filas:
            messagebox.showwarning("Resultado", "No se generaron filas.")
            return

        # -----------------------------
        # INSERTAR EN TABLA DE EXCEL
        # -----------------------------
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb[HOJA]

        if TABLA not in ws.tables:
            messagebox.showerror("Error", f"No se encontró la tabla '{TABLA}'.")
            return

        tabla = ws.tables[TABLA]
        min_col, min_row, max_col, max_row = range_boundaries(tabla.ref)

        fila_insercion = max_row + 1

        # Escribir datos nuevos
        for i, fila in enumerate(filas):
            for j, valor in enumerate(fila):
                ws.cell(row=fila_insercion + i, column=min_col + j, value=valor)

        # -----------------------------
        # COPIAR FORMULAS Y FORMATOS
        # -----------------------------
        fila_origen = max_row

        for col in range(min_col, max_col + 1):
            celda_origen = ws.cell(row=fila_origen, column=col)

            for i in range(len(filas)):
                celda_destino = ws.cell(row=fila_insercion + i, column=col)

                # Copiar fórmula si existe
                if celda_origen.value and isinstance(celda_origen.value, str) and celda_origen.value.startswith("="):
                    celda_destino.value = celda_origen.value

                # Copiar formato siempre
                celda_destino._style = copy(celda_origen._style)

        # -----------------------------
        # EXPANDIR TABLA
        # -----------------------------
        nuevo_max_row = max_row + len(filas)
        tabla.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{nuevo_max_row}"

        wb.save(ARCHIVO_EXCEL)

        messagebox.showinfo(
            "Proceso terminado",
            f"Proceso exitoso.\n\nFilas agregadas: {len(filas)}"
        )

    except Exception as e:
        messagebox.showerror("Error crítico", str(e))


# =============================
# INTERFAZ
# =============================
ventana = tk.Tk()
ventana.title("Extractor FedEx → Validación")
ventana.geometry("560x300")
ventana.resizable(False, False)

boton = tk.Button(
    ventana,
    text="Extraer guías y enviar a Validación",
    font=("Arial", 14, "bold"),
    padx=30,
    pady=15,
    command=ejecutar_proceso
)

boton.pack(expand=True)

ventana.mainloop()
