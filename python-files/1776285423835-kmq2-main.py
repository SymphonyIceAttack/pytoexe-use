import sys
import os
import sqlite3

try:
    import qtawesome as qta
    def _qicon(name):
        return qta.icon(name)
except ImportError:
    def _qicon(name):
        return QIcon()
from datetime import datetime

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QDialog, QFormLayout,
    QLineEdit, QComboBox, QLabel, QMessageBox, QHeaderView, QTabWidget,
    QAbstractItemView, QToolBar, QDateEdit, QTextEdit, QGroupBox,
    QStatusBar, QFileDialog, QSpinBox, QDoubleSpinBox, QDialogButtonBox,
    QSizePolicy, QFrame
)
from PyQt5.QtCore import Qt, QDate, QSize
from PyQt5.QtGui import QFont, QColor, QIcon

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "database.db")

QSS = """
QMainWindow, QWidget {
    background-color: #F8EBEB;
    font-family: "Georgia";
    font-size: 11px;
    color: #3D1A1A;
}
QToolBar {
    background-color: #8B3A3A;
    border: none;
    padding: 4px;
    spacing: 6px;
}
QToolBar QLabel {
    color: white;
    font-weight: bold;
    font-size: 13px;
    padding: 0 12px;
}
QToolBar QPushButton {
    background-color: transparent;
    color: white;
    border: 1px solid rgba(255,255,255,0.3);
    border-radius: 4px;
    padding: 5px 14px;
    font-weight: bold;
}
QToolBar QPushButton:hover {
    background-color: rgba(255,255,255,0.15);
    border-color: rgba(255,255,255,0.5);
}
QPushButton {
    background-color: #8B3A3A;
    color: white;
    border: none;
    border-radius: 4px;
    padding: 7px 18px;
    font-weight: bold;
    font-size: 11px;
}
QPushButton:hover {
    background-color: #7A2E2E;
}
QPushButton:pressed {
    background-color: #6A2424;
}
QPushButton:disabled {
    background-color: #B8A0A0;
    color: #E0D0D0;
}
QPushButton#btnDanger {
    background-color: #C0392B;
}
QPushButton#btnDanger:hover {
    background-color: #A93226;
}
QTabWidget::pane {
    border: 1px solid #D4A8A8;
    background-color: #F8EBEB;
    border-radius: 4px;
}
QTabBar::tab {
    background-color: #E5D0D0;
    border: 1px solid #D4A8A8;
    border-bottom: none;
    padding: 6px 16px;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
    margin-right: 2px;
    color: #3D1A1A;
}
QTabBar::tab:selected {
    background-color: #F8EBEB;
    font-weight: bold;
}
QTabBar::tab:hover:!selected {
    background-color: #EDD8D8;
}
QTableWidget {
    background-color: white;
    alternate-background-color: #F3E6E6;
    border: 1px solid #D4A8A8;
    border-radius: 4px;
    gridline-color: #DEC0C0;
    selection-background-color: #A85050;
    selection-color: white;
}
QTableWidget::item {
    padding: 4px 6px;
}
QHeaderView::section {
    background-color: #8B3A3A;
    color: white;
    padding: 6px;
    border: none;
    border-right: 1px solid #7A2E2E;
    font-weight: bold;
}
QLineEdit, QComboBox, QDateEdit, QSpinBox, QDoubleSpinBox, QTextEdit {
    border: 1px solid #D4A8A8;
    border-radius: 4px;
    padding: 5px 8px;
    background-color: white;
    font-size: 11px;
    color: #3D1A1A;
}
QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QSpinBox:focus,
QDoubleSpinBox:focus, QTextEdit:focus {
    border-color: #A85050;
}
QComboBox::drop-down {
    border: none;
    padding-right: 6px;
}
QComboBox QAbstractItemView {
    background-color: white;
    selection-background-color: #A85050;
    selection-color: white;
}
QGroupBox {
    font-weight: bold;
    border: 1px solid #D4A8A8;
    border-radius: 6px;
    margin-top: 12px;
    padding-top: 16px;
    background-color: rgba(255,255,255,0.5);
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 6px;
    color: #6B3030;
}
QStatusBar {
    background-color: #8B3A3A;
    color: white;
    font-size: 11px;
}
QDialog {
    background-color: #F8EBEB;
}
QLabel#infoLabel {
    color: #6B3030;
    font-style: italic;
    padding: 4px;
}
QFrame#separator {
    background-color: #D4A8A8;
    max-height: 1px;
}

QPushButton { border-radius: 8px; font-size: 13px; font-weight: bold; padding: 10px 26px; }
QPushButton#btnDanger { border-radius: 8px; font-size: 13px; }
QGroupBox { border-width: 2px; border-radius: 10px; }
QTableWidget { border-width: 2px; border-radius: 8px; }
QHeaderView::section { font-size: 13px; padding: 9px 12px; }
QTabBar::tab { font-size: 13px; padding: 11px 26px; border-top-left-radius: 10px; border-top-right-radius: 10px; }
QTabWidget::pane { border-width: 2px; border-radius: 8px; }
QLineEdit, QComboBox, QDateEdit, QSpinBox, QDoubleSpinBox { border-radius: 8px; min-height: 30px; font-size: 12px; border-width: 2px; }
QTextEdit { border-radius: 8px; border-width: 2px; font-size: 12px; }
QLabel { font-size: 12px; }
QDialog { border-radius: 8px; }
"""


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_connection()
    c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        login TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS halls (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        floor INTEGER,
        area_sqm REAL,
        description TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS exhibits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        category TEXT,
        year_created TEXT,
        hall_id INTEGER REFERENCES halls(id),
        condition TEXT,
        description TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS exhibitions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        start_date TEXT,
        end_date TEXT,
        hall_id INTEGER REFERENCES halls(id),
        status TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS movements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        exhibit_id INTEGER REFERENCES exhibits(id),
        from_location TEXT,
        to_location TEXT,
        date_move TEXT,
        reason TEXT,
        responsible TEXT
    )""")

    if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        c.execute("INSERT INTO users (login, password) VALUES (?, ?)",
                  ("admin", "admin"))

    if c.execute("SELECT COUNT(*) FROM halls").fetchone()[0] == 0:
        halls = [
            ("Зал русской живописи", 1, 180.0,
             "Картины XVIII–XIX веков, климат-контроль, влажность 50%"),
            ("Зал скульптуры", 1, 220.0,
             "Мраморные и бронзовые скульптуры, подиумы с подсветкой"),
            ("Зал археологии", 2, 150.0,
             "Витрины с подсветкой, сигнализация, контроль температуры"),
            ("Зал современного искусства", 2, 200.0,
             "Мультимедийное оборудование, проекторы, звук"),
            ("Временный выставочный зал", 3, 160.0,
             "Переносные стенды, направленное освещение"),
        ]
        c.executemany(
            "INSERT INTO halls (name, floor, area_sqm, description) VALUES (?,?,?,?)",
            halls)

    if c.execute("SELECT COUNT(*) FROM exhibits").fetchone()[0] == 0:
        exhibits = [
            ("Портрет купца Третьякова", "Живопись", "1876", 1, "Отличное",
             "Масло, холст, 80x60 см, автор — И.Н. Крамской"),
            ("Утро в сосновом лесу (копия)", "Живопись", "1889", 1, "Хорошее",
             "Масло, холст, 140x100 см, копия работы И.И. Шишкина"),
            ("Вид на Кремль с набережной", "Живопись", "1905", 1, "Отличное",
             "Акварель на бумаге, 50x40 см, неизвестный художник"),
            ("Бюст А.С. Пушкина", "Скульптура", "1937", 2, "Хорошее",
             "Бронза, высота 45 см, автор — Е.Ф. Белашова"),
            ("Давид (уменьшенная копия)", "Скульптура", "1960", 2, "Удовлетворительное",
             "Гипс, высота 180 см, учебная копия"),
            ("Глиняный кувшин", "Археология", "III в. до н.э.", 3, "Удовлетворительное",
             "Керамика, найден при раскопках в Крыму в 1978 г."),
            ("Бронзовый меч скифский", "Археология", "V в. до н.э.", 3, "Хорошее",
             "Бронза, длина 65 см, орнамент звериного стиля"),
            ("Золотая серьга сарматская", "Археология", "II в. н.э.", 3, "Отличное",
             "Золото, скифо-сарматская культура, Причерноморье"),
            ("Инсталляция «Время и пространство»", "Современное искусство", "2018", 4,
             "Отличное", "Смешанная техника, кинетическая инсталляция"),
            ("Цифровая проекция «Космос»", "Современное искусство", "2022", 4,
             "Отличное", "Видео-арт, проекция 360°, длительность 15 мин"),
            ("Фарфоровая ваза «Цветение»", "Декоративное искусство", "1850", 1,
             "Хорошее", "Императорский фарфоровый завод, роспись кобальтом"),
            ("Икона Казанской Богоматери", "Иконопись", "XVII в.", 1,
             "Удовлетворительное", "Дерево, темпера, 40x30 см, требует реставрации"),
            ("Каменный идол половецкий", "Археология", "X в.", 3,
             "Удовлетворительное", "Гранит, высота 90 см, Приазовье"),
            ("Пейзаж «Волга на закате»", "Живопись", "1910", 1, "Хорошее",
             "Масло, холст, 120x80 см, автор — А.К. Саврасов (школа)"),
            ("Мозаика «Весна в городе»", "Современное искусство", "2020", 4,
             "Отличное", "Цветное стекло, 200x150 см, авторская работа"),
            ("Статуэтка «Танцовщица»", "Скульптура", "1925", 2, "Хорошее",
             "Бронза, высота 35 см, стиль ар-деко"),
            ("Амфора чернофигурная (копия)", "Археология", "VI в. до н.э.", 3,
             "Отличное", "Керамика, музейная реплика, высота 50 см"),
        ]
        c.executemany(
            "INSERT INTO exhibits (name, category, year_created, hall_id, "
            "condition, description) VALUES (?,?,?,?,?,?)", exhibits)

    if c.execute("SELECT COUNT(*) FROM exhibitions").fetchone()[0] == 0:
        exhibitions = [
            ("Шедевры русской живописи XVIII–XIX вв.", "2026-01-15", "2026-04-15",
             1, "Открыта"),
            ("Античная археология Причерноморья", "2026-02-01", "2026-05-01",
             3, "Открыта"),
            ("Цифровое искусство XXI века", "2026-03-10", "2026-06-10",
             4, "Открыта"),
            ("Скульптура: от античности до модерна", "2026-04-01", "2026-07-01",
             2, "Планируется"),
            ("Детский фестиваль «Мир творчества»", "2026-05-01", "2026-05-15",
             5, "Планируется"),
            ("Ретроспектива фотографии XX века", "2025-09-01", "2025-12-31",
             5, "Завершена"),
            ("Золото скифов: находки последних лет", "2026-06-01", "2026-09-30",
             3, "Планируется"),
        ]
        c.executemany(
            "INSERT INTO exhibitions (name, start_date, end_date, hall_id, status) "
            "VALUES (?,?,?,?,?)", exhibitions)

    if c.execute("SELECT COUNT(*) FROM movements").fetchone()[0] == 0:
        movements = [
            (1, "Хранилище А-1", "Зал русской живописи", "2026-01-10",
             "Подготовка к выставке", "Петрова Анна Сергеевна"),
            (2, "Хранилище А-1", "Зал русской живописи", "2026-01-10",
             "Подготовка к выставке", "Петрова Анна Сергеевна"),
            (6, "Хранилище Б-2", "Зал археологии", "2026-01-28",
             "Постоянная экспозиция", "Кузнецов Игорь Владимирович"),
            (7, "Хранилище Б-2", "Зал археологии", "2026-01-28",
             "Постоянная экспозиция", "Кузнецов Игорь Владимирович"),
            (9, "Хранилище В-3", "Зал современного искусства", "2026-03-05",
             "Временная выставка", "Соколова Елена Дмитриевна"),
            (10, "Хранилище В-3", "Зал современного искусства", "2026-03-05",
             "Временная выставка", "Соколова Елена Дмитриевна"),
            (4, "Зал скульптуры", "Реставрационная мастерская", "2026-02-15",
             "Реставрация: трещина основания", "Волков Дмитрий Андреевич"),
            (5, "Зал скульптуры", "Реставрационная мастерская", "2026-02-20",
             "Реставрация: очистка поверхности", "Волков Дмитрий Андреевич"),
            (4, "Реставрационная мастерская", "Зал скульптуры", "2026-03-10",
             "Возврат после реставрации", "Волков Дмитрий Андреевич"),
            (11, "Хранилище А-1", "Зал русской живописи", "2026-01-12",
             "Подготовка к выставке", "Петрова Анна Сергеевна"),
            (13, "Хранилище Б-2", "Зал археологии", "2026-01-30",
             "Постоянная экспозиция", "Кузнецов Игорь Владимирович"),
            (15, "Хранилище В-3", "Зал современного искусства", "2026-03-08",
             "Временная выставка", "Соколова Елена Дмитриевна"),
            (3, "Зал русской живописи", "Временный выставочный зал", "2026-03-20",
             "Передача на гастрольную выставку", "Петрова Анна Сергеевна"),
            (16, "Хранилище А-2", "Зал скульптуры", "2026-02-01",
             "Пополнение экспозиции", "Морозова Татьяна Ивановна"),
        ]
        c.executemany(
            "INSERT INTO movements (exhibit_id, from_location, to_location, "
            "date_move, reason, responsible) VALUES (?,?,?,?,?,?)", movements)

    conn.commit()
    conn.close()


class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Авторизация")
        self.setFixedSize(340, 300)
        layout = QVBoxLayout(self)

        logo = QLabel()
        logo.setPixmap(_qicon("fa5s.landmark").pixmap(48, 48))
        logo.setAlignment(Qt.AlignCenter)
        layout.addWidget(logo)

        brand = QLabel("Музей — Управление фондами")
        brand.setAlignment(Qt.AlignCenter)
        brand.setStyleSheet("font-size: 14px; font-weight: bold; color: #8B3A3A; margin-bottom: 4px;")
        layout.addWidget(brand)

        title = QLabel("Музейный фонд")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet(
            "font-size: 20px; font-weight: bold; color: #8B3A3A; margin-bottom: 2px;")
        layout.addWidget(title)

        subtitle = QLabel("Авторизация в системе учёта")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("font-size: 11px; color: #6B3030; margin-bottom: 8px;")
        layout.addWidget(subtitle)

        form = QFormLayout()
        self.login_input = QLineEdit()
        self.login_input.setPlaceholderText("Введите логин")
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Введите пароль")
        self.password_input.setEchoMode(QLineEdit.Password)
        form.addRow("Логин:", self.login_input)
        form.addRow("Пароль:", self.password_input)
        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        self.btn_login = QPushButton(_qicon("fa5s.sign-in-alt"), " Войти")
        self.btn_login.clicked.connect(self.try_login)
        self.btn_cancel = QPushButton(_qicon("fa5s.times-circle"), " Отмена")
        self.btn_cancel.setObjectName("btnDanger")
        self.btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(self.btn_login)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)
        self.password_input.returnPressed.connect(self.try_login)
        self.authenticated = False

    def try_login(self):
        login = self.login_input.text().strip()
        password = self.password_input.text().strip()
        if not login or not password:
            QMessageBox.warning(self, "Ошибка", "Введите логин и пароль.")
            return
        conn = get_connection()
        row = conn.execute("SELECT id FROM users WHERE login=? AND password=?",
                           (login, password)).fetchone()
        conn.close()
        if row:
            self.authenticated = True
            self.accept()
        else:
            QMessageBox.warning(self, "Ошибка", "Неверный логин или пароль.")


class RecordDialog(QDialog):
    def __init__(self, parent, title, fields, values=None, combos=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(440)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.inputs = {}
        for key, label, wtype in fields:
            if wtype == "combo" and combos and key in combos:
                w = QComboBox()
                for cid, ctext in combos[key]:
                    w.addItem(ctext, cid)
                if values and key in values:
                    idx = w.findData(values[key])
                    if idx >= 0:
                        w.setCurrentIndex(idx)
            elif wtype == "date":
                w = QDateEdit()
                w.setCalendarPopup(True)
                w.setDisplayFormat("dd.MM.yyyy")
                if values and key in values and values[key]:
                    w.setDate(QDate.fromString(values[key], "yyyy-MM-dd"))
                else:
                    w.setDate(QDate.currentDate())
            elif wtype == "spin":
                w = QSpinBox()
                w.setRange(0, 999999)
                if values and key in values:
                    w.setValue(int(values[key]) if values[key] else 0)
            elif wtype == "double":
                w = QDoubleSpinBox()
                w.setRange(0, 9999999)
                w.setDecimals(2)
                if values and key in values:
                    w.setValue(float(values[key]) if values[key] else 0)
            elif wtype == "text":
                w = QTextEdit()
                w.setMaximumHeight(60)
                if values and key in values:
                    w.setPlainText(str(values[key]) if values[key] else "")
            else:
                w = QLineEdit()
                if values and key in values:
                    w.setText(str(values[key]) if values[key] else "")
            form.addRow(label + ":", w)
            self.inputs[key] = w
        layout.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("Сохранить")
        btns.button(QDialogButtonBox.Cancel).setText("Отмена")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def get_values(self):
        result = {}
        for key, w in self.inputs.items():
            if isinstance(w, QComboBox):
                result[key] = w.currentData()
            elif isinstance(w, QDateEdit):
                result[key] = w.date().toString("yyyy-MM-dd")
            elif isinstance(w, QSpinBox):
                result[key] = w.value()
            elif isinstance(w, QDoubleSpinBox):
                result[key] = w.value()
            elif isinstance(w, QTextEdit):
                result[key] = w.toPlainText().strip()
            else:
                result[key] = w.text().strip()
        return result


class CrudTab(QWidget):
    def __init__(self, parent, table_name, columns, headers, fields,
                 join_query=None, combos_fn=None):
        super().__init__(parent)
        self.table_name = table_name
        self.columns = columns
        self.headers = headers
        self.fields = fields
        self.join_query = join_query
        self.combos_fn = combos_fn
        self._build_ui()
        self.load_data()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        filter_group = QGroupBox("Поиск и фильтрация")
        filter_layout = QHBoxLayout(filter_group)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск...")
        self.search_input.textChanged.connect(self._filter_table)
        filter_layout.addWidget(QLabel("Поиск:"))
        filter_layout.addWidget(self.search_input)
        self.lbl_count = QLabel("Записей: 0")
        self.lbl_count.setStyleSheet("color: #6B3030; font-weight: bold;")
        filter_layout.addWidget(self.lbl_count)
        layout.addWidget(filter_group)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.doubleClicked.connect(self.edit_record)
        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()
        btn_add = QPushButton(_qicon("fa5s.plus-circle"), " Добавить")
        btn_add.clicked.connect(self.add_record)
        btn_edit = QPushButton(_qicon("fa5s.pen"), " Редактировать")
        btn_edit.clicked.connect(self.edit_record)
        btn_del = QPushButton(_qicon("fa5s.trash-alt"), " Удалить")
        btn_del.setObjectName("btnDanger")
        btn_del.clicked.connect(self.delete_record)
        btn_refresh = QPushButton("Обновить")
        btn_refresh.clicked.connect(self.load_data)
        btn_layout.addWidget(btn_add)
        btn_layout.addWidget(btn_edit)
        btn_layout.addWidget(btn_del)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_refresh)
        layout.addLayout(btn_layout)

    def load_data(self):
        conn = get_connection()
        if self.join_query:
            rows = conn.execute(self.join_query).fetchall()
        else:
            cols = ", ".join(["id"] + self.columns)
            rows = conn.execute(f"SELECT {cols} FROM {self.table_name}").fetchall()
        conn.close()
        self.table.setSortingEnabled(False)
        self.table.setColumnCount(len(self.headers) + 1)
        self.table.setHorizontalHeaderLabels(["ID"] + self.headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setColumnHidden(0, True)
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val is not None else "")
                if c == 0:
                    item.setData(Qt.UserRole, val)
                self.table.setItem(r, c, item)
        self.table.resizeColumnsToContents()
        self.table.setSortingEnabled(True)
        self.lbl_count.setText(f"Записей: {len(rows)}")

    def _filter_table(self, text):
        text = text.lower()
        visible = 0
        for r in range(self.table.rowCount()):
            match = not text
            for c in range(1, self.table.columnCount()):
                item = self.table.item(r, c)
                if item and text in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(r, not match)
            if match:
                visible += 1
        self.lbl_count.setText(f"Записей: {visible}")

    def _get_combos(self):
        if self.combos_fn:
            return self.combos_fn()
        return None

    def add_record(self):
        dlg = RecordDialog(self, "Добавить запись", self.fields,
                           combos=self._get_combos())
        if dlg.exec_() == QDialog.Accepted:
            vals = dlg.get_values()
            conn = get_connection()
            cols = ", ".join(vals.keys())
            placeholders = ", ".join(["?"] * len(vals))
            conn.execute(
                f"INSERT INTO {self.table_name} ({cols}) VALUES ({placeholders})",
                list(vals.values()))
            conn.commit()
            conn.close()
            self.load_data()

    def edit_record(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Информация", "Выберите запись.")
            return
        rec_id = self.table.item(row, 0).data(Qt.UserRole)
        conn = get_connection()
        cols_str = ", ".join(self.columns)
        db_row = conn.execute(
            f"SELECT {cols_str} FROM {self.table_name} WHERE id=?",
            (rec_id,)).fetchone()
        conn.close()
        if not db_row:
            return
        values = {}
        for i, (key, label, wtype) in enumerate(self.fields):
            values[key] = db_row[i]
        dlg = RecordDialog(self, "Редактировать запись", self.fields, values,
                           combos=self._get_combos())
        if dlg.exec_() == QDialog.Accepted:
            vals = dlg.get_values()
            conn = get_connection()
            sets = ", ".join(f"{k}=?" for k in vals.keys())
            conn.execute(f"UPDATE {self.table_name} SET {sets} WHERE id=?",
                         list(vals.values()) + [rec_id])
            conn.commit()
            conn.close()
            self.load_data()

    def delete_record(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Информация", "Выберите запись.")
            return
        rec_id = self.table.item(row, 0).data(Qt.UserRole)
        if QMessageBox.question(
                self, "Подтверждение", "Удалить выбранную запись?",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        conn = get_connection()
        try:
            conn.execute(
                f"DELETE FROM {self.table_name} WHERE id=?", (rec_id,))
            conn.commit()
        except sqlite3.IntegrityError:
            QMessageBox.warning(
                self, "Ошибка",
                "Невозможно удалить: запись связана с другими данными.")
        conn.close()
        self.load_data()


def _make_stat_card(caption, value, color="#8B3A3A"):
    frame = QFrame()
    frame.setStyleSheet(
        f"QFrame {{ background: white; border: 1px solid #D4A8A8; "
        f"border-radius: 6px; padding: 8px; }}")
    vl = QVBoxLayout(frame)
    vl.setContentsMargins(12, 8, 12, 8)
    lbl_val = QLabel(str(value))
    lbl_val.setAlignment(Qt.AlignCenter)
    lbl_val.setStyleSheet(
        f"font-size: 20px; font-weight: bold; color: {color}; border: none;")
    lbl_cap = QLabel(caption)
    lbl_cap.setAlignment(Qt.AlignCenter)
    lbl_cap.setStyleSheet("font-size: 10px; color: #6B3030; border: none;")
    vl.addWidget(lbl_val)
    vl.addWidget(lbl_cap)
    return frame


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Учёт экспонатов и фондов в музее")
        self.resize(1120, 740)
        self.setWindowIcon(_qicon("fa5s.landmark"))
        self._build_toolbar()
        self._build_ui()
        self._build_statusbar()

    def _build_toolbar(self):
        tb = QToolBar()
        tb.setMovable(False)
        tb.setIconSize(QSize(20, 20))
        lbl = QLabel("  Музей — экспонаты и фонды")
        tb.addWidget(lbl)
        stretch = QWidget()
        stretch.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        tb.addWidget(stretch)
        btn_export = QPushButton(_qicon("fa5s.file-excel"), " Экспорт в Excel")
        btn_export.clicked.connect(self._export_excel)
        tb.addWidget(btn_export)
        self.addToolBar(tb)

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(10, 10, 10, 4)

        self.stats_layout = QHBoxLayout()
        layout.addLayout(self.stats_layout)

        self.tabs = QTabWidget()

        self.tab_halls = CrudTab(
            self, "halls",
            ["name", "floor", "area_sqm", "description"],
            ["Название", "Этаж", "Площадь (м²)", "Описание"],
            [("name", "Название зала", "line"), ("floor", "Этаж", "spin"),
             ("area_sqm", "Площадь (м²)", "double"),
             ("description", "Описание", "text")]
        )
        self.tabs.addTab(self.tab_halls, _qicon("fa5s.door-open"), "Залы")

        self.tab_exhibits = CrudTab(
            self, "exhibits",
            ["name", "category", "year_created", "hall_id", "condition", "description"],
            ["Название", "Категория", "Год создания", "Зал", "Состояние", "Описание"],
            [("name", "Название экспоната", "line"),
             ("category", "Категория", "line"),
             ("year_created", "Год/период создания", "line"),
             ("hall_id", "Зал размещения", "combo"),
             ("condition", "Состояние", "line"),
             ("description", "Описание", "text")],
            join_query=(
                "SELECT e.id, e.name, e.category, e.year_created, "
                "h.name, e.condition, e.description "
                "FROM exhibits e LEFT JOIN halls h ON e.hall_id = h.id"),
            combos_fn=lambda: {"hall_id": self._get_halls_combo()}
        )
        self.tabs.addTab(self.tab_exhibits, _qicon("fa5s.gem"), "Экспонаты")

        self.tab_exhibitions = CrudTab(
            self, "exhibitions",
            ["name", "start_date", "end_date", "hall_id", "status"],
            ["Название", "Дата начала", "Дата окончания", "Зал", "Статус"],
            [("name", "Название выставки", "line"),
             ("start_date", "Дата начала", "date"),
             ("end_date", "Дата окончания", "date"),
             ("hall_id", "Зал проведения", "combo"),
             ("status", "Статус", "line")],
            join_query=(
                "SELECT ex.id, ex.name, ex.start_date, ex.end_date, "
                "h.name, ex.status "
                "FROM exhibitions ex LEFT JOIN halls h ON ex.hall_id = h.id"),
            combos_fn=lambda: {"hall_id": self._get_halls_combo()}
        )
        self.tabs.addTab(self.tab_exhibitions, _qicon("fa5s.images"), "Выставки")

        self.tab_movements = CrudTab(
            self, "movements",
            ["exhibit_id", "from_location", "to_location", "date_move",
             "reason", "responsible"],
            ["Экспонат", "Откуда", "Куда", "Дата", "Причина", "Ответственный"],
            [("exhibit_id", "Экспонат", "combo"),
             ("from_location", "Место отправления", "line"),
             ("to_location", "Место назначения", "line"),
             ("date_move", "Дата перемещения", "date"),
             ("reason", "Причина перемещения", "line"),
             ("responsible", "Ответственный сотрудник", "line")],
            join_query=(
                "SELECT m.id, e.name, m.from_location, m.to_location, "
                "m.date_move, m.reason, m.responsible "
                "FROM movements m LEFT JOIN exhibits e ON m.exhibit_id = e.id"),
            combos_fn=lambda: {"exhibit_id": self._get_exhibits_combo()}
        )
        self.tabs.addTab(self.tab_movements, _qicon("fa5s.exchange-alt"), "Перемещения")

        date_filter_grp = QGroupBox("Фильтр по дате")
        df_layout = QHBoxLayout(date_filter_grp)
        df_layout.addWidget(QLabel("С:"))
        self.mv_date_from = QDateEdit()
        self.mv_date_from.setCalendarPopup(True)
        self.mv_date_from.setDisplayFormat("dd.MM.yyyy")
        self.mv_date_from.setDate(QDate(2025, 1, 1))
        df_layout.addWidget(self.mv_date_from)
        df_layout.addWidget(QLabel("По:"))
        self.mv_date_to = QDateEdit()
        self.mv_date_to.setCalendarPopup(True)
        self.mv_date_to.setDisplayFormat("dd.MM.yyyy")
        self.mv_date_to.setDate(QDate.currentDate())
        df_layout.addWidget(self.mv_date_to)
        btn_df = QPushButton("Фильтр")
        btn_df.clicked.connect(self._filter_movements_by_date)
        df_layout.addWidget(btn_df)
        btn_df_reset = QPushButton("Сбросить")
        btn_df_reset.clicked.connect(lambda: self.tab_movements.load_data())
        df_layout.addWidget(btn_df_reset)
        self.tab_movements.layout().insertWidget(1, date_filter_grp)

        self.tab_search = self._build_search_tab()
        self.tabs.addTab(self.tab_search, _qicon("fa5s.search"), "Поиск")

        layout.addWidget(self.tabs, 1)

        self._update_stats()
        self.tabs.currentChanged.connect(lambda: self._update_stats())

    def _build_statusbar(self):
        sb = QStatusBar()
        self.setStatusBar(sb)
        sb.showMessage(
            "  Учёт экспонатов и фондов в музее v1.0 | "
            f"Дата: {datetime.now().strftime('%d.%m.%Y')}")

    def _update_stats(self):
        while self.stats_layout.count():
            child = self.stats_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        conn = get_connection()
        halls = conn.execute("SELECT COUNT(*) FROM halls").fetchone()[0]
        exhibits = conn.execute("SELECT COUNT(*) FROM exhibits").fetchone()[0]
        exhibitions = conn.execute("SELECT COUNT(*) FROM exhibitions").fetchone()[0]
        active = conn.execute(
            "SELECT COUNT(*) FROM exhibitions WHERE status='Открыта'"
        ).fetchone()[0]
        planned = conn.execute(
            "SELECT COUNT(*) FROM exhibitions WHERE status='Планируется'"
        ).fetchone()[0]
        movements = conn.execute("SELECT COUNT(*) FROM movements").fetchone()[0]
        conn.close()

        self.stats_layout.addWidget(_make_stat_card("Залов", halls))
        self.stats_layout.addWidget(_make_stat_card("Экспонатов", exhibits))
        self.stats_layout.addWidget(_make_stat_card("Выставок", exhibitions))
        self.stats_layout.addWidget(
            _make_stat_card("Открыто", active, "#27AE60"))
        self.stats_layout.addWidget(
            _make_stat_card("Планируется", planned, "#2980B9"))
        self.stats_layout.addWidget(
            _make_stat_card("Перемещений", movements))

    @staticmethod
    def _get_halls_combo():
        conn = get_connection()
        rows = conn.execute("SELECT id, name FROM halls").fetchall()
        conn.close()
        return rows

    @staticmethod
    def _get_exhibits_combo():
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, name || ' (' || category || ')' FROM exhibits"
        ).fetchall()
        conn.close()
        return rows

    def _filter_movements_by_date(self):
        d_from = self.mv_date_from.date().toString("yyyy-MM-dd")
        d_to = self.mv_date_to.date().toString("yyyy-MM-dd")
        conn = get_connection()
        rows = conn.execute(
            "SELECT m.id, e.name, m.from_location, m.to_location, "
            "m.date_move, m.reason, m.responsible "
            "FROM movements m LEFT JOIN exhibits e ON m.exhibit_id = e.id "
            "WHERE m.date_move BETWEEN ? AND ?", (d_from, d_to)).fetchall()
        conn.close()
        t = self.tab_movements.table
        t.setSortingEnabled(False)
        t.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val is not None else "")
                if c == 0:
                    item.setData(Qt.UserRole, val)
                t.setItem(r, c, item)
        t.setSortingEnabled(True)

    def _build_search_tab(self):
        tab = QWidget()
        lay = QVBoxLayout(tab)
        lay.setContentsMargins(6, 6, 6, 6)

        filter_grp = QGroupBox("Параметры поиска")
        fl = QHBoxLayout(filter_grp)

        self.srch_input = QLineEdit()
        self.srch_input.setPlaceholderText(
            "Поиск по названию, категории, описанию...")
        self.srch_input.textChanged.connect(self._search_execute)
        fl.addWidget(QLabel("Текст:"))
        fl.addWidget(self.srch_input, 1)

        fl.addWidget(QLabel("Категория:"))
        self.srch_cat_cb = QComboBox()
        self.srch_cat_cb.currentIndexChanged.connect(self._search_execute)
        fl.addWidget(self.srch_cat_cb)

        fl.addWidget(QLabel("Зал:"))
        self.srch_hall_cb = QComboBox()
        self.srch_hall_cb.currentIndexChanged.connect(self._search_execute)
        fl.addWidget(self.srch_hall_cb)

        lay.addWidget(filter_grp)

        self.srch_count_lbl = QLabel("Найдено: 0 экспонатов")
        self.srch_count_lbl.setStyleSheet(
            "font-weight: bold; color: #6B3030; font-size: 12px;")
        lay.addWidget(self.srch_count_lbl)

        self.srch_table = QTableWidget()
        self.srch_table.setAlternatingRowColors(True)
        self.srch_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.srch_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.srch_table.setSortingEnabled(True)
        self.srch_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.srch_table.doubleClicked.connect(self._search_show_detail)
        lay.addWidget(self.srch_table, 1)

        conn = get_connection()
        cats = conn.execute(
            "SELECT DISTINCT category FROM exhibits "
            "WHERE category IS NOT NULL ORDER BY category").fetchall()
        halls = conn.execute("SELECT id, name FROM halls").fetchall()
        conn.close()

        self.srch_cat_cb.blockSignals(True)
        self.srch_cat_cb.addItem("Все категории", None)
        for (cat,) in cats:
            self.srch_cat_cb.addItem(cat, cat)
        self.srch_cat_cb.blockSignals(False)

        self.srch_hall_cb.blockSignals(True)
        self.srch_hall_cb.addItem("Все залы", None)
        for hid, hname in halls:
            self.srch_hall_cb.addItem(hname, hid)
        self.srch_hall_cb.blockSignals(False)

        self._search_execute()
        return tab

    def _search_execute(self):
        text = self.srch_input.text().strip()
        cat = self.srch_cat_cb.currentData()
        hall_id = self.srch_hall_cb.currentData()

        query = (
            "SELECT e.id, e.name, e.category, e.year_created, "
            "h.name, e.condition, e.description "
            "FROM exhibits e LEFT JOIN halls h ON e.hall_id = h.id WHERE 1=1")
        params = []
        if text:
            query += (" AND (e.name LIKE ? OR e.category LIKE ? "
                      "OR e.description LIKE ?)")
            like = f"%{text}%"
            params.extend([like, like, like])
        if cat:
            query += " AND e.category = ?"
            params.append(cat)
        if hall_id:
            query += " AND e.hall_id = ?"
            params.append(hall_id)

        conn = get_connection()
        rows = conn.execute(query, params).fetchall()
        conn.close()

        headers = ["ID", "Название", "Категория", "Год", "Зал",
                    "Состояние", "Описание"]
        self.srch_table.setSortingEnabled(False)
        self.srch_table.setColumnCount(len(headers))
        self.srch_table.setHorizontalHeaderLabels(headers)
        self.srch_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.srch_table.setColumnHidden(0, True)
        self.srch_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val is not None else "")
                if c == 0:
                    item.setData(Qt.UserRole, val)
                self.srch_table.setItem(r, c, item)
        self.srch_table.resizeColumnsToContents()
        self.srch_table.setSortingEnabled(True)
        self.srch_count_lbl.setText(f"Найдено: {len(rows)} экспонатов")

    def _search_show_detail(self):
        row = self.srch_table.currentRow()
        if row < 0:
            return
        eid = self.srch_table.item(row, 0).data(Qt.UserRole)
        conn = get_connection()
        ex = conn.execute(
            "SELECT e.name, e.category, e.year_created, h.name, "
            "e.condition, e.description "
            "FROM exhibits e LEFT JOIN halls h ON e.hall_id = h.id "
            "WHERE e.id=?", (eid,)).fetchone()
        last_move = conn.execute(
            "SELECT to_location, date_move FROM movements "
            "WHERE exhibit_id=? ORDER BY date_move DESC LIMIT 1",
            (eid,)).fetchone()
        conn.close()
        if not ex:
            return

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Экспонат: {ex[0]}")
        dlg.setMinimumWidth(450)
        dl = QVBoxLayout(dlg)
        info = (
            f"<b>Название:</b> {ex[0]}<br>"
            f"<b>Категория:</b> {ex[1]}<br>"
            f"<b>Год создания:</b> {ex[2]}<br>"
            f"<b>Зал:</b> {ex[3]}<br>"
            f"<b>Состояние:</b> {ex[4]}<br>"
            f"<b>Описание:</b> {ex[5]}<br>")
        if last_move:
            info += (f"<br><b>Текущее расположение:</b> {last_move[0]}"
                     f" (от {last_move[1]})")
        lbl = QLabel(info)
        lbl.setWordWrap(True)
        lbl.setTextFormat(Qt.RichText)
        dl.addWidget(lbl)
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(dlg.accept)
        dl.addWidget(btn_close)
        dlg.exec_()

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт в Excel", "museum_report.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Ошибка",
                                "Модуль openpyxl не установлен.\n"
                                "Установите: pip install openpyxl")
            return

        wb = Workbook()
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="8B3A3A", end_color="8B3A3A",
                               fill_type="solid")
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        conn = get_connection()
        sheets = [
            ("Залы",
             "SELECT id, name, floor, area_sqm, description FROM halls",
             ["ID", "Название", "Этаж", "Площадь", "Описание"]),
            ("Экспонаты",
             "SELECT e.id, e.name, e.category, e.year_created, h.name, "
             "e.condition, e.description "
             "FROM exhibits e LEFT JOIN halls h ON e.hall_id = h.id",
             ["ID", "Название", "Категория", "Год", "Зал",
              "Состояние", "Описание"]),
            ("Выставки",
             "SELECT ex.id, ex.name, ex.start_date, ex.end_date, "
             "h.name, ex.status "
             "FROM exhibitions ex LEFT JOIN halls h ON ex.hall_id = h.id",
             ["ID", "Название", "Начало", "Окончание", "Зал", "Статус"]),
            ("Перемещения",
             "SELECT m.id, e.name, m.from_location, m.to_location, "
             "m.date_move, m.reason, m.responsible "
             "FROM movements m LEFT JOIN exhibits e ON m.exhibit_id = e.id",
             ["ID", "Экспонат", "Откуда", "Куда", "Дата",
              "Причина", "Ответственный"]),
        ]

        for i, (name, query, headers) in enumerate(sheets):
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = name
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin
            for ri, row in enumerate(conn.execute(query).fetchall(), 2):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border = thin
            for col in ws.columns:
                mx = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 50)

        conn.close()
        wb.save(path)
        QMessageBox.information(self, "Экспорт",
                                f"Данные успешно сохранены в:\n{path}")


def main():
    init_db()
    app = QApplication(sys.argv)
    app.setStyleSheet(QSS)
    app.setFont(QFont("Segoe UI", 11))
    login = LoginDialog()
    if login.exec_() != QDialog.Accepted or not login.authenticated:
        sys.exit(0)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
