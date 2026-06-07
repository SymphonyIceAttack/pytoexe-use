import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pdfplumber

# ========== CONFIGURAÇÕES ==========
TABELAS_INSS = {
    2020: {'teto': 6101.06, 'faixas': [
        {'limite': 1045.00, 'aliquota': 7.5},
        {'limite': 2089.60, 'aliquota': 9.0},
        {'limite': 3134.40, 'aliquota': 12.0},
        {'limite': 6101.06, 'aliquota': 14.0}
    ]},
    2021: {'teto': 6433.57, 'faixas': [
        {'limite': 1100.00, 'aliquota': 7.5},
        {'limite': 2203.48, 'aliquota': 9.0},
        {'limite': 3305.22, 'aliquota': 12.0},
        {'limite': 6433.57, 'aliquota': 14.0}
    ]},
    2022: {'teto': 7087.22, 'faixas': [
        {'limite': 1212.00, 'aliquota': 7.5},
        {'limite': 2427.35, 'aliquota': 9.0},
        {'limite': 3641.04, 'aliquota': 12.0},
        {'limite': 7087.22, 'aliquota': 14.0}
    ]},
    2023: {'teto': 7507.49, 'faixas': [
        {'limite': 1302.00, 'aliquota': 7.5},
        {'limite': 2571.29, 'aliquota': 9.0},
        {'limite': 3856.94, 'aliquota': 12.0},
        {'limite': 7507.49, 'aliquota': 14.0}
    ]},
    2024: {'teto': 7786.02, 'faixas': [
        {'limite': 1412.00, 'aliquota': 7.5},
        {'limite': 2666.68, 'aliquota': 9.0},
        {'limite': 4000.03, 'aliquota': 12.0},
        {'limite': 7786.02, 'aliquota': 14.0}
    ]},
    2025: {'teto': 8157.41, 'faixas': [
        {'limite': 1518.00, 'aliquota': 7.5},
        {'limite': 2793.88, 'aliquota': 9.0},
        {'limite': 4190.83, 'aliquota': 12.0},
        {'limite': 8157.41, 'aliquota': 14.0}
    ]}
}
PRESCRICAO_INICIO = "05/2019"
ALIQUOTA_COOPERADO = 20
LIMITE_DIVERGENCIA = 100.00

# ========== LEITOR CNIS ==========
class LeitorCNIS:
    def __init__(self, caminho_pdf):
        self.caminho = caminho_pdf
        self.dados = self._processar()
    
    def _extrair_texto(self):
        texto = ""
        with pdfplumber.open(self.caminho) as pdf:
            for page in pdf.pages:
                texto += page.extract_text() + "\n"
        return texto
    
    def _processar(self):
        texto = self._extrair_texto()
        identificacao = {}
        
        # Extrair dados do cliente
        match_nit = re.search(r'NIT:\s*([\d\.\-]+)', texto)
        match_cpf = re.search(r'CPF:\s*([\d\.\-]+)', texto)
        match_nome = re.search(r'Nome:\s*([A-Z\s]+?)(?=\s*Data de nascimento|\s*NIT)', texto, re.IGNORECASE)
        
        if match_nit:
            identificacao['nit'] = match_nit.group(1)
        if match_cpf:
            identificacao['cpf'] = match_cpf.group(1)
        if match_nome:
            identificacao['nome'] = match_nome.group(1).strip()
        
        # Extrair remunerações
        remuneracoes = {}
        padrao_rem = r'(\d{2}/\d{4})\s+([\d\.,]+)'
        matches = re.findall(padrao_rem, texto)
        
        for competencia, valor in matches:
            try:
                valor_limpo = float(valor.replace('.', '').replace(',', '.'))
                if competencia not in remuneracoes:
                    remuneracoes[competencia] = {'empregado': 0, 'cooperado': 0}
                remuneracoes[competencia]['empregado'] += valor_limpo
            except:
                continue
        
        return {'identificacao': identificacao, 'remuneracoes': remuneracoes}
    
    def obter_remuneracoes_por_competencia(self):
        return self.dados['remuneracoes']

# ========== LEITOR DIRF ==========
class LeitorDIRF:
    def __init__(self, caminho_pdf, ano):
        self.ano = ano
        with pdfplumber.open(caminho_pdf) as pdf:
            texto = ""
            for page in pdf.pages:
                texto += page.extract_text() + "\n"
        
        self.previdencia = 0
        padrao_prev = r'Prev\.\s*Oficial.*?(\d+\.?\d*,\d{2})'
        match = re.search(padrao_prev, texto, re.IGNORECASE)
        if match:
            self.previdencia = float(match.group(1).replace('.', '').replace(',', '.'))
    
    def obter_total_previdencia_oficial(self):
        return self.previdencia

# ========== CÁLCULO ==========
class CalculoRestituicao:
    def __init__(self, cnis, dirfs):
        self.cnis = cnis
        self.dirfs = dirfs
    
    def _competencia_valida(self, competencia):
        mes, ano = int(competencia.split('/')[0]), int(competencia.split('/')[1])
        mes_ref, ano_ref = 5, 2019
        return ano > ano_ref or (ano == ano_ref and mes >= mes_ref)
    
    def _calcular_inss(self, salario, ano):
        tabela = TABELAS_INSS.get(ano)
        if not tabela:
            return 0
        inss = 0
        salario_restante = salario
        limite_anterior = 0
        for faixa in tabela['faixas']:
            if salario_restante <= 0:
                break
            limite = faixa['limite']
            aliquota = faixa['aliquota'] / 100
            if salario > limite:
                valor_faixa = limite - limite_anterior
                inss += valor_faixa * aliquota
                salario_restante -= valor_faixa
            else:
                inss += salario_restante * aliquota
                salario_restante = 0
            limite_anterior = limite
        return round(inss, 2)
    
    def calcular(self):
        resultados = []
        remuneracoes = self.cnis.obter_remuneracoes_por_competencia()
        
        for competencia, rems in remuneracoes.items():
            if not self._competencia_valida(competencia):
                continue
            
            ano = int(competencia.split('/')[1])
            rem_total = rems.get('empregado', 0)
            
            inss_correto = self._calcular_inss(min(rem_total, TABELAS_INSS[ano]['teto']), ano)
            
            dirf = self.dirfs.get(ano)
            inss_pago = dirf.obter_total_previdencia_oficial() / 12 if dirf else 0
            
            diferenca = max(0, inss_pago - inss_correto)
            
            if diferenca > 0:
                resultados.append({
                    'competencia': competencia,
                    'ano': ano,
                    'valor_restituir': diferenca
                })
        
        return resultados

# ========== EXPORTADOR ==========
class Exportador:
    def __init__(self, resultados, dados_cliente):
        self.resultados = resultados
        self.dados_cliente = dados_cliente
        self.total = sum(r['valor_restituir'] for r in resultados)
    
    def exportar_excel(self, caminho):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Restituicao_INSS"
        
        ws['A1'] = "RESTITUIÇÃO TETO INSS"
        ws['A2'] = f"Cliente: {self.dados_cliente.get('nome', '')}"
        ws['A3'] = f"CPF: {self.dados_cliente.get('cpf', '')}"
        ws['A5'] = f"Total a Restituir: R$ {self.total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        ws['A7'] = "Competência"
        ws['B7'] = "Valor a Restituir"
        
        row = 8
        for r in self.resultados:
            ws[f'A{row}'] = r['competencia']
            ws[f'B{row}'] = r['valor_restituir']
            ws[f'B{row}'].number_format = 'R$ #,##0.00'
            row += 1
        
        wb.save(caminho)
        return caminho
    
    def gerar_espelho_perdcomp(self):
        linhas = []
        linhas.append("=" * 60)
        linhas.append("ESPELHO PARA PER/DCOMP - RESTITUIÇÃO TETO INSS")
        linhas.append("=" * 60)
        linhas.append(f"Cliente: {self.dados_cliente.get('nome', '')}")
        linhas.append(f"CPF: {self.dados_cliente.get('cpf', '')}")
        linhas.append("")
        
        for r in self.resultados:
            valor_str = f"R$ {r['valor_restituir']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            linhas.append(f"Competência: {r['competencia']} - Valor: {valor_str}")
        
        linhas.append("")
        total_str = f"TOTAL: R$ {self.total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        linhas.append(total_str)
        linhas.append("")
        linhas.append("Instruções: Acessar e-CAC -> PER/DCOMP Web")
        linhas.append("Tipo de crédito: Contribuição Previdenciária Indevida ou a Maior")
        
        return "\n".join(linhas)

# ========== INTERFACE ==========
class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Restituição Teto INSS - Mentoria Murilo")
        self.root.geometry("800x700")
        self.cnis = None
        self.dirfs = {}
        self.resultados = None
        self.setup_ui()
    
    def setup_ui(self):
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill=tk.BOTH, expand=True)
        
        titulo = ttk.Label(main, text="🔄 RESTITUIÇÃO TETO INSS", font=('Arial', 16, 'bold'))
        titulo.pack(pady=10)
        
        # Seleção de pasta
        frame = ttk.LabelFrame(main, text="Selecionar Pasta", padding=10)
        frame.pack(fill=tk.X, pady=10)
        
        self.lbl_pasta = ttk.Label(frame, text="Nenhuma pasta selecionada")
        self.lbl_pasta.pack(side=tk.LEFT, padx=5)
        
        btn = ttk.Button(frame, text="Selecionar Pasta", command=self.selecionar_pasta)
        btn.pack(side=tk.RIGHT, padx=5)
        
        # Status
        self.lbl_status = ttk.Label(main, text="Aguardando...")
        self.lbl_status.pack(pady=10)
        
        # Botão processar
        self.btn_processar = ttk.Button(main, text="🚀 PROCESSAR", command=self.processar, state=tk.DISABLED)
        self.btn_processar.pack(pady=10)
        
        # Resultados
        frame_res = ttk.LabelFrame(main, text="Resultados", padding=10)
        frame_res.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.txt_resultado = tk.Text(frame_res, height=15)
        self.txt_resultado.pack(fill=tk.BOTH, expand=True)
        
        # Exportar
        frame_exp = ttk.Frame(main)
        frame_exp.pack(fill=tk.X, pady=10)
        
        self.btn_excel = ttk.Button(frame_exp, text="Exportar Excel", command=self.exportar_excel, state=tk.DISABLED)
        self.btn_excel.pack(side=tk.LEFT, padx=5)
        
        self.btn_perdcomp = ttk.Button(frame_exp, text="Gerar PER/DCOMP", command=self.exportar_perdcomp, state=tk.DISABLED)
        self.btn_perdcomp.pack(side=tk.LEFT, padx=5)
    
    def selecionar_pasta(self):
        pasta = filedialog.askdirectory()
        if pasta:
            self.pasta = Path(pasta)
            self.lbl_pasta.config(text=str(pasta))
            self.arquivos = list(self.pasta.glob("*.pdf"))
            self.btn_processar.config(state=tk.NORMAL)
            self.lbl_status.config(text=f"Encontrados {len(self.arquivos)} PDFs")
    
    def processar(self):
        self.btn_processar.config(state=tk.DISABLED)
        self.lbl_status.config(text="Processando...")
        self.root.update()
        
        try:
            # Encontrar CNIS
            cnis_file = None
            for f in self.arquivos:
                if 'cnis' in f.name.lower():
                    cnis_file = f
                    break
            
            if not cnis_file:
                raise Exception("CNIS não encontrado")
            
            self.cnis = LeitorCNIS(str(cnis_file))
            
            # Ler DIRFs
            for f in self.arquivos:
                if 'dirf' in f.name.lower():
                    anos = re.findall(r'\d{4}', f.name)
                    if anos:
                        ano = int(anos[0])
                        self.dirfs[ano] = LeitorDIRF(str(f), ano)
            
            # Calcular
            calc = CalculoRestituicao(self.cnis, self.dirfs)
            self.resultados = calc.calcular()
            total = sum(r['valor_restituir'] for r in self.resultados)
            
            # Exibir
            self.txt_resultado.delete(1.0, tk.END)
            self.txt_resultado.insert(tk.END, "=" * 50 + "\n")
            self.txt_resultado.insert(tk.END, f"Cliente: {self.cnis.dados['identificacao'].get('nome', 'N/A')}\n")
            self.txt_resultado.insert(tk.END, f"CPF: {self.cnis.dados['identificacao'].get('cpf', 'N/A')}\n")
            self.txt_resultado.insert(tk.END, "=" * 50 + "\n\n")
            self.txt_resultado.insert(tk.END, f"TOTAL A RESTITUIR: R$ {total:,.2f}\n\n".replace(',', 'X').replace('.', ',').replace('X', '.'))
            
            for r in self.resultados:
                valor_str = f"R$ {r['valor_restituir']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                self.txt_resultado.insert(tk.END, f"{r['competencia']}: {valor_str}\n")
            
            self.btn_excel.config(state=tk.NORMAL)
            self.btn_perdcomp.config(state=tk.NORMAL)
            self.lbl_status.config(text=f"Concluído! Total: R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
            
        except Exception as e:
            self.lbl_status.config(text=f"Erro: {str(e)}")
            messagebox.showerror("Erro", str(e))
        
        self.btn_processar.config(state=tk.NORMAL)
    
    def exportar_excel(self):
        dados = self.cnis.dados.get('identificacao', {})
        exp = Exportador(self.resultados, dados)
        caminho = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if caminho:
            exp.exportar_excel(caminho)
            messagebox.showinfo("Sucesso", f"Salvo em:\n{caminho}")
    
    def exportar_perdcomp(self):
        dados = self.cnis.dados.get('identificacao', {})
        exp = Exportador(self.resultados, dados)
        espelho = exp.gerar_espelho_perdcomp()
        caminho = filedialog.asksaveasfilename(defaultextension=".txt")
        if caminho:
            with open(caminho, 'w', encoding='utf-8') as f:
                f.write(espelho)
            messagebox.showinfo("Sucesso", f"Salvo em:\n{caminho}")
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = App()
    app.run()