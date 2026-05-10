#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Телефонная книга — приложение для управления контактами
С интеграцией Microsoft Excel
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import subprocess
import sys

# ============================================================
# НАСТРОЙКИ
# ============================================================
DEFAULT_FILENAME = "Телефонная_книга.xlsx"
SHEET_NAME = "Контакты"

# Заголовки столбцов
HEADERS = [
    "№ п/п",
    "ФИО",
    "Место работы",
    "Должность",
    "Телефон 1",
    "Телефон 2",
    "Телефон 3",
    "Телефон 4",
    "Телефон 5",
    "Связанный контакт (Помощник/Приёмная)",
    "Номер автомобиля"
]

# ============================================================
# КЛАСС ДАННЫХ
# ============================================================
class Contact:
    def __init__(self, num=0, fio="", work="", position="", 
                 phone1="", phone2="", phone3="", phone4="", phone5="",
                 linked="", car=""):
        self.num = num
        self.fio = fio
        self.work = work
        self.position = position
        self.phone1 = phone1
        self.phone2 = phone2
        self.phone3 = phone3
        self.phone4 = phone4
        self.phone5 = phone5
        self.linked = linked
        self.car = car

    def to_list(self):
        return [
            self.num, self.fio, self.work, self.position,
            self.phone1, self.phone2, self.phone3, self.phone4, self.phone5,
            self.linked, self.car
        ]

    @classmethod
    def from_list(cls, data):
        if len(data) < 11:
            data.extend([""] * (11 - len(data)))
        return cls(
            num=data[0] if data[0] else 0,
            fio=str(data[1]),
            work=str(data[2]),
            position=str(data[3]),
            phone1=str(data[4]),
            phone2=str(data[5]),
            phone3=str(data[6]),
            phone4=str(data[7]),
            phone5=str(data[8]),
            linked=str(data[9]),
            car=str(data[10])
        )

# ============================================================
# РАБОТА С EXCEL
# ============================================================
class ExcelManager:
    def __init__(self, filename=None):
        self.filename = filename or DEFAULT_FILENAME
        self.workbook = None
        self.sheet = None

    def create_new(self):
        """Создать новую книгу Excel с заголовками"""
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = SHEET_NAME

        # Стили заголовков
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Записываем заголовки
        for col_idx, header in enumerate(HEADERS, 1):
            cell = self.sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Настройка ширины столбцов
        column_widths = [8, 25, 25, 20, 15, 15, 15, 15, 15, 35, 15]
        for i, width in enumerate(column_widths, 1):
            self.sheet.column_dimensions[get_column_letter(i)].width = width

        # Фиксируем заголовок
        self.sheet.freeze_panes = 'A2'

        self.save()
        return True

    def load(self):
        """Загрузить существующую книгу"""
        if not os.path.exists(self.filename):
            return False
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.sheet = self.workbook[SHEET_NAME]
            return True
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {str(e)}")
            return False

    def save(self):
        """Сохранить книгу"""
        try:
            self.workbook.save(self.filename)
            return True
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {str(e)}")
            return False

    def get_all_contacts(self):
        """Получить все контакты из Excel"""
        contacts = []
        if not self.sheet:
            return contacts

        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[1]:  # Проверяем, что ФИО не пустое
                contacts.append(Contact.from_list(list(row)))
        return contacts

    def add_contact(self, contact):
        """Добавить новый контакт"""
        if not self.sheet:
            return False

        next_row = self.sheet.max_row + 1
        contact.num = next_row - 1

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, value in enumerate(contact.to_list(), 1):
            cell = self.sheet.cell(row=next_row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        return self.save()

    def update_contact(self, row_num, contact):
        """Обновить контакт по номеру строки"""
        if not self.sheet:
            return False

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        excel_row = row_num + 1  # +1 из-за заголовка
        for col_idx, value in enumerate(contact.to_list(), 1):
            cell = self.sheet.cell(row=excel_row, column=col_idx, value=value)
            cell.border = thin_border

        return self.save()

    def delete_contact(self, row_num):
        """Удалить контакт"""
        if not self.sheet:
            return False

        excel_row = row_num + 1
        self.sheet.delete_rows(excel_row)

        # Перенумеровываем
        for idx, row in enumerate(self.sheet.iter_rows(min_row=2, max_col=1), start=1):
            row[0].value = idx

        return self.save()

    def open_in_excel(self):
        """Открыть файл в Microsoft Excel"""
        if not os.path.exists(self.filename):
            return False
        try:
            if sys.platform == "win32":
                os.startfile(self.filename)
            elif sys.platform == "darwin":
                subprocess.call(["open", self.filename])
            else:
                subprocess.call(["xdg-open", self.filename])
            return True
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть Excel: {str(e)}")
            return False

# ============================================================
# ГЛАВНОЕ ОКНО ПРИЛОЖЕНИЯ
# ============================================================
class PhoneBookApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📞 Телефонная книга")
        self.root.geometry("1400x700")
        self.root.minsize(1200, 600)

        # Менеджер Excel
        self.excel = ExcelManager()
        self.contacts = []
        self.selected_index = None

        self.setup_ui()
        self.check_or_create_file()

    def setup_ui(self):
        """Настройка интерфейса"""
        # Верхняя панель инструментов
        toolbar = tk.Frame(self.root, bg="#f0f0f0", height=50)
        toolbar.pack(fill=tk.X, padx=5, pady=5)

        btn_new = tk.Button(toolbar, text="➕ Новый контакт", command=self.show_add_dialog,
                           bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), padx=10)
        btn_new.pack(side=tk.LEFT, padx=5)

        btn_edit = tk.Button(toolbar, text="✏️ Редактировать", command=self.show_edit_dialog,
                            bg="#2196F3", fg="white", font=("Arial", 10, "bold"), padx=10)
        btn_edit.pack(side=tk.LEFT, padx=5)

        btn_delete = tk.Button(toolbar, text="🗑️ Удалить", command=self.delete_contact,
                              bg="#f44336", fg="white", font=("Arial", 10, "bold"), padx=10)
        btn_delete.pack(side=tk.LEFT, padx=5)

        btn_refresh = tk.Button(toolbar, text="🔄 Обновить", command=self.load_contacts,
                               bg="#FF9800", fg="white", font=("Arial", 10, "bold"), padx=10)
        btn_refresh.pack(side=tk.LEFT, padx=5)

        btn_excel = tk.Button(toolbar, text="📊 Открыть в Excel", command=self.open_excel,
                             bg="#9C27B0", fg="white", font=("Arial", 10, "bold"), padx=10)
        btn_excel.pack(side=tk.LEFT, padx=5)

        btn_select = tk.Button(toolbar, text="📁 Выбрать файл", command=self.select_file,
                              bg="#607D8B", fg="white", font=("Arial", 10, "bold"), padx=10)
        btn_select.pack(side=tk.RIGHT, padx=5)

        # Поиск
        search_frame = tk.Frame(self.root, bg="#f0f0f0")
        search_frame.pack(fill=tk.X, padx=5, pady=2)

        tk.Label(search_frame, text="🔍 Поиск:", bg="#f0f0f0", font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *args: self.filter_contacts())
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, font=("Arial", 10), width=40)
        search_entry.pack(side=tk.LEFT, padx=5)

        # Таблица контактов
        table_frame = tk.Frame(self.root)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Скроллбары
        scroll_y = tk.Scrollbar(table_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(table_frame, orient=tk.HORIZONTAL)

        # Treeview
        columns = ("num", "fio", "work", "position", "phone1", "phone2", 
                   "phone3", "phone4", "phone5", "linked", "car")

        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings",
                                yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)

        # Настройка столбцов
        col_settings = {
            "num": ("№", 40),
            "fio": ("ФИО", 200),
            "work": ("Место работы", 180),
            "position": ("Должность", 150),
            "phone1": ("Тел. 1", 120),
            "phone2": ("Тел. 2", 120),
            "phone3": ("Тел. 3", 120),
            "phone4": ("Тел. 4", 120),
            "phone5": ("Тел. 5", 120),
            "linked": ("Связанный контакт", 250),
            "car": ("Автомобиль", 120)
        }

        for col, (header, width) in col_settings.items():
            self.tree.heading(col, text=header)
            self.tree.column(col, width=width, minwidth=50)

        # Расположение
        self.tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Статусная строка
        self.status = tk.Label(self.root, text="Готово", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status.pack(side=tk.BOTTOM, fill=tk.X)

        # Привязка событий
        self.tree.bind("<Double-1>", lambda e: self.show_edit_dialog())
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

    def check_or_create_file(self):
        """Проверить наличие файла или создать новый"""
        if not os.path.exists(self.excel.filename):
            if messagebox.askyesno("Файл не найден", 
                f"Файл {self.excel.filename} не найден.\nСоздать новый?"):
                self.excel.create_new()
                self.status.config(text=f"Создан новый файл: {self.excel.filename}")
            else:
                self.select_file()
        else:
            self.load_contacts()

    def select_file(self):
        """Выбрать существующий файл Excel"""
        filename = filedialog.askopenfilename(
            title="Выберите файл телефонной книги",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
        )
        if filename:
            self.excel.filename = filename
            self.load_contacts()

    def load_contacts(self):
        """Загрузить контакты из Excel"""
        if self.excel.load():
            self.contacts = self.excel.get_all_contacts()
            self.refresh_table()
            self.status.config(text=f"Загружено контактов: {len(self.contacts)} | Файл: {self.excel.filename}")

    def refresh_table(self):
        """Обновить отображение таблицы"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        for contact in self.contacts:
            self.tree.insert("", tk.END, values=contact.to_list())

    def filter_contacts(self):
        """Фильтрация по поиску"""
        query = self.search_var.get().lower()

        for item in self.tree.get_children():
            self.tree.delete(item)

        for contact in self.contacts:
            values = contact.to_list()
            if any(query in str(v).lower() for v in values):
                self.tree.insert("", tk.END, values=values)

    def on_select(self, event):
        """Обработка выбора строки"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            self.selected_index = item["values"][0]  # Номер контакта

    def show_add_dialog(self):
        """Диалог добавления контакта"""
        dialog = ContactDialog(self.root, "Добавить контакт", self.save_new_contact)
        dialog.grab_set()

    def show_edit_dialog(self):
        """Диалог редактирования контакта"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Внимание", "Выберите контакт для редактирования")
            return

        item = self.tree.item(selection[0])
        values = item["values"]
        contact = Contact.from_list(values)

        dialog = ContactDialog(self.root, "Редактировать контакт", 
                              lambda c: self.save_edited_contact(contact.num, c),
                              contact)
        dialog.grab_set()

    def save_new_contact(self, contact):
        """Сохранить новый контакт"""
        if self.excel.add_contact(contact):
            self.load_contacts()
            messagebox.showinfo("Успех", "Контакт добавлен!")
            return True
        return False

    def save_edited_contact(self, num, contact):
        """Сохранить отредактированный контакт"""
        if self.excel.update_contact(num, contact):
            self.load_contacts()
            messagebox.showinfo("Успех", "Контакт обновлён!")
            return True
        return False

    def delete_contact(self):
        """Удалить выбранный контакт"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Внимание", "Выберите контакт для удаления")
            return

        item = self.tree.item(selection[0])
        num = item["values"][0]
        fio = item["values"][1]

        if messagebox.askyesno("Подтверждение", f"Удалить контакт "{fio}"?"):
            if self.excel.delete_contact(num):
                self.load_contacts()
                messagebox.showinfo("Успех", "Контакт удалён!")

    def open_excel(self):
        """Открыть файл в Excel для редактирования"""
        if self.excel.open_in_excel():
            self.status.config(text="Файл открыт в Microsoft Excel. После редактирования нажмите 'Обновить'")

# ============================================================
# ДИАЛОГ РЕДАКТИРОВАНИЯ КОНТАКТА
# ============================================================
class ContactDialog(tk.Toplevel):
    def __init__(self, parent, title, save_callback, contact=None):
        super().__init__(parent)
        self.title(title)
        self.geometry("500x550")
        self.resizable(False, False)
        self.save_callback = save_callback
        self.contact = contact or Contact()

        self.setup_ui()
        self.center_window()

    def setup_ui(self):
        """Настройка элементов диалога"""
        frame = tk.Frame(self, padx=20, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        # Поля ввода
        fields = [
            ("ФИО:", "fio", True),
            ("Место работы:", "work", False),
            ("Должность:", "position", False),
            ("Телефон 1:", "phone1", False),
            ("Телефон 2:", "phone2", False),
            ("Телефон 3:", "phone3", False),
            ("Телефон 4:", "phone4", False),
            ("Телефон 5:", "phone5", False),
            ("Связанный контакт (Помощник/Приёмная):", "linked", False),
            ("Номер автомобиля:", "car", False)
        ]

        self.entries = {}

        for label_text, attr, required in fields:
            row = tk.Frame(frame)
            row.pack(fill=tk.X, pady=3)

            label = tk.Label(row, text=label_text, width=35, anchor="w", font=("Arial", 10))
            label.pack(side=tk.LEFT)

            entry = tk.Entry(row, font=("Arial", 10), width=30)
            entry.pack(side=tk.RIGHT, fill=tk.X, expand=True)

            # Заполняем значения если редактирование
            value = getattr(self.contact, attr, "")
            entry.insert(0, str(value) if value else "")

            self.entries[attr] = entry

            if required:
                label.config(fg="red")

        # Кнопки
        btn_frame = tk.Frame(frame)
        btn_frame.pack(fill=tk.X, pady=20)

        tk.Button(btn_frame, text="💾 Сохранить", command=self.save,
                 bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), width=15).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="❌ Отмена", command=self.destroy,
                 bg="#f44336", fg="white", font=("Arial", 10, "bold"), width=15).pack(side=tk.RIGHT, padx=5)

    def center_window(self):
        """Центрирование окна"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

    def save(self):
        """Сохранение контакта"""
        fio = self.entries["fio"].get().strip()
        if not fio:
            messagebox.showwarning("Ошибка", "Поле 'ФИО' обязательно для заполнения!")
            return

        contact = Contact(
            fio=fio,
            work=self.entries["work"].get().strip(),
            position=self.entries["position"].get().strip(),
            phone1=self.entries["phone1"].get().strip(),
            phone2=self.entries["phone2"].get().strip(),
            phone3=self.entries["phone3"].get().strip(),
            phone4=self.entries["phone4"].get().strip(),
            phone5=self.entries["phone5"].get().strip(),
            linked=self.entries["linked"].get().strip(),
            car=self.entries["car"].get().strip()
        )

        if self.save_callback(contact):
            self.destroy()

# ============================================================
# ЗАПУСК ПРИЛОЖЕНИЯ
# ============================================================
if __name__ == "__main__":
    # Проверка зависимостей
    try:
        import openpyxl
    except ImportError:
        print("Установка необходимой библиотеки openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl

    root = tk.Tk()
    app = PhoneBookApp(root)
    root.mainloop()
