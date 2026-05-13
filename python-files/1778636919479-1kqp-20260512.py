# -*- coding: utf-8 -*-
"""
司法文书自动填充+自动打印程序
✅ 无pywin32依赖 | 纯Windows系统命令自动打印
✅ 仅扣划(4)需输入额外信息
✅ 自动打印至：实达 BP-3000II
✅ 不修改默认打印机 | 全自动无手动操作
"""
import os
import win32print
import win32api
from docx import Document
from openpyxl import load_workbook

# ====================== 核心配置 ======================
# 指定自动打印的打印机（必须和系统名称完全一致）
PRINTER_NAME = "实达BP-3000II"
# 模板文件名称
TEMPLATE_FILES = {
    "agree": "同意协办打印模板.docx",
    "freeze": "冻结特殊业务单打印模板.docx",
    "continue_freeze": "续冻特殊业务单打印模板.docx",
    "unfreeze": "解冻特殊业务单打印模板.docx",
    "deduct": "扣划特种转账单打印模板.xlsx"
}
# 填充后生成的文件
OUTPUT_FILES = {
    "agree": "同意协办_已填充.docx",
    "freeze": "冻结业务单_已填充.docx",
    "continue_freeze": "续冻业务单_已填充.docx",
    "unfreeze": "解冻业务单_已填充.docx",
    "deduct": "扣划单_已填充.xlsx"
}


# ====================================================
def check_printer_exist():
    """检测目标打印机是否真的存在"""
    printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL, None, 1)
    printer_list = [p[2] for p in printers]
    if PRINTER_NAME not in printer_list:
        print(f"❌ 错误：电脑里找不到打印机【{PRINTER_NAME}】")
        print(f"📋 本机已安装打印机列表：")
        for p in printer_list:
            print(f"   - {p}")
        return False
    print(f"✅ 检测到目标打印机：{PRINTER_NAME}")
    return True

def get_user_input():
    """分场景获取输入信息"""
    print("=" * 50)
    print("司法业务自动打印程序")
    print("1=冻结  2=续冻  3=解冻  4=扣划")
    print("=" * 50)

    # 选择业务类型
    while True:
        choice = input("请输入数字(1/2/3/4)：").strip()
        if choice in ["1", "2", "3", "4"]:
            break
        print("❌ 输入错误，请重新输入！")

    # 通用必填信息
    print("\n===== 请输入通用信息 =====")
    info = {
        "文书号": input("文书号：").strip(),
        "金额": input("金额：").strip(),
        "被执行人姓名": input("被执行人姓名：").strip(),
        "被执行账号": input("被执行账号：").strip(),
        "日期": input("日期（如：2025年12月25日）：").strip()
    }

    # 扣划专属信息（仅4需要）
    if choice == "4":
        print("\n===== 请输入扣划专属信息 =====")
        info["扣划去向"] = input("扣划去向：").strip()
        info["贷方账号"] = input("贷方账号：").strip()
        info["贷方开户行"] = input("贷方开户行：").strip()
        info["冻结序号"] = input("冻结序号：").strip()
        info["实际扣划金额"] = input("实际扣划金额：").strip()

    return choice, info


def fill_word_template(template_path, output_path, replace_dict):
    """填充Word文档"""
    try:
        doc = Document(template_path)
        for paragraph in doc.paragraphs:
            for key, value in replace_dict.items():
                placeholder = f"{{{{{key}}}}}"
                if placeholder in paragraph.text:
                    paragraph.text = paragraph.text.replace(placeholder, value)
        doc.save(output_path)
        print(f"✅ Word填充完成：{output_path}")
    except Exception as e:
        print(f"❌ Word填充失败：{str(e)}")


def fill_excel_template(template_path, output_path, info):
    """填充Excel表格"""
    try:
        wb = load_workbook(template_path)
        ws = wb.active
        ws["B11"] = info["被执行人姓名"]
        ws["D11"] = info["扣划去向"]
        ws["B12"] = info["被执行账号"]
        ws["D12"] = info["贷方账号"]
        ws["D13"] = info["贷方开户行"]
        ws["B14"] = info["金额"]

        # 填充备注文本
        text = "司法扣划，原冻结序号：冻结序号，应扣划：金额，实际扣划：实际扣划金额，余额不足"
        ws["B17"] = text.replace("冻结序号", info["冻结序号"]) \
            .replace("金额", info["金额"]) \
            .replace("实际扣划金额", info["实际扣划金额"])
        wb.save(output_path)
        print(f"✅ Excel填充完成：{output_path}")
    except Exception as e:
        print(f"❌ Excel填充失败：{str(e)}")


def print_file(file_path):
    """直接指定打印机打印，全程不动系统默认打印机"""
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在：{file_path}")
        return
    try:
        # 强制指定实达打印机打印，不修改系统默认
        win32api.ShellExecute(
            0,
            "printto",
            file_path,
            f'"{PRINTER_NAME}"',
            ".",
            0
        )
        print(f"✅ 打印任务已发送至：{PRINTER_NAME} -> {file_path}")
    except Exception as e:
        print(f"⚠️  打印调用提示：{str(e)}")
        print("如未自动打印，请手动打开文件选择打印机打印")

def main():
    # 1. 获取用户输入

    # 先检测打印机
    if not check_printer_exist():
        input("核对打印机名称后，按回车退出重启程序...")
        return

    choice, info = get_user_input()

    # 2. 填充同意协办模板（通用）
    fill_word_template(TEMPLATE_FILES["agree"], OUTPUT_FILES["agree"], {"日期": info["日期"]})

    # 3. 填充对应业务模板
    print_files = [OUTPUT_FILES["agree"]]
    print("\n开始填充业务模板...")

    if choice == "1":
        fill_word_template(TEMPLATE_FILES["freeze"], OUTPUT_FILES["freeze"], info)
        print_files.append(OUTPUT_FILES["freeze"])
    elif choice == "2":
        fill_word_template(TEMPLATE_FILES["continue_freeze"], OUTPUT_FILES["continue_freeze"], info)
        print_files.append(OUTPUT_FILES["continue_freeze"])
    elif choice == "3":
        fill_word_template(TEMPLATE_FILES["unfreeze"], OUTPUT_FILES["unfreeze"], info)
        print_files.append(OUTPUT_FILES["unfreeze"])
    elif choice == "4":
        fill_excel_template(TEMPLATE_FILES["deduct"], OUTPUT_FILES["deduct"], info)
        print_files.append(OUTPUT_FILES["deduct"])

    # 4. 【全自动打印】无任何手动操作
    print("\n开始自动打印...")
    for file in print_files:
        print_file(file)

    print("\n🎉 全部完成！文件已自动发送至打印机")
    input("按回车键退出程序...")


if __name__ == "__main__":
    main()