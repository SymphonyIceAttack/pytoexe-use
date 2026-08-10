import sys, csv, datetime, os
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox

# This import is intentionally explicit. The EXE build script uses PyInstaller
# collect-all/hidden-import so openpyxl is bundled inside the EXE.
try:
    import openpyxl
    from openpyxl import load_workbook
except Exception as exc:
    openpyxl = None
    OPENPYXL_ERROR = str(exc)

DATA_FILE = "条码数据.xlsx"
BG = "#F4F6F8"; CARD = "#FFFFFF"; TEXT = "#1F2937"; SUB = "#6B7280"
BLUE = "#2563EB"; GREEN = "#16A34A"; RED = "#DC2626"; YELLOW = "#D97706"; BORDER = "#E5E7EB"

class App:
    def __init__(self, root):
        self.root = root
        root.title("条码核对工具")
        root.geometry("1000x700")
        root.minsize(850, 600)
        root.configure(bg=BG)

        self.groups = {}
        self.current = None
        self.expected = set()
        self.scanned = set()
        self.input_var = tk.StringVar()
        self.group_var = tk.StringVar(value="等待扫描 A 组条码")
        self.count_var = tk.StringVar(value="0 / 0")
        self.status_var = tk.StringVar(value="准备就绪")
        self.detail_var = tk.StringVar(value="请扫描 A 组条码开始核对")
        self.clock_var = tk.StringVar()

        self.done = 0
        self.errors = 0
        self.scans = 0

        self.build_ui()
        self.load_excel()
        self.update_clock()
        root.after(150, self.focus_scan)

    def app_dir(self):
        return Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent

    def build_ui(self):
        header = tk.Frame(self.root, bg="#111827", height=76)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="▣  条码核对", bg="#111827", fg="white",
                 font=("Microsoft YaHei UI", 22, "bold")).pack(side="left", padx=28)
        tk.Label(header, text="LOCAL / OFFLINE", bg="#111827", fg="#9CA3AF",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=8)
        tk.Label(header, textvariable=self.clock_var, bg="#111827", fg="#D1D5DB",
                 font=("Segoe UI", 10)).pack(side="right", padx=28)

        main = tk.Frame(self.root, bg=BG)
        main.pack(fill="both", expand=True, padx=24, pady=20)

        left = tk.Frame(main, bg=BG)
        left.pack(side="left", fill="both", expand=True, padx=(0, 10))

        c = tk.Frame(left, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        c.pack(fill="x", pady=(0, 12))
        tk.Label(c, text="当前数据组", bg=CARD, fg=SUB,
                 font=("Microsoft YaHei UI", 11)).pack(anchor="w", padx=22, pady=(18, 2))
        tk.Label(c, textvariable=self.group_var, bg=CARD, fg=TEXT,
                 font=("Microsoft YaHei UI", 24, "bold")).pack(anchor="w", padx=22, pady=(0, 18))

        pc = tk.Frame(left, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        pc.pack(fill="x", pady=(0, 12))
        row = tk.Frame(pc, bg=CARD); row.pack(fill="x", padx=22, pady=(18, 8))
        tk.Label(row, text="核对进度", bg=CARD, fg=SUB,
                 font=("Microsoft YaHei UI", 11)).pack(side="left")
        tk.Label(row, textvariable=self.count_var, bg=CARD, fg=BLUE,
                 font=("Segoe UI", 16, "bold")).pack(side="right")
        self.progress = ttk.Progressbar(pc, maximum=100)
        self.progress.pack(fill="x", padx=22, pady=(0, 20), ipady=4)

        self.status_card = tk.Frame(left, bg="#EFF6FF", highlightbackground="#DBEAFE", highlightthickness=1)
        self.status_card.pack(fill="x", pady=(0, 12))
        self.status_title = tk.Label(self.status_card, textvariable=self.status_var,
                                     bg="#EFF6FF", fg=BLUE,
                                     font=("Microsoft YaHei UI", 25, "bold"))
        self.status_title.pack(anchor="w", padx=22, pady=(20, 4))
        self.status_detail = tk.Label(self.status_card, textvariable=self.detail_var,
                                      bg="#EFF6FF", fg="#4B5563",
                                      font=("Microsoft YaHei UI", 12),
                                      wraplength=590, justify="left")
        self.status_detail.pack(anchor="w", padx=22, pady=(0, 20))

        scan = tk.Frame(left, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        scan.pack(fill="x", pady=(0, 12))
        tk.Label(scan, text="扫描输入", bg=CARD, fg=SUB,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", padx=22, pady=(15, 5))
        self.entry = tk.Entry(scan, textvariable=self.input_var, font=("Consolas", 18),
                              bg="#F9FAFB", fg=TEXT, relief="flat",
                              highlightthickness=1, highlightbackground=BORDER,
                              highlightcolor=BLUE)
        self.entry.pack(fill="x", padx=22, pady=(0, 18), ipady=9)
        self.entry.bind("<Return>", self.process_scan)

        log = tk.Frame(left, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        log.pack(fill="both", expand=True)
        tk.Label(log, text="最近扫描", bg=CARD, fg=TEXT,
                 font=("Microsoft YaHei UI", 14, "bold")).pack(anchor="w", padx=20, pady=(14, 8))
        self.recent = tk.Listbox(log, bg=CARD, fg=TEXT, relief="flat",
                                 font=("Consolas", 10), highlightthickness=0)
        self.recent.pack(fill="both", expand=True, padx=15, pady=(0, 12))

        right = tk.Frame(main, bg=BG, width=300)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        stats = tk.Frame(right, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        stats.pack(fill="x", pady=(0, 12))
        tk.Label(stats, text="今日状态", bg=CARD, fg=TEXT,
                 font=("Microsoft YaHei UI", 15, "bold")).pack(anchor="w", padx=20, pady=(18, 15))
        self.done_label = self.stat(stats, "0", GREEN, "完成组")
        self.error_label = self.stat(stats, "0", RED, "错误次数")
        self.scan_label = self.stat(stats, "0", BLUE, "扫描次数")

        current = tk.Frame(right, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        current.pack(fill="both", expand=True)
        tk.Label(current, text="当前组条码", bg=CARD, fg=TEXT,
                 font=("Microsoft YaHei UI", 15, "bold")).pack(anchor="w", padx=20, pady=(18, 10))
        self.barcode_list = tk.Listbox(current, bg=CARD, fg=TEXT, relief="flat",
                                       font=("Consolas", 12), highlightthickness=0)
        self.barcode_list.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        for title, command, color in [
            ("重新加载 Excel", self.reload_excel, BLUE),
            ("清空当前组", self.reset_group, "#6B7280"),
            ("打开核对记录", self.open_logs, "#374151"),
        ]:
            tk.Button(current, text=title, command=command, bg=color, fg="white",
                      activebackground=color, activeforeground="white",
                      relief="flat", bd=0, font=("Microsoft YaHei UI", 10, "bold"),
                      pady=9).pack(fill="x", padx=15, pady=3)
        tk.Frame(current, bg=CARD, height=10).pack()

    def stat(self, parent, value, color, label):
        x = tk.Label(parent, text=value, bg=CARD, fg=color, font=("Segoe UI", 30, "bold"))
        x.pack(anchor="w", padx=20)
        tk.Label(parent, text=label, bg=CARD, fg=SUB,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", padx=20, pady=(0, 12))
        return x

    def load_excel(self):
        self.groups.clear()
        if openpyxl is None:
            self.set_status("Excel组件缺失", "error",
                            "打包时未正确包含 openpyxl。请使用本文件夹中的“一键生成EXE.bat”打包。")
            return

        path = self.app_dir() / DATA_FILE
        if not path.exists():
            self.set_status("找不到数据文件", "error",
                            f"请把 {DATA_FILE} 放在 EXE 同一文件夹。")
            return

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb["条码数据"] if "条码数据" in wb.sheetnames else wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                raise ValueError("Excel没有有效数据。")

            names = [str(v).strip() if v is not None else "" for v in header]
            a_idx = names.index("A组条码") if "A组条码" in names else 0
            c_idx = names.index("子条码") if "子条码" in names else 1

            for row in rows:
                if len(row) <= max(a_idx, c_idx):
                    continue
                a = "" if row[a_idx] is None else str(row[a_idx]).strip()
                child = "" if row[c_idx] is None else str(row[c_idx]).strip()
                if a and child:
                    self.groups.setdefault(a, set()).add(child)

            self.set_status("准备就绪", "normal",
                            f"已读取 {len(self.groups)} 个数据组，请扫描 A 组条码。")
        except Exception as exc:
            self.set_status("Excel读取失败", "error", str(exc))

    def process_scan(self, event=None):
        code = self.input_var.get().strip()
        self.input_var.set("")
        self.focus_scan()
        if not code:
            return

        self.scans += 1
        self.scan_label.config(text=str(self.scans))

        if self.current is None:
            if code in self.groups:
                self.current = code
                self.expected = set(self.groups[code])
                self.scanned.clear()
                self.group_var.set(code)
                self.update_progress()
                self.refresh_list()
                self.record(code, code, "A组开始")
                self.set_status("开始核对", "normal",
                                f"当前组共有 {len(self.expected)} 个子条码，请继续扫描。")
                self.beep("normal")
            else:
                self.errors += 1
                self.error_label.config(text=str(self.errors))
                self.record("", code, "A组不存在")
                self.set_status("✕ 不是有效 A 组", "error",
                                f"扫描到：{code}，Excel中没有这个 A 组。")
                self.beep("error")
            return

        if code in self.scanned:
            self.errors += 1
            self.error_label.config(text=str(self.errors))
            self.record(self.current, code, "重复")
            self.set_status("⚠ 重复扫描", "warning", f"{code} 已经核对过。")
            self.beep("warning")
            return

        if code not in self.expected:
            self.errors += 1
            self.error_label.config(text=str(self.errors))
            self.record(self.current, code, "错误")
            self.set_status("✕ 扫描错误", "error",
                            f"{code} 不属于当前组 {self.current}，当前进度不会受到影响。")
            self.beep("error")
            return

        self.scanned.add(code)
        self.record(self.current, code, "正确")
        self.update_progress()
        self.refresh_list()
        self.set_status("✓ 扫描正确", "success", f"{code} 属于 {self.current}。")
        self.beep("success")

        if self.scanned == self.expected:
            self.done += 1
            self.done_label.config(text=str(self.done))
            self.record(self.current, "", "核对完成")
            self.set_status("✓ 核对完成", "success",
                            f"{self.current} 的全部 {len(self.expected)} 个子条码已经核对完成。")
            self.beep("complete")
            self.root.after(1400, self.next_group)

    def update_progress(self):
        n, total = len(self.scanned), len(self.expected)
        self.count_var.set(f"{n} / {total}")
        self.progress["value"] = (n / total * 100) if total else 0

    def refresh_list(self):
        self.barcode_list.delete(0, tk.END)
        for code in sorted(self.expected):
            self.barcode_list.insert(tk.END, ("✓  " if code in self.scanned else "○  ") + code)

    def next_group(self):
        self.current = None
        self.expected.clear()
        self.scanned.clear()
        self.group_var.set("等待扫描 A 组条码")
        self.count_var.set("0 / 0")
        self.progress["value"] = 0
        self.refresh_list()
        self.set_status("准备下一组", "normal", "请扫描下一个 A 组条码。")
        self.focus_scan()

    def reset_group(self):
        if self.current is None:
            self.set_status("没有正在核对的组", "warning", "请先扫描 A 组条码。")
            return
        self.record(self.current, "", "手动清空当前组")
        self.scanned.clear()
        self.update_progress()
        self.refresh_list()
        self.set_status("已清空当前组", "warning", f"{self.current} 可以重新开始扫描。")
        self.focus_scan()

    def reload_excel(self):
        self.current = None
        self.expected.clear()
        self.scanned.clear()
        self.group_var.set("等待扫描 A 组条码")
        self.count_var.set("0 / 0")
        self.progress["value"] = 0
        self.refresh_list()
        self.load_excel()
        self.focus_scan()

    def record(self, group, code, result):
        try:
            folder = self.app_dir() / "核对记录"
            folder.mkdir(exist_ok=True)
            path = folder / (datetime.datetime.now().strftime("%Y-%m-%d") + ".csv")
            new = not path.exists()
            with path.open("a", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                if new:
                    writer.writerow(["时间", "A组条码", "扫描条码", "结果"])
                writer.writerow([
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    group, code, result
                ])
            self.recent.insert(0, f'{datetime.datetime.now().strftime("%H:%M:%S")}  {result:<8}  {code or group}')
            if self.recent.size() > 30:
                self.recent.delete(30, tk.END)
        except Exception:
            pass

    def open_logs(self):
        folder = self.app_dir() / "核对记录"
        folder.mkdir(exist_ok=True)
        try:
            os.startfile(str(folder))
        except Exception:
            messagebox.showinfo("核对记录", str(folder))

    def set_status(self, title, kind, detail):
        self.status_var.set(title)
        self.detail_var.set(detail)
        if kind == "success":
            bg, border, fg = "#ECFDF5", "#BBF7D0", GREEN
        elif kind == "error":
            bg, border, fg = "#FEF2F2", "#FECACA", RED
        elif kind == "warning":
            bg, border, fg = "#FFFBEB", "#FDE68A", YELLOW
        else:
            bg, border, fg = "#EFF6FF", "#DBEAFE", BLUE
        self.status_card.config(bg=bg, highlightbackground=border)
        self.status_title.config(bg=bg, fg=fg)
        self.status_detail.config(bg=bg)

    def beep(self, kind):
        try:
            import winsound
            sequences = {
                "normal": [(800, 80)],
                "success": [(1200, 90)],
                "error": [(500, 180), (350, 220)],
                "warning": [(750, 100), (750, 100)],
                "complete": [(900, 100), (1200, 120), (1500, 160)]
            }
            for freq, duration in sequences[kind]:
                winsound.Beep(freq, duration)
        except Exception:
            self.root.bell()

    def focus_scan(self):
        self.entry.focus_set()
        self.entry.selection_range(0, tk.END)

    def update_clock(self):
        self.clock_var.set(datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))
        self.root.after(1000, self.update_clock)

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
