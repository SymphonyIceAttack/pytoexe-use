import os
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Внутренний неизменяемый шаблон «по умолчанию»
DEFAULT_CONFIG = {
    "currency": "руб.",
    "price_brackets": [
        { "min": 1, "max": 100, "price_per_tag": 500 },
        { "min": 101, "max": 500, "price_per_tag": 400 },
        { "min": 501, "max": 2000, "price_per_tag": 300 },
        { "min": 2001, "max": 999999, "price_per_tag": 200 }
    ],
    "coefficients": {
        "equipment_type": {
            "Стандартное (Modbus/BACnet)": 1.0,
            "Нестандартное/Проприетарное": 1.3
        },
        "location": {
            "Без выезда (Удаленно)": 1.0,
            "С выездом на объект / Командировка": 1.25
        },
        "architecture": {
            "Одиночный сервер": 1.0,
            "Резервированный сервер (Redundancy)": 1.4
        }
    }
}

SAVE_PATH = os.path.join(os.path.expanduser("~"), ".scada_calc_settings.json")

class SCADACalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Калькулятор SCADA (Автономный)")
        self.root.geometry("620x600")
        
        self.last_calculation = None
        self.is_admin_authenticated = False  # Флаг авторизации администратора
        
        self.load_config()
        
        # Панель вкладок
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)
        
        self.tab_calc = ttk.Frame(self.notebook)
        self.tab_admin = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab_calc, text="📊 Расчет стоимости")
        self.notebook.add(self.tab_admin, text="⚙️ Редактор цен и коэфф.")
        
        # Отслеживание переключения вкладок
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        self.create_calc_tab()
        self.create_admin_tab()

    def load_config(self):
        if os.path.exists(SAVE_PATH):
            try:
                with open(SAVE_PATH, "r", encoding="utf-8") as f:
                    self.config = json.load(f)
            except Exception:
                self.config = DEFAULT_CONFIG.copy()
        else:
            self.config = DEFAULT_CONFIG.copy()

    def save_config(self):
        try:
            with open(SAVE_PATH, "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror("Ошибка сохранения", f"Не удалось записать данные: {e}")
            return False

    def get_base_price(self, tags):
        for bracket in self.config["price_brackets"]:
            if bracket["min"] <= tags <= bracket["max"]:
                return bracket["price_per_tag"]
        return 0

    def on_tab_changed(self, event):
        """Срабатывает при смене вкладок. Сбрасывает авторизацию при выходе из админки."""
        selected_tab = self.notebook.index(self.notebook.select())
        if selected_tab == 1:  # Перешли на вкладку админки
            self.render_admin_tab_content()
        else:
            # При уходе с вкладки админки сбрасываем сессию для безопасности
            self.is_admin_authenticated = False

    # ==========================================
    # ВКЛАДКА РАСЧЕТА
    # ==========================================
    def create_calc_tab(self):
        padding = {"padx": 15, "pady": 5}
        
        tk.Label(self.tab_calc, text="Количество тегов в системе:", font=("Arial", 10, "bold")).pack(**padding)
        self.entry_tags = tk.Entry(self.tab_calc, font=("Arial", 11), width=15)
        self.entry_tags.pack(**padding)

        tk.Label(self.tab_calc, text="Тип оборудования:").pack(**padding)
        self.combo_eq = ttk.Combobox(self.tab_calc, state="readonly", width=45)
        self.combo_eq.pack(**padding)

        tk.Label(self.tab_calc, text="Условия реализации:").pack(**padding)
        self.combo_loc = ttk.Combobox(self.tab_calc, state="readonly", width=45)
        self.combo_loc.pack(**padding)

        tk.Label(self.tab_calc, text="Архитектура системы:").pack(**padding)
        self.combo_arch = ttk.Combobox(self.tab_calc, state="readonly", width=45)
        self.combo_arch.pack(**padding)

        self.refresh_comboboxes()

        btn_calc = tk.Button(self.tab_calc, text="Рассчитать стоимость", command=self.calculate, bg="#2196F3", fg="white", font=("Arial", 10, "bold"), height=2)
        btn_calc.pack(pady=15)

        self.lbl_result = tk.Label(self.tab_calc, text="", font=("Arial", 11), justify="left", bd=1, relief="solid", padx=15, pady=10, bg="#f9f9f9")
        self.btn_export = tk.Button(self.tab_calc, text="🍏 Экспорт КП в Excel", command=self.export_to_excel, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), state="disabled")

    def refresh_comboboxes(self):
        self.combo_eq["values"] = list(self.config["coefficients"]["equipment_type"].keys())
        self.combo_eq.current(0) if self.config["coefficients"]["equipment_type"] else self.combo_eq.set("")
        
        self.combo_loc["values"] = list(self.config["coefficients"]["location"].keys())
        self.combo_loc.current(0) if self.config["coefficients"]["location"] else self.combo_loc.set("")
        
        self.combo_arch["values"] = list(self.config["coefficients"]["architecture"].keys())
        self.combo_arch.current(0) if self.config["coefficients"]["architecture"] else self.combo_arch.set("")

    def calculate(self):
        try:
            tags = int(self.entry_tags.get())
            if tags <= 0: raise ValueError
        except ValueError:
            messagebox.showwarning("Внимание", "Введите корректное число тегов (> 0)")
            return

        base_price = self.get_base_price(tags)
        base_cost = tags * base_price

        eq_name, loc_name, arch_name = self.combo_eq.get(), self.combo_loc.get(), self.combo_arch.get()
        if not eq_name or not loc_name or not arch_name:
            messagebox.showwarning("Внимание", "Заполните все коэффициенты в редакторе!")
            return

        k_eq = self.config["coefficients"]["equipment_type"][eq_name]
        k_loc = self.config["coefficients"]["location"][loc_name]
        k_arch = self.config["coefficients"]["architecture"][arch_name]
        total_k = k_eq * k_loc * k_arch

        total_cost = base_cost * total_k
        currency = self.config.get("currency", "руб.")

        result_text = (
            f"Базовая цена за 1 тег:  {base_price} {currency}\n"
            f"Стоимость базового объема:  {base_cost:,.2f} {currency}\n"
            f"Итоговый множитель (коэфф.):  {total_k:.2f}\n"
            f"--------------------------------------------------\n"
            f"ИТОГОВАЯ СТОИМОСТЬ СИСТЕМЫ:  {total_cost:,.2f} {currency}"
        )
        
        self.lbl_result.config(text=result_text)
        self.lbl_result.pack(pady=10, fill="x", padx=30)
        
        self.last_calculation = {
            "tags": tags, "base_price": base_price, "base_cost": base_cost,
            "eq_name": eq_name, "k_eq": k_eq, "loc_name": loc_name, "k_loc": k_loc,
            "arch_name": arch_name, "k_arch": k_arch, "total_cost": total_cost, "currency": currency
        }
        self.btn_export.config(state="normal")
        self.btn_export.pack(pady=5)

    def export_to_excel(self):
        if not self.last_calculation: return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Сохранить КП")
        if not file_path: return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Коммерческое предложение"
            ws.views.sheetView.showGridLines = True
            
            font_title = Font(name="Arial", size=16, bold=True, color="1F497D")
            font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            font_bold = Font(name="Arial", size=11, bold=True)
            font_regular = Font(name="Arial", size=11)
            fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            fill_accent = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
            thin_border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

            ws["A1"] = "КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ"
            ws["A1"].font = font_title
            ws["A2"] = "На разработку и внедрение системы диспетчеризации"
            ws["A2"].font = Font(name="Arial", size=11, italic=True)
            
            headers = ["Наименование параметра / условия", "Значение / Выбор", "Коэффициент / Цена"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num, value=header)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center")
            
            calc = self.last_calculation
            data = [
                ["Объем системы (количество тегов)", f"{calc['tags']} шт.", f"{calc['base_price']} {calc['currency']}/тег"],
                ["Базовая расчетная стоимость", "Расчет по прайсу", f"{calc['base_cost']:,.2f} {calc['currency']}"],
                ["Тип используемого оборудования", calc['eq_name'], f"x {calc['k_eq']}"],
                ["Условия проведения работ", calc['loc_name'], f"x {calc['k_loc']}"],
                ["Архитектура развертывания", calc['arch_name'], f"x {calc['k_arch']}"],
            ]
            
            for row_idx, row_data in enumerate(data, 5):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
