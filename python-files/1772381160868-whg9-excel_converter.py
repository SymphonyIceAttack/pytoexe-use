import os
import shutil
import tempfile
from datetime import datetime

from PySide6.QtCore import Qt
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QLabel, QPushButton, QFrame, QMessageBox
)

from openpyxl import load_workbook


def get_downloads_folder() -> str:
    """
    Windows에서 '다운로드' 폴더를 보통 USERPROFILE\\Downloads로 가정.
    (대부분의 환경에서 동작)
    """
    home = os.path.expanduser("~")
    downloads = os.path.join(home, "Downloads")
    return downloads


def safe_output_name(input_path: str) -> str:
    base = os.path.splitext(os.path.basename(input_path))[0]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_converted_{ts}.xlsx"


def convert_excel(input_path: str, output_path: str) -> None:
    """
    변환 규칙:
      - A1, B1 비우기
      - 2행부터:
          A행 = M열 값
          B행 = K열 값
      - 나머지 셀은 그대로 둠(원본 유지)
      - 기본적으로 활성 시트(active sheet)만 처리
    """
    wb = load_workbook(input_path)
    ws = wb.active

    # 1행 비우기
    ws["A1"].value = None
    ws["B1"].value = None

    max_row = ws.max_row

    # 2행부터 복사
    for r in range(2, max_row + 1):
        a_val = ws[f"M{r}"].value
        b_val = ws[f"K{r}"].value
        ws[f"A{r}"].value = a_val
        ws[f"B{r}"].value = b_val

    wb.save(output_path)


class DropBox(QFrame):
    """
    드래그&드롭 파일 받는 박스
    """
    def __init__(self, title: str, hint: str, accept_drop: bool):
        super().__init__()
        self.setFrameShape(QFrame.StyledPanel)
        self.setAcceptDrops(accept_drop)
        self.setMinimumSize(380, 260)

        self.title_label = QLabel(title)
        self.title_label.setFont(QFont("Segoe UI", 11, QFont.Bold))

        self.body_label = QLabel(hint)
        self.body_label.setWordWrap(True)
        self.body_label.setAlignment(Qt.AlignTop | Qt.AlignLeft)

        lay = QVBoxLayout(self)
        lay.addWidget(self.title_label)
        lay.addWidget(self.body_label)
        lay.addStretch(1)

        self.file_path = None  # 선택/드롭된 파일 경로

    def set_text(self, text: str):
        self.body_label.setText(text)

    def set_file(self, path: str):
        self.file_path = path
        self.set_text(f"파일:\n{path}")

    def clear(self):
        self.file_path = None

    def dragEnterEvent(self, event):
        if not self.acceptDrops():
            event.ignore()
            return
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                p = urls[0].toLocalFile()
                if p.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
                    event.acceptProposedAction()
                    return
        event.ignore()

    def dropEvent(self, event):
        if not self.acceptDrops():
            event.ignore()
            return
        urls = event.mimeData().urls()
        if not urls:
            event.ignore()
            return
        path = urls[0].toLocalFile()
        if not path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            QMessageBox.warning(self, "형식 오류", "엑셀 파일(.xlsx/.xlsm 등)만 가능합니다.")
            return
        self.set_file(path)
        event.acceptProposedAction()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("엑셀 열 변환기 (M→A, K→B)")
        self.resize(900, 420)

        self.temp_dir = tempfile.mkdtemp(prefix="excel_convert_")
        self.converted_path = None

        # 왼쪽: 변환된 파일 표시
        self.left_box = DropBox(
            title="(왼쪽) 변환된 파일",
            hint="아직 변환된 파일이 없습니다.\n오른쪽에 원본 엑셀을 드래그해서 넣고 '변환'을 누르세요.",
            accept_drop=False
        )

        # 오른쪽: 원본 업로드(드래그&드롭)
        self.right_box = DropBox(
            title="(오른쪽) 원본 엑셀 업로드",
            hint="여기에 원본 엑셀 파일을 드래그&드롭 하세요.\n지원: .xlsx, .xlsm",
            accept_drop=True
        )

        # 버튼들
        self.btn_convert = QPushButton("변환")
        self.btn_download = QPushButton("다운로드")
        self.btn_reset = QPushButton("초기화")

        self.btn_convert.clicked.connect(self.on_convert)
        self.btn_download.clicked.connect(self.on_download)
        self.btn_reset.clicked.connect(self.on_reset)

        self.btn_download.setEnabled(False)

        btn_row = QHBoxLayout()
        btn_row.addWidget(self.btn_convert)
        btn_row.addWidget(self.btn_download)
        btn_row.addWidget(self.btn_reset)
        btn_row.addStretch(1)

        boxes = QHBoxLayout()
        boxes.addWidget(self.left_box, 1)
        boxes.addWidget(self.right_box, 1)

        root = QVBoxLayout()
        root.addLayout(boxes)
        root.addLayout(btn_row)

        w = QWidget()
        w.setLayout(root)
        self.setCentralWidget(w)

    def show_error(self, title: str, msg: str):
        QMessageBox.critical(self, title, msg)

    def show_info(self, title: str, msg: str):
        QMessageBox.information(self, title, msg)

    def on_convert(self):
        src = self.right_box.file_path
        if not src or not os.path.isfile(src):
            self.show_error("원본 없음", "오른쪽 박스에 원본 엑셀 파일을 먼저 드래그&드롭 해주세요.")
            return

        try:
            out_name = safe_output_name(src)
            out_path = os.path.join(self.temp_dir, out_name)
            convert_excel(src, out_path)
            self.converted_path = out_path

            self.left_box.set_text(f"변환 완료!\n\n변환 파일:\n{out_path}\n\n이제 '다운로드'를 누르면 다운로드 폴더로 저장됩니다.")
            self.btn_download.setEnabled(True)
        except Exception as e:
            self.show_error("변환 실패", f"변환 중 오류가 발생했습니다.\n\n{e}")

    def on_download(self):
        if not self.converted_path or not os.path.isfile(self.converted_path):
            self.show_error("파일 없음", "다운로드할 변환 파일이 없습니다. 먼저 변환을 실행하세요.")
            return

        downloads = get_downloads_folder()
        if not os.path.isdir(downloads):
            self.show_error("다운로드 폴더 오류", f"다운로드 폴더를 찾을 수 없습니다:\n{downloads}")
            return

        try:
            dest = os.path.join(downloads, os.path.basename(self.converted_path))
            # 덮어쓰기 방지: 동일 이름이면 (1),(2) 붙이기
            if os.path.exists(dest):
                base, ext = os.path.splitext(dest)
                i = 1
                while True:
                    cand = f"{base} ({i}){ext}"
                    if not os.path.exists(cand):
                        dest = cand
                        break
                    i += 1

            shutil.copy2(self.converted_path, dest)
            self.show_info("다운로드 완료", f"다운로드 폴더에 저장했습니다:\n{dest}")
        except Exception as e:
            self.show_error("다운로드 실패", f"저장 중 오류가 발생했습니다.\n\n{e}")

    def on_reset(self):
        # UI 초기화
        self.right_box.clear()
        self.right_box.set_text("여기에 원본 엑셀 파일을 드래그&드롭 하세요.\n지원: .xlsx, .xlsm")

        self.left_box.clear()
        self.left_box.set_text("아직 변환된 파일이 없습니다.\n오른쪽에 원본 엑셀을 드래그해서 넣고 '변환'을 누르세요.")

        self.converted_path = None
        self.btn_download.setEnabled(False)


def main():
    app = QApplication([])
    win = MainWindow()
    win.show()
    app.exec()


if __name__ == "__main__":
    main()