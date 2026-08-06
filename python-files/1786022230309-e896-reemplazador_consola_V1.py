#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
REEMPLAZADOR MASIVO DE TEXTO EN DOCUMENTOS
==========================================
Busca y reemplaza palabras o frases dentro de todos los archivos
Word (.docx), Excel (.xlsx/.xlsm) y PowerPoint (.pptx) de un directorio.

Trabaja directamente sobre el XML interno del paquete OOXML, por lo que:
  - conserva formato, estilos, imagenes, graficos y macros intactos
  - alcanza cabeceras, pies de pagina, notas, comentarios, cuadros de texto,
    propiedades del documento, titulos de graficos y encabezados de impresion
  - resuelve el problema de los "runs" fragmentados de Word (una frase visible
    puede estar partida en 5 trozos de XML distintos)
  - mantiene un unico REGISTRO_CAMBIOS.xlsx dentro de _BACKUP_REEMPLAZO,
    con una fila por archivo y motivo del cambio, acumulando el historico
  - puede exportar a PDF los documentos modificados (requiere Office)

Uso:
    python reemplazador_consola.py                 -> modo interactivo
    python reemplazador_consola.py --help          -> modo linea de comandos

Requisitos: pip install lxml openpyxl pywin32
"""

import argparse
import csv
import datetime as _dt
import os
import re
import shutil
import sys
import zipfile

try:
    from lxml import etree
except ImportError:
    sys.exit("Falta la libreria lxml.  Instalala con:  pip install lxml")

try:
    import colorama
    colorama.init()
    C_OK, C_WARN, C_ERR, C_INFO, C_END = ("\033[92m", "\033[93m",
                                          "\033[91m", "\033[96m", "\033[0m")
except Exception:                                    # pragma: no cover
    C_OK = C_WARN = C_ERR = C_INFO = C_END = ""

VERSION = "1.0"

# ---------------------------------------------------------------------------
# Espacios de nombres OOXML
# ---------------------------------------------------------------------------
W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
A = "http://schemas.openxmlformats.org/drawingml/2006/main"
S = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XML = "http://www.w3.org/XML/1998/namespace"

EXTENSIONES = {".docx", ".docm", ".dotx", ".xlsx", ".xlsm", ".xltx", ".pptx", ".pptm"}
LEGADO = {".doc", ".xls", ".ppt"}          # formatos binarios antiguos
IGNORAR_CARPETAS = {"__pycache__", ".git", "_BACKUP_REEMPLAZO"}

# archivos generados por la propia herramienta: nunca deben procesarse,
# o el registro de cambios acabaria reescribiendose a si mismo
RX_PROPIOS = re.compile(
    r"^(REGISTRO_CAMBIOS|SIMULACION_|informe_reemplazo_|informe_dwg_)",
    re.IGNORECASE)


# ---------------------------------------------------------------------------
# Reglas de sustitucion
# ---------------------------------------------------------------------------
class Regla:
    """Una sustitucion: buscar -> reemplazar, ya compilada como expresion regular."""

    def __init__(self, buscar, reemplazar, sensible=False,
                 palabra_completa=False, es_regex=False):
        self.buscar = buscar
        self.reemplazar = reemplazar
        patron = buscar if es_regex else re.escape(buscar)
        if palabra_completa and not es_regex:
            patron = r"(?<!\w)" + patron + r"(?!\w)"
        flags = 0 if sensible else (re.IGNORECASE | re.UNICODE)
        self.rx = re.compile(patron, flags)
        self.sustituto = reemplazar if es_regex else reemplazar.replace("\\", "\\\\")

    def aplicar(self, texto):
        """Devuelve (texto_nuevo, n_sustituciones)."""
        return self.rx.subn(self.sustituto, texto)

    def __repr__(self):
        return f"{self.buscar!r} -> {self.reemplazar!r}"


def aplicar_reglas(texto, reglas):
    """Devuelve (texto_nuevo, [n_por_cada_regla])."""
    conteos = [0] * len(reglas)
    for i, r in enumerate(reglas):
        texto, n = r.aplicar(texto)
        conteos[i] = n
    return texto, conteos


def sumar(destino, origen):
    for i, v in enumerate(origen):
        destino[i] += v
    return destino


# ---------------------------------------------------------------------------
# Nucleo: sustitucion sobre un grupo de nodos de texto contiguos
# ---------------------------------------------------------------------------
def _reemplazar_en_nodos(nodos, reglas):
    """
    'nodos' es la lista ordenada de elementos de texto que juntos forman un
    parrafo visible (p.ej. todos los <w:t> de un <w:p>).

    Se concatena el texto completo, se aplican las reglas y luego se reparte
    el resultado: cada trozo intacto vuelve a su nodo original y cada
    sustitucion se deposita entera en el nodo donde empezaba la coincidencia.
    Asi se conserva el formato del resto del parrafo.

    Devuelve el numero de sustituciones realizadas.
    """
    conteos = [0] * len(reglas)
    if not nodos:
        return conteos

    trozos = [n.text or "" for n in nodos]
    completo = "".join(trozos)
    if not completo.strip():
        return conteos

    # --- aplicar las reglas arrastrando un mapa de posiciones ------------
    #     desplazamiento_mapa[i] = posicion en el texto ORIGINAL del
    #     caracter que ahora ocupa la posicion i
    trabajo = completo
    desplazamiento_mapa = list(range(len(completo)))
    for indice_regla, regla in enumerate(reglas):
        salida = []
        nuevo_mapa = []
        pos = 0
        for m in regla.rx.finditer(trabajo):
            conteos[indice_regla] += 1
            salida.append(trabajo[pos:m.start()])
            nuevo_mapa.extend(desplazamiento_mapa[pos:m.start()])
            sust = m.expand(regla.sustituto)
            salida.append(sust)
            # todo el texto insertado se atribuye al primer caracter original
            ancla = desplazamiento_mapa[m.start()] if m.start() < len(desplazamiento_mapa) \
                else (desplazamiento_mapa[-1] if desplazamiento_mapa else 0)
            nuevo_mapa.extend([ancla] * len(sust))
            pos = m.end()
        salida.append(trabajo[pos:])
        nuevo_mapa.extend(desplazamiento_mapa[pos:])
        trabajo = "".join(salida)
        desplazamiento_mapa = nuevo_mapa

    if sum(conteos) == 0 or trabajo == completo:
        return conteos

    # --- mapa caracter -> indice de nodo propietario ---------------------
    propietario = []
    for i, t in enumerate(trozos):
        propietario.extend([i] * len(t))

    # --- repartir el texto final entre los nodos -------------------------
    piezas = [[] for _ in nodos]
    for ch, pos_orig in zip(trabajo, desplazamiento_mapa):
        idx = propietario[pos_orig] if pos_orig < len(propietario) else len(nodos) - 1
        piezas[idx].append(ch)

    for nodo, pz in zip(nodos, piezas):
        texto = "".join(pz)
        nodo.text = texto
        if texto != texto.strip():
            nodo.set(f"{{{XML}}}space", "preserve")

    return conteos


def _reemplazar_directo(nodo, reglas):
    """Sustitucion simple sobre el texto de un unico nodo."""
    if nodo.text is None:
        return [0] * len(reglas)
    nuevo, conteos = aplicar_reglas(nodo.text, reglas)
    if sum(conteos):
        nodo.text = nuevo
        if nuevo != nuevo.strip():
            nodo.set(f"{{{XML}}}space", "preserve")
    return conteos


# ---------------------------------------------------------------------------
# Procesado de cada parte XML segun su tipo
# ---------------------------------------------------------------------------
def ubicacion_legible(nombre_parte):
    """Traduce el nombre interno de la parte XML a algo comprensible."""
    n = nombre_parte.lower()
    tabla = [
        ("word/document", "Cuerpo del documento"),
        ("word/header", "Encabezado"),
        ("word/footer", "Pie de pagina"),
        ("word/footnotes", "Notas al pie"),
        ("word/endnotes", "Notas al final"),
        ("word/comments", "Comentarios"),
        ("word/glossary", "Bloques de creacion"),
        ("xl/sharedstrings", "Celdas"),
        ("xl/worksheets", "Hoja de calculo"),
        ("xl/drawings", "Cuadros de texto / formas"),
        ("xl/charts", "Graficos"),
        ("ppt/slides", "Diapositivas"),
        ("ppt/slidelayouts", "Disenos de diapositiva"),
        ("ppt/slidemasters", "Patron de diapositivas"),
        ("ppt/notesslides", "Notas del orador"),
        ("docprops", "Propiedades del documento"),
    ]
    for clave, etiqueta in tabla:
        if n.startswith(clave):
            return etiqueta
    return nombre_parte


def procesar_xml(nombre_parte, datos, reglas):
    """Devuelve (bytes_nuevos_o_None, [n_por_regla], muestras)."""
    conteos = [0] * len(reglas)
    try:
        parser = etree.XMLParser(remove_blank_text=False, huge_tree=True)
        arbol = etree.fromstring(datos, parser)
    except etree.XMLSyntaxError:
        return None, conteos, [[] for _ in reglas]

    muestras = [[] for _ in reglas]
    parte = nombre_parte.lower()

    def anotar(antes, n):
        """Guarda un ejemplo del texto para cada regla que haya actuado."""
        for i, veces in enumerate(n):
            if veces and len(muestras[i]) < 2:
                muestras[i].append(antes.strip()[:140])

    def bloque(contenedor, etiqueta_texto):
        """Aplica las reglas a un parrafo completo y registra una muestra."""
        nodos = list(contenedor.iter(etiqueta_texto))
        antes = "".join(t.text or "" for t in nodos)
        n = _reemplazar_en_nodos(nodos, reglas)
        anotar(antes, n)
        sumar(conteos, n)

    # ---- Word: parrafos ------------------------------------------------
    if parte.startswith("word/"):
        for p in arbol.iter(f"{{{W}}}p"):
            bloque(p, f"{{{W}}}t")
        for p in arbol.iter(f"{{{A}}}p"):          # cuadros de texto DrawingML
            bloque(p, f"{{{A}}}t")

    # ---- Excel: cadenas compartidas, en linea, cabeceras y graficos -----
    elif parte.startswith("xl/"):
        for si in arbol.iter(f"{{{S}}}si", f"{{{S}}}is"):
            bloque(si, f"{{{S}}}t")
        for etiq in ("oddHeader", "evenHeader", "firstHeader",
                     "oddFooter", "evenFooter", "firstFooter"):
            for nodo in arbol.iter(f"{{{S}}}{etiq}"):
                antes = nodo.text or ""
                n = _reemplazar_directo(nodo, reglas)
                anotar(antes, n)
                sumar(conteos, n)
        for p in arbol.iter(f"{{{A}}}p"):
            bloque(p, f"{{{A}}}t")

    # ---- PowerPoint ----------------------------------------------------
    elif parte.startswith("ppt/"):
        for p in arbol.iter(f"{{{A}}}p"):
            bloque(p, f"{{{A}}}t")

    # ---- Propiedades del documento (titulo, asunto, palabras clave) -----
    elif parte.startswith("docprops/"):
        for nodo in arbol.iter():
            if isinstance(nodo.tag, str) and nodo.text and nodo.text.strip():
                etiqueta = etree.QName(nodo).localname.lower()
                if etiqueta in ("title", "subject", "keywords", "description",
                                "category", "company", "manager"):
                    antes = nodo.text
                    n = _reemplazar_directo(nodo, reglas)
                    anotar(antes, n)
                    sumar(conteos, n)

    if sum(conteos) == 0:
        return None, conteos, muestras
    return etree.tostring(arbol, xml_declaration=True,
                          encoding="UTF-8", standalone=True), conteos, muestras


PARTES_RELEVANTES = re.compile(
    r"^(word/(document|header\d*|footer\d*|footnotes|endnotes|comments)\d*\.xml"
    r"|word/glossary/document\.xml"
    r"|xl/(sharedStrings\.xml|worksheets/.+\.xml|drawings/.+\.xml|charts/.+\.xml)"
    r"|ppt/(slides|slideLayouts|slideMasters|notesSlides)/.+\.xml"
    r"|docProps/(core|app|custom)\.xml)$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Procesado de un archivo completo
# ---------------------------------------------------------------------------
def procesar_archivo(ruta, reglas, simulacion=True, carpeta_backup=None,
                     raiz=None):
    """Devuelve dict con el resultado del archivo."""
    res = {"ruta": ruta, "sustituciones": 0, "partes": [], "ubicaciones": [],
           "muestras": [[] for _ in reglas], "error": None, "backup": "",
           "por_regla": [0] * len(reglas)}
    try:
        with zipfile.ZipFile(ruta, "r") as z:
            items = z.infolist()
            contenidos = {i.filename: z.read(i.filename) for i in items}
    except Exception as e:
        res["error"] = f"no se pudo abrir: {e}"
        return res

    nuevos = {}
    for nombre, datos in contenidos.items():
        if not PARTES_RELEVANTES.match(nombre):
            continue
        salida, conteos, muestras = procesar_xml(nombre, datos, reglas)
        if sum(conteos):
            nuevos[nombre] = salida
            sumar(res["por_regla"], conteos)
            res["sustituciones"] += sum(conteos)
            res["partes"].append(f"{nombre} ({sum(conteos)})")
            etiqueta = ubicacion_legible(nombre)
            if etiqueta not in res["ubicaciones"]:
                res["ubicaciones"].append(etiqueta)
            for i, ejemplos in enumerate(muestras):
                for e in ejemplos:
                    if e not in res["muestras"][i] and len(res["muestras"][i]) < 2:
                        res["muestras"][i].append(e)

    if res["sustituciones"] == 0 or simulacion:
        return res

    # --- escritura real: copia de seguridad + reescritura del zip --------
    try:
        if carpeta_backup:
            # se replica la estructura de carpetas original dentro del backup
            rel = os.path.relpath(ruta, raiz) if raiz else os.path.basename(ruta)
            destino = os.path.join(carpeta_backup, rel)
            os.makedirs(os.path.dirname(destino), exist_ok=True)
            if os.path.exists(destino):
                # ya habia copia de una ejecucion anterior: se conserva la
                # primera (el original de verdad) y esta se marca con la fecha
                base, ext = os.path.splitext(destino)
                fecha = _dt.datetime.now().strftime("%Y%m%d")
                destino = f"{base}_{fecha}{ext}"
                k = 2
                while os.path.exists(destino):
                    destino = f"{base}_{fecha}_{k}{ext}"
                    k += 1
            shutil.copy2(ruta, destino)
            res["backup"] = destino

        temporal = ruta + ".tmp_reemplazo"
        with zipfile.ZipFile(ruta, "r") as origen, \
             zipfile.ZipFile(temporal, "w", zipfile.ZIP_DEFLATED) as salida_zip:
            for item in origen.infolist():
                datos = nuevos.get(item.filename, contenidos[item.filename])
                info = zipfile.ZipInfo(item.filename, date_time=item.date_time)
                info.compress_type = item.compress_type
                info.external_attr = item.external_attr
                info.internal_attr = item.internal_attr
                info.create_system = item.create_system
                salida_zip.writestr(info, datos)
        os.replace(temporal, ruta)
    except Exception as e:
        res["error"] = f"error al escribir: {e}"
        if os.path.exists(ruta + ".tmp_reemplazo"):
            os.remove(ruta + ".tmp_reemplazo")
    return res


# ---------------------------------------------------------------------------
# Recorrido del directorio
# ---------------------------------------------------------------------------
def recorrer(directorio, recursivo=True):
    encontrados, legado = [], []
    for raiz, dirs, ficheros in os.walk(directorio):
        dirs[:] = [d for d in dirs
                   if d not in IGNORAR_CARPETAS and not d.startswith("~")]
        for f in ficheros:
            if f.startswith("~$"):          # archivos temporales de Office
                continue
            if RX_PROPIOS.match(f):         # informes de esta herramienta
                continue
            ext = os.path.splitext(f)[1].lower()
            if ext in EXTENSIONES:
                encontrados.append(os.path.join(raiz, f))
            elif ext in LEGADO:
                legado.append(os.path.join(raiz, f))
        if not recursivo:
            break
    return sorted(encontrados), sorted(legado)


# ---------------------------------------------------------------------------
# Informe
# ---------------------------------------------------------------------------
def _registro_csv(filas, ruta_csv):
    """Alternativa en CSV si no esta disponible openpyxl."""
    nuevo = not os.path.isfile(ruta_csv)
    with open(ruta_csv, "a", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh, delimiter=";")
        if nuevo:
            w.writerow(["Fecha y hora", "Usuario", "Archivo", "Carpeta", "Tipo",
                        "Motivo: texto localizado", "Sustituido por",
                        "N. de cambios", "Donde", "Ejemplo", "Copia de seguridad"])
        for f in filas:
            w.writerow(f["datos"])
    return ruta_csv


TIPOS = {".docx": "Word", ".docm": "Word", ".dotx": "Word",
         ".xlsx": "Excel", ".xlsm": "Excel", ".xltx": "Excel",
         ".pptx": "PowerPoint", ".pptm": "PowerPoint"}


def escribir_registro(resultados, carpeta, reglas, destino_carpeta):
    """Envoltura tolerante a fallos: un problema al escribir el registro
    nunca debe impedir que la herramienta siga funcionando."""
    try:
        return _escribir_registro(resultados, carpeta, reglas, destino_carpeta)
    except Exception as e:
        return f"[no se pudo generar el registro: {e}]"


def _escribir_registro(resultados, carpeta, reglas, destino_carpeta):
    """
    Mantiene un unico REGISTRO_CAMBIOS.xlsx dentro de la carpeta de copia
    de seguridad. Si ya existe de ejecuciones anteriores, se le anaden las
    filas nuevas: el historial completo queda en un solo archivo.
    """
    momento = _dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    usuario = os.environ.get("USERNAME") or os.environ.get("USER") or ""
    if not usuario:
        try:
            usuario = os.getlogin()
        except Exception:
            usuario = "-"

    con_cambios = [r for r in resultados if r["sustituciones"]]

    # --- filas nuevas: una por archivo y motivo --------------------------
    filas = []
    for r in con_cambios:
        carpeta_rel = os.path.dirname(os.path.relpath(r["ruta"], carpeta)) or "."
        tipo = TIPOS.get(os.path.splitext(r["ruta"])[1].lower(), "")
        for idx, (regla, n) in enumerate(zip(reglas, r["por_regla"])):
            if not n:
                continue
            filas.append({
                "datos": [momento, usuario, os.path.basename(r["ruta"]),
                          carpeta_rel, tipo, regla.buscar, regla.reemplazar, n,
                          ", ".join(r["ubicaciones"]),
                          "  //  ".join(r["muestras"][idx]),
                          os.path.basename(r["backup"]) if r["backup"] else ""],
                "enlace_archivo": r["ruta"],
                "enlace_backup": r["backup"],
            })

    ruta = os.path.join(destino_carpeta, "REGISTRO_CAMBIOS.xlsx")

    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _registro_csv(filas, os.path.splitext(ruta)[0] + ".csv")

    CABECERAS = ["Fecha y hora", "Usuario", "Archivo", "Carpeta", "Tipo",
                 "Motivo: texto localizado", "Sustituido por", "N. de cambios",
                 "Donde", "Ejemplo del texto", "Copia de seguridad"]

    azul = PatternFill("solid", fgColor="1F4E79")
    blanco = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    normal = Font(name="Arial", size=10)
    enlace = Font(name="Arial", size=10, color="0563C1", underline="single")
    borde = Border(bottom=Side(style="thin", color="BFBFBF"))

    # --- abrir el registro existente o crearlo --------------------------
    if os.path.isfile(ruta):
        wb = load_workbook(ruta)
        ws = wb["Registro de cambios"] if "Registro de cambios" in wb.sheetnames \
            else wb.create_sheet("Registro de cambios")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registro de cambios"

    # ojo: leer ws.cell(1,1) crea la celda, asi que la cabecera se escribe
    # siempre en la fila 1 de forma explicita, nunca con append()
    if ws.cell(1, 1).value != CABECERAS[0]:
        for col, texto in enumerate(CABECERAS, 1):
            celda = ws.cell(1, col, texto)
            celda.fill = azul
            celda.font = blanco
            celda.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

    # --- anadir las filas de esta ejecucion -----------------------------
    fila = max(ws.max_row + 1, 2)
    for f in filas:
        for col, valor in enumerate(f["datos"], 1):
            ws.cell(fila, col, valor)
        c = ws.cell(fila, 3)
        c.hyperlink = f["enlace_archivo"]
        c.font = enlace
        if f["enlace_backup"]:
            cb = ws.cell(fila, 11)
            cb.hyperlink = f["enlace_backup"]
            cb.font = enlace
        for col in range(1, 12):
            celda = ws.cell(fila, col)
            if celda.font.color is None:
                celda.font = normal
            celda.border = borde
            celda.alignment = Alignment(vertical="top", wrap_text=(col == 10))
        fila += 1

    ultima = ws.max_row
    for i, ancho in enumerate([16, 14, 34, 26, 11, 34, 34, 13, 30, 46, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"
    if ultima >= 2:
        ws.auto_filter.ref = f"A1:K{ultima}"

    # --- rehacer el resumen sobre TODO el historico ---------------------
    if "Resumen" in wb.sheetnames:
        del wb["Resumen"]
    rs = wb.create_sheet("Resumen")
    titulo = Font(name="Arial", size=12, bold=True, color="1F4E79")
    etiqueta = Font(name="Arial", size=10, bold=True)

    # pares (buscado, sustituido) unicos de todo el registro
    pares, vistos, fechas = [], set(), set()
    for f in range(2, ultima + 1):
        if not isinstance(ws.cell(f, 8).value, (int, float)):
            continue                      # fila vacia o cabecera repetida
        par = (ws.cell(f, 6).value, ws.cell(f, 7).value)
        if par[0] is not None and par not in vistos:
            vistos.add(par)
            pares.append(par)
        if ws.cell(f, 1).value:
            fechas.add(str(ws.cell(f, 1).value)[:10])

    rs["A1"] = "REGISTRO DE CAMBIOS"
    rs["A1"].font = titulo
    f = 3
    for k, v in [("Ultima ejecucion", momento),
                 ("Usuario", usuario),
                 ("Carpeta procesada", os.path.abspath(carpeta)),
                 ("Archivos modificados (esta vez)", len(con_cambios)),
                 ("Lineas en el historico", max(ultima - 1, 0)),
                 ("Dias con cambios registrados", len(fechas))]:
        rs.cell(f, 1, k).font = etiqueta
        rs.cell(f, 2, v).font = normal
        f += 1

    f += 1
    rs.cell(f, 1, "Sustituciones acumuladas").font = titulo
    f += 1
    for col, texto in enumerate(["Texto localizado", "Sustituido por",
                                 "N. de cambios"], 1):
        c = rs.cell(f, col, texto)
        c.fill = azul
        c.font = blanco
    primera = f + 1
    for buscado, sustituido in pares:
        f += 1
        rs.cell(f, 1, buscado).font = normal
        rs.cell(f, 2, sustituido).font = normal
        # se suma desde la hoja de registro, nunca se escribe el numero a mano
        rs.cell(f, 3).value = (f"=SUMIFS('Registro de cambios'!$H:$H,"
                               f"'Registro de cambios'!$F:$F,A{f},"
                               f"'Registro de cambios'!$G:$G,B{f})")
        rs.cell(f, 3).font = normal
    f += 1
    rs.cell(f, 2, "TOTAL").font = etiqueta
    rs.cell(f, 3).value = (f"=SUM(C{primera}:C{f-1})" if pares else 0)
    rs.cell(f, 3).font = etiqueta

    errores = [r for r in resultados if r["error"]]
    if errores:
        f += 2
        rs.cell(f, 1, "Archivos con incidencias").font = titulo
        for r in errores:
            f += 1
            rs.cell(f, 1, r["ruta"]).font = normal
            rs.cell(f, 2, r["error"]).font = normal

    for col, ancho in zip("ABC", (42, 42, 16)):
        rs.column_dimensions[col].width = ancho
    wb.move_sheet("Resumen", offset=-(len(wb.sheetnames) - 1))
    wb.active = 0
    try:
        wb.save(ruta)
    except PermissionError:
        return (f"[registro abierto en Excel: cierralo y vuelve a ejecutar. "
                f"Los cambios SI se han aplicado]")
    return ruta


# ---------------------------------------------------------------------------
# Exportacion a PDF de los documentos modificados
# ---------------------------------------------------------------------------
def exportar_pdf(rutas, carpeta_destino, raiz):
    """
    Convierte a PDF los documentos indicados usando el propio Office
    instalado (Word, Excel, PowerPoint) mediante automatizacion COM.
    Cada aplicacion se abre una sola vez para todo el lote.
    Devuelve (n_correctos, [errores]).
    """
    try:
        import win32com.client as win32
    except ImportError:
        return 0, ["Falta pywin32. Instalalo con:  pip install pywin32"]

    os.makedirs(carpeta_destino, exist_ok=True)

    grupos = {"word": [], "excel": [], "ppt": []}
    for r in rutas:
        ext = os.path.splitext(r)[1].lower()
        if ext in (".docx", ".docm", ".dotx"):
            grupos["word"].append(r)
        elif ext in (".xlsx", ".xlsm", ".xltx"):
            grupos["excel"].append(r)
        elif ext in (".pptx", ".pptm"):
            grupos["ppt"].append(r)

    def destino_pdf(origen):
        rel = os.path.relpath(origen, raiz)
        salida = os.path.join(carpeta_destino, os.path.splitext(rel)[0] + ".pdf")
        os.makedirs(os.path.dirname(salida), exist_ok=True)
        return os.path.abspath(salida)

    correctos, errores = 0, []
    total = sum(len(v) for v in grupos.values())
    hechos = 0

    # ---------------------------- Word ----------------------------------
    if grupos["word"]:
        app = None
        try:
            app = win32.Dispatch("Word.Application")
            app.Visible = False
            app.DisplayAlerts = 0
            for origen in grupos["word"]:
                hechos += 1
                aviso(f"  [{hechos}/{total}] {os.path.basename(origen)}")
                doc = None
                try:
                    doc = app.Documents.Open(os.path.abspath(origen),
                                             ReadOnly=True, AddToRecentFiles=False)
                    doc.ExportAsFixedFormat(OutputFileName=destino_pdf(origen),
                                            ExportFormat=17)   # wdExportFormatPDF
                    correctos += 1
                except Exception as e:
                    errores.append(f"{os.path.basename(origen)}: {e}")
                finally:
                    if doc is not None:
                        try:
                            doc.Close(False)
                        except Exception:
                            pass
        except Exception as e:
            errores.append(f"No se pudo abrir Word: {e}")
        finally:
            if app is not None:
                try:
                    app.Quit()
                except Exception:
                    pass

    # ---------------------------- Excel ---------------------------------
    if grupos["excel"]:
        app = None
        try:
            app = win32.Dispatch("Excel.Application")
            app.Visible = False
            app.DisplayAlerts = False
            for origen in grupos["excel"]:
                hechos += 1
                aviso(f"  [{hechos}/{total}] {os.path.basename(origen)}")
                libro = None
                try:
                    libro = app.Workbooks.Open(os.path.abspath(origen),
                                               ReadOnly=True, UpdateLinks=0)
                    libro.ExportAsFixedFormat(0, destino_pdf(origen))  # 0 = PDF
                    correctos += 1
                except Exception as e:
                    errores.append(f"{os.path.basename(origen)}: {e}")
                finally:
                    if libro is not None:
                        try:
                            libro.Close(False)
                        except Exception:
                            pass
        except Exception as e:
            errores.append(f"No se pudo abrir Excel: {e}")
        finally:
            if app is not None:
                try:
                    app.Quit()
                except Exception:
                    pass

    # -------------------------- PowerPoint ------------------------------
    if grupos["ppt"]:
        app = None
        try:
            app = win32.Dispatch("PowerPoint.Application")
            for origen in grupos["ppt"]:
                hechos += 1
                aviso(f"  [{hechos}/{total}] {os.path.basename(origen)}")
                pres = None
                try:
                    pres = app.Presentations.Open(os.path.abspath(origen),
                                                  ReadOnly=True, WithWindow=False)
                    pres.SaveAs(destino_pdf(origen), 32)      # 32 = ppSaveAsPDF
                    correctos += 1
                except Exception as e:
                    errores.append(f"{os.path.basename(origen)}: {e}")
                finally:
                    if pres is not None:
                        try:
                            pres.Close()
                        except Exception:
                            pass
        except Exception as e:
            errores.append(f"No se pudo abrir PowerPoint: {e}")
        finally:
            if app is not None:
                try:
                    app.Quit()
                except Exception:
                    pass

    return correctos, errores


# ---------------------------------------------------------------------------
# Interfaz de consola
# ---------------------------------------------------------------------------
def banner():
    print(f"""{C_INFO}
================================================================
   REEMPLAZADOR MASIVO DE TEXTO EN DOCUMENTOS   v{VERSION}
   Word (.docx/.docm)  ·  Excel (.xlsx/.xlsm)  ·  PowerPoint
================================================================{C_END}""")


def aviso(texto=""):
    """print con volcado inmediato: evita que la salida parezca detenida."""
    print(texto, flush=True)


def preguntar(texto, defecto=None):
    sufijo = f" [{defecto}]" if defecto else ""
    r = input(f"{texto}{sufijo}: ").strip()
    return r or (defecto or "")


def modo_interactivo():
    banner()
    directorio = preguntar("Carpeta a procesar")
    directorio = directorio.strip('"').strip("'")
    if not os.path.isdir(directorio):
        print(f"{C_ERR}La carpeta no existe.{C_END}")
        return

    recursivo = preguntar("¿Incluir subcarpetas? (s/n)", "s").lower().startswith("s")

    def obtener_campos():
    return [
        "PROMOTOR",
        "FECHA",
        "NUM_EGR",
        "NUM_ESS",
        "NUM_PCC",
        "EXPEDIENTE",
        "PRESUPUESTO",
        "PRES.L",
        "PROYECTO",
        "DIRECCIÓN",
        "REPRESENTANTE",
        "ASES",
        "RES.EGR",
        "PRES.EGR",
        "PRES.EGR.L",
        "PRES.PCC",
        "ARQUITECTO",
        "DURACIÓN",
        "PRES.ESS"
    ]


reglas = []

print(f"\n{C_INFO}Introduce los valores para los campos.{C_END}")
print(f"{C_INFO}Deja vacío cualquier campo que no quieras sustituir.{C_END}\n")

for campo in obtener_campos():
    valor = input(f"  {campo}: ").strip()

    if valor:
        reglas.append((f"{{{{{campo}}}}}", valor))

if not reglas:
    print(f"{C_WARN}No se ha indicado ninguna sustitucion.{C_END}")
    return

    sensible = preguntar("¿Distinguir mayusculas/minusculas? (s/n)", "n").lower().startswith("s")
    completa = preguntar("¿Solo palabras completas? (s/n)", "n").lower().startswith("s")
    reglas = [Regla(b, r, sensible, completa) for b, r in reglas]

    archivos, legado = recorrer(directorio, recursivo)
    print(f"\n{C_INFO}{len(archivos)} archivo(s) compatibles encontrados.{C_END}")
    if legado:
        print(f"{C_WARN}{len(legado)} archivo(s) en formato antiguo (.doc/.xls) "
              f"se omitiran; conviertelos a .docx/.xlsx primero.{C_END}")
    if not archivos:
        return

    # ---------- 1) simulacion -------------------------------------------
    print(f"\n{C_INFO}--- SIMULACION (no se modifica nada) ---{C_END}")
    resultados = []
    for i, ruta in enumerate(archivos, 1):
        print(f"\r  Analizando {i}/{len(archivos)}...", end="", flush=True)
        resultados.append(procesar_archivo(ruta, reglas, simulacion=True,
                                           raiz=directorio))
    print("\r" + " " * 60 + "\r", end="")

    con_cambios = [r for r in resultados if r["sustituciones"]]
    total = sum(r["sustituciones"] for r in con_cambios)
    for r in con_cambios:
        print(f"  {C_OK}{r['sustituciones']:>4}{C_END}  {os.path.relpath(r['ruta'], directorio)}")
        planas = [m for grupo in r["muestras"] for m in grupo]
        for m in planas[:2]:
            print(f"        · {m}")
    errores = [r for r in resultados if r["error"]]
    for r in errores:
        print(f"  {C_ERR}ERROR{C_END} {r['ruta']}: {r['error']}")

    aviso(f"\n{C_INFO}TOTAL: {total} sustitucion(es) en "
          f"{len(con_cambios)} archivo(s).{C_END}")
    if not con_cambios:
        return

    # ---------- 2) confirmacion y aplicacion ----------------------------
    if preguntar("\n¿Aplicar los cambios? (s/n)", "n").lower() != "s":
        print("Cancelado. No se ha modificado ningun archivo.")
        return

    backup = os.path.join(directorio, "_BACKUP_REEMPLAZO")
    os.makedirs(backup, exist_ok=True)

    finales = []
    for i, r in enumerate(con_cambios, 1):
        print(f"\r  Aplicando {i}/{len(con_cambios)}...", end="", flush=True)
        finales.append(procesar_archivo(r["ruta"], reglas, False, backup,
                                        raiz=directorio))
    print("\r" + " " * 60 + "\r", end="")

    ok = sum(1 for r in finales if not r["error"])
    aviso(f"{C_OK}Hecho: {ok} archivo(s) modificados.{C_END}")
    aviso(f"Copias de seguridad en: {backup}")
    aviso("Generando registro de cambios...")
    aviso(f"Registro: {escribir_registro(finales, directorio, reglas, backup)}")

    # ---------- 3) exportacion opcional a PDF ---------------------------
    modificados = [r["ruta"] for r in finales if r["sustituciones"] and not r["error"]]
    if not modificados:
        return
    if not preguntar(f"\n¿Imprimir a PDF los {len(modificados)} documento(s) "
                     f"modificados? (s/n)", "n").lower().startswith("s"):
        return

    sugerida = os.path.join(directorio, "PDF")
    destino = preguntar("Carpeta donde guardar los PDF", sugerida)
    destino = destino.strip('"').strip("'")
    aviso(f"\n{C_INFO}Generando PDF (se abrira Office en segundo plano)...{C_END}")
    correctos, errores = exportar_pdf(modificados, destino, directorio)
    aviso(f"{C_OK}{correctos} PDF generado(s) en: {destino}{C_END}")
    for e in errores:
        aviso(f"  {C_ERR}{e}{C_END}")


def modo_argumentos(args):
    reglas = []
    if args.csv:
        with open(args.csv, encoding="utf-8-sig") as fh:
            for fila in csv.reader(fh, delimiter=";"):
                if len(fila) >= 2 and fila[0].strip():
                    reglas.append(Regla(fila[0], fila[1], args.sensible,
                                        args.palabra, args.regex))
    for par in args.reemplazo or []:
        if "=>" not in par:
            sys.exit(f"Formato incorrecto: {par}  (usa  \"viejo=>nuevo\")")
        b, r = par.split("=>", 1)
        reglas.append(Regla(b, r, args.sensible, args.palabra, args.regex))
    if not reglas:
        sys.exit("No se han definido sustituciones.")

    archivos, legado = recorrer(args.carpeta, not args.sin_subcarpetas)
    backup = None
    if args.aplicar and not args.sin_backup:
        backup = os.path.join(args.carpeta, "_BACKUP_REEMPLAZO")
        os.makedirs(backup, exist_ok=True)

    resultados = []
    for ruta in archivos:
        r = procesar_archivo(ruta, reglas, not args.aplicar, backup,
                             raiz=args.carpeta)
        resultados.append(r)
        if r["sustituciones"]:
            print(f"{r['sustituciones']:>4}  {ruta}")
    total = sum(r["sustituciones"] for r in resultados)
    estado = "APLICADAS" if args.aplicar else "SIMULADAS (usa --aplicar)"
    print(f"\nTOTAL: {total} sustitucion(es) {estado}")
    if legado:
        print(f"Omitidos {len(legado)} archivo(s) .doc/.xls (formato antiguo).")

    if backup:
        print("Registro de cambios:",
              escribir_registro(resultados, args.carpeta, reglas, backup))

    if args.pdf and args.aplicar:
        modificados = [r["ruta"] for r in resultados
                       if r["sustituciones"] and not r["error"]]
        if modificados:
            print(f"\nGenerando PDF de {len(modificados)} documento(s)...")
            correctos, errores = exportar_pdf(modificados, args.pdf, args.carpeta)
            print(f"{correctos} PDF generado(s) en: {args.pdf}")
            for e in errores:
                print(" ", e)


def desactivar_quickedit():
    """
    En Windows, el 'modo de edicion rapida' congela la salida de la consola
    en cuanto se hace clic dentro de la ventana, y la aplicacion parece
    colgada. Se desactiva al arrancar; si no se puede, no pasa nada.
    """
    if os.name != "nt":
        return
    try:
        import ctypes
        from ctypes import wintypes
        k32 = ctypes.windll.kernel32
        h = k32.GetStdHandle(-10)                 # STD_INPUT_HANDLE
        modo = wintypes.DWORD()
        if k32.GetConsoleMode(h, ctypes.byref(modo)):
            nuevo = (modo.value & ~0x0040) | 0x0080   # -QUICK_EDIT +EXTENDED
            k32.SetConsoleMode(h, nuevo)
    except Exception:
        pass


def main():
    desactivar_quickedit()
    if len(sys.argv) == 1:
        try:
            modo_interactivo()
        except KeyboardInterrupt:
            print("\nInterrumpido.")
        input("\nPulsa Enter para salir...")
        return

    p = argparse.ArgumentParser(description="Reemplazo masivo de texto en Word/Excel/PowerPoint.")
    p.add_argument("carpeta")
    p.add_argument("-r", "--reemplazo", action="append",
                   help='Par "texto viejo=>texto nuevo" (repetible)')
    p.add_argument("--csv", help="CSV con pares  buscar;reemplazar  (separador ;)")
    p.add_argument("--aplicar", action="store_true",
                   help="Escribe los cambios (sin este flag solo simula)")
    p.add_argument("--sensible", action="store_true", help="Distingue mayusculas")
    p.add_argument("--palabra", action="store_true", help="Solo palabras completas")
    p.add_argument("--regex", action="store_true", help="Interpreta como expresion regular")
    p.add_argument("--sin-subcarpetas", action="store_true")
    p.add_argument("--sin-backup", action="store_true")
    p.add_argument("--pdf", metavar="CARPETA",
                   help="Exporta a PDF los documentos modificados en esa carpeta")
    modo_argumentos(p.parse_args())


if __name__ == "__main__":
    main()
