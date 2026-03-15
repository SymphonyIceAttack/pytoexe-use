import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# Имя файла Excel
filename = 'data.xlsx'

# Проверка, существует ли файл, если нет — создать и добавить заголовки
if not os.path.exists(filename):
    wb = Workbook()
    ws = wb.active
    ws.append(["Время", "Координаты", "Тип БПЛА", "Частота", "Расчет", "Результат"])
    wb.save(filename)

def save_data():
    time = time_entry.get()
    coords = coords_entry.get()
    drone_type = type_entry.get()
    frequency = freq_entry.get()
    calculation = calc_entry.get()
    result = result_entry.get()

    if not (time and coords and drone_type and frequency and calculation and result):
        messagebox.showwarning("Ошибка", "Пожалуйста, заполните все поля")
        return

    # Открываем файл Excel и добавляем данные
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([time, coords, drone_type, frequency, calculation, result])
    wb.save(filename)

    messagebox.showinfo("Успех", "Данные сохранены в Excel файл")
    # Очистить поля после сохранения
    time_entry.delete(0, tk.END)
    coords_entry.delete(0, tk.END)
    type_entry.delete(0, tk.END)
    freq_entry.delete(0, tk.END)
    calc_entry.delete(0, tk.END)
    result_entry.delete(0, tk.END)

# Создаем окно формы
root = tk.Tk()
root.title("Ввод данных БПЛА")

tk.Label(root, text="Время").grid(row=0, column=0)
time_entry = tk.Entry(root)
time_entry.grid(row=0, column=1)

tk.Label(root, text="Координаты").grid(row=1, column=0)
coords_entry = tk.Entry(root)
coords_entry.grid(row=1, column=1)

tk.Label(root, text="Тип БПЛА").grid(row=2, column=0)
type_entry = tk.Entry(root)
type_entry.grid(row=2, column=1)

tk.Label(root, text="Частота").grid(row=3, column=0)
freq_entry = tk.Entry(root)
freq_entry.grid(row=3, column=1)

tk.Label(root, text="Расчет").grid(row=4, column=0)
calc_entry = tk.Entry(root)
calc_entry.grid(row=4, column=1)

tk.Label(root, text="Результат").grid(row=5, column=0)
result_entry = tk.Entry(root)
result_entry.grid(row=5, column=1)

submit_btn = tk.Button(root, text="Сохранить", command=save_data)
submit_btn.grid(row=6, column=0, columnspan=2)

root.mainloop()
