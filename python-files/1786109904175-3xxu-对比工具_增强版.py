#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 多条件对比工具 - 增强版
精确匹配：货品名称 + 客户 + 批号 + 入库类型
模糊匹配：规格（相似度 >= 阈值）
数值对比：验收数量
输出：带颜色高亮的 xlsx + 汇总统计
"""

import os
import sys
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========== 配置区 ==========
desktop = r'C:\Users\Administrator\Desktop'

file_a = os.path.join(desktop, '验收入库单数据.xlsx')
file_b = os.path.join(desktop, '马上放心数据.xlsx')
output_path = os.path.join(desktop, '对比结果_增强版.xlsx')

col_product  = '货品名称'
col_customer = '客户'
col_batch    = '批号'
col_in_type  = '入库类型'
col_spec     = '规格'
col_qty      = '验收数量'

SPEC_SIMILARITY_THRESHOLD = 0.85

# 颜色定义
FILL_GREEN  = PatternFill('solid', fgColor='C6EFCE')  # 完全一致
FILL_RED    = PatternFill('solid', fgColor='FFC7CE')  # 有差异
FILL_YELLOW = PatternFill('solid', fgColor='FFEB9C')  # 仅A有
FILL_BLUE   = PatternFill('solid', fgColor='BDD7EE')  # 仅B有
FONT_RED_BOLD = Font(color='9C0006', bold=True)
FONT_NORMAL   = Font()

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ========== 工具函数 ==========
def norm(s):
    """标准化字符串：去空格、转字符串、处理NaN"""
    if pd.isna(s):
        return ''
    return str(s).strip()

def norm_num(s):
    """标准化数字"""
    if pd.isna(s):
        return None
    try:
        return float(s)
    except:
        return None

def spec_similarity(a, b):
    """计算两个规格字符串的相似度"""
    a, b = norm(a), norm(b)
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()

def load_df(path):
    """加载Excel，统一列名"""
    df = pd.read_excel(path, dtype=str)
    df = df.fillna('')
    return df

def build_key(row):
    """构建精确匹配键"""
    return (
        norm(row.get(col_product, '')),
        norm(row.get(col_customer, '')),
        norm(row.get(col_batch, '')),
        norm(row.get(col_in_type, '')),
    )

def main():
    print("=" * 50)
    print("  Excel 多条件对比工具 - 增强版")
    print("=" * 50)
    print(f"\n📂 读取表A: {file_a}")
    print(f"📂 读取表B: {file_b}\n")

    if not os.path.exists(file_a):
        print(f"❌ 找不到文件: {file_a}")
        input("按回车键退出...")
        sys.exit(1)
    if not os.path.exists(file_b):
        print(f"❌ 找不到文件: {file_b}")
        input("按回车键退出...")
        sys.exit(1)

    df_a = load_df(file_a)
    df_b = load_df(file_b)

    print(f"✅ 表A 行数: {len(df_a)}")
    print(f"✅ 表B 行数: {len(df_b)}")

    # 构建索引
    a_index = {}
    for idx, row in df_a.iterrows():
        key = build_key(row)
        if key not in a_index:
            a_index[key] = []
        a_index[key].append(row)

    b_index = {}
    for idx, row in df_b.iterrows():
        key = build_key(row)
        if key not in b_index:
            b_index[key] = []
        b_index[key].append(row)

    # 获取所有key
    all_keys = set(list(a_index.keys()) + list(b_index.keys()))

    # 对比结果
    results = []
    matched_b_keys = set()

    for key in all_keys:
        a_rows = a_index.get(key, [])
        b_rows = b_index.get(key, [])
        prod, cust, batch, in_type = key

        if a_rows and b_rows:
            # 两边都有，逐行对比（取最小长度）
            max_len = max(len(a_rows), len(b_rows))
            for i in range(max_len):
                a_row = a_rows[i] if i < len(a_rows) else None
                b_row = b_rows[i] if i < len(b_rows) else None

                a_spec = norm(a_row.get(col_spec, '')) if a_row is not None else ''
                b_spec = norm(b_row.get(col_spec, '')) if b_row is not None else ''
                sim = spec_similarity(a_spec, b_spec) if (a_spec and b_spec) else (1.0 if a_spec == b_spec else 0.0)

                a_qty = norm_num(a_row.get(col_qty, '')) if a_row is not None else None
                b_qty = norm_num(b_row.get(col_qty, '')) if b_row is not None else None

                issues = []
                if sim < SPEC_SIMILARITY_THRESHOLD:
                    issues.append('规格不符')
                if a_qty is not None and b_qty is not None and a_qty != b_qty:
                    issues.append(f'数量不符(验收入库单={a_qty:.0f}, 马上放心={b_qty:.0f})')
                elif a_qty is None and b_qty is not None:
                    issues.append('验收入库单数量缺失')
                elif a_qty is not None and b_qty is None:
                    issues.append('马上放心数量缺失')

                if issues:
                    status = '❌ 存在差异'
                else:
                    status = '✅ 完全一致'

                results.append({
                    '比对结果': status,
                    '货品名称': prod,
                    '客户': cust,
                    '批号': batch,
                    '入库类型': in_type,
                    '验收入库单_规格': a_spec,
                    '马上放心_规格': b_spec,
                    '规格相似度': f'{sim*100:.1f}%',
                    '验收入库单_验收数量': a_qty if a_qty is not None else '',
                    '马上放心_验收数量': b_qty if b_qty is not None else '',
                    '数量差异': (a_qty - b_qty) if (a_qty is not None and b_qty is not None) else '',
                    '差异说明': '; '.join(issues) if issues else '无'
                })
        elif a_rows:
            for a_row in a_rows:
                results.append({
                    '比对结果': '⚠️ 仅验收入库单存在',
                    '货品名称': prod,
                    '客户': cust,
                    '批号': batch,
                    '入库类型': in_type,
                    '验收入库单_规格': norm(a_row.get(col_spec, '')),
                    '马上放心_规格': '',
                    '规格相似度': '',
                    '验收入库单_验收数量': norm_num(a_row.get(col_qty, '')) or '',
                    '马上放心_验收数量': '',
                    '数量差异': '',
                    '差异说明': f'马上放心数据中未找到(客户={cust}, 批号={batch}, 入库类型={in_type})'
                })
        elif b_rows:
            for b_row in b_rows:
                results.append({
                    '比对结果': '⚠️ 仅马上放心存在',
                    '货品名称': prod,
                    '客户': cust,
                    '批号': batch,
                    '入库类型': in_type,
                    '验收入库单_规格': '',
                    '马上放心_规格': norm(b_row.get(col_spec, '')),
                    '规格相似度': '',
                    '验收入库单_验收数量': '',
                    '马上放心_验收数量': norm_num(b_row.get(col_qty, '')) or '',
                    '数量差异': '',
                    '差异说明': f'验收入库单中未找到(客户={cust}, 批号={batch}, 入库类型={in_type})'
                })

    print(f"\n📊 对比完成，共 {len(results)} 条结果\n")

    # 统计
    cnt_ok = sum(1 for r in results if r['比对结果'] == '✅ 完全一致')
    cnt_diff = sum(1 for r in results if r['比对结果'] == '❌ 存在差异')
    cnt_only_a = sum(1 for r in results if r['比对结果'] == '⚠️ 仅验收入库单存在')
    cnt_only_b = sum(1 for r in results if r['比对结果'] == '⚠️ 仅马上放心存在')

    print(f"  ✅ 完全一致: {cnt_ok}")
    print(f"  ❌ 存在差异: {cnt_diff}")
    print(f"  ⚠️ 仅验收入库单: {cnt_only_a}")
    print(f"  ⚠️ 仅马上放心: {cnt_only_b}")

    # ========== 写入 Excel ==========
    wb = Workbook()
    ws = wb.active
    ws.title = '对比结果'

    headers = list(results[0].keys()) if results else []
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据行
    for row_data in results:
        row_values = [row_data[h] for h in headers]
        ws.append(row_values)
        row_idx = ws.max_row
        status = row_data['比对结果']

        if '✅' in status:
            fill = FILL_GREEN
        elif '❌' in status:
            fill = FILL_RED
        elif '仅验收入库单' in status:
            fill = FILL_YELLOW
        elif '仅马上放心' in status:
            fill = FILL_BLUE
        else:
            fill = PatternFill()

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        # 数量差异列加粗红字
        diff_col = headers.index('数量差异') + 1
        diff_cell = ws.cell(row=row_idx, column=diff_col)
        if diff_cell.value not in ('', None, 0, '0'):
            diff_cell.font = FONT_RED_BOLD

    # 列宽自适应
    col_widths = {
        '比对结果': 18, '货品名称': 18, '客户': 14, '批号': 16,
        '入库类型': 12, '验收入库单_规格': 22, '马上放心_规格': 22,
        '规格相似度': 12, '验收入库单_验收数量': 18, '马上放心_验收数量': 18,
        '数量差异': 12, '差异说明': 40
    }
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 14)

    # 冻结首行 + 筛选
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'

    # ========== Sheet2: 汇总统计 ==========
    ws2 = wb.create_sheet('汇总统计')
    ws2.append(['统计项', '数量'])
    for col_idx, cell in enumerate(ws2[1], 1):
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    stats = [
        ('✅ 完全一致', cnt_ok),
        ('❌ 存在差异（规格/数量）', cnt_diff),
        ('⚠️ 仅验收入库单存在', cnt_only_a),
        ('⚠️ 仅马上放心存在', cnt_only_b),
        ('合计', len(results)),
    ]
    for label, val in stats:
        ws2.append([label, val])
        row_idx = ws2.max_row
        for c in range(1, 3):
            cell = ws2.cell(row=row_idx, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if '❌' in label or '差异' in label:
                cell.font = FONT_RED_BOLD
                cell.fill = FILL_RED
            elif '✅' in label:
                cell.fill = FILL_GREEN
            elif '仅验收入库单' in label:
                cell.fill = FILL_YELLOW
            elif '仅马上放心' in label:
                cell.fill = FILL_BLUE

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 12

    # 保存
    wb.save(output_path)
    print(f"\n💾 结果已保存: {output_path}")
    print("=" * 50)
    print("  对比完成！按回车键退出...")
    print("=" * 50)
    input()

if __name__ == '__main__':
    main()
