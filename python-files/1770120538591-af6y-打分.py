import os
import re
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from tkinterdnd2 import DND_FILES, TkinterDnD

# 全局配置：缺交关键词、分数解析正则、Excel样式
ABSENT_KEYWORDS = ['缺', '空的']  # 缺交关键词（空格隔开）
SCORE_PATTERN = re.compile(r'(\d{2})$')  # 匹配文件名末尾两位数字
BORDER = Border(  # Excel单元格边框
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
ALIGN = Alignment(horizontal='center', vertical='center')  # 居中对齐

class ScoreRecorderApp:
    def __init__(self, root):
        self.root = root
        self.root.title('同学作业分数登记工具')
        self.root.geometry('600x400')
        self.root.resizable(False, False)

        # 拖放区域
        self.drop_frame = ttk.Frame(root, relief='solid', padding=50)
        self.drop_frame.pack(expand=True, fill='both', padx=20, pady=20)
        self.drop_label = ttk.Label(
            self.drop_frame,
            text='将目标文件夹拖入此处\n（内含作业文件夹、未交文件夹，作业下含同学文件夹）',
            font=('微软雅黑', 12),
            justify='center'
        )
        self.drop_label.pack()

        # 绑定拖放事件
        self.drop_frame.drop_target_register(DND_FILES)
        self.drop_frame.dnd_bind('<<Drop>>', self.on_drop)

        # 状态标签
        self.status_label = ttk.Label(
            root,
            text='等待拖入文件夹...',
            font=('微软雅黑', 10),
            foreground='#666666'
        )
        self.status_label.pack(pady=10)

        # 结果存储
        self.student_names = []  # 所有同学名（行）
        self.homework_names = []  # 所有作业名（列）
        self.score_data = {}  # 核心数据：{同学: {作业: [得分1, 得分2,...]}}
        self.absent_students = set()  # 未交作业的同学（从「未交」文件夹解析）

    def on_drop(self, event):
        """拖放文件夹后的核心处理逻辑"""
        # 获取拖入的文件夹路径（处理跨平台路径格式）
        target_dir = event.data.strip('{}').replace('/', '\\') if os.name == 'nt' else event.data.strip('{}')
        if not os.path.isdir(target_dir):
            messagebox.showerror('错误', '请拖入**文件夹**，而非文件！')
            return

        self.status_label.config(text='正在解析文件夹...', foreground='#333333')
        self.root.update()  # 刷新界面，显示状态

        try:
            # 步骤1：初始化数据，避免重复拖放的缓存问题
            self._init_data()
            # 步骤2：遍历目标文件夹，解析作业、同学、分数（适配新层级）
            self._parse_target_dir(target_dir)
            # 步骤3：生成Excel表格并保存
            self._generate_excel(target_dir)
            messagebox.showinfo('成功', f'分数表已生成！\n保存路径：{target_dir}\\同学作业分数登记表.xlsx')
            self.status_label.config(text='处理完成，可继续拖放其他文件夹...', foreground='#009900')
        except Exception as e:
            messagebox.showerror('处理失败', f'解析/生成Excel出错：\n{str(e)}')
            self.status_label.config(text='处理失败，请检查文件夹结构！', foreground='#ff0000')

    def _init_data(self):
        """初始化所有数据容器"""
        self.student_names = []
        self.homework_names = []
        self.score_data = {}
        self.absent_students = set()

    def _parse_target_dir(self, target_dir):
        """解析目标文件夹：适配新层级【作业文件夹(一级)→同学文件夹(二级)→分数文件】"""
        # 遍历目标文件夹下的所有一级子文件夹（作业文件夹/未交文件夹）
        for first_level_name in os.listdir(target_dir):
            first_level_path = os.path.join(target_dir, first_level_name)
            if not os.path.isdir(first_level_path):
                continue  # 跳过文件，只处理文件夹

            # 处理「未交」文件夹：提取未交同学名，加入未交集合
            if '未交' in first_level_name:
                self._parse_absent_folder(first_level_path)
                continue

            # 处理「作业文件夹」：提取作业名（文件夹名即为作业名）
            homework_name = first_level_name.strip()
            if not homework_name:
                continue
            if homework_name not in self.homework_names:
                self.homework_names.append(homework_name)  # 去重添加作业名

            # 遍历作业文件夹下的二级子文件夹（同学文件夹）
            for student_name in os.listdir(first_level_path):
                student_path = os.path.join(first_level_path, student_name)
                if not os.path.isdir(student_path):
                    continue  # 跳过文件，只处理同学子文件夹

                student_name = student_name.strip()
                if not student_name:
                    continue
                # 初始化同学的分数容器：首次出现则创建字典
                if student_name not in self.score_data:
                    self.student_names.append(student_name)
                    self.score_data[student_name] = {}
                # 初始化该同学当前作业的分数列表
                self.score_data[student_name][homework_name] = []

                # 遍历同学文件夹下的文件：提取分数/判断缺交
                self._parse_score_files(student_path, student_name, homework_name)

        # 去重+排序，方便查看
        self.student_names = sorted(list(set(self.student_names)))
        self.homework_names = sorted(self.homework_names)

    def _parse_absent_folder(self, absent_path):
        """解析未交文件夹：提取文件名中的同学名，加入未交集合"""
        for file_name in os.listdir(absent_path):
            file_path = os.path.join(absent_path, file_name)
            if os.path.isdir(file_path):
                continue  # 跳过子文件夹，只处理文件
            # 提取文件名中的同学名（去掉后缀，直接作为同学名）
            student_name = os.path.splitext(file_name)[0].strip()
            if student_name:
                self.absent_students.add(student_name)

    def _parse_score_files(self, student_path, student_name, homework_name):
        """解析同学文件夹下的文件：提取末尾两位数字为分数，判断缺交关键词"""
        for file_name in os.listdir(student_path):
            file_path = os.path.join(student_path, file_name)
            if os.path.isdir(file_path):
                continue  # 跳过子文件夹，只处理文件

            file_name_no_suffix = os.path.splitext(file_name)[0].strip()
            # 判断是否包含缺交关键词（空格隔开，只要有一个就判定缺交，不计分）
            if any(keyword in file_name_no_suffix.split() for keyword in ABSENT_KEYWORDS):
                continue

            # 匹配文件名末尾两位数字作为分数，匹配成功则加入分数列表
            score_match = SCORE_PATTERN.search(file_name_no_suffix)
            if score_match:
                score = int(score_match.group(1))
                self.score_data[student_name][homework_name].append(score)

    def _generate_excel(self, save_dir):
        """生成Excel表格：处理子列、未交标注、单元格样式（逻辑不变，适配新数据）"""
        wb = Workbook()
        ws = wb.active
        ws.title = '作业分数登记表'

        # 步骤1：构建Excel表头（作业名+子列，如「数学作业_1」「数学作业_2」）
        header_row = 1  # 表头行（作业主名称）
        sub_header_row = 2  # 子列表头行（1、2、3...）
        current_col = 2  # 第1列是同学名，从第2列开始是作业列
        header_mapping = {}  # 映射：(作业名, 子列索引) -> Excel列号，用于后续填值

        # 写同学名列标题
        ws.cell(row=header_row, column=1, value='同学姓名')
        ws.merge_cells(start_row=header_row, start_column=1, end_row=sub_header_row, end_column=1)
        ws.cell(row=header_row, column=1).alignment = ALIGN
        ws.cell(row=header_row, column=1).border = BORDER

        # 遍历作业名，生成主列+子列
        for homework in self.homework_names:
            # 找到该作业的最大子列数（所有同学中该作业的最多分数个数）
            max_sub_col = 0
            for student in self.score_data:
                if homework in self.score_data[student]:
                    max_sub_col = max(max_sub_col, len(self.score_data[student][homework]))
            max_sub_col = max(max_sub_col, 1)  # 至少1个子列，避免无数据

            # 合并单元格作为作业主标题
            start_col = current_col
            end_col = current_col + max_sub_col - 1
            ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=end_col)
            ws.cell(row=header_row, column=start_col, value=homework)
            ws.cell(row=header_row, column=start_col).alignment = ALIGN
            ws.cell(row=header_row, column=start_col).border = BORDER

            # 写子列标题并建立映射
            for sub_idx in range(1, max_sub_col + 1):
                header_mapping[(homework, sub_idx)] = current_col
                ws.cell(row=sub_header_row, column=current_col, value=str(sub_idx))
                ws.cell(row=sub_header_row, column=current_col).alignment = ALIGN
                ws.cell(row=sub_header_row, column=current_col).border = BORDER
                current_col += 1

        # 步骤2：填写同学名和分数/未交
        start_data_row = 3  # 从第3行开始填同学数据
        for student in self.student_names:
            # 填写同学名
            ws.cell(row=start_data_row, column=1, value=student)
            ws.cell(row=start_data_row, column=1).alignment = ALIGN
            ws.cell(row=start_data_row, column=1).border = BORDER

            # 遍历每个作业，填写对应子列的分数/未交
            for homework in self.homework_names:
                # 情况1：同学在未交集合中 → 该作业所有子列标未交
                if student in self.absent_students:
                    for sub_idx in header_mapping:
                        if sub_idx[0] == homework:
                            col = header_mapping[sub_idx]
                            ws.cell(row=start_data_row, column=col, value='未交')
                            ws.cell(row=start_data_row, column=col).alignment = ALIGN
                            ws.cell(row=start_data_row, column=col).border = BORDER
                    continue

                # 情况2：同学无该作业数据 → 标未交
                if homework not in self.score_data[student] or not self.score_data[student][homework]:
                    for sub_idx in header_mapping:
                        if sub_idx[0] == homework:
                            col = header_mapping[sub_idx]
                            ws.cell(row=start_data_row, column=col, value='未交')
                            ws.cell(row=start_data_row, column=col).alignment = ALIGN
                            ws.cell(row=start_data_row, column=col).border = BORDER
                    continue

                # 情况3：有分数 → 依次填写到子列，多余子列标未交
                score_list = self.score_data[student][homework]
                for sub_idx in header_mapping:
                    if sub_idx[0] == homework:
                        col = header_mapping[sub_idx]
                        sub_num = sub_idx[1]
                        if sub_num <= len(score_list):
                            ws.cell(row=start_data_row, column=col, value=score_list[sub_num - 1])
                        else:
                            ws.cell(row=start_data_row, column=col, value='未交')
                        ws.cell(row=start_data_row, column=col).alignment = ALIGN
                        ws.cell(row=start_data_row, column=col).border = BORDER

            start_data_row += 1

        # 步骤3：调整列宽（自适应内容，最大20）
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[col_letter].width = adjusted_width

        # 步骤4：保存Excel（覆盖已存在文件）
        save_path = os.path.join(save_dir, '同学作业分数登记表.xlsx')
        if os.path.exists(save_path):
            os.remove(save_path)
        wb.save(save_path)

if __name__ == '__main__':
    # 初始化带拖放功能的主窗口
    root = TkinterDnD.Tk()
    app = ScoreRecorderApp(root)
    root.mainloop()