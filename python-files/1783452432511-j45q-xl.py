import random
import os
from datetime import datetime
from openpyxl import load_workbook

# =========================
# ПАМЯТЬ
# =========================
DATE_FILE = "date_memory.txt"
TEXT_FILE = "text_memory.txt"
USED_FILE = "used_numbers.txt"

# =========================
# ТАБЛИЦА
# =========================
mapping = {
    1: 61.9615, 2: 80.1216, 3: 61.8605, 4: 54.7008, 5: 61.0426,
    6: 57.6513, 7: 67.8627, 8: 69.5534, 9: 71.8649, 10: 65.8709,
    11: 65.4677, 12: 67.4156, 13: 73.3084, 14: 71.9307, 15: 62.5712,
    16: 67.3564, 17: 59.3512, 18: 62.1583, 19: 68.5307, 20: 74.1168,
    21: 81.3475, 22: 79.1528, 23: 82.6164, 24: 75.3477, 25: 69.7546,
    26: 68.1427, 27: 77.4246, 28: 76.1835, 29: 64.5937, 30: 63.3746
}

# =========================
# UTIL
# =========================
def load_value(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read().strip()
    return None


def save_value(path, value):
    with open(path, "w", encoding="utf-8") as f:
        f.write(value)


def get_real_cell(ws, cell):
    for r in ws.merged_cells.ranges:
        if cell.coordinate in r:
            return ws.cell(r.min_row, r.min_col)
    return cell


# =========================
# УНИКАЛЬНЫЕ ЧИСЛА
# =========================
def get_unique_numbers():
    if os.path.exists(USED_FILE):
        with open(USED_FILE, "r") as f:
            used = set(map(int, f.read().split()))
    else:
        used = set()

    available = list(set(range(1, 31)) - used)

    if len(available) < 2:
        used = set()
        available = list(range(1, 31))

    nums = random.sample(available, 2)
    used.update(nums)

    with open(USED_FILE, "w") as f:
        f.write(" ".join(map(str, used)))

    return nums


# =========================
# ФАЙЛ
# =========================
file_path = input("Введите путь к Excel (.xlsx): ").strip()

if not os.path.exists(file_path) or not file_path.endswith(".xlsx"):
    print("❌ Неверный файл")
    exit()

# =========================
# ЦИКЛ РАБОТЫ
# =========================
while True:
    print("\n========================")
    print("🔁 Новый цикл (q = выход)")
    print("========================\n")

    saved_date = load_value(DATE_FILE)
    saved_text = load_value(TEXT_FILE)

    new_date = input("Дата " + f"{saved_date}" " (Enter = оставить): ").strip().replace(",", ".")
    if new_date.lower() == "q":
        break

    new_text = input("Текст " + f"{saved_text}" + " (Enter = оставить): ").strip()
    if new_text.lower() == "q":
        break

    ab = input("Введите a и b через пробел: ").strip().split()
    if ab and ab[0].lower() == "q":
        break
    if len(ab) != 2:
        print("❌ Нужно 2 значения")
        continue

    a, b = ab


    try:
        base_number = float(input("Введите число для M8/M9: ").replace(",", "."))
    except:
        print("❌ Ошибка числа")
        continue

    # =========================
    # ЗАГРУЗКА КАЖДЫЙ ЦИКЛ
    # =========================
    wb = load_workbook(file_path)
    # номер листа exel 0,1,2,3,4
    ws = wb.worksheets[1]

    # =========================
    # ДАТА
    # =========================
    if new_date:
        date_to_use = new_date
        save_value(DATE_FILE, new_date)
    else:
        date_to_use = load_value(DATE_FILE) or datetime.now().strftime("%Y-%m-%d")

    # =========================
    # ТЕКСТ
    # =========================
    if new_text:
        text_to_use = new_text
        save_value(TEXT_FILE, new_text)
    else:
        text_to_use = load_value(TEXT_FILE) or ""

    # =========================
    # A8
    # =========================
    get_real_cell(ws, ws["A8"]).value = date_to_use

    # =========================
    # A9
    # =========================
    get_real_cell(ws, ws["A9"]).value = f"64-20-10/{b}-26"

    # =========================
    # B8:B9
    # =========================
    ws.merge_cells("B8:B9")
    get_real_cell(ws, ws["B8"]).value = a

    # =========================
    # C8 C9
    # =========================
    nums = get_unique_numbers()
    get_real_cell(ws, ws["C8"]).value = nums[0]
    get_real_cell(ws, ws["C9"]).value = nums[1]

    # =========================
    # E8 E9
    # =========================
    get_real_cell(ws, ws["E8"]).value = float(mapping[nums[0]])
    get_real_cell(ws, ws["E9"]).value = float(mapping[nums[1]])

    ws["E8"].number_format = "0.0000"
    ws["E9"].number_format = "0.0000"

    # =========================
    # M8 M9
    # =========================
    offset = random.randint(0, 20 if base_number <= 500 else 50) / 10

    get_real_cell(ws, ws["M8"]).value = base_number + offset
    get_real_cell(ws, ws["M9"]).value = base_number - offset

    ws["M8"].number_format = "0.0"
    ws["M9"].number_format = "0.0"

    # =========================
    # D19:E19
    # =========================
    ws.merge_cells("D19:E19")
    get_real_cell(ws, ws["D19"]).value = text_to_use

    # =========================
    # СОХРАНЕНИЕ
    # =========================
    new_name = f"{a} 64-20-10-{b}-26.xlsx"
    new_path = os.path.join(os.path.dirname(file_path), new_name)

    wb.save(new_path)

    print(f"\n✅ Готово: {new_path}")