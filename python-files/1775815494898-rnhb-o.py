import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading

# 使用 openpyxl 代替 pandas，更轻量，打包成功率 100%
try:
    from openpyxl import load_workbook
except ImportError:
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("缺少库", "请在本地 CMD 执行: pip install openpyxl")
    exit()

def one_excel_to_js_lite(excel_path, js_path, log_func):
    try:
        # data_only=True 确保读取的是公式计算后的值而不是公式本身
        wb = load_workbook(excel_path, data_only=True)
        json_data = {}
        
        for sheet_name in wb.sheetnames:
            # 忽略以下划线开头的页签
            if sheet_name.startswith('_'):
                continue
            
            sheet = wb[sheet_name]
            # 第 11 行为列名 (openpyxl 行号从 1 开始)
            header_row = 11
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            if max_row < header_row:
                continue
                
            # 获取列名并记录哪些列不以下划线开头
            headers = []
            valid_col_indices = []
            for i in range(1, max_col + 1):
                col_name = sheet.cell(row=header_row, column=i).value
                if col_name and not str(col_name).startswith('_'):
                    headers.append(str(col_name))
                    valid_col_indices.append(i)
            
            # 读取数据
            rows_data = []
            for r in range(header_row + 1, max_row + 1):
                row_dict = {}
                has_data = False
                for idx, col_idx in enumerate(valid_col_indices):
                    val = sheet.cell(row=r, column=col_idx).value
                    # 处理空值，保持与之前逻辑一致
                    row_dict[headers[idx]] = val if val is not None else ""
                    if val is not None:
                        has_data = True
                if has_data: # 略过完全空的行
                    rows_data.append(row_dict)
            
            json_data[sheet_name] = rows_data
            
        if json_data:
            with open(js_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=4, ensure_ascii=False)
            return True
        return False
    except Exception as e:
        log_func(f"❌ 错误: {os.path.basename(excel_path)} -> {str(e)}")
        return False

class ExcelToJsonGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 转 JSON 转换器 (Lite版)")
        self.root.geometry("600x450")

        tk.Label(root, text="Excel 目录:").pack(pady=(10,0), padx=20, anchor="w")
        self.e_path = tk.Entry(root)
        self.e_path.pack(fill="x", padx=20, pady=5)
        tk.Button(root, text="选择...", command=lambda: self.get_path(self.e_path)).pack(anchor="e", padx=20)

        tk.Label(root, text="JSON 输出目录:").pack(pady=(10,0), padx=20, anchor="w")
        self.j_path = tk.Entry(root)
        self.j_path.pack(fill="x", padx=20, pady=5)
        tk.Button(root, text="选择...", command=lambda: self.get_path(self.j_path)).pack(anchor="e", padx=20)

        self.log_box = scrolledtext.ScrolledText(root, height=10)
        self.log_box.pack(fill="both", expand=True, padx=20, pady=10)

        self.btn = tk.Button(root, text="开始转换", bg="#4CAF50", fg="white", command=self.start)
        self.btn.pack(pady=10)

    def get_path(self, entry):
        p = filedialog.askdirectory()
        if p:
            entry.delete(0, tk.END)
            entry.insert(0, p)

    def log(self, msg):
        self.log_box.insert(tk.END, msg + "\n")
        self.log_box.see(tk.END)

    def start(self):
        threading.Thread(target=self.work, daemon=True).start()

    def work(self):
        src, dst = self.e_path.get(), self.j_path.get()
        if not src or not dst: return
        self.btn.config(state="disabled")
        self.log(">>> 正在扫描文件...")
        try:
            files = [f for f in os.listdir(src) if f.endswith('.xlsx') and not f.startswith('~$')]
            for f in files:
                if one_excel_to_js_lite(os.path.join(src, f), os.path.join(dst, f.replace('.xlsx', '.json')), self.log):
                    self.log(f"成功: {f}")
            self.log(">>> 处理完毕")
            messagebox.showinfo("完成", "转换任务已结束")
        finally:
            self.btn.config(state="normal")

if __name__ == "__main__":
    root = tk.Tk()
    ExcelToJsonGUI(root)
    root.mainloop()