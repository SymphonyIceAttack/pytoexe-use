#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pymysql
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import os
import sys
import argparse
import configparser

# ---------- НАСТРОЙКИ ПО УМОЛЧАНИЮ ----------
DEFAULT_CONFIG = {
    'Database': {
        'host': '172.25.3.49',
        'port': '3305',
        'user': 'root',
        'password': '',
        # Имена баз и таблиц жёстко заданы в коде, но можно вынести в конфиг
    },
    'Report': {
        'threshold_in': '08:00'   # порог опоздания (ЧЧ:ММ)
    }
}

CONFIG_FILE = 'config.ini'

# Имена БД и таблиц (фиксированы)
DB_LOG = 'tc-db-log'
DB_MAIN = 'tc-db-main'
TABLE_LOGS = 'logs'
TABLE_PERSONAL = 'personal'
# -----------------------------------------------------------------

def load_config():
    config = configparser.ConfigParser()
    config.read_dict(DEFAULT_CONFIG)
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
    return config

def parse_event_type(logdata):
    """Извлекает тип события из LOGDATA: 1 - выход, 2 - вход."""
    if not logdata:
        return None
    hex_str = logdata[2:] if logdata.startswith('0x') else logdata
    try:
        data = bytes.fromhex(hex_str)
        if len(data) > 3:
            return data[3]
    except ValueError:
        return None
    return None

def get_attendance(date, db_config):
    """
    Возвращает словарь {id: {'name': str, 'first_entry': datetime or None, 'last_exit': datetime or None}}
    только для сотрудников, отметившихся в этот день.
    """
    conn = None
    cursor = None
    try:
        conn = pymysql.connect(**db_config)
        cursor = conn.cursor()

        # Экранируем имена БД и таблиц (они содержат дефисы)
        logs_table = f"`{DB_LOG}`.`{TABLE_LOGS}`"
        personal_table = f"`{DB_MAIN}`.`{TABLE_PERSONAL}`"

        # 1. Получаем все логи за день
        sql_logs = f"""
            SELECT LOGTIME, EMPHINT, LOGDATA
            FROM {logs_table}
            WHERE DATE(LOGTIME) = %s
        """
        cursor.execute(sql_logs, (date,))
        rows = cursor.fetchall()

        if not rows:
            return {}

        # 2. Собираем EMPHINT и события
        emp_ids = set()
        events_by_emp = {}
        for logtime, emphint, logdata in rows:
            if emphint is None or emphint == 0:
                continue
            emp_ids.add(emphint)
            etype = parse_event_type(logdata)
            if etype not in (1, 2):
                continue
            events_by_emp.setdefault(emphint, []).append((logtime, etype))

        if not emp_ids:
            return {}

        # 3. Получаем имена сотрудников
        placeholders = ','.join(['%s'] * len(emp_ids))
        sql_names = f"""
            SELECT ID, NAME
            FROM {personal_table}
            WHERE ID IN ({placeholders})
        """
        cursor.execute(sql_names, list(emp_ids))
        name_map = {row[0]: row[1] for row in cursor.fetchall()}

        # 4. Определяем первый вход и последний выход для каждого
        result = {}
        for emp_id in emp_ids:
            events = events_by_emp.get(emp_id, [])
            if not events:
                continue
            first_entry = None
            last_exit = None
            for logtime, etype in events:
                if etype == 2:  # вход
                    if first_entry is None or logtime < first_entry:
                        first_entry = logtime
                elif etype == 1:  # выход
                    if last_exit is None or logtime > last_exit:
                        last_exit = logtime
            result[emp_id] = {
                'name': name_map.get(emp_id, f"ID {emp_id}"),
                'first_entry': first_entry,
                'last_exit': last_exit
            }
        return result

    except Exception as e:
        print(f"Ошибка при работе с БД: {e}", file=sys.stderr)
        raise
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def format_timedelta(td):
    if td is None:
        return ""
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    if hours > 0:
        return f"{hours} ч {minutes} мин"
    else:
        return f"{minutes} мин"

def save_report(attendance_data, threshold_time, filepath):
    wb = openpyxl.Workbook()

    # ---------- Лист 1: Все отметившиеся ----------
    ws_all = wb.active
    ws_all.title = "Все сотрудники"
    headers = ['№', 'ФИО', 'ВХОД', 'ВЫХОД']
    ws_all.append(headers)
    for col in range(1, 5):
        cell = ws_all.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    row_num = 2
    for idx, (emp_id, data) in enumerate(attendance_data.items(), start=1):
        ws_all.cell(row=row_num, column=1, value=idx)
        ws_all.cell(row=row_num, column=2, value=data['name'])

        # Вход
        cell_entry = ws_all.cell(row=row_num, column=3)
        if data['first_entry']:
            cell_entry.value = data['first_entry'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            cell_entry.value = "не отметился"
            cell_entry.fill = red_fill
            cell_entry.font = Font(color="FFFFFF")

        # Выход
        cell_exit = ws_all.cell(row=row_num, column=4)
        if data['last_exit']:
            cell_exit.value = data['last_exit'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            cell_exit.value = "не отметился"
            cell_exit.fill = red_fill
            cell_exit.font = Font(color="FFFFFF")

        row_num += 1

    # Автоширина
    for col in ws_all.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
        ws_all.column_dimensions[col_letter].width = max_len + 2

    # ---------- Лист 2: Опоздавшие ----------
    ws_late = wb.create_sheet("Опоздавшие")
    headers_late = ['№', 'ФИО', 'Время входа', 'Время выхода', 'Опоздание']
    ws_late.append(headers_late)
    for col in range(1, 6):
        cell = ws_late.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    late_employees = []
    for emp_id, data in attendance_data.items():
        first_entry = data['first_entry']
        if first_entry and first_entry.time() > threshold_time:
            late_employees.append({
                'name': data['name'],
                'entry': first_entry,
                'exit': data['last_exit'],
                'late_delta': first_entry - datetime.datetime.combine(first_entry.date(), threshold_time)
            })

    late_employees.sort(key=lambda x: x['entry'])

    row_num = 2
    for idx, item in enumerate(late_employees, start=1):
        ws_late.cell(row=row_num, column=1, value=idx)
        ws_late.cell(row=row_num, column=2, value=item['name'])
        ws_late.cell(row=row_num, column=3, value=item['entry'].strftime('%Y-%m-%d %H:%M:%S'))
        if item['exit']:
            ws_late.cell(row=row_num, column=4, value=item['exit'].strftime('%Y-%m-%d %H:%M:%S'))
        else:
            ws_late.cell(row=row_num, column=4, value="не отметился")
        ws_late.cell(row=row_num, column=5, value=format_timedelta(item['late_delta']))
        row_num += 1

    for col in ws_late.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
        ws_late.column_dimensions[col_letter].width = max_len + 2

    wb.save(filepath)
    print(f"Отчёт сохранён в {filepath}")

def main():
    config = load_config()
    db_config = {
        'host': config['Database']['host'],
        'port': int(config['Database']['port']),
        'user': config['Database']['user'],
        'password': config['Database']['password'],
        'database': DB_MAIN,  # Подключаемся к tc-db-main, но в запросах используем полные имена
        'charset': 'utf8mb4'
    }
    threshold_in_str = config['Report']['threshold_in']

    parser = argparse.ArgumentParser(description='Отчёт по явке сотрудников')
    parser.add_argument('--date', type=str, help='Дата ГГГГ-ММ-ДД (по умолчанию сегодня)')
    parser.add_argument('--threshold', type=str, help='Порог опоздания ЧЧ:ММ')
    parser.add_argument('--output', type=str, help='Имя выходного файла')
    args = parser.parse_args()

    if args.date:
        try:
            report_date = datetime.datetime.strptime(args.date, '%Y-%m-%d').date()
        except ValueError:
            print("Неверный формат даты. Используйте ГГГГ-ММ-ДД", file=sys.stderr)
            sys.exit(1)
    else:
        report_date = datetime.date.today()

    try:
        threshold_time = datetime.datetime.strptime(args.threshold if args.threshold else threshold_in_str, '%H:%M').time()
    except ValueError:
        print("Неверный формат времени. Используйте ЧЧ:ММ", file=sys.stderr)
        sys.exit(1)

    print(f"Дата: {report_date}, порог: {threshold_time}")

    try:
        attendance = get_attendance(report_date, db_config)
        if not attendance:
            print("За эту дату нет ни одного прохода. Отчёт не создан.")
            return

        filename = args.output if args.output else f"report_{report_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        save_report(attendance, threshold_time, filepath)

    except Exception as e:
        print(f"Критическая ошибка: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()