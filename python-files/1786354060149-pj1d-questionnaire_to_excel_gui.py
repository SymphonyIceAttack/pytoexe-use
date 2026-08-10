import re
import tkinter as tk
from tkinter import messagebox
from pathlib import Path
from openpyxl import load_workbook

EXCEL_FILE = Path(__file__).with_name("questionnaire_responses.xlsx")
SHEET_NAME = "Candidates"

FIELDS = [
    "Name",
    "Local Contact Number",
    "Age",
    "Nationality",
    "Current Employer",
    "Current Job Title",
    "Brands Dealt With",
    "Total Years of Experience",
    "Educational Background",
    "Current/Latest Salary",
    "Expected Salary",
    "Notice Period",
]

def parse_response(raw_text):
    text = raw_text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("–", "-").replace("—", "-")

    field_pattern = "|".join(re.escape(f) for f in sorted(FIELDS, key=len, reverse=True))

    pattern = re.compile(
        rf"(?im)^\s*({field_pattern})\s*(?::|-|=)\s*(.*?)\s*(?=\n\s*(?:{field_pattern})\s*(?::|-|=)|\Z)",
        re.DOTALL,
    )

    answers = {field: "" for field in FIELDS}

    for match in pattern.finditer(text):
        field = match.group(1)
        answers[field] = match.group(2).strip()

    return answers

def add_to_excel(answers):
    if not EXCEL_FILE.exists():
        messagebox.showerror(
            "Excel file not found",
            f"Couldn't find:\n{EXCEL_FILE.name}\n\n"
            "Put the Excel file in the same folder as this program."
        )
        return False

    try:
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME not in wb.sheetnames:
            messagebox.showerror("Sheet not found", f"Couldn't find the '{SHEET_NAME}' worksheet.")
            return False

        ws = wb[SHEET_NAME]
        headers = [cell.value for cell in ws[1]]
        header_to_col = {
            str(header).strip(): i + 1
            for i, header in enumerate(headers)
            if header is not None
        }

        # Find the first genuinely empty row rather than relying on max_row,
        # which can include the template's starter row.
        next_row = 2
        while any(ws.cell(next_row, col).value not in (None, "") for col in range(1, len(headers) + 1)):
            next_row += 1

        for field in FIELDS:
            if field in header_to_col:
                ws.cell(next_row, header_to_col[field], answers[field])

        wb.save(EXCEL_FILE)
        return True

    except PermissionError:
        messagebox.showerror(
            "Excel file is open",
            "Please close the Excel workbook and try again."
        )
        return False
    except Exception as e:
        messagebox.showerror("Could not save", str(e))
        return False

class App:
    def __init__(self, root):
        self.root = root
        root.title("Questionnaire → Excel")
        root.geometry("760x650")
        root.minsize(650, 550)

        title = tk.Label(
            root,
            text="Questionnaire → Excel",
            font=("Segoe UI", 18, "bold")
        )
        title.pack(pady=(15, 2))

        subtitle = tk.Label(
            root,
            text="Paste the questionnaire response below, review it, then add it to Excel.",
            font=("Segoe UI", 10)
        )
        subtitle.pack(pady=(0, 10))

        self.input_box = tk.Text(root, wrap="word", font=("Consolas", 10))
        self.input_box.pack(fill="both", expand=True, padx=15, pady=5)

        buttons = tk.Frame(root)
        buttons.pack(fill="x", padx=15, pady=10)

        tk.Button(
            buttons, text="Extract & Review", command=self.extract,
            font=("Segoe UI", 10, "bold"), padx=12, pady=7
        ).pack(side="left")

        tk.Button(
            buttons, text="Clear", command=self.clear,
            font=("Segoe UI", 10), padx=12, pady=7
        ).pack(side="left", padx=8)

        tk.Button(
            buttons, text="Exit", command=root.destroy,
            font=("Segoe UI", 10), padx=12, pady=7
        ).pack(side="right")

        self.status = tk.Label(root, text="Ready.", anchor="w", font=("Segoe UI", 9))
        self.status.pack(fill="x", padx=15, pady=(0, 10))

    def extract(self):
        raw = self.input_box.get("1.0", "end").strip()

        if not raw:
            messagebox.showwarning("Nothing to process", "Paste a questionnaire response first.")
            return

        answers = parse_response(raw)

        if not any(answers.values()):
            messagebox.showwarning(
                "No fields found",
                "I couldn't identify any of the expected fields.\n\n"
                "Make sure the response contains labels such as:\n"
                "Name: John Smith"
            )
            return

        self.show_review(answers)

    def show_review(self, answers):
        window = tk.Toplevel(self.root)
        window.title("Review Candidate")
        window.geometry("760x650")
        window.transient(self.root)
        window.grab_set()

        tk.Label(
            window, text="Review Candidate",
            font=("Segoe UI", 16, "bold")
        ).pack(pady=(15, 5))

        tk.Label(
            window,
            text="You can edit any value before adding it to Excel.",
            font=("Segoe UI", 10)
        ).pack(pady=(0, 10))

        frame = tk.Frame(window)
        frame.pack(fill="both", expand=True, padx=20)

        entries = {}

        for row, field in enumerate(FIELDS):
            tk.Label(
                frame, text=field + ":",
                anchor="w", width=28,
                font=("Segoe UI", 9, "bold")
            ).grid(row=row, column=0, sticky="nw", pady=3)

            entry = tk.Entry(frame, font=("Segoe UI", 9))
            entry.insert(0, answers[field])
            entry.grid(row=row, column=1, sticky="ew", pady=3)
            entries[field] = entry

        frame.columnconfigure(1, weight=1)

        buttons = tk.Frame(window)
        buttons.pack(fill="x", padx=20, pady=15)

        def save():
            final_answers = {field: entries[field].get() for field in FIELDS}

            if not final_answers["Name"].strip():
                if not messagebox.askyesno(
                    "Name is blank",
                    "The Name field is blank. Add this candidate anyway?",
                    parent=window
                ):
                    return

            if add_to_excel(final_answers):
                self.status.config(text=f"Added: {final_answers['Name'] or 'Unnamed candidate'}")
                messagebox.showinfo("Added", "Candidate added to Excel.", parent=window)
                window.destroy()
                self.clear()

        tk.Button(
            buttons, text="Add to Excel", command=save,
            font=("Segoe UI", 10, "bold"), padx=15, pady=7
        ).pack(side="left")

        tk.Button(
            buttons, text="Cancel", command=window.destroy,
            font=("Segoe UI", 10), padx=15, pady=7
        ).pack(side="left", padx=8)

    def clear(self):
        self.input_box.delete("1.0", "end")
        self.status.config(text="Ready.")
        self.input_box.focus_set()

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
