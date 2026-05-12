import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class KartBalancer:

    def __init__(self, root):
        self.root = root
        self.root.title("Kart Race Balancer")
        self.root.geometry("1100x700")

        self.last_assignments = []

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True)

        self.setup_page = ttk.Frame(self.notebook)
        self.results_page = ttk.Frame(self.notebook)

        self.notebook.add(self.setup_page, text="Настройка")
        self.notebook.add(self.results_page, text="Результаты")

        self.build_setup_page()
        self.build_results_page()

    def build_setup_page(self):

        title = tk.Label(
            self.setup_page,
            text="Распределение картов",
            font=("Arial", 22, "bold")
        )
        title.pack(pady=15)

        main_frame = tk.Frame(self.setup_page)
        main_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # КАРТЫ
        karts_frame = tk.LabelFrame(
            main_frame,
            text="Карты (от быстрого к медленному)",
            padx=10,
            pady=10
        )
        karts_frame.pack(side="left", fill="both", expand=True, padx=10)

        self.karts_text = tk.Text(karts_frame, width=40, height=25)
        self.karts_text.pack(fill="both", expand=True)

        # ГОНЩИКИ
        drivers_frame = tk.LabelFrame(
            main_frame,
            text="Гонщики (от быстрого к медленному)",
            padx=10,
            pady=10
        )
        drivers_frame.pack(side="left", fill="both", expand=True, padx=10)

        self.drivers_text = tk.Text(drivers_frame, width=40, height=25)
        self.drivers_text.pack(fill="both", expand=True)

        bottom_frame = tk.Frame(self.setup_page)
        bottom_frame.pack(pady=15)

        tk.Label(
            bottom_frame,
            text="Количество гонок:",
            font=("Arial", 12)
        ).pack(side="left", padx=5)

        self.races_var = tk.IntVar(value=3)

        races_spinbox = tk.Spinbox(
            bottom_frame,
            from_=1,
            to=20,
            width=5,
            textvariable=self.races_var,
            font=("Arial", 12)
        )
        races_spinbox.pack(side="left", padx=5)

        generate_btn = tk.Button(
            self.setup_page,
            text="Сгенерировать распределение",
            font=("Arial", 14, "bold"),
            bg="#2ecc71",
            fg="white",
            padx=20,
            pady=10,
            command=self.generate_assignments
        )
        generate_btn.pack(pady=10)

        export_btn = tk.Button(
            self.setup_page,
            text="Экспорт в Excel",
            font=("Arial", 12, "bold"),
            bg="#3498db",
            fg="white",
            padx=20,
            pady=8,
            command=self.export_to_excel
        )
        export_btn.pack(pady=5)

    def build_results_page(self):

        self.results_container = tk.Frame(self.results_page)
        self.results_container.pack(fill="both", expand=True, padx=20, pady=20)

    def get_clean_lines(self, text_widget):

        data = text_widget.get("1.0", tk.END)

        return [
            line.strip()
            for line in data.splitlines()
            if line.strip()
        ]

    def generate_assignments(self):

        drivers = self.get_clean_lines(self.drivers_text)
        karts = self.get_clean_lines(self.karts_text)

        race_count = self.races_var.get()

        if not drivers or not karts:
            messagebox.showerror(
                "Ошибка",
                "Заполните списки гонщиков и картов"
            )
            return

        if len(drivers) != len(karts):
            messagebox.showerror(
                "Ошибка",
                "Количество гонщиков должно совпадать с количеством картов"
            )
            return

        if race_count > len(karts):
            messagebox.showerror(
                "Ошибка",
                "Количество гонок не может быть больше количества картов"
            )
            return

        assignments = self.smart_distribution(
            drivers,
            karts,
            race_count
        )

        self.last_assignments = assignments

        self.show_results(assignments)

        self.notebook.select(self.results_page)

    def smart_distribution(self, drivers, karts, race_count):

        n = len(drivers)

        used_by_driver = defaultdict(set)

        all_races = []

        reversed_karts = list(reversed(karts))

        for race in range(race_count):

            race_result = {}

            shift = race % n

            rotated = (
                reversed_karts[shift:] +
                reversed_karts[:shift]
            )

            available = rotated.copy()

            for i, driver in enumerate(drivers):

                selected_kart = None

                for kart in available:

                    if kart not in used_by_driver[driver]:
                        selected_kart = kart
                        break

                if selected_kart is None:
                    selected_kart = available[0]

                race_result[driver] = selected_kart

                used_by_driver[driver].add(selected_kart)

                available.remove(selected_kart)

            all_races.append(race_result)

        return all_races

    def show_results(self, assignments):

        for widget in self.results_container.winfo_children():
            widget.destroy()

        canvas = tk.Canvas(self.results_container)

        scrollbar = ttk.Scrollbar(
            self.results_container,
            orient="vertical",
            command=canvas.yview
        )

        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window(
            (0, 0),
            window=scrollable_frame,
            anchor="nw"
        )

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        title = tk.Label(
            scrollable_frame,
            text="Результаты распределения",
            font=("Arial", 24, "bold")
        )
        title.pack(pady=20)

        for race_index, race_data in enumerate(assignments, start=1):

            race_frame = tk.LabelFrame(
                scrollable_frame,
                text=f"Гонка {race_index}",
                font=("Arial", 14, "bold"),
                padx=10,
                pady=10
            )
            race_frame.pack(fill="x", padx=20, pady=10)

            headers = tk.Frame(race_frame)
            headers.pack(fill="x", pady=5)

            tk.Label(
                headers,
                text="Гонщик",
                width=30,
                anchor="w",
                font=("Arial", 12, "bold")
            ).pack(side="left")

            tk.Label(
                headers,
                text="Карт",
                width=30,
                anchor="w",
                font=("Arial", 12, "bold")
            ).pack(side="left")

            ttk.Separator(
                race_frame,
                orient="horizontal"
            ).pack(fill="x", pady=5)

            for driver in sorted(race_data.keys()):

                kart = race_data[driver]

                row = tk.Frame(race_frame)
                row.pack(fill="x", pady=2)

                tk.Label(
                    row,
                    text=driver,
                    width=30,
                    anchor="w",
                    font=("Arial", 11)
                ).pack(side="left")

                tk.Label(
                    row,
                    text=kart,
                    width=30,
                    anchor="w",
                    font=("Arial", 11)
                ).pack(side="left")

    def export_to_excel(self):

        if not self.last_assignments:
            messagebox.showwarning(
                "Нет данных",
                "Сначала сгенерируйте распределение"
            )
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Сохранить Excel файл"
        )

        if not filepath:
            return

        wb = Workbook()
        ws = wb.active

        ws.title = "Kart Assignments"

        header_fill = PatternFill(
            start_color="2E86C1",
            end_color="2E86C1",
            fill_type="solid"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        center = Alignment(
            horizontal="center",
            vertical="center"
        )

        current_row = 1

        for race_index, race_data in enumerate(self.last_assignments, start=1):

            title_cell = ws.cell(row=current_row, column=1)
            title_cell.value = f"Гонка {race_index}"
            title_cell.font = Font(
                bold=True,
                size=14
            )

            current_row += 1

            headers = ["Гонщик", "Карт"]

            for col, header in enumerate(headers, start=1):

                cell = ws.cell(
                    row=current_row,
                    column=col
                )

                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            current_row += 1

            for driver in sorted(race_data.keys()):

                kart = race_data[driver]

                ws.cell(
                    row=current_row,
                    column=1
                ).value = driver

                ws.cell(
                    row=current_row,
                    column=2
                ).value = kart

                ws.cell(
                    row=current_row,
                    column=1
                ).alignment = center

                ws.cell(
                    row=current_row,
                    column=2
                ).alignment = center

                current_row += 1

            current_row += 2

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

        wb.save(filepath)

        messagebox.showinfo(
            "Успех",
            "Excel файл успешно сохранён!"
        )


if __name__ == "__main__":

    root = tk.Tk()

    app = KartBalancer(root)

    root.mainloop()