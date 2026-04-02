import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime
import json
import os

import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# 设置matplotlib中文显示
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'WenQuanYi Micro Hei', 'Arial Unicode MS', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# 美化后的颜色主题
COLORS = {
    'bg':         '#f0f4f8', 'card':       '#ffffff', 'border':     '#e2e8f0',
    'primary':    '#3b82f6', 'primary_h':  '#2563eb', 'success':    '#10b981',
    'success_h':  '#059669', 'warning':    '#f59e0b', 'warning_h':  '#d97706',
    'danger':     '#ef4444', 'danger_h':   '#dc2626', 'info':       '#8b5cf6',
    'info_h':     '#7c3aed', 'dark':       '#1e293b', 'text':       '#334155',
    'dim':        '#94a3b8', 'systolic':   '#ef4444', 'diastolic':  '#3b82f6',
    'heart_rate': '#f97316', 'spo2':       '#10b981', 'input_bg':   '#f8fafc',
    'input_focus':'#3b82f6',
}

class HealthRecordApp:
    def __init__(self, root):
        self.root = root
        self.root.title("健康数据管理系统 - 血压 / 血氧 / 心跳")
        self.root.geometry("1150x780")
        self.root.minsize(1050, 700)
        self.root.configure(bg=COLORS['bg'])

        self.data_file = os.path.join(os.path.expanduser("~"), "health_records.json")
        self.config_file = os.path.join(os.path.expanduser("~"), "health_config.json")
        
        self.data = self._load_data()
        self.ranges = self._load_config()
        self.current_person = None
        self.filter_mode = "all" # all, healthy, unhealthy

        self._build_ui()
        self._refresh_person_list()

    # ======================== 数据持久化 ========================
    def _load_data(self):
        if os.path.exists(self.data_file):
            try: return json.load(open(self.data_file, 'r', encoding='utf-8'))
            except: return {}
        return {}

    def _load_config(self):
        # 默认健康范围
        default_ranges = {
            'heart_rate_min': 60, 'heart_rate_max': 100,
            'systolic_min': 90, 'systolic_max': 140,
            'diastolic_min': 60, 'diastolic_max': 90,
            'spo2_min': 95, 'spo2_max': 100
        }
        if os.path.exists(self.config_file):
            try: return json.load(open(self.config_file, 'r', encoding='utf-8'))
            except: return default_ranges
        return default_ranges

    def _save_data(self):
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def _save_config(self):
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(self.ranges, f, ensure_ascii=False, indent=2)

    # ======================== UI 构建 ========================
    def _build_ui(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Beauty.Treeview", background=COLORS['card'], foreground=COLORS['text'], 
                        fieldbackground=COLORS['card'], borderwidth=0, font=("Microsoft YaHei", 10), rowheight=32)
        style.configure("Beauty.Treeview.Heading", background=COLORS['bg'], foreground=COLORS['dark'], 
                        font=("Microsoft YaHei", 10, "bold"), relief=tk.FLAT, borderwidth=0)
        style.map("Beauty.Treeview", background=[('selected', COLORS['primary'])], foreground=[('selected', COLORS['card'])])
        style.configure("TScrollbar", background=COLORS['border'], troughcolor=COLORS['bg'], borderwidth=0)

        main_frame = tk.Frame(self.root, bg=COLORS['bg'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # ---- 左侧面板 ----
        left_panel = tk.LabelFrame(main_frame, text="  👥 人员管理  ", font=("Microsoft YaHei", 11, "bold"),
                                   bg=COLORS['card'], fg=COLORS['dark'], padx=12, pady=12, width=240,
                                   bd=0, highlightbackground=COLORS['border'], highlightthickness=1)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 12))
        left_panel.pack_propagate(False)

        tk.Label(left_panel, text="姓名:", font=("Microsoft YaHei", 10), bg=COLORS['card'], fg=COLORS['text']).pack(anchor=tk.W, pady=(5, 0))
        self.name_entry = self._make_entry(left_panel)
        self.name_entry.pack(fill=tk.X, pady=(2, 5))
        self.name_entry.bind('<Return>', lambda e: self._add_person())

        btn_frame = tk.Frame(left_panel, bg=COLORS['card'])
        btn_frame.pack(fill=tk.X, pady=5)
        self._make_button(btn_frame, "➕ 添加", COLORS['success'], COLORS['success_h'], self._add_person).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 3))
        self._make_button(btn_frame, "🗑️ 删除", COLORS['danger'], COLORS['danger_h'], self._delete_person).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(3, 0))

        tk.Label(left_panel, text="人员列表:", font=("Microsoft YaHei", 10, "bold"), bg=COLORS['card'], fg=COLORS['dark']).pack(anchor=tk.W, pady=(12, 2))

        list_frame = tk.Frame(left_panel, bg=COLORS['card'])
        list_frame.pack(fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.person_listbox = tk.Listbox(list_frame, font=("Microsoft YaHei", 10), yscrollcommand=scrollbar.set,
                                         selectmode=tk.SINGLE, bg=COLORS['input_bg'], selectbackground=COLORS['primary'],
                                         selectforeground=COLORS['card'], relief=tk.FLAT, highlightthickness=1,
                                         highlightcolor=COLORS['primary'], highlightbackground=COLORS['border'], activestyle='none')
        self.person_listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.person_listbox.yview)
        self.person_listbox.bind('<<ListboxSelect>>', self._on_person_select)

        # 设置范围按钮
        self._make_button(left_panel, "⚙️ 设置健康范围", COLORS['info'], COLORS['info_h'], self._open_settings, width=24).pack(pady=(15, 0))

        # ---- 右侧面板 ----
        right_panel = tk.Frame(main_frame, bg=COLORS['bg'])
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 顶部录入
        input_frame = tk.LabelFrame(right_panel, text="  📝 数据录入  ", font=("Microsoft YaHei", 11, "bold"),
                                    bg=COLORS['card'], fg=COLORS['dark'], padx=15, pady=12,
                                    bd=0, highlightbackground=COLORS['border'], highlightthickness=1)
        input_frame.pack(fill=tk.X, pady=(0, 12))

        self.current_person_label = tk.Label(input_frame, text="当前人员: 未选择", font=("Microsoft YaHei", 11, "bold"),
                                             bg=COLORS['card'], fg=COLORS['danger'])
        self.current_person_label.grid(row=0, column=0, columnspan=6, sticky=tk.W, pady=(0, 8))

        labels = ["📅 时间:", "❤️ 心率:", "🔴 收缩压:", "🔵 舒张压:", "💚 血氧(%):", "📝 备注:"]
        self.entries = {}
        defaults = [self._now_str(), "", "", "", "", ""]
        for i, (lbl, default) in enumerate(zip(labels, defaults)):
            col = (i % 3) * 2
            row = (i // 3) + 1
            tk.Label(input_frame, text=lbl, font=("Microsoft YaHei", 10), bg=COLORS['card'], fg=COLORS['text']).grid(row=row, column=col, sticky=tk.E, padx=(5, 4), pady=5)
            entry = self._make_entry(input_frame, width=14)
            entry.insert(0, default)
            entry.grid(row=row, column=col + 1, sticky=tk.W, padx=(0, 12), pady=5)
            key = ['time', 'heart_rate', 'systolic', 'diastolic', 'spo2', 'note'][i]
            self.entries[key] = entry

        btn_row = tk.Frame(input_frame, bg=COLORS['card'])
        btn_row.grid(row=3, column=0, columnspan=6, pady=(10, 0))
        self._make_button(btn_row, "💾 保存记录", COLORS['primary'], COLORS['primary_h'], self._save_record, width=13).pack(side=tk.LEFT, padx=5)
        self._make_button(btn_row, "📊 查看图表", COLORS['info'], COLORS['info_h'], lambda: self._show_charts('today'), width=13).pack(side=tk.LEFT, padx=5)
        self._make_button(btn_row, "📈 健康统计", COLORS['warning'], COLORS['warning_h'], self._show_statistics, width=13).pack(side=tk.LEFT, padx=5)
        self._make_button(btn_row, "📤 导出Excel", COLORS['success'], COLORS['success_h'], self._export_excel, width=13).pack(side=tk.LEFT, padx=5)

        # 中部筛选与范围
        range_frame = tk.Frame(right_panel, bg=COLORS['card'], padx=12, pady=8,
                               highlightbackground=COLORS['border'], highlightthickness=1)
        range_frame.pack(fill=tk.X, pady=(0, 12))

        tk.Label(range_frame, text="数据筛选:", font=("Microsoft YaHei", 10, "bold"), bg=COLORS['card'], fg=COLORS['dark']).pack(side=tk.LEFT, padx=(0, 8))
        self.filter_var = tk.StringVar(value="all")
        for text, val, clr, clr_h in [("全部数据", "all", COLORS['primary'], COLORS['primary_h']), 
                                       ("🟢 健康", "healthy", COLORS['success'], COLORS['success_h']), 
                                       ("🔴 非健康", "unhealthy", COLORS['danger'], COLORS['danger_h'])]:
            rb = tk.Radiobutton(range_frame, text=text, variable=self.filter_var, value=val,
                                font=("Microsoft YaHei", 10, "bold"), bg=COLORS['card'], fg=clr,
                                selectcolor=COLORS['card'], activebackground=COLORS['card'], activeforeground=clr_h,
                                indicatoron=0, padx=12, pady=4, relief=tk.FLAT, bd=0, cursor="hand2")
            rb.pack(side=tk.LEFT, padx=4)
            # 悬浮效果
            rb.bind('<Enter>', lambda e, b=rb, c=clr: b.config(bg=c, fg='white'))
            rb.bind('<Leave>', lambda e, b=rb: b.config(bg=COLORS['card'], fg=COLORS['dark'] if val=="all" else clr))
            self.filter_var.trace_add('write', lambda *args: self._refresh_table())

        # 底部表格
        table_frame = tk.LabelFrame(right_panel, text="  📋 历史记录  ", font=("Microsoft YaHei", 11, "bold"),
                                    bg=COLORS['card'], fg=COLORS['dark'], padx=8, pady=8,
                                    bd=0, highlightbackground=COLORS['border'], highlightthickness=1)
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("状态", "时间", "心率", "收缩压", "舒张压", "血氧", "备注")
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=10, style="Beauty.Treeview")
        col_widths = [50, 150, 70, 70, 70, 70, 140]
        for col, w in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor=tk.CENTER, minwidth=50)

        tree_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview, style="TScrollbar")
        self.tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
        self.tree.pack(fill=tk.BOTH, expand=True, pady=5)

        table_btn_frame = tk.Frame(table_frame, bg=COLORS['card'])
        table_btn_frame.pack(fill=tk.X, pady=(5, 0))
        self._make_button(table_btn_frame, "✏️ 编辑选中", COLORS['warning'], COLORS['warning_h'], self._edit_record, width=13).pack(side=tk.LEFT, padx=5)
        self._make_button(table_btn_frame, "🗑️ 删除选中", COLORS['danger'], COLORS['danger_h'], self._delete_record, width=13).pack(side=tk.LEFT, padx=5)
        self._make_button(table_btn_frame, "🔄 刷新时间", COLORS['info'], COLORS['info_h'], self._refresh_time, width=13).pack(side=tk.LEFT, padx=5)

    # --- 辅助函数 ---
    def _make_entry(self, parent, width=None):
        entry = tk.Entry(parent, font=("Microsoft YaHei", 10), bg=COLORS['input_bg'], fg=COLORS['dark'],
                         relief=tk.FLAT, highlightbackground=COLORS['border'], highlightthickness=1, insertbackground=COLORS['dark'])
        if width: entry.configure(width=width)
        entry.bind('<FocusIn>', lambda e: entry.configure(highlightbackground=COLORS['input_focus'], highlightcolor=COLORS['input_focus']))
        entry.bind('<FocusOut>', lambda e: entry.configure(highlightbackground=COLORS['border'], highlightcolor=COLORS['border']))
        return entry

    def _make_button(self, parent, text, bg_color, hover_color, command, width=None):
        btn = tk.Button(parent, text=text, command=command, bg=bg_color, fg='white',
                        font=("Microsoft YaHei", 9, "bold"), relief=tk.FLAT, cursor='hand2',
                        activebackground=hover_color, activeforeground='white', padx=10, pady=6, bd=0)
        if width: btn.configure(width=width)
        btn.bind('<Enter>', lambda e: btn.configure(bg=hover_color))
        btn.bind('<Leave>', lambda e: btn.configure(bg=bg_color))
        return btn

    def _now_str(self): return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ======================== 核心逻辑：健康判定 ========================
    def _is_healthy(self, rec):
        """判断单条记录是否完全在健康范围内"""
        r = self.ranges
        try:
            hr_ok = r['heart_rate_min'] <= rec['heart_rate'] <= r['heart_rate_max']
            sys_ok = r['systolic_min'] <= rec['systolic'] <= r['systolic_max']
            dia_ok = r['diastolic_min'] <= rec['diastolic'] <= r['diastolic_max']
            spo2_ok = r['spo2_min'] <= rec['spo2'] <= r['spo2_max']
            return hr_ok and sys_ok and dia_ok and spo2_ok
        except (KeyError, TypeError):
            return False

    def _get_unhealthy_items(self, rec):
        """获取不健康的指标列表"""
        r = self.ranges
        items = []
        try:
            if not (r['heart_rate_min'] <= rec['heart_rate'] <= r['heart_rate_max']): items.append("心率")
            if not (r['systolic_min'] <= rec['systolic'] <= r['systolic_max']): items.append("高压")
            if not (r['diastolic_min'] <= rec['diastolic'] <= r['diastolic_max']): items.append("低压")
            if not (r['spo2_min'] <= rec['spo2'] <= r['spo2_max']): items.append("血氧")
        except: pass
        return items

    # ======================== 人员与记录管理 ========================
    def _refresh_person_list(self):
        self.person_listbox.delete(0, tk.END)
        for name in sorted(self.data.keys()):
            self.person_listbox.insert(tk.END, f"  {name} ({len(self.data[name])}条)")

    def _add_person(self):
        name = self.name_entry.get().strip()
        if not name: messagebox.showwarning("提示", "请输入姓名！"); return
        if name in self.data: messagebox.showwarning("提示", f"人员 '{name}' 已存在！"); return
        self.data[name] = []
        self._save_data()
        self.name_entry.delete(0, tk.END)
        self._refresh_person_list()
        idx = sorted(self.data.keys()).index(name)
        self.person_listbox.selection_set(idx)
        self._on_person_select(None)

    def _delete_person(self):
        if not self.current_person: return
        if messagebox.askyesno("确认删除", f"确定要删除 '{self.current_person}' 及其所有记录吗？"):
            del self.data[self.current_person]
            self._save_data()
            self.current_person = None
            self.current_person_label.config(text="当前人员: 未选择", fg=COLORS['danger'])
            self._refresh_person_list()
            self._clear_table()

    def _on_person_select(self, event):
        selection = self.person_listbox.curselection()
        if not selection: return
        idx = selection[0]
        name = sorted(self.data.keys())[idx]
        self.current_person = name
        self.current_person_label.config(text=f"当前人员: ✅ {name}", fg=COLORS['success'])
        self._refresh_table()

    def _refresh_time(self):
        self.entries['time'].delete(0, tk.END)
        self.entries['time'].insert(0, self._now_str())

    def _save_record(self):
        if not self.current_person: messagebox.showwarning("提示", "请先选择一个人员！"); return
        try:
            record = {
                'time': self.entries['time'].get().strip(),
                'heart_rate': int(self.entries['heart_rate'].get().strip()),
                'systolic': int(self.entries['systolic'].get().strip()),
                'diastolic': int(self.entries['diastolic'].get().strip()),
                'spo2': int(float(self.entries['spo2'].get().strip())),
                'note': self.entries['note'].get().strip(),
            }
        except ValueError:
            messagebox.showerror("错误", "请确保数值输入正确！"); return

        self.data[self.current_person].append(record)
        self.data[self.current_person].sort(key=lambda x: x['time'])
        self._save_data()

        for key in ['heart_rate', 'systolic', 'diastolic', 'spo2', 'note']: self.entries[key].delete(0, tk.END)
        self._refresh_time()
        self._refresh_table()
        self._refresh_person_list()
        
        status = "🟢 健康" if self._is_healthy(record) else "🔴 存在异常"
        messagebox.showinfo("成功", f"✅ 记录保存成功！\n当前判定: {status}")

    def _refresh_table(self):
        self._clear_table()
        if not self.current_person or self.current_person not in self.data: return
        records = self.data[self.current_person]
        mode = self.filter_var.get()
        
        for i, rec in enumerate(reversed(records)):
            real_idx = len(records) - 1 - i
            is_ok = self._is_healthy(rec)
            
            if mode == "healthy" and not is_ok: continue
            if mode == "unhealthy" and is_ok: continue
            
            status_text = "🟢" if is_ok else "🔴"
            self.tree.insert('', tk.END, values=(
                status_text, rec.get('time', ''), rec.get('heart_rate', ''), 
                rec.get('systolic', ''), rec.get('diastolic', ''), rec.get('spo2', ''), rec.get('note', '')
            ), iid=str(real_idx))

    def _clear_table(self):
        for item in self.tree.get_children(): self.tree.delete(item)

    def _edit_record(self):
        selection = self.tree.selection()
        if not selection: messagebox.showwarning("提示", "请先选择一条记录！"); return
        if not self.current_person: return
        idx = int(selection[0])
        records = self.data[self.current_person]
        if idx >= len(records): return
        record = records[idx]

        edit_win = tk.Toplevel(self.root)
        edit_win.title("✏️ 编辑记录")
        edit_win.geometry("420x400")
        edit_win.configure(bg=COLORS['bg'])
        edit_win.resizable(False, False)
        edit_win.transient(self.root)
        edit_win.grab_set()

        tk.Label(edit_win, text="编辑健康记录", font=("Microsoft YaHei", 14, "bold"), bg=COLORS['bg'], fg=COLORS['dark']).pack(pady=(20, 15))
        frame = tk.Frame(edit_win, bg=COLORS['card'], padx=20, pady=20, highlightbackground=COLORS['border'], highlightthickness=1)
        frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=(0, 15))

        edit_entries = {}
        for i, (label, key) in enumerate([("时间:", 'time'), ("心率:", 'heart_rate'), ("收缩压:", 'systolic'), ("舒张压:", 'diastolic'), ("血氧 (%):", 'spo2'), ("备注:", 'note')]):
            tk.Label(frame, text=label, font=("Microsoft YaHei", 10), bg=COLORS['card'], fg=COLORS['text']).grid(row=i, column=0, sticky=tk.E, padx=(5, 10), pady=6)
            entry = self._make_entry(frame, width=22)
            entry.insert(0, str(record.get(key, '')))
            entry.grid(row=i, column=1, pady=6)
            edit_entries[key] = entry

        def save_edit():
            try:
                record['time'] = edit_entries['time'].get().strip()
                record['heart_rate'] = int(edit_entries['heart_rate'].get().strip())
                record['systolic'] = int(edit_entries['systolic'].get().strip())
                record['diastolic'] = int(edit_entries['diastolic'].get().strip())
                record['spo2'] = int(float(edit_entries['spo2'].get().strip()))
                record['note'] = edit_entries['note'].get().strip()
            except ValueError:
                messagebox.showerror("错误", "请确保数字字段输入有效！", parent=edit_win); return
            self.data[self.current_person].sort(key=lambda x: x['time'])
            self._save_data()
            self._refresh_table()
            edit_win.destroy()

        btn_frame = tk.Frame(edit_win, bg=COLORS['bg'])
        btn_frame.pack(pady=(0, 20))
        self._make_button(btn_frame, "💾 保存", COLORS['success'], COLORS['success_h'], save_edit, width=14).pack(side=tk.LEFT, padx=12)
        self._make_button(btn_frame, "❌ 取消", COLORS['danger'], COLORS['danger_h'], edit_win.destroy, width=14).pack(side=tk.LEFT, padx=12)

    def _delete_record(self):
        selection = self.tree.selection()
        if not selection: return
        idx = int(selection[0])
        if messagebox.askyesno("确认删除", "确定要删除该条记录吗？"):
            self.data[self.current_person].pop(idx)
            self._save_data()
            self._refresh_table()
            self._refresh_person_list()

    # ======================== 健康范围设置 ========================
    def _open_settings(self):
        win = tk.Toplevel(self.root)
        win.title("⚙️ 设置健康范围标准")
        win.geometry("450x420")
        win.configure(bg=COLORS['bg'])
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        tk.Label(win, text="自定义健康指标范围", font=("Microsoft YaHei", 14, "bold"), bg=COLORS['bg'], fg=COLORS['dark']).pack(pady=(25, 15))
        
        card = tk.Frame(win, bg=COLORS['card'], padx=25, pady=20, highlightbackground=COLORS['border'], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=30, pady=(0, 15))

        settings_entries = {}
        fields = [
            ("❤️ 心率", 'heart_rate'), ("🔴 收缩压", 'systolic'), 
            ("🔵 舒张压", 'diastolic'), ("💚 血氧(%)", 'spo2')
        ]
        
        for i, (lbl, key) in enumerate(fields):
            tk.Label(card, text=lbl, font=("Microsoft YaHei", 11, "bold"), bg=COLORS['card'], fg=COLORS['text']).grid(row=i, column=0, sticky=tk.W, pady=8)
            tk.Label(card, text="最低:", font=("Microsoft YaHei", 10), bg=COLORS['card'], fg=COLORS['dim']).grid(row=i, column=1, padx=(15, 5))
            e_min = self._make_entry(card, width=6)
            e_min.insert(0, str(self.ranges[f"{key}_min"]))
            e_min.grid(row=i, column=2, pady=8)
            tk.Label(card, text="最高:", font=("Microsoft YaHei", 10), bg=COLORS['card'], fg=COLORS['dim']).grid(row=i, column=3, padx=(15, 5))
            e_max = self._make_entry(card, width=6)
            e_max.insert(0, str(self.ranges[f"{key}_max"]))
            e_max.grid(row=i, column=4, pady=8)
            settings_entries[key] = (e_min, e_max)

        def save_settings():
            try:
                for key, (e_min, e_max) in settings_entries.items():
                    self.ranges[f"{key}_min"] = int(e_min.get())
                    self.ranges[f"{key}_max"] = int(e_max.get())
                self._save_config()
                self._refresh_table() # 刷新表格状态
                win.destroy()
                messagebox.showinfo("成功", "✅ 健康范围已更新并保存！")
            except ValueError:
                messagebox.showerror("错误", "请输入有效的整数！", parent=win)

        btn_frame = tk.Frame(win, bg=COLORS['bg'])
        btn_frame.pack(pady=(0, 20))
        self._make_button(btn_frame, "💾 保存设置", COLORS['primary'], COLORS['primary_h'], save_settings, width=14).pack(side=tk.LEFT, padx=12)
        self._make_button(btn_frame, "取消", COLORS['dim'], COLORS['text'], win.destroy, width=14).pack(side=tk.LEFT, padx=12)

    # ======================== 数据可视化 (带颜色标识) ========================
    def _show_charts(self, time_range='today'):
        if not self.current_person or not self.data.get(self.current_person):
            messagebox.showwarning("提示", "请先选择人员并确保有数据！"); return

        records = self.data[self.current_person]
        now = datetime.datetime.now()
        today_str = now.strftime("%Y-%m-%d")

        if time_range == 'today':
            filtered = [r for r in records if r['time'].startswith(today_str)]; range_title = "今日"
        elif time_range == 'week':
            d = (now - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
            filtered = [r for r in records if r['time'].split()[0] >= d]; range_title = "近7天"
        else:
            d = (now - datetime.timedelta(days=30)).strftime("%Y-%m-%d")
            filtered = [r for r in records if r['time'].split()[0] >= d]; range_title = "近30天"

        if not filtered: messagebox.showinfo("提示", f"{range_title}暂无数据！"); return

        chart_win = tk.Toplevel(self.root)
        chart_win.title(f"📊 {self.current_person} - {range_title}健康数据")
        chart_win.geometry("1000x780")
        chart_win.configure(bg=COLORS['card'])

        fig = Figure(figsize=(12, 9), dpi=96, facecolor=COLORS['card'])
        times = [r['time'] for r in filtered]
        time_labels = [t.split(' ')[1][:5] if time_range == 'today' else t[:5] for t in times]
        x = list(range(len(time_labels)))
        
        sys_bp = [r['systolic'] for r in filtered]
        dia_bp = [r['diastolic'] for r in filtered]
        hr = [r['heart_rate'] for r in filtered]
        sp = [r['spo2'] for r in filtered]
        health_status = [self._is_healthy(r) for r in filtered]

        def style_ax(ax):
            ax.set_facecolor(COLORS['bg'])
            ax.grid(True, linestyle='--', alpha=0.4, color=COLORS['border'])
            ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
            ax.spines['left'].set_color(COLORS['border']); ax.spines['bottom'].set_color(COLORS['border'])
            ax.tick_params(colors=COLORS['text'], labelsize=8)

        def get_colors(status_list, default_color):
            return [COLORS['success'] if s else COLORS['danger'] for s in status_list]

        # 图1: 血压
        ax1 = fig.add_subplot(3, 1, 1)
        style_ax(ax1)
        # 用散点图代替折线图以实现多色
        sys_colors = get_colors(health_status, COLORS['systolic'])
        dia_colors = get_colors(health_status, COLORS['diastolic'])
        ax1.plot(x, sys_bp, color=COLORS['border'], linewidth=1, zorder=1) # 灰色连线
        ax1.plot(x, dia_bp, color=COLORS['border'], linewidth=1, zorder=1)
        ax1.scatter(x, sys_bp, color=sys_colors, s=60, zorder=3, label='收缩压', edgecolors='white', linewidth=0.5)
        ax1.scatter(x, dia_bp, color=dia_colors, s=60, zorder=3, label='舒张压', marker='s', edgecolors='white', linewidth=0.5)
        # 参考线
        ax1.axhspan(self.ranges['systolic_min'], self.ranges['systolic_max'], alpha=0.05, color=COLORS['systolic'])
        ax1.axhspan(self.ranges['diastolic_min'], self.ranges['diastolic_max'], alpha=0.05, color=COLORS['diastolic'])
        ax1.axhline(y=self.ranges['systolic_max'], color=COLORS['systolic'], linestyle='--', alpha=0.3)
        ax1.axhline(y=self.ranges['diastolic_min'], color=COLORS['diastolic'], linestyle='--', alpha=0.3)
        for xi, s, c in zip(x, sys_bp, sys_colors): ax1.text(xi, s+3, str(s), ha='center', fontsize=7, color=c, fontweight='bold')
        for xi, d, c in zip(x, dia_bp, dia_colors): ax1.text(xi, d-5, str(d), ha='center', fontsize=7, color=c, fontweight='bold')
        
        ax1.set_title("血压趋势 (🟢健康 🔴异常)", fontweight='bold', fontsize=12, color=COLORS['dark'])
        ax1.set_ylabel('血压', fontsize=10, color=COLORS['text'])
        ax1.legend(loc='upper right', fontsize=9, frameon=False)
        ax1.set_xticks(x); ax1.set_xticklabels(time_labels, rotation=30, ha='right')

        # 图2: 心率
        ax2 = fig.add_subplot(3, 1, 2)
        style_ax(ax2)
        hr_colors = get_colors(health_status, COLORS['heart_rate'])
        ax2.plot(x, hr, color=COLORS['border'], linewidth=1, zorder=1)
        ax2.scatter(x, hr, color=hr_colors, s=60, zorder=3, edgecolors='white', linewidth=0.5)
        ax2.axhspan(self.ranges['heart_rate_min'], self.ranges['heart_rate_max'], alpha=0.05, color=COLORS['heart_rate'])
        ax2.axhline(y=self.ranges['heart_rate_max'], color=COLORS['danger'], linestyle='--', alpha=0.3)
        ax2.axhline(y=self.ranges['heart_rate_min'], color=COLORS['success'], linestyle='--', alpha=0.3)
        for xi, h, c in zip(x, hr, hr_colors): ax2.text(xi, h+2, str(h), ha='center', fontsize=8, color=c, fontweight='bold')
        
        ax2.set_title("心率趋势 (🟢健康 🔴异常)", fontweight='bold', fontsize=12, color=COLORS['dark'])
        ax2.set_ylabel('心率', fontsize=10, color=COLORS['text'])
        ax2.set_xticks(x); ax2.set_xticklabels(time_labels, rotation=30, ha='right')

        # 图3: 血氧
        ax3 = fig.add_subplot(3, 1, 3)
        style_ax(ax3)
        sp_colors = get_colors(health_status, COLORS['spo2'])
        ax3.plot(x, sp, color=COLORS['border'], linewidth=1, zorder=1)
        ax3.scatter(x, sp, color=sp_colors, s=60, zorder=3, edgecolors='white', linewidth=0.5)
        ax3.axhline(y=self.ranges['spo2_min'], color=COLORS['danger'], linestyle='--', alpha=0.3)
        for xi, s, c in zip(x, sp, sp_colors): ax3.text(xi, s-2, f"{s}%", ha='center', fontsize=8, color=c, fontweight='bold')
        ax3.set_ylim(min(sp)-3, 101)
        
        ax3.set_title("血氧趋势 (🟢健康 🔴异常)", fontweight='bold', fontsize=12, color=COLORS['dark'])
        ax3.set_ylabel('血氧 (%)', fontsize=10, color=COLORS['text'])
        ax3.set_xticks(x); ax3.set_xticklabels(time_labels, rotation=30, ha='right')
        ax3.set_xlabel('时间', fontsize=10, color=COLORS['text'])

        fig.suptitle(f"{self.current_person} 的健康数据 - {range_title}", fontsize=15, fontweight='bold', y=0.98, color=COLORS['dark'])
        fig.tight_layout(rect=[0, 0, 1, 0.96])

        canvas = FigureCanvasTkAgg(fig, master=chart_win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        toolbar_frame = tk.Frame(chart_win, bg=COLORS['card'])
        toolbar_frame.pack(fill=tk.X, pady=(0, 10))
        self._make_button(toolbar_frame, "❌ 关闭", COLORS['danger'], COLORS['danger_h'], chart_win.destroy, width=12).pack(side=tk.RIGHT, padx=20)

    # ======================== 统计分析可视化 ========================
    def _show_statistics(self):
        if not self.current_person or not self.data.get(self.current_person):
            messagebox.showwarning("提示", "暂无数据可分析！"); return

        records = self.data[self.current_person]
        total = len(records)
        healthy_list = [r for r in records if self._is_healthy(r)]
        unhealthy_list = [r for r in records if not self._is_healthy(r)]
        h_cnt, uh_cnt = len(healthy_list), len(unhealthy_list)

        # 统计异常指标频次
        err_stats = {"心率异常": 0, "高压异常": 0, "低压异常": 0, "血氧异常": 0}
        for r in unhealthy_list:
            items = self._get_unhealthy_items(r)
            if "心率" in items: err_stats["心率异常"] += 1
            if "高压" in items: err_stats["高压异常"] += 1
            if "低压" in items: err_stats["低压异常"] += 1
            if "血氧" in items: err_stats["血氧异常"] += 1

        win = tk.Toplevel(self.root)
        win.title(f"📈 {self.current_person} - 健康统计分析")
        win.geometry("1000x750")
        win.configure(bg=COLORS['card'])

        fig = Figure(figsize=(13, 9), dpi=96, facecolor=COLORS['card'])

        def style_ax(ax):
            ax.set_facecolor(COLORS['bg'])
            ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
            ax.spines['left'].set_color(COLORS['border']); ax.spines['bottom'].set_color(COLORS['border'])
            ax.tick_params(colors=COLORS['text'], labelsize=9)

        # --- 图1: 健康率环形图 ---
        ax1 = fig.add_subplot(2, 2, 1)
        style_ax(ax1)
        sizes = [h_cnt, uh_cnt] if total > 0 else [1, 0]
        colors_pie = [COLORS['success'], COLORS['danger']]
        explode = (0.05, 0.05) if uh_cnt > 0 else (0.05, 0)
        
        wedges, texts, autotexts = ax1.pie(sizes, explode=explode, labels=['健康', '非健康'], 
                                           colors=colors_pie, autopct='%1.1f%%', startangle=90,
                                           pctdistance=0.75, textprops={'color': COLORS['dark'], 'fontweight': 'bold'})
        # 画中心圆变成环形图
        centre_circle = plt.Circle((0,0),0.50,fc=COLORS['card'])
        ax1.add_artist(centre_circle)
        ax1.text(0, 0, f"{h_cnt}/{total}\n记录", ha='center', va='center', fontsize=12, fontweight='bold', color=COLORS['dark'])
        ax1.set_title("总体健康率", fontweight='bold', fontsize=12, color=COLORS['dark'], pad=15)

        # --- 图2: 异常指标分布柱状图 ---
        ax2 = fig.add_subplot(2, 2, 2)
        style_ax(ax2)
        labels_err = list(err_stats.keys())
        values_err = list(err_stats.values())
        bars_colors = [COLORS['danger'] if v > 0 else COLORS['success'] for v in values_err]
        
        bars = ax2.bar(labels_err, values_err, color=bars_colors, width=0.6, edgecolor='white', linewidth=1)
        for bar, val in zip(bars, values_err):
            if val > 0: ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.2, str(val), 
                                 ha='center', va='bottom', fontweight='bold', color=COLORS['danger'])
        ax2.set_title("异常指标频次分布 (基于非健康记录)", fontweight='bold', fontsize=12, color=COLORS['dark'], pad=15)
        ax2.set_ylabel("异常次数", color=COLORS['text'])

        # --- 图3: 健康与非健康数值对比箱线图 ---
        ax3 = fig.add_subplot(2, 2, 3)
        style_ax(ax3)
        
        data_to_plot = []
        tick_labels = []
        metrics = [
            ('heart_rate', '心率', COLORS['heart_rate']),
            ('systolic', '收缩压', COLORS['systolic']),
            ('diastolic', '舒张压', COLORS['diastolic']),
            ('spo2', '血氧', COLORS['spo2'])
        ]
        
        for key, name, color in metrics:
            if h_cnt > 0: data_to_plot.append([r[key] for r in healthy_list]); tick_labels.append(f"{name}(健)")
            if uh_cnt > 0: data_to_plot.append([r[key] for r in unhealthy_list]); tick_labels.append(f"{name}(异)")

        if data_to_plot:
            bp = ax3.boxplot(data_to_plot, patch_artist=True, widths=0.6)
            for i, patch in enumerate(bp['boxes']):
                is_healthy_metric = "健" in tick_labels[i]
                patch.set_facecolor(COLORS['success'] if is_healthy_metric else COLORS['danger'])
                patch.set_alpha(0.6)
            ax3.set_xticklabels(tick_labels, fontsize=8)
        ax3.set_title("健康 vs 异常 数据分布对比", fontweight='bold', fontsize=12, color=COLORS['dark'], pad=15)
        ax3.grid(axis='y', linestyle='--', alpha=0.4)

        # --- 图4: 文本统计摘要 ---
        ax4 = fig.add_subplot(2, 2, 4)
        ax4.set_facecolor(COLORS['card'])
        ax4.axis('off')
        
        def calc_stats(lst, key):
            if not lst: return "无数据", "无数据", "无数据", "无数据"
            vals = [r[key] for r in lst]
            return round(sum(vals)/len(vals), 1), min(vals), max(vals), round((max(vals)-min(vals))/2, 1) if len(vals)>1 else 0

        summary_text = f"【 {self.current_person} 的健康数据摘要 】\n\n"
        summary_text += f"总记录数: {total}    🟢 健康: {h_cnt}    🔴 非健康: {uh_cnt}\n"
        summary_text += f"{'='*40}\n\n"
        
        for key, name, _ in metrics:
            avg_h, min_h, max_h, rng_h = calc_stats(healthy_list, key)
            avg_u, min_u, max_u, rng_u = calc_stats(unhealthy_list, key)
            summary_text += f"◉ {name}:\n"
            summary_text += f"  健康组 -> 均值:{avg_h}  范围:[{min_h} ~ {max_h}]\n"
            summary_text += f"  异常组 -> 均值:{avg_u}  范围:[{min_u} ~ {max_u}]\n\n"

        ax4.text(0.05, 0.95, summary_text, transform=ax4.transAxes, fontsize=10,
                 verticalalignment='top', fontfamily='Microsoft YaHei',
                 bbox=dict(boxstyle="round,pad=0.8", facecolor=COLORS['bg'], edgecolor=COLORS['border'], alpha=0.8),
                 color=COLORS['dark'], linespacing=1.5)

        fig.suptitle(f"{self.current_person} 健康统计分析报告", fontsize=16, fontweight='bold', y=0.99, color=COLORS['dark'])
        fig.tight_layout(rect=[0, 0, 1, 0.96])

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        toolbar_frame = tk.Frame(win, bg=COLORS['card'])
        toolbar_frame.pack(fill=tk.X, pady=(0, 10))
        self._make_button(toolbar_frame, "📤 导出本图", COLORS['primary'], COLORS['primary_h'], 
                          lambda: fig.savefig(f"统计报告_{self.current_person}.png", dpi=150, bbox_inches='tight'), width=12).pack(side=tk.LEFT, padx=20)
        self._make_button(toolbar_frame, "❌ 关闭", COLORS['danger'], COLORS['danger_h'], win.destroy, width=12).pack(side=tk.RIGHT, padx=20)

    # ======================== 数据导出 ========================
    def _export_excel(self):
        if not self.data: messagebox.showwarning("提示", "暂无数据可导出！"); return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("错误", "需要安装 openpyxl 库！\n请运行: pip install openpyxl"); return

        export_win = tk.Toplevel(self.root)
        export_win.title("📤 导出Excel")
        export_win.geometry("360x220")
        export_win.configure(bg=COLORS['bg'])
        export_win.resizable(False, False)
        export_win.transient(self.root)
        export_win.grab_set()

        tk.Label(export_win, text="选择导出范围", font=("Microsoft YaHei", 13, "bold"), bg=COLORS['bg'], fg=COLORS['dark']).pack(pady=(25, 15))
        export_var = tk.StringVar(value='current')
        for text, val in [("仅当前人员", 'current'), ("所有人员", 'all')]:
            tk.Radiobutton(export_win, text=text, variable=export_var, value=val,
                           font=("Microsoft YaHei", 11), bg=COLORS['bg'], fg=COLORS['text'], selectcolor=COLORS['card'],
                           activebackground=COLORS['bg'], activeforeground=COLORS['primary']).pack(anchor=tk.W, padx=70, pady=4)

        def do_export():
            wb = openpyxl.Workbook(); wb.remove(wb.active)
            h_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
            h_fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
            cell_font = Font(name='Microsoft YaHei', size=10)
            center = Alignment(horizontal='center', vertical='center')
            bdr = Border(left=Side('thin', color='D1D5DB'), right=Side('thin', color='D1D5DB'),
                         top=Side('thin', color='D1D5DB'), bottom=Side('thin', color='D1D5DB'))
            warn_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            good_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

            persons = [self.current_person] if export_var.get() == 'current' and self.current_person else sorted(self.data.keys())
            for p in persons:
                if p not in self.data or not self.data[p]: continue
                ws = wb.create_sheet(p[:31])
                ws.merge_cells('A1:G1')
                ws['A1'].value = f"{p} 的健康数据"; ws['A1'].font = Font(name='Microsoft YaHei', bold=True, size=14, color=COLORS['dark'])
                ws['A1'].alignment = center; ws.row_dimensions[1].height = 35
                
                headers = ["状态", "时间", "心率", "收缩压", "舒张压", "血氧(%)", "备注"]
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(row=2, column=c, value=h); cell.font = h_font; cell.fill = h_fill; cell.alignment = center; cell.border = bdr
                
                for r_idx, rec in enumerate(self.data[p], 3):
                    is_ok = self._is_healthy(rec)
                    vals = ["健康" if is_ok else "异常", rec.get('time',''), rec.get('heart_rate',''), rec.get('systolic',''), rec.get('diastolic',''), rec.get('spo2',''), rec.get('note','')]
                    for c_idx, v in enumerate(vals, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=v); cell.font = cell_font; cell.alignment = center; cell.border = bdr
                        if c_idx == 1: cell.fill = good_fill if is_ok else warn_fill

                widths = [8, 22, 12, 12, 12, 12, 25]
                for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

            path = filedialog.asksaveasfilename(parent=export_win, defaultextension='.xlsx', initialfile=f"健康数据_{datetime.datetime.now().strftime('%m%d_%H%M')}.xlsx", filetypes=[('Excel', '*.xlsx')])
            if path:
                wb.save(path); export_win.destroy()
                messagebox.showinfo("导出成功", f"✅ 已导出至:\n{path}")

        btn_frame = tk.Frame(export_win, bg=COLORS['bg'])
        btn_frame.pack(pady=20)
        self._make_button(btn_frame, "📤 导出", COLORS['success'], COLORS['success_h'], do_export, width=14).pack(side=tk.LEFT, padx=12)
        self._make_button(btn_frame, "❌ 取消", COLORS['danger'], COLORS['danger_h'], export_win.destroy, width=14).pack(side=tk.LEFT, padx=12)

def main():
    root = tk.Tk()
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except: pass
    app = HealthRecordApp(root)
    root.update_idletasks()
    root.geometry(f"+{(root.winfo_screenwidth() - 1150)//2}+{(root.winfo_screenheight() - 780)//2}")
    root.mainloop()

if __name__ == "__main__":
    main()
