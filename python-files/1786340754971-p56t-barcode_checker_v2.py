import sys
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox
import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None

APP_NAME = "条码核对"
DATA_FILE = "条码数据.xlsx"

BG = "#F4F6F8"
CARD = "#FFFFFF"
TEXT = "#1F2937"
SUBTEXT = "#6B7280"
BLUE = "#2563EB"
GREEN = "#16A34A"
RED = "#DC2626"
YELLOW = "#D97706"
BORDER = "#E5E7EB"

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("条码核对工具")
        self.root.geometry("980x680")
        self.root.minsize(850, 600)
        self.root.configure(bg=BG)

        self.groups = {}
        self.current_group = None
        self.expected = set()
        self.scanned = set()

        self.input_var = tk.StringVar()
        self.group_var = tk.StringVar(value="等待扫描 A 组条码")
        self.count_var = tk.StringVar(value="0 / 0")
        self.status_var = tk.StringVar(value="准备就绪")
        self.detail_var = tk.StringVar(value="请扫描 A 组条码开始核对")
        self.time_var = tk.StringVar()
        self.today_done = 0
        self.today_errors = 0

        self.build()
        self.load_data()
        self.update_clock()
        self.root.after(100, self.focus_scan)

    def build(self):
        # 顶部栏
        header = tk.Frame(self.root, bg="#111827", height=76)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(header, text="▣  条码核对", bg="#111827", fg="white",
                 font=("Microsoft YaHei UI", 22, "bold")).pack(side="left", padx=28)
        tk.Label(header, text="LOCAL  /  OFFLINE", bg="#111827", fg="#9CA3AF",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=8)

        self.time_label = tk.Label(header, textvariable=self.time_var, bg="#111827",
                                   fg="#D1D5DB", font=("Segoe UI", 10))
        self.time_label.pack(side="right", padx=28)

        # 主区
        main = tk.Frame(self.root, bg=BG)
        main.pack(fill="both", expand=True, padx=24, pady=20)

        # 左侧信息
        left = tk.Frame(main, bg=BG, width=600)
        left.pack(side="left", fill="both", expand=True, padx=(0, 10))
        left.pack_propagate(False)

        # 当前组卡片
        group_card = tk.Frame(left, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        group_card.pack(fill="x", pady=(0, 12))

        tk.Label(group_card, text="当前数据组", bg=CARD, fg=SUBTEXT,
                 font=("Microsoft YaHei UI", 11)).pack(anchor="w", padx=22, pady=(18, 2))
        tk.Label(group_card, textvariable=self.group_var, bg=CARD, fg=TEXT,
                 font=("Microsoft YaHei UI", 24, "bold")).pack(anchor="w", padx=22, pady=(0, 18))

        # 进度
        progress_card = tk.Frame(left, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        progress_card.pack(fill="x", pady=(0, 12))

        row = tk.Frame(progress_card, bg=CARD)
        row.pack(fill="x", padx=22, pady=(18, 8))
        tk.Label(row, text="核对进度", bg=CARD, fg=SUBTEXT,
                 font=("Microsoft YaHei UI", 11)).pack(side="left")
        tk.Label(row, textvariable=self.count_var, bg=CARD, fg=BLUE,
                 font=("Segoe UI", 16, "bold")).pack(side="right")

        self.progress = ttk.Progressbar(progress_card, mode="determinate", maximum=100)
        self.progress.pack(fill="x", padx=22, pady=(0, 20), ipady=4)

        # 状态卡
        self.status_card = tk.Frame(left, bg="#EFF6FF", highlightbackground="#DBEAFE", highlightthickness=1)
        self.status_card.pack(fill="x", pady=(0, 12))
        self.status_title = tk.Label(self.status_card, textvariable=self.status_var,
                                     bg="#EFF6FF", fg=BLUE, font=("Microsoft YaHei UI", 25, "bold"))
        self.status_title.pack(anchor="w", padx=22, pady=(20, 4))
        self.status_detail = tk.Label(self.status_card, textvariable=self.detail_var,
                                      bg="#EFF6FF", fg="#4B5563",
                                      font=("Microsoft YaHei UI", 12), wraplength=560, justify="left")
        self.status_detail.pack(anchor="w", padx=22, pady=(0, 20))

        # 扫描输入
        scan = tk.Frame(left, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        scan.pack(fill="x")

        tk.Label(scan, text="扫描输入", bg=CARD, fg=SUBTEXT,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", padx=22, pady=(15, 5))
        self.entry = tk.Entry(scan, textvariable=self.input_var, font=("Consolas", 18),
                              bg="#F9FAFB", fg=TEXT, relief="flat",
                              highlightthickness=1, highlightbackground=BORDER,
                              highlightcolor=BLUE, insertbackground=TEXT)
        self.entry.pack(fill="x", padx=22, pady=(0, 18), ipady=9)
        self.entry.bind("<Return>", self.scan)

        # 右侧
        right = tk.Frame(main, bg=BG, width=310)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        stats = tk.Frame(right, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        stats.pack(fill="x", pady=(0, 12))
        tk.Label(stats, text="今日状态", bg=CARD, fg=TEXT,
                 font=("Microsoft YaHei UI", 15, "bold")).pack(anchor="w", padx=20, pady=(18, 15))

        self.done_label = tk.Label(stats, text="0", bg=CARD, fg=GREEN,
                                   font=("Segoe UI", 30, "bold"))
        self.done_label.pack(anchor="w", padx=20)
        tk.Label(stats, text="完成组", bg=CARD, fg=SUBTEXT,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", padx=20, pady=(0, 15))

        self.error_label = tk.Label(stats, text="0", bg=CARD, fg=RED,
                                    font=("Segoe UI", 30, "bold"))
        self.error_label.pack(anchor="w", padx=20)
        tk.Label(stats, text="错误次数", bg=CARD, fg=SUBTEXT,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", padx=20, pady=(0, 20))

        list_card = tk.Frame(right, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        list_card.pack(fill="both", expand=True)

        tk.Label(list_card, text="当前组条码", bg=CARD, fg=TEXT,
                 font=("Microsoft YaHei UI", 15, "bold")).pack(anchor="w", padx=20, pady=(18, 10))

        self.listbox = tk.Listbox(list_card, bg=CARD, fg=TEXT, relief="flat",
                                  font=("Consolas", 12), activestyle="none",
                                  selectbackground="#DBEAFE", selectforeground=TEXT,
                                  highlightthickness=0)
        self.listbox.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        btns = tk.Frame(list_card, bg=CARD)
        btns.pack(fill="x", padx=15, pady=(0, 15))
        self.button(btns, "重新加载 Excel", self.reload, BLUE).pack(fill="x", pady=3)
        self.button(btns, "清空当前组", self.reset, "#6B7280").pack(fill="x", pady=3)

    def button(self, parent, text, command, color):
        b = tk.Button(parent, text=text, command=command, bg=color, fg="white",
                      activebackground=color, activeforeground="white",
                      relief="flat", bd=0, font=("Microsoft YaHei UI", 10, "bold"),
                      cursor="hand2", pady=9)
        return b

    def data_path(self):
        return Path(sys.executable).parent / DATA_FILE if getattr(sys, "frozen", False) else Path(__file__).parent / DATA_FILE

    def load_data(self):
        self.groups.clear()
        if openpyxl is None:
            self.set_status("读取失败", "error", "当前环境缺少 openpyxl")
            return
        path = self.data_path()
        if not path.exists():
            self.set_status("找不到数据文件", "error", f"请把 {DATA_FILE} 放在 EXE 同一文件夹")
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb["条码数据"] if "条码数据" in wb.sheetnames else wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            h = [str(x).strip() if x is not None else "" for x in header]
            a_idx = h.index("A组条码") if "A组条码" in h else 0
            c_idx = h.index("子条码") if "子条码" in h else 1
            for row in rows:
                if len(row) <= max(a_idx, c_idx): continue
                a = "" if row[a_idx] is None else str(row[a_idx]).strip()
                c = "" if row[c_idx] is None else str(row[c_idx]).strip()
                if a and c:
                    self.groups.setdefault(a, set()).add(c)
            self.set_status("准备就绪", "normal", f"已读取 {len(self.groups)} 个数据组，请扫描 A 组条码")
        except Exception as e:
            self.set_status("Excel读取失败", "error", str(e))

    def scan(self, event=None):
        code = self.input_var.get().strip()
        self.input_var.set("")
        self.focus_scan()
        if not code: return

        if self.current_group is None:
            if code in self.groups:
                self.current_group = code
                self.expected = set(self.groups[code])
                self.scanned = set()
                self.group_var.set(code)
                self.update_progress()
                self.refresh_list()
                self.set_status("开始核对", "normal", f"当前组共有 {len(self.expected)} 个子条码，请继续扫描")
                self.beep("normal")
            else:
                self.today_errors += 1
                self.error_label.config(text=str(self.today_errors))
                self.set_status("✕ 不是有效 A 组", "error", f"扫描到：{code}，数据表中没有这个 A 组")
                self.beep("error")
            return

        if code in self.scanned:
            self.today_errors += 1
            self.error_label.config(text=str(self.today_errors))
            self.set_status("⚠ 重复扫描", "warning", f"{code} 已经核对过，不需要再次扫描")
            self.beep("warning")
            return

        if code not in self.expected:
            self.today_errors += 1
            self.error_label.config(text=str(self.today_errors))
            self.set_status("✕ 扫描错误", "error", f"{code} 不属于当前组 {self.current_group}，当前进度不会受到影响")
            self.beep("error")
            return

        self.scanned.add(code)
        self.update_progress()
        self.refresh_list()
        self.set_status("✓ 扫描正确", "success", f"{code} 属于 {self.current_group}")
        self.beep("success")

        if self.scanned == self.expected:
            self.today_done += 1
            self.done_label.config(text=str(self.today_done))
            self.set_status("✓ 核对完成", "success", f"{self.current_group} 的全部 {len(self.expected)} 个子条码已经核对完成")
            self.beep("complete")
            self.root.after(1400, self.next_group)

    def next_group(self):
        self.current_group = None
        self.expected = set()
        self.scanned = set()
        self.group_var.set("等待扫描 A 组条码")
        self.count_var.set("0 / 0")
        self.progress["value"] = 0
        self.refresh_list()
        self.set_status("准备下一组", "normal", "请扫描下一个 A 组条码")
        self.focus_scan()

    def refresh_list(self):
        self.listbox.delete(0, tk.END)
        for code in sorted(self.expected):
            mark = "✓" if code in self.scanned else "○"
            self.listbox.insert(tk.END, f"{mark}  {code}")

    def update_progress(self):
        n = len(self.scanned)
        total = len(self.expected)
        self.count_var.set(f"{n} / {total}")
        self.progress["value"] = (n / total * 100) if total else 0

    def reset(self):
        if self.current_group is None:
            self.set_status("没有正在核对的组", "warning", "请先扫描 A 组条码")
            return
        self.scanned.clear()
        self.update_progress()
        self.refresh_list()
        self.set_status("已清空当前组", "warning", f"{self.current_group} 可以重新开始扫描")
        self.focus_scan()

    def reload(self):
        self.current_group = None
        self.expected = set()
        self.scanned = set()
        self.group_var.set("等待扫描 A 组条码")
        self.count_var.set("0 / 0")
        self.progress["value"] = 0
        self.refresh_list()
        self.load_data()
        self.focus_scan()

    def set_status(self, title, kind, detail):
        self.status_var.set(title)
        self.detail_var.set(detail)
        if kind == "success":
            bg, border, fg = "#ECFDF5", "#BBF7D0", GREEN
        elif kind == "error":
            bg, border, fg = "#FEF2F2", "#FECACA", RED
        elif kind == "warning":
            bg, border, fg = "#FFFBEB", "#FDE68A", YELLOW
        else:
            bg, border, fg = "#EFF6FF", "#DBEAFE", BLUE
        self.status_card.config(bg=bg, highlightbackground=border)
        self.status_title.config(bg=bg, fg=fg)
        self.status_detail.config(bg=bg)

    def beep(self, kind):
        # Windows系统音：不同状态使用不同频率/次数
        try:
            import winsound
            if kind == "success":
                winsound.Beep(1200, 90)
            elif kind == "error":
                winsound.Beep(500, 180)
                winsound.Beep(350, 220)
            elif kind == "warning":
                winsound.Beep(750, 100)
                winsound.Beep(750, 100)
            elif kind == "complete":
                winsound.Beep(900, 100)
                winsound.Beep(1200, 120)
                winsound.Beep(1500, 160)
            else:
                winsound.Beep(800, 80)
        except Exception:
            self.root.bell()

    def focus_scan(self):
        self.entry.focus_set()
        self.entry.selection_range(0, tk.END)

    def update_clock(self):
        self.time_var.set(datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))
        self.root.after(1000, self.update_clock)

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
