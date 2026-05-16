from __future__ import annotations

import csv
import json
import os
import platform
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except Exception:
    TkinterDnD = tk  # type: ignore
    DND_FILES = None


APP_NAME = "ISMS Secure Asset Command Center"
APP_SLUG = "ISMS_SACCC"

BG = "#07111d"
PANEL = "#0d1b2a"
PANEL_2 = "#10263d"
BORDER = "#1f3550"
TEXT = "#e6f1ff"
MUTED = "#95a4bd"
ACCENT = "#00d4ff"
ACCENT_2 = "#8b5cf6"
SUCCESS = "#22c55e"
WARN = "#f59e0b"
DANGER = "#ef4444"

CANONICAL_FIELDS = [
    "hostname","employee_code","employee_name","imei","region","employee_department",
    "bitlocker_status","bitlocker_key_c","bitlocker_key_d","ram","storage",
    "windows_version","edr_type","admin_password","asset_status","remarks",
    "current_holder","current_location","lifecycle_stage","last_updated_by","last_updated_at",
]

ALIASES = {
    "hostname": {"hostname", "host name", "device name", "computer name", "asset hostname"},
    "employee_code": {"employee code", "emp code", "empid", "employeeid", "code"},
    "employee_name": {"employee name", "name", "full name"},
    "imei": {"imei", "serial", "serial number"},
    "region": {"region"},
    "employee_department": {"employee department", "department", "dept"},
    "bitlocker_status": {"bitlocker status", "bitlocker", "encryption"},
    "bitlocker_key_c": {"bitlocker key of c drive", "bitlocker key c", "c drive key"},
    "bitlocker_key_d": {"bitlocker key of d drive", "bitlocker key d", "d drive key"},
    "ram": {"ram", "memory"},
    "storage": {"storage", "disk", "ssd"},
    "windows_version": {"windows version", "os version"},
    "edr_type": {"edr type", "edr", "endpoint protection"},
    "admin_password": {"admin password", "administrator password", "local admin password"},
    "asset_status": {"asset status", "status"},
    "remarks": {"remarks", "notes", "note"},
    "current_holder": {"current holder", "holder"},
    "current_location": {"current location", "location"},
    "lifecycle_stage": {"lifecycle stage", "stage"},
    "last_updated_by": {"last updated by"},
    "last_updated_at": {"last updated at"},
    "manager": {"manager"},
    "left_date": {"left date"},
    "left_reason": {"left reason"},
    "status": {"status"},
    "assigned_assets": {"assigned assets"},
    "exit_history": {"exit history"},
}

REQUIRED_IMPORT_FIELDS = {"hostname", "employee_code", "employee_name"}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_json(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return norm_text(value)


def app_data_root() -> Path:
    if platform.system().lower().startswith("win"):
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
    else:
        base = Path.home() / ".local" / "share"
    return base / APP_SLUG


@dataclass(frozen=True)
class Paths:
    root: Path
    data_dir: Path
    backup_dir: Path
    import_dir: Path
    db_path: Path
    config_path: Path


def resolve_paths() -> Paths:
    root = app_data_root()
    data_dir = root / "data"
    backup_dir = root / "backups"
    import_dir = root / "imports"
    for p in (root, data_dir, backup_dir, import_dir):
        p.mkdir(parents=True, exist_ok=True)
    return Paths(root=root, data_dir=data_dir, backup_dir=backup_dir, import_dir=import_dir,
                 db_path=data_dir / "isms_saccc.sqlite", config_path=data_dir / "config.json")


SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS employees (
    employee_code TEXT PRIMARY KEY,
    employee_name TEXT NOT NULL,
    department TEXT,
    manager TEXT,
    region TEXT,
    status TEXT NOT NULL DEFAULT 'active',
    left_date TEXT,
    left_reason TEXT,
    assigned_assets INTEGER NOT NULL DEFAULT 0,
    exit_history TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS assets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    hostname TEXT,
    employee_code TEXT,
    employee_name TEXT,
    imei TEXT,
    region TEXT,
    employee_department TEXT,
    bitlocker_status TEXT,
    bitlocker_key_c TEXT,
    bitlocker_key_d TEXT,
    ram TEXT,
    storage TEXT,
    windows_version TEXT,
    edr_type TEXT,
    admin_password TEXT,
    asset_status TEXT,
    remarks TEXT,
    current_holder TEXT,
    current_location TEXT,
    lifecycle_stage TEXT,
    last_updated_by TEXT,
    last_updated_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE UNIQUE INDEX IF NOT EXISTS idx_assets_hostname_unique
ON assets(hostname)
WHERE hostname IS NOT NULL AND hostname <> '';

CREATE UNIQUE INDEX IF NOT EXISTS idx_assets_imei_unique
ON assets(imei)
WHERE imei IS NOT NULL AND imei <> '';

CREATE TABLE IF NOT EXISTS movements (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    asset_id INTEGER NOT NULL,
    from_location TEXT,
    to_location TEXT,
    from_holder TEXT,
    to_holder TEXT,
    reason TEXT,
    timestamp TEXT NOT NULL,
    changed_by TEXT,
    FOREIGN KEY(asset_id) REFERENCES assets(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS changes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    entity_type TEXT NOT NULL,
    entity_id TEXT NOT NULL,
    field_name TEXT NOT NULL,
    old_value TEXT,
    new_value TEXT,
    changed_by TEXT,
    timestamp TEXT NOT NULL,
    note TEXT
);

CREATE TABLE IF NOT EXISTS imports (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    filename TEXT NOT NULL,
    imported_at TEXT NOT NULL,
    rows_total INTEGER NOT NULL,
    rows_inserted INTEGER NOT NULL,
    rows_updated INTEGER NOT NULL,
    duplicates INTEGER NOT NULL,
    invalid_rows INTEGER NOT NULL
);
"""


class Database:
    def __init__(self, path: Path):
        self.path = path
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA journal_mode=WAL;")
        self.conn.execute("PRAGMA synchronous=NORMAL;")
        self.conn.execute("PRAGMA foreign_keys=ON;")

    def close(self):
        try:
            self.conn.commit()
        finally:
            self.conn.close()

    def initialize(self):
        self.conn.executescript(SCHEMA_SQL)
        self.conn.commit()

    def set_setting(self, key: str, value: str):
        self.conn.execute(
            "INSERT INTO settings(key, value) VALUES(?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
        )
        self.conn.commit()

    def get_setting(self, key: str, default: str | None = None) -> str | None:
        row = self.conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row[0] if row else default

    def mark_setup_done(self):
        self.set_setting("setup_done", "1")

    def needs_initial_import(self) -> bool:
        count_assets = self.conn.execute("SELECT COUNT(*) FROM assets").fetchone()[0]
        count_employees = self.conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        return self.get_setting("setup_done", "0") != "1" and count_assets == 0 and count_employees == 0

    def _log_changes(self, entity_type: str, entity_id: str, old: dict[str, Any], new: dict[str, Any], changed_by: str, note: str = ""):
        keys = sorted(set(old) | set(new))
        for key in keys:
            old_value = safe_json(old.get(key))
            new_value = safe_json(new.get(key))
            if old_value != new_value:
                self.conn.execute(
                    "INSERT INTO changes(entity_type, entity_id, field_name, old_value, new_value, changed_by, timestamp, note) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (entity_type, entity_id, key, old_value, new_value, changed_by, now_iso(), note),
                )
        self.conn.commit()

    def create_change(self, entity_type: str, entity_id: str, field_name: str, old_value: Any, new_value: Any, changed_by: str, note: str = ""):
        self.conn.execute(
            "INSERT INTO changes(entity_type, entity_id, field_name, old_value, new_value, changed_by, timestamp, note) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (entity_type, entity_id, field_name, safe_json(old_value), safe_json(new_value), changed_by, now_iso(), note),
        )
        self.conn.commit()

    def _clean_asset(self, payload: dict[str, Any]) -> dict[str, Any]:
        data = {field: norm_text(payload.get(field)) for field in CANONICAL_FIELDS}
        data.setdefault("asset_status", "active")
        data.setdefault("lifecycle_stage", "inventory")
        data.setdefault("current_location", data.get("current_location") or "Inventory")
        data.setdefault("current_holder", data.get("current_holder") or "")
        data.setdefault("last_updated_by", payload.get("last_updated_by") or "system")
        data["last_updated_at"] = payload.get("last_updated_at") or now_iso()
        return data

    def find_asset_by_identity(self, hostname: str = "", imei: str = "", employee_code: str = ""):
        clauses = []
        params = []
        if hostname:
            clauses.append("hostname=?")
            params.append(hostname)
        if imei:
            clauses.append("imei=?")
            params.append(imei)
        if employee_code:
            clauses.append("employee_code=?")
            params.append(employee_code)
        if not clauses:
            return None
        return self.conn.execute(f"SELECT * FROM assets WHERE {' OR '.join(clauses)} LIMIT 1", params).fetchone()

    def upsert_asset(self, payload: dict[str, Any], changed_by: str = "system") -> int:
        now = now_iso()
        data = self._clean_asset(payload)
        existing = self.find_asset_by_identity(data.get("hostname"), data.get("imei"), data.get("employee_code"))
        if existing:
            asset_id = int(existing["id"])
            old = dict(existing)
            self.conn.execute(
                """
                UPDATE assets SET
                    hostname=:hostname, employee_code=:employee_code, employee_name=:employee_name, imei=:imei,
                    region=:region, employee_department=:employee_department, bitlocker_status=:bitlocker_status,
                    bitlocker_key_c=:bitlocker_key_c, bitlocker_key_d=:bitlocker_key_d, ram=:ram, storage=:storage,
                    windows_version=:windows_version, edr_type=:edr_type, admin_password=:admin_password,
                    asset_status=:asset_status, remarks=:remarks, current_holder=:current_holder,
                    current_location=:current_location, lifecycle_stage=:lifecycle_stage,
                    last_updated_by=:last_updated_by, last_updated_at=:last_updated_at, updated_at=:updated_at
                WHERE id=:id
                """,
                {**data, "id": asset_id, "updated_at": now},
            )
            new = dict(old)
            new.update(data)
            new["updated_at"] = now
            self._log_changes("asset", str(asset_id), old, new, changed_by, note="Asset update")
        else:
            self.conn.execute(
                """
                INSERT INTO assets(hostname, employee_code, employee_name, imei, region, employee_department,
                    bitlocker_status, bitlocker_key_c, bitlocker_key_d, ram, storage, windows_version,
                    edr_type, admin_password, asset_status, remarks, current_holder, current_location,
                    lifecycle_stage, last_updated_by, last_updated_at, created_at, updated_at)
                VALUES (:hostname, :employee_code, :employee_name, :imei, :region, :employee_department,
                    :bitlocker_status, :bitlocker_key_c, :bitlocker_key_d, :ram, :storage, :windows_version,
                    :edr_type, :admin_password, :asset_status, :remarks, :current_holder, :current_location,
                    :lifecycle_stage, :last_updated_by, :last_updated_at, :created_at, :updated_at)
                """,
                {**data, "created_at": now, "updated_at": now},
            )
            asset_id = int(self.conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            self._log_changes("asset", str(asset_id), {}, data, changed_by, note="Asset created")
        self.conn.commit()
        return asset_id

    def upsert_employee(self, payload: dict[str, Any]):
        now = now_iso()
        code = norm_text(payload.get("employee_code"))
        if not code:
            return
        existing = self.conn.execute("SELECT * FROM employees WHERE employee_code=?", (code,)).fetchone()
        data = {
            "employee_code": code,
            "employee_name": norm_text(payload.get("employee_name")),
            "department": norm_text(payload.get("department") or payload.get("employee_department")),
            "manager": norm_text(payload.get("manager")),
            "region": norm_text(payload.get("region")),
            "status": norm_text(payload.get("status") or "active").lower(),
            "left_date": norm_text(payload.get("left_date")),
            "left_reason": norm_text(payload.get("left_reason")),
            "assigned_assets": int(payload.get("assigned_assets") or 0),
            "exit_history": norm_text(payload.get("exit_history")),
            "updated_at": now,
        }
        if existing:
            old = dict(existing)
            self.conn.execute(
                """
                UPDATE employees SET employee_name=:employee_name, department=:department, manager=:manager,
                region=:region, status=:status, left_date=:left_date, left_reason=:left_reason,
                assigned_assets=:assigned_assets, exit_history=:exit_history, updated_at=:updated_at
                WHERE employee_code=:employee_code
                """,
                data,
            )
            new = dict(old)
            new.update(data)
            self._log_changes("employee", code, old, new, payload.get("changed_by", "system"), note="Employee update")
        else:
            self.conn.execute(
                """
                INSERT INTO employees(employee_code, employee_name, department, manager, region, status,
                    left_date, left_reason, assigned_assets, exit_history, created_at, updated_at)
                VALUES (:employee_code, :employee_name, :department, :manager, :region, :status,
                    :left_date, :left_reason, :assigned_assets, :exit_history, :created_at, :updated_at)
                """,
                {**data, "created_at": now},
            )
            self._log_changes("employee", code, {}, data, payload.get("changed_by", "system"), note="Employee created")
        self.conn.commit()

    def list_assets(self, search: str = "", status: str = "", region: str = "", lifecycle: str = "") -> list[dict[str, Any]]:
        sql = "SELECT * FROM assets WHERE 1=1"
        params: list[Any] = []
        if search:
            sql += """ AND (
                hostname LIKE ? OR employee_code LIKE ? OR employee_name LIKE ? OR imei LIKE ? OR current_holder LIKE ?
                OR current_location LIKE ? OR asset_status LIKE ? OR lifecycle_stage LIKE ?
            )"""
            t = f"%{search}%"
            params.extend([t] * 8)
        if status:
            sql += " AND COALESCE(asset_status,'')=?"
            params.append(status)
        if region:
            sql += " AND COALESCE(region,'')=?"
            params.append(region)
        if lifecycle:
            sql += " AND COALESCE(lifecycle_stage,'')=?"
            params.append(lifecycle)
        sql += " ORDER BY datetime(last_updated_at) DESC, id DESC"
        return [dict(r) for r in self.conn.execute(sql, params).fetchall()]

    def list_employees(self, search: str = "", status: str = "") -> list[dict[str, Any]]:
        sql = "SELECT * FROM employees WHERE 1=1"
        params = []
        if search:
            sql += """ AND (
                employee_code LIKE ? OR employee_name LIKE ? OR department LIKE ? OR manager LIKE ? OR region LIKE ?
            )"""
            t = f"%{search}%"
            params.extend([t] * 5)
        if status:
            sql += " AND COALESCE(status,'')=?"
            params.append(status)
        sql += " ORDER BY datetime(updated_at) DESC, employee_name ASC"
        return [dict(r) for r in self.conn.execute(sql, params).fetchall()]

    def asset_detail(self, asset_id: int):
        r = self.conn.execute("SELECT * FROM assets WHERE id=?", (asset_id,)).fetchone()
        return dict(r) if r else None

    def employee_detail(self, employee_code: str):
        r = self.conn.execute("SELECT * FROM employees WHERE employee_code=?", (employee_code,)).fetchone()
        return dict(r) if r else None

    def movements(self, limit: int = 300) -> list[dict[str, Any]]:
        return [dict(r) for r in self.conn.execute(
            """SELECT m.*, a.hostname, a.employee_code, a.employee_name
               FROM movements m LEFT JOIN assets a ON a.id=m.asset_id
               ORDER BY datetime(m.timestamp) DESC, m.id DESC LIMIT ?""", (limit,)
        ).fetchall()]

    def changes(self, limit: int = 500) -> list[dict[str, Any]]:
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM changes ORDER BY datetime(timestamp) DESC, id DESC LIMIT ?", (limit,)
        ).fetchall()]

    def asset_movements(self, asset_id: int) -> list[dict[str, Any]]:
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM movements WHERE asset_id=? ORDER BY datetime(timestamp) DESC, id DESC", (asset_id,)
        ).fetchall()]

    def asset_changes(self, asset_id: int) -> list[dict[str, Any]]:
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM changes WHERE entity_type='asset' AND entity_id=? ORDER BY datetime(timestamp) DESC, id DESC", (str(asset_id),)
        ).fetchall()]

    def employee_changes(self, employee_code: str) -> list[dict[str, Any]]:
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM changes WHERE entity_type='employee' AND entity_id=? ORDER BY datetime(timestamp) DESC, id DESC", (str(employee_code),)
        ).fetchall()]

    def create_movement(self, payload: dict[str, Any]):
        now = now_iso()
        self.conn.execute(
            """INSERT INTO movements(asset_id, from_location, to_location, from_holder, to_holder, reason, timestamp, changed_by)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                payload["asset_id"], payload.get("from_location"), payload.get("to_location"),
                payload.get("from_holder"), payload.get("to_holder"), payload.get("reason"), now, payload.get("changed_by") or "system"
            )
        )
        self.conn.execute(
            """UPDATE assets SET current_location=?, current_holder=?, lifecycle_stage=?, asset_status=?,
               last_updated_by=?, last_updated_at=?, updated_at=? WHERE id=?""",
            (
                payload.get("to_location"), payload.get("to_holder"), payload.get("lifecycle_stage") or payload.get("asset_status") or "assigned",
                payload.get("asset_status") or "active", payload.get("changed_by") or "system", now, now, payload["asset_id"]
            )
        )
        self.conn.commit()
        self._log_changes("asset", str(payload["asset_id"]),
                          {"current_location": payload.get("from_location"), "current_holder": payload.get("from_holder")},
                          {"current_location": payload.get("to_location"), "current_holder": payload.get("to_holder")},
                          payload.get("changed_by") or "system", note=payload.get("reason") or "Movement")

    def delete_asset(self, asset_id: int, changed_by: str = "system"):
        row = self.asset_detail(asset_id)
        if row:
            self.create_change("asset", str(asset_id), "__deleted__", json.dumps(row, default=str), "", changed_by, note="Asset deleted")
        self.conn.execute("DELETE FROM assets WHERE id=?", (asset_id,))
        self.conn.commit()

    def delete_employee(self, employee_code: str, changed_by: str = "system"):
        row = self.employee_detail(employee_code)
        if row:
            self.create_change("employee", employee_code, "__deleted__", json.dumps(row, default=str), "", changed_by, note="Employee deleted")
        self.conn.execute("DELETE FROM employees WHERE employee_code=?", (employee_code,))
        self.conn.commit()

    def counts(self) -> dict[str, int]:
        assets = self.conn.execute("SELECT * FROM assets").fetchall()
        employees = self.conn.execute("SELECT * FROM employees").fetchall()
        active = sum(1 for r in employees if norm_text(r["status"]).lower() == "active")
        exited = sum(1 for r in employees if norm_text(r["status"]).lower() != "active")
        bitlocker = sum(1 for r in assets if norm_text(r["bitlocker_status"]).lower() in {"on", "enabled", "compliant", "yes", "true"})
        edr = sum(1 for r in assets if norm_text(r["edr_type"]).lower() not in {"", "other", "none", "n/a"})
        missing = sum(1 for r in assets if norm_text(r["edr_type"]).lower() in {"", "other", "none", "n/a"})
        enc_fail = sum(1 for r in assets if norm_text(r["bitlocker_status"]).lower() in {"off", "failed", "no", "false"})
        alerts = sum(1 for r in assets if norm_text(r["asset_status"]).lower() in {"missing", "lost", "retired", "disposed"}) + enc_fail + missing
        return {
            "total_assets": len(assets),
            "active_employees": active,
            "exited_employees": exited,
            "bitlocker_compliance": bitlocker,
            "edr_coverage": edr,
            "missing_edr": missing,
            "encryption_failures": enc_fail,
            "alerts": alerts,
            "recent_changes": len(self.changes(20)),
            "recent_movements": len(self.movements(20)),
        }

    def trends(self) -> dict[str, list[tuple[str, int]]]:
        mv = self.conn.execute(
            "SELECT substr(timestamp,1,10) d, COUNT(*) c FROM movements GROUP BY d ORDER BY d DESC LIMIT 14"
        ).fetchall()
        ch = self.conn.execute(
            "SELECT substr(timestamp,1,10) d, COUNT(*) c FROM changes GROUP BY d ORDER BY d DESC LIMIT 14"
        ).fetchall()
        return {
            "movements": list(reversed([(r["d"], r["c"]) for r in mv])),
            "changes": list(reversed([(r["d"], r["c"]) for r in ch])),
        }

    def all_tables_export(self) -> dict[str, list[dict[str, Any]]]:
        return {
            "assets": [dict(r) for r in self.conn.execute("SELECT * FROM assets ORDER BY id")],
            "employees": [dict(r) for r in self.conn.execute("SELECT * FROM employees ORDER BY employee_name")],
            "movements": [dict(r) for r in self.conn.execute("SELECT * FROM movements ORDER BY id")],
            "changes": [dict(r) for r in self.conn.execute("SELECT * FROM changes ORDER BY id")],
            "imports": [dict(r) for r in self.conn.execute("SELECT * FROM imports ORDER BY id")],
            "settings": [dict(r) for r in self.conn.execute("SELECT * FROM settings ORDER BY key")],
        }

    def wipe_and_restore(self, data: dict[str, list[dict[str, Any]]]):
        self.conn.executescript("""
        DELETE FROM movements;
        DELETE FROM changes;
        DELETE FROM assets;
        DELETE FROM employees;
        DELETE FROM imports;
        DELETE FROM settings;
        """)
        self.conn.commit()
        for row in data.get("employees", []):
            self.conn.execute(
                """INSERT INTO employees(employee_code, employee_name, department, manager, region, status, left_date, left_reason,
                assigned_assets, exit_history, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    row.get("employee_code"), row.get("employee_name"), row.get("department"), row.get("manager"), row.get("region"),
                    row.get("status") or "active", row.get("left_date"), row.get("left_reason"), row.get("assigned_assets") or 0,
                    row.get("exit_history"), row.get("created_at") or now_iso(), row.get("updated_at") or now_iso()
                )
            )
        for row in data.get("assets", []):
            self.conn.execute(
                """INSERT INTO assets(id, hostname, employee_code, employee_name, imei, region, employee_department, bitlocker_status,
                bitlocker_key_c, bitlocker_key_d, ram, storage, windows_version, edr_type, admin_password, asset_status, remarks,
                current_holder, current_location, lifecycle_stage, last_updated_by, last_updated_at, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    row.get("id"), row.get("hostname"), row.get("employee_code"), row.get("employee_name"), row.get("imei"), row.get("region"),
                    row.get("employee_department"), row.get("bitlocker_status"), row.get("bitlocker_key_c"), row.get("bitlocker_key_d"), row.get("ram"),
                    row.get("storage"), row.get("windows_version"), row.get("edr_type"), row.get("admin_password"), row.get("asset_status"),
                    row.get("remarks"), row.get("current_holder"), row.get("current_location"), row.get("lifecycle_stage"), row.get("last_updated_by"),
                    row.get("last_updated_at"), row.get("created_at") or now_iso(), row.get("updated_at") or now_iso()
                )
            )
        for row in data.get("movements", []):
            self.conn.execute(
                "INSERT INTO movements(id, asset_id, from_location, to_location, from_holder, to_holder, reason, timestamp, changed_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (row.get("id"), row.get("asset_id"), row.get("from_location"), row.get("to_location"), row.get("from_holder"), row.get("to_holder"), row.get("reason"), row.get("timestamp") or now_iso(), row.get("changed_by"))
            )
        for row in data.get("changes", []):
            self.conn.execute(
                "INSERT INTO changes(id, entity_type, entity_id, field_name, old_value, new_value, changed_by, timestamp, note) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (row.get("id"), row.get("entity_type"), row.get("entity_id"), row.get("field_name"), row.get("old_value"), row.get("new_value"), row.get("changed_by"), row.get("timestamp") or now_iso(), row.get("note"))
            )
        for row in data.get("imports", []):
            self.conn.execute(
                "INSERT INTO imports(id, filename, imported_at, rows_total, rows_inserted, rows_updated, duplicates, invalid_rows) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (row.get("id"), row.get("filename"), row.get("imported_at") or now_iso(), row.get("rows_total") or 0, row.get("rows_inserted") or 0, row.get("rows_updated") or 0, row.get("duplicates") or 0, row.get("invalid_rows") or 0)
            )
        for row in data.get("settings", []):
            self.conn.execute("INSERT INTO settings(key, value) VALUES(?, ?)", (row.get("key"), row.get("value")))
        self.conn.commit()


def normalize_header(h: str) -> str:
    return re.sub(r"[\s\-_]+", " ", (h or "").strip().lower())


def auto_map_headers(headers: list[str]) -> dict[str, str]:
    header_map = {}
    normalized = {normalize_header(h): h for h in headers}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                header_map[normalized[alias]] = canonical
                break
    for h in headers:
        n = normalize_header(h)
        if n in CANONICAL_FIELDS or n in {"manager", "left_date", "left_reason", "status", "assigned_assets", "exit_history"}:
            header_map[h] = n
    return header_map


def parse_csv_file(path: Path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        mapping = auto_map_headers(headers)
        rows = []
        for raw in reader:
            row = {}
            for src, canonical in mapping.items():
                row[canonical] = norm_text(raw.get(src))
            rows.append(row)
        return rows, headers, mapping


def import_rows(db: Database, rows: list[dict[str, str]], source_name: str = "consolidated_import.csv", changed_by: str = "system") -> dict[str, int]:
    total = len(rows)
    inserted = 0
    updated = 0
    duplicates = 0
    invalid = 0
    seen = set()
    existing = {norm_text(v).lower() for r in db.list_assets() for v in (r.get("hostname"), r.get("imei"), r.get("employee_code")) if norm_text(v)}
    for row in rows:
        if not any(row.get(k) for k in REQUIRED_IMPORT_FIELDS):
            invalid += 1
            continue
        key = (row.get("hostname") or row.get("imei") or row.get("employee_code") or "").lower()
        if key in seen:
            duplicates += 1
            continue
        if key and key in existing:
            duplicates += 1
        seen.add(key)
        try:
            db.upsert_asset(row, changed_by=changed_by)
            db.upsert_employee({
                "employee_code": row.get("employee_code"),
                "employee_name": row.get("employee_name"),
                "department": row.get("employee_department"),
                "manager": row.get("manager"),
                "region": row.get("region"),
                "status": row.get("status") or "active",
                "left_date": row.get("left_date"),
                "left_reason": row.get("left_reason"),
                "assigned_assets": row.get("assigned_assets") or 0,
                "exit_history": row.get("exit_history") or "",
                "changed_by": changed_by,
            })
            inserted += 1
        except Exception:
            invalid += 1
    db.conn.execute(
        "INSERT INTO imports(filename, imported_at, rows_total, rows_inserted, rows_updated, duplicates, invalid_rows) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (source_name, now_iso(), total, inserted, updated, duplicates, invalid),
    )
    db.conn.commit()
    db.mark_setup_done()
    return {"rows_total": total, "rows_inserted": inserted, "rows_updated": updated, "duplicates": duplicates, "invalid_rows": invalid}


def export_csv(path: Path, rows: list[dict[str, Any]]):
    headers = list(rows[0].keys()) if rows else CANONICAL_FIELDS
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})


def export_json(path: Path, data: dict[str, Any]):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


def restore_json(path: Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def export_xlsx(path: Path, tables: dict[str, list[dict[str, Any]]]):
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    for sheet_name, rows in tables.items():
        ws = wb.create_sheet(sheet_name[:31])
        if not rows:
            ws.append(["No records"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        for col in ws.columns:
            letter = col[0].column_letter
            width = 0
            for cell in col:
                width = max(width, len(str(cell.value)) if cell.value is not None else 0)
            ws.column_dimensions[letter].width = min(width + 3, 42)
    wb.save(path)


class ToplevelDialog(tk.Toplevel):
    def __init__(self, parent, title: str, size: str):
        super().__init__(parent)
        self.title(title)
        self.geometry(size)
        self.configure(bg=BG)
        self.transient(parent)
        self.grab_set()
        self.result = None
        self._center(parent)
        self.protocol("WM_DELETE_WINDOW", self.close)

    def _center(self, parent):
        self.update_idletasks()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        px, py = parent.winfo_rootx(), parent.winfo_rooty()
        w, h = self.winfo_width(), self.winfo_height()
        self.geometry(f"+{px + (pw - w)//2}+{py + (ph - h)//2}")

    def close(self):
        try:
            self.grab_release()
        except Exception:
            pass
        self.destroy()


class AssetDialog(ToplevelDialog):
    FIELDS = [
        ("hostname", "Hostname"), ("employee_code", "Employee Code"), ("employee_name", "Employee Name"), ("imei", "IMEI"),
        ("region", "Region"), ("employee_department", "Employee Department"), ("bitlocker_status", "BitLocker Status"),
        ("bitlocker_key_c", "BitLocker Key C"), ("bitlocker_key_d", "BitLocker Key D"), ("ram", "RAM"), ("storage", "Storage"),
        ("windows_version", "Windows Version"), ("edr_type", "EDR Type"), ("admin_password", "Admin Password"),
        ("asset_status", "Asset Status"), ("remarks", "Remarks"), ("current_holder", "Current Holder"),
        ("current_location", "Current Location"), ("lifecycle_stage", "Lifecycle Stage"), ("last_updated_by", "Last Updated By"),
    ]
    def __init__(self, parent, asset: dict[str, Any] | None = None):
        super().__init__(parent, "Asset Detail", "1020x760")
        self.asset = asset or {}
        self.vars: dict[str, tk.StringVar] = {}
        self._build()

    def _build(self):
        head = tk.Frame(self, bg=BG)
        head.pack(fill="x", padx=18, pady=(14, 8))
        tk.Label(head, text="Asset Detail", fg=TEXT, bg=BG, font=("Segoe UI", 18, "bold")).pack(side="left")
        tk.Label(head, text="Edit, move, and review history", fg=MUTED, bg=BG, font=("Segoe UI", 10)).pack(side="left", padx=12)
        body = tk.Frame(self, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        body.pack(fill="both", expand=True, padx=18, pady=8)
        left = tk.Frame(body, bg=PANEL)
        left.pack(side="left", fill="both", expand=True)
        canvas = tk.Canvas(left, bg=PANEL, highlightthickness=0)
        scroll = ttk.Scrollbar(left, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=PANEL)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window = canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(window, width=e.width))
        for idx, (key, label) in enumerate(self.FIELDS):
            tk.Label(inner, text=label, fg=MUTED, bg=PANEL, anchor="w", font=("Segoe UI", 9, "bold")).grid(row=idx, column=0, sticky="ew", padx=14, pady=(10, 2))
            var = tk.StringVar(value=norm_text(self.asset.get(key)))
            self.vars[key] = var
            tk.Entry(inner, textvariable=var, bg="#091522", fg=TEXT, insertbackground=TEXT, relief="flat").grid(row=idx, column=1, sticky="ew", padx=14, pady=(10, 2), ipady=5)
        inner.columnconfigure(1, weight=1)
        actions = tk.Frame(self, bg=BG)
        actions.pack(fill="x", padx=18, pady=12)
        tk.Button(actions, text="Save", command=self.save, bg=ACCENT, fg="#00111d", activebackground="#6cecff", relief="flat", font=("Segoe UI", 10, "bold")).pack(side="right", padx=8)
        tk.Button(actions, text="Cancel", command=self.close, bg="#24364a", fg=TEXT, activebackground="#34495e", relief="flat").pack(side="right")

    def save(self):
        self.result = {k: v.get().strip() for k, v in self.vars.items()}
        self.result["last_updated_at"] = now_iso()
        self.close()


class EmployeeDialog(ToplevelDialog):
    FIELDS = [
        ("employee_code", "Employee Code"), ("employee_name", "Employee Name"), ("department", "Department"), ("manager", "Manager"),
        ("region", "Region"), ("status", "Status"), ("left_date", "Left Date"), ("left_reason", "Left Reason"),
        ("assigned_assets", "Assigned Assets"), ("exit_history", "Exit History"),
    ]
    def __init__(self, parent, employee: dict[str, Any] | None = None):
        super().__init__(parent, "Employee Detail", "860x620")
        self.employee = employee or {}
        self.vars = {}
        self._build()

    def _build(self):
        tk.Label(self, text="Employee Detail", fg=TEXT, bg=BG, font=("Segoe UI", 18, "bold")).pack(anchor="w", padx=18, pady=(14, 8))
        body = tk.Frame(self, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        body.pack(fill="both", expand=True, padx=18, pady=8)
        for idx, (key, label) in enumerate(self.FIELDS):
            tk.Label(body, text=label, fg=MUTED, bg=PANEL, anchor="w", font=("Segoe UI", 9, "bold")).grid(row=idx, column=0, sticky="ew", padx=14, pady=(10, 2))
            var = tk.StringVar(value=norm_text(self.employee.get(key)))
            self.vars[key] = var
            tk.Entry(body, textvariable=var, bg="#091522", fg=TEXT, insertbackground=TEXT, relief="flat").grid(row=idx, column=1, sticky="ew", padx=14, pady=(10, 2), ipady=5)
        body.columnconfigure(1, weight=1)
        actions = tk.Frame(self, bg=BG)
        actions.pack(fill="x", padx=18, pady=12)
        tk.Button(actions, text="Save", command=self.save, bg=ACCENT, fg="#00111d", activebackground="#6cecff", relief="flat", font=("Segoe UI", 10, "bold")).pack(side="right", padx=8)
        tk.Button(actions, text="Cancel", command=self.close, bg="#24364a", fg=TEXT, activebackground="#34495e", relief="flat").pack(side="right")

    def save(self):
        self.result = {k: v.get().strip() for k, v in self.vars.items()}
        self.close()


class MovementDialog(ToplevelDialog):
    def __init__(self, parent, asset: dict[str, Any]):
        super().__init__(parent, "Movement Workflow", "760x560")
        self.asset = asset
        self.vars = {}
        self._build()

    def _build(self):
        tk.Label(self, text=f"Move {self.asset.get('hostname') or self.asset.get('imei') or 'Asset'}", fg=TEXT, bg=BG, font=("Segoe UI", 18, "bold")).pack(anchor="w", padx=18, pady=(14, 8))
        body = tk.Frame(self, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        body.pack(fill="both", expand=True, padx=18, pady=8)
        items = [
            ("from_location", self.asset.get("current_location", "")),
            ("to_location", ""),
            ("from_holder", self.asset.get("current_holder", "")),
            ("to_holder", ""),
            ("reason", ""),
            ("changed_by", self.asset.get("last_updated_by", "system")),
            ("asset_status", self.asset.get("asset_status", "active")),
            ("lifecycle_stage", self.asset.get("lifecycle_stage", "assigned")),
        ]
        for idx, (key, default) in enumerate(items):
            tk.Label(body, text=key.replace("_", " ").title(), fg=MUTED, bg=PANEL, anchor="w", font=("Segoe UI", 9, "bold")).grid(row=idx, column=0, sticky="ew", padx=14, pady=(10, 2))
            var = tk.StringVar(value=default)
            self.vars[key] = var
            tk.Entry(body, textvariable=var, bg="#091522", fg=TEXT, insertbackground=TEXT, relief="flat").grid(row=idx, column=1, sticky="ew", padx=14, pady=(10, 2), ipady=5)
        body.columnconfigure(1, weight=1)
        actions = tk.Frame(self, bg=BG)
        actions.pack(fill="x", padx=18, pady=12)
        tk.Button(actions, text="Apply Movement", command=self.save, bg=ACCENT, fg="#00111d", activebackground="#6cecff", relief="flat", font=("Segoe UI", 10, "bold")).pack(side="right", padx=8)
        tk.Button(actions, text="Cancel", command=self.close, bg="#24364a", fg=TEXT, activebackground="#34495e", relief="flat").pack(side="right")

    def save(self):
        self.result = {k: v.get().strip() for k, v in self.vars.items()}
        self.close()


class ExitDialog(ToplevelDialog):
    def __init__(self, parent, employee: dict[str, Any]):
        super().__init__(parent, "Employee Exit Workflow", "760x420")
        self.employee = employee
        self.vars = {}
        self._build()

    def _build(self):
        tk.Label(self, text=f"Exit workflow: {self.employee.get('employee_name')}", fg=TEXT, bg=BG, font=("Segoe UI", 18, "bold")).pack(anchor="w", padx=18, pady=(14, 8))
        body = tk.Frame(self, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        body.pack(fill="both", expand=True, padx=18, pady=8)
        items = [("left_date", "Left Date"), ("left_reason", "Left Reason"), ("exit_history", "Exit History")]
        for idx, (key, label) in enumerate(items):
            tk.Label(body, text=label, fg=MUTED, bg=PANEL, anchor="w", font=("Segoe UI", 9, "bold")).grid(row=idx, column=0, sticky="ew", padx=14, pady=(10, 2))
            var = tk.StringVar()
            self.vars[key] = var
            tk.Entry(body, textvariable=var, bg="#091522", fg=TEXT, insertbackground=TEXT, relief="flat").grid(row=idx, column=1, sticky="ew", padx=14, pady=(10, 2), ipady=5)
        body.columnconfigure(1, weight=1)
        actions = tk.Frame(self, bg=BG)
        actions.pack(fill="x", padx=18, pady=12)
        tk.Button(actions, text="Mark Left", command=self.save, bg=ACCENT, fg="#00111d", activebackground="#6cecff", relief="flat", font=("Segoe UI", 10, "bold")).pack(side="right", padx=8)
        tk.Button(actions, text="Cancel", command=self.close, bg="#24364a", fg=TEXT, activebackground="#34495e", relief="flat").pack(side="right")

    def save(self):
        self.result = {k: v.get().strip() for k, v in self.vars.items()}
        self.close()


class ImportWizard(ToplevelDialog):
    def __init__(self, parent, db: Database, paths: Paths):
        super().__init__(parent, f"{APP_NAME} — First Launch Import", "1120x760")
        self.db = db
        self.paths = paths
        self.file_path: Path | None = None
        self.rows: list[dict[str, str]] = []
        self.headers: list[str] = []
        self.mapping: dict[str, str] = {}
        self.preview_rows: list[dict[str, str]] = []
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=BG)
        top.pack(fill="x", padx=18, pady=(14, 6))
        tk.Label(top, text="Import your consolidated CSV", bg=BG, fg=TEXT, font=("Segoe UI", 20, "bold")).pack(anchor="w")
        tk.Label(top, text="Drag and drop a CSV or browse to one. The import will validate rows, detect duplicates, and seed the local database.", bg=BG, fg=MUTED, font=("Segoe UI", 10)).pack(anchor="w", pady=(4, 0))
        self.drop = tk.Label(self, text="Drop CSV here", bg="#0a1522", fg=ACCENT, relief="solid", bd=1, font=("Segoe UI", 16, "bold"), height=3)
        self.drop.pack(fill="x", padx=18, pady=10)
        if DND_FILES is not None and hasattr(self, "drop_target_register"):
            try:
                self.drop.drop_target_register(DND_FILES)
                self.drop.dnd_bind("<<Drop>>", self.on_drop)
            except Exception:
                pass
        self.drop.bind("<Button-1>", lambda e: self.browse())
        controls = tk.Frame(self, bg=BG)
        controls.pack(fill="x", padx=18, pady=(0, 8))
        tk.Button(controls, text="Browse CSV", command=self.browse, bg="#24364a", fg=TEXT, relief="flat", activebackground="#34495e").pack(side="left")
        tk.Button(controls, text="Load Sample CSV", command=self.load_sample, bg="#24364a", fg=TEXT, relief="flat", activebackground="#34495e").pack(side="left", padx=8)
        self.file_label = tk.Label(controls, text="No file selected", bg=BG, fg=MUTED)
        self.file_label.pack(side="left", padx=12)
        split = tk.PanedWindow(self, orient="horizontal", sashrelief="flat", bg=BG, bd=0)
        split.pack(fill="both", expand=True, padx=18, pady=8)
        left = tk.Frame(split, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        right = tk.Frame(split, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        split.add(left, minsize=520)
        split.add(right, minsize=420)
        tk.Label(left, text="Preview", bg=PANEL, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=12, pady=(10, 4))
        self.tree = ttk.Treeview(left, show="headings", height=18)
        self.tree.pack(fill="both", expand=True, padx=12, pady=8)
        tk.Label(right, text="Import summary", bg=PANEL, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=12, pady=(10, 4))
        self.summary = tk.Text(right, bg="#08111b", fg=TEXT, insertbackground=TEXT, relief="flat", height=24, wrap="word")
        self.summary.pack(fill="both", expand=True, padx=12, pady=8)
        footer = tk.Frame(self, bg=BG)
        footer.pack(fill="x", padx=18, pady=12)
        tk.Button(footer, text="Run Import", command=self.run_import, bg=ACCENT, fg="#00111d", relief="flat", activebackground="#6cecff", font=("Segoe UI", 10, "bold")).pack(side="right", padx=8)
        tk.Button(footer, text="Continue without CSV", command=self.finish_empty, bg="#24364a", fg=TEXT, relief="flat", activebackground="#34495e").pack(side="right")

    def log(self, text: str):
        self.summary.insert("end", text + "\n")
        self.summary.see("end")

    def on_drop(self, event):
        raw = event.data
        path = raw.strip("{}").split("}\n{")[0].strip("{}").strip()
        self.load_file(Path(path))

    def browse(self):
        file = filedialog.askopenfilename(title="Select consolidated CSV", filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if file:
            self.load_file(Path(file))

    def load_sample(self):
        sample = self.paths.import_dir / "sample_consolidated_assets.csv"
        if sample.exists():
            self.load_file(sample)
        else:
            messagebox.showinfo("Sample CSV", f"Sample CSV not found at:\n{sample}")

    def load_file(self, path: Path):
        try:
            self.file_path = path
            self.file_label.config(text=str(path))
            self.rows, self.headers, self.mapping = parse_csv_file(path)
            self.preview_rows = self.rows[:25]
            self.populate_preview()
            self.summary.delete("1.0", "end")
            self.log(f"Loaded: {path.name}")
            self.log(f"Rows detected: {len(self.rows)}")
            self.log(f"Columns: {', '.join(self.headers)}")
            if self.mapping:
                self.log("Auto-map:")
                for src, dst in self.mapping.items():
                    self.log(f"  {src} -> {dst}")
            valid, invalid, messages = validate_rows(self.rows, self.db.list_assets())
            self.log(f"Valid rows: {len(valid)}")
            self.log(f"Invalid rows: {len(invalid)}")
            self.log(f"Duplicate or existing rows flagged: {sum(1 for r in valid if r.get('_duplicate') == '1')}")
            if messages:
                self.log("Validation notes:")
                for m in messages[:20]:
                    self.log(f"  • {m}")
        except Exception as e:
            messagebox.showerror("CSV Load Failed", str(e))

    def populate_preview(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        cols = list(self.preview_rows[0].keys()) if self.preview_rows else []
        self.tree["columns"] = cols
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor="w")
        for row in self.preview_rows:
            self.tree.insert("", "end", values=[row.get(c, "") for c in cols])

    def run_import(self):
        if not self.rows:
            messagebox.showwarning("No CSV", "Please select a CSV before importing.")
            return
        result = import_rows(self.db, self.rows, source_name=self.file_path.name if self.file_path else "consolidated_import.csv")
        messagebox.showinfo("Import complete", f"Imported: {result['rows_inserted']}\nDuplicates: {result['duplicates']}\nInvalid rows: {result['invalid_rows']}")
        self.result = result
        self.close()

    def finish_empty(self):
        self.db.mark_setup_done()
        self.result = {"rows_total": 0, "rows_inserted": 0, "rows_updated": 0, "duplicates": 0, "invalid_rows": 0}
        self.close()


class DashboardFrame(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.cards = {}
        self.chart_mv = None
        self.chart_ch = None
        self.build()

    def build(self):
        header = tk.Frame(self, bg=BG)
        header.pack(fill="x", padx=18, pady=(16, 8))
        tk.Label(header, text="ISMS Secure Asset Command Center", bg=BG, fg=TEXT, font=("Segoe UI", 20, "bold")).pack(anchor="w")
        tk.Label(header, text="Offline asset tracking, change governance, movement history, and endpoint compliance", bg=BG, fg=MUTED, font=("Segoe UI", 10)).pack(anchor="w", pady=(4, 0))
        grid = tk.Frame(self, bg=BG)
        grid.pack(fill="x", padx=18, pady=8)
        labels = [
            ("total_assets", "Total Assets"), ("active_employees", "Active Employees"), ("exited_employees", "Exited Employees"),
            ("bitlocker_compliance", "BitLocker Compliant"), ("edr_coverage", "EDR Coverage"), ("alerts", "Alerts"),
        ]
        for i, (key, title) in enumerate(labels):
            card = tk.Frame(grid, bg=PANEL, highlightthickness=1, highlightbackground=BORDER, padx=14, pady=14)
            card.grid(row=i // 3, column=i % 3, sticky="nsew", padx=8, pady=8)
            tk.Label(card, text=title, bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(anchor="w")
            value = tk.Label(card, text="0", bg=PANEL, fg=TEXT, font=("Segoe UI", 24, "bold"))
            value.pack(anchor="w", pady=(6, 0))
            self.cards[key] = value
        for c in range(3):
            grid.columnconfigure(c, weight=1)
        bottom = tk.Frame(self, bg=BG)
        bottom.pack(fill="both", expand=True, padx=18, pady=8)
        left = tk.Frame(bottom, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        right = tk.Frame(bottom, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        right.pack(side="left", fill="both", expand=True, padx=(8, 0))
        self.fig1 = Figure(figsize=(4.6, 2.8), dpi=100, facecolor=PANEL)
        self.ax1 = self.fig1.add_subplot(111)
        self.fig2 = Figure(figsize=(4.6, 2.8), dpi=100, facecolor=PANEL)
        self.ax2 = self.fig2.add_subplot(111)
        self.canvas1 = FigureCanvasTkAgg(self.fig1, master=left)
        self.canvas1.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=8)
        self.canvas2 = FigureCanvasTkAgg(self.fig2, master=right)
        self.canvas2.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=8)
        self.update()

    def update(self):
        counts = self.app.db.counts()
        for key, widget in self.cards.items():
            widget.config(text=str(counts.get(key, 0)))
        trends = self.app.db.trends()
        self.ax1.clear()
        self.ax1.set_facecolor(PANEL)
        if trends["movements"]:
            x = [d for d, _ in trends["movements"]]
            y = [c for _, c in trends["movements"]]
            self.ax1.plot(x, y, color=ACCENT, linewidth=2.5, marker="o")
            self.ax1.set_title("Movement Trend", color=TEXT)
            self.ax1.tick_params(axis="x", rotation=35, colors=MUTED)
            self.ax1.tick_params(axis="y", colors=MUTED)
        else:
            self.ax1.text(0.5, 0.5, "No movement data yet", ha="center", va="center", color=MUTED)
        self.fig1.tight_layout()
        self.canvas1.draw_idle()
        self.ax2.clear()
        self.ax2.set_facecolor(PANEL)
        labels = ["Compliant", "Non-Compliant", "Missing EDR"]
        vals = [counts.get("bitlocker_compliance", 0), counts.get("encryption_failures", 0), counts.get("missing_edr", 0)]
        if sum(vals) > 0:
            self.ax2.pie(vals, labels=labels, autopct="%1.0f%%", startangle=90, textprops={"color": TEXT})
            self.ax2.set_title("Compliance Snapshot", color=TEXT)
        else:
            self.ax2.text(0.5, 0.5, "No asset data yet", ha="center", va="center", color=MUTED)
        self.fig2.tight_layout()
        self.canvas2.draw_idle()


class AssetFrame(ttk.Frame):
    COLS = ["id", "hostname", "employee_code", "employee_name", "imei", "region", "asset_status", "current_location", "current_holder", "lifecycle_stage", "bitlocker_status", "edr_type", "last_updated_at"]
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.current_rows: list[dict[str, Any]] = []
        self.build()

    def build(self):
        top = tk.Frame(self, bg=BG)
        top.pack(fill="x", padx=18, pady=(16, 8))
        tk.Label(top, text="Asset Management", bg=BG, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(side="left")
        controls = tk.Frame(top, bg=BG)
        controls.pack(side="right")
        self.search = tk.StringVar()
        self.status = tk.StringVar(value="")
        self.region = tk.StringVar(value="")
        self.lifecycle = tk.StringVar(value="")
        tk.Entry(controls, textvariable=self.search, width=28, bg="#091522", fg=TEXT, insertbackground=TEXT, relief="flat").pack(side="left", padx=4, ipady=4)
        ttk.Combobox(controls, textvariable=self.status, width=12, values=["", "active", "missing", "retired", "disposed"], state="readonly").pack(side="left", padx=4)
        ttk.Combobox(controls, textvariable=self.lifecycle, width=14, values=["", "inventory", "assigned", "returned", "bin", "retired"], state="readonly").pack(side="left", padx=4)
        tk.Button(controls, text="Refresh", command=self.refresh, bg="#24364a", fg=TEXT, relief="flat").pack(side="left", padx=4)
        tk.Button(controls, text="Add/Edit", command=self.edit_selected, bg=ACCENT, fg="#00111d", relief="flat").pack(side="left", padx=4)
        tk.Button(controls, text="Movement", command=self.move_selected, bg="#24364a", fg=TEXT, relief="flat").pack(side="left", padx=4)
        tk.Button(controls, text="Delete", command=self.delete_selected, bg=DANGER, fg="white", relief="flat").pack(side="left", padx=4)
        body = tk.PanedWindow(self, orient="horizontal", sashrelief="flat", bg=BG, bd=0)
        body.pack(fill="both", expand=True, padx=18, pady=8)
        left = tk.Frame(body, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        right = tk.Frame(body, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        body.add(left, minsize=680)
        body.add(right, minsize=360)
        self.tree = ttk.Treeview(left, columns=self.COLS, show="headings", height=18)
        vs = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        hs = ttk.Scrollbar(left, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
        self.tree.pack(side="top", fill="both", expand=True, padx=8, pady=(8, 0))
        vs.pack(side="right", fill="y")
        hs.pack(side="bottom", fill="x")
        for col in self.COLS:
            self.tree.heading(col, text=col.replace("_", " ").title(), command=lambda c=col: self.sort_by(c))
            self.tree.column(col, width=120 if col not in {"remarks", "id"} else 60, anchor="w")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.detail = tk.Text(right, bg="#08111b", fg=TEXT, insertbackground=TEXT, relief="flat", wrap="word")
        self.detail.pack(fill="both", expand=True, padx=10, pady=10)
        action_row = tk.Frame(right, bg=PANEL)
        action_row.pack(fill="x", padx=10, pady=(0, 10))
        tk.Button(action_row, text="Save Row", command=self.edit_selected, bg=ACCENT, fg="#00111d", relief="flat").pack(side="left")
        tk.Button(action_row, text="Move", command=self.move_selected, bg="#24364a", fg=TEXT, relief="flat").pack(side="left", padx=6)
        tk.Button(action_row, text="Delete", command=self.delete_selected, bg=DANGER, fg="white", relief="flat").pack(side="left")
        self.refresh()

    def refresh(self):
        self.current_rows = self.app.db.list_assets(self.search.get().strip(), self.status.get().strip(), self.region.get().strip(), self.lifecycle.get().strip())
        for i in self.tree.get_children():
            self.tree.delete(i)
        for row in self.current_rows:
            self.tree.insert("", "end", iid=str(row["id"]), values=[row.get(c, "") for c in self.COLS])
        self.update_detail(None)
        self.app.dashboard.update()

    def sort_by(self, col: str):
        self.current_rows.sort(key=lambda r: norm_text(r.get(col)).lower())
        for i in self.tree.get_children():
            self.tree.delete(i)
        for row in self.current_rows:
            self.tree.insert("", "end", iid=str(row["id"]), values=[row.get(c, "") for c in self.COLS])

    def selected_id(self) -> int | None:
        sel = self.tree.selection()
        return int(sel[0]) if sel else None

    def on_select(self, event=None):
        asset_id = self.selected_id()
        if asset_id:
            self.update_detail(self.app.db.asset_detail(asset_id))

    def update_detail(self, row: dict[str, Any] | None):
        self.detail.delete("1.0", "end")
        if not row:
            self.detail.insert("end", "Select an asset to view details and history.\n")
            return
        lines = [
            f"Hostname: {row.get('hostname')}", f"Employee Code: {row.get('employee_code')}", f"Employee Name: {row.get('employee_name')}",
            f"IMEI: {row.get('imei')}", f"Region: {row.get('region')}", f"Department: {row.get('employee_department')}",
            f"BitLocker: {row.get('bitlocker_status')}", f"EDR: {row.get('edr_type')}", f"Status: {row.get('asset_status')}",
            f"Holder: {row.get('current_holder')}", f"Location: {row.get('current_location')}", f"Lifecycle: {row.get('lifecycle_stage')}",
            f"Last Updated By: {row.get('last_updated_by')}", f"Last Updated At: {row.get('last_updated_at')}", f"Remarks: {row.get('remarks')}",
            "", "Recent Movements:",
        ]
        for mv in self.app.db.asset_movements(row["id"] )[:5]:
            lines.append(f"- {mv.get('timestamp')} | {mv.get('from_location')} → {mv.get('to_location')} | {mv.get('reason')}")
        lines.append("\nChange Log:")
        for ch in self.app.db.asset_changes(row["id"] )[:8]:
            lines.append(f"- {ch.get('timestamp')} | {ch.get('field_name')}: {ch.get('old_value')} → {ch.get('new_value')}")
        self.detail.insert("end", "\n".join(lines))

    def edit_selected(self):
        asset_id = self.selected_id()
        asset = self.app.db.asset_detail(asset_id) if asset_id else {}
        dialog = AssetDialog(self, asset)
        self.wait_window(dialog)
        if dialog.result is not None:
            try:
                asset_id = self.app.db.upsert_asset(dialog.result, changed_by=dialog.result.get("last_updated_by") or "system")
                self.refresh()
                self.tree.selection_set(str(asset_id))
                self.tree.see(str(asset_id))
            except Exception as e:
                messagebox.showerror("Save failed", str(e))

    def move_selected(self):
        asset_id = self.selected_id()
        if not asset_id:
            messagebox.showwarning("No selection", "Select an asset first.")
            return
        asset = self.app.db.asset_detail(asset_id)
        dialog = MovementDialog(self, asset)
        self.wait_window(dialog)
        if dialog.result:
            payload = dict(dialog.result)
            payload["asset_id"] = asset_id
            try:
                self.app.db.create_movement(payload)
                self.refresh()
            except Exception as e:
                messagebox.showerror("Movement failed", str(e))

    def delete_selected(self):
        asset_id = self.selected_id()
        if not asset_id:
            return
        if messagebox.askyesno("Confirm delete", "Delete this asset permanently?"):
            try:
                self.app.db.delete_asset(asset_id, changed_by="user")
                self.refresh()
            except Exception as e:
                messagebox.showerror("Delete failed", str(e))


class EmployeeFrame(ttk.Frame):
    COLS = ["employee_code", "employee_name", "department", "manager", "region", "status", "left_date", "assigned_assets", "updated_at"]
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.current_rows = []
        self.build()

    def build(self):
        top = tk.Frame(self, bg=BG)
        top.pack(fill="x", padx=18, pady=(16, 8))
        tk.Label(top, text="Employee Management", bg=BG, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(side="left")
        controls = tk.Frame(top, bg=BG)
        controls.pack(side="right")
        self.search = tk.StringVar()
        self.status = tk.StringVar(value="")
        tk.Entry(controls, textvariable=self.search, width=28, bg="#091522", fg=TEXT, insertbackground=TEXT, relief="flat").pack(side="left", padx=4, ipady=4)
        ttk.Combobox(controls, textvariable=self.status, width=12, values=["", "active", "left"], state="readonly").pack(side="left", padx=4)
        tk.Button(controls, text="Refresh", command=self.refresh, bg="#24364a", fg=TEXT, relief="flat").pack(side="left", padx=4)
        tk.Button(controls, text="Add/Edit", command=self.edit_selected, bg=ACCENT, fg="#00111d", relief="flat").pack(side="left", padx=4)
        tk.Button(controls, text="Exit", command=self.exit_selected, bg=WARN, fg="#111111", relief="flat").pack(side="left", padx=4)
        tk.Button(controls, text="Delete", command=self.delete_selected, bg=DANGER, fg="white", relief="flat").pack(side="left", padx=4)
        body = tk.PanedWindow(self, orient="horizontal", sashrelief="flat", bg=BG, bd=0)
        body.pack(fill="both", expand=True, padx=18, pady=8)
        left = tk.Frame(body, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        right = tk.Frame(body, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        body.add(left, minsize=700)
        body.add(right, minsize=340)
        self.tree = ttk.Treeview(left, columns=self.COLS, show="headings", height=18)
        self.tree.pack(side="top", fill="both", expand=True, padx=8, pady=(8, 0))
        vs = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        hs = ttk.Scrollbar(left, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
        vs.pack(side="right", fill="y")
        hs.pack(side="bottom", fill="x")
        for col in self.COLS:
            self.tree.heading(col, text=col.replace("_", " ").title())
            self.tree.column(col, width=120 if col not in {"assigned_assets"} else 80)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.detail = tk.Text(right, bg="#08111b", fg=TEXT, insertbackground=TEXT, relief="flat", wrap="word")
        self.detail.pack(fill="both", expand=True, padx=10, pady=10)
        action_row = tk.Frame(right, bg=PANEL)
        action_row.pack(fill="x", padx=10, pady=(0, 10))
        tk.Button(action_row, text="Save Row", command=self.edit_selected, bg=ACCENT, fg="#00111d", relief="flat").pack(side="left")
        tk.Button(action_row, text="Exit Workflow", command=self.exit_selected, bg=WARN, fg="#111111", relief="flat").pack(side="left", padx=6)
        self.refresh()

    def refresh(self):
        self.current_rows = self.app.db.list_employees(self.search.get().strip(), self.status.get().strip())
        for i in self.tree.get_children():
            self.tree.delete(i)
        for row in self.current_rows:
            self.tree.insert("", "end", iid=row["employee_code"], values=[row.get(c, "") for c in self.COLS])
        self.update_detail(None)

    def selected_code(self) -> str | None:
        sel = self.tree.selection()
        return sel[0] if sel else None

    def on_select(self, event=None):
        code = self.selected_code()
        if code:
            self.update_detail(self.app.db.employee_detail(code))

    def update_detail(self, row: dict[str, Any] | None):
        self.detail.delete("1.0", "end")
        if not row:
            self.detail.insert("end", "Select an employee to view details and history.\n")
            return
        lines = [
            f"Employee Code: {row.get('employee_code')}", f"Employee Name: {row.get('employee_name')}", f"Department: {row.get('department')}",
            f"Manager: {row.get('manager')}", f"Region: {row.get('region')}", f"Status: {row.get('status')}",
            f"Left Date: {row.get('left_date')}", f"Left Reason: {row.get('left_reason')}",
            f"Assigned Assets: {row.get('assigned_assets')}", f"Exit History: {row.get('exit_history')}",
            "", "Recent Changes:",
        ]
        for ch in self.app.db.employee_changes(row["employee_code"])[:10]:
            lines.append(f"- {ch.get('timestamp')} | {ch.get('field_name')}: {ch.get('old_value')} → {ch.get('new_value')}")
        self.detail.insert("end", "\n".join(lines))

    def edit_selected(self):
        code = self.selected_code()
        emp = self.app.db.employee_detail(code) if code else {}
        dialog = EmployeeDialog(self, emp)
        self.wait_window(dialog)
        if dialog.result is not None:
            try:
                self.app.db.upsert_employee(dialog.result)
                self.refresh()
            except Exception as e:
                messagebox.showerror("Save failed", str(e))

    def exit_selected(self):
        code = self.selected_code()
        if not code:
            messagebox.showwarning("No selection", "Select an employee first.")
            return
        emp = self.app.db.employee_detail(code)
        dialog = ExitDialog(self, emp)
        self.wait_window(dialog)
        if dialog.result:
            try:
                payload = dict(emp)
                payload.update(dialog.result)
                payload["status"] = "left"
                self.app.db.upsert_employee(payload)
                # Set assigned assets count from current DB snapshot
                assigned = len([a for a in self.app.db.list_assets() if norm_text(a.get("employee_code")) == code])
                self.app.db.conn.execute("UPDATE employees SET assigned_assets=? WHERE employee_code=?", (assigned, code))
                self.app.db.conn.commit()
                self.refresh()
            except Exception as e:
                messagebox.showerror("Exit failed", str(e))

    def delete_selected(self):
        code = self.selected_code()
        if not code:
            return
        if messagebox.askyesno("Confirm delete", "Delete this employee permanently?"):
            try:
                self.app.db.delete_employee(code, changed_by="user")
                self.refresh()
            except Exception as e:
                messagebox.showerror("Delete failed", str(e))


class HistoryFrame(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.build()

    def build(self):
        top = tk.Frame(self, bg=BG)
        top.pack(fill="x", padx=18, pady=(16, 8))
        tk.Label(top, text="Audit / Movement History", bg=BG, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(side="left")
        tk.Button(top, text="Refresh", command=self.refresh, bg="#24364a", fg=TEXT, relief="flat").pack(side="right")
        container = tk.PanedWindow(self, orient="horizontal", bg=BG, bd=0)
        container.pack(fill="both", expand=True, padx=18, pady=8)
        left = tk.Frame(container, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        right = tk.Frame(container, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        container.add(left, minsize=610)
        container.add(right, minsize=610)
        tk.Label(left, text="Recent Movements", bg=PANEL, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=10, pady=(10, 4))
        self.mov = ttk.Treeview(left, columns=("timestamp", "hostname", "from_location", "to_location", "reason", "changed_by"), show="headings", height=15)
        self.mov.pack(fill="both", expand=True, padx=10, pady=8)
        for c in self.mov["columns"]:
            self.mov.heading(c, text=c.replace("_", " ").title())
            self.mov.column(c, width=120)
        tk.Label(right, text="Recent Change Log", bg=PANEL, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=10, pady=(10, 4))
        self.chg = ttk.Treeview(right, columns=("timestamp", "entity_type", "entity_id", "field_name", "old_value", "new_value", "changed_by"), show="headings", height=15)
        self.chg.pack(fill="both", expand=True, padx=10, pady=8)
        for c in self.chg["columns"]:
            self.chg.heading(c, text=c.replace("_", " ").title())
            self.chg.column(c, width=120)
        self.refresh()

    def refresh(self):
        for t in (self.mov, self.chg):
            for i in t.get_children():
                t.delete(i)
        for row in self.app.db.movements(250):
            self.mov.insert("", "end", values=(row.get("timestamp"), row.get("hostname"), row.get("from_location"), row.get("to_location"), row.get("reason"), row.get("changed_by")))
        for row in self.app.db.changes(400):
            self.chg.insert("", "end", values=(row.get("timestamp"), row.get("entity_type"), row.get("entity_id"), row.get("field_name"), row.get("old_value"), row.get("new_value"), row.get("changed_by")))


class ImportExportFrame(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.build()

    def build(self):
        top = tk.Frame(self, bg=BG)
        top.pack(fill="x", padx=18, pady=(16, 8))
        tk.Label(top, text="Import / Export / Backup", bg=BG, fg=TEXT, font=("Segoe UI", 18, "bold")).pack(side="left")
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=18, pady=8)
        left = tk.Frame(body, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        right = tk.Frame(body, bg=PANEL, highlightthickness=1, highlightbackground=BORDER)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        right.pack(side="left", fill="both", expand=True, padx=(8, 0))
        self.log = tk.Text(left, bg="#08111b", fg=TEXT, insertbackground=TEXT, relief="flat", wrap="word")
        self.log.pack(fill="both", expand=True, padx=10, pady=10)
        buttons = tk.Frame(right, bg=PANEL)
        buttons.pack(fill="both", expand=True, padx=10, pady=10)
        actions = [
            ("Import CSV", self.import_csv),
            ("Export Assets CSV", self.export_assets_csv),
            ("Export Employees CSV", self.export_employees_csv),
            ("Export Full Excel", self.export_excel),
            ("Backup JSON", self.backup_json),
            ("Restore JSON", self.restore_json_action),
        ]
        for i, (label, cmd) in enumerate(actions):
            tk.Button(buttons, text=label, command=cmd, bg="#24364a", fg=TEXT, relief="flat", activebackground="#34495e", font=("Segoe UI", 10, "bold")).pack(fill="x", pady=5)
        self.write("Local import/export tools are ready. All files remain on the laptop.")

    def write(self, text: str):
        self.log.insert("end", text + "\n")
        self.log.see("end")

    def import_csv(self):
        file = filedialog.askopenfilename(title="Select consolidated CSV", filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not file:
            return
        try:
            rows, headers, mapping = parse_csv_file(Path(file))
            result = import_rows(self.app.db, rows, source_name=Path(file).name)
            self.write(f"Imported {result['rows_inserted']} rows from {file}")
            self.write(f"Duplicates: {result['duplicates']} | Invalid: {result['invalid_rows']}")
            self.app.refresh_all()
        except Exception as e:
            messagebox.showerror("Import failed", str(e))

    def export_assets_csv(self):
        file = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")], title="Save assets CSV")
        if file:
            export_csv(Path(file), self.app.db.list_assets())
            self.write(f"Assets exported to {file}")

    def export_employees_csv(self):
        file = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")], title="Save employees CSV")
        if file:
            export_csv(Path(file), self.app.db.list_employees())
            self.write(f"Employees exported to {file}")

    def export_excel(self):
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save workbook")
        if file:
            export_xlsx(Path(file), self.app.db.all_tables_export())
            self.write(f"Workbook exported to {file}")

    def backup_json(self):
        file = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")], title="Save backup JSON")
        if file:
            export_json(Path(file), self.app.db.all_tables_export())
            self.write(f"Backup saved to {file}")

    def restore_json_action(self):
        file = filedialog.askopenfilename(title="Restore from backup JSON", filetypes=[("JSON files", "*.json")])
        if not file:
            return
        if not messagebox.askyesno("Confirm restore", "This will replace all local data. Continue?"):
            return
        try:
            data = restore_json(Path(file))
            self.app.db.wipe_and_restore(data)
            self.write(f"Restored backup from {file}")
            self.app.refresh_all()
        except Exception as e:
            messagebox.showerror("Restore failed", str(e))


class MainWindow:
    def __init__(self, root: tk.Tk, db: Database, paths: Paths):
        self.root = root
        self.db = db
        self.paths = paths
        self.root.title(APP_NAME)
        self.root.configure(bg=BG)
        self.root.geometry("1600x960")
        self.root.minsize(1400, 860)
        self.root.deiconify()
        self.root.protocol("WM_DELETE_WINDOW", self.close)
        self._style()
        self.build_layout()
        self.refresh_all()

    def _style(self):
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(".", background=BG, foreground=TEXT)
        style.configure("TFrame", background=BG)
        style.configure("Card.TFrame", background=PANEL)
        style.configure("TLabel", background=BG, foreground=TEXT)
        style.configure("Treeview", background="#08111b", fieldbackground="#08111b", foreground=TEXT, rowheight=28, bordercolor=BORDER)
        style.configure("Treeview.Heading", background="#12304f", foreground=TEXT, relief="flat", font=("Segoe UI", 9, "bold"))
        style.map("Treeview", background=[("selected", "#14365c")], foreground=[("selected", TEXT)])
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(14, 8), background=PANEL, foreground=TEXT)
        style.map("TNotebook.Tab", background=[("selected", PANEL_2)])

    def build_layout(self):
        root = tk.Frame(self.root, bg=BG)
        root.pack(fill="both", expand=True)
        sidebar = tk.Frame(root, bg="#08111b", width=250, highlightthickness=1, highlightbackground=BORDER)
        sidebar.pack(side="left", fill="y")
        self.content = tk.Frame(root, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)
        title = tk.Frame(sidebar, bg="#08111b")
        title.pack(fill="x", padx=16, pady=(18, 12))
        tk.Label(title, text="ISMS", bg="#08111b", fg=ACCENT, font=("Segoe UI", 16, "bold")).pack(anchor="w")
        tk.Label(title, text="Secure Asset Command Center", bg="#08111b", fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        tk.Label(title, text="Offline Windows Desktop", bg="#08111b", fg=MUTED, font=("Segoe UI", 9)).pack(anchor="w")
        self.nav = {}
        for name in ["Dashboard", "Assets", "Employees", "History", "Import / Export"]:
            b = tk.Button(sidebar, text=name, command=lambda n=name: self.show(n), bg="#0d1b2a", fg=TEXT, activebackground="#12304f", relief="flat", font=("Segoe UI", 10, "bold"))
            b.pack(fill="x", padx=12, pady=6, ipady=8)
            self.nav[name] = b
        tk.Label(sidebar, text="All data stored locally\nNo browser\nNo server\nNo cloud", bg="#08111b", fg=MUTED, font=("Segoe UI", 9), justify="left").pack(anchor="w", padx=16, pady=18)
        self.frame_host = tk.Frame(self.content, bg=BG)
        self.frame_host.pack(fill="both", expand=True)
        self.frames = {
            "Dashboard": DashboardFrame(self.frame_host, self),
            "Assets": AssetFrame(self.frame_host, self),
            "Employees": EmployeeFrame(self.frame_host, self),
            "History": HistoryFrame(self.frame_host, self),
            "Import / Export": ImportExportFrame(self.frame_host, self),
        }
        for frame in self.frames.values():
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.dashboard = self.frames["Dashboard"]
        self.show("Dashboard")

    def show(self, name: str):
        frame = self.frames[name]
        frame.lift()
        for k, b in self.nav.items():
            b.config(bg="#12304f" if k == name else "#0d1b2a")
        if hasattr(frame, "update"):
            try:
                frame.update()
            except TypeError:
                pass
        if hasattr(frame, "refresh"):
            try:
                frame.refresh()
            except Exception:
                pass

    def refresh_all(self):
        for name, frame in self.frames.items():
            if hasattr(frame, "refresh"):
                try:
                    frame.refresh()
                except Exception:
                    pass
        if hasattr(self.dashboard, "update"):
            self.dashboard.update()

    def close(self):
        self.db.close()
        self.root.destroy()


def seed_sample_csv(paths: Paths):
    sample = paths.import_dir / "sample_consolidated_assets.csv"
    if sample.exists():
        return sample
    rows = [
        {
            "hostname": "LAP-ISMS-001",
            "employee_code": "EMP1001",
            "employee_name": "Aarav Mehta",
            "imei": "SN-001-A1",
            "region": "North",
            "employee_department": "IT",
            "bitlocker_status": "On",
            "bitlocker_key_c": "CK-001",
            "bitlocker_key_d": "DK-001",
            "ram": "16 GB",
            "storage": "512 GB SSD",
            "windows_version": "Windows 11 Pro",
            "edr_type": "CS",
            "admin_password": "Stored Securely",
            "asset_status": "active",
            "remarks": "Primary laptop",
            "current_holder": "Aarav Mehta",
            "current_location": "Office",
            "lifecycle_stage": "assigned",
            "last_updated_by": "admin",
            "last_updated_at": now_iso(),
            "manager": "Nisha Rao",
            "department": "IT",
            "status": "active",
            "assigned_assets": 1,
            "exit_history": "",
        },
        {
            "hostname": "LAP-ISMS-002",
            "employee_code": "EMP1002",
            "employee_name": "Mira Shah",
            "imei": "SN-002-B2",
            "region": "West",
            "employee_department": "Finance",
            "bitlocker_status": "On",
            "bitlocker_key_c": "CK-002",
            "bitlocker_key_d": "DK-002",
            "ram": "16 GB",
            "storage": "1 TB SSD",
            "windows_version": "Windows 11 Pro",
            "edr_type": "TM",
            "admin_password": "Stored Securely",
            "asset_status": "active",
            "remarks": "Finance endpoint",
            "current_holder": "Mira Shah",
            "current_location": "Branch Office",
            "lifecycle_stage": "assigned",
            "last_updated_by": "admin",
            "last_updated_at": now_iso(),
            "manager": "Karan Patel",
            "department": "Finance",
            "status": "active",
            "assigned_assets": 1,
            "exit_history": "",
        },
    ]
    with open(sample, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return sample


def main():
    paths = resolve_paths()
    seed_sample_csv(paths)
    db = Database(paths.db_path)
    db.initialize()
    root = TkinterDnD.Tk() if hasattr(TkinterDnD, "Tk") else tk.Tk()
    root.withdraw()
    root.update_idletasks()
    try:
        if db.needs_initial_import():
            wizard = ImportWizard(root, db, paths)
            root.wait_window(wizard)
        app = MainWindow(root, db, paths)
        root.deiconify()
        root.mainloop()
    finally:
        try:
            db.close()
        except Exception:
            pass


if __name__ == "__main__":
    main()
