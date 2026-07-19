# ================= ГЛОБАЛЬНЫЙ ПЕРЕХВАТ ОШИБОК =================
import sys, os, traceback, platform, socket, getpass, subprocess
from datetime import datetime
from io import BytesIO
from pathlib import Path

# Попытка импорта библиотек, без которых не обойтись
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from Crypto.Cipher import AES
    from Crypto.Protocol.KDF import PBKDF2
    from Crypto.Random import get_random_bytes
except Exception as e:
    # Если какая-то из библиотек не установлена, сразу покажем ошибку и не закроем окно
    print("ОШИБКА: не установлены необходимые библиотеки.")
    print("Установите их командой: pip install openpyxl pycryptodome")
    print(f"Техническая информация: {e}")
    input("Нажмите Enter для выхода...")
    sys.exit(1)

# ================= ПАРАМЕТРЫ ШИФРОВАНИЯ =================
PASSWORD = "OtP1nd2j3hL9dxeREJOlQkUvMRPdICH0m2K7BQh8sVuY8XRT3ugUpPHdw1xnLEvh"
SALT = b"MyS@lt!2024_"
ITERATIONS = 10_000
KEY_SIZE = 32
OUTPUT_DIR = r"C:\Temp"

def log_and_pause():
    """Вызывается при критической ошибке: выводит traceback на экран и в лог, ждёт Enter."""
    print("\n" + "=" * 60)
    print("КРИТИЧЕСКАЯ ОШИБКА! Окно останется открытым до нажатия Enter.")
    traceback.print_exc()
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        with open(os.path.join(OUTPUT_DIR, "error.log"), "w", encoding="utf-8") as f:
            traceback.print_exc(file=f)
        print(f"\nПодробности сохранены в {OUTPUT_DIR}\\error.log")
    except:
        pass
    input("\nНажмите Enter для выхода...")

# ================= ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =================
def run_cmd(command):
    """Безопасный вызов subprocess, возвращает строку вывода или пустую строку."""
    try:
        return subprocess.check_output(command, shell=True, text=True, stderr=subprocess.STDOUT).strip()
    except:
        return ""

def get_hostname():
    return platform.node()

def get_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

def get_user():
    return getpass.getuser()

# ================= ФУНКЦИИ СБОРА ДАННЫХ =================

def collect_overview():
    """Основная информация (титульный лист)."""
    data = {
        "Report Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Hostname": get_hostname(),
        "IP Address": get_ip(),
        "User": get_user(),
        "OS": platform.platform(),
        "Architecture": platform.architecture()[0],
        "Processor": platform.processor(),
    }
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                            r"SOFTWARE\Microsoft\Windows NT\CurrentVersion") as key:
            data["OS Name"] = winreg.QueryValueEx(key, "ProductName")[0]
            try:
                data["ReleaseId"] = winreg.QueryValueEx(key, "DisplayVersion")[0]
            except:
                data["ReleaseId"] = ""
            data["Build"] = winreg.QueryValueEx(key, "CurrentBuild")[0]
    except:
        data["OS Name"] = platform.platform()
    out = run_cmd("systeminfo")
    for line in out.splitlines():
        if "Domain:" in line:
            data["Domain"] = line.split(":")[1].strip()
            break
    else:
        data["Domain"] = "Unknown"
    return data

def collect_system_info():
    """Детальная информация о системе (без psutil)."""
    info = {}
    # Производитель и модель
    out = run_cmd("wmic computersystem get manufacturer,model /format:csv")
    if out:
        lines = out.splitlines()
        if len(lines) >= 2:
            parts = [p.strip() for p in lines[1].split(',')]
            if len(parts) >= 3:
                info["Manufacturer"] = parts[2]
                info["Model"] = parts[3] if len(parts) > 3 else ""
    # ОЗУ
    out = run_cmd("wmic computersystem get totalphysicalmemory /format:csv")
    if out:
        lines = out.splitlines()
        if len(lines) >= 2:
            val = lines[1].split(',')[-1].strip()
            try:
                ram = int(val)
                info["Total RAM (GB)"] = round(ram / (1024**3), 1)
            except:
                pass
    # Серийный номер
    out = run_cmd("wmic bios get serialnumber /format:csv")
    if out:
        lines = out.splitlines()
        if len(lines) >= 2:
            sn = lines[1].split(',')[-1].strip()
            if sn:
                info["Serial Number"] = sn
    # Дата установки ОС
    out = run_cmd("wmic os get installdate /format:csv")
    if out:
        lines = out.splitlines()
        if len(lines) >= 2:
            ds = lines[1].split(',')[-1].strip()[:14]
            if ds:
                try:
                    dt = datetime.strptime(ds, "%Y%m%d%H%M%S")
                    info["Install Date"] = dt.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    pass
    return info

def collect_users_and_groups():
    """Локальные пользователи и группы."""
    users = []
    admins = []
    out = run_cmd('wmic useraccount where "localaccount=true" get name')
    for line in out.splitlines():
        name = line.strip()
        if name and name != "Name":
            users.append(name)
    out = run_cmd("net localgroup administrators")
    capture = False
    for line in out.splitlines():
        if "---" in line:
            capture = True
            continue
        if capture and line.strip() and not line.startswith("Команда выполнена"):
            admins.append(line.strip())
    groups = []
    out = run_cmd("net localgroup")
    for line in out.splitlines():
        if line.startswith("*"):
            group_name = line[1:].strip()
            members_out = run_cmd(f'net localgroup "{group_name}"')
            members = []
            capture = False
            for mline in members_out.splitlines():
                if "---" in mline:
                    capture = True
                    continue
                if capture and mline.strip() and not mline.startswith("Команда выполнена"):
                    members.append(mline.strip())
            groups.append({"Group": group_name, "Members": ", ".join(members)})
    user_list = [{"Username": u, "IsAdmin": u in admins} for u in users]
    return user_list, groups

def collect_network():
    """Сетевые настройки, RDP, hosts, DNS, netstat."""
    rdp = {}
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                            r"SYSTEM\CurrentControlSet\Control\Terminal Server") as key:
            fDeny = winreg.QueryValueEx(key, "fDenyTSConnections")[0]
        rdp["RDP Enabled"] = fDeny == 0
    except:
        rdp["RDP Enabled"] = "Unknown"
    netstat = []
    out = run_cmd("netstat -ano")
    for line in out.splitlines():
        if line.strip():
            netstat.append({"Line": line.strip()})
    hosts = []
    hosts_path = r"C:\Windows\System32\drivers\etc\hosts"
    if os.path.exists(hosts_path):
        with open(hosts_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    hosts.append({"Entry": line})
    dns_cache = []
    out = run_cmd("ipconfig /displaydns")
    record = {}
    for line in out.splitlines():
        if not line.strip():
            if record:
                dns_cache.append(record)
                record = {}
        elif ":" in line:
            k, v = line.split(":", 1)
            record[k.strip()] = v.strip()
    if record:
        dns_cache.append(record)
    return rdp, netstat, hosts, dns_cache

def collect_installed_software():
    """Установленные программы из реестра."""
    import winreg
    software = []
    paths = [
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall",
        r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"
    ]
    for uninstall_path in paths:
        try:
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, uninstall_path, 0, winreg.KEY_READ)
            i = 0
            while True:
                try:
                    subkey_name = winreg.EnumKey(key, i)
                    subkey = winreg.OpenKey(key, subkey_name)
                    try:
                        name = winreg.QueryValueEx(subkey, "DisplayName")[0]
                    except:
                        name = None
                    if name:
                        version = publisher = install_date = ""
                        try: version = winreg.QueryValueEx(subkey, "DisplayVersion")[0]
                        except: pass
                        try: publisher = winreg.QueryValueEx(subkey, "Publisher")[0]
                        except: pass
                        try: install_date = winreg.QueryValueEx(subkey, "InstallDate")[0]
                        except: pass
                        software.append({
                            "Name": name,
                            "Version": version,
                            "Publisher": publisher,
                            "InstallDate": install_date,
                        })
                    winreg.CloseKey(subkey)
                    i += 1
                except OSError:
                    break
            winreg.CloseKey(key)
        except:
            pass
    return software

def collect_portable_apps():
    """Портативные приложения."""
    known_names = {sw["Name"].lower().replace(" ", "") for sw in collect_installed_software()}
    portable_paths = [
        os.environ.get("ProgramFiles", "C:\\Program Files"),
        os.environ.get("ProgramFiles(x86)", "C:\\Program Files (x86)"),
        os.path.join(os.environ.get("LOCALAPPDATA", ""), "Programs"),
    ]
    portable = []
    for base in portable_paths:
        if not base or not os.path.exists(base):
            continue
        try:
            for entry in os.scandir(base):
                if entry.is_dir():
                    exe_file = None
                    for f in os.scandir(entry.path):
                        if f.is_file() and f.name.lower().endswith(".exe"):
                            exe_file = f.path
                            break
                    if exe_file:
                        name = os.path.splitext(os.path.basename(exe_file))[0]
                        if name.lower().replace(" ", "") in known_names:
                            continue
                        version = manufacturer = ""
                        try:
                            escaped = exe_file.replace("\\", "\\\\")
                            info = run_cmd(f'wmic datafile where name="{escaped}" get version, manufacturer')
                            lines = info.splitlines()
                            if len(lines) >= 2:
                                parts = lines[1].split(None, 1)
                                version = parts[0] if parts else ""
                                manufacturer = parts[1] if len(parts) > 1 else ""
                        except:
                            pass
                        portable.append({
                            "Name": name,
                            "Version": version,
                            "Publisher": manufacturer,
                            "Path": entry.path,
                        })
        except:
            pass
    return portable

def collect_autoruns():
    """Точки автозагрузки."""
    import winreg
    entries = []
    hives = {"HKCU": winreg.HKEY_CURRENT_USER, "HKLM": winreg.HKEY_LOCAL_MACHINE}
    subpaths = [
        r"Software\Microsoft\Windows\CurrentVersion\Run",
        r"Software\Microsoft\Windows\CurrentVersion\RunOnce",
        r"Software\Microsoft\Windows\CurrentVersion\RunServices",
        r"Software\Microsoft\Windows\CurrentVersion\RunServicesOnce",
    ]
    for hive_name, hive in hives.items():
        for subpath in subpaths:
            try:
                key = winreg.OpenKey(hive, subpath, 0, winreg.KEY_READ)
                i = 0
                while True:
                    try:
                        name, value, _ = winreg.EnumValue(key, i)
                        entries.append({"Hive": hive_name, "Path": subpath, "Name": name, "Value": str(value)})
                        i += 1
                    except OSError:
                        break
                winreg.CloseKey(key)
            except:
                pass
    # Winlogon спецключи
    special = [
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Winlogon", "Shell"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Winlogon", "Userinit"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Windows", "AppInit_DLLs"),
    ]
    for hive, path, value_name in special:
        try:
            key = winreg.OpenKey(hive, path, 0, winreg.KEY_READ)
            value, _ = winreg.QueryValueEx(key, value_name)
            entries.append({"Hive": "HKLM", "Path": path, "Name": value_name, "Value": str(value)})
            winreg.CloseKey(key)
        except:
            pass
    return entries

def collect_processes():
    """Процессы через tasklist."""
    proc_list = []
    out = run_cmd("tasklist /FO CSV /NH")
    for line in out.splitlines():
        parts = [p.strip('"') for p in line.split('","')]
        if len(parts) >= 5:
            proc_list.append({
                "PID": parts[1],
                "Name": parts[0],
                "Path": "",
                "User": parts[-1] if parts[-1] != 'N/A' else "",
            })
    return proc_list

def collect_services():
    """Службы через sc query."""
    services = []
    out = run_cmd("sc query state= all")
    current = {}
    for line in out.splitlines():
        if line.startswith("SERVICE_NAME:"):
            if current:
                services.append(current)
            current = {"Name": line.split(":",1)[1].strip()}
        elif line.startswith("DISPLAY_NAME:"):
            current["DisplayName"] = line.split(":",1)[1].strip()
        elif "STATE" in line and ":" in line:
            parts = line.split(":")[1].strip().split()
            if parts:
                code = parts[0]
                status = "STOPPED" if code == "1" else "RUNNING" if code == "4" else "UNKNOWN"
                current["Status"] = status
        elif "START_TYPE" in line and ":" in line:
            parts = line.split(":")[1].strip().split()
            if parts:
                start = {"2":"AUTO", "3":"MANUAL", "4":"DISABLED"}.get(parts[0], "UNKNOWN")
                current["StartType"] = start
        elif line.startswith("        BINARY_PATH_NAME   :"):
            current["BinaryPath"] = line.split(":",1)[1].strip()
    if current:
        services.append(current)
    return services

def collect_drivers():
    """Драйверы через driverquery."""
    out = run_cmd("driverquery /v /fo csv")
    drivers = []
    if out:
        lines = out.splitlines()
        if lines:
            headers = [h.strip('"') for h in lines[0].split(',')]
            for line in lines[1:]:
                vals = [v.strip('" ') for v in line.split(',')]
                if len(vals) >= len(headers):
                    drivers.append(dict(zip(headers, vals)))
    return drivers

def collect_scheduled_tasks():
    """Задачи планировщика."""
    out = run_cmd("schtasks /query /fo CSV /v")
    tasks = []
    if out:
        lines = out.splitlines()
        if not lines:
            return tasks
        headers = [h.strip('"') for h in lines[0].split(',')]
        for line in lines[1:]:
            vals = [v.strip('" ') for v in line.split(',')]
            if len(vals) >= len(headers):
                tasks.append(dict(zip(headers, vals)))
    return tasks

def collect_security_events():
    """События безопасности за 24 часа."""
    out = run_cmd(
        'wevtutil qe Security /q:"*[System[(EventID=4624 or EventID=4625 or EventID=4672 or EventID=1102 or EventID=7045 or EventID=7036 or EventID=1001 or EventID=41) and TimeCreated[timediff(@SystemTime) <= 86400000]]]" /c:100 /rd:true /f:text'
    )
    events = []
    record = {}
    for line in out.splitlines():
        if not line.strip():
            if record:
                events.append(record)
                record = {}
        elif ":" in line:
            k, v = line.split(":", 1)
            record[k.strip()] = v.strip()
    if record:
        events.append(record)
    return events

def collect_disk_encryption():
    """Диски и BitLocker."""
    disks = []
    out = run_cmd("wmic logicaldisk where drivetype=3 get deviceid,size,freespace,filesystem /format:csv")
    if out:
        lines = out.splitlines()
        for line in lines[1:]:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) >= 5:
                try:
                    size = int(parts[3]) if parts[3] else 0
                    free = int(parts[2]) if parts[2] else 0
                except:
                    size = free = 0
                disks.append({
                    "Device": parts[1],
                    "Mount": parts[1],
                    "FS": parts[4] if len(parts) > 4 else "",
                    "TotalGB": round(size / (1024**3), 2) if size else 0,
                    "FreeGB": round(free / (1024**3), 2) if free else 0,
                })
    bitlocker = []
    out = run_cmd("manage-bde -status")
    vol = {}
    for line in out.splitlines():
        if "Volume:" in line:
            if vol:
                bitlocker.append(vol)
            vol = {"Volume": line.split(":",1)[1].strip()}
        elif "Conversion Status:" in line:
            vol["ConversionStatus"] = line.split(":",1)[1].strip()
        elif "Protection Method:" in line:
            vol["ProtectionMethod"] = line.split(":",1)[1].strip()
    if vol:
        bitlocker.append(vol)
    return disks, bitlocker

def collect_environment():
    return [{"Variable": k, "Value": v} for k, v in os.environ.items()]

def collect_usb_history():
    import winreg
    devices = []
    path = r"SYSTEM\CurrentControlSet\Enum\USBSTOR"
    try:
        key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, path, 0, winreg.KEY_READ)
        i = 0
        while True:
            try:
                devices.append({"Device": winreg.EnumKey(key, i)})
                i += 1
            except OSError:
                break
        winreg.CloseKey(key)
    except:
        pass
    return devices

def collect_unsigned_drivers():
    unsigned = []
    ps_cmd = r'Get-ChildItem -Path C:\Windows\System32\drivers -Filter *.sys -ErrorAction SilentlyContinue | Get-AuthenticodeSignature | Where-Object { $_.Status -ne "Valid" } | Select-Object -ExpandProperty Path'
    out = run_cmd(["powershell", "-Command", ps_cmd])
    for line in out.splitlines():
        if line.strip():
            unsigned.append({"Path": line.strip()})
    return unsigned

def collect_legacy_tasks():
    files = []
    for folder in [r"C:\Windows\Tasks", r"C:\Windows\System32\Tasks"]:
        if os.path.exists(folder):
            for root, dirs, filenames in os.walk(folder):
                for f in filenames:
                    files.append({"Path": os.path.join(root, f)})
    return files

def collect_applocker():
    out = run_cmd('powershell -Command "Get-AppLockerPolicy -Effective -ErrorAction SilentlyContinue | Select-Object -ExpandProperty RuleCollections | Out-String"')
    if out.strip():
        return [{"RawPolicy": out.strip()}]
    return []

# ================= ФОРМИРОВАНИЕ EXCEL =================
def style_sheet(ws, headers=None):
    if headers:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    ws.freeze_panes = 'A2'

def create_workbook(overview_data, blocks):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    # Титульный лист
    ws_over = wb.create_sheet("Overview", 0)
    ws_over.merge_cells('A1:B1')
    c = ws_over['A1']
    c.value = "Host Audit Report"
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_over.row_dimensions[1].height = 30
    ws_over.merge_cells('A2:B2')
    ws_over['A2'].value = f"Generated: {overview_data.get('Report Generated', '')}"
    ws_over['A2'].font = Font(italic=True, size=10)
    row = 4
    for k, v in overview_data.items():
        if k == "Report Generated":
            continue
        ws_over.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws_over.cell(row=row, column=2, value=str(v))
        row += 1
    ws_over.column_dimensions['A'].width = 25
    ws_over.column_dimensions['B'].width = 50
    ws_over.freeze_panes = 'A4'
    # Остальные листы
    for name, (headers, rows) in blocks.items():
        ws = wb.create_sheet(name)
        if rows:
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h, "") for h in headers])
            style_sheet(ws, headers)
        else:
            ws.append(["No data"])
            ws.merge_cells('A1:Z1')
            ws['A1'].font = Font(italic=True, color="808080")
    return wb

def encrypt_data(data):
    key = PBKDF2(PASSWORD, SALT, dkLen=KEY_SIZE, count=ITERATIONS)
    iv = get_random_bytes(16)
    cipher = AES.new(key, AES.MODE_CBC, iv)
    pad_len = 16 - len(data) % 16
    if pad_len == 0:
        pad_len = 16
    padded = data + bytes([pad_len] * pad_len)
    return iv + cipher.encrypt(padded)

# ================= ГЛАВНАЯ ФУНКЦИЯ =================
def main():
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        hostname = get_hostname()
        enc_path = os.path.join(OUTPUT_DIR, f"{hostname}-report.enc")

        print("Сбор информации о системе...")
        overview = collect_overview()
        system_info = collect_system_info()
        users, groups = collect_users_and_groups()
        rdp, netstat, hosts, dns_cache = collect_network()
        installed = collect_installed_software()
        portable = collect_portable_apps()
        autoruns = collect_autoruns()
        processes = collect_processes()
        services = collect_services()
        drivers = collect_drivers()
        tasks = collect_scheduled_tasks()
        events = collect_security_events()
        disks, bitlocker = collect_disk_encryption()
        env = collect_environment()
        usb = collect_usb_history()
        unsigned = collect_unsigned_drivers()
        legacy = collect_legacy_tasks()
        applocker = collect_applocker()

        blocks = {
            "System Info": (["Key", "Value"], [{"Key": k, "Value": v} for k, v in system_info.items()]),
            "Users & Groups": (["Group", "Members"], groups),
            "Local Users": (["Username", "IsAdmin"], users),
            "RDP": (["Property", "Value"], [{"Property": "RDP Enabled", "Value": str(rdp.get("RDP Enabled", ""))}]),
            "Netstat": (["Line"], netstat),
            "Hosts": (["Entry"], hosts),
            "DNS Cache": (["Record"], [{"Record": str(r)} for r in dns_cache]),
            "Installed Software": (["Name", "Version", "Publisher", "InstallDate"], installed),
            "Portable Apps": (["Name", "Version", "Publisher", "Path"], portable),
            "Autoruns": (["Hive", "Path", "Name", "Value"], autoruns),
            "Processes": (["PID", "Name", "Path", "User"], processes),
            "Services": (["Name", "DisplayName", "Status", "StartType", "BinaryPath"], services),
            "Drivers": (["Module Name", "Display Name", "Driver Type", "Link Date"],
                         [{"Module Name": d.get("Module Name",""), "Display Name": d.get("Display Name",""),
                           "Driver Type": d.get("Driver Type",""), "Link Date": d.get("Link Date","")} for d in drivers]),
            "Scheduled Tasks": (list(tasks[0].keys()) if tasks else [], tasks),
            "Security Events": (list(events[0].keys()) if events else [], events),
            "Disks": (list(disks[0].keys()) if disks else [], disks),
            "BitLocker": (["Volume", "ConversionStatus", "ProtectionMethod"], bitlocker),
            "Environment": (["Variable", "Value"], env),
            "USB History": (["Device"], usb),
            "Unsigned Drivers": (["Path"], unsigned),
            "Legacy Tasks": (["Path"], legacy),
            "AppLocker": (["RawPolicy"], applocker),
        }

        print("Создание Excel...")
        wb = create_workbook(overview, blocks)
        buf = BytesIO()
        wb.save(buf)
        xlsx_data = buf.getvalue()

        print("Шифрование...")
        encrypted = encrypt_data(xlsx_data)
        with open(enc_path, "wb") as f:
            f.write(encrypted)
        print(f"Отчёт успешно сохранён: {enc_path}")
    except Exception:
        log_and_pause()
    finally:
        input("\nНажмите Enter для выхода...")

if __name__ == "__main__":
    main()