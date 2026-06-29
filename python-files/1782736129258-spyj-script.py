import pandas as pd
import requests
import datetime
import time
import json
import tkinter as tk
from openpyxl import load_workbook # Импортируем load_workbooks
from tkinter import filedialog

def send_api_requests_with_delay_py(excel_file_path: str, sheet_name: str):
    """
    Отправляет POST-запросы к API, используя данные из Excel файла,
    и записывает результаты (статус или ошибку) обратно в исходный файл,
    предполагая, что данные (ИНН) находятся в колонке B, а результаты будут записаны в колонку D.

    Args:
        excel_file_path (str): Путь к файлу Excel.
        sheet_name (str, optional): Название листа в Excel файле. По умолчанию 'Sheet1'.
    """

    api_url = "https://statusnpd.nalog.ru/api/v1/tracker/taxpayer_status"
    delay_seconds = 31  # Задержка между запросами в секундах
    

    try:
        inn_column_letter = input("Введите букву колонки с ИНН: ")
        target_column_letter = input("Введите букву колонки с результатами: ")
        # Читаем данные из Excel, указывая колонку с ИНН
        # header=0, чтобы первая строка использовалась как заголовки
        df_inn = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=0, usecols=inn_column_letter)

        # Проверяем, была ли успешно прочитана колонка с ИНН
        if df_inn.empty:
            print(f"Ошибка: Лист '{sheet_name}' в файле '{excel_file_path}' пуст или колонка '{inn_column_letter}' с ИНН не найдена/пуста.")
            return

        # Получаем текущую системную дату и форматируем ее как YYYY-MM-DD
        request_date = datetime.datetime.now().strftime("%Y-%m-%d")

        # Список для хранения результатов (статус/ошибка)
        results_list = []

        # Итерируемся по строкам DataFrame с ИНН
        for index, row in df_inn.iterrows():
            # Получаем ИНН из единственной колонки (индекс 0)
            inn = row.iloc[0]

            # Проверяем, что ИНН не является пустым значением (NaN) и не пустой строкой
            if pd.notna(inn) and str(inn).strip() != "":
                inn_str = str(inn).strip() # Приводим ИНН к строковому типу и убираем пробелы
                print(f"Обрабатывается строка {index + 1}: ИНН = {inn_str}")

                # Конструируем JSON-payload для API запроса
                payload = {
                    "inn": inn_str,
                    "requestDate": request_date
                }
                json_payload = json.dumps(payload) # Сериализуем словарь в строку JSON

                status_value = None
                error_message = None

                try:
                    # Отправляем POST-запрос на API
                    headers = {'Content-Type': 'application/json'}
                    response = requests.post(api_url, data=json_payload, headers=headers, timeout=60) # Устанавливаем таймаут 60 секунд

                    # Проверяем статус ответа HTTP. Если статус код не 2xx, будет вызвано исключение.
                    response.raise_for_status()

                    # Парсим JSON-ответ и извлекаем значение статуса
                    response_json = response.json()
                    status_value = response_json.get("status", "Статус не найден") # Получаем значение status, или значение по умолчанию

                    print(f"  ---> Успешно: Статус = {status_value}")

                except requests.exceptions.Timeout:
                    error_message = "Ошибка: Таймаут запроса."
                    print(f"  ---> Ошибка: {error_message}")
                except requests.exceptions.HTTPError as http_err:
                    error_message = f"Ошибка HTTP: {http_err} - {response.text}"
                    print(f"  ---> Ошибка: {error_message}")
                except requests.exceptions.RequestException as req_err:
                    error_message = f"Ошибка запроса: {req_err}"
                    print(f"  ---> Ошибка: {error_message}")
                except json.JSONDecodeError:
                    error_message = "Ошибка: Не удалось декодировать JSON из ответа."
                    print(f"  ---> Ошибка: {error_message}")
                except Exception as e:
                    error_message = f"Непредвиденная ошибка при обработке запроса: {e}"
                    print(f"  ---> Ошибка: {error_message}")

                # Добавляем результат (статус или ошибку) в список
                results_list.append(status_value if status_value is not None else error_message)

                # Применяем задержку, если это не последний запрос
                if index < df_inn.shape[0] - 1:
                    print(f"  Задержка {delay_seconds} секунд...")
                    time.sleep(delay_seconds)
            else:
                print(f"Строка {index + 1}: ИНН пуст или некорректен, пропускается.")
                # Добавляем запись с пометкой об ошибке, если ИНН был пустым
                results_list.append("Ошибка: Пустой ИНН")

        # --- ОБНОВЛЕНИЕ ИСХОДНОГО EXCEL ФАЙЛА ---
        try:
            workbook = load_workbook(excel_file_path)

            # Получаем лист, если он существует
            if sheet_name not in workbook.sheetnames:
                print(f"Ошибка: Лист '{sheet_name}' не найден в файле '{excel_file_path}'.")
                return

            sheet = workbook[sheet_name]

            # Преобразуем букву колонки в числовой индекс Excel (A=1, B=2, ...)
            target_col_index = 0
            if len(target_column_letter) == 1:
                target_col_index = ord(target_column_letter.upper()) - ord('A') + 1
            else:
                # Обработка многобуквенных индексов (AA, AB, ...)
                # TODO: Реализовать более сложную логику, если нужно
                print(f"Предупреждение: Нестандартный формат буквы колонки '{target_column_letter}'. Попытка использовать как однобуквенный индекс.")
                if target_column_letter.upper() in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
                    target_col_index = ord(target_column_letter.upper()) - ord('A') + 1
                else:
                    print(f"Ошибка: Недопустимая буква колонки '{target_column_letter}'.")
                    return

            # Проверяем, существует ли уже заголовок в целевой колонке.
            header_cell = sheet.cell(row=1, column=target_col_index)
            if header_cell.value is None:
                header_cell.value = "StatusResult" # Устанавливаем заголовок, если его нет
                print(f"Заголовок 'StatusResult' записан в колонку {target_column_letter}.")
            else:
                print(f"Колонка {target_column_letter} уже существует с заголовком '{header_cell.value}'. Результаты будут добавлены в нее.")

            # Записываем результаты в целевую колонку, начиная со второй строки (index + 2)
            for i, result in enumerate(results_list):
                sheet.cell(row=i + 2, column=target_col_index).value = result

            # Сохраняем изменения в книгу
            workbook.save(excel_file_path)
            print(f"\nРезультаты успешно записаны/обновлены в колонку {target_column_letter} файла: {excel_file_path}")

        except Exception as e:
                print(f"\nОшибка при обновлении файла Excel: {e}")
            
        finally:
            if 'writer' in locals() and writer:
                writer.close()

            print("Обработка завершена.")

    except FileNotFoundError:
        print(f"Ошибка: Файл Excel не найден по пути: {excel_file_path}")
    except ImportError:
        print("Ошибка: Необходимая библиотека 'openpyxl' не установлена.")
        print("Пожалуйста, установите ее: pip install openpyxl")
    except Exception as e:
        print(f"Произошла непредвиденная ошибка: {e}")

if __name__ == "__main__":

    root = tk.Tk()
    root.withdraw()  
    file_path = filedialog.askopenfilename()

    if file_path:
        print("Путь к файлу:", file_path)
    else:
        print("Файл не выбран.")

    root.destroy()

    my_excel_file_path = file_path
    my_sheet_name = input("Введите название листа: ")
    send_api_requests_with_delay_py(excel_file_path=my_excel_file_path, sheet_name=my_sheet_name)
