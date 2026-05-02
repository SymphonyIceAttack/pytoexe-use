import os
import glob
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

def normalize_time(val):
    """Приводит значение ячейки к строке ЧЧ:ММ, если оно является временем."""
    if val is None:
        return None
    if isinstance(val, (datetime.time, datetime.datetime)):
        return val.strftime("%H:%M")
    if isinstance(val, str):
        val = val.strip()
        parts = val.split(":")
        if len(parts) >= 2:
            try:
                return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
            except ValueError:
                return None
    if isinstance(val, (int, float)):
        if 0 <= val < 1:  # Excel хранит время как долю суток
            total_minutes = int(round(val * 24 * 60))
            h, m = divmod(total_minutes, 60)
            return f"{h:02d}:{m:02d}"
    return None

def main():
    # Директория, где лежит скрипт
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Ищем Excel-файлы, исключая временные копии (~$...)
    all_excel = glob.glob(os.path.join(script_dir, "*.xlsx")) + \
                glob.glob(os.path.join(script_dir, "*.xls"))
    excel_files = [f for f in all_excel if not os.path.basename(f).startswith("~$")]
    
    if not excel_files:
        print("❌ Ошибка: Excel файлы не найдены в папке со скриптом.")
        input("⏸ Нажмите Enter для выхода... ")
        return

    # 1️⃣ Поиск файла по первому слову "Отчет"
    target_file = None
    for f in excel_files:
        fname = os.path.basename(f)
        name_parts = os.path.splitext(fname)[0].split()
        first_word = name_parts[0] if name_parts else ""
        if first_word.lower() == "отчет":
            target_file = f
            break

    if target_file is None:
        print("❌ Ошибка: Файл, имя которого начинается со слова 'Отчет', не найден.")
        input("⏸ Нажмите Enter для выхода... ")
        return

    print(f"✅ Найден файл: {os.path.basename(target_file)}")

    try:
        wb = load_workbook(target_file)
        ws = wb.active  # Открываем активный лист

        # 🔥 УСТАНАВЛИВАЕМ ВЫСОТУ ВСЕХ СТРОК (20 пикселей ≈ 15 пунктов)
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 15  # 15 пунктов ≈ 20 пикселей

        # 🔥 УСТАНАВЛИВАЕМ ШИРИНУ СТОЛБЦА B (124 пикселя ≈ 17 символов)
        ws.column_dimensions['B'].width = 17

        # 🔥 УСТАНАВЛИВАЕМ ШИРИНУ СТОЛБЦА AG (305 пикселей ≈ 43 символа)
        ws.column_dimensions['AG'].width = 43

        # 2️⃣ Определяем, какие столбцы обрабатывать
        ignore_headers = {"фамилия", "имя", "общая величина преждевременности ухода"}
        cols_to_process = []
        for col in range(1, ws.max_column + 1):
            # 🔥 ИСКЛЮЧАЕМ СТОЛБЕЦ AG (33-й столбец)
            col_letter = get_column_letter(col)
            if col_letter == "AG":
                continue
            header = ws.cell(row=1, column=col).value
            if header and str(header).strip().lower() in ignore_headers:
                continue
            cols_to_process.append(col)
            
        print(f"📊 Будут обработаны столбцы (индексы): {cols_to_process}")
        processed_count = 0

        # Проходим по строкам, начиная со 2-й (1-я = заголовки)
        for row in range(2, ws.max_row + 1):
            for col in cols_to_process:
                cell = ws.cell(row=row, column=col)
                time_str = normalize_time(cell.value)

                if time_str is not None:
                    if time_str == "01:00":
                        cell.fill = YELLOW_FILL
                    elif time_str == "00:00":
                        continue  # Игнорируем по условию
                    else:
                        cell.fill = GREEN_FILL
                    processed_count += 1

        wb.save(target_file)
        print(f"✨ Готово! Заливка применена к {processed_count} ячейкам. Файл сохранён.")

    except Exception as e:
        print(f"💥 Произошла ошибка при обработке: {e}")

    # Пауза в консоли
    input("⏸ Нажмите Enter для выхода... ")

if __name__ == "__main__":
    main()