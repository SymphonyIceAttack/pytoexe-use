import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import subprocess
import sys

# --- 自动修复逻辑：如果没有库，就原地安装 ---
def install_requirements():
    try:
        import openpyxl
    except ImportError:
        # 尝试自动安装 openpyxl
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl
        except Exception as e:
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("环境错误", f"程序尝试自动安装库失败，请检查网络或手动执行: pip install openpyxl\n错误信息: {str(e)}")
            sys.exit()

# 执行安装检测
install_requirements()
from openpyxl import load_workbook

# --- 转换核心逻辑 ---
def one_excel_to_js_lite(excel_path, js_path, log_func):
    try:
        wb = load_workbook(excel_path, data_only=True)
        json_data = {}
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('_'): continue
            sheet = wb[sheet_name]
            header_row = 11 # 你的需求：第11行为列名
            max_row = sheet.max_row
            if max_row < header_row: continue
            
            headers = []
            valid_cols = []
            for i in range(1, sheet.max_column + 1):
                col_name = sheet.cell(row=header_row, column=i).value
                if col_name and not str(col_name).startswith('_'):
                    headers.append(str(col_name))
                    valid_cols.append(i)
            
            rows = []
            for r in range(header_row + 1, max_row + 1):
                row_dict = {}
                has_val = False
                for idx, col_idx in enumerate(valid_cols):
                    val = sheet.cell(row=r, column=col_idx).value
                    row_dict[headers[idx]] = val if val is not None else ""
                    if val is not None: has_val = True
                if has_val: rows.append(row_dict)
            json_data[sheet_name] = rows
            
        if json_data:
            with open(js_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=4, ensure_ascii=False)
            return True
        return False
    except Exception as e:
        log_func(f"❌ 出错: {os.path.basename(excel_path)} -> {str(e)}")
        return False

# --- GUI 界面 ---
class ConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel2Json 自动修复版")
        self.root.geometry("550x400")
        
        tk.Label(root, text="Excel 目录:").pack(pady=(10,0), padx=20, anchor="w")
        self.e_ent = tk.Entry(root)
        self.e_ent.pack(fill="x", padx=20)
        tk.Button(root, text="浏览", command=lambda: self.set_path(self.e_ent)).pack(anchor="e", padx=20)

        tk.Label(root, text="JSON 输出目录:").pack(pady=(10,0), padx=20, anchor="w")
        self.j_ent = tk.Entry(root)
        self.j_ent.pack(fill="x", padx=20)
        tk.Button(root, text="浏览", command=lambda: self.set_path(self.j_ent)).pack(anchor="e", padx=20)

        self.log_area = scrolledtext.ScrolledText(root, height=8)
        self.log_area.pack(fill="both", expand=True, padx=20, pady=10)

        self.btn = tk.Button(root, text="开始处理", bg="#2196F3", fg="white", height=2, command=self.start)
        self.btn.pack(fill="x", padx=20, pady=10)

    def set_path(self, ent):
        p = filedialog.askdirectory()
        if p: ent.delete(0, tk.END); ent.insert(0, p)

    def log(self, m):
        self.log_area.insert(tk.END, m + "\n"); self.log_area.see(tk.END)

    def start(self):
        threading.Thread(target=self.run, daemon=True).start()

    def run(self):
        s, d = self.e_ent.get(), self.j_ent.get()
        if not s or not d: return
        self.btn.config(state="disabled")
        try:
            files = [f for f in os.listdir(s) if f.endswith('.xlsx') and not f.startswith('~$')]
            for f in files:
                if one_excel_to_js_lite(os.path.join(s, f), os.path.join(d, f.replace('.xlsx', '.json')), self.log):
                    self.log(f"✅ 处理成功: {f}")
            self.log("✨ 任务完成")
        finally:
            self.btn.config(state="normal")

if __name__ == "__main__":
    r = tk.Tk()
    ConverterGUI(r)
    r.mainloop()